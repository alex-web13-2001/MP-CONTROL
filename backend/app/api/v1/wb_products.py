"""
WB Products API endpoints.

GET  /products/wb?shop_id=X&page=1&per_page=25&sort=revenue_7d&order=desc&filter=all&search=&period=7
PATCH /products/wb/cost  — update cost price for a WB product (by vendor_code)
POST /products/wb/cost/bulk — bulk upload cost prices via Excel
GET  /products/wb/cost/template — download Excel template
"""
import io
import logging
from datetime import date, datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.shop import Shop
from app.models.user import User
from sqlalchemy import select

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/products", tags=["WB Products"])


class WBCostUpdateRequest(BaseModel):
    shop_id: int
    vendor_code: str
    cost_price: float
    packaging_cost: float = 0


# ── WB Products List ──────────────────────────────────────

@router.get("/wb")
async def get_wb_products(
    shop_id: int = Query(...),
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=10, le=100),
    sort: str = Query("revenue_7d"),
    order: str = Query("desc"),
    filter: str = Query("all"),
    search: str = Query(""),
    period: int = Query(7, ge=1, le=366),
    date_from: Optional[date] = Query(None, description="Custom range start (YYYY-MM-DD)"),
    date_to: Optional[date] = Query(None, description="Custom range end (YYYY-MM-DD)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Get WB products list with sales, advertising, stock and margin data.
    """
    # Verify shop ownership
    shop_result = await db.execute(
        select(Shop).where(
            Shop.id == shop_id,
            Shop.user_id == current_user.id,
            Shop.marketplace == "wildberries",
        )
    )
    shop = shop_result.scalar_one_or_none()
    if not shop:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="WB магазин не найден")

    # ── Dates ─────────────────────────────────────────────
    MSK = timezone(timedelta(hours=3))
    today = datetime.now(MSK).date()
    if date_from and date_to:
        d_start = date_from
        d_end = date_to
    else:
        d_end = today
        d_start = today - timedelta(days=period - 1)
    span = (d_end - d_start).days + 1
    d_prev_end = d_start - timedelta(days=1)
    d_prev_start = d_prev_end - timedelta(days=span - 1)

    from app.core.clickhouse import get_clickhouse_client
    ch = get_clickhouse_client()

    # ── 1. Orders (current + prev period) ─────────────────
    try:
        orders_result = ch.query("""
            SELECT
                nm_id,
                any(supplier_article)  AS vendor_code,
                countIf(toDate(addHours(date, 3)) >= {d_start:Date} AND is_cancel = 0)  AS orders_cur,
                sumIf(price_with_disc, toDate(addHours(date, 3)) >= {d_start:Date} AND is_cancel = 0)  AS revenue_cur,
                countIf(toDate(addHours(date, 3)) < {d_start:Date} AND is_cancel = 0)   AS orders_prev,
                sumIf(price_with_disc, toDate(addHours(date, 3)) < {d_start:Date} AND is_cancel = 0)   AS revenue_prev,
                countIf(toDate(addHours(date, 3)) >= {d_start:Date} AND is_cancel = 1)  AS cancels_cur
            FROM mms_analytics.fact_orders_raw FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND toDate(addHours(date, 3)) >= {d_prev_start:Date}
              AND toDate(addHours(date, 3)) <= {d_end:Date}
            GROUP BY nm_id
        """, parameters={
            "shop_id": shop_id,
            "d_start": d_start,
            "d_end": d_end,
            "d_prev_start": d_prev_start,
        })
        orders_map = {}
        for r in orders_result.result_rows:
            nm_id = int(r[0])
            orders_cur = int(r[2])
            cancels_cur = int(r[6])
            total_with_cancels = orders_cur + cancels_cur
            orders_map[nm_id] = {
                "vendor_code": str(r[1] or ""),
                "orders_7d": orders_cur,
                "revenue_7d": float(r[3]),
                "orders_prev": int(r[4]),
                "revenue_prev": float(r[5]),
                "cancels": cancels_cur,
                "cancel_rate": round(cancels_cur / total_with_cancels * 100, 1) if total_with_cancels > 0 else 0.0,
            }
    except Exception as e:
        logger.warning("CH orders query failed: %s", e)
        orders_map = {}

    # ── 2. Ads (current period) ────────────────────────────
    try:
        ads_result = ch.query("""
            SELECT
                nm_id,
                sum(spend)  AS ad_spend,
                sum(views)  AS views,
                sum(clicks) AS clicks
            FROM mms_analytics.fact_advert_stats_v3 FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND date >= {d_start:Date}
              AND date <= {d_end:Date}
            GROUP BY nm_id
        """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": d_end})
        ads_map = {}
        for r in ads_result.result_rows:
            ads_map[int(r[0])] = {
                "ad_spend_7d": float(r[1]),
                "ad_views": int(r[2]),
                "ad_clicks": int(r[3]),
            }
    except Exception as e:
        logger.warning("CH ads query failed: %s", e)
        ads_map = {}

    # ── 2b. Active ad campaigns (status=9) per nm_id ────────
    active_ad_nm_ids: set[int] = set()
    try:
        active_ads_result = ch.query("""
            SELECT DISTINCT lb.nm_id
            FROM mms_analytics.log_wb_bids lb
            WHERE lb.shop_id = {shop_id:UInt32}
              AND lb.advert_id IN (
                  SELECT advert_id
                  FROM mms_analytics.dim_advert_campaigns FINAL
                  WHERE shop_id = {shop_id:UInt32} AND status = 9
              )
              AND lb.timestamp >= now() - INTERVAL 2 DAY
        """, parameters={"shop_id": shop_id})
        for r in active_ads_result.result_rows:
            active_ad_nm_ids.add(int(r[0]))
    except Exception as e:
        logger.warning("CH active ads query failed: %s", e)

    # ── 3. Stocks (latest snapshot) ────────────────────────
    try:
        stocks_result = ch.query("""
            SELECT
                nm_id,
                sumIf(quantity, NOT startsWith(warehouse_name, 'FBS:')) AS stock_fbo,
                sumIf(quantity, startsWith(warehouse_name, 'FBS:'))     AS stock_fbs
            FROM mms_analytics.fact_inventory_snapshot FINAL
            WHERE shop_id = {shop_id:UInt32}
            GROUP BY nm_id
        """, parameters={"shop_id": shop_id})
        stocks_map = {}
        for r in stocks_result.result_rows:
            stocks_map[int(r[0])] = {
                "stock_fbo": int(r[1]),
                "stock_fbs": int(r[2]),
            }
    except Exception as e:
        logger.warning("CH stocks query failed: %s", e)
        stocks_map = {}

    # ── 4. WB Finance data from fact_finances (SOURCE OF TRUTH) ──────
    # Используем ФАКТИЧЕСКИЕ данные из финансовых отчётов WB.
    # Поля идентичны /finances/wb/products (finances.py):
    #   revenue  = retail_price_withdisc_rub (розничная цена)
    #   payout   = payout_amount (ppvz_for_pay, к перечислению)
    #   logistics = wb_delivery_rub (доставка)
    #   storage  = storage_fee (хранение — обычно 0 по товарам, распределяется пропорционально)
    # NB: ppvz_for_pay = revenue - commission - acquiring. 
    #     Логистика, хранение, удержания — ОТДЕЛЬНЫЕ расходы, НЕ включены в payout!
    try:
        fees_result = ch.query("""
            SELECT
                toUInt64(external_id)  AS nm_id,
                -- Revenue (retail_price_withdisc_rub, как в финотчёте)
                sumIf(JSONExtractFloat(raw_payload, 'retail_price_withdisc_rub'),
                    operation_type = 'Продажа') AS revenue_rpw,
                sumIf(JSONExtractFloat(raw_payload, 'retail_price_withdisc_rub'),
                    operation_type = 'Возврат') AS returns_rpw,
                -- Payout (payout_amount = ppvz_for_pay)
                sumIf(payout_amount, operation_type = 'Продажа')
                  - sumIf(payout_amount, operation_type = 'Возврат') AS payout,
                -- Qty
                sumIf(quantity, operation_type = 'Продажа' AND quantity > 0) AS qty,
                -- Fee components
                sum(wb_delivery_rub)    AS logistics,
                sum(storage_fee)        AS storage,
                sum(acceptance_fee)     AS acceptance,
                sum(penalty_total)      AS fines,
                sum(wb_acquiring)       AS acquiring
            FROM mms_analytics.fact_finances FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND marketplace = 1
              AND event_date >= {d_start:Date}
              AND event_date <= {d_end:Date}
            GROUP BY nm_id
            HAVING revenue_rpw > 0 OR payout != 0 OR logistics != 0
        """, parameters={
            "shop_id": shop_id,
            "d_start": d_start,
            "d_end": d_end,
        })
        fees_map = {}
        for r in fees_result.result_rows:
            nm = int(r[0])
            rev_rpw = float(r[1] or 0)
            ret_rpw = abs(float(r[2] or 0))
            fees_map[nm] = {
                "fin_revenue": rev_rpw - ret_rpw,    # net revenue
                "payout": float(r[3] or 0),
                "qty": int(r[4] or 0),
                "logistics": abs(float(r[5] or 0)),
                "storage": abs(float(r[6] or 0)),
                "acceptance": abs(float(r[7] or 0)),
                "fines": abs(float(r[8] or 0)),
                "acquiring": abs(float(r[9] or 0)),
                "deductions": 0.0,  # будет заполнено из отдельного запроса
            }
    except Exception as e:
        logger.warning("CH fees query failed: %s", e)
        fees_map = {}

    # ── 4b. Non-ad deductions per product (bonus_type_name parsing) ──
    # Удержания в fact_finances приходят с external_id=0, но nm_id записан
    # в bonus_type_name: "товар NNNNNN". Парсим и привязываем к товарам.
    # ИСКЛЮЧАЕМ рекламные удержания (ВБ Продвижение) — ads берём из fact_advert_stats_v3.
    nm_to_vc_map = {}  # для связки nm_id -> vendor_code
    unlinked_ded = 0.0
    try:
        ded_result = ch.query("""
            SELECT
                toUInt64OrZero(extract(
                    JSONExtractString(raw_payload, 'bonus_type_name'),
                    'товар\\s+(\\d+)'
                )) AS parsed_nm_id,
                sum(abs(JSONExtractFloat(raw_payload, 'deduction'))) AS ded_amount
            FROM mms_analytics.fact_finances FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND marketplace = 1
              AND event_date >= {d_start:Date}
              AND event_date <= {d_end:Date}
              AND abs(JSONExtractFloat(raw_payload, 'deduction')) > 0
              -- Исключаем рекламные удержания (ВБ Продвижение)
              AND positionCaseInsensitiveUTF8(
                    JSONExtractString(raw_payload, 'bonus_type_name'), 'продвижение') = 0
            GROUP BY parsed_nm_id
        """, parameters={
            "shop_id": shop_id,
            "d_start": d_start,
            "d_end": d_end,
        })
        for r in ded_result.result_rows:
            nm = int(r[0] or 0)
            ded_val = float(r[1] or 0)
            if nm and nm in fees_map:
                fees_map[nm]["deductions"] += ded_val
            else:
                unlinked_ded += ded_val
    except Exception as e:
        logger.warning("CH deductions by nm_id query failed: %s", e)

    # ── 4c. Proportional distribution: ONLY storage (NOT deductions) ──
    # Storage: WB не привязывает к конкретному товару → распределяем пропорционально
    # Unlinked deductions (отзывы за баллы, прочие): НЕ распределяем по товарам!
    #   Они идут только в общий итог P&L, но НЕ в расходы конкретных товаров.
    unknown_fees = fees_map.pop(0, None)
    undist_storage = unknown_fees.get("storage", 0) if unknown_fees else 0

    # Storage → пропорциональное распределение
    if undist_storage > 0:
        total_rev = sum(f["fin_revenue"] for f in fees_map.values() if f.get("fin_revenue", 0) > 0)
        if total_rev > 0:
            for nm, f in fees_map.items():
                rev = f.get("fin_revenue", 0)
                if rev > 0:
                    share = rev / total_rev
                    f["storage"] += round(undist_storage * share, 2)

    # ── 4d. Prev period revenue from fact_finances ────────────────────
    try:
        prev_result = ch.query("""
            SELECT
                toUInt64(external_id) AS nm_id,
                sumIf(JSONExtractFloat(raw_payload, 'retail_price_withdisc_rub'),
                    operation_type = 'Продажа')
                - sumIf(JSONExtractFloat(raw_payload, 'retail_price_withdisc_rub'),
                    operation_type = 'Возврат') AS fin_revenue_prev
            FROM mms_analytics.fact_finances FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND marketplace = 1
              AND event_date >= {d_prev_start:Date}
              AND event_date <= {d_prev_end:Date}
            GROUP BY nm_id
            HAVING fin_revenue_prev > 0
        """, parameters={
            "shop_id": shop_id,
            "d_prev_start": d_prev_start,
            "d_prev_end": d_prev_end,
        })
        for r in prev_result.result_rows:
            nm = int(r[0])
            if nm in fees_map:
                fees_map[nm]["fin_revenue_prev"] = float(r[1])
            else:
                fees_map[nm] = {"fin_revenue_prev": float(r[1])}
    except Exception as e:
        logger.warning("CH prev period query failed: %s", e)

    # ── 4e. Fallback Chain — estimate fees for products without fact_finances ──
    # Только для товаров с заказами (fact_orders_raw) но БЕЗ fact_finances
    estimated_nm_ids: set[int] = set()
    needs_estimation = set()
    for nm_id, od in orders_map.items():
        if od.get("orders_7d", 0) > 0 and (
            nm_id not in fees_map or fees_map[nm_id].get("payout", 0) == 0
        ):
            needs_estimation.add(nm_id)

    if needs_estimation:
        # Step 1: Per-product historical PER-UNIT rates (last 90 days)
        try:
            hist_result = ch.query("""
                SELECT
                    toUInt64(external_id)  AS nm_id,
                    sum(payout_amount) / nullIf(
                        sumIf(quantity, operation_type = 'Продажа' AND quantity > 0), 0
                    ) AS payout_per_unit,
                    sum(wb_delivery_rub) / nullIf(
                        sumIf(quantity, operation_type = 'Продажа' AND quantity > 0), 0
                    ) AS logistics_per_unit
                FROM mms_analytics.fact_finances FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND marketplace = 1
                  AND event_date >= today() - 90
                  AND toUInt64(external_id) IN {nm_ids:Array(UInt64)}
                GROUP BY nm_id
                HAVING sumIf(quantity, operation_type = 'Продажа' AND quantity > 0) > 0
            """, parameters={
                "shop_id": shop_id,
                "nm_ids": list(needs_estimation),
            })
            for r in hist_result.result_rows:
                nm = int(r[0])
                od = orders_map.get(nm, {})
                cur_orders = od.get("orders_7d", 0)
                ppu = float(r[1] or 0)
                lpu = abs(float(r[2] or 0))

                fees_map[nm] = {
                    "fin_revenue": od.get("revenue_7d", 0.0),
                    "payout": round(ppu * cur_orders, 2),
                    "qty": cur_orders,
                    "logistics": round(lpu * cur_orders, 2),
                    "storage": 0.0,
                    "acceptance": 0.0,
                    "fines": 0.0,
                    "acquiring": 0.0,
                    "deductions": 0.0,
                }
                estimated_nm_ids.add(nm)
                needs_estimation.discard(nm)
        except Exception as e:
            logger.warning("CH historical rates query failed: %s", e)

        # Step 2: Shop-wide average PER-UNIT rates (final fallback)
        if needs_estimation:
            try:
                avg_result = ch.query("""
                    SELECT
                        sum(payout_amount) / nullIf(
                            sumIf(quantity, operation_type = 'Продажа' AND quantity > 0), 0
                        ),
                        sum(wb_delivery_rub) / nullIf(
                            sumIf(quantity, operation_type = 'Продажа' AND quantity > 0), 0
                        )
                    FROM mms_analytics.fact_finances FINAL
                    WHERE shop_id = {shop_id:UInt32}
                      AND marketplace = 1
                      AND event_date >= today() - 90
                """, parameters={"shop_id": shop_id})

                avg_row = avg_result.result_rows[0] if avg_result.result_rows else None
                if avg_row and avg_row[0] is not None:
                    avg_ppu = float(avg_row[0] or 0)
                    avg_lpu = abs(float(avg_row[1] or 0))

                    for nm_id in list(needs_estimation):
                        od = orders_map.get(nm_id, {})
                        cur_orders = od.get("orders_7d", 0)
                        fees_map[nm_id] = {
                            "fin_revenue": od.get("revenue_7d", 0.0),
                            "payout": round(avg_ppu * cur_orders, 2),
                            "qty": cur_orders,
                            "logistics": round(avg_lpu * cur_orders, 2),
                            "storage": 0.0,
                            "acceptance": 0.0,
                            "fines": 0.0,
                            "acquiring": 0.0,
                            "deductions": 0.0,
                        }
                        estimated_nm_ids.add(nm_id)
            except Exception as e:
                logger.warning("CH shop-wide avg rates query failed: %s", e)

    # ── Unit economics from fact_finances (rolling 30d) for estimated profit ──
    unit_rate_map: dict[int, dict] = {}  # nm_id → {avg_payout, avg_logistics}
    try:
        unit_econ_data = ch.query("""
            SELECT
                toUInt64(external_id) AS nm_id,
                sum(payout_amount) / nullIf(
                    sumIf(quantity, operation_type = 'Продажа' AND quantity > 0), 0
                ) AS avg_payout_per_unit,
                sum(wb_delivery_rub) / nullIf(
                    sumIf(quantity, operation_type = 'Продажа' AND quantity > 0), 0
                ) AS avg_logistics_per_unit
            FROM mms_analytics.fact_finances FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND marketplace = 1
              AND event_date >= today() - 30
            GROUP BY nm_id
            HAVING sumIf(quantity, operation_type = 'Продажа' AND quantity > 0) > 0
        """, parameters={"shop_id": shop_id}).result_rows
        for r in unit_econ_data:
            unit_rate_map[int(r[0])] = {
                "avg_payout": float(r[1] or 0),
                "avg_logistics": abs(float(r[2] or 0)),
            }
    except Exception as e:
        logger.warning("CH unit economics query failed: %s", e)

    ch.close()

    # ── 5. Product catalog from PostgreSQL (dim_products) ──
    pg_result = await db.execute(
        text("""
            SELECT dp.nm_id,
                   dp.name,
                   COALESCE(dp.main_image_url, '') AS image_url,
                   COALESCE(dp.current_price, 0)   AS current_price,
                   COALESCE(dp.vendor_code, '')     AS vendor_code,
                   COALESCE(pc.cost_price, 0)       AS cost_price,
                   COALESCE(pc.packaging_cost, 0)   AS packaging_cost
            FROM dim_products dp
            LEFT JOIN product_costs pc
                ON pc.shop_id = dp.shop_id AND pc.offer_id = dp.vendor_code
            WHERE dp.shop_id = :shop_id
        """),
        {"shop_id": shop_id},
    )

    # ── 6. Merge all data ────────────────────────────────────
    all_nm_ids = set(orders_map.keys()) | set(ads_map.keys()) | set(fees_map.keys())
    pg_rows = pg_result.fetchall()
    pg_map = {}
    for row in pg_rows:
        nm_id = int(row[0])
        pg_map[nm_id] = {
            "nm_id": nm_id,
            "name": row[1] or "",
            "image_url": row[2],
            "current_price": float(row[3]),
            "vendor_code": row[4],
            "cost_price": float(row[5]),
            "packaging_cost": float(row[6]),
        }
        all_nm_ids.add(nm_id)

    products = []
    for nm_id in all_nm_ids:
        info = pg_map.get(nm_id, {})
        orders = orders_map.get(nm_id, {})
        ads = ads_map.get(nm_id, {})
        stocks = stocks_map.get(nm_id, {})

        vendor_code = info.get("vendor_code") or orders.get("vendor_code") or str(nm_id)
        ad_spend_7d = ads.get("ad_spend_7d", 0.0)
        cost_price = info.get("cost_price", 0.0)
        packaging_cost = info.get("packaging_cost", 0.0)

        # ── Оперативные данные из fact_orders_raw ────
        orders_raw = orders.get("orders_7d", 0)
        revenue_raw = orders.get("revenue_7d", 0.0)
        orders_prev_raw = orders.get("orders_prev", 0)
        revenue_prev_raw = orders.get("revenue_prev", 0.0)

        # ── Данные из fact_finances (информационные — для fee breakdown) ────
        fees = fees_map.get(nm_id, {})
        revenue_prev = revenue_prev_raw

        # Всегда используем fact_orders_raw для revenue и orders
        # (fact_finances отстаёт на неделю, расхождение с Dashboard/Sales)
        orders_7d = orders_raw
        revenue_7d = revenue_raw

        avg_price = round(revenue_7d / orders_7d, 2) if orders_7d > 0 else 0.0

        # ── Fees из fact_finances (БЕЗ масштабирования) ────
        # Масштабирование удалено — оно создавало расхождение с финотчётом.
        # Если fact_finances нет — все fees = 0 (до появления отчёта).
        payout = fees.get("payout", 0.0)
        fee_logistics = fees.get("logistics", 0.0)
        fee_storage = fees.get("storage", 0.0)
        fee_acceptance = fees.get("acceptance", 0.0)
        fee_deductions = fees.get("deductions", 0.0)
        fee_fines = fees.get("fines", 0.0)

        # mp_fees = все расходы маркетплейса (логистика+хранение+удержания+приёмка+штрафы)
        # НЕ включает комиссию (она уже в payout) и НЕ включает рекламу (отдельно)
        mp_fees = round(fee_logistics + fee_storage + fee_deductions + fee_acceptance + fee_fines, 2)
        mp_fees_percent = round(mp_fees / revenue_7d * 100, 1) if revenue_7d > 0 else 0.0

        current_price = info.get("current_price", 0.0)
        sales_amount = round(revenue_7d, 2)

        # ── DRR ───────────────────────────────────────────
        drr = round(ad_spend_7d / sales_amount * 100, 1) if sales_amount > 0 else 0.0

        # ── Revenue delta ─────────────────────────────────
        if revenue_prev > 0:
            revenue_delta = round((revenue_7d - revenue_prev) / revenue_prev * 100, 1)
        elif revenue_7d > 0:
            revenue_delta = 100.0
        else:
            revenue_delta = 0.0

        # ── Расчётная прибыль (unit economics из fact_finances 30d) ─────
        # unit_profit = avg_payout_per_unit - avg_logistics_per_unit - COGS
        # gross_profit = unit_profit × orders - ad_spend
        # Аналогичная формула используется на Dashboard для единообразия.
        gross_profit = None
        margin = None
        unit_rates = unit_rate_map.get(nm_id, {})
        avg_payout_pu = unit_rates.get("avg_payout", 0)
        avg_logistics_pu = unit_rates.get("avg_logistics", 0)
        if orders_7d > 0 and avg_payout_pu > 0:
            unit_cogs = cost_price + packaging_cost
            unit_profit = avg_payout_pu - avg_logistics_pu - unit_cogs
            gross_profit = round(unit_profit * orders_7d - ad_spend_7d, 2)
            if sales_amount > 0:
                margin = round(gross_profit / sales_amount * 100, 1)

        p = {
            "nm_id": nm_id,
            "vendor_code": vendor_code,
            "name": info.get("name", vendor_code),
            "image_url": info.get("image_url", ""),
            "current_price": current_price,
            "cost_price": cost_price,
            "packaging_cost": packaging_cost,
            # P&L waterfall
            "avg_price": avg_price,
            "sales_amount": sales_amount,
            "orders_7d": orders_7d,
            "revenue_7d": revenue_7d,
            "orders_prev": 0,
            "revenue_prev": revenue_prev,
            "revenue_delta": revenue_delta,
            "payout": round(payout, 2),
            # Ads (from fact_advert_stats_v3)
            "ad_spend_7d": ad_spend_7d,
            "ad_views": ads.get("ad_views", 0),
            "ad_clicks": ads.get("ad_clicks", 0),
            "drr": drr,
            # Stocks
            "stock_fbo": stocks.get("stock_fbo", 0),
            "stock_fbs": stocks.get("stock_fbs", 0),
            # Cancellations
            "cancels": orders.get("cancels", 0),
            "cancel_rate": orders.get("cancel_rate", 0.0),
            # WB Marketplace fees (logistics+storage+deductions+acceptance+fines)
            "mp_fees": mp_fees,
            "mp_fees_percent": mp_fees_percent,
            "mp_fees_logistics": round(fee_logistics, 2),
            "mp_fees_storage": round(fee_storage, 2),
            "mp_fees_deductions": round(fee_deductions, 2),
            "mp_fees_acceptance": round(fee_acceptance, 2),
            "mp_fees_fines": round(fee_fines, 2),
            # Profit
            "gross_profit": gross_profit,
            "margin": margin,
            # Fee source indicator for frontend ≈ badge
            "fees_source": "estimated" if nm_id in estimated_nm_ids else "actual",
            # Active ad campaigns (status=9) indicator
            "has_active_ads": nm_id in active_ad_nm_ids,
        }
        products.append(p)


    # ── 6. Search ────────────────────────────────────────────
    if search:
        s = search.lower()
        products = [
            p for p in products
            if s in p["name"].lower() or s in p["vendor_code"].lower()
        ]

    # ── 7. Filter ────────────────────────────────────────────
    if filter == "with_ads":
        products = [p for p in products if p["ad_spend_7d"] > 0]
    elif filter == "no_ads":
        products = [p for p in products if p["ad_spend_7d"] == 0 and (p["orders_7d"] > 0 or p["revenue_7d"] > 0)]
    elif filter == "leaders":
        median = sorted([p["revenue_7d"] for p in products])[len(products) // 2] if products else 0
        products = [p for p in products if p["revenue_7d"] >= median and p["revenue_7d"] > 0]
    elif filter == "falling":
        products = [p for p in products if p["revenue_delta"] < -10]
    elif filter == "problems":
        products = [p for p in products if p["stock_fbo"] + p["stock_fbs"] == 0 and p["revenue_7d"] > 0]
    elif filter == "in_stock":
        products = [p for p in products if p["stock_fbo"] + p["stock_fbs"] > 0]
    elif filter == "no_stock":
        products = [p for p in products if p["stock_fbo"] + p["stock_fbs"] == 0]
    elif filter == "archived":
        # WB API не отдаёт флаг is_archived → показываем пустой список
        products = []

    # ── 7b. Hide ghost products (no catalog entry + no sales) ──
    products = [
        p for p in products
        if p["nm_id"] in pg_map  # есть в каталоге
        or p["orders_7d"] > 0    # или есть продажи за период
        or p["payout"] > 0       # или есть выплаты
    ]

    # ── 8. Sort ──────────────────────────────────────────────
    SORT_FIELDS = {
        "revenue_7d", "orders_7d", "ad_spend_7d", "drr",
        "revenue_delta", "stock_fbo", "stock_fbs", "current_price",
        "gross_profit", "margin",
    }
    sort_key = sort if sort in SORT_FIELDS else "revenue_7d"
    reverse = (order == "desc")

    def sort_val(p):
        v = p.get(sort_key)
        if v is None:
            return float("-inf") if reverse else float("inf")
        return v

    products.sort(key=sort_val, reverse=reverse)

    # ── 9. Paginate ──────────────────────────────────────────
    # Compute totals across ALL filtered products (before pagination)
    t_stocks = 0
    t_orders = 0
    t_sales = 0.0
    t_revenue = 0.0
    t_payout = 0.0
    t_ad_spend = 0.0
    t_mp_fees = 0.0
    t_mp_fees_logistics = 0.0
    t_mp_fees_storage = 0.0
    t_mp_fees_deductions = 0.0
    t_mp_fees_acceptance = 0.0
    t_mp_fees_fines = 0.0
    t_total_cogs = 0.0
    t_profit = 0.0
    t_profit_count = 0
    t_cancels = 0
    for p in products:
        t_stocks += p["stock_fbo"] + p["stock_fbs"]
        t_orders += p["orders_7d"]
        t_sales += p["sales_amount"]
        t_revenue += p["revenue_7d"]
        t_payout += p["payout"]
        t_ad_spend += p["ad_spend_7d"]
        t_mp_fees += p["mp_fees"]
        t_mp_fees_logistics += p.get("mp_fees_logistics", 0)
        t_mp_fees_storage += p.get("mp_fees_storage", 0)
        t_mp_fees_deductions += p.get("mp_fees_deductions", 0)
        t_mp_fees_acceptance += p.get("mp_fees_acceptance", 0)
        t_mp_fees_fines += p.get("mp_fees_fines", 0)
        t_cancels += p.get("cancels", 0)
        # COGS для итого-строки
        if p["cost_price"] > 0 and p["orders_7d"] > 0:
            t_total_cogs += (p["cost_price"] + p.get("packaging_cost", 0)) * p["orders_7d"]
        if p["gross_profit"] is not None:
            t_profit += p["gross_profit"]
            t_profit_count += 1

    totals = {
        "count": len(products),
        "stocks": t_stocks,
        "orders": t_orders,
        "sales": round(t_sales, 2),
        "revenue": round(t_revenue, 2),
        "payout": round(t_payout, 2),
        "avg_price": round(t_payout / t_orders, 2) if t_orders > 0 else 0,
        "total_cogs": round(t_total_cogs, 2),
        "ad_spend": round(t_ad_spend, 2),
        "drr": round(t_ad_spend / t_sales * 100, 1) if t_sales > 0 else 0,
        "returns_pct": 0,
        "cancels": t_cancels,
        "cancel_rate": round(t_cancels / (t_orders + t_cancels) * 100, 1) if (t_orders + t_cancels) > 0 else 0,
        # Только расходы, привязанные к товарам (без общих удержаний)
        "mp_fees": round(t_mp_fees, 2),
        "mp_fees_logistics": round(t_mp_fees_logistics, 2),
        "mp_fees_storage": round(t_mp_fees_storage, 2),
        "mp_fees_deductions": round(t_mp_fees_deductions, 2),
        "mp_fees_acceptance": round(t_mp_fees_acceptance, 2),
        "mp_fees_fines": round(t_mp_fees_fines, 2),
        "mp_fees_pct": round(t_mp_fees / t_revenue * 100, 1) if t_revenue > 0 else 0,
        "profit": round(t_profit, 2),
        "profit_pct": round(t_profit / t_sales * 100, 1) if t_sales > 0 and t_profit_count > 0 else 0,
        "profit_count": t_profit_count,
    }

    total = len(products)
    offset = (page - 1) * per_page
    page_products = products[offset: offset + per_page]

    cost_missing = sum(1 for p in products if p["cost_price"] == 0)

    return {
        "total": total,
        "page": page,
        "per_page": per_page,
        "products": page_products,
        "cost_missing_count": cost_missing,
        "totals": totals,
    }


# ── PATCH cost price (single product) ─────────────────────

@router.patch("/wb/cost")
async def update_wb_cost(
    req: WBCostUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update cost price for a single WB product (by vendor_code)."""
    shop_result = await db.execute(
        select(Shop).where(Shop.id == req.shop_id, Shop.user_id == current_user.id)
    )
    if not shop_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Shop not found")

    await db.execute(
        text("""
            INSERT INTO product_costs (shop_id, offer_id, cost_price, packaging_cost)
            VALUES (:shop_id, :offer_id, :cost_price, :packaging_cost)
            ON CONFLICT (shop_id, offer_id) DO UPDATE SET
                cost_price = EXCLUDED.cost_price,
                packaging_cost = EXCLUDED.packaging_cost,
                updated_at = NOW()
        """),
        {
            "shop_id": req.shop_id,
            "offer_id": req.vendor_code,
            "cost_price": req.cost_price,
            "packaging_cost": req.packaging_cost,
        },
    )
    await db.commit()
    return {"ok": True, "vendor_code": req.vendor_code, "cost_price": req.cost_price}


# ── POST bulk Excel upload ─────────────────────────────────

@router.post("/wb/cost/bulk")
async def upload_wb_cost_excel(
    shop_id: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Bulk upload WB cost prices from Excel.
    Format: column A = vendor_code (артикул), column B = cost_price.
    """
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    if not shop_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Shop not found")

    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате .xlsx")

    try:
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="Пустой Excel файл")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка чтения Excel: {e}")

    updated = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=2, values_only=True), start=1):
        vendor_code_raw, cost_raw = row
        if vendor_code_raw is None or cost_raw is None:
            continue
        vendor_code = str(vendor_code_raw).strip()
        if not vendor_code:
            continue
        try:
            # Handle Russian locale: non-breaking spaces as thousands separator, comma as decimal
            if isinstance(cost_raw, str):
                cost_str = cost_raw.replace('\xa0', '').replace(' ', '').replace(',', '.')
                cost_price = float(cost_str)
            else:
                cost_price = float(cost_raw)
            if cost_price < 0:
                errors.append(f"Строка {row_idx}: отрицательная цена ({cost_price})")
                continue
        except (ValueError, TypeError):
            errors.append(f"Строка {row_idx}: невалидная цена '{cost_raw}'")
            continue

        await db.execute(
            text("""
                INSERT INTO product_costs (shop_id, offer_id, cost_price)
                VALUES (:shop_id, :offer_id, :cost_price)
                ON CONFLICT (shop_id, offer_id) DO UPDATE SET
                    cost_price = EXCLUDED.cost_price,
                    updated_at = NOW()
            """),
            {"shop_id": shop_id, "offer_id": vendor_code, "cost_price": cost_price},
        )
        updated += 1

    await db.commit()
    return {"ok": True, "updated": updated, "errors": errors[:20]}


# ── GET Excel template ─────────────────────────────────────

@router.get("/wb/cost/template")
async def download_wb_cost_template(
    shop_id: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download Excel template with all WB vendor_codes for cost price entry."""
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    if not shop_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Shop not found")

    result = await db.execute(
        text("""
            SELECT dp.vendor_code,
                   COALESCE(pc.cost_price, 0) AS cost_price,
                   dp.name
            FROM dim_products dp
            LEFT JOIN product_costs pc
                ON pc.shop_id = dp.shop_id AND pc.offer_id = dp.vendor_code
            WHERE dp.shop_id = :shop_id
              AND dp.vendor_code IS NOT NULL
              AND dp.vendor_code != ''
            ORDER BY dp.name
        """),
        {"shop_id": shop_id},
    )
    rows = result.fetchall()

    import openpyxl
    wb_xl = openpyxl.Workbook()
    ws = wb_xl.active
    ws.title = "Себестоимость WB"
    ws.append(["Артикул продавца", "Себестоимость", "Название (справочно)"])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 50
    for r in rows:
        ws.append([r[0], float(r[1]), r[2] or ""])

    buf = io.BytesIO()
    wb_xl.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=wb_cost_template_{shop_id}.xlsx"},
    )
