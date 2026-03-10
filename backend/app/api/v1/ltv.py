"""
LTV Analysis API endpoints (Ozon only).

GET /sales/ozon/ltv           — KPI + cohorts + SKU table + time distribution
GET /sales/ozon/ltv/chain     — Purchase chain for a specific SKU (L1→L5)
"""
import logging
from datetime import date, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.shop import Shop
from app.models.user import User

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/sales", tags=["LTV Analysis"])


# ── Helpers ────────────────────────────────────────────────────

import math


def _sf(v) -> float:
    """Safe float: NaN/Inf -> 0.0 for JSON compliance."""
    try:
        f = float(v) if v is not None else 0.0
        return 0.0 if (math.isnan(f) or math.isinf(f)) else f
    except (TypeError, ValueError):
        return 0.0


def _ltv_dates(
    period: str,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> tuple[date, date]:
    """Return (start, end) dates for LTV analysis."""
    today = date.today()
    if date_from and date_to:
        return date_from, date_to
    mapping = {
        "30d": 30,
        "90d": 90,
        "6m": 180,
        "1y": 365,
        "all": 730,
    }
    days = mapping.get(period, 180)
    return today - timedelta(days=days), today


async def _verify_ozon_shop(db: AsyncSession, shop_id: int, user: User) -> Shop:
    """Verify shop exists and belongs to user."""
    result = await db.execute(
        select(Shop).where(
            Shop.id == shop_id,
            Shop.user_id == user.id,
            Shop.marketplace == "ozon",
        )
    )
    shop = result.scalar_one_or_none()
    if not shop:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ozon магазин не найден",
        )
    return shop


# ══════════════════════════════════════════════════════════════
# Main LTV endpoint — KPI + Cohorts + SKU table + Distribution
# ══════════════════════════════════════════════════════════════


@router.get("/ozon/ltv")
async def get_ozon_ltv(
    shop_id: int = Query(..., description="Shop ID"),
    period: str = Query("6m", description="Period: 30d, 90d, 6m, 1y, all"),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Ozon customer LTV analysis.

    Returns KPI metrics, cohort retention matrix, SKU repeat purchase table,
    and time-between-purchases distribution.

    Source: fact_ozon_orders FINAL — client_id = splitByChar('-', posting_number)[1]
    """
    await _verify_ozon_shop(db, shop_id, current_user)
    start_date, end_date = _ltv_dates(period, date_from, date_to)

    from app.core.clickhouse import get_clickhouse_client

    try:
        ch = get_clickhouse_client()
        params = {"shop_id": shop_id, "start_date": start_date, "end_date": end_date}

        # ══════════════════════════════════════════════
        # 1. KPI — unique clients, repeat, avg LTV, etc
        # ══════════════════════════════════════════════
        kpi_result = ch.query("""
            WITH clients AS (
                SELECT
                    splitByChar('-', posting_number)[1] AS client_id,
                    count(DISTINCT order_number) AS orders,
                    sum(price * quantity) AS client_rev,
                    min(toDate(addHours(in_process_at, 3))) AS first_order_date,
                    max(toDate(addHours(in_process_at, 3))) AS last_order_date
                FROM mms_analytics.fact_ozon_orders FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND toDate(addHours(in_process_at, 3)) >= {start_date:Date}
                  AND toDate(addHours(in_process_at, 3)) <= {end_date:Date}
                GROUP BY client_id
            )
            SELECT
                count() AS total_clients,
                countIf(orders >= 2) AS repeat_clients,
                round(avg(client_rev), 0) AS avg_ltv,
                round(sum(client_rev) / nullIf(sum(orders), 0), 0) AS avg_check,
                round(avg(orders), 2) AS avg_orders_per_client,
                sum(client_rev) AS total_revenue,
                round(avgIf(client_rev / nullIf(orders, 0), orders >= 2), 0) AS avg_check_repeat
            FROM clients
        """, parameters=params).first_row

        kpi = {
            "total_clients": int(kpi_result[0] or 0),
            "repeat_clients": int(kpi_result[1] or 0),
            "repeat_rate": round(int(kpi_result[1] or 0) / max(int(kpi_result[0] or 0), 1) * 100, 1),
            "avg_ltv": _sf(kpi_result[2]),
            "avg_check": _sf(kpi_result[3]),
            "avg_orders_per_client": _sf(kpi_result[4]),
            "total_revenue": _sf(kpi_result[5]),
            "avg_check_repeat": _sf(kpi_result[6]),
        }

        # ══════════════════════════════════════════════
        # 2. Cohort Retention Matrix
        # ══════════════════════════════════════════════
        cohort_rows = ch.query("""
            WITH
                client_orders AS (
                    SELECT
                        splitByChar('-', posting_number)[1] AS client_id,
                        order_number,
                        min(toDate(addHours(in_process_at, 3))) AS order_date
                    FROM mms_analytics.fact_ozon_orders FINAL
                    WHERE shop_id = {shop_id:UInt32}
                      AND toDate(addHours(in_process_at, 3)) >= {start_date:Date}
                      AND toDate(addHours(in_process_at, 3)) <= {end_date:Date}
                    GROUP BY client_id, order_number
                ),
                first_orders AS (
                    SELECT client_id, min(order_date) AS first_date
                    FROM client_orders
                    GROUP BY client_id
                ),
                cohort_data AS (
                    SELECT
                        fo.client_id,
                        toStartOfMonth(fo.first_date) AS cohort_month,
                        dateDiff('month', fo.first_date, co.order_date) AS month_offset
                    FROM first_orders fo
                    JOIN client_orders co ON fo.client_id = co.client_id
                )
            SELECT
                toString(cohort_month) AS cohort,
                month_offset,
                count(DISTINCT client_id) AS clients
            FROM cohort_data
            GROUP BY cohort_month, month_offset
            ORDER BY cohort_month, month_offset
        """, parameters=params).result_rows

        # Build cohort matrix
        cohorts: dict[str, dict] = {}
        for row in cohort_rows:
            cohort_key = row[0][:7]  # YYYY-MM
            offset = int(row[1])
            clients = int(row[2])
            if cohort_key not in cohorts:
                cohorts[cohort_key] = {"cohort": cohort_key, "size": 0, "months": {}}
            if offset == 0:
                cohorts[cohort_key]["size"] = clients
            cohorts[cohort_key]["months"][str(offset)] = clients

        # Calculate retention percentages
        cohort_matrix = []
        for key in sorted(cohorts.keys()):
            c = cohorts[key]
            size = c["size"]
            months_data = {}
            for offset_str, count in c["months"].items():
                months_data[offset_str] = {
                    "clients": count,
                    "rate": round(count / max(size, 1) * 100, 1),
                }
            cohort_matrix.append({
                "cohort": key,
                "size": size,
                "months": months_data,
            })

        # ══════════════════════════════════════════════
        # 3. SKU repeat purchase summary table
        # ══════════════════════════════════════════════
        sku_rows = ch.query("""
            WITH
                sku_clients AS (
                    SELECT
                        sku,
                        offer_id,
                        product_name,
                        splitByChar('-', posting_number)[1] AS client_id,
                        order_number,
                        toDate(addHours(in_process_at, 3)) AS order_date,
                        sum(price * quantity) AS order_revenue,
                        sum(quantity) AS qty
                    FROM mms_analytics.fact_ozon_orders FINAL
                    WHERE shop_id = {shop_id:UInt32}
                      AND toDate(addHours(in_process_at, 3)) >= {start_date:Date}
                      AND toDate(addHours(in_process_at, 3)) <= {end_date:Date}
                    GROUP BY sku, offer_id, product_name, client_id, order_number, order_date
                ),
                sku_client_agg AS (
                    SELECT
                        sku,
                        argMax(offer_id, order_date) AS offer_id,
                        any(product_name) AS product_name,
                        client_id,
                        count() AS purchases,
                        sum(order_revenue) AS client_revenue,
                        sum(qty) AS client_qty,
                        min(order_date) AS first_buy,
                        max(order_date) AS last_buy
                    FROM sku_clients
                    GROUP BY sku, client_id
                )
            SELECT
                sku,
                argMax(offer_id, last_buy) AS offer_id,
                any(product_name) AS name,
                count() AS total_buyers,
                sum(client_qty) AS total_qty,
                sum(client_revenue) AS total_revenue,
                countIf(purchases >= 2) AS repeat_buyers,
                countIf(purchases >= 3) AS buyers_3plus,
                round(countIf(purchases >= 2) / nullIf(count(), 0) * 100, 1) AS conv_2,
                round(countIf(purchases >= 3) / nullIf(countIf(purchases >= 2), 0) * 100, 1) AS conv_3,
                round(avg(if(purchases >= 2, dateDiff('day', first_buy, last_buy) / (purchases - 1), 0)), 0) AS avg_days_between,
                round(avgIf(client_revenue, purchases >= 2), 0) AS avg_ltv_repeat
            FROM sku_client_agg
            GROUP BY sku
            HAVING total_buyers >= 1
            ORDER BY repeat_buyers DESC, total_revenue DESC
            LIMIT 100
        """, parameters=params).result_rows

        sku_table = []
        for r in sku_rows:
            sku_table.append({
                "sku": int(r[0]),
                "offer_id": str(r[1]),
                "name": str(r[2])[:80],
                "total_buyers": int(r[3] or 0),
                "total_qty": int(r[4] or 0),
                "total_revenue": _sf(r[5]),
                "repeat_buyers": int(r[6] or 0),
                "buyers_3plus": int(r[7] or 0),
                "conv_to_2": _sf(r[8]),
                "conv_to_3": _sf(r[9]),
                "avg_days_between": int(_sf(r[10])),
                "avg_ltv_repeat": _sf(r[11]),
                "image_url": "",
            })

        # ── Enrich with images from PostgreSQL ──
        if sku_table:
            sku_ids = [s["sku"] for s in sku_table]
            from sqlalchemy import text as sa_text
            img_result = await db.execute(
                sa_text(
                    "SELECT sku, main_image_url FROM dim_ozon_products "
                    "WHERE shop_id = :shop_id AND sku = ANY(:skus)"
                ),
                {"shop_id": shop_id, "skus": sku_ids},
            )
            img_map = {int(row[0]): (row[1] or "") for row in img_result.fetchall()}
            for item in sku_table:
                item["image_url"] = img_map.get(item["sku"], "")

        # ══════════════════════════════════════════════
        # 4. Time-to-repeat distribution (histogram data)
        # ══════════════════════════════════════════════
        dist_rows = ch.query("""
            WITH
                client_orders AS (
                    SELECT
                        splitByChar('-', posting_number)[1] AS client_id,
                        order_number,
                        min(toDate(addHours(in_process_at, 3))) AS order_date
                    FROM mms_analytics.fact_ozon_orders FINAL
                    WHERE shop_id = {shop_id:UInt32}
                      AND toDate(addHours(in_process_at, 3)) >= {start_date:Date}
                      AND toDate(addHours(in_process_at, 3)) <= {end_date:Date}
                    GROUP BY client_id, order_number
                ),
                with_next AS (
                    SELECT
                        client_id,
                        order_date,
                        leadInFrame(order_date, 1, toDate('1970-01-01'))
                            OVER (PARTITION BY client_id ORDER BY order_date
                                  ROWS BETWEEN CURRENT ROW AND UNBOUNDED FOLLOWING)
                            AS next_date
                    FROM client_orders
                ),
                gaps AS (
                    SELECT dateDiff('day', order_date, next_date) AS days_gap
                    FROM with_next
                    WHERE next_date > toDate('1970-01-01')
                      AND next_date > order_date
                )
            SELECT
                multiIf(
                    days_gap <= 7, '0-7',
                    days_gap <= 14, '8-14',
                    days_gap <= 30, '15-30',
                    days_gap <= 60, '31-60',
                    days_gap <= 90, '61-90',
                    '90+'
                ) AS bucket,
                count() AS cnt,
                round(avg(days_gap), 1) AS avg_days
            FROM gaps
            GROUP BY bucket
            ORDER BY
                multiIf(
                    bucket = '0-7', 1,
                    bucket = '8-14', 2,
                    bucket = '15-30', 3,
                    bucket = '31-60', 4,
                    bucket = '61-90', 5,
                    6
                )
        """, parameters=params).result_rows

        distribution = [
            {"bucket": str(r[0]), "count": int(r[1]), "avg_days": float(r[2])}
            for r in dist_rows
        ]

        # ══════════════════════════════════════════════
        # 5. Monthly new vs repeat buyers
        # ══════════════════════════════════════════════
        monthly_rows = ch.query("""
            WITH
                client_orders AS (
                    SELECT
                        splitByChar('-', posting_number)[1] AS client_id,
                        order_number,
                        min(toDate(addHours(in_process_at, 3))) AS order_date,
                        sum(price * quantity) AS order_revenue
                    FROM mms_analytics.fact_ozon_orders FINAL
                    WHERE shop_id = {shop_id:UInt32}
                      AND toDate(addHours(in_process_at, 3)) >= {start_date:Date}
                      AND toDate(addHours(in_process_at, 3)) <= {end_date:Date}
                    GROUP BY client_id, order_number
                ),
                first_purchase AS (
                    SELECT client_id, min(order_date) AS first_date
                    FROM client_orders
                    GROUP BY client_id
                ),
                monthly_agg AS (
                    SELECT
                        toStartOfMonth(co.order_date) AS month,
                        co.client_id,
                        fp.first_date,
                        sum(co.order_revenue) AS revenue
                    FROM client_orders co
                    JOIN first_purchase fp ON co.client_id = fp.client_id
                    GROUP BY month, co.client_id, fp.first_date
                )
            SELECT
                toString(month) AS m,
                count() AS total,
                countIf(toStartOfMonth(first_date) = month) AS new_buyers,
                countIf(toStartOfMonth(first_date) < month) AS repeat_buyers,
                round(sumIf(revenue, toStartOfMonth(first_date) = month), 0) AS new_revenue,
                round(sumIf(revenue, toStartOfMonth(first_date) < month), 0) AS repeat_revenue
            FROM monthly_agg
            GROUP BY month
            ORDER BY month
        """, parameters=params).result_rows

        monthly_buyers = [
            {
                "month": str(r[0])[:7],
                "total": int(r[1] or 0),
                "new_buyers": int(r[2] or 0),
                "repeat_buyers": int(r[3] or 0),
                "new_revenue": _sf(r[4]),
                "repeat_revenue": _sf(r[5]),
            }
            for r in monthly_rows
        ]

        ch.close()

        return {
            "shop_id": shop_id,
            "period": period,
            "date_range": {"start": str(start_date), "end": str(end_date)},
            "kpi": kpi,
            "cohort_matrix": cohort_matrix,
            "sku_table": sku_table,
            "time_distribution": distribution,
            "monthly_buyers": monthly_buyers,
        }

    except Exception as e:
        logger.error("LTV analysis error: %s", e, exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Ошибка анализа LTV: {str(e)}",
        )


# ══════════════════════════════════════════════════════════════
# Purchase Chain endpoint — L1→L5 for specific SKU
# ══════════════════════════════════════════════════════════════


@router.get("/ozon/ltv/chain")
async def get_ozon_purchase_chain(
    shop_id: int = Query(..., description="Shop ID"),
    sku: int = Query(..., description="SKU to analyze chain for"),
    period: str = Query("6m", description="Period: 30d, 90d, 6m, 1y, all"),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Purchase chain analysis for a specific SKU.

    Shows what customers buy in their 2nd, 3rd, 4th, 5th purchases
    after initially buying this SKU. Tracks cross-sell and up-sell patterns.
    """
    await _verify_ozon_shop(db, shop_id, current_user)
    start_date, end_date = _ltv_dates(period, date_from, date_to)

    from app.core.clickhouse import get_clickhouse_client

    try:
        ch = get_clickhouse_client()
        params = {
            "shop_id": shop_id,
            "target_sku": sku,
            "start_date": start_date,
            "end_date": end_date,
        }

        # ── Step 1: Get all client orders with numbered sequence ──
        chain_data = ch.query("""
            WITH
                -- All orders per client, one row per (client, order, date)
                all_orders AS (
                    SELECT
                        splitByChar('-', posting_number)[1] AS client_id,
                        order_number,
                        sku,
                        argMax(offer_id, toDate(addHours(in_process_at, 3))) AS offer_id,
                        argMax(product_name, toDate(addHours(in_process_at, 3))) AS product_name,
                        min(toDate(addHours(in_process_at, 3))) AS order_date,
                        sum(price * quantity) AS revenue,
                        sum(quantity) AS qty
                    FROM mms_analytics.fact_ozon_orders FINAL
                    WHERE shop_id = {shop_id:UInt32}
                      AND toDate(addHours(in_process_at, 3)) >= {start_date:Date}
                      AND toDate(addHours(in_process_at, 3)) <= {end_date:Date}
                    GROUP BY client_id, order_number, sku
                ),
                -- Clients who bought the target SKU
                target_clients AS (
                    SELECT DISTINCT client_id
                    FROM all_orders
                    WHERE sku = {target_sku:UInt64}
                ),
                -- Number each client's orders chronologically
                client_numbered AS (
                    SELECT
                        ao.client_id,
                        ao.order_number,
                        ao.sku,
                        ao.offer_id,
                        ao.product_name,
                        ao.order_date,
                        ao.revenue,
                        ao.qty,
                        dense_rank() OVER (
                            PARTITION BY ao.client_id
                            ORDER BY ao.order_date, ao.order_number
                        ) AS purchase_num
                    FROM all_orders ao
                    JOIN target_clients tc ON ao.client_id = tc.client_id
                ),
                -- Find which purchase_num the target SKU was first bought
                first_target AS (
                    SELECT client_id, min(purchase_num) AS target_pnum
                    FROM client_numbered
                    WHERE sku = {target_sku:UInt64}
                    GROUP BY client_id
                ),
                -- Reindex: L1 = target purchase, L2 = next, etc
                reindexed AS (
                    SELECT
                        cn.client_id,
                        cn.sku,
                        cn.offer_id,
                        cn.product_name,
                        cn.order_date,
                        cn.revenue,
                        cn.qty,
                        cn.purchase_num - ft.target_pnum + 1 AS level
                    FROM client_numbered cn
                    JOIN first_target ft ON cn.client_id = ft.client_id
                    WHERE cn.purchase_num >= ft.target_pnum
                      AND cn.purchase_num < ft.target_pnum + 5
                )
            SELECT
                level,
                sku,
                argMax(offer_id, order_date) AS offer_id,
                argMax(product_name, order_date) AS name,
                count(DISTINCT client_id) AS buyers,
                sum(qty) AS total_qty,
                round(sum(revenue), 0) AS total_revenue,
                round(avg(revenue), 0) AS avg_revenue
            FROM reindexed
            GROUP BY level, sku
            ORDER BY level, buyers DESC
        """, parameters=params).result_rows

        # ── Step 2: Get L1 stats for context ──
        l1_stats = ch.query("""
            SELECT
                count(DISTINCT splitByChar('-', posting_number)[1]) AS total_buyers,
                sum(quantity) AS total_qty,
                round(sum(price * quantity), 0) AS total_revenue,
                round(avg(price), 0) AS avg_price,
                argMax(offer_id, toDate(addHours(in_process_at, 3))) AS offer_id,
                argMax(product_name, toDate(addHours(in_process_at, 3))) AS name
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND sku = {target_sku:UInt64}
              AND toDate(addHours(in_process_at, 3)) >= {start_date:Date}
              AND toDate(addHours(in_process_at, 3)) <= {end_date:Date}
        """, parameters=params).first_row

        # ── Step 3: Time between purchases ──
        time_between = ch.query("""
            WITH
                client_orders AS (
                    SELECT
                        splitByChar('-', posting_number)[1] AS client_id,
                        order_number,
                        min(toDate(addHours(in_process_at, 3))) AS order_date
                    FROM mms_analytics.fact_ozon_orders FINAL
                    WHERE shop_id = {shop_id:UInt32}
                      AND toDate(addHours(in_process_at, 3)) >= {start_date:Date}
                      AND toDate(addHours(in_process_at, 3)) <= {end_date:Date}
                      AND splitByChar('-', posting_number)[1] IN (
                          SELECT DISTINCT splitByChar('-', posting_number)[1]
                          FROM mms_analytics.fact_ozon_orders FINAL
                          WHERE shop_id = {shop_id:UInt32} AND sku = {target_sku:UInt64}
                      )
                    GROUP BY client_id, order_number
                ),
                numbered AS (
                    SELECT
                        client_id,
                        order_date,
                        row_number() OVER (PARTITION BY client_id ORDER BY order_date) AS rn
                    FROM client_orders
                )
            SELECT
                round(avgIf(
                    dateDiff('day', n1.order_date, n2.order_date),
                    n1.rn = 1 AND n2.rn = 2
                ), 0) AS avg_days_1_to_2,
                round(avgIf(
                    dateDiff('day', n1.order_date, n2.order_date),
                    n1.rn = 2 AND n2.rn = 3
                ), 0) AS avg_days_2_to_3,
                round(avgIf(
                    dateDiff('day', n1.order_date, n2.order_date),
                    n1.rn = 3 AND n2.rn = 4
                ), 0) AS avg_days_3_to_4,
                round(avgIf(
                    dateDiff('day', n1.order_date, n2.order_date),
                    n1.rn = 4 AND n2.rn = 5
                ), 0) AS avg_days_4_to_5
            FROM numbered n1
            JOIN numbered n2 ON n1.client_id = n2.client_id
            WHERE n2.rn = n1.rn + 1
        """, parameters=params).first_row

        ch.close()

        # ── Build response ──
        # Group chain_data by level
        levels: dict[int, list] = {}
        for row in chain_data:
            lvl = int(row[0])
            if lvl not in levels:
                levels[lvl] = []
            levels[lvl].append({
                "sku": int(row[1]),
                "offer_id": str(row[2]),
                "name": str(row[3])[:80],
                "buyers": int(row[4] or 0),
                "total_qty": int(row[5] or 0),
                "total_revenue": _sf(row[6]),
                "avg_revenue": _sf(row[7]),
            })

        # Calculate conversions
        l1_total = int(l1_stats[0] or 0)
        chain_levels = []
        for lvl_num in range(1, 6):
            products = levels.get(lvl_num, [])
            lvl_total_buyers = sum(p["buyers"] for p in products)
            prev_total = l1_total if lvl_num == 1 else sum(
                p["buyers"] for p in levels.get(lvl_num - 1, [])
            )

            # Add percentage from L1 to each product
            for p in products:
                p["pct_of_l1"] = round(p["buyers"] / max(l1_total, 1) * 100, 1)
                p["pct_of_level"] = round(
                    p["buyers"] / max(lvl_total_buyers, 1) * 100, 1
                )

            chain_levels.append({
                "level": lvl_num,
                "total_buyers": lvl_total_buyers,
                "conversion_from_prev": round(
                    lvl_total_buyers / max(prev_total, 1) * 100, 1
                ) if lvl_num > 1 else 100.0,
                "conversion_from_l1": round(
                    lvl_total_buyers / max(l1_total, 1) * 100, 1
                ),
                "products": products[:10],  # Top 10 per level
            })

        # ── Enrich with canonical offer_id from PostgreSQL ──
        all_chain_skus = set()
        all_chain_skus.add(sku)  # L1 SKU
        for cl in chain_levels:
            for p in cl["products"]:
                all_chain_skus.add(p["sku"])

        if all_chain_skus:
            from sqlalchemy import text as sa_text
            pg_result = await db.execute(
                sa_text(
                    "SELECT sku, offer_id FROM dim_ozon_products "
                    "WHERE shop_id = :shop_id AND sku = ANY(:skus)"
                ),
                {"shop_id": shop_id, "skus": list(all_chain_skus)},
            )
            canon_map = {int(row[0]): str(row[1]) for row in pg_result.fetchall()}
            # Override offer_id in chain products
            for cl in chain_levels:
                for p in cl["products"]:
                    if p["sku"] in canon_map:
                        p["offer_id"] = canon_map[p["sku"]]
            # Override L1 offer_id
            l1_offer_id = canon_map.get(sku, str(l1_stats[4]))
        else:
            l1_offer_id = str(l1_stats[4])

        return {
            "shop_id": shop_id,
            "target_sku": sku,
            "period": period,
            "date_range": {"start": str(start_date), "end": str(end_date)},
            "l1": {
                "sku": sku,
                "offer_id": l1_offer_id,
                "name": str(l1_stats[5])[:80],
                "total_buyers": l1_total,
                "total_qty": int(l1_stats[1] or 0),
                "total_revenue": _sf(l1_stats[2]),
                "avg_price": _sf(l1_stats[3]),
            },
            "chain": chain_levels,
            "avg_days_between": {
                "l1_to_l2": int(_sf(time_between[0])),
                "l2_to_l3": int(_sf(time_between[1])),
                "l3_to_l4": int(_sf(time_between[2])),
                "l4_to_l5": int(_sf(time_between[3])),
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error("Purchase chain error: %s", e, exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Ошибка цепочки продаж: {str(e)}",
        )


# ══════════════════════════════════════════════════════════════
# Excel Export — LTV full report
# ══════════════════════════════════════════════════════════════

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _build_ltv_xlsx(
    ltv_data: dict,
    sku_retention: list[dict],
    shop_name: str,
    marketplace: str,
    sku_transitions: list[dict] | None = None,
) -> BytesIO:
    """Build Excel workbook with full LTV report (7 sheets)."""
    wb = Workbook()

    # ── Styles ──
    HDR_FILL = PatternFill("solid", fgColor="1F2937")
    HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F2937")
    SUBTITLE_FONT = Font(name="Calibri", size=11, color="6B7280")
    MONEY_FMT = '#,##0'
    PCT_FMT = '0.0%'
    NUM_FMT = '#,##0'
    THIN_BORDER = Border(bottom=Side(style="thin", color="E5E7EB"))
    GREEN_FILL = PatternFill("solid", fgColor="D1FAE5")
    VIOLET_FILL = PatternFill("solid", fgColor="EDE9FE")
    AMBER_FILL = PatternFill("solid", fgColor="FEF3C7")

    # Heatmap fills for retention %
    def retention_fill(pct: float):
        if pct >= 50: return PatternFill("solid", fgColor="7C3AED")
        if pct >= 30: return PatternFill("solid", fgColor="8B5CF6")
        if pct >= 20: return PatternFill("solid", fgColor="A78BFA")
        if pct >= 10: return PatternFill("solid", fgColor="C4B5FD")
        if pct >= 5:  return PatternFill("solid", fgColor="DDD6FE")
        if pct > 0:   return PatternFill("solid", fgColor="EDE9FE")
        return None

    def retention_font(pct: float):
        if pct >= 20:
            return Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        return Font(name="Calibri", bold=True, color="1F2937", size=11)

    def _hdr_row(ws, row, headers, widths):
        for i, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=row, column=i, value=h)
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[c.column_letter].width = w

    kpi = ltv_data.get("kpi", {})
    dr = ltv_data.get("date_range", {})
    period_str = f'{dr.get("start", "?")} — {dr.get("end", "?")}'

    # ══════════════════════════════════════════════
    # Sheet 1: KPI Summary
    # ══════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "📊 KPI"
    ws1.sheet_properties.tabColor = "8B5CF6"

    ws1.merge_cells("A1:D1")
    ws1["A1"] = f"LTV-анализ — {shop_name} ({marketplace})"
    ws1["A1"].font = TITLE_FONT
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:D2")
    ws1["A2"] = f"Период: {period_str}"
    ws1["A2"].font = SUBTITLE_FONT

    kpi_items = [
        ("Уникальные клиенты", kpi.get("total_clients", 0), NUM_FMT),
        ("Повторные клиенты", kpi.get("repeat_clients", 0), NUM_FMT),
        ("Доля повторных (%)", kpi.get("repeat_rate", 0) / 100, PCT_FMT),
        ("Средний LTV", kpi.get("avg_ltv", 0), MONEY_FMT),
        ("Средний чек", kpi.get("avg_check", 0), MONEY_FMT),
        ("Заказов / клиент", kpi.get("avg_orders_per_client", 0), '0.00'),
        ("Общая выручка", kpi.get("total_revenue", 0), MONEY_FMT),
    ]
    _hdr_row(ws1, 4, ["Метрика", "Значение"], [30, 22])
    for i, (label, val, fmt) in enumerate(kpi_items, 5):
        ws1.cell(row=i, column=1, value=label).font = Font(name="Calibri", size=11, bold=True)
        c = ws1.cell(row=i, column=2, value=val)
        c.number_format = fmt
        c.font = Font(name="Calibri", size=12, bold=True, color="7C3AED")
        c.alignment = Alignment(horizontal="right")
        ws1.cell(row=i, column=1).border = THIN_BORDER
        c.border = THIN_BORDER

    # ══════════════════════════════════════════════
    # Sheet 2: Monthly Buyers
    # ══════════════════════════════════════════════
    monthly = ltv_data.get("monthly_buyers", [])
    ws2 = wb.create_sheet("📅 Месяцы")
    ws2.sheet_properties.tabColor = "10B981"

    ws2.merge_cells("A1:H1")
    ws2["A1"] = "Новые и повторные покупатели по месяцам"
    ws2["A1"].font = TITLE_FONT
    ws2.row_dimensions[1].height = 28

    _hdr_row(ws2, 3, [
        "Месяц", "Всего", "Новые", "Повторные",
        "% новых", "% повтор.", "Выручка новых", "Выручка повтор.",
    ], [14, 12, 12, 12, 10, 10, 18, 18])

    for i, m in enumerate(monthly, 4):
        total = m.get("total", 0) or 1
        new_b = m.get("new_buyers", 0)
        rep_b = m.get("repeat_buyers", 0)
        ws2.cell(row=i, column=1, value=m.get("month", "")).font = Font(name="Calibri", bold=True, size=11)
        ws2.cell(row=i, column=2, value=total).number_format = NUM_FMT
        c_new = ws2.cell(row=i, column=3, value=new_b)
        c_new.number_format = NUM_FMT
        c_new.fill = GREEN_FILL
        c_rep = ws2.cell(row=i, column=4, value=rep_b)
        c_rep.number_format = NUM_FMT
        c_rep.fill = VIOLET_FILL
        ws2.cell(row=i, column=5, value=new_b / total).number_format = PCT_FMT
        ws2.cell(row=i, column=6, value=rep_b / total).number_format = PCT_FMT
        ws2.cell(row=i, column=7, value=m.get("new_revenue", 0)).number_format = MONEY_FMT
        ws2.cell(row=i, column=8, value=m.get("repeat_revenue", 0)).number_format = MONEY_FMT
        for col in range(1, 9):
            ws2.cell(row=i, column=col).border = THIN_BORDER

    # ══════════════════════════════════════════════
    # Sheet 3: SKU Retention Map (1→5 purchase)
    # ══════════════════════════════════════════════
    ws3 = wb.create_sheet("🗺️ Retention по SKU")
    ws3.sheet_properties.tabColor = "F59E0B"

    ws3.merge_cells("A1:L1")
    ws3["A1"] = "Карта удержания по товарам (покупка 1 → 5)"
    ws3["A1"].font = TITLE_FONT
    ws3.row_dimensions[1].height = 28

    ws3.merge_cells("A2:L2")
    ws3["A2"] = "Сколько клиентов возвращаются за покупкой 2, 3, 4, 5 после первой"
    ws3["A2"].font = SUBTITLE_FONT

    ret_headers = [
        "Артикул", "Название", "Покупателей (1)", "Покупка 2", "% → 2",
        "Покупка 3", "% → 3", "Покупка 4", "% → 4",
        "Покупка 5", "% → 5", "Ср. дней между",
    ]
    ret_widths = [18, 35, 14, 12, 10, 12, 10, 12, 10, 12, 10, 14]
    _hdr_row(ws3, 4, ret_headers, ret_widths)

    for i, sku in enumerate(sku_retention, 5):
        ws3.cell(row=i, column=1, value=sku.get("offer_id", "")).font = Font(name="Calibri", size=10, bold=True)
        ws3.cell(row=i, column=2, value=sku.get("name", "")[:60]).font = Font(name="Calibri", size=10)
        b1 = sku.get("buyers_1", 0)
        b2 = sku.get("buyers_2", 0)
        b3 = sku.get("buyers_3", 0)
        b4 = sku.get("buyers_4", 0)
        b5 = sku.get("buyers_5", 0)
        p2 = b2 / max(b1, 1) * 100
        p3 = b3 / max(b2, 1) * 100
        p4 = b4 / max(b3, 1) * 100
        p5 = b5 / max(b4, 1) * 100

        ws3.cell(row=i, column=3, value=b1).number_format = NUM_FMT
        ws3.cell(row=i, column=3).font = Font(name="Calibri", bold=True, size=11)

        for col, buyers, pct in [(4, b2, p2), (6, b3, p3), (8, b4, p4), (10, b5, p5)]:
            c_b = ws3.cell(row=i, column=col, value=buyers)
            c_b.number_format = NUM_FMT
            c_p = ws3.cell(row=i, column=col + 1, value=pct / 100)
            c_p.number_format = PCT_FMT
            fill = retention_fill(pct)
            if fill:
                c_p.fill = fill
                c_p.font = retention_font(pct)

        avg_d = sku.get("avg_days", 0)
        ws3.cell(row=i, column=12, value=f"{avg_d} дн" if avg_d > 0 else "—")
        for col in range(1, 13):
            ws3.cell(row=i, column=col).border = THIN_BORDER

    # ══════════════════════════════════════════════
    # Sheet 4: SKU Table (full)
    # ══════════════════════════════════════════════
    sku_table = ltv_data.get("sku_table", [])
    ws4 = wb.create_sheet("📦 Товары")
    ws4.sheet_properties.tabColor = "6366F1"

    ws4.merge_cells("A1:I1")
    ws4["A1"] = "Таблица повторных покупок по товарам"
    ws4["A1"].font = TITLE_FONT
    ws4.row_dimensions[1].height = 28

    ws4.merge_cells("A2:I2")
    ws4["A2"] = "Покупатели, вернувшиеся за повторной покупкой того же товара"
    ws4["A2"].font = SUBTITLE_FONT

    sku_headers = [
        "Артикул", "Название товара",
        "Всего покупателей", "Из них повторных",
        "Повтор в 2-ю покупку", "Повтор в 3-ю покупку",
        "Ср. дней между покупками", "Ср. чек повторных ₽", "Общая выручка ₽",
    ]
    sku_widths = [18, 38, 17, 17, 19, 19, 22, 20, 18]
    _hdr_row(ws4, 4, sku_headers, sku_widths)

    for i, s in enumerate(sku_table, 5):
        ws4.cell(row=i, column=1, value=s.get("offer_id", "")).font = Font(name="Calibri", size=10, bold=True)
        ws4.cell(row=i, column=2, value=s.get("name", "")[:60]).font = Font(name="Calibri", size=10)
        ws4.cell(row=i, column=3, value=s.get("total_buyers", 0)).number_format = NUM_FMT
        c_rep = ws4.cell(row=i, column=4, value=s.get("repeat_buyers", 0))
        c_rep.number_format = NUM_FMT
        c_rep.font = Font(name="Calibri", bold=True, color="7C3AED", size=11)
        ws4.cell(row=i, column=5, value=s.get("conv_to_2", 0) / 100).number_format = PCT_FMT
        ws4.cell(row=i, column=6, value=s.get("conv_to_3", 0) / 100).number_format = PCT_FMT
        avg_d = s.get("avg_days_between", 0)
        c_days = ws4.cell(row=i, column=7, value=avg_d if avg_d > 0 else None)
        if avg_d > 0:
            c_days.number_format = '0'
        ws4.cell(row=i, column=8, value=s.get("avg_ltv_repeat", 0)).number_format = MONEY_FMT
        ws4.cell(row=i, column=9, value=s.get("total_revenue", 0)).number_format = MONEY_FMT
        for col in range(1, 10):
            ws4.cell(row=i, column=col).border = THIN_BORDER

    # ══════════════════════════════════════════════
    # Sheet 5: Cohort Matrix
    # ══════════════════════════════════════════════
    cohorts = ltv_data.get("cohort_matrix", [])
    ws5 = wb.create_sheet("🔄 Когорты")
    ws5.sheet_properties.tabColor = "EC4899"

    ws5.merge_cells("A1:I1")
    ws5["A1"] = "Когортная матрица Retention"
    ws5["A1"].font = TITLE_FONT
    ws5.row_dimensions[1].height = 28

    if cohorts:
        max_offset = max(
            int(k) for c in cohorts for k in c.get("months", {}).keys()
        )
        n_months = min(max_offset + 1, 7)

        headers_c = ["Когорта", "Размер"] + [f"+{m} мес" if m > 0 else "Мес 0" for m in range(n_months)]
        widths_c = [14, 12] + [12] * n_months
        _hdr_row(ws5, 3, headers_c, widths_c)

        for i, c in enumerate(cohorts, 4):
            ws5.cell(row=i, column=1, value=c.get("cohort", "")).font = Font(name="Calibri", bold=True, size=11)
            ws5.cell(row=i, column=2, value=c.get("size", 0)).number_format = NUM_FMT
            months = c.get("months", {})
            for m_idx in range(n_months):
                m_data = months.get(str(m_idx))
                if m_data:
                    rate = m_data.get("rate", 0)
                    cell = ws5.cell(row=i, column=3 + m_idx, value=rate / 100)
                    cell.number_format = PCT_FMT
                    fill = retention_fill(rate)
                    if fill:
                        cell.fill = fill
                        cell.font = retention_font(rate)
            for col in range(1, 3 + n_months):
                ws5.cell(row=i, column=col).border = THIN_BORDER

    # ══════════════════════════════════════════════
    # Sheet 6: Time Distribution
    # ══════════════════════════════════════════════
    time_dist = ltv_data.get("time_distribution", [])
    ws6 = wb.create_sheet("⏱ Время")
    ws6.sheet_properties.tabColor = "F97316"

    ws6.merge_cells("A1:D1")
    ws6["A1"] = "Время до повторной покупки"
    ws6["A1"].font = TITLE_FONT
    ws6.row_dimensions[1].height = 28

    _hdr_row(ws6, 3, ["Период (дней)", "Кол-во", "Ср. дней", "Доля"], [16, 12, 12, 12])
    total_dist = sum(d.get("count", 0) for d in time_dist) or 1
    for i, d in enumerate(time_dist, 4):
        ws6.cell(row=i, column=1, value=d.get("bucket", "")).font = Font(name="Calibri", bold=True, size=11)
        ws6.cell(row=i, column=2, value=d.get("count", 0)).number_format = NUM_FMT
        ws6.cell(row=i, column=3, value=d.get("avg_days", 0)).number_format = '0.0'
        ws6.cell(row=i, column=4, value=d.get("count", 0) / total_dist).number_format = PCT_FMT
        for col in range(1, 5):
            ws6.cell(row=i, column=col).border = THIN_BORDER

    # ══════════════════════════════════════════════════════════
    # Sheet 7: Cross-SKU Transitions
    # ══════════════════════════════════════════════════════════
    if sku_transitions:
        ws7 = wb.create_sheet("🔀 Переходы")
        ws7.sheet_properties.tabColor = "8B5CF6"

        ws7.merge_cells("A1:G1")
        ws7["A1"] = "Куда переходят покупатели (покупка 1 → 5)"
        ws7["A1"].font = TITLE_FONT
        ws7.row_dimensions[1].height = 28

        ws7.merge_cells("A2:G2")
        ws7["A2"] = "После покупки исходного товара, какие товары клиенты покупают дальше (топ-3 на каждом уровне)"
        ws7["A2"].font = SUBTITLE_FONT

        t_headers = [
            "Исходный товар (1-я покупка)",
            "Покупателей",
            "Покупка №",
            "Товар перехода",
            "Артикул перехода",
            "Покупателей",
            "% от исходных",
        ]
        t_widths = [38, 14, 12, 38, 20, 14, 14]
        _hdr_row(ws7, 4, t_headers, t_widths)

        LEVEL_FILL = {
            2: PatternFill("solid", fgColor="EDE9FE"),
            3: PatternFill("solid", fgColor="DBEAFE"),
            4: PatternFill("solid", fgColor="D1FAE5"),
            5: PatternFill("solid", fgColor="FEF3C7"),
        }
        SAME_SKU_FONT = Font(name="Calibri", bold=True, color="7C3AED", size=10)
        NORMAL_FONT = Font(name="Calibri", size=10)

        row_idx = 5
        current_source = None
        for t in sku_transitions:
            source = t["source_name"]
            is_new_source = source != current_source
            if is_new_source:
                if current_source is not None:
                    row_idx += 1  # blank separator
                current_source = source

            c1 = ws7.cell(row=row_idx, column=1)
            if is_new_source:
                c1.value = source
                c1.font = Font(name="Calibri", bold=True, size=11)
            c2 = ws7.cell(row=row_idx, column=2)
            if is_new_source:
                c2.value = t.get("source_buyers", 0)
                c2.number_format = NUM_FMT

            level = t.get("level", 2)
            ws7.cell(row=row_idx, column=3, value=f"{level}-я").font = Font(name="Calibri", bold=True, size=10)
            fill = LEVEL_FILL.get(level)
            if fill:
                ws7.cell(row=row_idx, column=3).fill = fill

            is_same = t.get("is_same_sku", False)
            name_font = SAME_SKU_FONT if is_same else NORMAL_FONT
            name_val = t.get("target_name", "")
            if is_same:
                name_val = "⭐ " + name_val
            ws7.cell(row=row_idx, column=4, value=name_val[:50]).font = name_font
            ws7.cell(row=row_idx, column=5, value=t.get("target_offer_id", "")).font = NORMAL_FONT
            ws7.cell(row=row_idx, column=6, value=t.get("buyers", 0)).number_format = NUM_FMT

            source_b = t.get("source_buyers", 1) or 1
            pct_val = t.get("buyers", 0) / source_b
            c_pct = ws7.cell(row=row_idx, column=7, value=pct_val)
            c_pct.number_format = PCT_FMT
            if is_same:
                c_pct.font = SAME_SKU_FONT

            for col in range(1, 8):
                ws7.cell(row=row_idx, column=col).border = THIN_BORDER

            row_idx += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════
# Ozon LTV Excel endpoint
# ══════════════════════════════════════════════════════════════

@router.get("/ozon/ltv/xlsx")
async def export_ozon_ltv_xlsx(
    shop_id: int = Query(...),
    period: str = Query("6m"),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export full Ozon LTV analysis as Excel file."""
    # Get main LTV data
    ltv_data = await get_ozon_ltv(
        shop_id=shop_id, period=period,
        date_from=date_from, date_to=date_to,
        db=db, current_user=current_user,
    )

    # Get shop name
    result = await db.execute(select(Shop).where(Shop.id == shop_id))
    shop = result.scalar_one_or_none()
    shop_name = shop.name if shop else f"Shop #{shop_id}"

    # Get per-SKU retention funnel (1→5 purchases) from ClickHouse
    start_date, end_date = _ltv_dates(period, date_from, date_to)
    from app.core.clickhouse import get_clickhouse_client

    try:
        ch = get_clickhouse_client()
        params = {"shop_id": shop_id, "start_date": start_date, "end_date": end_date}

        retention_rows = ch.query("""
            WITH
                /* All client-order-date combinations */
                all_orders AS (
                    SELECT
                        offer_id AS sku,
                        splitByChar('-', posting_number)[1] AS client_id,
                        toDate(addHours(in_process_at, 3)) AS order_date,
                        order_number
                    FROM mms_analytics.fact_ozon_orders FINAL
                    WHERE shop_id = {shop_id:UInt32}
                      AND toDate(addHours(in_process_at, 3)) >= {start_date:Date}
                      AND toDate(addHours(in_process_at, 3)) <= {end_date:Date}
                ),
                /* Unique client orders (deduplicate by order_number) */
                client_orders AS (
                    SELECT
                        client_id,
                        min(order_date) AS order_date
                    FROM all_orders
                    GROUP BY client_id, order_number
                ),
                /* For each SKU: which clients bought it and when first */
                sku_clients AS (
                    SELECT
                        sku AS target_sku,
                        client_id,
                        min(order_date) AS first_target_date
                    FROM all_orders
                    GROUP BY sku, client_id
                ),
                /* All client's orders from first target purchase onwards, numbered */
                client_chain AS (
                    SELECT
                        sc.target_sku,
                        sc.client_id,
                        co.order_date,
                        sc.first_target_date,
                        dense_rank() OVER (
                            PARTITION BY sc.target_sku, sc.client_id
                            ORDER BY co.order_date
                        ) AS purchase_num
                    FROM sku_clients sc
                    INNER JOIN client_orders co
                        ON co.client_id = sc.client_id
                       AND co.order_date >= sc.first_target_date
                ),
                /* Per target_sku + client: summary stats */
                client_summary AS (
                    SELECT
                        target_sku,
                        client_id,
                        max(purchase_num) AS total_purchases,
                        first_target_date,
                        max(order_date) AS last_order_date
                    FROM client_chain
                    GROUP BY target_sku, client_id, first_target_date
                ),
                /* SKU metadata */
                sku_meta AS (
                    SELECT
                        sku,
                        any(product_name) AS product_name,
                        any(toString(offer_id)) AS offer_id_str
                    FROM (
                        SELECT offer_id AS sku, product_name, offer_id
                        FROM mms_analytics.fact_ozon_orders FINAL
                        WHERE shop_id = {shop_id:UInt32}
                          AND toDate(addHours(in_process_at, 3)) >= {start_date:Date}
                          AND toDate(addHours(in_process_at, 3)) <= {end_date:Date}
                    )
                    GROUP BY sku
                )
            SELECT
                cs.target_sku AS sku,
                sm.offer_id_str AS offer_id,
                sm.product_name AS name,
                count() AS b1,
                countIf(total_purchases >= 2) AS b2,
                countIf(total_purchases >= 3) AS b3,
                countIf(total_purchases >= 4) AS b4,
                countIf(total_purchases >= 5) AS b5,
                round(avgIf(
                    dateDiff('day', first_target_date, last_order_date)
                        / greatest(total_purchases - 1, 1),
                    total_purchases >= 2
                ), 0) AS avg_days
            FROM client_summary cs
            LEFT JOIN sku_meta sm ON sm.sku = cs.target_sku
            GROUP BY cs.target_sku, sm.offer_id_str, sm.product_name
            HAVING b1 >= 3
            ORDER BY b2 DESC, b1 DESC
            LIMIT 100
        """, parameters=params).result_rows

        ch.close()

        sku_retention = []
        for r in retention_rows:
            sku_retention.append({
                "sku": str(r[0]),
                "offer_id": str(r[1] or r[0]),
                "name": str(r[2] or "")[:60],
                "buyers_1": int(r[3] or 0),
                "buyers_2": int(r[4] or 0),
                "buyers_3": int(r[5] or 0),
                "buyers_4": int(r[6] or 0),
                "buyers_5": int(r[7] or 0),
                "avg_days": int(_sf(r[8])),
            })

    except HTTPException:
        raise
    except Exception as e:
        logger.error("LTV XLSX retention query error: %s", e, exc_info=True)
        sku_retention = []

    # ── Cross-SKU transitions for top SKUs ──
    sku_transitions = []
    try:
        top_skus = sorted(sku_retention, key=lambda x: x["buyers_2"], reverse=True)[:15]
        top_sku_ids = [s["sku"] for s in top_skus if s["buyers_2"] > 0]

        if top_sku_ids:
            ch2 = get_clickhouse_client()
            chain_rows = ch2.query("""
                WITH
                    all_orders AS (
                        SELECT
                            splitByChar('-', posting_number)[1] AS client_id,
                            order_number,
                            sku,
                            argMax(offer_id, toDate(addHours(in_process_at, 3))) AS offer_id,
                            argMax(product_name, toDate(addHours(in_process_at, 3))) AS product_name,
                            min(toDate(addHours(in_process_at, 3))) AS order_date
                        FROM mms_analytics.fact_ozon_orders FINAL
                        WHERE shop_id = {shop_id:UInt32}
                          AND toDate(addHours(in_process_at, 3)) >= {start_date:Date}
                          AND toDate(addHours(in_process_at, 3)) <= {end_date:Date}
                        GROUP BY client_id, order_number, sku
                    ),
                    target_clients AS (
                        SELECT sku AS target_sku, client_id
                        FROM all_orders
                        WHERE sku IN {target_skus:Array(UInt64)}
                        GROUP BY sku, client_id
                    ),
                    client_numbered AS (
                        SELECT
                            tc.target_sku,
                            ao.client_id,
                            ao.sku,
                            ao.offer_id,
                            ao.product_name,
                            ao.order_date,
                            dense_rank() OVER (
                                PARTITION BY tc.target_sku, ao.client_id
                                ORDER BY ao.order_date, ao.order_number
                            ) AS purchase_num
                        FROM all_orders ao
                        JOIN target_clients tc ON ao.client_id = tc.client_id
                    ),
                    first_target AS (
                        SELECT target_sku, client_id, min(purchase_num) AS tpn
                        FROM client_numbered
                        WHERE sku = target_sku
                        GROUP BY target_sku, client_id
                    ),
                    reindexed AS (
                        SELECT
                            cn.target_sku,
                            cn.client_id,
                            cn.sku,
                            cn.offer_id,
                            cn.product_name,
                            cn.purchase_num - ft.tpn + 1 AS level
                        FROM client_numbered cn
                        JOIN first_target ft
                          ON cn.target_sku = ft.target_sku
                         AND cn.client_id = ft.client_id
                        WHERE cn.purchase_num >= ft.tpn
                          AND cn.purchase_num < ft.tpn + 5
                    )
                SELECT
                    target_sku,
                    level,
                    sku,
                    any(offer_id) AS offer_id,
                    any(product_name) AS name,
                    count(DISTINCT client_id) AS buyers
                FROM reindexed
                WHERE level BETWEEN 2 AND 5
                GROUP BY target_sku, level, sku
                HAVING buyers >= 1
                ORDER BY target_sku, level, buyers DESC
            """, parameters={
                **params,
                "target_skus": top_sku_ids,
            }).result_rows
            ch2.close()

            # Build sku_transitions list (top-3 per level per source)
            sku_info = {s["sku"]: s for s in sku_retention}
            from collections import defaultdict
            grouped = defaultdict(list)
            for r in chain_rows:
                grouped[(int(r[0]), int(r[1]))].append(r)

            for src in top_skus:
                s_sku = src["sku"]
                src_info = sku_info.get(s_sku, src)
                for lvl in range(2, 6):
                    rows = grouped.get((s_sku, lvl), [])
                    for r in rows[:3]:  # top-3 per level
                        sku_transitions.append({
                            "source_sku": s_sku,
                            "source_name": src_info.get("offer_id", str(s_sku)),
                            "source_buyers": src_info.get("buyers_1", 0),
                            "level": lvl,
                            "target_sku": int(r[2]),
                            "target_offer_id": str(r[3]),
                            "target_name": str(r[4])[:50],
                            "buyers": int(r[5]),
                            "is_same_sku": int(r[2]) == s_sku,
                        })
    except Exception as e:
        logger.error("LTV XLSX transitions query error: %s", e, exc_info=True)

    buf = _build_ltv_xlsx(ltv_data, sku_retention, shop_name, "Ozon", sku_transitions)
    filename = f"LTV_Ozon_{shop_name}_{period}.xlsx"
    from urllib.parse import quote
    encoded = quote(filename)
    from starlette.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )
