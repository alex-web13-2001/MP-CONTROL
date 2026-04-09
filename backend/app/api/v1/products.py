"""
Products API endpoints.

GET  /products/ozon?shop_id=X&page=1&per_page=25&sort=revenue_7d&order=desc&filter=all&search=
PATCH /products/ozon/cost  — update cost price for a product
POST /products/ozon/cost/bulk — bulk upload cost prices via Excel
"""
import io
import logging
from datetime import date, timedelta
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from pydantic import BaseModel
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.shop import Shop
from app.models.user import User

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/products", tags=["Products"])


# ── Request / Response schemas ────────────────────────────

class CostUpdateRequest(BaseModel):
    shop_id: int
    offer_id: str
    cost_price: float
    packaging_cost: float = 0


class CostUpdateResponse(BaseModel):
    ok: bool
    offer_id: str
    cost_price: float


# ── Ozon Products List ────────────────────────────────────

@router.get("/ozon")
async def get_ozon_products(
    shop_id: int = Query(...),
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=10, le=100),
    sort: str = Query("revenue_7d"),
    order: str = Query("desc"),
    filter: str = Query("all"),
    search: str = Query(""),
    period: int = Query(7, ge=1, le=366),
    date_from: Optional[date] = Query(None, description="Custom range start (YYYY-MM-DD)"),
    date_to: Optional[date] = Query(None, description="Custom range end (YYYY-MM-DD)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Ozon products list with aggregated analytics from 8+ data sources."""

    # Verify shop ownership
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "ozon":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Shop not found")

    from app.core.clickhouse import get_clickhouse_client

    try:
        ch = get_clickhouse_client()
    except Exception as e:
        logger.error("ClickHouse connection error: %s", e)
        raise HTTPException(status_code=500, detail="Analytics unavailable")

    today = date.today()
    if date_from and date_to:
        d_start = date_from
        d_end = date_to
    else:
        d_end = today
        d_start = today - timedelta(days=period - 1)
    span = (d_end - d_start).days + 1
    d_prev_start = d_start - timedelta(days=span)
    d_prev_end = d_start - timedelta(days=1)
    d30_start = today - timedelta(days=29)

    # ────────────────────────────────────────────────────
    # 1. Base catalog from PostgreSQL (dim_ozon_products)
    # ────────────────────────────────────────────────────
    pg_result = await db.execute(
        text("""
            SELECT p.product_id, p.offer_id, p.sku, p.name, p.barcode,
                   COALESCE(NULLIF(p.primary_image_url, ''), p.main_image_url, '') AS image_url,
                   p.price, p.old_price, p.min_price, p.marketing_price,
                   p.stocks_fbo, p.stocks_fbs,
                   p.price_index_color, p.price_index_value, p.competitor_min_price,
                   p.status, p.moderate_status, p.status_name,
                   p.is_archived, p.volume_weight, p.vat,
                   p.model_id, p.model_count,
                   COALESCE(c.images_count, 0) AS images_count,
                   COALESCE(c.title_hash, '') AS title_hash,
                   COALESCE(cost.cost_price, 0) AS cost_price,
                   COALESCE(cost.packaging_cost, 0) AS packaging_cost
            FROM dim_ozon_products p
            LEFT JOIN dim_ozon_product_content c
                ON c.shop_id = p.shop_id AND c.product_id = p.product_id
            LEFT JOIN product_costs cost
                ON cost.shop_id = p.shop_id AND cost.offer_id = p.offer_id
            WHERE p.shop_id = :shop_id
            ORDER BY p.name
        """),
        {"shop_id": shop_id},
    )
    rows = pg_result.fetchall()

    if not rows:
        return {
            "products": [],
            "total": 0,
            "page": page,
            "per_page": per_page,
            "cost_missing_count": 0,
        }

    # Build products dict keyed by offer_id
    products_map = {}
    all_offer_ids = []
    all_skus = []
    all_product_ids = []

    for r in rows:
        oid = r[1]  # offer_id
        products_map[oid] = {
            "product_id": r[0],
            "offer_id": oid,
            "sku": r[2],
            "name": r[3] or oid,
            "barcode": r[4],
            "image_url": r[5] or "",
            "price": float(r[6] or 0),
            "old_price": float(r[7] or 0),
            "min_price": float(r[8] or 0),
            "marketing_price": float(r[9] or 0),
            "stocks_fbo": r[10] or 0,
            "stocks_fbs": r[11] or 0,
            "price_index_color": r[12] or "",
            "price_index_value": float(r[13] or 0),
            "competitor_min_price": float(r[14] or 0),
            "status": r[15] or "",
            "moderate_status": r[16] or "",
            "status_name": r[17] or "",
            "is_archived": r[18] or False,
            "volume_weight": float(r[19] or 0),
            "vat": float(r[20] or 0),
            "model_id": r[21],
            "model_count": r[22] or 0,
            "images_count": r[23] or 0,
            "cost_price": float(r[25] or 0),
            "packaging_cost": float(r[26] or 0),
            # Will be filled from CH
            "orders_7d": 0,
            "revenue_7d": 0.0,
            "revenue_list": 0.0,
            "orders_prev_7d": 0,
            "revenue_delta": 0.0,
            "orders_30d": 0,
            "ad_spend_7d": 0.0,
            "drr": 0.0,
            "returns_30d": 0,
            "cancels": 0,
            "cancel_rate": 0.0,
            "content_rating": 0.0,
            "commission_percent": 0.0,
            "fbo_logistics": 0.0,
            "margin": None,
            "margin_percent": None,
            "payout_period": 0.0,
            "payout_prev": 0.0,
            "gross_profit": None,
            "gross_profit_percent": None,
            "gross_profit_prev": None,
            "gross_profit_delta": None,
            "mp_fees": 0.0,
            "mp_fees_percent": 0.0,
            "mp_fees_storage": 0.0,
            "mp_fees_other": 0.0,
            "avg_price": 0.0,
            "sales_amount": 0.0,
            "period": period,
            "events": [],
            "promotions": [],
        }
        all_offer_ids.append(oid)
        if r[2]:
            all_skus.append(r[2])
        all_product_ids.append(r[0])

    # ────────────────────────────────────────────────────
    # 2. Orders (period) + prev period from ClickHouse
    # ────────────────────────────────────────────────────
    try:
        orders_result = ch.query("""
            SELECT offer_id,
                   sumIf(quantity, order_date >= {d_start:Date} AND order_date <= {d_end:Date}) AS orders_period,
                   sumIf(price * quantity, order_date >= {d_start:Date} AND order_date <= {d_end:Date}) AS revenue_list,
                   sumIf(payout, order_date >= {d_start:Date} AND order_date <= {d_end:Date}) AS payout_period,
                   sumIf(quantity, order_date >= {d_prev_start:Date} AND order_date <= {d_prev_end:Date}) AS orders_prev,
                   sumIf(payout, order_date >= {d_prev_start:Date} AND order_date <= {d_prev_end:Date}) AS payout_prev,
                   sumIf(quantity, order_date >= {d_start:Date} AND order_date <= {d_end:Date} AND payout > 0) AS payout_qty
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= {d_prev_start:Date}
              AND status NOT IN ('cancelled', 'canceled')
            GROUP BY offer_id
        """, parameters={
            "shop_id": shop_id,
            "d_start": d_start,
            "d_end": d_end,
            "d_prev_start": d_prev_start,
            "d_prev_end": d_prev_end,
        })
        for r in orders_result.result_rows:
            oid = r[0]
            if oid in products_map:
                products_map[oid]["orders_7d"] = r[1]
                products_map[oid]["revenue_7d"] = float(r[2])       # price × qty (list price, like Ozon admin)
                products_map[oid]["payout_period"] = float(r[3])    # payout from orders (for profit calc)
                products_map[oid]["payout_prev"] = float(r[5])      # prev period payout
                # avg_price = payout per unit (real money received per piece)
                # NOTE: Ozon removed buyer price from Seller API (Nov 2025)
                # Use payout_qty (orders with payout > 0) to avoid dividing by undelivered orders
                payout_qty = r[6] if r[6] else 0
                if payout_qty > 0 and float(r[3]) > 0:
                    products_map[oid]["avg_price"] = round(float(r[3]) / payout_qty, 2)
                prev_orders = r[4]
                products_map[oid]["orders_prev_7d"] = prev_orders
                if prev_orders > 0:
                    products_map[oid]["revenue_delta"] = round(
                        (r[1] - prev_orders) / prev_orders * 100, 1
                    )
                elif r[1] > 0:
                    products_map[oid]["revenue_delta"] = 100.0
    except Exception as e:
        logger.warning("CH orders query failed: %s", e)

    # ────────────────────────────────────────────────────
    # 2b. Cancellations (period) from ClickHouse
    # ────────────────────────────────────────────────────
    try:
        cancels_result = ch.query("""
            SELECT offer_id,
                   sumIf(quantity, order_date >= {d_start:Date} AND order_date <= {d_end:Date}) AS cancels_period
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= {d_start:Date}
              AND order_date <= {d_end:Date}
              AND status IN ('cancelled', 'canceled')
            GROUP BY offer_id
        """, parameters={
            "shop_id": shop_id,
            "d_start": d_start,
            "d_end": d_end,
        })
        for r in cancels_result.result_rows:
            oid = r[0]
            if oid in products_map:
                cancels_val = int(r[1] or 0)
                products_map[oid]["cancels"] = cancels_val
                orders_val = products_map[oid]["orders_7d"]
                total_with_cancels = orders_val + cancels_val
                if total_with_cancels > 0:
                    products_map[oid]["cancel_rate"] = round(cancels_val / total_with_cancels * 100, 1)
    except Exception as e:
        logger.warning("CH cancels query failed: %s", e)

    # ────────────────────────────────────────────────────
    # 3. Ads 7d from ClickHouse (keyed by SKU, not offer_id)
    # ────────────────────────────────────────────────────
    # Build sku → offer_id mapping for ads lookup
    sku_to_offer = {}
    for oid, p in products_map.items():
        if p["sku"]:
            sku_to_offer[p["sku"]] = oid

    try:
        ads_result = ch.query("""
            SELECT sku,
                   sum(money_spent) AS ad_spend,
                   sum(views) AS views,
                   sum(clicks) AS clicks,
                   sum(orders) AS ad_orders,
                   sum(revenue) AS ad_revenue
            FROM mms_analytics.fact_ozon_ad_daily FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND dt >= {d_start:Date}
              AND dt <= {d_end:Date}
            GROUP BY sku
        """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": d_end})
        for r in ads_result.result_rows:
            oid = sku_to_offer.get(r[0])
            if oid and oid in products_map:
                products_map[oid]["ad_spend_7d"] = float(r[1])
                rev = products_map[oid]["revenue_7d"]
                if rev > 0:
                    products_map[oid]["drr"] = round(float(r[1]) / rev * 100, 1)
    except Exception as e:
        logger.warning("CH ads query failed: %s", e)

    # ────────────────────────────────────────────────────
    # 3b. Active ad campaigns per SKU (Ozon)
    # ────────────────────────────────────────────────────
    # Ozon has no dim_campaigns table — use log_ozon_bids presence
    # as a proxy: if bids were logged in the last 2 days, campaign is active.
    active_ad_skus: set[int] = set()
    try:
        active_ads_result = ch.query("""
            SELECT DISTINCT sku
            FROM mms_analytics.log_ozon_bids
            WHERE shop_id = {shop_id:UInt32}
              AND timestamp >= now() - INTERVAL 2 DAY
        """, parameters={"shop_id": shop_id})
        for r in active_ads_result.result_rows:
            active_ad_skus.add(int(r[0]))
    except Exception as e:
        logger.warning("CH active Ozon ads query failed: %s", e)

    # ────────────────────────────────────────────────────
    # 4. Returns 30d from ClickHouse
    # ────────────────────────────────────────────────────
    try:
        returns_result = ch.query("""
            SELECT offer_id,
                   sum(quantity) AS returns_count
            FROM mms_analytics.fact_ozon_returns FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND dt >= {d30_start:Date}
            GROUP BY offer_id
        """, parameters={"shop_id": shop_id, "d30_start": d30_start})
        for r in returns_result.result_rows:
            if r[0] in products_map:
                products_map[r[0]]["returns_30d"] = r[1]
    except Exception as e:
        logger.warning("CH returns query failed: %s", e)

    # Also get 30d orders for return rate calculation
    try:
        orders30_result = ch.query("""
            SELECT offer_id,
                   sum(quantity) AS orders_30d
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= {d30_start:Date}
              AND status NOT IN ('cancelled', 'canceled')
            GROUP BY offer_id
        """, parameters={"shop_id": shop_id, "d30_start": d30_start})
        for r in orders30_result.result_rows:
            if r[0] in products_map:
                products_map[r[0]]["orders_30d"] = r[1]
    except Exception as e:
        logger.warning("CH orders_30d query failed: %s", e)

    # ────────────────────────────────────────────────────
    # 5. Commissions (latest) from ClickHouse
    # ────────────────────────────────────────────────────
    try:
        comm_result = ch.query("""
            SELECT offer_id,
                   argMax(sales_percent, dt) AS sales_pct,
                   argMax(fbo_fulfillment_amount, dt) AS fbo_logistics
            FROM mms_analytics.fact_ozon_commissions FINAL
            WHERE shop_id = {shop_id:UInt32}
            GROUP BY offer_id
        """, parameters={"shop_id": shop_id})
        for r in comm_result.result_rows:
            if r[0] in products_map:
                products_map[r[0]]["commission_percent"] = float(r[1])
                products_map[r[0]]["fbo_logistics"] = float(r[2])
    except Exception as e:
        logger.warning("CH commissions query failed: %s", e)

    # ────────────────────────────────────────────────────
    # 6. Content rating (latest) from ClickHouse
    # ────────────────────────────────────────────────────
    try:
        rating_result = ch.query("""
            SELECT sku,
                   argMax(rating, dt) AS rating
            FROM mms_analytics.fact_ozon_content_rating FINAL
            WHERE shop_id = {shop_id:UInt32}
            GROUP BY sku
        """, parameters={"shop_id": shop_id})
        # Map sku → product (already built above for ads)
        for r in rating_result.result_rows:
            oid = sku_to_offer.get(r[0])
            if oid:
                products_map[oid]["content_rating"] = float(r[1])
    except Exception as e:
        logger.warning("CH content rating query failed: %s", e)

    # ────────────────────────────────────────────────────
    # 7. Active promotions from ClickHouse
    # ────────────────────────────────────────────────────
    try:
        promo_result = ch.query("""
            SELECT product_id, promo_type
            FROM mms_analytics.fact_ozon_promotions FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND is_enabled = 1
              AND dt >= {d30_start:Date}
        """, parameters={"shop_id": shop_id, "d30_start": d30_start})
        pid_to_offer = {}
        for oid, p in products_map.items():
            pid_to_offer[p["product_id"]] = oid
        for r in promo_result.result_rows:
            oid = pid_to_offer.get(r[0])
            if oid:
                products_map[oid]["promotions"].append(r[1])
    except Exception as e:
        logger.warning("CH promotions query failed: %s", e)

    # ────────────────────────────────────────────────────
    # 8. Events from PostgreSQL event_log (last 30 days)
    # ────────────────────────────────────────────────────
    try:
        events_result = await db.execute(
            text("""
                SELECT nm_id, event_type, old_value, new_value, created_at
                FROM event_log
                WHERE shop_id = :shop_id
                  AND created_at >= :since
                ORDER BY created_at DESC
            """),
            {"shop_id": shop_id, "since": today - timedelta(days=30)},
        )
        for ev in events_result:
            # nm_id maps to product_id for Ozon events
            pid = ev[0]
            oid = pid_to_offer.get(pid) if pid else None
            if oid and oid in products_map:
                products_map[oid]["events"].append({
                    "type": ev[1],
                    "old_value": ev[2],
                    "new_value": ev[3],
                    "date": ev[4].isoformat() if ev[4] else None,
                })
    except Exception as e:
        logger.warning("PG events query failed: %s", e)

    # ────────────────────────────────────────────────────
    # 9. Price changes from ClickHouse (last 30d)
    # ────────────────────────────────────────────────────
    try:
        price_result = ch.query("""
            SELECT offer_id,
                   groupArray(price) AS prices,
                   groupArray(dt) AS dates
            FROM (
                SELECT offer_id, price, dt
                FROM mms_analytics.fact_ozon_prices FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND dt >= {d30_start:Date}
                ORDER BY dt
            )
            GROUP BY offer_id
        """, parameters={"shop_id": shop_id, "d30_start": d30_start})
        for r in price_result.result_rows:
            oid = r[0]
            if oid in products_map and len(r[1]) >= 2:
                prices = r[1]
                dates = r[2]
                for i in range(1, len(prices)):
                    if prices[i] != prices[i - 1]:
                        direction = "PRICE_UP" if prices[i] > prices[i - 1] else "PRICE_DOWN"
                        products_map[oid]["events"].append({
                            "type": direction,
                            "old_value": str(prices[i - 1]),
                            "new_value": str(prices[i]),
                            "date": str(dates[i]),
                        })
    except Exception as e:
        logger.warning("CH price changes query failed: %s", e)

    # ────────────────────────────────────────────────────
    # 10. Detailed financials from fact_ozon_transactions
    #     Uses accruals_for_sale (REAL seller revenue) and
    #     sale_commission / services_total for fee breakdown.
    #     revenue_7d (price × qty) is LIST PRICE, not seller revenue!
    # ────────────────────────────────────────────────────
    try:
        txn_result = ch.query("""
            SELECT sku,
                   sumIf(accruals_for_sale, toDate(operation_date) >= {d_start:Date} AND toDate(operation_date) <= {d_end:Date}) AS rev_cur,
                   sumIf(abs(sale_commission), toDate(operation_date) >= {d_start:Date} AND toDate(operation_date) <= {d_end:Date}) AS comm_cur,
                   sumIf(abs(services_total), toDate(operation_date) >= {d_start:Date} AND toDate(operation_date) <= {d_end:Date}) AS svc_cur,
                   sumIf(amount, toDate(operation_date) >= {d_start:Date} AND toDate(operation_date) <= {d_end:Date}) AS payout_cur,
                   sumIf(accruals_for_sale, toDate(operation_date) >= {d_prev_start:Date} AND toDate(operation_date) <= {d_prev_end:Date}) AS rev_prev,
                   sumIf(abs(sale_commission), toDate(operation_date) >= {d_prev_start:Date} AND toDate(operation_date) <= {d_prev_end:Date}) AS comm_prev,
                   sumIf(abs(services_total), toDate(operation_date) >= {d_prev_start:Date} AND toDate(operation_date) <= {d_prev_end:Date}) AS svc_prev,
                   sumIf(amount, toDate(operation_date) >= {d_prev_start:Date} AND toDate(operation_date) <= {d_prev_end:Date}) AS payout_prev
            FROM mms_analytics.fact_ozon_transactions FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND category = 'Revenue'
              AND sku > 0
              AND toDate(operation_date) >= {d_prev_start:Date}
              AND toDate(operation_date) <= {d_end:Date}
            GROUP BY sku
        """, parameters={
            "shop_id": shop_id, "d_start": d_start, "d_end": d_end,
            "d_prev_start": d_prev_start, "d_prev_end": d_prev_end,
        })
        for r in txn_result.result_rows:
            oid = sku_to_offer.get(r[0])
            if oid and oid in products_map:
                products_map[oid]["txn_revenue"] = float(r[1])        # accruals_for_sale (real revenue)
                products_map[oid]["txn_commission"] = float(r[2])     # sale_commission (abs)
                products_map[oid]["txn_logistics"] = float(r[3])      # services_total (abs) — logistics+processing
                products_map[oid]["txn_payout"] = float(r[4])         # net amount
                products_map[oid]["txn_revenue_prev"] = float(r[5])
                products_map[oid]["txn_commission_prev"] = float(r[6])
                products_map[oid]["txn_logistics_prev"] = float(r[7])
                products_map[oid]["txn_payout_prev"] = float(r[8])
    except Exception as e:
        logger.warning("CH transactions query failed: %s", e)

    # ── 10b. Bulk charges (Acquiring, Storage) — not per-product in Ozon ──
    # Distribute proportionally by txn_revenue share
    try:
        bulk_result = ch.query("""
            SELECT
                category,
                sumIf(abs(amount), toDate(operation_date) >= {d_start:Date} AND toDate(operation_date) <= {d_end:Date}) AS total_cur,
                sumIf(abs(amount), toDate(operation_date) >= {d_prev_start:Date} AND toDate(operation_date) <= {d_prev_end:Date}) AS total_prev
            FROM mms_analytics.fact_ozon_transactions FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND toDate(operation_date) >= {d_prev_start:Date}
              AND toDate(operation_date) <= {d_end:Date}
              AND category IN ('Acquiring', 'Storage')
            GROUP BY category
        """, parameters={
            "shop_id": shop_id, "d_start": d_start, "d_end": d_end,
            "d_prev_start": d_prev_start, "d_prev_end": d_prev_end,
        })
        bulk_cur = {}
        bulk_prev = {}
        for r in bulk_result.result_rows:
            bulk_cur[r[0]] = float(r[1] or 0)
            bulk_prev[r[0]] = float(r[2] or 0)

        total_txn_rev = sum(p.get("txn_revenue", 0) for p in products_map.values())
        total_txn_rev_prev = sum(p.get("txn_revenue_prev", 0) for p in products_map.values())

        for oid, p in products_map.items():
            for cat in ("Acquiring", "Storage"):
                key = cat.lower()
                rev = p.get("txn_revenue", 0)
                rev_p = p.get("txn_revenue_prev", 0)
                if cat in bulk_cur and total_txn_rev > 0 and rev > 0:
                    share = rev / total_txn_rev
                    p.setdefault("txn_" + key, 0)
                    p["txn_" + key] = round(bulk_cur[cat] * share, 2)
                if cat in bulk_prev and total_txn_rev_prev > 0 and rev_p > 0:
                    share = rev_p / total_txn_rev_prev
                    p.setdefault("txn_" + key + "_prev", 0)
                    p["txn_" + key + "_prev"] = round(bulk_prev[cat] * share, 2)
    except Exception as e:
        logger.warning("CH bulk charges query failed: %s", e)

    # ────────────────────────────────────────────────────
    # Calculate margin & gross profit
    # Uses REAL revenue from transactions (accruals_for_sale),
    # NOT the inflated list price (price × quantity).
    # ────────────────────────────────────────────────────
    for p in products_map.values():
        cost = p["cost_price"] + p["packaging_cost"]
        if cost > 0 and p["price"] > 0:
            p["margin"] = round(cost, 2)
            p["margin_percent"] = round(cost / p["price"] * 100, 1)

        # ── Real revenue from transactions ──
        txn_revenue = p.get("txn_revenue", 0)
        txn_commission = p.get("txn_commission", 0)
        txn_logistics = p.get("txn_logistics", 0)
        txn_acquiring = p.get("txn_acquiring", 0)
        txn_storage = p.get("txn_storage", 0)

        # ── Marketplace fees = commission + logistics + acquiring + storage ──
        if txn_revenue > 0:
            total_fees = txn_commission + txn_logistics + txn_acquiring + txn_storage
            p["mp_fees"] = round(total_fees, 2)
            p["mp_fees_commission"] = round(txn_commission, 2)
            p["mp_fees_logistics"] = round(txn_logistics + txn_acquiring + txn_storage, 2)
            p["mp_fees_percent"] = round(total_fees / txn_revenue * 100, 1)

            # Override revenue_7d with REAL revenue for display
            p["revenue_7d"] = round(txn_revenue, 2)

        # ── Gross profit = real_revenue - COGS - total_fees - ad_spend ──
        if cost > 0 and p["orders_7d"] > 0 and txn_revenue > 0:
            cogs = cost * p["orders_7d"]
            total_fees_val = p.get("mp_fees", 0)
            gp = txn_revenue - cogs - total_fees_val - p["ad_spend_7d"]
            p["gross_profit"] = round(gp, 2)
            if txn_revenue > 0:
                p["gross_profit_percent"] = round(gp / txn_revenue * 100, 1)
            # Prev-period gross profit for delta
            txn_revenue_prev = p.get("txn_revenue_prev", 0)
            if p.get("orders_prev_7d", 0) > 0 and txn_revenue_prev > 0:
                cogs_prev = cost * p["orders_prev_7d"]
                comm_prev = p.get("txn_commission_prev", 0)
                log_prev = p.get("txn_logistics_prev", 0)
                acq_prev = p.get("txn_acquiring_prev", 0)
                sto_prev = p.get("txn_storage_prev", 0)
                fees_prev = comm_prev + log_prev + acq_prev + sto_prev
                gp_prev = txn_revenue_prev - cogs_prev - fees_prev - p.get("ad_spend_prev", 0)
                p["gross_profit_prev"] = round(gp_prev, 2)
                if gp_prev != 0:
                    p["gross_profit_delta"] = round((gp - gp_prev) / abs(gp_prev) * 100, 1)
                elif gp > 0:
                    p["gross_profit_delta"] = 100.0

        # Recalculate DRR with real revenue (revenue_7d now = txn_revenue)
        if p["ad_spend_7d"] > 0 and p["revenue_7d"] > 0:
            p["drr"] = round(p["ad_spend_7d"] / p["revenue_7d"] * 100, 1)

        # sales_amount = avg_price × orders (payout-based total)
        if p["avg_price"] > 0 and p["orders_7d"] > 0:
            p["sales_amount"] = round(p["avg_price"] * p["orders_7d"], 2)

        # Active ad campaigns indicator
        p["has_active_ads"] = (p.get("sku") or 0) in active_ad_skus

    # ────────────────────────────────────────────────────
    # Apply filter
    # ────────────────────────────────────────────────────
    products_list = list(products_map.values())

    if filter == "in_stock":
        products_list = [p for p in products_list if p["stocks_fbo"] + p["stocks_fbs"] > 0]
    elif filter == "no_stock":
        products_list = [p for p in products_list if p["stocks_fbo"] + p["stocks_fbs"] == 0 and not p["is_archived"]]
    elif filter == "with_ads":
        products_list = [p for p in products_list if p["ad_spend_7d"] > 0]
    elif filter == "no_ads":
        products_list = [p for p in products_list if p["ad_spend_7d"] == 0 and not p["is_archived"]]
    elif filter == "problems":
        products_list = [p for p in products_list if (
            p["drr"] > 20 or
            (p["stocks_fbo"] + p["stocks_fbs"] == 0 and not p["is_archived"]) or
            p["price_index_color"] in ("NON_PROFIT",)
        )]
    elif filter == "archived":
        products_list = [p for p in products_list if p["is_archived"]]

    # Apply search
    if search:
        q = search.lower()
        products_list = [
            p for p in products_list
            if q in (p["name"] or "").lower()
            or q in (p["offer_id"] or "").lower()
            or q in str(p["sku"] or "")
            or q in (p["barcode"] or "").lower()
        ]

    # Sort — push null/zero to the end regardless of direction
    # Use (primary_key, offer_id) tuple for STABLE pagination across pages
    _DESC_NULL = float('-inf')  # for desc: nulls go last (smallest)
    _ASC_NULL = float('inf')   # for asc: nulls go last (largest)
    _null = _DESC_NULL if order == "desc" else _ASC_NULL

    sort_key_map = {
        "revenue_7d": lambda p: (p["revenue_7d"] if p["revenue_7d"] else _null, p["offer_id"]),
        "orders_7d": lambda p: (p["orders_7d"] if p["orders_7d"] else _null, p["offer_id"]),
        "stocks": lambda p: ((p["stocks_fbo"] + p["stocks_fbs"]) or _null, p["offer_id"]),
        "price": lambda p: (p["price"] if p["price"] else _null, p["offer_id"]),
        "margin": lambda p: (p["margin"] if p["margin"] is not None else _null, p["offer_id"]),
        "gross_profit": lambda p: (p["gross_profit"] if p["gross_profit"] is not None else _null, p["offer_id"]),
        "drr": lambda p: (p["drr"] if p["drr"] else _null, p["offer_id"]),
        "returns": lambda p: (p["returns_30d"] if p["returns_30d"] else _null, p["offer_id"]),
        "name": lambda p: ((p["name"] or "").lower(), p["offer_id"]),
        "content_rating": lambda p: (p["content_rating"] if p["content_rating"] else _null, p["offer_id"]),
    }
    sort_fn = sort_key_map.get(sort, sort_key_map["revenue_7d"])
    products_list.sort(key=sort_fn, reverse=(order == "desc"))

    # Count cost missing
    cost_missing = sum(1 for p in products_map.values() if p["cost_price"] == 0 and not p["is_archived"])

    # ── Compute totals across ALL filtered products (before pagination) ──
    t_stocks = 0
    t_orders = 0
    t_revenue = 0.0
    t_payout = 0.0
    t_ad_spend = 0.0
    t_mp_fees = 0.0
    t_mp_fees_commission = 0.0
    t_mp_fees_logistics = 0.0
    t_total_cogs = 0.0
    t_profit = 0.0
    t_profit_count = 0
    t_returns = 0
    t_orders_30d = 0
    t_cancels = 0
    for p in products_list:
        t_stocks += p["stocks_fbo"] + p["stocks_fbs"]
        t_orders += p["orders_7d"]
        t_revenue += p["revenue_7d"]
        t_payout += p.get("payout_period", 0)
        t_ad_spend += p["ad_spend_7d"]
        t_mp_fees += p["mp_fees"]
        t_mp_fees_commission += p.get("mp_fees_commission", 0)
        t_mp_fees_logistics += p.get("mp_fees_logistics", 0)
        t_returns += p["returns_30d"]
        t_orders_30d += p.get("orders_30d", 0)
        t_cancels += p.get("cancels", 0)
        # COGS для итого-строки
        if p["cost_price"] > 0 and p["orders_7d"] > 0:
            t_total_cogs += (p["cost_price"] + p.get("packaging_cost", 0)) * p["orders_7d"]
        if p["gross_profit"] is not None:
            t_profit += p["gross_profit"]
            t_profit_count += 1

    totals = {
        "count": len(products_list),
        "stocks": t_stocks,
        "orders": t_orders,
        "revenue": round(t_revenue, 2),
        "payout": round(t_payout, 2),
        "avg_price": round(t_payout / t_orders, 2) if t_orders > 0 and t_payout > 0 else 0,
        "total_cogs": round(t_total_cogs, 2),
        "ad_spend": round(t_ad_spend, 2),
        "drr": round(t_ad_spend / t_revenue * 100, 1) if t_revenue > 0 else 0,
        "returns_pct": round(t_returns / t_orders_30d * 100, 1) if t_orders_30d > 0 else 0,
        "cancels": t_cancels,
        "cancel_rate": round(t_cancels / (t_orders + t_cancels) * 100, 1) if (t_orders + t_cancels) > 0 else 0,
        "mp_fees": round(t_mp_fees, 2),
        "mp_fees_commission": round(t_mp_fees_commission, 2),
        "mp_fees_logistics": round(t_mp_fees_logistics, 2),
        "mp_fees_pct": round(t_mp_fees / t_revenue * 100, 1) if t_revenue > 0 else 0,
        "profit": round(t_profit, 2),
        "profit_pct": round(t_profit / t_revenue * 100, 1) if t_revenue > 0 and t_profit_count > 0 else 0,
        "profit_count": t_profit_count,
    }

    # Paginate
    total = len(products_list)
    start = (page - 1) * per_page
    end = start + per_page
    page_items = products_list[start:end]

    # Trim events to last 5 per product
    for p in page_items:
        p["events"] = sorted(p["events"], key=lambda e: e.get("date", ""), reverse=True)[:5]

    return {
        "products": page_items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "cost_missing_count": cost_missing,
        "period": period,
        "totals": totals,
    }


# ── Ozon Prices (for Prices page) ────────────────────────

@router.get("/ozon/prices")
async def get_ozon_prices(
    shop_id: int = Query(...),
    search: str = Query(""),
    sort: str = Query("name"),
    order: str = Query("asc"),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=10, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Ozon product prices with profit calculations for the Prices page.
    
    Similar to WB /products/wb/prices, returns:
    - Catalog data (prices, images, names)
    - Stocks (FBO/FBS)
    - Profit per unit (from fact_ozon_transactions, 30d average)
    - Ad spend, DRR, profit_with_ads (from fact_ozon_ad_daily + fact_ozon_orders)
    - Price index data
    """
    # Verify shop ownership
    shop_result = await db.execute(
        select(Shop).where(
            Shop.id == shop_id,
            Shop.user_id == current_user.id,
            Shop.marketplace == "ozon",
        )
    )
    shop = shop_result.scalar_one_or_none()
    if not shop:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ozon магазин не найден")

    # ── 1. Catalog from PostgreSQL ──
    pg_result = await db.execute(
        text("""
            SELECT p.product_id, p.offer_id, p.sku, p.name,
                   COALESCE(NULLIF(p.primary_image_url, ''), p.main_image_url, '') AS image_url,
                   p.price, p.old_price, p.min_price, p.marketing_price,
                   p.stocks_fbo, p.stocks_fbs,
                   p.price_index_color, p.price_index_value, p.competitor_min_price,
                   p.is_archived,
                   COALESCE(cost.cost_price, 0) AS cost_price,
                   COALESCE(cost.packaging_cost, 0) AS packaging_cost
            FROM dim_ozon_products p
            LEFT JOIN product_costs cost
                ON cost.shop_id = p.shop_id AND cost.offer_id = p.offer_id
            WHERE p.shop_id = :shop_id
              AND NOT COALESCE(p.is_archived, false)
            ORDER BY p.name
        """),
        {"shop_id": shop_id},
    )
    rows = pg_result.fetchall()

    if not rows:
        return {
            "products": [],
            "total": 0,
            "page": page,
            "per_page": per_page,
            "cost_missing_count": 0,
        }

    # Build products list + lookup maps
    products_map = {}
    sku_to_offer = {}
    for r in rows:
        oid = r[1]  # offer_id
        sku = r[2]
        products_map[oid] = {
            "product_id": r[0],
            "offer_id": oid,
            "sku": sku,
            "name": r[3] or oid,
            "image_url": r[4] or "",
            # Ozon price fields:
            # price = base seller price (before marketplace discounts)
            # old_price = "strike-through" price (highest/original)
            # marketing_price = real buyer-facing price (after all discounts)
            # min_price = minimum allowed price
            "price": float(r[5] or 0),        # base seller price
            "old_price": float(r[6] or 0),     # original/old price
            "min_price": float(r[7] or 0),     # minimum allowed
            "marketing_price": float(r[8] or 0),  # real buyer price
            "stocks_fbo": r[9] or 0,
            "stocks_fbs": r[10] or 0,
            "price_index_color": r[11] or "",
            "price_index_value": float(r[12] or 0),
            "competitor_min_price": float(r[13] or 0),
            "cost_price": float(r[15] or 0),
            "packaging_cost": float(r[16] or 0),
            # Will be filled from CH
            "profit_per_unit": None,
            "profit_source": None,
            "profit_with_ads": None,
            "ad_spend_30d": 0,
            "drr": None,
        }
        if sku:
            sku_to_offer[sku] = oid

    # ── 2. Profit per unit from fact_ozon_transactions (7d) ──
    # Uses last 7 days instead of 30 to reflect recent pricing changes accurately
    from app.core.clickhouse import get_clickhouse_client
    try:
        ch = get_clickhouse_client()
    except Exception as e:
        logger.error("ClickHouse connection error: %s", e)
        raise HTTPException(status_code=500, detail="Analytics unavailable")

    profit_map: dict = {}
    try:
        d7_start = date.today() - timedelta(days=6)
        txn_result = ch.query("""
            SELECT sku,
                   sum(accruals_for_sale) / nullIf(
                       sumIf(1, category = 'Revenue' AND accruals_for_sale > 0), 0
                   ) AS avg_revenue_per_unit,
                   sum(abs(sale_commission)) / nullIf(
                       sumIf(1, category = 'Revenue' AND accruals_for_sale > 0), 0
                   ) AS avg_commission_per_unit,
                   sum(abs(services_total)) / nullIf(
                       sumIf(1, category = 'Revenue' AND accruals_for_sale > 0), 0
                   ) AS avg_logistics_per_unit,
                   sumIf(1, category = 'Revenue' AND accruals_for_sale > 0) AS sales_count
            FROM mms_analytics.fact_ozon_transactions FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND category = 'Revenue'
              AND sku > 0
              AND toDate(operation_date) >= {d7_start:Date}
            GROUP BY sku
            HAVING sumIf(1, category = 'Revenue' AND accruals_for_sale > 0) > 0
        """, parameters={"shop_id": shop_id, "d7_start": d7_start})
        for r in txn_result.result_rows:
            oid = sku_to_offer.get(r[0])
            if oid:
                profit_map[oid] = {
                    "avg_revenue": float(r[1] or 0),
                    "avg_commission": float(r[2] or 0),
                    "avg_logistics": float(r[3] or 0),
                    "sales_count": int(r[4] or 0),
                }
    except Exception as e:
        logger.warning("CH profit query failed: %s", e)

    # ── 3. Ad spend per SKU (7d) + orders/revenue for DRR ──
    # Uses last 7 days to reflect current DRR after price changes
    ad_spend_map: dict[int, float] = {}
    orders_revenue_map: dict[int, tuple] = {}  # offer_id → (revenue, orders)
    try:
        d7_start = date.today() - timedelta(days=6)
        ad_result = ch.query("""
            SELECT sku,
                   round(sum(money_spent), 2) AS ad_spend_7d
            FROM mms_analytics.fact_ozon_ad_daily FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND dt >= {d7_start:Date}
            GROUP BY sku
        """, parameters={"shop_id": shop_id, "d7_start": d7_start})
        for r in ad_result.result_rows:
            ad_spend_map[r[0]] = float(r[1])

        # Revenue for DRR (7d)
        rev_result = ch.query("""
            SELECT offer_id,
                   round(sum(price * quantity), 2) AS revenue_7d,
                   sum(quantity) AS orders_7d
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= {d7_start:Date}
              AND status NOT IN ('cancelled', 'canceled')
            GROUP BY offer_id
        """, parameters={"shop_id": shop_id, "d7_start": d7_start})
        for r in rev_result.result_rows:
            if r[0] in products_map:
                orders_revenue_map[r[0]] = (float(r[1]), int(r[2]))
    except Exception as e:
        logger.warning("CH ad_spend/revenue query failed: %s", e)

    # ── 3b. Last sale date per product ──
    last_sale_map: dict[str, str] = {}  # offer_id → ISO date string
    try:
        last_sale_result = ch.query("""
            SELECT
                offer_id,
                max(order_date) AS last_sale_date
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND status NOT IN ('cancelled', 'canceled')
            GROUP BY offer_id
        """, parameters={"shop_id": shop_id})
        for r in last_sale_result.result_rows:
            last_sale_map[r[0]] = str(r[1])
    except Exception as e:
        logger.warning("CH last_sale query failed: %s", e)

    ch.close()

    # ── 4. Calculate profit & merge ──
    products = []
    for oid, p in products_map.items():
        cost = p["cost_price"] + p["packaging_cost"]
        marketing_price = p["marketing_price"]
        base_price = p["price"]

        # Profit per unit (without ads):
        # Option 1: from transactions (accurate)
        # Option 2: fallback — estimated fees ~30% of marketing_price
        profit_per_unit = None
        profit_source = None

        if cost > 0 and oid in profit_map:
            pm = profit_map[oid]
            # avg_revenue - avg_commission - avg_logistics - cost
            profit_per_unit = round(pm["avg_revenue"] - pm["avg_commission"] - pm["avg_logistics"] - cost, 2)
            profit_source = "finance"
        elif cost > 0 and marketing_price > 0:
            estimated_fees = round(marketing_price * 0.30, 2)
            profit_per_unit = round(marketing_price - cost - estimated_fees, 2)
            profit_source = "estimated"

        # Ad spend & DRR (7d — reflects recent price changes)
        sku = p.get("sku")
        ad_spend_7d = ad_spend_map.get(sku, 0) if sku else 0
        rev_data = orders_revenue_map.get(oid)
        revenue_7d = rev_data[0] if rev_data else 0
        orders_7d = rev_data[1] if rev_data else 0
        drr = round(ad_spend_7d / revenue_7d * 100, 1) if revenue_7d > 0 and ad_spend_7d > 0 else None

        # Profit with ads = profit_per_unit - (ad_spend / orders)
        ad_per_unit = round(ad_spend_7d / orders_7d, 2) if orders_7d > 0 and ad_spend_7d > 0 else 0
        profit_with_ads = round(profit_per_unit - ad_per_unit, 2) if profit_per_unit is not None and ad_per_unit > 0 else None

        # Discount calculation:
        # marketing_price = real buyer price (after all discounts)
        # price = base seller price
        # old_price = original "crossed-out" price
        discount_base = p["old_price"] if p["old_price"] > 0 else base_price
        buyer_price = marketing_price if marketing_price > 0 else base_price
        discount_pct = round((1 - buyer_price / discount_base) * 100) if discount_base > 0 and buyer_price < discount_base else 0

        products.append({
            "product_id": p["product_id"],
            "offer_id": oid,
            "sku": p["sku"],
            "name": p["name"],
            "image_url": p["image_url"],
            # Price fields — correctly mapped for UI
            "price": base_price,               # base seller price
            "old_price": p["old_price"],         # original/strike-through
            "marketing_price": marketing_price,  # real buyer price
            "min_price": p["min_price"],
            "discount_pct": discount_pct,
            # Price index
            "price_index_color": p["price_index_color"],
            "price_index_value": p["price_index_value"],
            "competitor_min_price": p["competitor_min_price"],
            # Cost
            "cost_price": p["cost_price"],
            "packaging_cost": p["packaging_cost"],
            # Profit
            "profit_per_unit": profit_per_unit,
            "profit_source": profit_source,
            "profit_with_ads": profit_with_ads,
            # Ads (7d window for up-to-date DRR)
            "ad_spend_30d": ad_spend_7d,
            "drr": drr,
            # Stocks
            "stock_fbo": p["stocks_fbo"],
            "stock_fbs": p["stocks_fbs"],
            # Last sale
            "last_sale_date": last_sale_map.get(oid),
        })

    # ── 5. Search ──
    if search:
        s = search.lower()
        products = [
            p for p in products
            if s in p["name"].lower()
            or s in p["offer_id"].lower()
            or s in str(p.get("sku") or "")
        ]

    # ── 6. Sort ──
    SORT_FIELDS = {
        "name", "price", "marketing_price", "old_price",
        "cost_price", "profit_per_unit",
        "stock_fbo", "stock_fbs", "last_sale_date",
    }
    sort_key = sort if sort in SORT_FIELDS else "name"
    reverse = (order == "desc")

    def sort_val(p):
        v = p.get(sort_key)
        if v is None:
            return float("-inf") if reverse else float("inf")
        if isinstance(v, str):
            return v.lower()
        return v

    products.sort(key=sort_val, reverse=reverse)

    # ── 7. Totals ──
    cost_missing = sum(1 for p in products if p["cost_price"] == 0)

    # ── 8. Paginate ──
    total = len(products)
    offset = (page - 1) * per_page
    page_products = products[offset: offset + per_page]

    return {
        "products": page_products,
        "total": total,
        "page": page,
        "per_page": per_page,
        "cost_missing_count": cost_missing,
    }


# ── Update Cost Price ─────────────────────────────────────

@router.patch("/ozon/cost")
async def update_ozon_cost(
    body: CostUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update cost price for an Ozon product."""

    # Verify shop ownership
    shop_result = await db.execute(
        select(Shop).where(Shop.id == body.shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop:
        raise HTTPException(status_code=404, detail="Shop not found")

    # Trim offer_id to prevent space duplicates
    offer_id = body.offer_id.strip()

    # Upsert cost
    await db.execute(
        text("""
            INSERT INTO product_costs (shop_id, offer_id, cost_price, packaging_cost)
            VALUES (:shop_id, :offer_id, :cost_price, :packaging_cost)
            ON CONFLICT (shop_id, offer_id) DO UPDATE SET
                cost_price = EXCLUDED.cost_price,
                packaging_cost = EXCLUDED.packaging_cost,
                updated_at = NOW()
        """),
        {
            "shop_id": body.shop_id,
            "offer_id": offer_id,
            "cost_price": body.cost_price,
            "packaging_cost": body.packaging_cost,
        },
    )
    await db.commit()

    # Check if cost > price for this product → warn
    warning = None
    price_result = await db.execute(
        text("SELECT price FROM dim_ozon_products WHERE shop_id = :sid AND offer_id = :oid LIMIT 1"),
        {"sid": body.shop_id, "oid": offer_id},
    )
    row = price_result.fetchone()
    if row and row[0] and body.cost_price > float(row[0]):
        warning = f"Себестоимость ({body.cost_price}₽) выше цены продажи ({float(row[0])}₽)"

    return CostUpdateResponse(
        ok=True,
        offer_id=offer_id,
        cost_price=body.cost_price,
    )


# ── Bulk Upload Cost Prices via Excel ─────────────────────

@router.post("/ozon/cost/bulk")
async def upload_cost_excel(
    shop_id: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Bulk upload cost prices from Excel.
    Expected format: column A = offer_id (артикул), column B = cost_price.
    No headers required.
    """
    # Verify shop ownership
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop:
        raise HTTPException(status_code=404, detail="Shop not found")

    # Validate file type
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате .xlsx")

    try:
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="Пустой Excel файл")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка чтения Excel: {e}")

    updated = 0
    errors = []
    total_rows = 0
    skipped_empty = 0

    for row_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=2, values_only=True), start=1):
        total_rows += 1
        offer_id_raw, cost_raw = row
        if offer_id_raw is None or cost_raw is None:
            skipped_empty += 1
            continue

        offer_id = str(offer_id_raw).strip()
        if not offer_id:
            skipped_empty += 1
            continue

        try:
            # Handle Russian locale: non-breaking spaces as thousands separator, comma as decimal
            if isinstance(cost_raw, str):
                cost_str = cost_raw.replace('\xa0', '').replace(' ', '').replace(',', '.')
                cost_price = float(cost_str)
            else:
                cost_price = float(cost_raw)
            if cost_price < 0:
                errors.append(f"Строка {row_idx}: отрицательная цена ({cost_price})")
                continue
        except (ValueError, TypeError):
            errors.append(f"Строка {row_idx}: невалидная цена '{cost_raw}'")
            continue

        await db.execute(
            text("""
                INSERT INTO product_costs (shop_id, offer_id, cost_price)
                VALUES (:shop_id, :offer_id, :cost_price)
                ON CONFLICT (shop_id, offer_id) DO UPDATE SET
                    cost_price = EXCLUDED.cost_price,
                    updated_at = NOW()
            """),
            {"shop_id": shop_id, "offer_id": offer_id, "cost_price": cost_price},
        )
        updated += 1

    await db.commit()
    logger.info(
        "[cost-bulk] shop=%s: total_rows=%d updated=%d skipped_empty=%d errors=%d",
        shop_id, total_rows, updated, skipped_empty, len(errors),
    )

    # Check for cost > price warnings
    warnings = []
    if updated > 0:
        warn_result = await db.execute(
            text("""
                SELECT c.offer_id, c.cost_price, p.price
                FROM product_costs c
                JOIN dim_ozon_products p ON p.shop_id = c.shop_id AND p.offer_id = c.offer_id
                WHERE c.shop_id = :shop_id AND c.cost_price > p.price AND p.price > 0
            """),
            {"shop_id": shop_id},
        )
        for r in warn_result.fetchall():
            warnings.append(f"{r[0]}: с/с {float(r[1]):.0f}₽ > цена {float(r[2]):.0f}₽")

    return {
        "ok": True,
        "updated": updated,
        "errors": errors[:20],
        "warnings": warnings[:20],
    }


# ── Download Cost Template Excel ──────────────────────────

@router.get("/ozon/cost/template")
async def download_cost_template(
    shop_id: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download Excel template with all offer_ids for cost price entry."""
    from fastapi.responses import StreamingResponse

    # Verify shop ownership
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop:
        raise HTTPException(status_code=404, detail="Shop not found")

    # Get all products with current cost
    result = await db.execute(
        text("""
            SELECT p.offer_id,
                   COALESCE(cost.cost_price, 0) AS cost_price,
                   p.name
            FROM dim_ozon_products p
            LEFT JOIN product_costs cost
                ON cost.shop_id = p.shop_id AND cost.offer_id = p.offer_id
            WHERE p.shop_id = :shop_id
              AND NOT COALESCE(p.is_archived, false)
            ORDER BY p.name
        """),
        {"shop_id": shop_id},
    )
    rows = result.fetchall()

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Себестоимость"

    # Headers
    ws.append(["Артикул", "Себестоимость", "Название (справочно)"])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 50

    for r in rows:
        ws.append([r[0], float(r[1]), r[2] or ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=cost_template_{shop_id}.xlsx"},
    )
