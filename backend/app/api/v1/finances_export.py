"""
Finances Export — Excel report generation for Ozon & WB.

GET /finances/ozon/excel?shop_id=X&date_from=Y&date_to=Z
    → Returns .xlsx with 6 sheets: Сводка, Транзакции, По неделям,
      По месяцам, По товарам (SKU), Расходы детально
"""
import io
import logging
from datetime import date, timedelta
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.shop import Shop
from app.models.user import User

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/finances", tags=["Finances Export"])

# ── Styles ────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT = Font(name="Calibri", size=10)
MONEY_FMT = '#,##0'
MONEY_FMT_2 = '#,##0.00'
PCT_FMT = '0.0%'
DATE_FMT = 'DD.MM.YYYY'
THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)
ALT_ROW_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
DELTA_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")  # grey-200 for delta columns
DELTA_FILL_ALT = PatternFill(start_color="D1D5DB", end_color="D1D5DB", fill_type="solid")  # grey-300 for alt rows in delta columns
DELTA_HDR_FILL = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")  # grey-600 header
GREEN_FONT = Font(name="Calibri", bold=True, color="16A34A", size=10)
RED_FONT = Font(name="Calibri", bold=True, color="DC2626", size=10)
GREEN_FONT_SM = Font(name="Calibri", bold=True, color="16A34A", size=9)
RED_FONT_SM = Font(name="Calibri", bold=True, color="DC2626", size=9)
GREY_FONT_SM = Font(name="Calibri", color="6B7280", size=9)
PROFIT_GREEN = Font(name="Calibri", bold=True, color="16A34A", size=11)
PROFIT_RED = Font(name="Calibri", bold=True, color="DC2626", size=11)


def _style_header_row(ws, row_num: int, col_count: int):
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_row(ws, row_num: int, col_count: int, is_alt: bool = False):
    """Apply data row styling."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = ALT_ROW_FILL


def _style_total_row(ws, row_num: int, col_count: int):
    """Apply total row styling."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 40):
    """Auto-adjust column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _safe_delta(current: float, previous: float) -> float:
    if previous == 0:
        return 100.0 if current > 0 else (-100.0 if current < 0 else 0.0)
    return round((current - previous) / abs(previous) * 100, 1)


MONTHS_SHORT_RU = {
    1: "янв.", 2: "фев.", 3: "март.", 4: "апр.", 5: "мая", 6: "июн.",
    7: "июл.", 8: "авг.", 9: "сен.", 10: "окт.", 11: "нояб.", 12: "дек.",
}


def _fmt_date_ru(d) -> str:
    """Format date as '2 март.' for Excel readability."""
    if isinstance(d, str):
        d = date.fromisoformat(d)
    return f"{d.day} {MONTHS_SHORT_RU.get(d.month, str(d.month))}"


CAT_MAP = {
    "Logistics": "logistics",
    "Storage": "storage",
    "Acquiring": "acquiring",
    "Refund": "refunds",
    "Penalty": "penalties",
    "Compensation": "compensation",
    "Marketing": "marketing",
}


# ══════════════════════════════════════════════════════════════
# Endpoint
# ══════════════════════════════════════════════════════════════

@router.get("/ozon/excel")
async def export_ozon_excel(
    shop_id: int = Query(..., description="Shop ID"),
    date_from: date = Query(..., description="Start date"),
    date_to: date = Query(..., description="End date"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Generate and download comprehensive Ozon Excel financial report."""

    # ── Verify shop ──
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "ozon":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Shop not found")

    from app.core.clickhouse import get_clickhouse_client

    try:
        ch = get_clickhouse_client()
    except Exception as e:
        logger.error("ClickHouse connection error: %s", e)
        raise HTTPException(status_code=500, detail="Analytics unavailable")

    d_start = date_from
    d_end = date_to
    span = (d_end - d_start).days + 1
    d_prev_start = d_start - timedelta(days=span)
    d_prev_end = d_start - timedelta(days=1)

    # ══════════════════════════════════════════════════════════
    # Fetch all data from ClickHouse
    # ══════════════════════════════════════════════════════════

    # 1. Aggregate KPI (revenue, commission, services, payout, orders)
    txn_totals = ch.query("""
        SELECT
            sumIf(accruals_for_sale,
                toDate(operation_date) >= {d_start:Date} AND toDate(operation_date) <= {d_end:Date}
                AND category = 'Revenue') AS accruals_cur,
            sumIf(accruals_for_sale,
                toDate(operation_date) >= {d_prev_start:Date} AND toDate(operation_date) <= {d_prev_end:Date}
                AND category = 'Revenue') AS accruals_prev,
            sumIf(sale_commission,
                toDate(operation_date) >= {d_start:Date} AND toDate(operation_date) <= {d_end:Date}
                AND category = 'Revenue') AS comm_cur,
            sumIf(sale_commission,
                toDate(operation_date) >= {d_prev_start:Date} AND toDate(operation_date) <= {d_prev_end:Date}
                AND category = 'Revenue') AS comm_prev,
            sumIf(services_total,
                toDate(operation_date) >= {d_start:Date} AND toDate(operation_date) <= {d_end:Date}
                AND category = 'Revenue') AS svc_cur,
            sumIf(services_total,
                toDate(operation_date) >= {d_prev_start:Date} AND toDate(operation_date) <= {d_prev_end:Date}
                AND category = 'Revenue') AS svc_prev,
            sumIf(amount,
                toDate(operation_date) >= {d_start:Date} AND toDate(operation_date) <= {d_end:Date}
                ) AS pay_cur,
            sumIf(amount,
                toDate(operation_date) >= {d_prev_start:Date} AND toDate(operation_date) <= {d_prev_end:Date}
                ) AS pay_prev,
            countIf(
                toDate(operation_date) >= {d_start:Date} AND toDate(operation_date) <= {d_end:Date}
                AND category = 'Revenue') AS ord_cur,
            countIf(
                toDate(operation_date) >= {d_prev_start:Date} AND toDate(operation_date) <= {d_prev_end:Date}
                AND category = 'Revenue') AS ord_prev
        FROM mms_analytics.fact_ozon_transactions FINAL
        WHERE shop_id = {shop_id:UInt32}
          AND toDate(operation_date) >= {d_prev_start:Date}
          AND toDate(operation_date) <= {d_end:Date}
    """, parameters={
        "shop_id": shop_id,
        "d_start": d_start, "d_end": d_end,
        "d_prev_start": d_prev_start, "d_prev_end": d_prev_end,
    })

    r = txn_totals.result_rows[0] if txn_totals.result_rows else [0]*10
    revenue_cur = float(r[0] or 0)
    revenue_prev = float(r[1] or 0)
    commission_cur = abs(float(r[2] or 0))
    commission_prev = abs(float(r[3] or 0))
    services_cur = abs(float(r[4] or 0))
    services_prev = abs(float(r[5] or 0))
    payout_cur = float(r[6] or 0)
    payout_prev = float(r[7] or 0)
    orders_cur = int(r[8] or 0)
    orders_prev = int(r[9] or 0)

    # 2. Bulk expense categories
    bulk_cur = {"logistics": 0.0, "storage": 0.0, "acquiring": 0.0, "refunds": 0.0,
                "penalties": 0.0, "compensation": 0.0, "marketing": 0.0, "other": 0.0}
    bulk_prev = dict(bulk_cur)

    cat_result = ch.query("""
        SELECT
            category,
            sumIf(amount, toDate(operation_date) >= {d_start:Date} AND toDate(operation_date) <= {d_end:Date}) AS total_cur,
            sumIf(amount, toDate(operation_date) >= {d_prev_start:Date} AND toDate(operation_date) <= {d_prev_end:Date}) AS total_prev
        FROM mms_analytics.fact_ozon_transactions FINAL
        WHERE shop_id = {shop_id:UInt32}
          AND toDate(operation_date) >= {d_prev_start:Date}
          AND toDate(operation_date) <= {d_end:Date}
          AND category NOT IN ('Revenue')
        GROUP BY category
    """, parameters={
        "shop_id": shop_id,
        "d_start": d_start, "d_end": d_end,
        "d_prev_start": d_prev_start, "d_prev_end": d_prev_end,
    })
    for r in cat_result.result_rows:
        key = CAT_MAP.get(r[0], "other")
        if key in bulk_cur:
            bulk_cur[key] = float(r[1] or 0)
            bulk_prev[key] = float(r[2] or 0)
        else:
            bulk_cur["other"] += float(r[1] or 0)
            bulk_prev["other"] += float(r[2] or 0)

    # 2b. Detailed per-operation_type breakdown (for summary P&L)
    detail_breakdown = ch.query("""
        SELECT
            operation_type,
            category,
            sumIf(abs(amount), toDate(operation_date) >= {d_start:Date} AND toDate(operation_date) <= {d_end:Date}) AS cur_amount,
            sumIf(abs(amount), toDate(operation_date) >= {d_prev_start:Date} AND toDate(operation_date) <= {d_prev_end:Date}) AS prev_amount
        FROM mms_analytics.fact_ozon_transactions FINAL
        WHERE shop_id = {shop_id:UInt32}
          AND toDate(operation_date) >= {d_prev_start:Date}
          AND toDate(operation_date) <= {d_end:Date}
          AND category NOT IN ('Revenue')
        GROUP BY operation_type, category
        ORDER BY cur_amount DESC
    """, parameters={
        "shop_id": shop_id,
        "d_start": d_start, "d_end": d_end,
        "d_prev_start": d_prev_start, "d_prev_end": d_prev_end,
    })

    # Parse detailed breakdown into named buckets
    # FBO/logistics sub-items
    fbo_crossdocking_cur, fbo_crossdocking_prev = 0.0, 0.0
    fbo_intake_cur, fbo_intake_prev = 0.0, 0.0
    fbo_other_cur, fbo_other_prev = 0.0, 0.0
    # Marketing sub-items
    ads_cpc_cur, ads_cpc_prev = 0.0, 0.0
    ads_review_cur, ads_review_prev = 0.0, 0.0
    ads_other_cur, ads_other_prev = 0.0, 0.0
    # Other sub-items
    cashback_cur, cashback_prev = 0.0, 0.0
    fines_cur, fines_prev = 0.0, 0.0
    packaging_cur, packaging_prev = 0.0, 0.0
    other_misc_cur, other_misc_prev = 0.0, 0.0

    FBO_CROSSDOCKING_OPS = {'MarketplaceServiceItemCrossdocking'}
    FBO_INTAKE_OPS = {
        'OperationMarketplaceSupplyAdditional',
        'OperationMarketplaceSupplyExpirationDateProcessing',
        'OperationMarketplaceServiceSupplyInboundCargoShortage',
        'OperationMarketplaceServiceSupplyInboundSupplyShortage',
        'OperationMarketplaceServiceSupplyInboundCargoSurplus',
    }
    ADS_CPC_OPS = {'OperationMarketplaceCostPerClick'}
    ADS_REVIEW_OPS = {'OperationPointsForReviews'}
    CASHBACK_OPS = {'OperationMarketplaceServicePremiumCashbackIndividualPoints'}
    FINE_OPS = {'DefectFineShipmentDelay', 'DefectFineShipmentDelayRated', 'DefectFineCancellation'}
    PACKAGING_OPS = {'OperationMarketplacePackageMaterialsProvision', 'OperationMarketplacePackageRedistribution'}

    for r in detail_breakdown.result_rows:
        op_type = r[0]
        cat = r[1]
        c_val = float(r[2] or 0)
        p_val = float(r[3] or 0)

        if cat == 'Logistics':
            if op_type in FBO_CROSSDOCKING_OPS:
                fbo_crossdocking_cur += c_val; fbo_crossdocking_prev += p_val
            elif op_type in FBO_INTAKE_OPS:
                fbo_intake_cur += c_val; fbo_intake_prev += p_val
            else:
                fbo_other_cur += c_val; fbo_other_prev += p_val
        elif cat == 'Marketing':
            if op_type in ADS_CPC_OPS:
                ads_cpc_cur += c_val; ads_cpc_prev += p_val
            elif op_type in ADS_REVIEW_OPS:
                ads_review_cur += c_val; ads_review_prev += p_val
            else:
                ads_other_cur += c_val; ads_other_prev += p_val
        elif cat in ('Other', 'Penalty'):
            if op_type in CASHBACK_OPS:
                cashback_cur += c_val; cashback_prev += p_val
            elif op_type in FINE_OPS:
                fines_cur += c_val; fines_prev += p_val
            elif op_type in PACKAGING_OPS:
                packaging_cur += c_val; packaging_prev += p_val
            else:
                other_misc_cur += c_val; other_misc_prev += p_val

    # 3. Ad spend
    ad_spend_cur = abs(bulk_cur.get("marketing", 0))
    ad_spend_prev = abs(bulk_prev.get("marketing", 0))

    # 4. COGS
    cogs_cur = 0.0
    cogs_prev = 0.0
    try:
        sku_map_result = await db.execute(
            text("SELECT sku, offer_id FROM dim_ozon_products WHERE shop_id = :shop_id AND sku > 0"),
            {"shop_id": shop_id},
        )
        sku_to_offer = {int(r[0]): r[1] for r in sku_map_result.fetchall()}

        # Load barcodes for per-SKU sheet
        barcode_result = await db.execute(
            text("SELECT sku, barcode FROM dim_ozon_products WHERE shop_id = :shop_id AND sku > 0 AND barcode IS NOT NULL AND barcode != ''"),
            {"shop_id": shop_id},
        )
        sku_to_barcode = {int(r[0]): r[1] for r in barcode_result.fetchall()}

        cost_result = await db.execute(
            text("""SELECT offer_id, COALESCE(cost_price, 0) + COALESCE(packaging_cost, 0) AS total_cost
                    FROM product_costs WHERE shop_id = :shop_id AND (cost_price > 0 OR packaging_cost > 0)"""),
            {"shop_id": shop_id},
        )
        cost_map = {r[0]: float(r[1]) for r in cost_result.fetchall()}

        if cost_map and sku_to_offer:
            sku_cost_map = {}
            for sku, offer_id in sku_to_offer.items():
                cost = cost_map.get(offer_id, 0)
                if cost > 0:
                    sku_cost_map[sku] = cost

            if sku_cost_map:
                sku_list = list(sku_cost_map.keys())
                cogs_ch = ch.query("""
                    SELECT toDate(operation_date) AS dt, sku, count() AS qty
                    FROM mms_analytics.fact_ozon_transactions FINAL
                    WHERE shop_id = {shop_id:UInt32}
                      AND toDate(operation_date) >= {d_prev_start:Date}
                      AND toDate(operation_date) <= {d_end:Date}
                      AND category = 'Revenue'
                      AND sku IN {skus:Array(UInt64)}
                    GROUP BY dt, sku
                """, parameters={
                    "shop_id": shop_id,
                    "d_prev_start": d_prev_start, "d_end": d_end,
                    "skus": sku_list,
                })
                for r in cogs_ch.result_rows:
                    row_date = r[0]
                    sku = int(r[1])
                    qty = int(r[2] or 0)
                    cost = sku_cost_map.get(sku, 0)
                    if d_start <= row_date <= d_end:
                        cogs_cur += cost * qty
                    elif d_prev_start <= row_date <= d_prev_end:
                        cogs_prev += cost * qty
    except Exception as e:
        logger.warning("COGS calc failed: %s", e)

    # Derived metrics
    # Logistics total = delivery to customers (services_total) + FBO (crossdocking + intake + other)
    fbo_total_cur = fbo_crossdocking_cur + fbo_intake_cur + fbo_other_cur
    fbo_total_prev = fbo_crossdocking_prev + fbo_intake_prev + fbo_other_prev
    logistics_total_cur = services_cur + fbo_total_cur
    logistics_total_prev = services_prev + fbo_total_prev
    # Operating expenses = acquiring + refunds + storage + other/fines/cashback/packaging
    opex_cur = abs(bulk_cur["acquiring"]) + abs(bulk_cur["refunds"]) + abs(bulk_cur["storage"]) + abs(bulk_cur["other"]) + abs(bulk_cur["penalties"]) + abs(bulk_cur["compensation"])
    opex_prev = abs(bulk_prev["acquiring"]) + abs(bulk_prev["refunds"]) + abs(bulk_prev["storage"]) + abs(bulk_prev["other"]) + abs(bulk_prev["penalties"]) + abs(bulk_prev["compensation"])
    # Total MP fees = commission + logistics + opex
    mp_fees_cur = commission_cur + logistics_total_cur + opex_cur
    mp_fees_prev = commission_prev + logistics_total_prev + opex_prev
    profit_cur = revenue_cur - mp_fees_cur - ad_spend_cur - cogs_cur
    profit_prev = revenue_prev - mp_fees_prev - ad_spend_prev - cogs_prev

    # 5. All raw transactions for the period
    raw_txns = ch.query("""
        SELECT
            operation_id, toDate(operation_date) AS dt, operation_type, operation_type_name,
            category, sku, item_name, amount, accruals_for_sale,
            sale_commission, services_total, posting_number, delivery_schema
        FROM mms_analytics.fact_ozon_transactions FINAL
        WHERE shop_id = {shop_id:UInt32}
          AND toDate(operation_date) >= {d_start:Date}
          AND toDate(operation_date) <= {d_end:Date}
        ORDER BY toDate(operation_date), category, operation_type
    """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": d_end})

    # 6. Weekly aggregation — FULL RETROSPECTIVE (all weeks since shop creation)
    weekly_data = ch.query("""
        SELECT
            toMonday(toDate(operation_date)) AS week_start,
            toMonday(toDate(operation_date)) + 6 AS week_end,
            countIf(category = 'Revenue') AS qty,
            sumIf(accruals_for_sale, category = 'Revenue') AS sales,
            sumIf(abs(amount), category = 'Refund') AS returns,
            abs(sumIf(sale_commission, category = 'Revenue')) AS commission,
            sumIf(amount, category = 'Compensation') AS compensations,
            abs(sumIf(services_total, category = 'Revenue')) AS other_services,
            abs(sumIf(amount, category = 'Marketing')) AS marketing,
            abs(sumIf(amount, category IN ('Penalty', 'Other'))) AS other_charges,
            sumIf(abs(amount), category = 'Logistics' AND operation_type IN (
                'MarketplaceServiceItemCrossdocking',
                'OperationMarketplaceSupplyAdditional',
                'OperationMarketplaceSupplyExpirationDateProcessing',
                'OperationMarketplaceServiceSupplyInboundCargoShortage',
                'OperationMarketplaceServiceSupplyInboundSupplyShortage'
            )) AS fbo_services,
            abs(sumIf(amount, category = 'Acquiring')) AS acquiring,
            sumIf(abs(amount), category = 'Logistics' AND operation_type NOT IN (
                'MarketplaceServiceItemCrossdocking',
                'OperationMarketplaceSupplyAdditional',
                'OperationMarketplaceSupplyExpirationDateProcessing',
                'OperationMarketplaceServiceSupplyInboundCargoShortage',
                'OperationMarketplaceServiceSupplyInboundSupplyShortage'
            )) AS delivery_services,
            abs(sumIf(amount, category = 'Storage')) AS storage,
            sum(amount) AS payout
        FROM mms_analytics.fact_ozon_transactions FINAL
        WHERE shop_id = {shop_id:UInt32}
        GROUP BY week_start
        ORDER BY week_start
    """, parameters={"shop_id": shop_id})

    # 7. Monthly aggregation — FULL RETROSPECTIVE
    monthly_data = ch.query("""
        SELECT
            toYYYYMM(toDate(operation_date)) AS ym,
            min(toDate(operation_date)) AS m_start,
            max(toDate(operation_date)) AS m_end,
            countIf(category = 'Revenue') AS qty,
            sumIf(accruals_for_sale, category = 'Revenue') AS sales,
            sumIf(abs(amount), category = 'Refund') AS returns,
            abs(sumIf(sale_commission, category = 'Revenue')) AS commission,
            sumIf(amount, category = 'Compensation') AS compensations,
            abs(sumIf(services_total, category = 'Revenue')) AS other_services,
            abs(sumIf(amount, category = 'Marketing')) AS marketing,
            abs(sumIf(amount, category IN ('Penalty', 'Other'))) AS other_charges,
            sumIf(abs(amount), category = 'Logistics' AND operation_type IN (
                'MarketplaceServiceItemCrossdocking',
                'OperationMarketplaceSupplyAdditional',
                'OperationMarketplaceSupplyExpirationDateProcessing',
                'OperationMarketplaceServiceSupplyInboundCargoShortage',
                'OperationMarketplaceServiceSupplyInboundSupplyShortage'
            )) AS fbo_services,
            abs(sumIf(amount, category = 'Acquiring')) AS acquiring,
            sumIf(abs(amount), category = 'Logistics' AND operation_type NOT IN (
                'MarketplaceServiceItemCrossdocking',
                'OperationMarketplaceSupplyAdditional',
                'OperationMarketplaceSupplyExpirationDateProcessing',
                'OperationMarketplaceServiceSupplyInboundCargoShortage',
                'OperationMarketplaceServiceSupplyInboundSupplyShortage'
            )) AS delivery_services,
            abs(sumIf(amount, category = 'Storage')) AS storage,
            sum(amount) AS payout
        FROM mms_analytics.fact_ozon_transactions FINAL
        WHERE shop_id = {shop_id:UInt32}
        GROUP BY ym
        ORDER BY ym
    """, parameters={"shop_id": shop_id})

    # 8. Per-SKU P&L
    sku_data = ch.query("""
        SELECT
            sku,
            any(item_name) AS name,
            sumIf(accruals_for_sale, category = 'Revenue') AS revenue,
            countIf(category = 'Revenue') AS qty,
            abs(sumIf(sale_commission, category = 'Revenue')) AS commission,
            abs(sumIf(services_total, category = 'Revenue')) AS services,
            sum(amount) AS payout
        FROM mms_analytics.fact_ozon_transactions FINAL
        WHERE shop_id = {shop_id:UInt32}
          AND toDate(operation_date) >= {d_start:Date}
          AND toDate(operation_date) <= {d_end:Date}
          AND sku > 0
        GROUP BY sku
        ORDER BY revenue DESC
    """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": d_end})

    # 9. Detailed expenses by operation_type
    expense_detail = ch.query("""
        SELECT
            operation_type,
            operation_type_name,
            category,
            count() AS cnt,
            round(sum(amount), 2) AS total,
            round(avg(amount), 2) AS avg_amount
        FROM mms_analytics.fact_ozon_transactions FINAL
        WHERE shop_id = {shop_id:UInt32}
          AND toDate(operation_date) >= {d_start:Date}
          AND toDate(operation_date) <= {d_end:Date}
        GROUP BY operation_type, operation_type_name, category
        ORDER BY abs(sum(amount)) DESC
    """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": d_end})

    # ══════════════════════════════════════════════════════════
    # Build Excel workbook
    # ══════════════════════════════════════════════════════════

    wb = Workbook()

    # ── Sheet 1: Сводка (P&L) ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Сводка"
    ws1.sheet_properties.tabColor = "2563EB"

    # Additional styles for sections
    SECTION_HDR_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    SECTION_HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    SUBTOTAL_FILL = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    SUBTOTAL_FONT = Font(name="Calibri", bold=True, size=10)
    SUB_ITEM_FONT = Font(name="Calibri", size=10, color="4B5563")
    INDENT = "    "  # visual indent for sub-items

    def _write_section_header(ws, row, label, cols=5):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        ws.cell(row=row, column=1, value=label).font = SECTION_HDR_FONT
        for c in range(1, cols + 1):
            ws.cell(row=row, column=c).fill = SECTION_HDR_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER

    def _write_kpi_row(ws, row, label, cur, prev, is_money=True, is_sub=False, is_total=False, is_profit=False, cols=5):
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=round(cur, 2) if is_money else cur)
        c3 = ws.cell(row=row, column=3, value=round(prev, 2) if is_money else prev)
        delta = _safe_delta(cur, prev)
        c4 = ws.cell(row=row, column=4, value=f"{'+' if delta > 0 else ''}{delta}%")
        # % of revenue
        pct_cur = cur / revenue_cur * 100 if revenue_cur > 0 else 0
        c5 = ws.cell(row=row, column=5, value=round(pct_cur, 1) if is_money else "")

        if is_money:
            c2.number_format = MONEY_FMT
            c3.number_format = MONEY_FMT
        if is_money and c5.value != "":
            c5.number_format = '0.0"%"'
        for c in [c2, c3, c4, c5]:
            c.alignment = Alignment(horizontal='right')

        if delta > 0:
            c4.font = GREEN_FONT
        elif delta < 0:
            c4.font = RED_FONT

        if is_sub:
            c1.font = SUB_ITEM_FONT
        elif is_total:
            _style_total_row(ws, row, cols)
        elif is_profit:
            _style_total_row(ws, row, cols)
            c2.font = PROFIT_GREEN if cur >= 0 else PROFIT_RED
            c3.font = PROFIT_GREEN if prev >= 0 else PROFIT_RED
        else:
            c1.font = NORMAL_FONT

        if not is_total and not is_profit:
            _style_data_row(ws, row, cols)

    # Title
    ws1.merge_cells('A1:E1')
    ws1['A1'] = f"Финансовый отчёт Ozon — {shop.name}"
    ws1['A1'].font = Font(name="Calibri", bold=True, size=16, color="2563EB")
    ws1.merge_cells('A2:E2')
    ws1['A2'] = f"Период: {d_start.strftime('%d.%m.%Y')} — {d_end.strftime('%d.%m.%Y')} ({span} дн.)"
    ws1['A2'].font = Font(name="Calibri", size=11, color="6B7280")

    # Column headers
    row = 4
    for ci, hdr in enumerate(["Показатель", "Текущий период", "Предыдущий период", "Изм. %", "% выр."], 1):
        ws1.cell(row=row, column=ci, value=hdr)
    _style_header_row(ws1, row, 5)

    row = 5

    # ═══ ВЫРУЧКА И ЗАКАЗЫ ═══
    _write_section_header(ws1, row, "ВЫРУЧКА И ЗАКАЗЫ")
    row += 1
    avg_check_cur = revenue_cur / orders_cur if orders_cur > 0 else 0
    avg_check_prev = revenue_prev / orders_prev if orders_prev > 0 else 0
    _write_kpi_row(ws1, row, "Выручка (продажи)", revenue_cur, revenue_prev); row += 1
    _write_kpi_row(ws1, row, "Заказы", orders_cur, orders_prev, is_money=False); row += 1
    _write_kpi_row(ws1, row, "Средний чек", avg_check_cur, avg_check_prev); row += 1

    # ═══ КОМИССИЯ OZON ═══
    row += 1
    _write_section_header(ws1, row, "КОМИССИЯ OZON")
    row += 1
    _write_kpi_row(ws1, row, "Комиссия с продаж", commission_cur, commission_prev); row += 1

    # ═══ ДОСТАВКА И ЛОГИСТИКА ═══
    row += 1
    _write_section_header(ws1, row, "ДОСТАВКА И ЛОГИСТИКА")
    row += 1
    _write_kpi_row(ws1, row, "Доставка покупателям", services_cur, services_prev); row += 1
    if fbo_crossdocking_cur > 0 or fbo_crossdocking_prev > 0:
        _write_kpi_row(ws1, row, f"{INDENT}Кроссдокинг", fbo_crossdocking_cur, fbo_crossdocking_prev, is_sub=True); row += 1
    if fbo_intake_cur > 0 or fbo_intake_prev > 0:
        _write_kpi_row(ws1, row, f"{INDENT}Приёмка товаров", fbo_intake_cur, fbo_intake_prev, is_sub=True); row += 1
    if fbo_other_cur > 0 or fbo_other_prev > 0:
        _write_kpi_row(ws1, row, f"{INDENT}Прочие FBO-услуги", fbo_other_cur, fbo_other_prev, is_sub=True); row += 1
    _write_kpi_row(ws1, row, "Итого логистика", logistics_total_cur, logistics_total_prev, is_total=True); row += 1

    # ═══ ОПЕРАЦИОННЫЕ РАСХОДЫ ═══
    row += 1
    _write_section_header(ws1, row, "ОПЕРАЦИОННЫЕ РАСХОДЫ")
    row += 1
    _write_kpi_row(ws1, row, "Эквайринг", abs(bulk_cur["acquiring"]), abs(bulk_prev["acquiring"])); row += 1
    _write_kpi_row(ws1, row, "Возвраты и отмены", abs(bulk_cur["refunds"]), abs(bulk_prev["refunds"])); row += 1
    _write_kpi_row(ws1, row, "Хранение", abs(bulk_cur["storage"]), abs(bulk_prev["storage"])); row += 1
    if fines_cur > 0 or fines_prev > 0:
        _write_kpi_row(ws1, row, "Штрафы (задержки, отмены)", fines_cur, fines_prev); row += 1
    if cashback_cur > 0 or cashback_prev > 0:
        _write_kpi_row(ws1, row, "Баллы покупателям / кэшбек", cashback_cur, cashback_prev); row += 1
    if packaging_cur > 0 or packaging_prev > 0:
        _write_kpi_row(ws1, row, "Упаковка", packaging_cur, packaging_prev); row += 1
    if abs(bulk_cur["compensation"]) > 0 or abs(bulk_prev["compensation"]) > 0:
        _write_kpi_row(ws1, row, "Компенсации", abs(bulk_cur["compensation"]), abs(bulk_prev["compensation"])); row += 1
    if other_misc_cur > 0 or other_misc_prev > 0:
        _write_kpi_row(ws1, row, "Прочее", other_misc_cur, other_misc_prev); row += 1
    _write_kpi_row(ws1, row, "Итого операционные", opex_cur, opex_prev, is_total=True); row += 1

    # ═══ РЕКЛАМА ═══
    row += 1
    _write_section_header(ws1, row, "РЕКЛАМА")
    row += 1
    if ads_cpc_cur > 0 or ads_cpc_prev > 0:
        _write_kpi_row(ws1, row, "Продвижение (CPC)", ads_cpc_cur, ads_cpc_prev); row += 1
    if ads_review_cur > 0 or ads_review_prev > 0:
        _write_kpi_row(ws1, row, "Баллы за отзывы", ads_review_cur, ads_review_prev); row += 1
    if ads_other_cur > 0 or ads_other_prev > 0:
        _write_kpi_row(ws1, row, "Прочая реклама", ads_other_cur, ads_other_prev); row += 1
    _write_kpi_row(ws1, row, "Итого реклама", ad_spend_cur, ad_spend_prev, is_total=True); row += 1
    drr_cur = ad_spend_cur / revenue_cur * 100 if revenue_cur > 0 else 0
    drr_prev = ad_spend_prev / revenue_prev * 100 if revenue_prev > 0 else 0
    ws1.cell(row=row, column=1, value=f"{INDENT}ДРР (доля рекл. расх.)").font = SUB_ITEM_FONT
    ws1.cell(row=row, column=2, value=f"{round(drr_cur, 1)}%").alignment = Alignment(horizontal='right')
    ws1.cell(row=row, column=3, value=f"{round(drr_prev, 1)}%").alignment = Alignment(horizontal='right')
    _style_data_row(ws1, row, 5)
    row += 1

    # ═══ ИТОГИ ═══
    row += 1
    _write_section_header(ws1, row, "ИТОГИ")
    row += 1
    _write_kpi_row(ws1, row, "Всего удержано Ozon", mp_fees_cur, mp_fees_prev, is_total=True); row += 1
    _write_kpi_row(ws1, row, "К перечислению", payout_cur, payout_prev); row += 1
    _write_kpi_row(ws1, row, "Реклама", ad_spend_cur, ad_spend_prev); row += 1
    _write_kpi_row(ws1, row, "Себестоимость (COGS)", cogs_cur, cogs_prev); row += 1
    total_expenses_cur = mp_fees_cur + ad_spend_cur + cogs_cur
    total_expenses_prev = mp_fees_prev + ad_spend_prev + cogs_prev
    _write_kpi_row(ws1, row, "Всего расходов", total_expenses_cur, total_expenses_prev, is_total=True); row += 1
    _write_kpi_row(ws1, row, "ЧИСТАЯ ПРИБЫЛЬ", profit_cur, profit_prev, is_profit=True); row += 1
    margin_cur = profit_cur / revenue_cur * 100 if revenue_cur > 0 else 0
    margin_prev = profit_prev / revenue_prev * 100 if revenue_prev > 0 else 0
    ws1.cell(row=row, column=1, value=f"{INDENT}Маржинальность").font = SUB_ITEM_FONT
    ws1.cell(row=row, column=2, value=f"{round(margin_cur, 1)}%").alignment = Alignment(horizontal='right')
    ws1.cell(row=row, column=3, value=f"{round(margin_prev, 1)}%").alignment = Alignment(horizontal='right')
    _style_data_row(ws1, row, 5)
    row += 2

    # ═══ СТРУКТУРА РАСХОДОВ (% от выручки) — с обоими периодами ═══
    ws1.cell(row=row, column=1, value="СТРУКТУРА РАСХОДОВ (% от выручки)").font = SECTION_HDR_FONT
    ws1.cell(row=row, column=1).fill = SECTION_HDR_FILL
    for ci, hdr in enumerate(["", "Тек. период", "% выр.", "Пред. период", "% выр."], 1):
        ws1.cell(row=row, column=ci, value=hdr if ci > 1 else "СТРУКТУРА РАСХОДОВ (% от выручки)")
    _style_header_row(ws1, row, 5)
    row += 1

    pct_rows = [
        ("Комиссия", commission_cur, commission_prev),
        ("Доставка покупателям", services_cur, services_prev),
        ("Кроссдокинг / приёмка (FBO)", fbo_total_cur, fbo_total_prev),
        ("Эквайринг", abs(bulk_cur["acquiring"]), abs(bulk_prev["acquiring"])),
        ("Возвраты", abs(bulk_cur["refunds"]), abs(bulk_prev["refunds"])),
        ("Хранение", abs(bulk_cur["storage"]), abs(bulk_prev["storage"])),
        ("Штрафы / кэшбек / прочее", fines_cur + cashback_cur + packaging_cur + other_misc_cur,
                                      fines_prev + cashback_prev + packaging_prev + other_misc_prev),
        ("Реклама", ad_spend_cur, ad_spend_prev),
        ("Себестоимость", cogs_cur, cogs_prev),
    ]

    # Total row at end
    total_exp_cur = sum(v[1] for v in pct_rows)
    total_exp_prev = sum(v[2] for v in pct_rows)

    for i, (label, cur_val, prev_val) in enumerate(pct_rows):
        pct_c = cur_val / revenue_cur * 100 if revenue_cur > 0 else 0
        pct_p = prev_val / revenue_prev * 100 if revenue_prev > 0 else 0
        ws1.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws1.cell(row=row, column=2, value=round(cur_val, 2)).number_format = MONEY_FMT
        ws1.cell(row=row, column=3, value=round(pct_c, 1))
        ws1.cell(row=row, column=3).number_format = '0.0"%"'
        ws1.cell(row=row, column=4, value=round(prev_val, 2)).number_format = MONEY_FMT
        ws1.cell(row=row, column=5, value=round(pct_p, 1))
        ws1.cell(row=row, column=5).number_format = '0.0"%"'
        for c in range(2, 6):
            ws1.cell(row=row, column=c).alignment = Alignment(horizontal='right')
        _style_data_row(ws1, row, 5, is_alt=(i % 2 == 1))
        row += 1

    # Total expenses row
    pct_tc = total_exp_cur / revenue_cur * 100 if revenue_cur > 0 else 0
    pct_tp = total_exp_prev / revenue_prev * 100 if revenue_prev > 0 else 0
    ws1.cell(row=row, column=1, value="Итого расходов").font = TOTAL_FONT
    ws1.cell(row=row, column=2, value=round(total_exp_cur, 2)).number_format = MONEY_FMT
    ws1.cell(row=row, column=3, value=round(pct_tc, 1))
    ws1.cell(row=row, column=3).number_format = '0.0"%"'
    ws1.cell(row=row, column=4, value=round(total_exp_prev, 2)).number_format = MONEY_FMT
    ws1.cell(row=row, column=5, value=round(pct_tp, 1))
    ws1.cell(row=row, column=5).number_format = '0.0"%"'
    _style_total_row(ws1, row, 5)
    row += 1

    # Profit row
    pct_prof_c = profit_cur / revenue_cur * 100 if revenue_cur > 0 else 0
    pct_prof_p = profit_prev / revenue_prev * 100 if revenue_prev > 0 else 0
    ws1.cell(row=row, column=1, value="Прибыль").font = TOTAL_FONT
    c2 = ws1.cell(row=row, column=2, value=round(profit_cur, 2))
    c2.number_format = MONEY_FMT
    c2.font = PROFIT_GREEN if profit_cur >= 0 else PROFIT_RED
    ws1.cell(row=row, column=3, value=round(pct_prof_c, 1))
    ws1.cell(row=row, column=3).number_format = '0.0"%"'
    c4 = ws1.cell(row=row, column=4, value=round(profit_prev, 2))
    c4.number_format = MONEY_FMT
    c4.font = PROFIT_GREEN if profit_prev >= 0 else PROFIT_RED
    ws1.cell(row=row, column=5, value=round(pct_prof_p, 1))
    ws1.cell(row=row, column=5).number_format = '0.0"%"'
    _style_total_row(ws1, row, 5)

    _auto_width(ws1)
    ws1.column_dimensions['A'].width = 32
    ws1.freeze_panes = 'A5'

    # ── Sheet 2: Все транзакции ───────────────────────────────
    ws2 = wb.create_sheet("Транзакции")
    ws2.sheet_properties.tabColor = "16A34A"

    txn_headers = [
        "ID операции", "Дата", "Тип операции", "Описание",
        "Категория", "SKU", "Товар", "Сумма",
        "Начисления (продажа)", "Комиссия", "Сервисы",
        "Номер отправления", "Схема доставки"
    ]
    for col, h in enumerate(txn_headers, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header_row(ws2, 1, len(txn_headers))

    for i, r in enumerate(raw_txns.result_rows):
        row_num = i + 2
        ws2.cell(row=row_num, column=1, value=r[0])  # operation_id
        ws2.cell(row=row_num, column=2, value=str(r[1])).number_format = DATE_FMT  # dt
        ws2.cell(row=row_num, column=3, value=r[2])   # operation_type
        ws2.cell(row=row_num, column=4, value=r[3])   # operation_type_name
        ws2.cell(row=row_num, column=5, value=r[4])   # category
        ws2.cell(row=row_num, column=6, value=r[5] if r[5] else "")  # sku
        ws2.cell(row=row_num, column=7, value=r[6])   # item_name
        ws2.cell(row=row_num, column=8, value=float(r[7] or 0)).number_format = MONEY_FMT_2  # amount
        ws2.cell(row=row_num, column=9, value=float(r[8] or 0)).number_format = MONEY_FMT_2  # accruals
        ws2.cell(row=row_num, column=10, value=float(r[9] or 0)).number_format = MONEY_FMT_2  # commission
        ws2.cell(row=row_num, column=11, value=float(r[10] or 0)).number_format = MONEY_FMT_2  # services
        ws2.cell(row=row_num, column=12, value=r[11] or "")  # posting_number
        ws2.cell(row=row_num, column=13, value=r[12] or "")  # delivery_schema
        _style_data_row(ws2, row_num, len(txn_headers), is_alt=(i % 2 == 1))

    # Totals row
    total_row = len(raw_txns.result_rows) + 2
    ws2.cell(row=total_row, column=1, value="ИТОГО")
    ws2.cell(row=total_row, column=7, value=f"{len(raw_txns.result_rows)} операций")
    # Sum formulas
    for col in [8, 9, 10, 11]:
        ws2.cell(row=total_row, column=col,
                 value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row - 1})")
        ws2.cell(row=total_row, column=col).number_format = MONEY_FMT_2
    _style_total_row(ws2, total_row, len(txn_headers))

    _auto_width(ws2)
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(txn_headers))}{total_row}"

    # ── Sheet 3: По неделям (FULL RETROSPECTIVE) ──────────────
    ws3 = wb.create_sheet("По неделям")
    ws3.sheet_properties.tabColor = "EA580C"

    # Headers: money columns + paired (%, Δ%) columns
    # Each % column is followed by a grey Δ column showing change vs previous week
    wk_base_headers = [
        "Год", "Нед.", "Начало", "Конец", "Кол-во",
        "Продажи", "Возвраты", "Комиссия", "Компенсации",
        "Доставка", "Реклама", "Прочее",
        "Приёмка/FBO", "Эквайринг", "Хранение",
        "К перечисл.", "С/С", "Прибыль",
    ]  # cols 1..18
    # Percentage pairs: (label, Δlabel) — each pair = 2 columns
    wk_pct_pairs = [
        ("Комис%", "Δ"), ("Дост%", "Δ"), ("Рекл%", "Δ"),
        ("Приём%", "Δ"), ("С/С%", "Δ"), ("Приб%", "Δ"),
    ]  # 6 pairs = 12 columns → cols 19..30
    wk_headers = wk_base_headers + [h for pair in wk_pct_pairs for h in pair]
    for col, h in enumerate(wk_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
    _style_header_row(ws3, 1, len(wk_headers))
    # Style Δ headers with grey fill
    for pi in range(len(wk_pct_pairs)):
        delta_col = 19 + pi * 2 + 1  # 20, 22, 24, 26, 28, 30
        c = ws3.cell(row=1, column=delta_col)
        c.fill = DELTA_HDR_FILL
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    # Build COGS per week
    weekly_cogs = {}
    if sku_to_offer and cost_map:
        try:
            cogs_wk_ch = ch.query("""
                SELECT toMonday(toDate(operation_date)) AS ws, sku, countIf(category = 'Revenue') AS qty
                FROM mms_analytics.fact_ozon_transactions FINAL
                WHERE shop_id = {shop_id:UInt32} AND category = 'Revenue' AND sku > 0
                GROUP BY ws, sku
            """, parameters={"shop_id": shop_id})
            for r in cogs_wk_ch.result_rows:
                ws_key = str(r[0])
                sku = int(r[1] or 0)
                qty = int(r[2] or 0)
                offer_id = sku_to_offer.get(sku, "")
                unit_cost = cost_map.get(offer_id, cost_map.get(offer_id.lower() if isinstance(offer_id, str) else "", 0))
                if unit_cost > 0 and qty > 0:
                    weekly_cogs[ws_key] = weekly_cogs.get(ws_key, 0) + unit_cost * qty
        except Exception as e:
            logger.warning("Weekly COGS query failed: %s", e)

    wk_totals = {k: 0.0 for k in [
        "sales", "returns", "commission", "compensations", "delivery",
        "marketing", "other_charges", "fbo_services", "acquiring",
        "storage", "payout", "cogs", "profit"
    ]}
    wk_totals["qty"] = 0

    # Track previous week percentages for delta calculation
    prev_wk_pcts = None  # list of 6 pct values from previous row

    # weekly_data columns: 0=week_start, 1=week_end, 2=qty, 3=sales, 4=returns,
    # 5=commission, 6=compensations, 7=other_services(=delivery), 8=marketing, 9=other_charges,
    # 10=fbo_services, 11=acquiring, 12=delivery_services(≈0, merged into fbo), 13=storage, 14=payout
    for i, r in enumerate(weekly_data.result_rows):
        row_num = i + 2
        ws_date = r[0]
        we_date = r[1]
        iso_cal = ws_date.isocalendar() if hasattr(ws_date, 'isocalendar') else date.fromisoformat(str(ws_date)).isocalendar()

        qty = int(r[2] or 0)
        sales = float(r[3] or 0)
        returns_ = float(r[4] or 0)
        commission_ = float(r[5] or 0)
        compensations_ = float(r[6] or 0)
        delivery_ = float(r[7] or 0)
        marketing_ = float(r[8] or 0)
        other_charges_ = float(r[9] or 0)
        fbo_ = float(r[10] or 0) + float(r[12] or 0)
        acquiring_ = float(r[11] or 0)
        storage_ = float(r[13] or 0)
        payout_ = float(r[14] or 0)
        cogs_wk = weekly_cogs.get(str(ws_date), 0)
        profit_wk = sales - abs(commission_) - abs(fbo_) - abs(delivery_) - abs(acquiring_) - abs(storage_) - abs(marketing_) - abs(other_charges_) - abs(returns_) - abs(compensations_) - cogs_wk

        ws3.cell(row=row_num, column=1, value=iso_cal[0])
        ws3.cell(row=row_num, column=2, value=iso_cal[1])
        ws3.cell(row=row_num, column=3, value=_fmt_date_ru(ws_date))
        ws3.cell(row=row_num, column=4, value=_fmt_date_ru(we_date))
        ws3.cell(row=row_num, column=5, value=qty)

        money_vals = [sales, returns_, commission_, compensations_, delivery_,
                      marketing_, other_charges_, fbo_, acquiring_, storage_,
                      payout_, cogs_wk, profit_wk]
        money_keys = ["sales", "returns", "commission", "compensations", "delivery",
                       "marketing", "other_charges", "fbo_services", "acquiring",
                       "storage", "payout", "cogs", "profit"]

        for j, v in enumerate(money_vals):
            c = ws3.cell(row=row_num, column=6 + j, value=round(v, 2))
            c.number_format = MONEY_FMT
            wk_totals[money_keys[j]] += v

        wk_totals["qty"] += qty

        # Profit coloring
        ws3.cell(row=row_num, column=18).font = GREEN_FONT if profit_wk >= 0 else RED_FONT

        # Percentage columns: Комис%, Дост%, Рекл%, Приём%, С/С%, Приб%
        # Each followed by delta column (grey fill)
        is_alt = (i % 2 == 1)
        if sales > 0:
            cur_pcts = [
                round(commission_ / sales * 100, 1),
                round(delivery_ / sales * 100, 1),
                round(marketing_ / sales * 100, 1),
                round(fbo_ / sales * 100, 1),
                round(cogs_wk / sales * 100, 1),
                round(profit_wk / sales * 100, 1),
            ]
        else:
            cur_pcts = [0, 0, 0, 0, 0, 0]

        for pi, pct_val in enumerate(cur_pcts):
            pct_col = 19 + pi * 2       # % column: 19, 21, 23, 25, 27, 29
            delta_col = 19 + pi * 2 + 1 # Δ column: 20, 22, 24, 26, 28, 30

            # Write % value
            pc = ws3.cell(row=row_num, column=pct_col, value=pct_val)
            pc.number_format = '0.0"%"'
            if pi == 5:  # Profit%: higher is better
                pc.font = GREEN_FONT if pct_val >= 0 else RED_FONT
            else:  # Expense%: lower is better
                pc.font = RED_FONT if pct_val > 35 else (GREEN_FONT if pct_val < 20 else NORMAL_FONT)

            # Write Δ value (change vs previous week in pp)
            dc = ws3.cell(row=row_num, column=delta_col)
            dc.fill = DELTA_FILL_ALT if is_alt else DELTA_FILL
            if prev_wk_pcts is not None:
                delta_pp = round(pct_val - prev_wk_pcts[pi], 1)
                dc.value = delta_pp
                dc.number_format = '+0.0;-0.0;0.0'
                if pi == 5:  # Profit: increase is good
                    dc.font = GREEN_FONT_SM if delta_pp > 0 else (RED_FONT_SM if delta_pp < 0 else GREY_FONT_SM)
                else:  # Expenses: decrease is good
                    dc.font = GREEN_FONT_SM if delta_pp < 0 else (RED_FONT_SM if delta_pp > 0 else GREY_FONT_SM)
            else:
                dc.value = "—"
                dc.font = GREY_FONT_SM

        prev_wk_pcts = cur_pcts

        _style_data_row(ws3, row_num, 18, is_alt=is_alt)  # style only money part

    # Totals
    total_row = len(weekly_data.result_rows) + 2
    ws3.cell(row=total_row, column=1, value="ИТОГО")
    ws3.cell(row=total_row, column=5, value=wk_totals["qty"])
    for j, k in enumerate(money_keys):
        ws3.cell(row=total_row, column=6 + j, value=round(wk_totals[k], 2)).number_format = MONEY_FMT

    ts = wk_totals["sales"] or 1
    total_pcts_wk = [
        round(wk_totals["commission"] / ts * 100, 1),
        round(wk_totals["delivery"] / ts * 100, 1),
        round(wk_totals["marketing"] / ts * 100, 1),
        round(wk_totals["fbo_services"] / ts * 100, 1),
        round(wk_totals["cogs"] / ts * 100, 1),
        round(wk_totals["profit"] / ts * 100, 1),
    ]
    for pi, pv in enumerate(total_pcts_wk):
        pct_col = 19 + pi * 2
        ws3.cell(row=total_row, column=pct_col, value=pv).number_format = '0.0"%"'
    _style_total_row(ws3, total_row, len(wk_headers))

    _auto_width(ws3)
    ws3.freeze_panes = 'A2'

    # ── Sheet 4: По месяцам (FULL RETROSPECTIVE) ──────────────
    ws4 = wb.create_sheet("По месяцам")
    ws4.sheet_properties.tabColor = "7C3AED"

    MONTHS_RU = {1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
                 7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"}

    # Headers: money columns + paired (%, Δ%) columns — same structure as weekly
    mo_base_headers = [
        "Месяц", "Начало", "Конец", "Кол-во",
        "Продажи", "Возвраты", "Комиссия", "Компенсации",
        "Доставка", "Реклама", "Прочее",
        "Приёмка/FBO", "Эквайринг", "Хранение",
        "К перечисл.", "С/С", "Прибыль",
    ]  # cols 1..17
    mo_pct_pairs = [
        ("Комис%", "Δ"), ("Дост%", "Δ"), ("Рекл%", "Δ"),
        ("Приём%", "Δ"), ("С/С%", "Δ"), ("Приб%", "Δ"),
    ]  # 6 pairs = 12 columns → cols 18..29
    mo_headers = mo_base_headers + [h for pair in mo_pct_pairs for h in pair]
    for col, h in enumerate(mo_headers, 1):
        ws4.cell(row=1, column=col, value=h)
    _style_header_row(ws4, 1, len(mo_headers))
    # Style Δ headers with grey fill
    for pi in range(len(mo_pct_pairs)):
        delta_col = 18 + pi * 2 + 1  # 19, 21, 23, 25, 27, 29
        c = ws4.cell(row=1, column=delta_col)
        c.fill = DELTA_HDR_FILL
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    # Build COGS per month
    monthly_cogs = {}
    if sku_to_offer and cost_map:
        try:
            cogs_mo_ch = ch.query("""
                SELECT toYYYYMM(toDate(operation_date)) AS ym, sku, countIf(category = 'Revenue') AS qty
                FROM mms_analytics.fact_ozon_transactions FINAL
                WHERE shop_id = {shop_id:UInt32} AND category = 'Revenue' AND sku > 0
                GROUP BY ym, sku
            """, parameters={"shop_id": shop_id})
            for r in cogs_mo_ch.result_rows:
                ym_key = int(r[0])
                sku = int(r[1] or 0)
                qty = int(r[2] or 0)
                offer_id = sku_to_offer.get(sku, "")
                unit_cost = cost_map.get(offer_id, cost_map.get(offer_id.lower() if isinstance(offer_id, str) else "", 0))
                if unit_cost > 0 and qty > 0:
                    monthly_cogs[ym_key] = monthly_cogs.get(ym_key, 0) + unit_cost * qty
        except Exception as e:
            logger.warning("Monthly COGS query failed: %s", e)

    prev_mo_pcts = None  # list of 6 pct values from previous month

    # monthly_data columns: 0=ym, 1=m_start, 2=m_end, 3=qty, 4=sales, 5=returns,
    # 6=commission, 7=compensations, 8=other_services(=delivery), 9=marketing, 10=other_charges,
    # 11=fbo_services, 12=acquiring, 13=delivery_services(≈0, merged into fbo), 14=storage, 15=payout
    for i, r in enumerate(monthly_data.result_rows):
        row_num = i + 2
        ym = int(r[0])
        year = ym // 100
        month = ym % 100
        month_name = f"{MONTHS_RU.get(month, str(month))} {year}"

        qty = int(r[3] or 0)
        sales_mo = float(r[4] or 0)
        returns_mo = float(r[5] or 0)
        commission_mo = float(r[6] or 0)
        compensations_mo = float(r[7] or 0)
        delivery_mo = float(r[8] or 0)
        marketing_mo = float(r[9] or 0)
        other_charges_mo = float(r[10] or 0)
        fbo_mo = float(r[11] or 0) + float(r[13] or 0)
        acquiring_mo = float(r[12] or 0)
        storage_mo = float(r[14] or 0)
        payout_mo = float(r[15] or 0)
        cogs_mo = monthly_cogs.get(ym, 0)
        profit_mo = sales_mo - abs(commission_mo) - abs(fbo_mo) - abs(delivery_mo) - abs(acquiring_mo) - abs(storage_mo) - abs(marketing_mo) - abs(other_charges_mo) - abs(returns_mo) - abs(compensations_mo) - cogs_mo

        ws4.cell(row=row_num, column=1, value=month_name)
        ws4.cell(row=row_num, column=2, value=str(r[1]))
        ws4.cell(row=row_num, column=3, value=str(r[2]))
        ws4.cell(row=row_num, column=4, value=qty)

        money_vals_mo = [sales_mo, returns_mo, commission_mo, compensations_mo, delivery_mo,
                         marketing_mo, other_charges_mo, fbo_mo, acquiring_mo, storage_mo,
                         payout_mo, cogs_mo, profit_mo]

        for j, v in enumerate(money_vals_mo):
            c = ws4.cell(row=row_num, column=5 + j, value=round(v, 2))
            c.number_format = MONEY_FMT

        # Profit coloring
        ws4.cell(row=row_num, column=17).font = GREEN_FONT if profit_mo >= 0 else RED_FONT

        # Percentage columns: Комис%, Дост%, Рекл%, Приём%, С/С%, Приб%
        # Each followed by delta column (grey fill)
        is_alt = (i % 2 == 1)
        if sales_mo > 0:
            cur_mo_pcts = [
                round(commission_mo / sales_mo * 100, 1),
                round(delivery_mo / sales_mo * 100, 1),
                round(marketing_mo / sales_mo * 100, 1),
                round(fbo_mo / sales_mo * 100, 1),
                round(cogs_mo / sales_mo * 100, 1),
                round(profit_mo / sales_mo * 100, 1),
            ]
        else:
            cur_mo_pcts = [0, 0, 0, 0, 0, 0]

        for pi, pct_val in enumerate(cur_mo_pcts):
            pct_col = 18 + pi * 2       # % column: 18, 20, 22, 24, 26, 28
            delta_col = 18 + pi * 2 + 1 # Δ column: 19, 21, 23, 25, 27, 29

            # Write % value
            pc = ws4.cell(row=row_num, column=pct_col, value=pct_val)
            pc.number_format = '0.0"%"'
            if pi == 5:  # Profit%: higher is better
                pc.font = GREEN_FONT if pct_val >= 0 else RED_FONT
            else:  # Expense%: lower is better
                pc.font = RED_FONT if pct_val > 35 else (GREEN_FONT if pct_val < 20 else NORMAL_FONT)

            # Write Δ value (change vs previous month in pp)
            dc = ws4.cell(row=row_num, column=delta_col)
            dc.fill = DELTA_FILL_ALT if is_alt else DELTA_FILL
            if prev_mo_pcts is not None:
                delta_pp = round(pct_val - prev_mo_pcts[pi], 1)
                dc.value = delta_pp
                dc.number_format = '+0.0;-0.0;0.0'
                if pi == 5:  # Profit: increase is good
                    dc.font = GREEN_FONT_SM if delta_pp > 0 else (RED_FONT_SM if delta_pp < 0 else GREY_FONT_SM)
                else:  # Expenses: decrease is good
                    dc.font = GREEN_FONT_SM if delta_pp < 0 else (RED_FONT_SM if delta_pp > 0 else GREY_FONT_SM)
            else:
                dc.value = "—"
                dc.font = GREY_FONT_SM

        prev_mo_pcts = cur_mo_pcts

        _style_data_row(ws4, row_num, 17, is_alt=is_alt)  # style only money part

    _auto_width(ws4)
    ws4.freeze_panes = 'A2'

    # ── Sheet 5: По товарам (SKU) — расширенный P&L ─────────
    ws5 = wb.create_sheet("По товарам (SKU)")
    ws5.sheet_properties.tabColor = "DB2777"

    sku_headers = [
        "SKU", "Артикул", "Штрих-код", "Название", "Кол-во", "Выручка",
        "Комиссия", "Сервисы", "Логистика", "Реклама", "ДРР %",
        "Хранение", "Нетто (к перечисл.)",
        "С/С", "Прибыль", "Маржа %"
    ]
    for col, h in enumerate(sku_headers, 1):
        ws5.cell(row=1, column=col, value=h)
    _style_header_row(ws5, 1, len(sku_headers))

    # Build sku→cost map for COGS in per-SKU
    sku_cost_map_all = {}
    try:
        if sku_to_offer:
            for sku, offer_id in sku_to_offer.items():
                cost = cost_map.get(offer_id, 0)
                if cost > 0:
                    sku_cost_map_all[sku] = cost
    except Exception:
        pass

    # ── Per-SKU logistics: proportional distribution ──────────
    # Ozon API does NOT link logistics to SKU (sku=0 for all Logistics
    # transactions). We distribute total logistics proportionally by revenue.
    sku_logistics_map = {}
    total_logistics_period = 0.0
    try:
        log_ch = ch.query("""
            SELECT abs(sum(amount)) AS total_logistics
            FROM mms_analytics.fact_ozon_transactions FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND toDate(operation_date) >= {d_start:Date}
              AND toDate(operation_date) <= {d_end:Date}
              AND category = 'Logistics'
        """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": d_end})
        if log_ch.result_rows:
            total_logistics_period = float(log_ch.result_rows[0][0] or 0)
    except Exception:
        pass

    # Build revenue map for proportional logistics distribution
    sku_revenue_for_logistics = {}
    total_sku_revenue = 0.0
    for r in sku_data.result_rows:
        sku = int(r[0] or 0)
        rev = float(r[2] or 0)
        if rev > 0:
            sku_revenue_for_logistics[sku] = rev
            total_sku_revenue += rev

    if total_logistics_period > 0 and total_sku_revenue > 0:
        for sku, rev in sku_revenue_for_logistics.items():
            sku_logistics_map[sku] = total_logistics_period * (rev / total_sku_revenue)

    # ── Per-SKU ad spend from fact_ozon_ad_daily ──────────────
    # Ozon transaction API has sku=0 for Marketing. Real per-SKU ad
    # data comes from Ozon Performance API → fact_ozon_ad_daily.
    from app.services.ozon_finance_queries import get_ad_costs_by_sku, get_placement_costs_by_sku
    sku_ad_map = get_ad_costs_by_sku(ch, shop_id, d_start, d_end)

    # ── Per-SKU storage from fact_ozon_placement_cost ─────────
    # Ozon transaction API has sku=0 for Storage. Real per-SKU storage
    # data comes from Ozon placement report → fact_ozon_placement_cost.
    offer_storage_map = get_placement_costs_by_sku(ch, shop_id, d_start, d_end)

    sku_totals = {"qty": 0, "revenue": 0, "commission": 0, "services": 0,
                  "logistics": 0, "ad_spend": 0, "storage": 0, "payout": 0,
                  "cogs": 0, "profit": 0}

    # Filter out SKUs with zero quantity (no Revenue transactions — only refunds/other)
    sku_rows = [(r_idx, r) for r_idx, r in enumerate(sku_data.result_rows) if int(r[3] or 0) > 0]

    for i, (r_idx, r) in enumerate(sku_rows):
        row_num = i + 2
        sku = int(r[0] or 0)
        name = r[1] or ""
        revenue_sku = float(r[2] or 0)
        qty = int(r[3] or 0)
        comm = float(r[4] or 0)
        svcs = float(r[5] or 0)
        payout_sku = float(r[6] or 0)
        logistics_sku = sku_logistics_map.get(sku, 0)
        ad_sku = sku_ad_map.get(sku, 0)
        offer_id_for_sku = sku_to_offer.get(sku, "")
        barcode_for_sku = sku_to_barcode.get(sku, "")
        storage_sku = offer_storage_map.get(offer_id_for_sku, 0) if offer_id_for_sku else 0
        cogs_sku = sku_cost_map_all.get(sku, 0) * qty
        drr_pct = round(ad_sku / revenue_sku * 100, 1) if revenue_sku > 0 else 0
        # Unified profit formula: revenue - all expenses - cogs
        profit_sku = revenue_sku - abs(comm) - abs(svcs) - logistics_sku - ad_sku - storage_sku - cogs_sku

        ws5.cell(row=row_num, column=1, value=sku)
        ws5.cell(row=row_num, column=2, value=offer_id_for_sku)
        ws5.cell(row=row_num, column=3, value=barcode_for_sku)
        ws5.cell(row=row_num, column=4, value=name)
        ws5.cell(row=row_num, column=5, value=qty)
        ws5.cell(row=row_num, column=6, value=round(revenue_sku, 2)).number_format = MONEY_FMT
        ws5.cell(row=row_num, column=7, value=round(comm, 2)).number_format = MONEY_FMT
        ws5.cell(row=row_num, column=8, value=round(svcs, 2)).number_format = MONEY_FMT
        ws5.cell(row=row_num, column=9, value=round(logistics_sku, 2)).number_format = MONEY_FMT
        ws5.cell(row=row_num, column=10, value=round(ad_sku, 2)).number_format = MONEY_FMT
        drr_c = ws5.cell(row=row_num, column=11, value=drr_pct)
        drr_c.number_format = '0.0"%"'
        drr_c.font = RED_FONT if drr_pct > 30 else (GREEN_FONT if drr_pct < 15 else NORMAL_FONT)
        ws5.cell(row=row_num, column=12, value=round(storage_sku, 2)).number_format = MONEY_FMT
        ws5.cell(row=row_num, column=13, value=round(payout_sku, 2)).number_format = MONEY_FMT
        ws5.cell(row=row_num, column=14, value=round(cogs_sku, 2)).number_format = MONEY_FMT
        pc = ws5.cell(row=row_num, column=15, value=round(profit_sku, 2))
        pc.number_format = MONEY_FMT
        pc.font = GREEN_FONT if profit_sku >= 0 else RED_FONT
        margin = profit_sku / revenue_sku * 100 if revenue_sku > 0 else 0
        mc = ws5.cell(row=row_num, column=16, value=round(margin, 1))
        mc.number_format = '0.0"%"'
        mc.font = GREEN_FONT if margin >= 15 else (RED_FONT if margin < 5 else NORMAL_FONT)

        sku_totals["qty"] += qty
        sku_totals["revenue"] += revenue_sku
        sku_totals["commission"] += comm
        sku_totals["services"] += svcs
        sku_totals["logistics"] += logistics_sku
        sku_totals["ad_spend"] += ad_sku
        sku_totals["storage"] += storage_sku
        sku_totals["payout"] += payout_sku
        sku_totals["cogs"] += cogs_sku
        sku_totals["profit"] += profit_sku

        _style_data_row(ws5, row_num, len(sku_headers), is_alt=(i % 2 == 1))

    # Totals
    total_row = len(sku_rows) + 2
    ws5.cell(row=total_row, column=1, value="ИТОГО")
    ws5.cell(row=total_row, column=4, value=f"{len(sku_rows)} товаров")
    ws5.cell(row=total_row, column=5, value=sku_totals["qty"])
    ws5.cell(row=total_row, column=6, value=round(sku_totals["revenue"], 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=7, value=round(sku_totals["commission"], 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=8, value=round(sku_totals["services"], 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=9, value=round(sku_totals["logistics"], 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=10, value=round(sku_totals["ad_spend"], 2)).number_format = MONEY_FMT
    # DRR% total
    total_drr = round(sku_totals["ad_spend"] / sku_totals["revenue"] * 100, 1) if sku_totals["revenue"] > 0 else 0
    ws5.cell(row=total_row, column=11, value=total_drr).number_format = '0.0"%"'
    ws5.cell(row=total_row, column=12, value=round(sku_totals["storage"], 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=13, value=round(sku_totals["payout"], 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=14, value=round(sku_totals["cogs"], 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=15, value=round(sku_totals["profit"], 2)).number_format = MONEY_FMT
    tm = sku_totals["profit"] / sku_totals["revenue"] * 100 if sku_totals["revenue"] > 0 else 0
    ws5.cell(row=total_row, column=16, value=round(tm, 1)).number_format = '0.0"%"'
    _style_total_row(ws5, total_row, len(sku_headers))

    _auto_width(ws5)
    ws5.freeze_panes = 'A2'
    ws5.auto_filter.ref = f"A1:{get_column_letter(len(sku_headers))}{total_row}"

    # ── Sheet 6: Расходы детально ────────────────────────────
    ws6 = wb.create_sheet("Расходы детально")
    ws6.sheet_properties.tabColor = "D97706"

    exp_headers = [
        "Тип операции (API)", "Описание", "Категория",
        "Кол-во", "Общая сумма", "Средняя сумма", "% от выручки"
    ]
    for col, h in enumerate(exp_headers, 1):
        ws6.cell(row=1, column=col, value=h)
    _style_header_row(ws6, 1, len(exp_headers))

    for i, r in enumerate(expense_detail.result_rows):
        row_num = i + 2
        ws6.cell(row=row_num, column=1, value=r[0])  # operation_type
        ws6.cell(row=row_num, column=2, value=r[1])  # operation_type_name
        ws6.cell(row=row_num, column=3, value=r[2])  # category
        ws6.cell(row=row_num, column=4, value=int(r[3] or 0))
        ws6.cell(row=row_num, column=5, value=float(r[4] or 0)).number_format = MONEY_FMT_2
        ws6.cell(row=row_num, column=6, value=float(r[5] or 0)).number_format = MONEY_FMT_2
        pct = abs(float(r[4] or 0)) / revenue_cur * 100 if revenue_cur > 0 else 0
        ws6.cell(row=row_num, column=7, value=round(pct, 2)).number_format = '0.00"%"'

        _style_data_row(ws6, row_num, len(exp_headers), is_alt=(i % 2 == 1))

    _auto_width(ws6)
    ws6.freeze_panes = 'A2'
    ws6.auto_filter.ref = f"A1:{get_column_letter(len(exp_headers))}{len(expense_detail.result_rows) + 1}"

    # ══════════════════════════════════════════════════════════
    # Save to bytes buffer and return
    # ══════════════════════════════════════════════════════════

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Ozon_Finance_{shop.name}_{d_start}_{d_end}.xlsx"
    encoded = quote(filename)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


# ══════════════════════════════════════════════════════════════
# WB Excel Export
# ══════════════════════════════════════════════════════════════

WB_HEADER_FILL = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")


def _style_wb_header_row(ws, row_num: int, col_count: int):
    """WB purple header styling."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = WB_HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


@router.get("/wb/excel")
async def export_wb_excel(
    shop_id: int = Query(..., description="Shop ID"),
    date_from: date = Query(..., description="Start date"),
    date_to: date = Query(..., description="End date"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Generate and download comprehensive WB Excel financial report."""

    # ── Verify shop ──
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "wildberries":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Shop not found")

    from app.core.clickhouse import get_clickhouse_client

    try:
        ch = get_clickhouse_client()
    except Exception as e:
        logger.error("ClickHouse connection error: %s", e)
        raise HTTPException(status_code=500, detail="Analytics unavailable")

    d_start = date_from
    d_end = date_to
    span = (d_end - d_start).days + 1
    d_prev_start = d_start - timedelta(days=span)
    d_prev_end = d_start - timedelta(days=1)

    # ══════════════════════════════════════════════════════
    # 1. Summary KPIs from fact_finances
    # ══════════════════════════════════════════════════════
    revenue_cur = revenue_prev = 0.0
    payout_cur = payout_prev = 0.0
    commission_cur = commission_prev = 0.0
    logistics_cur = logistics_prev = 0.0
    storage_cur = storage_prev = 0.0
    acquiring_cur = acquiring_prev = 0.0
    acceptance_cur = acceptance_prev = 0.0
    deductions_cur = deductions_prev = 0.0
    deductions_ads_cur = deductions_ads_prev = 0.0
    penalties_cur = penalties_prev = 0.0
    returns_cur = returns_prev = 0.0
    orders_cur = orders_prev = 0

    try:
        fin = ch.query("""
            SELECT
                sumIf(JSONExtractFloat(raw_payload, 'retail_price_withdisc_rub'), operation_type='Продажа' AND event_date>={d_start:Date} AND event_date<={d_end:Date})
                 - sumIf(JSONExtractFloat(raw_payload, 'retail_price_withdisc_rub'), operation_type='Возврат' AND event_date>={d_start:Date} AND event_date<={d_end:Date}) AS rev_cur,
                sumIf(JSONExtractFloat(raw_payload, 'retail_price_withdisc_rub'), operation_type='Продажа' AND event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date})
                 - sumIf(JSONExtractFloat(raw_payload, 'retail_price_withdisc_rub'), operation_type='Возврат' AND event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date}) AS rev_prev,

                sumIf(payout_amount, operation_type='Продажа' AND event_date>={d_start:Date} AND event_date<={d_end:Date})
                 - sumIf(payout_amount, operation_type='Возврат' AND event_date>={d_start:Date} AND event_date<={d_end:Date}) AS pay_cur,
                sumIf(payout_amount, operation_type='Продажа' AND event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date})
                 - sumIf(payout_amount, operation_type='Возврат' AND event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date}) AS pay_prev,

                sumIf(wb_delivery_rub, event_date>={d_start:Date} AND event_date<={d_end:Date}) AS log_cur,
                sumIf(wb_delivery_rub, event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date}) AS log_prev,
                sumIf(storage_fee, event_date>={d_start:Date} AND event_date<={d_end:Date}) AS stor_cur,
                sumIf(storage_fee, event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date}) AS stor_prev,
                sumIf(wb_acquiring, event_date>={d_start:Date} AND event_date<={d_end:Date}) AS acq_cur,
                sumIf(wb_acquiring, event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date}) AS acq_prev,
                sumIf(acceptance_fee, event_date>={d_start:Date} AND event_date<={d_end:Date}) AS acc_cur,
                sumIf(acceptance_fee, event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date}) AS acc_prev,
                sumIf(JSONExtractFloat(raw_payload,'deduction'), event_date>={d_start:Date} AND event_date<={d_end:Date}
                    AND positionCaseInsensitiveUTF8(JSONExtractString(raw_payload,'bonus_type_name'),'продвижение')=0) AS ded_cur,
                sumIf(JSONExtractFloat(raw_payload,'deduction'), event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date}
                    AND positionCaseInsensitiveUTF8(JSONExtractString(raw_payload,'bonus_type_name'),'продвижение')=0) AS ded_prev,
                sumIf(JSONExtractFloat(raw_payload,'deduction'), event_date>={d_start:Date} AND event_date<={d_end:Date}
                    AND positionCaseInsensitiveUTF8(JSONExtractString(raw_payload,'bonus_type_name'),'продвижение')>0) AS ded_ads_cur,
                sumIf(JSONExtractFloat(raw_payload,'deduction'), event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date}
                    AND positionCaseInsensitiveUTF8(JSONExtractString(raw_payload,'bonus_type_name'),'продвижение')>0) AS ded_ads_prev,
                sumIf(penalty_total, event_date>={d_start:Date} AND event_date<={d_end:Date} AND operation_type!='Удержание') AS pen_cur,
                sumIf(penalty_total, event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date} AND operation_type!='Удержание') AS pen_prev,
                sumIf(JSONExtractFloat(raw_payload,'retail_price_withdisc_rub'), operation_type='Возврат' AND event_date>={d_start:Date} AND event_date<={d_end:Date}) AS ret_cur,
                sumIf(JSONExtractFloat(raw_payload,'retail_price_withdisc_rub'), operation_type='Возврат' AND event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date}) AS ret_prev,
                sumIf(quantity, operation_type='Продажа' AND quantity>0 AND event_date>={d_start:Date} AND event_date<={d_end:Date}) AS ord_cur,
                sumIf(quantity, operation_type='Продажа' AND quantity>0 AND event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date}) AS ord_prev
            FROM mms_analytics.fact_finances FINAL
            WHERE shop_id={shop_id:UInt32} AND marketplace=1
              AND event_date>={d_prev_start:Date} AND event_date<={d_end:Date}
        """, parameters={
            "shop_id": shop_id,
            "d_start": d_start, "d_end": d_end,
            "d_prev_start": d_prev_start, "d_prev_end": d_prev_end,
        })
        if fin.result_rows:
            r = fin.result_rows[0]
            revenue_cur, revenue_prev = float(r[0] or 0), float(r[1] or 0)
            payout_cur, payout_prev = float(r[2] or 0), float(r[3] or 0)
            logistics_cur, logistics_prev = abs(float(r[4] or 0)), abs(float(r[5] or 0))
            storage_cur, storage_prev = abs(float(r[6] or 0)), abs(float(r[7] or 0))
            acquiring_cur, acquiring_prev = abs(float(r[8] or 0)), abs(float(r[9] or 0))
            acceptance_cur, acceptance_prev = abs(float(r[10] or 0)), abs(float(r[11] or 0))
            deductions_cur, deductions_prev = abs(float(r[12] or 0)), abs(float(r[13] or 0))
            deductions_ads_cur, deductions_ads_prev = abs(float(r[14] or 0)), abs(float(r[15] or 0))
            penalties_cur, penalties_prev = abs(float(r[16] or 0)), abs(float(r[17] or 0))
            returns_cur, returns_prev = abs(float(r[18] or 0)), abs(float(r[19] or 0))
            orders_cur, orders_prev = int(r[20] or 0), int(r[21] or 0)
            commission_cur = max(revenue_cur - payout_cur, 0)
            commission_prev = max(revenue_prev - payout_prev, 0)
    except Exception as e:
        logger.warning("WB Excel KPI query failed: %s", e)

    # 2. Ad spend from fact_advert_stats_v3
    ad_spend_cur = ad_spend_prev = 0.0
    try:
        ads = ch.query("""
            SELECT
                sumIf(spend, date>={d_start:Date} AND date<={d_end:Date}) AS cur,
                sumIf(spend, date>={d_prev_start:Date} AND date<={d_prev_end:Date}) AS prev
            FROM mms_analytics.fact_advert_stats_v3 FINAL
            WHERE shop_id={shop_id:UInt32}
              AND date>={d_prev_start:Date} AND date<={d_end:Date}
        """, parameters={
            "shop_id": shop_id,
            "d_start": d_start, "d_end": d_end,
            "d_prev_start": d_prev_start, "d_prev_end": d_prev_end,
        })
        if ads.result_rows:
            ad_spend_cur = float(ads.result_rows[0][0] or 0)
            ad_spend_prev = float(ads.result_rows[0][1] or 0)
    except Exception:
        pass

    # 2b. MAX-reconciliation: use the higher of fact_finances vs fact_advert_stats_v3
    # fact_finances misses ad spend when WB ad balance is topped up manually
    deductions_ads_cur = max(deductions_ads_cur, ad_spend_cur)
    deductions_ads_prev = max(deductions_ads_prev, ad_spend_prev)

    # 3. COGS
    cogs_cur = cogs_prev = 0.0
    cost_map = {}
    try:
        cost_result = await db.execute(
            text("""
                SELECT offer_id, COALESCE(cost_price,0)+COALESCE(packaging_cost,0) AS total_cost
                FROM product_costs WHERE shop_id=:shop_id AND (cost_price>0 OR packaging_cost>0)
            """), {"shop_id": shop_id},
        )
        cost_map = {r[0].lower(): float(r[1]) for r in cost_result.fetchall()}

        if cost_map:
            cogs_ch = ch.query("""
                SELECT
                    vendor_code,
                    sumIf(quantity, operation_type='Продажа' AND event_date>={d_start:Date} AND event_date<={d_end:Date})
                     - sumIf(quantity, operation_type='Возврат' AND event_date>={d_start:Date} AND event_date<={d_end:Date}) AS qty_cur,
                    sumIf(quantity, operation_type='Продажа' AND event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date})
                     - sumIf(quantity, operation_type='Возврат' AND event_date>={d_prev_start:Date} AND event_date<={d_prev_end:Date}) AS qty_prev
                FROM mms_analytics.fact_finances FINAL
                WHERE shop_id={shop_id:UInt32} AND marketplace=1
                  AND operation_type IN ('Продажа','Возврат') AND vendor_code!=''
                  AND event_date>={d_prev_start:Date} AND event_date<={d_end:Date}
                GROUP BY vendor_code
            """, parameters={
                "shop_id": shop_id,
                "d_start": d_start, "d_end": d_end,
                "d_prev_start": d_prev_start, "d_prev_end": d_prev_end,
            })
            for r in cogs_ch.result_rows:
                vc = str(r[0] or "").lower()
                unit = cost_map.get(vc, 0)
                if unit > 0:
                    cogs_cur += unit * max(0, int(r[1] or 0))
                    cogs_prev += unit * max(0, int(r[2] or 0))
    except Exception:
        pass

    # Derived
    # NOTE: For WB, ВБ Промо (deductions_ads) IS the advertising — no separate external ads.
    # bank = payout minus all fees & deductions (including WB Promo)
    bank_cur = payout_cur - logistics_cur - storage_cur - acceptance_cur - deductions_cur - deductions_ads_cur - penalties_cur
    bank_prev = payout_prev - logistics_prev - storage_prev - acceptance_prev - deductions_prev - deductions_ads_prev - penalties_prev
    profit_cur = bank_cur - cogs_cur
    profit_prev = bank_prev - cogs_prev

    # ══════════════════════════════════════════════════════
    # 4. Daily dynamics
    # ══════════════════════════════════════════════════════
    daily_data = []
    try:
        dd = ch.query("""
            SELECT
                event_date AS dt,
                sumIf(JSONExtractFloat(raw_payload,'retail_price_withdisc_rub'), operation_type='Продажа')
                 - sumIf(JSONExtractFloat(raw_payload,'retail_price_withdisc_rub'), operation_type='Возврат') AS rev,
                sumIf(payout_amount, operation_type='Продажа') - sumIf(payout_amount, operation_type='Возврат') AS pay,
                sum(wb_delivery_rub) AS log,
                sum(storage_fee) AS stor,
                sum(wb_acquiring) AS acq,
                sum(acceptance_fee) AS acc,
                sumIf(JSONExtractFloat(raw_payload,'deduction'),
                    positionCaseInsensitiveUTF8(JSONExtractString(raw_payload,'bonus_type_name'),'продвижение')=0) AS ded,
                sumIf(JSONExtractFloat(raw_payload,'deduction'),
                    positionCaseInsensitiveUTF8(JSONExtractString(raw_payload,'bonus_type_name'),'продвижение')>0) AS ded_ads,
                sumIf(quantity, operation_type='Продажа' AND quantity>0) AS orders
            FROM mms_analytics.fact_finances FINAL
            WHERE shop_id={shop_id:UInt32} AND marketplace=1
              AND event_date>={d_start:Date} AND event_date<={d_end:Date}
            GROUP BY dt ORDER BY dt
        """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": d_end})
        daily_data = dd.result_rows
    except Exception:
        pass

    # 4b. Daily ad spend from fact_advert_stats_v3 for MAX-reconciliation
    daily_ad_spend = {}  # date_str -> spend
    try:
        dad = ch.query("""
            SELECT date, sum(spend) AS s
            FROM mms_analytics.fact_advert_stats_v3 FINAL
            WHERE shop_id={shop_id:UInt32}
              AND date>={d_start:Date} AND date<={d_end:Date}
            GROUP BY date
        """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": d_end})
        for r in dad.result_rows:
            daily_ad_spend[str(r[0])] = float(r[1] or 0)
    except Exception:
        pass

    # 5. Weekly FULL RETROSPECTIVE
    weekly_data = []
    try:
        wd = ch.query("""
            SELECT
                toMonday(event_date) AS ws,
                toMonday(event_date)+6 AS we,
                sumIf(quantity, operation_type='Продажа' AND quantity>0) AS qty,
                sumIf(JSONExtractFloat(raw_payload,'retail_price_withdisc_rub'), operation_type='Продажа')
                 - sumIf(JSONExtractFloat(raw_payload,'retail_price_withdisc_rub'), operation_type='Возврат') AS rev,
                sumIf(payout_amount, operation_type='Продажа') - sumIf(payout_amount, operation_type='Возврат') AS pay,
                sumIf(JSONExtractFloat(raw_payload,'retail_price_withdisc_rub'), operation_type='Возврат') AS rets,
                sum(wb_delivery_rub) AS log,
                sum(storage_fee) AS stor,
                sum(wb_acquiring) AS acq,
                sum(acceptance_fee) AS acc,
                sumIf(JSONExtractFloat(raw_payload,'deduction'),
                    positionCaseInsensitiveUTF8(JSONExtractString(raw_payload,'bonus_type_name'),'продвижение')=0) AS ded,
                sumIf(JSONExtractFloat(raw_payload,'deduction'),
                    positionCaseInsensitiveUTF8(JSONExtractString(raw_payload,'bonus_type_name'),'продвижение')>0) AS ded_ads
            FROM mms_analytics.fact_finances FINAL
            WHERE shop_id={shop_id:UInt32} AND marketplace=1
            GROUP BY ws ORDER BY ws
        """, parameters={"shop_id": shop_id})
        weekly_data = wd.result_rows
    except Exception:
        pass

    # 5b. Weekly ad spend from fact_advert_stats_v3 for MAX-reconciliation
    weekly_ad_spend = {}  # monday_str -> spend
    try:
        wad = ch.query("""
            SELECT toMonday(date) AS ws, sum(spend) AS s
            FROM mms_analytics.fact_advert_stats_v3 FINAL
            WHERE shop_id={shop_id:UInt32}
            GROUP BY ws
        """, parameters={"shop_id": shop_id})
        for r in wad.result_rows:
            weekly_ad_spend[str(r[0])] = float(r[1] or 0)
    except Exception:
        pass

    # Weekly COGS
    cogs_by_week = {}
    if cost_map:
        try:
            cw = ch.query("""
                SELECT toMonday(event_date) AS ws, vendor_code,
                    sumIf(quantity, operation_type='Продажа') - sumIf(quantity, operation_type='Возврат') AS qty
                FROM mms_analytics.fact_finances FINAL
                WHERE shop_id={shop_id:UInt32} AND marketplace=1
                  AND operation_type IN ('Продажа','Возврат') AND vendor_code!=''
                GROUP BY ws, vendor_code
            """, parameters={"shop_id": shop_id})
            for r in cw.result_rows:
                ws_key = str(r[0])
                vc = str(r[1] or "").lower()
                qty = max(0, int(r[2] or 0))
                unit = cost_map.get(vc, 0)
                if unit > 0 and qty > 0:
                    cogs_by_week[ws_key] = cogs_by_week.get(ws_key, 0) + unit * qty
        except Exception:
            pass

    # 6. Monthly FULL RETROSPECTIVE
    monthly_data = []
    try:
        md = ch.query("""
            SELECT
                toYYYYMM(event_date) AS ym,
                min(event_date) AS m_start, max(event_date) AS m_end,
                sumIf(quantity, operation_type='Продажа' AND quantity>0) AS qty,
                sumIf(JSONExtractFloat(raw_payload,'retail_price_withdisc_rub'), operation_type='Продажа')
                 - sumIf(JSONExtractFloat(raw_payload,'retail_price_withdisc_rub'), operation_type='Возврат') AS rev,
                sumIf(payout_amount, operation_type='Продажа') - sumIf(payout_amount, operation_type='Возврат') AS pay,
                sumIf(JSONExtractFloat(raw_payload,'retail_price_withdisc_rub'), operation_type='Возврат') AS rets,
                sum(wb_delivery_rub) AS log,
                sum(storage_fee) AS stor,
                sum(wb_acquiring) AS acq,
                sum(acceptance_fee) AS acc,
                sumIf(JSONExtractFloat(raw_payload,'deduction'),
                    positionCaseInsensitiveUTF8(JSONExtractString(raw_payload,'bonus_type_name'),'продвижение')=0) AS ded,
                sumIf(JSONExtractFloat(raw_payload,'deduction'),
                    positionCaseInsensitiveUTF8(JSONExtractString(raw_payload,'bonus_type_name'),'продвижение')>0) AS ded_ads
            FROM mms_analytics.fact_finances FINAL
            WHERE shop_id={shop_id:UInt32} AND marketplace=1
            GROUP BY ym ORDER BY ym
        """, parameters={"shop_id": shop_id})
        monthly_data = md.result_rows
    except Exception:
        pass

    # 6b. Monthly ad spend from fact_advert_stats_v3 for MAX-reconciliation
    monthly_ad_spend = {}  # yyyymm_int -> spend
    try:
        mad = ch.query("""
            SELECT toYYYYMM(date) AS ym, sum(spend) AS s
            FROM mms_analytics.fact_advert_stats_v3 FINAL
            WHERE shop_id={shop_id:UInt32}
            GROUP BY ym
        """, parameters={"shop_id": shop_id})
        for r in mad.result_rows:
            monthly_ad_spend[int(r[0])] = float(r[1] or 0)
    except Exception:
        pass


    # Monthly COGS
    cogs_by_month = {}
    if cost_map:
        try:
            cm = ch.query("""
                SELECT toYYYYMM(event_date) AS ym, vendor_code,
                    sumIf(quantity, operation_type='Продажа') - sumIf(quantity, operation_type='Возврат') AS qty
                FROM mms_analytics.fact_finances FINAL
                WHERE shop_id={shop_id:UInt32} AND marketplace=1
                  AND operation_type IN ('Продажа','Возврат') AND vendor_code!=''
                GROUP BY ym, vendor_code
            """, parameters={"shop_id": shop_id})
            for r in cm.result_rows:
                ym_key = int(r[0])
                vc = str(r[1] or "").lower()
                qty = max(0, int(r[2] or 0))
                unit = cost_map.get(vc, 0)
                if unit > 0 and qty > 0:
                    cogs_by_month[ym_key] = cogs_by_month.get(ym_key, 0) + unit * qty
        except Exception:
            pass

    # 7. Per-SKU data from fact_finances (base: qty, rev, pay, log, acc)
    #    Ad spend from fact_advert_stats_v3, storage from fact_wb_paid_storage
    sku_data = {}
    nm_to_vc = {}  # nm_id → vendor_code mapping for ad spend
    try:
        sku_q = ch.query("""
            SELECT
                vendor_code,
                JSONExtractUInt(raw_payload, 'nm_id') AS nm_id,
                sumIf(quantity, operation_type='Продажа' AND quantity>0) AS qty,
                sumIf(JSONExtractFloat(raw_payload,'retail_price_withdisc_rub'), operation_type='Продажа')
                 - sumIf(JSONExtractFloat(raw_payload,'retail_price_withdisc_rub'), operation_type='Возврат') AS rev,
                sumIf(payout_amount, operation_type='Продажа') - sumIf(payout_amount, operation_type='Возврат') AS pay,
                sum(wb_delivery_rub) AS log,
                sum(acceptance_fee) AS acc
            FROM mms_analytics.fact_finances FINAL
            WHERE shop_id={shop_id:UInt32} AND marketplace=1
              AND event_date>={d_start:Date} AND event_date<={d_end:Date}
              AND vendor_code!=''
            GROUP BY vendor_code, nm_id ORDER BY rev DESC
        """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": d_end})
        for r in sku_q.result_rows:
            vc = str(r[0] or "")
            nm = int(r[1] or 0)
            if nm and vc:
                nm_to_vc[nm] = vc
            if vc not in sku_data:
                sku_data[vc] = {
                    "qty": 0, "revenue": 0, "payout": 0,
                    "logistics": 0, "storage": 0, "acquiring": 0,
                    "acceptance": 0, "deductions": 0, "ad_spend": 0,
                }
            sku_data[vc]["qty"] += int(r[2] or 0)
            sku_data[vc]["revenue"] += float(r[3] or 0)
            sku_data[vc]["payout"] += float(r[4] or 0)
            sku_data[vc]["logistics"] += abs(float(r[5] or 0))
            sku_data[vc]["acceptance"] += abs(float(r[6] or 0))
    except Exception:
        pass

    # 7b. Per-SKU ad spend from fact_advert_stats_v3 — ACTUAL spend per nm_id
    try:
        sku_ads = ch.query("""
            SELECT nm_id, sum(spend) AS s
            FROM mms_analytics.fact_advert_stats_v3 FINAL
            WHERE shop_id={shop_id:UInt32}
              AND date>={d_start:Date} AND date<={d_end:Date}
            GROUP BY nm_id
        """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": d_end})
        unmatched_ads = 0.0
        for r in sku_ads.result_rows:
            nm = int(r[0] or 0)
            spend = float(r[1] or 0)
            vc = nm_to_vc.get(nm)
            if vc and vc in sku_data:
                sku_data[vc]["ad_spend"] += spend
            else:
                unmatched_ads += spend
        if unmatched_ads > 0:
            sku_data["__реклама_без_привязки__"] = {
                "qty": 0, "revenue": 0, "payout": 0,
                "logistics": 0, "storage": 0, "acquiring": 0,
                "acceptance": 0, "deductions": 0, "ad_spend": unmatched_ads,
            }
    except Exception:
        pass

    # 7c. Per-SKU storage from fact_wb_paid_storage — ACTUAL storage per vendor_code
    #     (WB doesn't bind storage_fee to vendor_code in fact_finances — always 0)
    try:
        paid_storage_q = ch.query("""
            SELECT
                vendor_code,
                round(SUM(warehouse_price), 2) AS storage_total
            FROM mms_analytics.fact_wb_paid_storage FINAL
            WHERE shop_id={shop_id:UInt32}
              AND dt>={d_start:Date} AND dt<={d_end:Date}
              AND vendor_code != ''
            GROUP BY vendor_code
            HAVING storage_total != 0
        """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": d_end})
        sku_lower_map = {k.lower(): k for k in sku_data}
        for r in paid_storage_q.result_rows:
            vc_raw = str(r[0] or "").strip()
            stor_val = abs(float(r[1] or 0))
            if not vc_raw or stor_val <= 0:
                continue
            original_key = sku_lower_map.get(vc_raw.lower())
            if original_key:
                sku_data[original_key]["storage"] = stor_val
    except Exception:
        pass

    # Product names from dim_products (PG)
    names_map = {}
    barcode_map = {}  # vendor_code -> barcode/GTIN
    try:
        nm = await db.execute(
            text("SELECT vendor_code, name FROM dim_products WHERE shop_id=:shop_id"),
            {"shop_id": shop_id},
        )
        for r in nm.fetchall():
            if r[1]:
                names_map[r[0]] = r[1]
    except Exception:
        pass

    # Fallback: get names + barcode from fact_finances raw_payload
    # (dim_products may be empty for WB shops)
    # Build case-insensitive lookup from dim_products names first
    names_lower = {k.lower(): v for k, v in names_map.items()}
    try:
        names_q = ch.query("""
            SELECT
                vendor_code,
                any(JSONExtractString(raw_payload, 'subject_name')) AS subject,
                any(JSONExtractString(raw_payload, 'sa_name')) AS sa_name,
                any(JSONExtractString(raw_payload, 'barcode')) AS barcode
            FROM mms_analytics.fact_finances FINAL
            WHERE shop_id={shop_id:UInt32} AND marketplace=1
              AND event_date>={d_start:Date} AND event_date<={d_end:Date}
              AND vendor_code != ''
            GROUP BY vendor_code
        """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": d_end})
        for r in names_q.result_rows:
            vc = str(r[0] or "").strip()
            subject = str(r[1] or "").strip()
            sa_name = str(r[2] or "").strip()
            barcode = str(r[3] or "").strip()
            # Only use fallback name if dim_products doesn't have it (case-insensitive)
            if vc and vc.lower() not in names_lower:
                # subject_name is category (e.g. "Корм сухой"), sa_name is vendor_code
                # Use as last resort only
                name = subject if subject else sa_name
                if name and name.lower() != vc.lower():
                    names_map[vc] = name
                    names_lower[vc.lower()] = name
            if vc and barcode:
                barcode_map[vc] = barcode
    except Exception:
        pass

    # Build case-insensitive lookup for barcodes
    barcode_lower = {k.lower(): v for k, v in barcode_map.items()}

    # 8. Expense detail (for "Расходы детально" sheet)
    # NOTE: penalty_total for 'Удержание' ops duplicates deduction — use raw penalty only
    # Exclude Продажа/Возврат — those are revenue operations, not expenses
    expense_detail_rows = []
    try:
        exp = ch.query("""
            SELECT
                operation_type,
                JSONExtractString(raw_payload, 'bonus_type_name') AS bonus_type,
                count() AS cnt,
                sum(JSONExtractFloat(raw_payload, 'delivery_rub')) AS delivery_total,
                sum(storage_fee) AS storage_total,
                sum(acceptance_fee) AS acceptance_total,
                sumIf(penalty_total, operation_type != 'Удержание') AS penalty_clean,
                sum(JSONExtractFloat(raw_payload, 'deduction')) AS deduction_total,
                sum(wb_acquiring) AS acquiring_total
            FROM mms_analytics.fact_finances FINAL
            WHERE shop_id={shop_id:UInt32} AND marketplace=1
              AND event_date>={d_start:Date} AND event_date<={d_end:Date}
              AND operation_type NOT IN ('Продажа', 'Возврат')
            GROUP BY operation_type, bonus_type
            ORDER BY operation_type, bonus_type
        """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": d_end})
        expense_detail_rows = exp.result_rows
    except Exception:
        pass

    # ══════════════════════════════════════════════════════
    # BUILD WORKBOOK
    # ══════════════════════════════════════════════════════
    wb = Workbook()

    # ── Sheet 1: Сводка (секционный P&L как Ozon) ──
    ws1 = wb.active
    ws1.title = "Сводка"
    ws1.sheet_properties.tabColor = "7C3AED"

    # WB-specific styles for sections (purple theme)
    WB_SECTION_HDR_FILL = PatternFill(start_color="5B21B6", end_color="5B21B6", fill_type="solid")
    WB_SECTION_HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    WB_SUBTOTAL_FILL = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    WB_SUBTOTAL_FONT = Font(name="Calibri", bold=True, size=10)
    WB_SUB_ITEM_FONT = Font(name="Calibri", size=10, color="4B5563")
    WB_INDENT = "    "

    def _wb_section_header(ws, row, label, cols=5):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        ws.cell(row=row, column=1, value=label).font = WB_SECTION_HDR_FONT
        for c in range(1, cols + 1):
            ws.cell(row=row, column=c).fill = WB_SECTION_HDR_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER

    def _wb_kpi_row(ws, row, label, cur, prev, is_money=True, is_sub=False, is_total=False, is_profit=False, cols=5):
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=round(cur, 2) if is_money else cur)
        c3 = ws.cell(row=row, column=3, value=round(prev, 2) if is_money else prev)
        delta = _safe_delta(cur, prev)
        c4 = ws.cell(row=row, column=4, value=f"{'+' if delta > 0 else ''}{delta}%")
        pct_cur = cur / revenue_cur * 100 if revenue_cur > 0 and is_money else 0
        c5 = ws.cell(row=row, column=5, value=round(pct_cur, 1) if is_money else "")
        if is_money:
            c2.number_format = MONEY_FMT
            c3.number_format = MONEY_FMT
        if is_money and c5.value != "":
            c5.number_format = '0.0"%"'
        for c in [c2, c3, c4, c5]:
            c.alignment = Alignment(horizontal='right')
        if delta > 0:
            c4.font = GREEN_FONT
        elif delta < 0:
            c4.font = RED_FONT
        if is_sub:
            c1.font = WB_SUB_ITEM_FONT
        elif is_total:
            _style_total_row(ws, row, cols)
        elif is_profit:
            _style_total_row(ws, row, cols)
            c2.font = PROFIT_GREEN if cur >= 0 else PROFIT_RED
            c3.font = PROFIT_GREEN if prev >= 0 else PROFIT_RED
        else:
            c1.font = NORMAL_FONT
        if not is_total and not is_profit:
            _style_data_row(ws, row, cols)

    # Title
    ws1.merge_cells('A1:E1')
    ws1['A1'] = f"Финансовый отчёт WB — {shop.name}"
    ws1['A1'].font = Font(name="Calibri", bold=True, size=16, color="7C3AED")
    ws1.merge_cells('A2:E2')
    ws1['A2'] = f"Период: {d_start.strftime('%d.%m.%Y')} — {d_end.strftime('%d.%m.%Y')} ({span} дн.)"
    ws1['A2'].font = Font(name="Calibri", size=11, color="6B7280")

    # Column headers
    row = 4
    for ci, hdr in enumerate(["Показатель", "Текущий период", "Предыдущий период", "Изм. %", "% выр."], 1):
        ws1.cell(row=row, column=ci, value=hdr)
    _style_wb_header_row(ws1, row, 5)

    row = 5

    # Derived for summary
    avg_check_cur = revenue_cur / orders_cur if orders_cur > 0 else 0
    avg_check_prev = revenue_prev / orders_prev if orders_prev > 0 else 0
    # Total MP fees for WB: commission + logistics + storage + acceptance + deductions + deductions_ads + penalties
    total_mp_fees_cur = commission_cur + logistics_cur + storage_cur + acceptance_cur + deductions_cur + deductions_ads_cur + penalties_cur
    total_mp_fees_prev = commission_prev + logistics_prev + storage_prev + acceptance_prev + deductions_prev + deductions_ads_prev + penalties_prev
    opex_cur = storage_cur + acceptance_cur + deductions_cur + penalties_cur
    opex_prev = storage_prev + acceptance_prev + deductions_prev + penalties_prev
    total_expenses_cur = total_mp_fees_cur + cogs_cur
    total_expenses_prev = total_mp_fees_prev + cogs_prev

    # ═══ ВЫРУЧКА И ЗАКАЗЫ ═══
    _wb_section_header(ws1, row, "ВЫРУЧКА И ЗАКАЗЫ"); row += 1
    _wb_kpi_row(ws1, row, "Выручка (продажи)", revenue_cur, revenue_prev); row += 1
    _wb_kpi_row(ws1, row, "Возвраты", returns_cur, returns_prev); row += 1
    _wb_kpi_row(ws1, row, "Заказы", orders_cur, orders_prev, is_money=False); row += 1
    _wb_kpi_row(ws1, row, "Средний чек", avg_check_cur, avg_check_prev); row += 1

    # ═══ КОМИССИЯ + СКИДКИ ═══
    row += 1
    _wb_section_header(ws1, row, "КОМИССИЯ + СКИДКИ"); row += 1
    _wb_kpi_row(ws1, row, "Комиссия + SPP скидки", commission_cur, commission_prev); row += 1
    _wb_kpi_row(ws1, row, "Эквайринг", acquiring_cur, acquiring_prev); row += 1

    # ═══ ЛОГИСТИКА ═══
    row += 1
    _wb_section_header(ws1, row, "ЛОГИСТИКА"); row += 1
    _wb_kpi_row(ws1, row, "Доставка / логистика", logistics_cur, logistics_prev); row += 1
    _wb_kpi_row(ws1, row, "Итого логистика", logistics_cur, logistics_prev, is_total=True); row += 1

    # ═══ ОПЕРАЦИОННЫЕ РАСХОДЫ ═══
    row += 1
    _wb_section_header(ws1, row, "ОПЕРАЦИОННЫЕ РАСХОДЫ"); row += 1
    _wb_kpi_row(ws1, row, "Хранение", storage_cur, storage_prev); row += 1
    _wb_kpi_row(ws1, row, "Платная приёмка", acceptance_cur, acceptance_prev); row += 1
    _wb_kpi_row(ws1, row, "Удержания (прочие)", deductions_cur, deductions_prev); row += 1
    if penalties_cur > 0 or penalties_prev > 0:
        _wb_kpi_row(ws1, row, "Штрафы", penalties_cur, penalties_prev); row += 1
    _wb_kpi_row(ws1, row, "Итого операционные", opex_cur, opex_prev, is_total=True); row += 1

    # ═══ РЕКЛАМА ═══
    row += 1
    _wb_section_header(ws1, row, "РЕКЛАМА (ВБ ПРОМО)"); row += 1
    _wb_kpi_row(ws1, row, "Итого реклама", deductions_ads_cur, deductions_ads_prev); row += 1
    # Show source breakdown for transparency if there's a discrepancy
    if ad_spend_cur > 0 or ad_spend_prev > 0:
        _wb_kpi_row(ws1, row, f"{WB_INDENT}Факт. расход (API кампаний)", ad_spend_cur, ad_spend_prev, is_sub=True); row += 1
    drr_cur = deductions_ads_cur / revenue_cur * 100 if revenue_cur > 0 else 0
    drr_prev = deductions_ads_prev / revenue_prev * 100 if revenue_prev > 0 else 0
    ws1.cell(row=row, column=1, value=f"{WB_INDENT}ДРР (доля рекл. расх.)").font = WB_SUB_ITEM_FONT
    ws1.cell(row=row, column=2, value=f"{round(drr_cur, 1)}%").alignment = Alignment(horizontal='right')
    ws1.cell(row=row, column=3, value=f"{round(drr_prev, 1)}%").alignment = Alignment(horizontal='right')
    _style_data_row(ws1, row, 5)
    row += 1

    # ═══ ИТОГИ ═══
    row += 1
    _wb_section_header(ws1, row, "ИТОГИ"); row += 1
    _wb_kpi_row(ws1, row, "Всего удержано WB", total_mp_fees_cur, total_mp_fees_prev, is_total=True); row += 1
    _wb_kpi_row(ws1, row, "К перечислению", payout_cur, payout_prev); row += 1
    _wb_kpi_row(ws1, row, "Себестоимость (COGS)", cogs_cur, cogs_prev); row += 1
    _wb_kpi_row(ws1, row, "Всего расходов", total_expenses_cur, total_expenses_prev, is_total=True); row += 1
    _wb_kpi_row(ws1, row, "ЧИСТАЯ ПРИБЫЛЬ", profit_cur, profit_prev, is_profit=True); row += 1
    margin_cur = profit_cur / revenue_cur * 100 if revenue_cur > 0 else 0
    margin_prev = profit_prev / revenue_prev * 100 if revenue_prev > 0 else 0
    ws1.cell(row=row, column=1, value=f"{WB_INDENT}Маржинальность").font = WB_SUB_ITEM_FONT
    ws1.cell(row=row, column=2, value=f"{round(margin_cur, 1)}%").alignment = Alignment(horizontal='right')
    ws1.cell(row=row, column=3, value=f"{round(margin_prev, 1)}%").alignment = Alignment(horizontal='right')
    _style_data_row(ws1, row, 5)
    row += 2

    # ═══ СТРУКТУРА РАСХОДОВ (% от выручки) ═══
    ws1.cell(row=row, column=1, value="СТРУКТУРА РАСХОДОВ (% от выручки)").font = WB_SECTION_HDR_FONT
    ws1.cell(row=row, column=1).fill = WB_SECTION_HDR_FILL
    for ci, hdr in enumerate(["", "Тек. период", "% выр.", "Пред. период", "% выр."], 1):
        ws1.cell(row=row, column=ci, value=hdr if ci > 1 else "СТРУКТУРА РАСХОДОВ (% от выручки)")
    _style_wb_header_row(ws1, row, 5)
    row += 1

    pct_rows = [
        ("Комиссия + скидки", commission_cur, commission_prev),
        ("Логистика", logistics_cur, logistics_prev),
        ("Хранение", storage_cur, storage_prev),
        ("Приёмка", acceptance_cur, acceptance_prev),
        ("Удержания (прочие)", deductions_cur, deductions_prev),
        ("ВБ Промо", deductions_ads_cur, deductions_ads_prev),
        ("Штрафы", penalties_cur, penalties_prev),
        ("Себестоимость", cogs_cur, cogs_prev),
    ]

    total_exp_c = sum(v[1] for v in pct_rows)
    total_exp_p = sum(v[2] for v in pct_rows)

    for i, (label, cur_val, prev_val) in enumerate(pct_rows):
        pct_c = cur_val / revenue_cur * 100 if revenue_cur > 0 else 0
        pct_p = prev_val / revenue_prev * 100 if revenue_prev > 0 else 0
        ws1.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws1.cell(row=row, column=2, value=round(cur_val, 2)).number_format = MONEY_FMT
        ws1.cell(row=row, column=3, value=round(pct_c, 1))
        ws1.cell(row=row, column=3).number_format = '0.0"%"'
        ws1.cell(row=row, column=4, value=round(prev_val, 2)).number_format = MONEY_FMT
        ws1.cell(row=row, column=5, value=round(pct_p, 1))
        ws1.cell(row=row, column=5).number_format = '0.0"%"'
        for c in range(2, 6):
            ws1.cell(row=row, column=c).alignment = Alignment(horizontal='right')
        _style_data_row(ws1, row, 5, is_alt=(i % 2 == 1))
        row += 1

    # Total expenses row
    pct_tc = total_exp_c / revenue_cur * 100 if revenue_cur > 0 else 0
    pct_tp = total_exp_p / revenue_prev * 100 if revenue_prev > 0 else 0
    ws1.cell(row=row, column=1, value="Итого расходов").font = TOTAL_FONT
    ws1.cell(row=row, column=2, value=round(total_exp_c, 2)).number_format = MONEY_FMT
    ws1.cell(row=row, column=3, value=round(pct_tc, 1))
    ws1.cell(row=row, column=3).number_format = '0.0"%"'
    ws1.cell(row=row, column=4, value=round(total_exp_p, 2)).number_format = MONEY_FMT
    ws1.cell(row=row, column=5, value=round(pct_tp, 1))
    ws1.cell(row=row, column=5).number_format = '0.0"%"'
    _style_total_row(ws1, row, 5)
    row += 1

    # Profit row
    pct_prof_c = profit_cur / revenue_cur * 100 if revenue_cur > 0 else 0
    pct_prof_p = profit_prev / revenue_prev * 100 if revenue_prev > 0 else 0
    ws1.cell(row=row, column=1, value="Прибыль").font = TOTAL_FONT
    c2 = ws1.cell(row=row, column=2, value=round(profit_cur, 2))
    c2.number_format = MONEY_FMT
    c2.font = PROFIT_GREEN if profit_cur >= 0 else PROFIT_RED
    ws1.cell(row=row, column=3, value=round(pct_prof_c, 1))
    ws1.cell(row=row, column=3).number_format = '0.0"%"'
    c4 = ws1.cell(row=row, column=4, value=round(profit_prev, 2))
    c4.number_format = MONEY_FMT
    c4.font = PROFIT_GREEN if profit_prev >= 0 else PROFIT_RED
    ws1.cell(row=row, column=5, value=round(pct_prof_p, 1))
    ws1.cell(row=row, column=5).number_format = '0.0"%"'
    _style_total_row(ws1, row, 5)

    _auto_width(ws1)
    ws1.column_dimensions['A'].width = 32
    ws1.freeze_panes = 'A5'


    # ── Sheet 2: По неделям (FULL RETROSPECTIVE with Δ-columns) ──
    ws3 = wb.create_sheet("По неделям")
    ws3.sheet_properties.tabColor = "EA580C"

    # Headers: money columns + paired (%, Δ%) columns
    wk_base_headers = [
        "Год", "Нед.", "Начало", "Конец", "Кол-во",
        "Выручка", "К перечисл.", "Возвраты", "Логистика", "Хранение",
        "Приёмка", "Удержания", "ВБ Промо",
        "С/С", "Прибыль",
    ]  # cols 1..15
    wk_pct_pairs = [
        ("Комис%", "Δ"), ("Логист%", "Δ"), ("ВБПромо%", "Δ"),
        ("С/С%", "Δ"), ("Приб%", "Δ"),
    ]  # 5 pairs = 10 columns → cols 16..25
    wk_headers = wk_base_headers + [h for pair in wk_pct_pairs for h in pair]
    for col, h in enumerate(wk_headers, 1):
        ws3.cell(row=1, column=col, value=h)
    _style_wb_header_row(ws3, 1, len(wk_headers))
    # Style Δ headers with grey fill
    for pi in range(len(wk_pct_pairs)):
        delta_col = 16 + pi * 2 + 1  # 17, 19, 21, 23, 25
        c = ws3.cell(row=1, column=delta_col)
        c.fill = DELTA_HDR_FILL
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    wk_totals = {"qty": 0, "revenue": 0, "payout": 0, "returns": 0,
                 "logistics": 0, "storage": 0, "acceptance": 0,
                 "deductions": 0, "ded_ads": 0,
                 "cogs": 0, "profit": 0}
    prev_wk_pcts = None

    for i, r in enumerate(weekly_data):
        rn = i + 2
        ws_d = r[0]
        we_d = r[1]
        qty = int(r[2] or 0)
        rev = float(r[3] or 0)
        pay = float(r[4] or 0)
        rets = abs(float(r[5] or 0))
        log = abs(float(r[6] or 0))
        stor = abs(float(r[7] or 0))
        acq = abs(float(r[8] or 0))
        acc = abs(float(r[9] or 0))
        ded = abs(float(r[10] or 0))
        ded_ads_fin = abs(float(r[11] or 0))
        ded_ads_stats = weekly_ad_spend.get(str(ws_d), 0)
        ded_ads = max(ded_ads_fin, ded_ads_stats)  # MAX-reconciliation
        cogs_wk = cogs_by_week.get(str(ws_d), 0)
        comm = max(rev - pay, 0)
        profit_wk = pay - log - stor - acc - ded - ded_ads - cogs_wk

        ws3.cell(row=rn, column=1, value=ws_d.year if hasattr(ws_d, 'year') else int(str(ws_d)[:4]))
        ws3.cell(row=rn, column=2, value=ws_d.isocalendar()[1] if hasattr(ws_d, 'isocalendar') else 0)
        ws3.cell(row=rn, column=3, value=_fmt_date_ru(ws_d))
        ws3.cell(row=rn, column=4, value=_fmt_date_ru(we_d))
        ws3.cell(row=rn, column=5, value=qty)

        money_vals = [rev, pay, rets, log, stor, acc, ded, ded_ads, cogs_wk, profit_wk]
        money_keys = ["revenue", "payout", "returns", "logistics", "storage", "acceptance",
                      "deductions", "ded_ads", "cogs", "profit"]
        for j, v in enumerate(money_vals):
            c = ws3.cell(row=rn, column=6 + j, value=round(v, 2))
            c.number_format = MONEY_FMT
            wk_totals[money_keys[j]] += v
        wk_totals["qty"] += qty

        # Profit coloring
        ws3.cell(row=rn, column=15).font = GREEN_FONT if profit_wk >= 0 else RED_FONT

        # Percentage columns: Комис%, Логист%, ВБПромо%, С/С%, Приб%
        is_alt = (i % 2 == 1)
        if rev > 0:
            cur_pcts = [
                round(comm / rev * 100, 1),
                round(log / rev * 100, 1),
                round(ded_ads / rev * 100, 1),
                round(cogs_wk / rev * 100, 1),
                round(profit_wk / rev * 100, 1),
            ]
        else:
            cur_pcts = [0, 0, 0, 0, 0]

        for pi, pct_val in enumerate(cur_pcts):
            pct_col = 16 + pi * 2       # 16, 18, 20, 22, 24
            delta_col = 16 + pi * 2 + 1 # 17, 19, 21, 23, 25

            pc = ws3.cell(row=rn, column=pct_col, value=pct_val)
            pc.number_format = '0.0"%"'
            if pi == 4:  # Profit%
                pc.font = GREEN_FONT if pct_val >= 0 else RED_FONT
            else:
                pc.font = RED_FONT if pct_val > 35 else (GREEN_FONT if pct_val < 20 else NORMAL_FONT)

            dc = ws3.cell(row=rn, column=delta_col)
            dc.fill = DELTA_FILL_ALT if is_alt else DELTA_FILL
            if prev_wk_pcts is not None:
                delta_pp = round(pct_val - prev_wk_pcts[pi], 1)
                dc.value = delta_pp
                dc.number_format = '+0.0;-0.0;0.0'
                if pi == 4:
                    dc.font = GREEN_FONT_SM if delta_pp > 0 else (RED_FONT_SM if delta_pp < 0 else GREY_FONT_SM)
                else:
                    dc.font = GREEN_FONT_SM if delta_pp < 0 else (RED_FONT_SM if delta_pp > 0 else GREY_FONT_SM)
            else:
                dc.value = "—"
                dc.font = GREY_FONT_SM

        prev_wk_pcts = cur_pcts
        _style_data_row(ws3, rn, 15, is_alt=is_alt)

    # Totals row
    total_row = len(weekly_data) + 2
    ws3.cell(row=total_row, column=1, value="ИТОГО")
    ws3.cell(row=total_row, column=5, value=wk_totals["qty"])
    for j, k in enumerate(money_keys):
        ws3.cell(row=total_row, column=6 + j, value=round(wk_totals[k], 2)).number_format = MONEY_FMT

    ts = wk_totals["revenue"] or 1
    total_comm_wk = max(wk_totals["revenue"] - wk_totals["payout"], 0)
    total_pcts_wk = [
        round(total_comm_wk / ts * 100, 1),
        round(wk_totals["logistics"] / ts * 100, 1),
        round(wk_totals["ded_ads"] / ts * 100, 1),
        round(wk_totals["cogs"] / ts * 100, 1),
        round(wk_totals["profit"] / ts * 100, 1),
    ]
    for pi, pv in enumerate(total_pcts_wk):
        pct_col = 16 + pi * 2
        ws3.cell(row=total_row, column=pct_col, value=pv).number_format = '0.0"%"'
    _style_total_row(ws3, total_row, len(wk_headers))

    _auto_width(ws3)
    ws3.freeze_panes = 'E2'

    # ── Sheet 3: По месяцам (FULL RETROSPECTIVE with Δ-columns) ──
    ws4 = wb.create_sheet("По месяцам")
    ws4.sheet_properties.tabColor = "7C3AED"

    MONTHS_RU = {1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
                 7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"}

    mo_base_headers = [
        "Месяц", "Начало", "Конец", "Кол-во",
        "Выручка", "К перечисл.", "Возвраты", "Логистика", "Хранение",
        "Приёмка", "Удержания", "ВБ Промо",
        "С/С", "Прибыль",
    ]  # cols 1..14
    mo_pct_pairs = [
        ("Комис%", "Δ"), ("Логист%", "Δ"), ("ВБПромо%", "Δ"),
        ("С/С%", "Δ"), ("Приб%", "Δ"),
    ]  # 5 pairs = 10 columns → cols 15..24
    mo_headers = mo_base_headers + [h for pair in mo_pct_pairs for h in pair]
    for col, h in enumerate(mo_headers, 1):
        ws4.cell(row=1, column=col, value=h)
    _style_wb_header_row(ws4, 1, len(mo_headers))
    for pi in range(len(mo_pct_pairs)):
        delta_col = 15 + pi * 2 + 1  # 16, 18, 20, 22, 24
        c = ws4.cell(row=1, column=delta_col)
        c.fill = DELTA_HDR_FILL
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    mo_totals = {"qty": 0, "revenue": 0, "payout": 0, "returns": 0,
                 "logistics": 0, "storage": 0, "acceptance": 0,
                 "deductions": 0, "ded_ads": 0,
                 "cogs": 0, "profit": 0}
    prev_mo_pcts = None

    for i, r in enumerate(monthly_data):
        rn = i + 2
        ym = int(r[0])
        year = ym // 100
        month = ym % 100
        month_name = f"{MONTHS_RU.get(month, str(month))} {year}"

        qty = int(r[3] or 0)
        rev = float(r[4] or 0)
        pay = float(r[5] or 0)
        rets = abs(float(r[6] or 0))
        log = abs(float(r[7] or 0))
        stor = abs(float(r[8] or 0))
        acq = abs(float(r[9] or 0))
        acc = abs(float(r[10] or 0))
        ded = abs(float(r[11] or 0))
        ded_ads_fin = abs(float(r[12] or 0))
        ded_ads_stats = monthly_ad_spend.get(ym, 0)
        ded_ads = max(ded_ads_fin, ded_ads_stats)  # MAX-reconciliation
        cogs_mo = cogs_by_month.get(ym, 0)
        comm = max(rev - pay, 0)
        profit_mo = pay - log - stor - acc - ded - ded_ads - cogs_mo

        ws4.cell(row=rn, column=1, value=month_name)
        ws4.cell(row=rn, column=2, value=str(r[1]))
        ws4.cell(row=rn, column=3, value=str(r[2]))
        ws4.cell(row=rn, column=4, value=qty)

        money_vals_mo = [rev, pay, rets, log, stor, acc, ded, ded_ads, cogs_mo, profit_mo]
        money_keys_mo = ["revenue", "payout", "returns", "logistics", "storage", "acceptance",
                         "deductions", "ded_ads", "cogs", "profit"]
        for j, v in enumerate(money_vals_mo):
            c = ws4.cell(row=rn, column=5 + j, value=round(v, 2))
            c.number_format = MONEY_FMT
            mo_totals[money_keys_mo[j]] += v
        mo_totals["qty"] += qty

        # Profit coloring
        ws4.cell(row=rn, column=14).font = GREEN_FONT if profit_mo >= 0 else RED_FONT

        # Percentage columns with Δ
        is_alt = (i % 2 == 1)
        if rev > 0:
            cur_mo_pcts = [
                round(comm / rev * 100, 1),
                round(log / rev * 100, 1),
                round(ded_ads / rev * 100, 1),
                round(cogs_mo / rev * 100, 1),
                round(profit_mo / rev * 100, 1),
            ]
        else:
            cur_mo_pcts = [0, 0, 0, 0, 0]

        for pi, pct_val in enumerate(cur_mo_pcts):
            pct_col = 15 + pi * 2       # 15, 17, 19, 21, 23
            delta_col = 15 + pi * 2 + 1 # 16, 18, 20, 22, 24

            pc = ws4.cell(row=rn, column=pct_col, value=pct_val)
            pc.number_format = '0.0"%"'
            if pi == 4:
                pc.font = GREEN_FONT if pct_val >= 0 else RED_FONT
            else:
                pc.font = RED_FONT if pct_val > 35 else (GREEN_FONT if pct_val < 20 else NORMAL_FONT)

            dc = ws4.cell(row=rn, column=delta_col)
            dc.fill = DELTA_FILL_ALT if is_alt else DELTA_FILL
            if prev_mo_pcts is not None:
                delta_pp = round(pct_val - prev_mo_pcts[pi], 1)
                dc.value = delta_pp
                dc.number_format = '+0.0;-0.0;0.0'
                if pi == 4:
                    dc.font = GREEN_FONT_SM if delta_pp > 0 else (RED_FONT_SM if delta_pp < 0 else GREY_FONT_SM)
                else:
                    dc.font = GREEN_FONT_SM if delta_pp < 0 else (RED_FONT_SM if delta_pp > 0 else GREY_FONT_SM)
            else:
                dc.value = "—"
                dc.font = GREY_FONT_SM

        prev_mo_pcts = cur_mo_pcts
        _style_data_row(ws4, rn, 14, is_alt=is_alt)

    # Totals row
    total_row_mo = len(monthly_data) + 2
    ws4.cell(row=total_row_mo, column=1, value="ИТОГО")
    ws4.cell(row=total_row_mo, column=4, value=mo_totals["qty"])
    for j, k in enumerate(money_keys_mo):
        ws4.cell(row=total_row_mo, column=5 + j, value=round(mo_totals[k], 2)).number_format = MONEY_FMT

    ts_mo = mo_totals["revenue"] or 1
    total_comm_mo = max(mo_totals["revenue"] - mo_totals["payout"], 0)
    total_pcts_mo = [
        round(total_comm_mo / ts_mo * 100, 1),
        round(mo_totals["logistics"] / ts_mo * 100, 1),
        round(mo_totals["ded_ads"] / ts_mo * 100, 1),
        round(mo_totals["cogs"] / ts_mo * 100, 1),
        round(mo_totals["profit"] / ts_mo * 100, 1),
    ]
    for pi, pv in enumerate(total_pcts_mo):
        pct_col = 15 + pi * 2
        ws4.cell(row=total_row_mo, column=pct_col, value=pv).number_format = '0.0"%"'
    _style_total_row(ws4, total_row_mo, len(mo_headers))

    _auto_width(ws4)
    ws4.freeze_panes = 'D2'

    # ── Sheet 5: По товарам (SKU) — с ДРР% ──
    ws5 = wb.create_sheet("По товарам")
    ws5.sheet_properties.tabColor = "10B981"

    sku_headers = [
        "Артикул", "Название", "GTIN", "Кол-во", "Выручка",
        "К перечисл.", "Логистика", "Хранение", "Приёмка",
        "Удержания", "Реклама", "ДРР%", "С/С", "Прибыль", "Маржа%",
    ]
    for col, h in enumerate(sku_headers, 1):
        ws5.cell(row=1, column=col, value=h)
    _style_wb_header_row(ws5, 1, len(sku_headers))

    sorted_skus = sorted(sku_data.items(), key=lambda x: x[1]["revenue"], reverse=True)
    for i, (vc, sd) in enumerate(sorted_skus):
        rn = i + 2
        unit_cost = cost_map.get(vc.lower(), 0)
        cogs_sku = unit_cost * sd["qty"] if unit_cost > 0 else 0
        ads_sku = sd.get("ad_spend", 0)
        profit_sku = sd["payout"] - sd["logistics"] - sd["storage"] - sd["acceptance"] - sd["deductions"] - ads_sku - cogs_sku
        margin = profit_sku / sd["revenue"] * 100 if sd["revenue"] > 0 else 0
        drr_sku = ads_sku / sd["revenue"] * 100 if sd["revenue"] > 0 else 0

        vc_lower = vc.lower()
        ws5.cell(row=rn, column=1, value=vc)
        ws5.cell(row=rn, column=2, value=names_map.get(vc, names_lower.get(vc_lower, "")))
        ws5.cell(row=rn, column=3, value=barcode_map.get(vc, barcode_lower.get(vc_lower, "")))
        ws5.cell(row=rn, column=4, value=sd["qty"])
        ws5.cell(row=rn, column=5, value=round(sd["revenue"], 2)).number_format = MONEY_FMT
        ws5.cell(row=rn, column=6, value=round(sd["payout"], 2)).number_format = MONEY_FMT
        ws5.cell(row=rn, column=7, value=round(sd["logistics"], 2)).number_format = MONEY_FMT
        ws5.cell(row=rn, column=8, value=round(sd["storage"], 2)).number_format = MONEY_FMT
        ws5.cell(row=rn, column=9, value=round(sd["acceptance"], 2)).number_format = MONEY_FMT
        ws5.cell(row=rn, column=10, value=round(sd["deductions"], 2)).number_format = MONEY_FMT
        ws5.cell(row=rn, column=11, value=round(ads_sku, 2)).number_format = MONEY_FMT
        cell_drr = ws5.cell(row=rn, column=12, value=round(drr_sku, 1))
        cell_drr.number_format = '0.0"%"'
        cell_drr.font = RED_FONT if drr_sku > 30 else (GREEN_FONT if drr_sku < 15 else NORMAL_FONT)
        ws5.cell(row=rn, column=13, value=round(cogs_sku, 2)).number_format = MONEY_FMT
        ws5.cell(row=rn, column=14, value=round(profit_sku, 2)).number_format = MONEY_FMT
        cell_m = ws5.cell(row=rn, column=15, value=round(margin, 1))
        cell_m.number_format = '0.0"%"'
        cell_m.font = GREEN_FONT if margin >= 0 else RED_FONT

        _style_data_row(ws5, rn, len(sku_headers), is_alt=(i % 2 == 1))

    # Totals row
    total_row = len(sorted_skus) + 2
    ws5.cell(row=total_row, column=1, value="ИТОГО")
    t_qty = sum(s["qty"] for s in sku_data.values())
    t_rev = sum(s["revenue"] for s in sku_data.values())
    t_pay = sum(s["payout"] for s in sku_data.values())
    t_log = sum(s["logistics"] for s in sku_data.values())
    t_stor = sum(s["storage"] for s in sku_data.values())
    t_acc = sum(s["acceptance"] for s in sku_data.values())
    t_ded = sum(s["deductions"] for s in sku_data.values())
    t_ads_sku = sum(s.get("ad_spend", 0) for s in sku_data.values())
    t_cogs_sku = sum(cost_map.get(vc.lower(), 0) * s["qty"] for vc, s in sku_data.items())
    t_profit_sku = t_pay - t_log - t_stor - t_acc - t_ded - t_ads_sku - t_cogs_sku
    t_margin = t_profit_sku / t_rev * 100 if t_rev > 0 else 0
    t_drr = t_ads_sku / t_rev * 100 if t_rev > 0 else 0

    ws5.cell(row=total_row, column=4, value=t_qty)
    ws5.cell(row=total_row, column=5, value=round(t_rev, 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=6, value=round(t_pay, 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=7, value=round(t_log, 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=8, value=round(t_stor, 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=9, value=round(t_acc, 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=10, value=round(t_ded, 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=11, value=round(t_ads_sku, 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=12, value=round(t_drr, 1)).number_format = '0.0"%"'
    ws5.cell(row=total_row, column=13, value=round(t_cogs_sku, 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=14, value=round(t_profit_sku, 2)).number_format = MONEY_FMT
    ws5.cell(row=total_row, column=15, value=round(t_margin, 1)).number_format = '0.0"%"'
    _style_total_row(ws5, total_row, len(sku_headers))

    _auto_width(ws5)
    ws5.freeze_panes = 'A2'
    ws5.auto_filter.ref = f"A1:{chr(64 + len(sku_headers))}{total_row - 1}"

    # ── Sheet 6: Расходы детально (с % от выручки) ──
    ws6 = wb.create_sheet("Расходы детально")
    ws6.sheet_properties.tabColor = "D97706"

    exp_headers = [
        "Тип операции", "Тип бонуса/удержания", "Записей",
        "Логистика", "Хранение", "Приёмка",
        "Штрафы", "Удержания", "Эквайринг",
        "Итого", "% выр.",
    ]
    for col, h in enumerate(exp_headers, 1):
        ws6.cell(row=1, column=col, value=h)
    _style_wb_header_row(ws6, 1, len(exp_headers))

    exp_grand_total = 0.0
    for i, r in enumerate(expense_detail_rows):
        rn = i + 2
        ws6.cell(row=rn, column=1, value=str(r[0] or ""))
        ws6.cell(row=rn, column=2, value=str(r[1] or "—"))
        ws6.cell(row=rn, column=3, value=int(r[2] or 0))
        vals = [abs(float(r[j] or 0)) for j in range(3, 9)]
        for j, v in enumerate(vals):
            ws6.cell(row=rn, column=4 + j, value=v).number_format = MONEY_FMT_2
        row_total = sum(vals)
        exp_grand_total += row_total
        ws6.cell(row=rn, column=10, value=round(row_total, 2)).number_format = MONEY_FMT_2
        pct_rev = row_total / revenue_cur * 100 if revenue_cur > 0 else 0
        c_pct = ws6.cell(row=rn, column=11, value=round(pct_rev, 2))
        c_pct.number_format = '0.00"%"'
        c_pct.font = RED_FONT if pct_rev > 5 else NORMAL_FONT
        _style_data_row(ws6, rn, len(exp_headers), is_alt=(i % 2 == 1))

    # Totals row for expenses
    exp_total_row = len(expense_detail_rows) + 2
    ws6.cell(row=exp_total_row, column=1, value="ИТОГО")
    ws6.cell(row=exp_total_row, column=10, value=round(exp_grand_total, 2)).number_format = MONEY_FMT_2
    pct_total = exp_grand_total / revenue_cur * 100 if revenue_cur > 0 else 0
    ws6.cell(row=exp_total_row, column=11, value=round(pct_total, 1)).number_format = '0.0"%"'
    _style_total_row(ws6, exp_total_row, len(exp_headers))

    _auto_width(ws6)
    ws6.freeze_panes = 'A2'
    ws6.auto_filter.ref = f"A1:{chr(64 + len(exp_headers))}{exp_total_row - 1}"

    # ══════════════════════════════════════════════════════
    # Save
    # ══════════════════════════════════════════════════════
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"WB_Finance_{shop.name}_{d_start}_{d_end}.xlsx"
    encoded = quote(filename)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )

