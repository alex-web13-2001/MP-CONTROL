"""
Warehouses API — supply recommendations.

GET /warehouses/ozon/supply       — JSON recommendations per SKU × cluster
GET /warehouses/ozon/supply/export — Excel download
"""
import io
import logging
from datetime import date, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.shop import Shop

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/warehouses", tags=["Warehouses"])


# ══════════════════════════════════════════════════════════════
# Delivery time matrix (hours) - source: Ozon normative 01/2026
# ══════════════════════════════════════════════════════════════

DELIVERY_HOURS: dict[str, dict[str, int]] = {
    "Москва, МО и Дальние регионы": {"Москва, МО и Дальние регионы": 28, "Санкт-Петербург и СЗО": 60, "Тверь": 45, "Ярославль": 45, "Воронеж": 60, "Казань": 60, "Самара": 75, "Саратов": 60, "Оренбург": 75, "Екатеринбург": 75, "Пермь": 75, "Уфа": 75, "Красноярск": 135, "Новосибирск": 105, "Омск": 105, "Тюмень": 90, "Ростов": 60, "Краснодар": 60, "Махачкала": 75, "Невинномысск": 75, "Дальний Восток": 150, "Калининград": 60, "Беларусь": 45, "Астана": 75, "Алматы": 75},
    "Санкт-Петербург и СЗО": {"Москва, МО и Дальние регионы": 60, "Санкт-Петербург и СЗО": 28, "Тверь": 60, "Ярославль": 60, "Воронеж": 75, "Казань": 75, "Самара": 90, "Саратов": 75, "Оренбург": 75, "Екатеринбург": 90, "Пермь": 90, "Уфа": 75, "Красноярск": 150, "Новосибирск": 120, "Омск": 120, "Тюмень": 90, "Ростов": 75, "Краснодар": 75, "Махачкала": 105, "Невинномысск": 105, "Дальний Восток": 150, "Калининград": 60, "Беларусь": 75, "Астана": 90, "Алматы": 90},
    "Тверь": {"Москва, МО и Дальние регионы": 60, "Санкт-Петербург и СЗО": 60, "Тверь": 28, "Ярославль": 45, "Воронеж": 60, "Казань": 60, "Самара": 75, "Саратов": 60, "Оренбург": 75, "Екатеринбург": 75, "Пермь": 75, "Уфа": 75, "Красноярск": 150, "Новосибирск": 120, "Омск": 120, "Тюмень": 90, "Ростов": 75, "Краснодар": 75, "Махачкала": 75, "Невинномысск": 75, "Дальний Восток": 150, "Калининград": 60, "Беларусь": 60, "Астана": 75, "Алматы": 90},
    "Ярославль": {"Москва, МО и Дальние регионы": 60, "Санкт-Петербург и СЗО": 60, "Тверь": 45, "Ярославль": 28, "Воронеж": 75, "Казань": 60, "Самара": 75, "Саратов": 75, "Оренбург": 90, "Екатеринбург": 75, "Пермь": 75, "Уфа": 75, "Красноярск": 150, "Новосибирск": 120, "Омск": 120, "Тюмень": 90, "Ростов": 75, "Краснодар": 75, "Махачкала": 75, "Невинномысск": 75, "Дальний Восток": 150, "Калининград": 60, "Беларусь": 60, "Астана": 75, "Алматы": 90},
    "Воронеж": {"Москва, МО и Дальние регионы": 60, "Санкт-Петербург и СЗО": 60, "Тверь": 60, "Ярославль": 60, "Воронеж": 28, "Казань": 60, "Самара": 60, "Саратов": 45, "Оренбург": 75, "Екатеринбург": 75, "Пермь": 75, "Уфа": 75, "Красноярск": 135, "Новосибирск": 105, "Омск": 105, "Тюмень": 90, "Ростов": 45, "Краснодар": 60, "Махачкала": 60, "Невинномысск": 60, "Дальний Восток": 150, "Калининград": 75, "Беларусь": 60, "Астана": 75, "Алматы": 90},
    "Казань": {"Москва, МО и Дальние регионы": 60, "Санкт-Петербург и СЗО": 60, "Тверь": 60, "Ярославль": 60, "Воронеж": 60, "Казань": 28, "Самара": 45, "Саратов": 60, "Оренбург": 60, "Екатеринбург": 45, "Пермь": 45, "Уфа": 45, "Красноярск": 120, "Новосибирск": 105, "Омск": 90, "Тюмень": 75, "Ростов": 75, "Краснодар": 75, "Махачкала": 90, "Невинномысск": 90, "Дальний Восток": 150, "Калининград": 75, "Беларусь": 60, "Астана": 75, "Алматы": 90},
    "Самара": {"Москва, МО и Дальние регионы": 60, "Санкт-Петербург и СЗО": 75, "Тверь": 60, "Ярославль": 60, "Воронеж": 45, "Казань": 45, "Самара": 28, "Саратов": 45, "Оренбург": 45, "Екатеринбург": 60, "Пермь": 60, "Уфа": 45, "Красноярск": 120, "Новосибирск": 105, "Омск": 105, "Тюмень": 75, "Ростов": 75, "Краснодар": 75, "Махачкала": 90, "Невинномысск": 90, "Дальний Восток": 150, "Калининград": 75, "Беларусь": 75, "Астана": 75, "Алматы": 90},
    "Саратов": {"Москва, МО и Дальние регионы": 75, "Санкт-Петербург и СЗО": 75, "Тверь": 75, "Ярославль": 90, "Воронеж": 60, "Казань": 60, "Самара": 28, "Саратов": 28, "Оренбург": 75, "Екатеринбург": 75, "Пермь": 75, "Уфа": 75, "Красноярск": 120, "Новосибирск": 105, "Омск": 105, "Тюмень": 75, "Ростов": 75, "Краснодар": 75, "Махачкала": 90, "Невинномысск": 90, "Дальний Восток": 150, "Калининград": 75, "Беларусь": 75, "Астана": 75, "Алматы": 90},
    "Оренбург": {"Москва, МО и Дальние регионы": 75, "Санкт-Петербург и СЗО": 75, "Тверь": 75, "Ярославль": 90, "Воронеж": 60, "Казань": 60, "Самара": 80, "Саратов": 80, "Оренбург": 28, "Екатеринбург": 60, "Пермь": 75, "Уфа": 45, "Красноярск": 120, "Новосибирск": 105, "Омск": 90, "Тюмень": 75, "Ростов": 75, "Краснодар": 75, "Махачкала": 90, "Невинномысск": 90, "Дальний Восток": 150, "Калининград": 90, "Беларусь": 75, "Астана": 75, "Алматы": 90},
    "Екатеринбург": {"Москва, МО и Дальние регионы": 75, "Санкт-Петербург и СЗО": 90, "Тверь": 105, "Ярославль": 75, "Воронеж": 75, "Казань": 75, "Самара": 75, "Саратов": 60, "Оренбург": 60, "Екатеринбург": 28, "Пермь": 45, "Уфа": 45, "Красноярск": 90, "Новосибирск": 75, "Омск": 75, "Тюмень": 45, "Ростов": 75, "Краснодар": 75, "Махачкала": 90, "Невинномысск": 90, "Дальний Восток": 150, "Калининград": 105, "Беларусь": 90, "Астана": 45, "Алматы": 45},
    "Пермь": {"Москва, МО и Дальние регионы": 75, "Санкт-Петербург и СЗО": 90, "Тверь": 105, "Ярославль": 75, "Воронеж": 75, "Казань": 60, "Самара": 60, "Саратов": 45, "Оренбург": 28, "Екатеринбург": 45, "Пермь": 28, "Уфа": 45, "Красноярск": 90, "Новосибирск": 75, "Омск": 75, "Тюмень": 60, "Ростов": 75, "Краснодар": 75, "Махачкала": 90, "Невинномысск": 90, "Дальний Восток": 150, "Калининград": 90, "Беларусь": 90, "Астана": 60, "Алматы": 60},
    "Уфа": {"Москва, МО и Дальние регионы": 75, "Санкт-Петербург и СЗО": 75, "Тверь": 75, "Ярославль": 75, "Воронеж": 60, "Казань": 45, "Самара": 45, "Саратов": 60, "Оренбург": 45, "Екатеринбург": 45, "Пермь": 45, "Уфа": 28, "Красноярск": 120, "Новосибирск": 105, "Омск": 90, "Тюмень": 75, "Ростов": 75, "Краснодар": 75, "Махачкала": 90, "Невинномысск": 90, "Дальний Восток": 150, "Калининград": 90, "Беларусь": 75, "Астана": 75, "Алматы": 90},
    "Красноярск": {"Москва, МО и Дальние регионы": 135, "Санкт-Петербург и СЗО": 150, "Тверь": 135, "Ярославль": 135, "Воронеж": 120, "Казань": 120, "Самара": 90, "Саратов": 90, "Оренбург": 75, "Екатеринбург": 75, "Пермь": 75, "Уфа": 75, "Красноярск": 28, "Новосибирск": 45, "Омск": 45, "Тюмень": 60, "Ростов": 120, "Краснодар": 120, "Махачкала": 120, "Невинномысск": 120, "Дальний Восток": 45, "Калининград": 150, "Беларусь": 150, "Астана": 105, "Алматы": 120},
    "Новосибирск": {"Москва, МО и Дальние регионы": 105, "Санкт-Петербург и СЗО": 120, "Тверь": 105, "Ярославль": 135, "Воронеж": 105, "Казань": 75, "Самара": 90, "Саратов": 105, "Оренбург": 75, "Екатеринбург": 60, "Пермь": 75, "Уфа": 75, "Красноярск": 45, "Новосибирск": 28, "Омск": 45, "Тюмень": 60, "Ростов": 90, "Краснодар": 90, "Махачкала": 120, "Невинномысск": 120, "Дальний Восток": 150, "Калининград": 105, "Беларусь": 105, "Астана": 75, "Алматы": 75},
    "Омск": {"Москва, МО и Дальние регионы": 105, "Санкт-Петербург и СЗО": 120, "Тверь": 105, "Ярославль": 120, "Воронеж": 105, "Казань": 90, "Самара": 105, "Саратов": 105, "Оренбург": 75, "Екатеринбург": 75, "Пермь": 75, "Уфа": 75, "Красноярск": 45, "Новосибирск": 45, "Омск": 28, "Тюмень": 60, "Ростов": 120, "Краснодар": 120, "Махачкала": 105, "Невинномысск": 105, "Дальний Восток": 45, "Калининград": 135, "Беларусь": 120, "Астана": 75, "Алматы": 75},
    "Тюмень": {"Москва, МО и Дальние регионы": 90, "Санкт-Петербург и СЗО": 90, "Тверь": 90, "Ярославль": 120, "Воронеж": 90, "Казань": 90, "Самара": 90, "Саратов": 90, "Оренбург": 75, "Екатеринбург": 45, "Пермь": 60, "Уфа": 75, "Красноярск": 75, "Новосибирск": 60, "Омск": 60, "Тюмень": 28, "Ростов": 90, "Краснодар": 90, "Махачкала": 105, "Невинномысск": 105, "Дальний Восток": 150, "Калининград": 120, "Беларусь": 90, "Астана": 45, "Алматы": 45},
    "Ростов": {"Москва, МО и Дальние регионы": 75, "Санкт-Петербург и СЗО": 75, "Тверь": 75, "Ярославль": 90, "Воронеж": 45, "Казань": 75, "Самара": 90, "Саратов": 60, "Оренбург": 75, "Екатеринбург": 75, "Пермь": 75, "Уфа": 75, "Красноярск": 135, "Новосибирск": 120, "Омск": 120, "Тюмень": 90, "Ростов": 28, "Краснодар": 45, "Махачкала": 60, "Невинномысск": 45, "Дальний Восток": 150, "Калининград": 75, "Беларусь": 90, "Астана": 105, "Алматы": 105},
    "Краснодар": {"Москва, МО и Дальние регионы": 75, "Санкт-Петербург и СЗО": 75, "Тверь": 75, "Ярославль": 90, "Воронеж": 60, "Казань": 75, "Самара": 90, "Саратов": 60, "Оренбург": 75, "Екатеринбург": 75, "Пермь": 75, "Уфа": 75, "Красноярск": 135, "Новосибирск": 120, "Омск": 120, "Тюмень": 90, "Ростов": 45, "Краснодар": 28, "Махачкала": 60, "Невинномысск": 60, "Дальний Восток": 150, "Калининград": 75, "Беларусь": 90, "Астана": 90, "Алматы": 90},
    "Махачкала": {"Москва, МО и Дальние регионы": 75, "Санкт-Петербург и СЗО": 75, "Тверь": 75, "Ярославль": 75, "Воронеж": 60, "Казань": 90, "Самара": 105, "Саратов": 105, "Оренбург": 90, "Екатеринбург": 90, "Пермь": 90, "Уфа": 90, "Красноярск": 135, "Новосибирск": 120, "Омск": 90, "Тюмень": 90, "Ростов": 60, "Краснодар": 60, "Махачкала": 28, "Невинномысск": 45, "Дальний Восток": 150, "Калининград": 90, "Беларусь": 120, "Астана": 90, "Алматы": 120},
    "Невинномысск": {"Москва, МО и Дальние регионы": 75, "Санкт-Петербург и СЗО": 75, "Тверь": 75, "Ярославль": 75, "Воронеж": 60, "Казань": 90, "Самара": 105, "Саратов": 105, "Оренбург": 90, "Екатеринбург": 90, "Пермь": 90, "Уфа": 90, "Красноярск": 135, "Новосибирск": 120, "Омск": 90, "Тюмень": 90, "Ростов": 45, "Краснодар": 45, "Махачкала": 45, "Невинномысск": 28, "Дальний Восток": 150, "Калининград": 90, "Беларусь": 90, "Астана": 90, "Алматы": 120},
    "Дальний Восток": {"Москва, МО и Дальние регионы": 150, "Санкт-Петербург и СЗО": 150, "Тверь": 150, "Ярославль": 150, "Воронеж": 150, "Казань": 150, "Самара": 150, "Саратов": 150, "Оренбург": 150, "Екатеринбург": 150, "Пермь": 150, "Уфа": 150, "Красноярск": 120, "Новосибирск": 150, "Омск": 150, "Тюмень": 150, "Ростов": 150, "Краснодар": 150, "Махачкала": 150, "Невинномысск": 150, "Дальний Восток": 28, "Калининград": 150, "Беларусь": 150, "Астана": 150, "Алматы": 150},
    "Калининград": {"Москва, МО и Дальние регионы": 999, "Санкт-Петербург и СЗО": 999, "Тверь": 999, "Ярославль": 999, "Воронеж": 999, "Казань": 999, "Самара": 999, "Саратов": 999, "Оренбург": 999, "Екатеринбург": 999, "Пермь": 999, "Уфа": 999, "Красноярск": 999, "Новосибирск": 999, "Омск": 999, "Тюмень": 999, "Ростов": 999, "Краснодар": 999, "Махачкала": 999, "Невинномысск": 999, "Дальний Восток": 999, "Калининград": 28, "Беларусь": 999, "Астана": 999, "Алматы": 999},
    "Беларусь": {"Москва, МО и Дальние регионы": 60, "Санкт-Петербург и СЗО": 60, "Тверь": 60, "Ярославль": 60, "Воронеж": 60, "Казань": 75, "Самара": 90, "Саратов": 75, "Оренбург": 105, "Екатеринбург": 90, "Пермь": 90, "Уфа": 105, "Красноярск": 150, "Новосибирск": 120, "Омск": 120, "Тюмень": 90, "Ростов": 75, "Краснодар": 75, "Махачкала": 105, "Невинномысск": 105, "Дальний Восток": 150, "Калининград": 60, "Беларусь": 28, "Астана": 90, "Алматы": 90},
    "Астана": {"Москва, МО и Дальние регионы": 90, "Санкт-Петербург и СЗО": 120, "Тверь": 90, "Ярославль": 150, "Воронеж": 90, "Казань": 105, "Самара": 105, "Саратов": 90, "Оренбург": 105, "Екатеринбург": 60, "Пермь": 60, "Уфа": 105, "Красноярск": 105, "Новосибирск": 75, "Омск": 75, "Тюмень": 75, "Ростов": 105, "Краснодар": 105, "Махачкала": 150, "Невинномысск": 150, "Дальний Восток": 150, "Калининград": 120, "Беларусь": 105, "Астана": 28, "Алматы": 28},
    "Алматы": {"Москва, МО и Дальние регионы": 90, "Санкт-Петербург и СЗО": 120, "Тверь": 90, "Ярославль": 150, "Воронеж": 90, "Казань": 105, "Самара": 105, "Саратов": 90, "Оренбург": 105, "Екатеринбург": 60, "Пермь": 60, "Уфа": 105, "Красноярск": 105, "Новосибирск": 75, "Омск": 75, "Тюмень": 75, "Ростов": 105, "Краснодар": 105, "Махачкала": 150, "Невинномысск": 150, "Дальний Восток": 150, "Калининград": 120, "Беларусь": 105, "Астана": 28, "Алматы": 28},
}


def _resolve_hub(demand_cluster: str) -> tuple[str, int]:
    """
    For a demand cluster, find the optimal hub (source) with min delivery time.
    Priority: self (28h) is always best; returns (hub_name, hours).
    """
    best_hub = demand_cluster
    best_hours = 28  # self-delivery
    for source, dests in DELIVERY_HOURS.items():
        hours = dests.get(demand_cluster, 999)
        if hours < best_hours:
            best_hours = hours
            best_hub = source
    return best_hub, best_hours


# Объединённые группы: hub → [clusters served]
# Вместо 25 отдельных поставок — 9 точек отгрузки
CONSOLIDATED_GROUPS: dict[str, list[str]] = {
    "Москва, МО и Дальние регионы": ["Москва, МО и Дальние регионы", "Тверь", "Ярославль", "Беларусь"],
    "Санкт-Петербург и СЗО": ["Санкт-Петербург и СЗО"],
    "Казань": ["Казань", "Самара", "Уфа"],
    "Екатеринбург": ["Екатеринбург", "Пермь", "Тюмень", "Оренбург"],
    "Воронеж": ["Воронеж", "Саратов"],
    "Ростов": ["Ростов", "Краснодар", "Невинномысск", "Махачкала"],
    "Красноярск": ["Красноярск", "Новосибирск", "Омск", "Дальний Восток"],
    "Калининград": ["Калининград"],
    "Астана": ["Астана", "Алматы"],
}

# Reverse mapping: cluster → consolidated hub
_CLUSTER_TO_GROUP = {cl: hub for hub, cls in CONSOLIDATED_GROUPS.items() for cl in cls}


# ══════════════════════════════════════════════════════════════
# Ozon РФЦ → Кластер доставки (для per-warehouse stock)
# ══════════════════════════════════════════════════════════════

WAREHOUSE_TO_CLUSTER: dict[str, str] = {
    # Москва
    "ГРИВНО_РФЦ": "Москва, МО и Дальние регионы",
    "ДОМОДЕДОВО_РФЦ": "Москва, МО и Дальние регионы",
    "ЖУКОВСКИЙ_РФЦ": "Москва, МО и Дальние регионы",
    "ПЕТРОВСКОЕ_РФЦ": "Москва, МО и Дальние регионы",
    "СОФЬИНО_РФЦ": "Москва, МО и Дальние регионы",
    "Дедовск": "Москва, МО и Дальние регионы",
    "ПУШКИНО_1_РФЦ": "Москва, МО и Дальние регионы",
    # СПб
    "СПБ_КОЛПИНО_РФЦ": "Санкт-Петербург и СЗО",
    "СПБ_БУГРЫ_РФЦ": "Санкт-Петербург и СЗО",
    "Санкт_Петербург_РФЦ": "Санкт-Петербург и СЗО",
    "САНКТ-ПЕТЕРБУРГ_РФЦ": "Санкт-Петербург и СЗО",
    # Поволжье
    "Казань_РФЦ_НОВЫЙ": "Казань",
    "КАЗАНЬ_РФЦ_НОВЫЙ": "Казань",
    "НИЖНИЙ_НОВГОРОД_РФЦ": "Казань",
    "САМАРА_РФЦ": "Самара",
    "САРАТОВ_РФЦ": "Саратов",
    # Урал
    "Екатеринбург_РФЦ_НОВЫЙ": "Екатеринбург",
    "ЕКАТЕРИНБУРГ_РФЦ_НОВЫЙ": "Екатеринбург",
    "УФА_РФЦ": "Уфа",
    # ЦФО / Черноземье
    "ВОРОНЕЖ_2_РФЦ": "Воронеж",
    "ВОРОНЕЖ_МРФЦ": "Воронеж",
    # Юг
    "РОСТОВ_НА_ДОНУ_2_РФЦ": "Ростов",
    "КРАСНОДАР_2_РФЦ": "Краснодар",
    "НЕВИННОМЫССК_РФЦ": "Невинномысск",
    "НОВОРОССИЙСК_МРФЦ": "Краснодар",
    "АДЫГЕЙСК_РФЦ": "Краснодар",
    # Сибирь / ДВ
    "Новосибирск_РФЦ_НОВЫЙ": "Новосибирск",
    "НОВОСИБИРСК_РФЦ_НОВЫЙ": "Новосибирск",
    "КРАСНОЯРСК_МРФЦ": "Красноярск",
    # Калининград
    "КАЛИНИНГРАД_МРФЦ": "Калининград",
}

# Reverse: cluster → list of warehouses serving it
CLUSTER_TO_WAREHOUSES_OZON: dict[str, list[str]] = {}
for _wh, _cl in WAREHOUSE_TO_CLUSTER.items():
    CLUSTER_TO_WAREHOUSES_OZON.setdefault(_cl, []).append(_wh)


# ══════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════

def _compute_supply(
    ch,
    shop_id: int,
    sales_period: int,
    target_days: int,
    safety: float,
    use_ad_boost: bool,
    product_info: dict,
):
    """
    Core computation: returns (kpi_dict, items_list).
    product_info: {offer_id: {name, image_url, sku}} from PostgreSQL.
    """

    today = date.today()
    d_sales_start = today - timedelta(days=sales_period)

    # ── 1. FBO stocks per warehouse (latest snapshot) ───────────
    fbo_rows = ch.query("""
        SELECT offer_id, warehouse_name,
               sum(free_to_sell) AS fbo_free,
               sum(reserved)     AS fbo_reserved
        FROM mms_analytics.fact_ozon_warehouse_stocks FINAL
        WHERE shop_id = {shop_id:UInt32}
          AND warehouse_type = 'fbo'
          AND dt = (
              SELECT max(dt)
              FROM mms_analytics.fact_ozon_warehouse_stocks
              WHERE shop_id = {shop_id:UInt32} AND warehouse_type = 'fbo'
          )
        GROUP BY offer_id, warehouse_name
    """, parameters={"shop_id": shop_id})

    # Per-warehouse stock: {offer_id: {warehouse_name: free_qty}}
    fbo_by_wh: dict[str, dict[str, int]] = {}
    # Totals (for backward compat): {offer_id: {free, reserved, wh_count}}
    fbo_stock: dict[str, dict] = {}
    for r in fbo_rows.result_rows:
        offer_id, wh_name, free, reserved = r[0], r[1], r[2], r[3]
        # Per-warehouse
        fbo_by_wh.setdefault(offer_id, {})[wh_name] = free
        # Totals
        if offer_id not in fbo_stock:
            fbo_stock[offer_id] = {"free": 0, "reserved": 0, "wh_count": 0}
        fbo_stock[offer_id]["free"] += free
        fbo_stock[offer_id]["reserved"] += reserved
        fbo_stock[offer_id]["wh_count"] += 1

    # Per-cluster stock: aggregate warehouse stock by cluster using WAREHOUSE_TO_CLUSTER
    # {offer_id: {cluster_name: total_free}}
    fbo_by_cluster: dict[str, dict[str, int]] = {}
    for offer_id, wh_stocks in fbo_by_wh.items():
        cluster_stocks: dict[str, int] = {}
        for wh_name, free_qty in wh_stocks.items():
            cluster = WAREHOUSE_TO_CLUSTER.get(wh_name)
            if cluster:
                cluster_stocks[cluster] = cluster_stocks.get(cluster, 0) + free_qty
            else:
                # Unknown warehouse — assign to closest match or skip
                logger.warning("Unknown Ozon warehouse '%s' — not in WAREHOUSE_TO_CLUSTER", wh_name)
        fbo_by_cluster[offer_id] = cluster_stocks


    # ── 2. Sales per SKU × cluster ───────────────────────────
    sales_rows = ch.query("""
        SELECT offer_id, cluster_to,
               sum(quantity)              AS qty,
               round(sum(price*quantity)) AS revenue,
               count(DISTINCT toDate(order_date)) AS active_days
        FROM mms_analytics.fact_ozon_orders FINAL
        WHERE shop_id = {shop_id:UInt32}
          AND order_date >= {d_start:Date}
          AND status NOT IN ('cancelled', 'canceled')
        GROUP BY offer_id, cluster_to
        ORDER BY offer_id, qty DESC
    """, parameters={"shop_id": shop_id, "d_start": d_sales_start})

    sku_clusters = {}   # offer_id → [{cluster, qty, revenue, active_days}]
    sku_totals = {}     # offer_id → total qty
    for r in sales_rows.result_rows:
        offer, cluster, qty, rev, days = r
        sku_clusters.setdefault(offer, [])
        sku_clusters[offer].append({
            "cluster": cluster,
            "qty": qty,
            "revenue": float(rev),
            "active_days": days,
        })
        sku_totals[offer] = sku_totals.get(offer, 0) + qty

    # ── 2b. Actual warehouse consumption (for cross-cluster analysis) ──
    #   Which warehouse actually shipped each order? This reveals
    #   cross-cluster drain: e.g. Moscow warehouse serving Kazan orders.
    cross_rows = ch.query("""
        SELECT warehouse_name, offer_id, cluster_to, sum(quantity) AS qty
        FROM mms_analytics.fact_ozon_orders FINAL
        WHERE shop_id = {shop_id:UInt32}
          AND order_date >= {d_start:Date}
          AND status NOT IN ('cancelled', 'canceled')
          AND warehouse_name != ''
        GROUP BY warehouse_name, offer_id, cluster_to
    """, parameters={"shop_id": shop_id, "d_start": d_sales_start})

    # wh_consumption: {offer_id: {source_cluster: {dest_cluster: qty}}}
    wh_consumption: dict[str, dict[str, dict[str, int]]] = {}
    for r in cross_rows.result_rows:
        wh_name, offer, dest_cluster, qty = r[0], r[1], r[2], r[3]
        src_cluster = WAREHOUSE_TO_CLUSTER.get(wh_name)
        if not src_cluster:
            continue
        wh_consumption.setdefault(offer, {}).setdefault(src_cluster, {})
        wh_consumption[offer][src_cluster][dest_cluster] = (
            wh_consumption[offer][src_cluster].get(dest_cluster, 0) + qty
        )

    # ── 3. Ad boost (7d vs prev 7d) ─────────────────────────
    boost_map = {}
    if use_ad_boost:
        sales_7d_rows = ch.query("""
            SELECT offer_id,
                   sumIf(quantity, order_date >= today() - 7)  AS qty_7d,
                   sumIf(quantity, order_date <  today() - 7)  AS qty_prev7d
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= today() - 14
              AND status NOT IN ('cancelled', 'canceled')
            GROUP BY offer_id
        """, parameters={"shop_id": shop_id})
        s7 = {r[0]: {"q7": r[1], "qp7": r[2]} for r in sales_7d_rows.result_rows}

        # Build sku→offer_id mapping for ads
        offer_to_sku = {}
        for oid, info in product_info.items():
            if info.get("sku"):
                offer_to_sku[oid] = info["sku"]
        sku_to_offer = {v: k for k, v in offer_to_sku.items()}

        sku_list = list(sku_to_offer.keys())
        ad_data = {}
        if sku_list:
            ads_rows = ch.query("""
                SELECT sku,
                       sumIf(money_spent, dt >= today() - 7)  AS ad_7d,
                       sumIf(money_spent, dt <  today() - 7 AND dt >= today() - 14) AS ad_prev7d,
                       sumIf(views, dt >= today() - 7)        AS views_7d,
                       sumIf(clicks, dt >= today() - 7)       AS clicks_7d,
                       sumIf(orders, dt >= today() - 7)       AS ad_orders_7d,
                       sumIf(add_to_cart, dt >= today() - 7)  AS carts_7d
                FROM mms_analytics.fact_ozon_ad_daily FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND dt >= today() - 14
                GROUP BY sku
            """, parameters={"shop_id": shop_id})
            for r in ads_rows.result_rows:
                oid = sku_to_offer.get(r[0])
                if oid:
                    ad_data[oid] = {
                        "spend_7d": float(r[1]),
                        "spend_prev": float(r[2]),
                        "views": r[3], "clicks": r[4],
                        "orders": r[5], "carts": r[6],
                    }

        for offer in sku_totals:
            s = s7.get(offer, {"q7": 0, "qp7": 0})
            a = ad_data.get(offer, {"spend_7d": 0})
            has_ads = a["spend_7d"] > 0
            if has_ads and s["qp7"] > 0 and s["q7"] > s["qp7"]:
                boost_map[offer] = min(s["q7"] / s["qp7"], 2.0)
            elif has_ads and s["qp7"] == 0 and s["q7"] > 0:
                boost_map[offer] = 1.3
            else:
                boost_map[offer] = 1.0
    else:
        ad_data = {}

    # ── 4. Build recommendations ─────────────────────────────
    items = []
    total_need = 0
    critical_count = 0
    attention_count = 0
    weighted_days_sum = 0.0
    weighted_rev_sum = 0.0

    all_offers = set(list(sku_totals.keys()) + list(fbo_stock.keys()))

    for offer in all_offers:
        stock_info = fbo_stock.get(offer, {"free": 0, "reserved": 0, "wh_count": 0})
        stock = stock_info["free"]
        sold = sku_totals.get(offer, 0)
        boost = boost_map.get(offer, 1.0)
        daily = sold / sales_period if sales_period > 0 else 0
        boosted_daily = daily * boost
        days_supply = stock / boosted_daily if boosted_daily > 0 else 9999

        info = product_info.get(offer, {})
        ad = ad_data.get(offer, {})
        revenue = sum(c["revenue"] for c in sku_clusters.get(offer, []))

        # Status
        if days_supply < 14:
            item_status = "critical"
            critical_count += 1
        elif days_supply < target_days:
            item_status = "attention"
            attention_count += 1
        else:
            item_status = "ok"

        # Weighted days for KPI
        if revenue > 0:
            weighted_days_sum += days_supply * revenue
            weighted_rev_sum += revenue

        # Per-cluster recs — use REAL stock on the cluster's warehouse(s)
        clusters_out = []
        item_need = 0
        offer_cluster_stocks = fbo_by_cluster.get(offer, {})
        for cl in sku_clusters.get(offer, []):
            share = cl["qty"] / sold if sold > 0 else 0
            cl_daily = (cl["qty"] / sales_period) * boost

            # Real stock on the warehouse(s) serving this demand cluster
            wh_stock = offer_cluster_stocks.get(cl["cluster"], 0)
            cl_need = max(0, round(cl_daily * target_days * safety - wh_stock))
            item_need += cl_need
            hub_name, hub_hours = _resolve_hub(cl["cluster"])

            # List actual warehouses with stock for this cluster
            cluster_whs = CLUSTER_TO_WAREHOUSES_OZON.get(cl["cluster"], [])
            wh_with_stock = []
            offer_wh_stocks = fbo_by_wh.get(offer, {})
            for cwh in cluster_whs:
                cwh_qty = offer_wh_stocks.get(cwh, 0)
                if cwh_qty > 0:
                    wh_with_stock.append(f"{cwh} ({cwh_qty})")

            clusters_out.append({
                "cluster": cl["cluster"],
                "sold": cl["qty"],
                "share": round(share * 100, 1),
                "daily": round(cl["qty"] / sales_period, 2),
                "daily_boosted": round(cl_daily, 2),
                "wh_stock": wh_stock,
                "est_stock": wh_stock,  # backward compat — same as wh_stock now
                "need": cl_need,
                "revenue": cl["revenue"],
                "hub": hub_name,
                "hub_hours": hub_hours,
                "warehouses": wh_with_stock,
            })

        total_need += item_need

        items.append({
            "offer_id": offer,
            "name": info.get("name", offer),
            "image_url": info.get("image_url", ""),
            "sku": info.get("sku", 0),
            "sold": sold,
            "revenue": revenue,
            "fbo_stock": stock,
            "fbo_reserved": stock_info["reserved"],
            "fbo_warehouses": stock_info["wh_count"],
            "daily_avg": round(daily, 2),
            "boost": round(boost, 2),
            "boosted_daily": round(boosted_daily, 2),
            "days_supply": round(days_supply, 1),
            "status": item_status,
            "total_need": item_need,
            # Ad info
            "ad_spend_7d": ad.get("spend_7d", 0),
            "ad_views_7d": ad.get("views", 0),
            "ad_clicks_7d": ad.get("clicks", 0),
            "ad_orders_7d": ad.get("orders", 0),
            "ad_carts_7d": ad.get("carts", 0),
            "clusters": clusters_out,
        })

    # ── 4b. Cross-cluster analysis: effective_days & post_restock_days ──
    #   For each cluster with stock, check if the warehouse ALSO ships
    #   to other clusters (cross-drain). If so, the real stock horizon
    #   is shorter than days_supply suggests.
    for item in items:
        offer = item["offer_id"]
        offer_cross = wh_consumption.get(offer, {})

        for cl in item["clusters"]:
            cluster = cl["cluster"]
            wh_stock = cl.get("wh_stock", 0)

            # What did this cluster's warehouse actually ship?
            consumption = offer_cross.get(cluster, {})
            own_qty = consumption.get(cluster, 0)       # orders TO own cluster
            total_qty = sum(consumption.values())       # ALL orders FROM this wh
            cross_qty = total_qty - own_qty             # orders to OTHER clusters

            total_daily = total_qty / sales_period if sales_period > 0 else 0
            own_daily = own_qty / sales_period if sales_period > 0 else 0

            # effective_days: stock / real consumption (includes cross-drain)
            eff_days = wh_stock / total_daily if total_daily > 0 else 9999
            # post_restock_days: if all empty clusters get restocked, only own demand remains
            post_days = wh_stock / own_daily if own_daily > 0 else 9999

            # Cross-cluster details
            cross_clusters = []
            for dest_cl, q in consumption.items():
                if dest_cl != cluster and q > 0:
                    cross_clusters.append({
                        "cluster": dest_cl,
                        "qty": q,
                        "daily": round(q / sales_period, 2),
                    })
            cross_clusters.sort(key=lambda x: x["qty"], reverse=True)

            cl["effective_days"] = round(eff_days, 1)
            cl["post_restock_days"] = round(post_days, 1)
            cl["cross_consumption"] = cross_qty
            cl["cross_clusters"] = cross_clusters

        # Recompute item-level effective_days: weighted by revenue
        total_rev = sum(c["revenue"] for c in item["clusters"])
        if total_rev > 0:
            item_eff_days = sum(
                c["effective_days"] * c["revenue"] for c in item["clusters"]
            ) / total_rev
        else:
            item_eff_days = item["days_supply"]
        item["effective_days"] = round(item_eff_days, 1)

        # Re-evaluate status using effective_days if it's worse
        if item["effective_days"] < item["days_supply"]:
            if item["effective_days"] < 14 and item["status"] != "critical":
                item["status"] = "critical"
                critical_count += 1
                if item.get("_was_attention"):
                    attention_count -= 1
            elif item["effective_days"] < target_days and item["status"] == "ok":
                item["status"] = "attention"
                attention_count += 1

    # Sort: critical first, then by days_supply ascending
    items.sort(key=lambda x: (
        0 if x["status"] == "critical" else 1 if x["status"] == "attention" else 2,
        min(x["days_supply"], x.get("effective_days", x["days_supply"])),
    ))

    avg_days = round(weighted_days_sum / weighted_rev_sum, 1) if weighted_rev_sum > 0 else 0

    kpi = {
        "total_need": total_need,
        "critical_count": critical_count,
        "attention_count": attention_count,
        "avg_days_supply": avg_days,
        "total_fbo": sum(s["free"] for s in fbo_stock.values()),
        "total_sku": len(all_offers),
    }

    # ── 5. Build hub-level aggregation ──────────────────────────
    hubs_map: dict[str, dict] = {}  # hub_name → {items: [...], total_need, total_revenue}
    for item in items:
        for cl in item["clusters"]:
            if cl["need"] <= 0:
                continue
            hub = cl["hub"]
            if hub not in hubs_map:
                hubs_map[hub] = {"hub": hub, "items": [], "total_need": 0, "total_revenue": 0.0}
            hubs_map[hub]["items"].append({
                "offer_id": item["offer_id"],
                "name": item["name"],
                "image_url": item["image_url"],
                "cluster": cl["cluster"],
                "need": cl["need"],
                "wh_stock": cl.get("wh_stock", 0),
                "revenue": cl["revenue"],
                "hub_hours": cl["hub_hours"],
                "daily_boosted": cl["daily_boosted"],
            })
            hubs_map[hub]["total_need"] += cl["need"]
            hubs_map[hub]["total_revenue"] += cl["revenue"]

    hubs = sorted(hubs_map.values(), key=lambda h: h["total_need"], reverse=True)

    return kpi, items, hubs


# ══════════════════════════════════════════════════════════════
# GET /warehouses/ozon/supply
# ══════════════════════════════════════════════════════════════

@router.get("/ozon/supply")
async def get_ozon_supply(
    shop_id: int = Query(..., description="Shop ID"),
    sales_period: int = Query(30, ge=7, le=90, description="Sales period for avg daily calc"),
    target_days: int = Query(60, ge=14, le=90, description="Target stock horizon"),
    safety: float = Query(1.15, ge=1.0, le=2.0, description="Safety coefficient"),
    use_ad_boost: bool = Query(True, description="Apply ad boost coefficient"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Supply recommendations per SKU × cluster for Ozon FBO."""

    # Verify shop
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "ozon":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Shop not found")

    from app.core.clickhouse import get_clickhouse_client
    try:
        ch = get_clickhouse_client()
    except Exception as e:
        logger.error("ClickHouse connection error: %s", e)
        raise HTTPException(status_code=500, detail="Analytics unavailable")

    # Product info from PostgreSQL
    pg_result = await db.execute(text("""
        SELECT offer_id, name, sku,
               COALESCE(NULLIF(primary_image_url, ''), main_image_url, '') AS image_url
        FROM dim_ozon_products
        WHERE shop_id = :shop_id
    """), {"shop_id": shop_id})

    product_info = {}
    for r in pg_result.fetchall():
        product_info[r[0]] = {"name": r[1] or r[0], "sku": r[2], "image_url": r[3] or ""}

    kpi, items, hubs = _compute_supply(
        ch, shop_id, sales_period, target_days, safety, use_ad_boost, product_info,
    )

    return {
        "shop_id": shop_id,
        "sales_period": sales_period,
        "target_days": target_days,
        "safety": safety,
        "use_ad_boost": use_ad_boost,
        "kpi": kpi,
        "items": items,
        "hubs": hubs,
    }


# ══════════════════════════════════════════════════════════════
# GET /warehouses/ozon/supply/export — Excel
# ══════════════════════════════════════════════════════════════

@router.get("/ozon/supply/export")
async def export_ozon_supply(
    shop_id: int = Query(...),
    sales_period: int = Query(30, ge=7, le=90),
    target_days: int = Query(60, ge=14, le=90),
    safety: float = Query(1.15, ge=1.0, le=2.0),
    use_ad_boost: bool = Query(True),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download supply recommendations as Excel."""

    # Verify shop
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "ozon":
        raise HTTPException(status_code=404, detail="Shop not found")

    from app.core.clickhouse import get_clickhouse_client
    ch = get_clickhouse_client()

    pg_result = await db.execute(text("""
        SELECT offer_id, name, sku,
               COALESCE(NULLIF(primary_image_url, ''), main_image_url, '') AS image_url
        FROM dim_ozon_products
        WHERE shop_id = :shop_id
    """), {"shop_id": shop_id})
    product_info = {}
    for r in pg_result.fetchall():
        product_info[r[0]] = {"name": r[1] or r[0], "sku": r[2], "image_url": r[3] or ""}

    kpi, items, hubs = _compute_supply(
        ch, shop_id, sales_period, target_days, safety, use_ad_boost, product_info,
    )

    # ── Build Excel ──────────────────────────────────────────
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # === Sheet 1: Рекомендации ===
    ws = wb.active
    ws.title = "Рекомендации"

    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    need_fill = PatternFill("solid", fgColor="FFF2CC")
    sku_fill = PatternFill("solid", fgColor="D9E2F3")
    sku_font = Font(bold=True, size=11)
    boost_font = Font(bold=True, color="FF6600")
    thin = Side(style="thin", color="D0D0D0")
    num_fmt = "#,##0"
    pct_fmt = "0.0%"

    from openpyxl.utils import get_column_letter

    headers = [
        ("Статус", 12), ("Артикул", 24), ("Кластер", 40),
        ("Склад отгрузки", 28), ("Доставка, ч", 10),
        ("Продано", 10), ("Доля", 8), ("Ежедн.", 8),
        ("Boost", 7), ("Ежедн×b", 8), ("FBO", 8),
        ("Сток РФЦ", 10), ("Склады", 32), ("Дн.зап", 8),
        ("Реал.зап", 9), ("Кросс", 36), ("ПОСТАВИТЬ", 12),
        ("Выручка", 12), ("Рекл₽", 10), ("Показы", 8),
        ("Клики", 7), ("Корзины", 8), ("Р.заказы", 8),
    ]
    for ci, (name, w) in enumerate(headers, 1):
        c = ws.cell(1, ci, name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    row = 2
    for item in items:
        status_label = {"critical": "🔴 Крит.", "attention": "🟡 Вним.", "ok": "🟢 Норма"}[item["status"]]
        s_fill = {"critical": red_fill, "attention": yellow_fill, "ok": green_fill}[item["status"]]
        boost_str = f"×{item['boost']:.2f}" if item["boost"] > 1.0 else "-"
        first = True

        for cl in item["clusters"]:
            r = row
            if first:
                ws.cell(r, 1, status_label).fill = s_fill
                ws.cell(r, 1).font = Font(bold=True, size=10)
                c = ws.cell(r, 2, item["offer_id"])
                c.font = sku_font
                c.fill = sku_fill
            else:
                ws.cell(r, 1, "")
                ws.cell(r, 2, "")

            ws.cell(r, 3, cl["cluster"])
            ws.cell(r, 4, cl["hub"])
            c_dh = ws.cell(r, 5, cl["hub_hours"])
            if cl["hub_hours"] <= 28:
                c_dh.font = Font(bold=True, color="00AA00")
            elif cl["hub_hours"] <= 45:
                c_dh.font = Font(bold=True, color="CC8800")
            ws.cell(r, 6, cl["sold"]).number_format = num_fmt
            ws.cell(r, 7, cl["share"] / 100).number_format = pct_fmt
            ws.cell(r, 8, cl["daily"]).number_format = "0.00"

            if first:
                c = ws.cell(r, 9, boost_str)
                if item["boost"] > 1.0:
                    c.font = boost_font
            else:
                ws.cell(r, 9, "")

            ws.cell(r, 10, cl["daily_boosted"]).number_format = "0.00"

            if first:
                ws.cell(r, 11, item["fbo_stock"]).number_format = num_fmt
            else:
                ws.cell(r, 11, "")

            # Col 12: Real stock on this cluster's warehouse(s)
            ws.cell(r, 12, cl.get("wh_stock", 0)).number_format = num_fmt

            # Col 13: Warehouse names with stock
            wh_list = cl.get("warehouses", [])
            ws.cell(r, 13, ", ".join(wh_list) if wh_list else "—")

            if first:
                c = ws.cell(r, 14, round(item["days_supply"], 1))
                c.number_format = "0.0"
                if item["days_supply"] < 14:
                    c.font = Font(bold=True, color="CC0000")
            else:
                ws.cell(r, 14, "")

            # Col 15: Реал.зап (effective_days with cross-drain)
            eff_d = cl.get("effective_days", 9999)
            if first and eff_d < 9999:
                c = ws.cell(r, 15, eff_d)
                c.number_format = "0.0"
                if eff_d < item["days_supply"] * 0.8:  # significantly lower
                    c.font = Font(bold=True, color="CC0000")
                    c.fill = PatternFill("solid", fgColor="FFC7CE")
                elif eff_d < item["days_supply"]:
                    c.font = Font(bold=True, color="CC8800")
            elif first:
                ws.cell(r, 15, "—")
            else:
                ws.cell(r, 15, "")

            # Col 16: Кросс (cross-cluster drain description)
            cross_cls = cl.get("cross_clusters", [])
            if cross_cls:
                cross_desc = ", ".join(
                    f"{cc['cluster']} +{cc['daily']}/д" for cc in cross_cls[:3]
                )
                if len(cross_cls) > 3:
                    cross_desc += f" +{len(cross_cls)-3} ещё"
                ws.cell(r, 16, cross_desc).font = Font(italic=True, size=9, color="CC6600")
            else:
                ws.cell(r, 16, "—")

            c = ws.cell(r, 17, cl["need"])
            c.number_format = num_fmt
            if cl["need"] > 0:
                c.font = Font(bold=True, size=12, color="CC0000")
                c.fill = need_fill

            ws.cell(r, 18, cl["revenue"]).number_format = num_fmt

            if first:
                ws.cell(r, 19, item.get("ad_spend_7d", 0)).number_format = num_fmt
                ws.cell(r, 20, item.get("ad_views_7d", 0)).number_format = num_fmt
                ws.cell(r, 21, item.get("ad_clicks_7d", 0)).number_format = num_fmt
                ws.cell(r, 22, item.get("ad_carts_7d", 0)).number_format = num_fmt
                ws.cell(r, 23, item.get("ad_orders_7d", 0)).number_format = num_fmt
                for ci2 in range(1, len(headers) + 1):
                    ws.cell(r, ci2).border = Border(top=Side(style="medium", color="2F5496"))

            first = False
            row += 1

    # === Sheet 2: Сводка ===
    ws2 = wb.create_sheet("Сводка")
    s_hdr = [
        ("Статус", 12), ("Артикул", 24), ("Продано", 10), ("Выручка", 14),
        ("FBO", 8), ("Ежедн.", 8), ("Boost", 7), ("Дн.зап", 8),
        ("ПОСТАВИТЬ", 12), ("Кластеров", 9),
    ]
    for ci, (name, w) in enumerate(s_hdr, 1):
        c = ws2.cell(1, ci, name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"

    for ri, item in enumerate(items, 2):
        sl = {"critical": "🔴 Крит.", "attention": "🟡 Вним.", "ok": "🟢 Норма"}[item["status"]]
        sf = {"critical": red_fill, "attention": yellow_fill, "ok": green_fill}[item["status"]]
        vals = [
            sl, item["offer_id"], item["sold"], item["revenue"],
            item["fbo_stock"], item["daily_avg"],
            f"×{item['boost']:.2f}" if item["boost"] > 1 else "-",
            round(item["days_supply"], 1), item["total_need"], len(item["clusters"]),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(ri, ci, v)
            if ci == 1:
                c.fill = sf
                c.font = Font(bold=True)
            if ci in (3, 4, 5, 9):
                c.number_format = num_fmt
            if ci == 9 and v > 0:
                c.font = Font(bold=True, size=12, color="CC0000")
                c.fill = need_fill

    # === Sheet 3: Методология ===
    ws3 = wb.create_sheet("Методология")
    ws3.column_dimensions["A"].width = 90
    lines = [
        ("МЕТОДОЛОГИЯ РЕКОМЕНДАЦИИ ПОСТАВКИ FBO", True, 14),
        ("", False, 11),
        (f"Магазин: {shop.name or shop_id} | Период продаж: {sales_period}д | Target: {target_days}д | Safety: {safety} | Ad boost: {'ДА' if use_ad_boost else 'НЕТ'}", False, 11),
        ("", False, 11),
        ("═══ ФОРМУЛА ═══", True, 12),
        ("daily_cluster = продажи_кластер / период × ad_boost", False, 11),
        (f"target_stock = daily_cluster × {target_days} дней × {safety} (safety)", False, 11),
        ("supply_need = max(0, target_stock − РЕАЛЬНЫЙ_сток_на_РФЦ_кластера)", False, 11),
        ("", False, 11),
        ("═══ СТОКИ ═══", True, 12),
        ("• Стоки берутся ПО КАЖДОМУ СКЛАДУ из fact_ozon_warehouse_stocks", False, 11),
        ("• Склады привязаны к кластерам: КАЗАНЬ_РФЦ → Казань, САМАРА_РФЦ → Самара и т.д.", False, 11),
        ("• «Сток РФЦ» = реальный FBO остаток на складе(ах) обслуживающих этот кластер", False, 11),
        ("", False, 11),
        ("═══ ПАРАМЕТРЫ ═══", True, 12),
        (f"• Target = {target_days} дней", False, 11),
        (f"• Safety = ×{safety} ({round((safety-1)*100)}%)", False, 11),
        ("• Распределение: по реальным стокам на складах кластеров", False, 11),
        ("", False, 11),
        ("═══ AD BOOST ═══", True, 12),
        ("Реклама + рост продаж → boost = min(рост_7д, 2.0x)", False, 11),
        ("Реклама + новый SKU → boost = 1.3x", False, 11),
        ("Нет рекламы → boost = 1.0", False, 11),
        ("", False, 11),
        ("═══ СТАТУСЫ ═══", True, 12),
        ("🔴 Критический — запас < 14 дней", False, 11),
        (f"🟡 Внимание — запас < {target_days} дней", False, 11),
        (f"🟢 Норма — запас ≥ {target_days} дней", False, 11),
        ("", False, 11),
        ("═══ КРОСС-КЛАСТЕРНЫЙ АНАЛИЗ ═══", True, 12),
        ("• Анализируем warehouse_name из fact_ozon_orders — с какого склада реально уехал заказ", False, 11),
        ("• Если склад кластера А отгружает заказы в кластер Б (потому что на Б пусто) —", False, 11),
        ("  это «кросс-слив»: сток А расходуется быстрее расчётного", False, 11),
        ("• «Реал.зап» = сток / ФАКТИЧЕСКИЙ расход склада (включая чужие заказы)", False, 11),
        ("• «Кросс» = какие кластеры «паразитируют» на этом складе и сколько шт/день", False, 11),
        ("• Если «Реал.зап» << «Дн.зап» — склад тратится быстрее чем кажется!", False, 11),
        ("• После поставки в пустые кластеры — кросс-нагрузка уйдёт, и расход нормализуется", False, 11),
    ]
    for ri, (t, bold, sz) in enumerate(lines, 1):
        c = ws3.cell(ri, 1, t)
        c.font = Font(bold=bold, size=sz)
        if "═══" in t:
            c.fill = PatternFill("solid", fgColor="D9E2F3")

    # === Sheet 4: Поставка по кластерам ===
    ws4 = wb.create_sheet("Поставка по кластерам")
    h4_headers = [
        ("Склад отгрузки (хаб)", 28), ("Артикул", 24), ("Название", 44),
        ("Кластер спроса", 28), ("Время доставки, ч", 14),
        ("Ежедн.×b", 10), ("Сток РФЦ", 10), ("ПОСТАВИТЬ", 12), ("Выручка", 12),
    ]
    for ci, (name, w) in enumerate(h4_headers, 1):
        c = ws4.cell(1, ci, name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.freeze_panes = "A2"

    hub_fill = PatternFill("solid", fgColor="BDD7EE")
    hub_font = Font(bold=True, size=11, color="1F4E79")
    row4 = 2
    for hub_data in hubs:
        # Hub header row
        c = ws4.cell(row4, 1, f"📦 {hub_data['hub']}")
        c.font = hub_font
        c.fill = hub_fill
        ws4.cell(row4, 8, hub_data["total_need"]).font = Font(bold=True, size=12, color="CC0000")
        ws4.cell(row4, 8).number_format = num_fmt
        ws4.cell(row4, 9, round(hub_data["total_revenue"])).number_format = num_fmt
        for ci2 in range(1, len(h4_headers) + 1):
            ws4.cell(row4, ci2).fill = hub_fill
        row4 += 1

        # Items in this hub
        for hi in hub_data["items"]:
            ws4.cell(row4, 1, "")
            ws4.cell(row4, 2, hi["offer_id"])
            ws4.cell(row4, 3, hi["name"])
            ws4.cell(row4, 4, hi["cluster"])
            ws4.cell(row4, 5, hi["hub_hours"])
            ws4.cell(row4, 6, hi["daily_boosted"]).number_format = "0.00"
            ws4.cell(row4, 7, hi.get("wh_stock", 0)).number_format = num_fmt
            c = ws4.cell(row4, 8, hi["need"])
            c.number_format = num_fmt
            if hi["need"] > 0:
                c.font = Font(bold=True, color="CC0000")
                c.fill = need_fill
            ws4.cell(row4, 9, round(hi["revenue"])).number_format = num_fmt
            row4 += 1

        row4 += 1  # blank row between hubs

    # === Sheet 5: Объединённые кластеры ===
    ws5 = wb.create_sheet("Объединённые кластеры")
    h5_headers = [
        ("Хаб отгрузки", 30), ("Обслуживаемые кластеры", 50),
        ("Артикул", 24), ("Название", 44),
        ("Кластер спроса", 28), ("Время доставки, ч", 14),
        ("Сток РФЦ", 10), ("ПОСТАВИТЬ", 12), ("Выручка", 12),
    ]
    for ci, (name, w) in enumerate(h5_headers, 1):
        c = ws5.cell(1, ci, name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws5.column_dimensions[get_column_letter(ci)].width = w
    ws5.freeze_panes = "A2"

    # Build consolidated data: group_hub → {sku → [{cluster, need, revenue, hours}]}
    consol: dict[str, dict] = {}  # hub → {items: {offer_id → [...]}, total_need, ...}
    for item in items:
        for cl in item["clusters"]:
            if cl["need"] <= 0:
                continue
            group_hub = _CLUSTER_TO_GROUP.get(cl["cluster"], cl["cluster"])
            if group_hub not in consol:
                consol[group_hub] = {"items": {}, "total_need": 0, "total_revenue": 0.0}
            g = consol[group_hub]
            g["total_need"] += cl["need"]
            g["total_revenue"] += cl["revenue"]
            oid = item["offer_id"]
            if oid not in g["items"]:
                g["items"][oid] = {"name": item["name"], "clusters": []}
            delivery_h = DELIVERY_HOURS.get(group_hub, {}).get(cl["cluster"], 999)
            g["items"][oid]["clusters"].append({
                "cluster": cl["cluster"],
                "need": cl["need"],
                "wh_stock": cl.get("wh_stock", 0),
                "revenue": cl["revenue"],
                "hours": delivery_h,
            })

    # Sort by total_need desc
    sorted_groups = sorted(consol.items(), key=lambda x: x[1]["total_need"], reverse=True)

    grp_fill = PatternFill("solid", fgColor="E2EFDA")
    grp_font = Font(bold=True, size=12, color="375623")
    sub_fill = PatternFill("solid", fgColor="F2F7ED")
    row5 = 2

    for group_hub, gdata in sorted_groups:
        members = CONSOLIDATED_GROUPS.get(group_hub, [group_hub])
        members_str = ", ".join(members)

        # Group header row
        c = ws5.cell(row5, 1, f"📦 {group_hub}")
        c.font = grp_font
        ws5.cell(row5, 2, members_str).font = Font(italic=True, size=10, color="666666")
        ws5.cell(row5, 8, gdata["total_need"]).font = Font(bold=True, size=13, color="CC0000")
        ws5.cell(row5, 8).number_format = num_fmt
        ws5.cell(row5, 9, round(gdata["total_revenue"])).number_format = num_fmt
        for ci2 in range(1, len(h5_headers) + 1):
            ws5.cell(row5, ci2).fill = grp_fill
        row5 += 1

        # Items in this group
        for oid, odata in gdata["items"].items():
            first_cl = True
            for cl_info in odata["clusters"]:
                if first_cl:
                    ws5.cell(row5, 3, oid).font = Font(bold=True, size=10)
                    ws5.cell(row5, 4, odata["name"])
                else:
                    ws5.cell(row5, 3, "")
                    ws5.cell(row5, 4, "")

                ws5.cell(row5, 5, cl_info["cluster"])
                c_h = ws5.cell(row5, 6, cl_info["hours"])
                if cl_info["hours"] <= 28:
                    c_h.font = Font(bold=True, color="00AA00")
                elif cl_info["hours"] <= 45:
                    c_h.font = Font(bold=True, color="CC8800")
                else:
                    c_h.font = Font(color="CC0000")

                ws5.cell(row5, 7, cl_info.get("wh_stock", 0)).number_format = num_fmt

                c = ws5.cell(row5, 8, cl_info["need"])
                c.number_format = num_fmt
                if cl_info["need"] > 0:
                    c.font = Font(bold=True, color="CC0000")
                    c.fill = need_fill

                ws5.cell(row5, 9, round(cl_info["revenue"])).number_format = num_fmt

                first_cl = False
                row5 += 1

        row5 += 1  # blank row

    # === Sheet 6: Анализ логистики (по объединённым группам) ===
    THRESHOLD_HOURS = 29  # Ozon recommended avg delivery time

    ws6 = wb.create_sheet("Анализ логистики")
    h6_headers = [
        ("Артикул", 24), ("Название", 40),
        ("Хаб отгрузки", 28), ("Кластер спроса", 28),
        ("Продано (шт)", 10), ("Время от хаба, ч", 12),
        ("Превышает 29ч", 10), ("Влияние (шт×ч)", 14), ("Доля влияния %", 12),
        ("Ср. время (взвеш.)", 14), ("Рекомендация", 55),
    ]
    for ci, (name, w) in enumerate(h6_headers, 1):
        c = ws6.cell(1, ci, name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws6.column_dimensions[get_column_letter(ci)].width = w
    ws6.freeze_panes = "A2"

    # Threshold info row
    ws6.cell(2, 1, "⚙ Методика Ozon: порог 29ч").font = Font(bold=True, size=11, color="1F4E79")
    ws6.cell(2, 2, (
        "Анализ по объединённым группам. "
        "Когда товар грузится в 1 хаб (напр. Москва), "
        "маршруты в дальние кластеры (Тверь 45ч, Ярославль 45ч) увеличивают среднее время. "
        "Доля влияния показывает, какие маршруты дороже всего."
    )).font = Font(italic=True, size=9, color="666666")
    for ci2 in range(1, len(h6_headers) + 1):
        ws6.cell(2, ci2).fill = PatternFill("solid", fgColor="DAEEF3")
    row6 = 4

    high_fill = PatternFill("solid", fgColor="FCE4D6")
    critical_font = Font(bold=True, color="CC0000")
    warn_font = Font(bold=True, color="CC8800")
    ok_font = Font(color="00AA00")

    for item in items:
        if not item["clusters"] or item["sold"] <= 0:
            continue

        # Group clusters by consolidated hub
        hub_groups: dict[str, list] = {}  # hub_name → [{cluster, vol, hours_from_hub}]
        for cl in item["clusters"]:
            if cl["sold"] <= 0:
                continue
            group_hub = _CLUSTER_TO_GROUP.get(cl["cluster"], cl["cluster"])
            hours_from_hub = DELIVERY_HOURS.get(group_hub, {}).get(cl["cluster"], 60)
            if group_hub not in hub_groups:
                hub_groups[group_hub] = []
            hub_groups[group_hub].append({
                "cluster": cl["cluster"],
                "vol": cl["sold"],
                "hours": hours_from_hub,
            })

        if not hub_groups:
            continue

        # Calculate influence across ALL routes for this SKU
        all_routes = []
        for hub_name, members in hub_groups.items():
            for m in members:
                exceeds = m["hours"] > THRESHOLD_HOURS
                influence = m["vol"] * m["hours"] if exceeds else 0
                all_routes.append({
                    "hub": hub_name,
                    "cluster": m["cluster"],
                    "vol": m["vol"],
                    "hours": m["hours"],
                    "exceeds": exceeds,
                    "influence": influence,
                })

        total_influence = sum(r["influence"] for r in all_routes)
        total_vol = sum(r["vol"] for r in all_routes)
        weighted_hours = sum(r["vol"] * r["hours"] for r in all_routes)
        avg_hours = weighted_hours / total_vol if total_vol > 0 else 0

        # Generate recommendation
        recommendation = ""
        problem_routes = [r for r in all_routes if r["exceeds"]]
        if not problem_routes or avg_hours <= THRESHOLD_HOURS:
            recommendation = "✅ Среднее время ≤29ч — оптимально"
        else:
            # Find top problem route and suggest direct shipment
            top = max(problem_routes, key=lambda r: r["influence"])
            pct = (top["influence"] / total_influence * 100) if total_influence > 0 else 0
            dest = top["cluster"]
            local_h = DELIVERY_HOURS.get(dest, {}).get(dest, 28)
            if top["hub"] != dest and local_h <= THRESHOLD_HOURS:
                recommendation = (
                    f"⚠ Отделить {top['vol']} шт → напрямую в {dest} "
                    f"({top['hours']}ч → {local_h}ч). "
                    f"Снизит влияние на {pct:.0f}%"
                )
            else:
                recommendation = (
                    f"⚠ {top['hub']}→{dest} ({top['hours']}ч) = "
                    f"{pct:.0f}% влияния. Среднее {avg_hours:.0f}ч"
                )

        # SKU header row
        sku_hdr_fill = PatternFill("solid", fgColor="D6E4F0")
        ws6.cell(row6, 1, item["offer_id"]).font = Font(bold=True, size=11)
        ws6.cell(row6, 2, item["name"]).font = Font(bold=True, size=10)
        c_avg = ws6.cell(row6, 10, round(avg_hours, 1))
        c_avg.number_format = "0.0"
        c_avg.font = (
            critical_font if avg_hours > 45
            else warn_font if avg_hours > THRESHOLD_HOURS
            else ok_font
        )
        ws6.cell(row6, 11, recommendation).font = (
            Font(bold=True, size=10, color="CC6600")
            if avg_hours > THRESHOLD_HOURS
            else Font(size=10)
        )
        for ci2 in range(1, len(h6_headers) + 1):
            ws6.cell(row6, ci2).fill = sku_hdr_fill
        row6 += 1

        # Route rows sorted by influence desc
        for r in sorted(all_routes, key=lambda x: x["influence"], reverse=True):
            share_pct = (r["influence"] / total_influence * 100) if total_influence > 0 else 0

            ws6.cell(row6, 3, r["hub"])
            ws6.cell(row6, 4, r["cluster"])
            ws6.cell(row6, 5, r["vol"]).number_format = num_fmt

            c_h = ws6.cell(row6, 6, r["hours"])
            c_h.font = (
                ok_font if r["hours"] <= 28
                else warn_font if r["hours"] <= 45
                else critical_font
            )

            ws6.cell(row6, 7, "Да" if r["exceeds"] else "—").font = (
                Font(bold=True, color="CC0000") if r["exceeds"]
                else Font(color="999999")
            )

            ws6.cell(row6, 8, r["influence"]).number_format = num_fmt
            c_s = ws6.cell(row6, 9, round(share_pct, 1) if total_influence > 0 else 0)
            c_s.number_format = "0.0"
            if share_pct >= 40:
                c_s.font = critical_font
                c_s.fill = high_fill
            elif share_pct >= 20:
                c_s.font = warn_font

            row6 += 1

        row6 += 1  # blank separator

    # ── Save & return ────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"supply_{shop_id}_{target_days}d.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ══════════════════════════════════════════════════════════════
# WB Supply — рекомендации поставок с учётом хранения
# ══════════════════════════════════════════════════════════════

def _parse_ru_float(s: str) -> float:
    """Parse Russian-formatted float: '0,12' → 0.12"""
    if not s or s == "-":
        return 0.0
    return float(s.replace(",", "."))


# ═══════════════════════════════════════════════════════════════
# Маппинг: склад WB → его «домашний» федеральный округ
# Используется для кросс-складского анализа:
# если warehouse_name.okrug != oblast_okrug_name → кросс-слив
# ═══════════════════════════════════════════════════════════════
WAREHOUSE_TO_OKRUG: dict[str, str] = {
    # ЦФО (Московская обл. + Центральная Россия)
    "Котовск": "Центральный федеральный округ",
    "Котовск: Питание": "Центральный федеральный округ",
    "Электросталь": "Центральный федеральный округ",
    "Электросталь: Питание": "Центральный федеральный округ",
    "Рязань (Тюшевское)": "Центральный федеральный округ",
    "Рязань (Тюшевское): Питание": "Центральный федеральный округ",
    "Истра": "Центральный федеральный округ",
    "Подольск 3": "Центральный федеральный округ",
    "Подольск 4": "Центральный федеральный округ",
    "Коледино": "Центральный федеральный округ",
    "Белые Столбы": "Центральный федеральный округ",
    "Обухово": "Центральный федеральный округ",
    "Обухово 2": "Центральный федеральный округ",
    "Домодедово-2": "Центральный федеральный округ",
    "Домодедово 2": "Центральный федеральный округ",
    "Домодедово 2: Питание": "Центральный федеральный округ",
    "Пушкино": "Центральный федеральный округ",
    "Чехов 1": "Центральный федеральный округ",
    "Чехов 2": "Центральный федеральный округ",
    "Ногинск": "Центральный федеральный округ",
    "Воронеж": "Центральный федеральный округ",
    "Старый Оскол": "Центральный федеральный округ",
    # СЗО
    "Санкт-Петербург Уткина Заводь": "Северо-Западный федеральный округ",
    "Никольское": "Северо-Западный федеральный округ",
    "Шушары: Питание": "Северо-Западный федеральный округ",
    # ПФО
    "Самара (Новосемейкино)": "Приволжский федеральный округ",
    "Новосемейкино": "Приволжский федеральный округ",
    "Новосемейкино: Питание": "Приволжский федеральный округ",
    "Казань": "Приволжский федеральный округ",
    "Казань: Питание": "Приволжский федеральный округ",
    "Пенза": "Приволжский федеральный округ",
    "Сарапул": "Приволжский федеральный округ",
    # ЮФО
    "Краснодар": "Южный федеральный округ",
    "Краснодар (Тихорецкая)": "Южный федеральный округ",
    "Краснодар (Тихорецкая): Питание": "Южный федеральный округ",
    "Волгоград": "Южный федеральный округ",
    "Волгоград: Питание": "Южный федеральный округ",
    "Ростов, Гайдара 8": "Южный федеральный округ",
    # СКФО
    "Невинномысск": "Северо-Кавказский федеральный округ",
    # Урал
    "Екатеринбург - Перспективная 14": "Уральский федеральный округ",
    "Екатеринбург (Перспективная): Питание": "Уральский федеральный округ",
    "Екатеринбург - Испытателей 14г": "Уральский федеральный округ",
    # Сибирь
    "Новосибирск": "Сибирский федеральный округ",
    "Новосибирск СГТ": "Сибирский федеральный округ",
    # ДВ
    "Владивосток": "Дальневосточный федеральный округ",
    "Владивосток СГТ": "Дальневосточный федеральный округ",
}

# ═══════════════════════════════════════════════════════════════
# Маппинг: округ покупателя → ближайшие склады WB
# Содержит ВСЕ варианты: обычные, : Питание, СГТ
# Фильтрация по типу товара происходит в runtime
# ═══════════════════════════════════════════════════════════════
REGION_TO_WAREHOUSES: dict[str, list[str]] = {
    "Центральный федеральный округ": [
        "Котовск", "Котовск: Питание",
        "Электросталь", "Электросталь: Питание",
        "Рязань (Тюшевское)", "Рязань (Тюшевское): Питание",
        "Истра", "Подольск 3", "Коледино",
        "Белые Столбы", "Обухово",
        "Домодедово-2", "Домодедово 2: Питание",
    ],
    "Северо-Западный федеральный округ": [
        "Санкт-Петербург Уткина Заводь", "Никольское",
        "Шушары: Питание",
    ],
    "Приволжский федеральный округ": [
        "Новосемейкино", "Новосемейкино: Питание",
        "Казань", "Казань: Питание",
        "Пенза",
    ],
    "Южный федеральный округ": [
        "Краснодар (Тихорецкая)", "Краснодар (Тихорецкая): Питание",
        "Волгоград", "Волгоград: Питание",
    ],
    "Северо-Кавказский федеральный округ": [
        "Невинномысск",
        "Краснодар (Тихорецкая)", "Краснодар (Тихорецкая): Питание",
    ],
    "Уральский федеральный округ": [
        "Екатеринбург - Перспективная 14",
        "Екатеринбург (Перспективная): Питание",
        "Екатеринбург - Испытателей 14г",
    ],
    "Сибирский федеральный округ": [
        "Новосибирск", "Новосибирск СГТ",
    ],
    "Дальневосточный федеральный округ": [
        "Владивосток СГТ",
    ],
}

# ── Warehouse type classification ──
_FOOD_SUFFIX = ": Питание"
_SGT_SUFFIX = "СГТ"
_MAX_BOX_WEIGHT_KG = 25  # Ограничение коробов по весу

# Список категорий WB, которые нуждаются в складах «Питание» (Меркурий / пищевая продукция)
_FOOD_CATEGORIES = {
    "Товары для животных",
    "Продукты питания",
    "Здоровое питание",
    "Детское питание",
    "Корма",
}


def _classify_product_wh_type(
    nm_id: int,
    product_categories: dict[int, str],
    wh_stocks: dict[str, dict],
    wh_sales: dict[str, dict],
) -> str:
    """
    Определяет тип товара:
    1) По категории из fact_orders_raw (приоритет)
    2) По складам с стоками/продажами (fallback)
    """
    # 1. По категории
    cat = product_categories.get(nm_id, "")
    if cat in _FOOD_CATEGORIES:
        return "food"

    # 2. По названию склада (если есть на складе с суффиксом)
    all_wh_names = set(list(wh_stocks.keys()) + list(wh_sales.keys()))
    for wh in all_wh_names:
        if _FOOD_SUFFIX in wh:
            return "food"
    for wh in all_wh_names:
        if _SGT_SUFFIX in wh:
            return "sgt"
    return "normal"


def _filter_wh_for_type(
    warehouse_name: str,
    product_type: str,
) -> bool:
    """
    Проверяет, подходит ли склад для типа товара.
    - food  → только ': Питание'
    - sgt   → только 'СГТ'
    - normal → БЕЗ ': Питание' и БЕЗ 'СГТ'
    """
    is_food_wh = _FOOD_SUFFIX in warehouse_name
    is_sgt_wh = _SGT_SUFFIX in warehouse_name

    if product_type == "food":
        return is_food_wh
    elif product_type == "sgt":
        return is_sgt_wh
    else:  # normal
        return not is_food_wh and not is_sgt_wh


def _match_warehouse(name: str, available: set[str]) -> str | None:
    """
    Точное сопоставление склада из маппинга с доступными складами в тарифах.
    Не используем startswith чтобы 'Котовск' не матчил 'Котовск: Питание'.
    """
    # 1. Точное совпадение
    if name in available:
        return name
    # 2. Расширенный поиск только для СГТ (отличаются названия)
    for avail_wh in available:
        # Матч только если оба одного типа
        name_is_food = _FOOD_SUFFIX in name
        avail_is_food = _FOOD_SUFFIX in avail_wh
        name_is_sgt = _SGT_SUFFIX in name
        avail_is_sgt = _SGT_SUFFIX in avail_wh
        if name_is_food != avail_is_food or name_is_sgt != avail_is_sgt:
            continue  # разные типы — не матчим
        # Одинаковый тип — проверяем базовое название
        if avail_wh.startswith(name) or name.startswith(avail_wh.split(":")[0].split(" СГТ")[0]):
            return avail_wh
    return None


def _estimate_weight_kg(vol_liters: float) -> float:
    """Оценка веса из объёма (~0.5 кг/л)."""
    return vol_liters * 0.5


async def _build_wb_supply_data(
    shop_id: int, sales_period: int, target_days: int, safety: float,
    use_ad_boost: bool, db: AsyncSession,
) -> dict:
    """
    Common data builder for WB supply (used by both JSON and XLSX endpoints).
    Returns dict with keys: items, tariffs, kpi, warehouse_summary.
    """
    from app.core.clickhouse import get_clickhouse_client

    ch = get_clickhouse_client()
    today = date.today()
    d_sales_start = today - timedelta(days=sales_period)

    # ── 1. WB Stocks (latest inventory snapshot) ─────────────
    stock_rows = ch.query(f"""
        SELECT
            nm_id,
            warehouse_name,
            argMax(quantity, fetched_at)  AS qty,
            argMax(price, fetched_at)     AS price,
            argMax(discount, fetched_at)  AS discount
        FROM mms_analytics.fact_inventory_snapshot
        WHERE shop_id = {shop_id}
          AND warehouse_name NOT LIKE 'FBS:%'
        GROUP BY nm_id, warehouse_name
        HAVING qty > 0
    """).result_rows

    stocks_by_nm: dict[int, dict] = {}
    for row in stock_rows:
        nm_id = row[0]
        if nm_id not in stocks_by_nm:
            stocks_by_nm[nm_id] = {}
        stocks_by_nm[nm_id][row[1]] = {
            "qty": row[2], "price": float(row[3]), "discount": row[4]
        }

    # ── 2. WB Orders by warehouse ────────────────────────────
    sales_rows = ch.query(f"""
        SELECT
            nm_id,
            warehouse_name,
            count()          AS orders_count,
            sum(price_with_disc) AS revenue
        FROM mms_analytics.fact_orders_raw
        WHERE shop_id = {shop_id}
          AND date >= '{d_sales_start}'
          AND date <= '{today}'
          AND is_cancel = 0
        GROUP BY nm_id, warehouse_name
    """).result_rows

    sales_by_nm: dict[int, dict] = {}
    for row in sales_rows:
        nm_id = row[0]
        if nm_id not in sales_by_nm:
            sales_by_nm[nm_id] = {}
        sales_by_nm[nm_id][row[1]] = {
            "orders": row[2], "revenue": float(row[3])
        }

    # ── 2b. Ad boost (7d vs prev 7d) ─────────────────────────
    boost_map: dict[int, float] = {}  # nm_id → boost coefficient
    if use_ad_boost:
        try:
            # Sales 7d vs prev 7d per nm_id
            s7_rows = ch.query(f"""
                SELECT nm_id,
                       countIf(date >= today() - 7)  AS qty_7d,
                       countIf(date <  today() - 7)  AS qty_prev7d
                FROM mms_analytics.fact_orders_raw
                WHERE shop_id = {shop_id}
                  AND date >= today() - 14
                  AND is_cancel = 0
                GROUP BY nm_id
            """).result_rows
            s7 = {int(r[0]): {"q7": r[1], "qp7": r[2]} for r in s7_rows}

            # Ad spend per nm_id (7d)
            ad_rows = ch.query(f"""
                SELECT nm_id,
                       sumIf(spend, date >= today() - 7)  AS ad_7d
                FROM mms_analytics.fact_advert_stats_v3
                WHERE shop_id = {shop_id}
                  AND date >= today() - 14
                GROUP BY nm_id
            """).result_rows
            ad_spend_7d = {int(r[0]): float(r[1]) for r in ad_rows}

            # Calculate boost
            all_boost_nms = set(list(s7.keys()) + list(ad_spend_7d.keys()))
            for nm in all_boost_nms:
                s = s7.get(nm, {"q7": 0, "qp7": 0})
                has_ads = ad_spend_7d.get(nm, 0) > 0
                if has_ads and s["qp7"] > 0 and s["q7"] > s["qp7"]:
                    boost_map[nm] = min(s["q7"] / s["qp7"], 2.0)
                elif has_ads and s["qp7"] == 0 and s["q7"] > 0:
                    boost_map[nm] = 1.3
                else:
                    boost_map[nm] = 1.0
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning("WB ad boost query failed: %s", e)

    # ── 2c. Demand by buyer region ───────────────────────────
    # Считаем заказы по nm_id × округ покупателя (а не склад отгрузки!)
    regional_demand: dict[int, dict[str, int]] = {}  # nm_id → {округ → qty}
    product_categories: dict[int, str] = {}  # nm_id → category
    try:
        rd_rows = ch.query(f"""
            SELECT nm_id, oblast_okrug_name, count() AS cnt,
                   any(category) AS cat
            FROM mms_analytics.fact_orders_raw
            WHERE shop_id = {shop_id}
              AND date >= '{d_sales_start}'
              AND date <= '{today}'
              AND is_cancel = 0
              AND oblast_okrug_name != ''
            GROUP BY nm_id, oblast_okrug_name
        """).result_rows
        for row in rd_rows:
            nm = int(row[0])
            if nm not in regional_demand:
                regional_demand[nm] = {}
            regional_demand[nm][row[1]] = row[2]
            if row[3] and nm not in product_categories:
                product_categories[nm] = row[3]
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("WB regional demand query failed: %s", e)

    # ── 2d. Cross-warehouse consumption (actual warehouse → buyer okrug) ──
    # Which warehouse shipped to which buyer region?
    # If warehouse's home okrug != buyer okrug → cross-warehouse drain
    wh_consumption: dict[int, dict[str, dict[str, int]]] = {}  # nm_id → {wh → {okrug → qty}}
    try:
        cross_rows = ch.query(f"""
            SELECT nm_id, warehouse_name, oblast_okrug_name,
                   count() AS qty
            FROM mms_analytics.fact_orders_raw
            WHERE shop_id = {shop_id}
              AND date >= '{d_sales_start}'
              AND date <= '{today}'
              AND is_cancel = 0
              AND warehouse_name != ''
              AND oblast_okrug_name != ''
            GROUP BY nm_id, warehouse_name, oblast_okrug_name
        """).result_rows
        for row in cross_rows:
            nm = int(row[0])
            wh_name, okrug, qty = row[1], row[2], row[3]
            wh_consumption.setdefault(nm, {}).setdefault(wh_name, {})[okrug] = qty
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("WB cross-warehouse query failed: %s", e)

    # ── 3. Product info from PostgreSQL ──────────────────────
    prod_result = await db.execute(text("""
        SELECT nm_id, vendor_code, name,
               COALESCE(length, 0), COALESCE(width, 0), COALESCE(height, 0),
               COALESCE(current_price, 0)
        FROM dim_products
        WHERE shop_id = :sid
    """), {"sid": shop_id})
    prod_rows = prod_result.fetchall()

    product_info: dict[int, dict] = {}
    for r in prod_rows:
        l, w, h = float(r[3]), float(r[4]), float(r[5])
        vol_liters = (l * w * h) / 1000.0
        if vol_liters > 10000:
            vol_liters = (l * w * h) / 1_000_000.0
        product_info[r[0]] = {
            "vendor_code": r[1] or "",
            "name": r[2] or "",
            "length": l,
            "width": w,
            "height": h,
            "vol_liters": max(vol_liters, 0.1),
            "price": float(r[6]),
        }

    # ── 3b. Image URLs from Redis ────────────────────────────
    import os
    import redis as redis_lib
    redis_url = os.getenv("REDIS_URL", "redis://redis:6379/0")
    try:
        r = redis_lib.Redis.from_url(redis_url, decode_responses=True)
        for nm_id in product_info:
            img = r.get(f"state:image_url:{shop_id}:{nm_id}")
            product_info[nm_id]["image_url"] = img or ""
        r.close()
    except Exception:
        for nm_id in product_info:
            product_info[nm_id]["image_url"] = ""

    # ── 4. WB Tariffs ────────────────────────────────────────
    tariffs: dict[str, dict] = {}
    try:
        tariff_rows = ch.query(f"""
            SELECT
                warehouse_name,
                argMax(storage_coef, updated_at) AS storage_coef,
                argMax(storage_base_liter, updated_at) AS storage_base_liter,
                argMax(storage_additional_liter, updated_at) AS storage_add_liter,
                argMax(delivery_coef, updated_at) AS delivery_coef,
                argMax(delivery_base_liter, updated_at) AS delivery_base_liter,
                argMax(delivery_additional_liter, updated_at) AS delivery_add_liter,
                argMax(coefficient, updated_at) AS acceptance_coef,
                argMax(allow_unload, updated_at) AS allow_unload
            FROM mms_analytics.fact_wb_acceptance_tariffs
            WHERE box_type_id = 2
              AND dt >= today()
            GROUP BY warehouse_name
        """).result_rows

        for r_t in tariff_rows:
            tariffs[r_t[0]] = {
                "storage_coef": _parse_ru_float(r_t[1]),
                "storage_base_liter": _parse_ru_float(r_t[2]),
                "storage_add_liter": _parse_ru_float(r_t[3]),
                "delivery_coef": _parse_ru_float(r_t[4]),
                "delivery_base_liter": _parse_ru_float(r_t[5]),
                "delivery_add_liter": _parse_ru_float(r_t[6]),
                "acceptance_coef": float(r_t[7]),
                "allow_unload": r_t[8],
            }
    except Exception:
        # Таблица wb_acceptance_tariffs может отсутствовать — работаем с дефолтами
        tariffs = {}

    # ── 4b. Actual paid storage from fact_wb_paid_storage ─────
    # Real per-SKU per-warehouse daily storage cost (avg last 14d)
    actual_storage_per_day: dict[int, dict[str, float]] = {}  # nm_id → {warehouse → cost_per_day}
    has_actual_storage = False
    try:
        ps_rows = ch.query(f"""
            SELECT nm_id, warehouse,
                   SUM(warehouse_price) AS cost_period,
                   count(DISTINCT dt) AS days_cnt
            FROM mms_analytics.fact_wb_paid_storage FINAL
            WHERE shop_id = {shop_id}
              AND dt >= today() - 14
              AND dt < today()
            GROUP BY nm_id, warehouse
            HAVING cost_period != 0
        """).result_rows
        if ps_rows:
            has_actual_storage = True
        for r in ps_rows:
            nm_id_ps = int(r[0])
            wh_ps = r[1]
            cost_period = float(r[2])
            days_cnt = max(int(r[3]), 1)
            cost_per_day = cost_period / days_cnt  # average daily cost per unit-like
            actual_storage_per_day.setdefault(nm_id_ps, {})[wh_ps] = cost_per_day
        import logging
        logging.getLogger(__name__).info(
            "WB supply: loaded actual storage for %d nm_ids from fact_wb_paid_storage",
            len(actual_storage_per_day))
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("WB supply: paid storage query failed: %s", e)

    # ── 5. Build recommendations per SKU × warehouse ─────────
    # Используем РЕГИОНАЛЬНЫЙ спрос: округ покупателя → ближайший склад
    all_nm_ids = set(
        list(stocks_by_nm.keys()) +
        list(sales_by_nm.keys()) +
        list(regional_demand.keys())
    )

    # Собираем все доступные склады (из тарифов + маппинга)
    available_warehouses = set(tariffs.keys())

    items = []
    for nm_id in sorted(all_nm_ids):
        pinfo = product_info.get(nm_id, {})
        vol = pinfo.get("vol_liters", 1.0)
        vendor = pinfo.get("vendor_code", str(nm_id))
        name = pinfo.get("name", "")
        image_url = pinfo.get("image_url", "")

        wh_stocks = stocks_by_nm.get(nm_id, {})
        wh_sales = sales_by_nm.get(nm_id, {})
        nm_regions = regional_demand.get(nm_id, {})

        # ── 5a. Определяем тип товара (по категории + складам) ──
        product_type = _classify_product_wh_type(nm_id, product_categories, wh_stocks, wh_sales)
        est_weight = _estimate_weight_kg(vol)

        if est_weight > _MAX_BOX_WEIGHT_KG and product_type == "normal":
            product_type = "sgt"

        # ── 5b. Распределяем региональный спрос по складам ──
        demand_by_wh: dict[str, dict] = {}
        for region, qty in nm_regions.items():
            target_whs = REGION_TO_WAREHOUSES.get(region, [])
            typed_whs = [w for w in target_whs if _filter_wh_for_type(w, product_type)]
            if not typed_whs:
                typed_whs = target_whs

            placed = False
            for twh in typed_whs:
                matched_wh = _match_warehouse(twh, available_warehouses)
                if matched_wh:
                    if matched_wh not in demand_by_wh:
                        demand_by_wh[matched_wh] = {"regional_orders": 0, "regions": []}
                    demand_by_wh[matched_wh]["regional_orders"] += qty
                    demand_by_wh[matched_wh]["regions"].append(region)
                    placed = True
                    break
            if not placed:
                fallback_wh = typed_whs[0] if typed_whs else "Котовск"
                if fallback_wh not in demand_by_wh:
                    demand_by_wh[fallback_wh] = {"regional_orders": 0, "regions": []}
                demand_by_wh[fallback_wh]["regional_orders"] += qty
                demand_by_wh[fallback_wh]["regions"].append(region)

        # Объединяем склады и фильтруем по типу товара
        raw_wh = set(
            list(wh_stocks.keys()) +
            list(wh_sales.keys()) +
            list(demand_by_wh.keys())
        )
        # Для food — оставляем только ': Питание' + склады с текущим стоком (показываем но need=0)
        # Для normal — исключаем ': Питание' и 'СГТ'
        all_wh = set()
        for wh in raw_wh:
            is_stock_or_sales = wh in wh_stocks or wh in wh_sales
            if is_stock_or_sales:
                # Всегда показываем склады с текущими стоками/продажами
                all_wh.add(wh)
            elif _filter_wh_for_type(wh, product_type):
                # Новые склады — только подходящего типа
                all_wh.add(wh)

        total_sold = sum(s.get("orders", 0) for s in wh_sales.values())
        total_stock = sum(s.get("qty", 0) for s in wh_stocks.values())
        daily_avg = total_sold / max(sales_period, 1)
        boost = boost_map.get(nm_id, 1.0)
        boosted_daily = daily_avg * boost

        turnover_days = total_stock / boosted_daily if boosted_daily > 0 else 999

        warehouses = []
        for wh in sorted(all_wh):
            stock = wh_stocks.get(wh, {}).get("qty", 0)
            orders = wh_sales.get(wh, {}).get("orders", 0)
            revenue = wh_sales.get(wh, {}).get("revenue", 0)
            wh_daily = orders / max(sales_period, 1)

            # For food/SGT variants: include sales from paired regular warehouse
            # because WB books actual sales under "Котовск" not "Котовск: Питание"
            paired_orders = 0
            if _FOOD_SUFFIX in wh:
                base_name = wh.replace(_FOOD_SUFFIX, "").strip()
                paired_orders = wh_sales.get(base_name, {}).get("orders", 0)
            elif _SGT_SUFFIX in wh:
                base_name = wh.replace(" СГТ", "").strip()
                paired_orders = wh_sales.get(base_name, {}).get("orders", 0)
            paired_daily = paired_orders / max(sales_period, 1)

            # Региональный спрос → daily для этого склада
            rd = demand_by_wh.get(wh, {})
            regional_orders = rd.get("regional_orders", 0)
            regional_daily = regional_orders / max(sales_period, 1)
            demand_regions = rd.get("regions", [])

            # Используем МАКСИМУМ из (фактический daily, региональный daily, парный daily)
            # чтобы не занизить если склад уже отгружает больше
            effective_daily = max(wh_daily, regional_daily, paired_daily)
            effective_daily_boosted = effective_daily * boost

            # For food/SGT warehouses: account for stock at paired regular warehouse
            # "Котовск: Питание" and "Котовск" are the SAME physical location
            paired_stock = 0
            if _FOOD_SUFFIX in wh or _SGT_SUFFIX in wh:
                base_name = wh.replace(_FOOD_SUFFIX, "").replace(" СГТ", "").strip()
                paired_stock = wh_stocks.get(base_name, {}).get("qty", 0)
            elif product_type in ("food", "sgt"):
                # Regular warehouse for a food product — check if food variant exists
                food_variant = f"{wh}{_FOOD_SUFFIX}"
                sgt_variant = f"{wh} СГТ"
                paired_stock = (
                    wh_stocks.get(food_variant, {}).get("qty", 0) +
                    wh_stocks.get(sgt_variant, {}).get("qty", 0)
                )

            combined_stock = stock + paired_stock

            # Food/SGT products: need=0 on REGULAR warehouses.
            # All supply should go to the paired :Питание / СГТ variant.
            # Regular warehouse rows stay visible (show stock/sales) but don't generate supply.
            is_wrong_type_wh = (
                (product_type == "food" and _FOOD_SUFFIX not in wh) or
                (product_type == "sgt" and _SGT_SUFFIX not in wh)
            )
            if is_wrong_type_wh:
                need = 0
            else:
                need = max(0, int(effective_daily_boosted * target_days * safety) - combined_stock)

            t = tariffs.get(wh, {})
            stor_base = t.get("storage_base_liter", 0)
            stor_add = t.get("storage_add_liter", 0)

            # Use REAL paid storage data if available, fallback to tariff
            real_cost = actual_storage_per_day.get(nm_id, {}).get(wh)
            if real_cost is not None and stock > 0:
                # real_cost = daily cost for this nm_id on this warehouse (summed over all units)
                # storage_per_day = cost per unit per day
                storage_per_day = real_cost / stock
                storage_source = "actual"
            elif has_actual_storage:
                # Shop has actual paid storage data but this SKU/warehouse combo is missing
                # → likely zero cost (not charged). Don't use tariff fallback.
                storage_per_day = 0
                storage_source = "actual"
            else:
                if vol <= 1:
                    storage_per_day = stor_base * vol
                else:
                    storage_per_day = stor_base + stor_add * (vol - 1)
                storage_source = "tariff"

            wh_turnover = stock / effective_daily if effective_daily > 0 else 999
            ac = t.get("acceptance_coef", 0)

            # Store paired data for Excel "Поставка по складам" (food/SGT fix)
            paired_revenue = 0.0
            if _FOOD_SUFFIX in wh:
                base_name = wh.replace(_FOOD_SUFFIX, "").strip()
                paired_revenue = wh_sales.get(base_name, {}).get("revenue", 0)
            elif _SGT_SUFFIX in wh:
                base_name = wh.replace(" СГТ", "").strip()
                paired_revenue = wh_sales.get(base_name, {}).get("revenue", 0)

            warehouses.append({
                "warehouse": wh,
                "stock": stock,
                "orders": orders,
                "paired_orders": paired_orders,
                "regional_orders": regional_orders,
                "demand_regions": list(set(demand_regions)),
                "daily_boosted": round(effective_daily_boosted, 2),
                "revenue": revenue,
                "paired_revenue": round(paired_revenue, 2),
                "daily": round(wh_daily, 2),
                "regional_daily": round(regional_daily, 2),
                "need": need,
                "storage_per_day": round(storage_per_day, 4),
                "storage_per_month": round(storage_per_day * 30, 2),
                "storage_coef": t.get("storage_coef", 0),
                "storage_source": storage_source,
                "acceptance_coef": ac,
                "acceptance": "Без коэфф." if ac <= 0 or ac == -1 else f"x{ac:.0f}",
                "turnover_days": round(wh_turnover, 1),
            })

        # ── 5c. Cross-warehouse drain analysis ──
        # Enrich each warehouse with effective_days based on actual consumption
        nm_cross = wh_consumption.get(nm_id, {})
        item_effective_days = 9999
        for wh_item in warehouses:
            wh_name = wh_item["warehouse"]
            wh_okrug_data = nm_cross.get(wh_name, {})
            if not wh_okrug_data:
                wh_item["effective_days"] = wh_item["turnover_days"]
                wh_item["cross_daily"] = 0.0
                wh_item["cross_okrugs"] = []
                continue

            home_okrug = WAREHOUSE_TO_OKRUG.get(wh_name, "")
            own_qty = 0
            cross_qty = 0
            cross_details = []  # [{okrug, qty, daily}]
            for dest_okrug, qty in wh_okrug_data.items():
                if dest_okrug == home_okrug:
                    own_qty += qty
                else:
                    cross_qty += qty
                    cross_details.append({
                        "okrug": dest_okrug,
                        "qty": qty,
                        "daily": round(qty / max(sales_period, 1), 2),
                    })

            total_actual = own_qty + cross_qty
            actual_daily = total_actual / max(sales_period, 1)
            own_daily = own_qty / max(sales_period, 1)
            cross_daily_total = cross_qty / max(sales_period, 1)

            stock = wh_item["stock"]
            eff_days = stock / actual_daily if actual_daily > 0 else 9999

            wh_item["effective_days"] = round(eff_days, 1)
            wh_item["cross_daily"] = round(cross_daily_total, 2)
            wh_item["cross_okrugs"] = sorted(cross_details, key=lambda x: x["qty"], reverse=True)

            if eff_days < item_effective_days:
                item_effective_days = eff_days

        # Status — хранение платное с 1-го дня, overstock = оборот > target_days
        if daily_avg == 0 and total_stock == 0:
            status = "ok"
        elif turnover_days > target_days and total_stock > 0:
            status = "overstock"
        elif turnover_days < 14:
            status = "critical"
        elif turnover_days < target_days:
            status = "attention"
        else:
            status = "ok"

        # Re-evaluate status using effective_days (cross-warehouse drain)
        if item_effective_days < 9999:
            if item_effective_days < 14 and status in ("ok", "attention", "overstock"):
                status = "critical"
            elif item_effective_days < target_days and status == "ok":
                status = "attention"

        # ── 5d. Cross-drain re-balance ──
        # If we recommend supply to a regional warehouse, reduce
        # central warehouse need by the cross-drain portion going
        # to that region. This prevents double-counting demand.
        #
        # For food/SGT: only subtract cross-drain if the target region
        # has a food-compatible warehouse. If not, the cross-drain
        # is unavoidable and must stay on the central warehouse.
        regional_wh_with_need = set()
        for wh_item in warehouses:
            if wh_item["need"] > 0:
                wh_okrug = WAREHOUSE_TO_OKRUG.get(wh_item["warehouse"], "")
                if wh_okrug:
                    regional_wh_with_need.add(wh_okrug)

        if regional_wh_with_need and nm_cross:
            for wh_item in warehouses:
                wh_name = wh_item["warehouse"]
                home_okrug = WAREHOUSE_TO_OKRUG.get(wh_name, "")
                if not home_okrug:
                    continue
                wh_okrug_data = nm_cross.get(wh_name, {})
                if not wh_okrug_data:
                    continue

                # Calculate how much cross-drain goes to regions we're supplying
                reducible_cross = 0
                for dest_okrug, qty in wh_okrug_data.items():
                    if dest_okrug == home_okrug:
                        continue  # own region, not cross
                    if dest_okrug not in regional_wh_with_need:
                        continue  # not supplying that region
                    # For food/SGT: check if target region has compatible warehouse
                    if product_type in ("food", "sgt"):
                        target_whs = REGION_TO_WAREHOUSES.get(dest_okrug, [])
                        has_compatible = any(
                            _filter_wh_for_type(tw, product_type)
                            for tw in target_whs
                        )
                        if not has_compatible:
                            continue  # no food/SGT warehouse → cross unavoidable
                    reducible_cross += qty

                if reducible_cross > 0 and wh_item["need"] > 0:
                    # Reduce effective_daily by cross portion
                    cross_daily_reduce = reducible_cross / max(sales_period, 1)
                    current_eff_daily = wh_item["daily_boosted"]
                    # New effective daily = max(own_demand, 0)
                    new_eff_daily = max(0, current_eff_daily - cross_daily_reduce * boost)
                    # Get combined stock for this warehouse
                    wh_stock = wh_item["stock"]
                    ps = 0
                    if _FOOD_SUFFIX in wh_name or _SGT_SUFFIX in wh_name:
                        bn = wh_name.replace(_FOOD_SUFFIX, "").replace(" СГТ", "").strip()
                        ps = wh_stocks.get(bn, {}).get("qty", 0)
                    elif product_type in ("food", "sgt"):
                        fv = f"{wh_name}{_FOOD_SUFFIX}"
                        sv = f"{wh_name} СГТ"
                        ps = wh_stocks.get(fv, {}).get("qty", 0) + wh_stocks.get(sv, {}).get("qty", 0)
                    cs = wh_stock + ps

                    is_wrong = (
                        (product_type == "food" and _FOOD_SUFFIX not in wh_name) or
                        (product_type == "sgt" and _SGT_SUFFIX not in wh_name)
                    )
                    if not is_wrong:
                        new_need = max(0, int(new_eff_daily * target_days * safety) - cs)
                        wh_item["need"] = new_need
                        wh_item["daily_boosted"] = round(new_eff_daily, 2)

        # ── Global cap: prevent overstocking ──
        # Sum of per-warehouse needs must not exceed the SKU-level target.
        # Without this cap, regional demand + paired sales + actual sales
        # can independently inflate each warehouse's need, resulting in
        # a total 2-3× higher than what's actually required.
        global_target = max(0, int(boosted_daily * target_days * safety) - total_stock)
        raw_total_need = sum(w["need"] for w in warehouses)

        if raw_total_need > global_target and raw_total_need > 0:
            # Proportionally scale down each warehouse's need
            scale = global_target / raw_total_need
            remainder = 0.0
            for w in warehouses:
                if w["need"] > 0:
                    exact = w["need"] * scale + remainder
                    w["need"] = int(exact)
                    remainder = exact - w["need"]

        total_need = sum(w["need"] for w in warehouses)
        storage_cost_month = sum(w["storage_per_month"] * w["stock"] for w in warehouses)

        items.append({
            "nm_id": nm_id,
            "vendor_code": vendor,
            "name": name,
            "image_url": image_url,
            "vol_liters": round(vol, 2),
            "total_sold": total_sold,
            "total_stock": total_stock,
            "daily_avg": round(daily_avg, 2),
            "boost": round(boost, 2),
            "boosted_daily": round(boosted_daily, 2),
            "turnover_days": round(turnover_days, 1),
            "effective_days": round(item_effective_days, 1) if item_effective_days < 9999 else None,
            "total_need": total_need,
            "status": status,
            "storage_cost_month": round(storage_cost_month, 2),
            "product_type": product_type,
            "warehouses": warehouses,
        })

    # ── 6. KPI ───────────────────────────────────────────────
    kpi = {
        "total_need": sum(it["total_need"] for it in items),
        "critical_count": sum(1 for it in items if it["status"] == "critical"),
        "attention_count": sum(1 for it in items if it["status"] == "attention"),
        "overstock_count": sum(1 for it in items if it["status"] == "overstock"),
        "avg_days_supply": round(
            sum(it["turnover_days"] for it in items if it["turnover_days"] < 999) /
            max(sum(1 for it in items if it["turnover_days"] < 999), 1),
            1
        ),
        "total_stock": sum(it["total_stock"] for it in items),
        "total_sku": len(items),
        "total_storage_month": round(sum(it["storage_cost_month"] for it in items), 2),
    }

    # ── 7. Warehouse summary ─────────────────────────────────
    wh_agg: dict[str, dict] = {}
    for it in items:
        for wh in it["warehouses"]:
            wn = wh["warehouse"]
            if wn not in wh_agg:
                wh_agg[wn] = {
                    "warehouse": wn,
                    "total_stock": 0, "total_orders": 0,
                    "total_need": 0, "total_revenue": 0,
                    "items_count": 0,
                    "storage_coef": wh.get("storage_coef", 0),
                    "acceptance": wh.get("acceptance", ""),
                }
            wh_agg[wn]["total_stock"] += wh["stock"]
            wh_agg[wn]["total_orders"] += wh["orders"]
            wh_agg[wn]["total_need"] += wh["need"]
            wh_agg[wn]["total_revenue"] += wh["revenue"]
            wh_agg[wn]["items_count"] += 1

    warehouse_summary = sorted(wh_agg.values(), key=lambda x: x["total_orders"], reverse=True)

    return {
        "items": items,
        "tariffs": tariffs,
        "kpi": kpi,
        "warehouse_summary": warehouse_summary,
    }


@router.get("/wb/supply")
async def get_wb_supply(
    shop_id: int = Query(..., description="Shop ID"),
    sales_period: int = Query(30, ge=7, le=90, description="Sales period for avg daily calc"),
    target_days: int = Query(45, ge=14, le=60, description="Target stock days (max 60 = free storage)"),
    safety: float = Query(1.15, ge=1.0, le=2.0, description="Safety coefficient"),
    use_ad_boost: bool = Query(True, description="Apply ad boost coefficient"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    WB Supply JSON — recommendations per SKU × warehouse.
    Returns kpi, items with nested warehouses, warehouse_summary.
    """
    shop = await db.get(Shop, shop_id)
    if not shop or shop.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Shop not found")
    if shop.marketplace != "wildberries":
        raise HTTPException(status_code=400, detail="Shop is not Wildberries")

    result = await _build_wb_supply_data(shop_id, sales_period, target_days, safety, use_ad_boost, db)

    return {
        "shop_id": shop_id,
        "sales_period": sales_period,
        "target_days": target_days,
        "safety": safety,
        "use_ad_boost": use_ad_boost,
        "kpi": result["kpi"],
        "items": result["items"],
        "warehouse_summary": result["warehouse_summary"],
    }


@router.get("/wb/supply/xlsx")
async def get_wb_supply_xlsx(
    shop_id: int = Query(..., description="Shop ID"),
    sales_period: int = Query(30, ge=7, le=90, description="Sales period for avg daily calc"),
    target_days: int = Query(45, ge=14, le=60, description="Target stock days (max 60 = free storage)"),
    safety: float = Query(1.15, ge=1.0, le=2.0, description="Safety coefficient"),
    use_ad_boost: bool = Query(True, description="Apply ad boost coefficient"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    WB Supply Excel: recommendations per SKU × warehouse with storage cost analysis.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    shop = await db.get(Shop, shop_id)
    if not shop or shop.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Shop not found")
    if shop.marketplace != "wildberries":
        raise HTTPException(status_code=400, detail="Shop is not Wildberries")

    result = await _build_wb_supply_data(shop_id, sales_period, target_days, safety, use_ad_boost, db)
    items = result["items"]
    tariffs = result["tariffs"]

    # ══════════════════════════════════════════════════════════
    # Генерация Excel
    # ══════════════════════════════════════════════════════════
    wb_xlsx = Workbook()

    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    num_fmt = "#,##0"
    money_fmt = "#,##0.00"
    warn_font = Font(bold=True, color="CC8800")
    critical_font = Font(bold=True, color="CC0000")
    ok_font = Font(color="00AA00")
    blue_font = Font(bold=True, color="0070C0")
    gray_font = Font(color="999999")
    sku_hdr_fill = PatternFill("solid", fgColor="D6E4F0")

    def _fmt_turnover(td):
        """Форматирование оборачиваемости: число → '25 дн' или 'Нет продаж'."""
        return f"{td:.0f} дн" if td < 999 else "Нет продаж"

    def _fmt_acceptance(ac):
        """Форматирование коэффициента приёмки: -1/0 → Без коэфф., >0 → x5."""
        if ac == -1 or ac == 0:
            return "Без коэфф."
        return f"x{ac:.0f}"

    def _acceptance_font(ac):
        if ac > 5:
            return critical_font
        elif ac > 0:
            return warn_font
        return ok_font

    # ══ Лист 1: Рекомендации по складам ══
    ws1 = wb_xlsx.active
    ws1.title = "Рекомендации по складам"
    h1 = [
        ("Артикул", 18), ("Название товара", 35), ("Склад WB", 28),
        ("Продано, шт", 12), ("Продаж в день", 10), ("Остаток, шт", 12),
        ("Оборачиваемость, дн", 14), ("Реал.зап, дн", 12), ("Кросс", 30),
        ("Нужно довезти, шт", 14), ("Выручка, руб", 14),
        ("Хранение, руб/день", 12), ("Хранение, руб/мес", 12),
        ("Коэфф. хранения", 12), ("Приёмка", 14),
    ]
    for ci, (n, w) in enumerate(h1, 1):
        c = ws1.cell(1, ci, n)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.freeze_panes = "A2"

    ws1.cell(2, 1, f"Параметры расчёта: период продаж {sales_period} дн, горизонт {target_days} дн, запас прочности ×{safety} | Ad boost: {'ДА' if use_ad_boost else 'НЕТ'}").font = Font(bold=True, size=10, color="1F4E79")
    ws1.cell(2, 2, f"WB: хранение платное с 1-го дня, коэфф. фиксируется на 60 дн").font = Font(italic=True, size=9, color="666666")
    for ci2 in range(1, len(h1) + 1):
        ws1.cell(2, ci2).fill = PatternFill("solid", fgColor="DAEEF3")
    row1 = 4

    for item in items:
        if not item["warehouses"]:
            continue

        ws1.cell(row1, 1, item["vendor_code"]).font = Font(bold=True, size=11)
        ws1.cell(row1, 2, item["name"]).font = Font(bold=True, size=10)
        td_val = item["turnover_days"]
        ws1.cell(row1, 7, _fmt_turnover(td_val)).font = (
            critical_font if td_val > 60 else warn_font if td_val > 45 else ok_font
        )
        # Item-level effective_days
        eff = item.get("effective_days")
        if eff is not None:
            ws1.cell(row1, 8, f"{eff:.0f} дн").font = (
                critical_font if eff < 14 else warn_font if eff < target_days else ok_font
            )
        else:
            ws1.cell(row1, 8, "—").font = gray_font
        for ci2 in range(1, len(h1) + 1):
            ws1.cell(row1, ci2).fill = sku_hdr_fill
        row1 += 1

        for wh in sorted(item["warehouses"], key=lambda x: x["orders"], reverse=True):
            ws1.cell(row1, 3, wh["warehouse"])
            ws1.cell(row1, 4, wh["orders"]).number_format = num_fmt
            ws1.cell(row1, 5, wh["daily"]).number_format = "0.00"
            ws1.cell(row1, 6, wh["stock"]).number_format = num_fmt
            ws1.cell(row1, 7, _fmt_turnover(wh["turnover_days"])).font = (
                critical_font if wh["turnover_days"] > 60
                else warn_font if wh["turnover_days"] > 45
                else ok_font
            )

            # Реал.зап — effective_days per warehouse
            wh_eff = wh.get("effective_days", wh["turnover_days"])
            if wh_eff < 9999:
                ws1.cell(row1, 8, f"{wh_eff:.0f} дн").font = (
                    critical_font if wh_eff < 14 else warn_font if wh_eff < target_days else ok_font
                )
            else:
                ws1.cell(row1, 8, "—").font = gray_font

            # Кросс — cross-warehouse drain details
            cross_okrugs = wh.get("cross_okrugs", [])
            cross_daily = wh.get("cross_daily", 0)
            if cross_okrugs and cross_daily > 0:
                # Shorten okrug names for readability
                parts = []
                for co in cross_okrugs[:3]:  # top-3
                    short_okrug = co["okrug"].replace(" федеральный округ", "").strip()
                    parts.append(f"{short_okrug} +{co['daily']}/д")
                cross_text = ", ".join(parts)
                ws1.cell(row1, 9, cross_text).font = warn_font
            else:
                ws1.cell(row1, 9, "—").font = gray_font

            ws1.cell(row1, 10, wh["need"]).number_format = num_fmt
            if wh["need"] > 0:
                ws1.cell(row1, 10).font = blue_font
            ws1.cell(row1, 11, wh["revenue"]).number_format = money_fmt
            ws1.cell(row1, 12, wh["storage_per_day"]).number_format = "0.0000"
            ws1.cell(row1, 13, wh["storage_per_month"]).number_format = money_fmt
            if wh["storage_per_month"] > 50:
                ws1.cell(row1, 13).font = warn_font

            sc = wh.get("storage_coef", 0)
            ac = wh.get("acceptance_coef", 0)
            ws1.cell(row1, 14, f"{sc:.0f}%" if sc else "нет данных")
            ws1.cell(row1, 15, _fmt_acceptance(ac)).font = _acceptance_font(ac)

            row1 += 1
        row1 += 1

    # ══ Лист 2: Сводка по товарам ══
    ws2 = wb_xlsx.create_sheet("Сводка по товарам")
    h2 = [
        ("Артикул", 18), ("Название товара", 35), ("Объём, литры", 10),
        ("Продано, шт", 12), ("Продаж в день", 10), ("Остаток, шт", 12),
        ("Оборачиваемость, дн", 14), ("Хранение, руб/мес", 14),
        ("Рекомендация", 55),
    ]
    for ci, (n, w) in enumerate(h2, 1):
        c = ws2.cell(1, ci, n)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"
    row2 = 2

    for item in items:
        if item["total_sold"] == 0 and item["total_stock"] == 0:
            continue

        total_storage_month = sum(wh["storage_per_month"] * wh["stock"]
                                  for wh in item["warehouses"])
        td = item["turnover_days"]

        if td > 90:
            rec = f"Критично! Более 90 дней оборота. Нужна распродажа или вывоз товара."
            rec_font = critical_font
        elif td > target_days:
            rec = f"Перезатарка: {td:.0f} дней оборота. Превышает горизонт поставки ({target_days} дн). Хранение платное с 1-го дня!"
            rec_font = warn_font
        elif td < 14:
            rec = f"Мало товара! Всего {td:.0f} дней запаса. Срочно нужна поставка!"
            rec_font = blue_font
        elif item["daily_avg"] > 0:
            rec = f"В норме: {td:.0f} дней запаса (цель — до {target_days} дней)"
            rec_font = ok_font
        else:
            rec = "Нет продаж за период анализа"
            rec_font = gray_font

        ws2.cell(row2, 1, item["vendor_code"])
        ws2.cell(row2, 2, item["name"])
        ws2.cell(row2, 3, item["vol_liters"]).number_format = "0.00"
        ws2.cell(row2, 4, item["total_sold"]).number_format = num_fmt
        ws2.cell(row2, 5, item["daily_avg"]).number_format = "0.00"
        ws2.cell(row2, 6, item["total_stock"]).number_format = num_fmt
        ws2.cell(row2, 7, _fmt_turnover(td)).font = (
            critical_font if td > 60 else warn_font if td > 45 else ok_font
        )
        ws2.cell(row2, 8, round(total_storage_month, 2)).number_format = money_fmt
        ws2.cell(row2, 9, rec).font = rec_font
        row2 += 1

    # ══ Лист 3: Поставка по складам ══
    # Группировка: склад → список SKU с need > 0
    ws_wh = wb_xlsx.create_sheet("Поставка по складам")
    h_wh = [
        ("Склад WB", 32), ("Артикул", 18), ("Название товара", 35),
        ("Заказы прямые", 12), ("Заказы региональные", 14),
        ("Ежедн. эфф.", 10), ("Остаток", 10), ("ПОСТАВИТЬ", 14),
        ("Выручка, руб", 14), ("Тип товара", 14),
    ]
    for ci, (n, w) in enumerate(h_wh, 1):
        c = ws_wh.cell(1, ci, n)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws_wh.column_dimensions[get_column_letter(ci)].width = w
    ws_wh.freeze_panes = "A2"

    # Собираем данные: склад → [{sku info}]
    wh_groups: dict[str, list[dict]] = {}
    for item in items:
        for wh in item.get("warehouses", []):
            if wh["need"] <= 0:
                continue
            wh_name = wh["warehouse"]
            if wh_name not in wh_groups:
                wh_groups[wh_name] = []
            # For food/SGT: include paired orders and revenue
            # WB books sales under base name ("Котовск"), not "Котовск: Питание"
            effective_orders = wh.get("orders", 0) + wh.get("paired_orders", 0)
            effective_revenue = wh.get("revenue", 0) + wh.get("paired_revenue", 0)
            wh_groups[wh_name].append({
                "vendor_code": item["vendor_code"],
                "name": item["name"],
                "orders": effective_orders,
                "regional_orders": wh.get("regional_orders", 0),
                "daily_boosted": wh.get("daily_boosted", 0),
                "stock": wh["stock"],
                "need": wh["need"],
                "revenue": effective_revenue,
                "product_type": item.get("product_type", ""),
            })

    wh_hdr_fill = PatternFill("solid", fgColor="BDD7EE")
    wh_hdr_font = Font(bold=True, size=11, color="1F4E79")
    need_fill = PatternFill("solid", fgColor="FFF2CC")
    regional_font = Font(bold=True, color="8B5CF6")
    row_wh = 2

    for wh_name in sorted(wh_groups.keys(), key=lambda w: sum(i["need"] for i in wh_groups[w]), reverse=True):
        sku_list = wh_groups[wh_name]
        total_need = sum(i["need"] for i in sku_list)
        total_revenue = sum(i["revenue"] for i in sku_list)
        total_regional = sum(i["regional_orders"] for i in sku_list)
        total_direct = sum(i["orders"] for i in sku_list)
        sku_count = len(sku_list)

        # Заголовок склада
        wh_label = f"📦 {wh_name}  —  {sku_count} SKU"
        c = ws_wh.cell(row_wh, 1, wh_label)
        c.font = wh_hdr_font
        c.fill = wh_hdr_fill
        ws_wh.cell(row_wh, 4, total_direct).number_format = num_fmt
        ws_wh.cell(row_wh, 5, total_regional).number_format = num_fmt
        if total_regional > 0:
            ws_wh.cell(row_wh, 5).font = regional_font
        ws_wh.cell(row_wh, 8, total_need).font = Font(bold=True, size=12, color="CC0000")
        ws_wh.cell(row_wh, 8).number_format = num_fmt
        ws_wh.cell(row_wh, 9, round(total_revenue)).number_format = num_fmt
        for ci2 in range(1, len(h_wh) + 1):
            ws_wh.cell(row_wh, ci2).fill = wh_hdr_fill
        row_wh += 1

        # SKU внутри склада
        for si in sorted(sku_list, key=lambda x: x["need"], reverse=True):
            ws_wh.cell(row_wh, 1, "")
            ws_wh.cell(row_wh, 2, si["vendor_code"])
            ws_wh.cell(row_wh, 3, si["name"])
            ws_wh.cell(row_wh, 4, si["orders"]).number_format = num_fmt
            c5 = ws_wh.cell(row_wh, 5, si["regional_orders"])
            c5.number_format = num_fmt
            if si["regional_orders"] > 0:
                c5.font = regional_font
            ws_wh.cell(row_wh, 6, si["daily_boosted"]).number_format = "0.00"
            ws_wh.cell(row_wh, 7, si["stock"]).number_format = num_fmt
            c = ws_wh.cell(row_wh, 8, si["need"])
            c.number_format = num_fmt
            c.font = Font(bold=True, color="CC0000")
            c.fill = need_fill
            ws_wh.cell(row_wh, 9, round(si["revenue"])).number_format = num_fmt
            pt = si.get("product_type", "")
            pt_label = {"food": "🍖 Питание", "sgt": "📦 СГТ"}.get(pt, "Обычный")
            ws_wh.cell(row_wh, 10, pt_label)
            row_wh += 1

        row_wh += 1  # пустая строка между складами

    if not wh_groups:
        ws_wh.cell(row_wh, 1, "Нет товаров к поставке")
        ws_wh.cell(row_wh, 1).font = Font(bold=True, size=12, color="00AA00")

    # ══ Лист 4: Тарифы складов WB ══
    ws3 = wb_xlsx.create_sheet("Тарифы складов WB")
    h3 = [
        ("Склад", 30), ("Коэфф. хранения", 14),
        ("Хранение: базовый, руб/литр", 14), ("Хранение: доп литры, руб/литр", 14),
        ("Коэфф. логистики", 14),
        ("Логистика: базовый, руб/литр", 14), ("Логистика: доп литры, руб/литр", 14),
        ("Приёмка", 14), ("Можно отгружать", 12),
    ]
    for ci, (n, w) in enumerate(h3, 1):
        c = ws3.cell(1, ci, n)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.freeze_panes = "A2"
    row3 = 2

    for wh_name in sorted(tariffs.keys()):
        t = tariffs[wh_name]
        if "Маркетплейс" in wh_name:
            continue

        ws3.cell(row3, 1, wh_name)
        ws3.cell(row3, 2, f"{t['storage_coef']:.0f}%" if t['storage_coef'] else "нет данных")
        ws3.cell(row3, 3, t['storage_base_liter']).number_format = "0.00"
        ws3.cell(row3, 4, t['storage_add_liter']).number_format = "0.00"
        ws3.cell(row3, 5, f"{t['delivery_coef']:.0f}%" if t['delivery_coef'] else "нет данных")
        ws3.cell(row3, 6, t['delivery_base_liter']).number_format = "0.00"
        ws3.cell(row3, 7, t['delivery_add_liter']).number_format = "0.00"

        ac = t['acceptance_coef']
        ws3.cell(row3, 8, _fmt_acceptance(ac)).font = _acceptance_font(ac)
        ws3.cell(row3, 9, "Да" if t['allow_unload'] else "Нет").font = ok_font if t['allow_unload'] else critical_font
        row3 += 1

    # ══ Лист 4: Риск перезатаривания ══
    ws4 = wb_xlsx.create_sheet("Риск перезатаривания")
    h4 = [
        ("Артикул", 18), ("Название товара", 35), ("Объём, литры", 10),
        ("Остаток, шт", 12), ("Продаж в день", 10), ("Оборачиваемость, дн", 14),
        ("Превышение горизонта, дн", 14), ("Доп расходы хранение, руб/мес", 16),
        ("Рекомендация", 55),
    ]
    for ci, (n, w) in enumerate(h4, 1):
        c = ws4.cell(1, ci, n)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.freeze_panes = "A2"
    row4 = 2

    risk_items = [it for it in items if it["turnover_days"] > target_days and it["total_stock"] > 0]
    risk_items.sort(key=lambda x: x["turnover_days"], reverse=True)

    for item in risk_items:
        excess_days = max(0, item["turnover_days"] - target_days)
        # Ежемесячная стоимость хранения текущего stока
        storage_per_month = sum(
            wh["storage_per_day"] * wh["stock"] * 30 for wh in item["warehouses"] if wh["stock"] > 0
        )
        # Количество избыточных единиц
        excess_qty = 0
        if item["daily_avg"] > 0:
            target_stock = int(item["daily_avg"] * target_days * safety)
            excess_qty = max(0, item["total_stock"] - target_stock)

        if item["turnover_days"] > 90:
            rec = f"Критично! {item['total_stock']} шт (излишек ~{excess_qty}) лежат более 90 дней. Рекомендуем распродажу или возврат товара!"
        elif item["turnover_days"] > target_days:
            rec = f"Перезатарка: {excess_days:.0f} дней сверх горизонта ({target_days} дн), излишек ~{excess_qty} шт. Хранение платное. Ускорьте продажи или снизьте запас."
        else:
            rec = f"Внимание: оборачиваемость {item['turnover_days']:.0f} дн. Контролируйте запасы."

        ws4.cell(row4, 1, item["vendor_code"])
        ws4.cell(row4, 2, item["name"])
        ws4.cell(row4, 3, item["vol_liters"]).number_format = "0.00"
        ws4.cell(row4, 4, item["total_stock"]).number_format = num_fmt
        ws4.cell(row4, 5, item["daily_avg"]).number_format = "0.00"
        ws4.cell(row4, 6, _fmt_turnover(item["turnover_days"])).font = (
            critical_font if item["turnover_days"] > 60 else warn_font
        )
        ws4.cell(row4, 7, f"{excess_days:.0f} дн / ~{excess_qty} шт").number_format = num_fmt
        ws4.cell(row4, 8, round(storage_per_month, 2)).number_format = money_fmt
        if storage_per_month > 0:
            ws4.cell(row4, 8).font = critical_font
        ws4.cell(row4, 9, rec).font = Font(size=10)
        row4 += 1

    if not risk_items:
        ws4.cell(row4, 1, "Нет товаров с риском перезатаривания")
        ws4.cell(row4, 1).font = Font(bold=True, size=12, color="00AA00")

    # ── Сохранение и отправка ────────────────────────────────
    buf = io.BytesIO()
    wb_xlsx.save(buf)
    buf.seek(0)

    fname = f"wb_supply_{shop_id}_{target_days}d.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ══════════════════════════════════════════════════════════════
# Ozon: Warehouse Analytics — аналитика по складам
# ══════════════════════════════════════════════════════════════

# Маппинг warehouse_name → cluster (для расчёта скорости доставки)
_WAREHOUSE_TO_CLUSTER: dict[str, str] = {
    # Московский регион
    "ДОМОДЕДОВО_РФЦ": "Москва, МО и Дальние регионы",
    "ЖУКОВСКИЙ_РФЦ": "Москва, МО и Дальние регионы",
    "ГРИВНО_РФЦ": "Москва, МО и Дальние регионы",
    "СОФЬИНО_РФЦ": "Москва, МО и Дальние регионы",
    "НОГИНСК_РФЦ": "Москва, МО и Дальние регионы",
    "ПЕТРОВСКОЕ_РФЦ": "Москва, МО и Дальние регионы",
    "Дедовск": "Москва, МО и Дальние регионы",
    # СПБ
    "СПБ_БУГРЫ_РФЦ": "Санкт-Петербург и СЗО",
    "СПБ_КОЛПИНО_РФЦ": "Санкт-Петербург и СЗО",
    "Санкт_Петербург_РФЦ": "Санкт-Петербург и СЗО",
    "САНКТ-ПЕТЕРБУРГ_РФЦ": "Санкт-Петербург и СЗО",
    # Юг
    "РОСТОВ_НА_ДОНУ_2_РФЦ": "Ростов",
    "АДЫГЕЙСК_РФЦ": "Краснодар",
    "НЕВИННОМЫССК_РФЦ": "Невинномысск",
    # Центр
    "ВОРОНЕЖ_2_РФЦ": "Воронеж",
    "САРАТОВ_РФЦ": "Саратов",
    # Поволжье
    "Казань_РФЦ_НОВЫЙ": "Казань",
    "КАЗАНЬ_РФЦ_НОВЫЙ": "Казань",
    "САМАРА_РФЦ": "Самара",
    "УФА_РФЦ": "Уфа",
    # Урал / Сибирь
    "Екатеринбург_РФЦ_НОВЫЙ": "Екатеринбург",
    "КРАСНОЯРСК_МРФЦ": "Красноярск",
    "НИЖНИЙ_НОВГОРОД_РФЦ": "Казань",
    "НОВОСИБИРСК_РФЦ": "Новосибирск",
}


def _get_cluster_for_warehouse(wh_name: str) -> str:
    """Resolve warehouse name to delivery cluster."""
    if not wh_name:
        return "Неизвестный"
    # Exact match
    cluster = _WAREHOUSE_TO_CLUSTER.get(wh_name)
    if cluster:
        return cluster
    # Fuzzy fallback by keywords
    wh_upper = wh_name.upper()
    keyword_map = {
        "ДОМОДЕДОВО": "Москва, МО и Дальние регионы",
        "ЖУКОВСКИЙ": "Москва, МО и Дальние регионы",
        "ГРИВНО": "Москва, МО и Дальние регионы",
        "СОФЬИНО": "Москва, МО и Дальние регионы",
        "НОГИНСК": "Москва, МО и Дальние регионы",
        "ПЕТРОВСКОЕ": "Москва, МО и Дальние регионы",
        "ДЕДОВСК": "Москва, МО и Дальние регионы",
        "ТВЕРЬ": "Тверь",
        "СПБ": "Санкт-Петербург и СЗО",
        "ПЕТЕРБУРГ": "Санкт-Петербург и СЗО",
        "РОСТОВ": "Ростов",
        "АДЫГЕЙСК": "Краснодар",
        "КРАСНОДАР": "Краснодар",
        "НЕВИННОМЫССК": "Невинномысск",
        "ВОРОНЕЖ": "Воронеж",
        "САРАТОВ": "Саратов",
        "КАЗАНЬ": "Казань",
        "САМАРА": "Самара",
        "УФА": "Уфа",
        "ЕКАТЕРИНБУРГ": "Екатеринбург",
        "КРАСНОЯРСК": "Красноярск",
        "НОВОСИБИРСК": "Новосибирск",
        "НИЖНИЙ": "Казань",
        "ПЕРМЬ": "Пермь",
        "ТЮМЕНЬ": "Тюмень",
        "ОМСК": "Омск",
    }
    for kw, cl in keyword_map.items():
        if kw in wh_upper:
            return cl
    return "Неизвестный"


# Тарифы хранения Ozon (руб/литр/день)
OZON_STORAGE_TARIFFS = {
    "free": {"max_days": 160, "rate": 0.0},
    "low": {"max_days": 180, "rate": 0.75},
    "high": {"max_days": 99999, "rate": 1.50},
}


# ══════════════════════════════════════════════════════════════
# GET /warehouses/ozon/storage  — Ozon storage analytics
# ══════════════════════════════════════════════════════════════

@router.get("/ozon/storage")
async def ozon_storage_analytics(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Storage analytics for Ozon FBO — per-SKU turnover, storage cost estimation,
    zone classification (free / warning / paid).
    Format compatible with WB StorageSkusTable for component reuse.
    """
    import os
    import clickhouse_connect

    shop = await db.get(Shop, shop_id)
    if not shop or shop.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Shop not found")
    if shop.marketplace != "ozon":
        raise HTTPException(status_code=400, detail="Only Ozon shops supported")

    ch_host = os.getenv("CLICKHOUSE_HOST", "clickhouse")
    ch_port = int(os.getenv("CLICKHOUSE_PORT", 8123))
    ch_user = os.getenv("CLICKHOUSE_USER", "default")
    ch_pass = os.getenv("CLICKHOUSE_PASSWORD", "")
    ch = clickhouse_connect.get_client(
        host=ch_host, port=ch_port, username=ch_user, password=ch_pass,
        database="mms_analytics",
    )

    try:
        # ── 1. Current FBO stocks per SKU × warehouse ──
        stocks_query = ch.query("""
            SELECT sku, offer_id, product_name, warehouse_name,
                   free_to_sell, reserved
            FROM fact_ozon_warehouse_stocks FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND dt = (SELECT max(dt) FROM fact_ozon_warehouse_stocks WHERE shop_id = {shop_id:UInt32})
              AND warehouse_type = 'fbo'
              AND free_to_sell > 0
        """, parameters={"shop_id": shop_id})

        sku_agg: dict[int, dict] = {}
        for row in stocks_query.result_rows:
            sku_id = int(row[0])
            if sku_id not in sku_agg:
                sku_agg[sku_id] = {
                    "sku": sku_id,
                    "offer_id": row[1],
                    "name": row[2],
                    "total_stock": 0,
                    "total_reserved": 0,
                    "warehouses": [],
                }
            sku_agg[sku_id]["total_stock"] += int(row[4])
            sku_agg[sku_id]["total_reserved"] += int(row[5])
            sku_agg[sku_id]["warehouses"].append({
                "warehouse_name": row[3],
                "stock": int(row[4]),
                "reserved": int(row[5]),
            })

        if not sku_agg:
            return {
                "kpi": {
                    "total_skus": 0, "total_stock": 0,
                    "total_storage": 0, "avg_turnover_days": None,
                    "paid_zone_skus": 0, "warning_zone_skus": 0,
                    "period_days": period,
                },
                "storage_skus": [],
            }

        # ── 2. Orders per SKU (last N days) ──
        sku_ids = list(sku_agg.keys())
        orders_query = ch.query("""
            SELECT sku, sum(quantity) as sold, sum(price * quantity) as revenue
            FROM fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= today() - {period:UInt32}
              AND status != 'cancelled'
              AND sku IN {sku_ids:Array(UInt64)}
            GROUP BY sku
        """, parameters={"shop_id": shop_id, "period": period, "sku_ids": sku_ids})

        sales_map: dict[int, dict] = {}
        for row in orders_query.result_rows:
            sales_map[int(row[0])] = {"sold": int(row[1]), "revenue": float(row[2])}

        # ── 3. Turnover data from fact_ozon_turnover (Ozon API) ──
        turnover_map: dict[int, dict] = {}
        try:
            tr = ch.query("""
                SELECT sku, days_of_supply, avg_daily_sales,
                       stock_fbo, turnover_category
                FROM fact_ozon_turnover FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND dt = (SELECT max(dt) FROM fact_ozon_turnover WHERE shop_id = {shop_id:UInt32})
            """, parameters={"shop_id": shop_id})
            for row in tr.result_rows:
                turnover_map[int(row[0])] = {
                    "days_of_supply": float(row[1]),
                    "avg_daily_sales": float(row[2]),
                    "stock_fbo": int(row[3]),
                    "turnover_category": row[4],
                }
        except Exception:
            pass

        # ── 4. Ad data for SKUs ──
        ad_map: dict[int, dict] = {}
        try:
            ad_query = ch.query("""
                SELECT sku,
                       sum(money_spent) AS spend_30d,
                       sum(orders) AS orders_30d,
                       sumIf(money_spent, dt >= today() - 7) AS spend_7d
                FROM mms_analytics.fact_ozon_ad_daily FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND dt >= today() - 30
                  AND sku IN {sku_ids:Array(UInt64)}
                GROUP BY sku
            """, parameters={"shop_id": shop_id, "sku_ids": sku_ids})
            for r in ad_query.result_rows:
                ad_map[int(r[0])] = {
                    "has_ads": float(r[3]) > 0,
                    "spend_30d": float(r[1]),
                    "orders_30d": int(r[2]),
                }
        except Exception:
            pass

        # ── 5. Product volume from volume_weight ──
        # Ozon dim_ozon_products: depth/height/width often = 0
        # But volume_weight (kg) × 2.87 ≈ volume in liters
        # (coefficient reverse-engineered from Ozon seller dashboard)
        VW_TO_LITERS = 2.87
        dims_map: dict[int, float] = {}  # sku → volume in liters
        try:
            await db.rollback()
        except Exception:
            pass
        try:
            vol_rows = await db.execute(
                text("""SELECT sku, volume_weight, depth, height, width
                        FROM dim_ozon_products
                        WHERE shop_id = :sid AND sku IS NOT NULL"""),
                {"sid": shop_id},
            )
            for vr in vol_rows.fetchall():
                if not vr[0]:
                    continue
                d, h, w = float(vr[2] or 0), float(vr[3] or 0), float(vr[4] or 0)
                vw = float(vr[1] or 0)
                # Priority: actual dimensions > volume_weight estimate
                if d > 0 and h > 0 and w > 0:
                    volume_liters = (d * h * w) / 1_000_000
                elif vw > 0:
                    volume_liters = vw * VW_TO_LITERS
                else:
                    volume_liters = 0.5  # fallback
                dims_map[int(vr[0])] = max(round(volume_liters, 2), 0.1)
        except Exception:
            try:
                await db.rollback()
            except Exception:
                pass

        # ── 5b. ACTUAL placement costs from Ozon report ──
        # Data is stored per-day × per-warehouse, aggregate SUM by offer_id
        actual_costs: dict[str, float] = {}  # offer_id → total placement_cost
        # Per-warehouse cost: offer_id → {warehouse_name → cost}
        actual_wh_costs: dict[str, dict[str, float]] = {}
        actual_period = None
        try:
            pc = ch.query("""
                SELECT offer_id,
                       sum(placement_cost) AS total_cost,
                       min(period_from) AS p_from,
                       max(period_to) AS p_to
                FROM fact_ozon_placement_cost FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND period_to = (
                      SELECT max(period_to)
                      FROM fact_ozon_placement_cost FINAL
                      WHERE shop_id = {shop_id:UInt32}
                  )
                GROUP BY offer_id
            """, parameters={"shop_id": shop_id})
            for row in pc.result_rows:
                actual_costs[str(row[0])] = float(row[1])
                if actual_period is None:
                    actual_period = {"from": str(row[2]), "to": str(row[3])}

            # Per-warehouse breakdown for expanded detail view
            pwc = ch.query("""
                SELECT offer_id, warehouse_name,
                       sum(placement_cost) AS wh_cost
                FROM fact_ozon_placement_cost FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND period_to = (
                      SELECT max(period_to)
                      FROM fact_ozon_placement_cost FINAL
                      WHERE shop_id = {shop_id:UInt32}
                  )
                GROUP BY offer_id, warehouse_name
            """, parameters={"shop_id": shop_id})
            for row in pwc.result_rows:
                oid = str(row[0])
                wh_name = str(row[1])
                if oid not in actual_wh_costs:
                    actual_wh_costs[oid] = {}
                actual_wh_costs[oid][wh_name] = float(row[2])
        except Exception:
            logger.exception("Error fetching fact_ozon_placement_cost for shop %s", shop_id)

        has_actual_data = len(actual_costs) > 0

        # ── 5c. FALLBACK: total storage from fact_ozon_transactions ──
        # If fact_ozon_placement_cost is empty (e.g. migration dropped data,
        # sync not yet completed, or Ozon API 429), get total storage cost
        # from fact_ozon_transactions which is always populated by sync_ozon_finance.
        fallback_total_storage = 0.0
        if not has_actual_data:
            try:
                txn_storage = ch.query("""
                    SELECT sum(abs(amount)) AS total_storage,
                           min(toDate(operation_date)) AS min_dt,
                           max(toDate(operation_date)) AS max_dt
                    FROM mms_analytics.fact_ozon_transactions FINAL
                    WHERE shop_id = {shop_id:UInt32}
                      AND category = 'Storage'
                      AND toDate(operation_date) >= today() - {period:UInt32}
                """, parameters={"shop_id": shop_id, "period": period})
                for row in txn_storage.result_rows:
                    fallback_val = float(row[0] or 0)
                    if fallback_val > 0:
                        fallback_total_storage = fallback_val
                        has_actual_data = True
                        actual_period = {"from": str(row[1]), "to": str(row[2])}
                        logger.info(
                            "Ozon storage fallback from transactions: %.2f for shop %s (period %s→%s)",
                            fallback_val, shop_id, row[1], row[2],
                        )
            except Exception:
                logger.exception("Error fetching storage fallback from fact_ozon_transactions for shop %s", shop_id)

        # ── 6. Build per-SKU storage data ──
        # Ozon tariff: ~0.14 ₽/L/day for turnover > 160 days
        # (calibrated from Ozon seller dashboard reverse-engineering)
        TARIFF_PER_LITER_DAY = 0.14
        storage_skus = []
        total_est_cost = 0.0
        total_actual_cost = 0.0

        for sku_id, info in sku_agg.items():
            sales = sales_map.get(sku_id, {"sold": 0, "revenue": 0})
            daily_sales = sales["sold"] / period if period > 0 else 0

            # Prefer Ozon turnover data, fallback to calculated
            tr_data = turnover_map.get(sku_id)
            if tr_data and tr_data["days_of_supply"] > 0:
                turnover_days = tr_data["days_of_supply"]
            else:
                turnover_days = info["total_stock"] / daily_sales if daily_sales > 0 else 99999

            # Volume and cost estimation
            vol_liters = dims_map.get(sku_id, 0.5)
            est_daily_cost = info["total_stock"] * vol_liters * TARIFF_PER_LITER_DAY
            est_monthly_cost = est_daily_cost * 30

            # Zone classification (Ozon: free < 120d, warning 120-160d, paid > 160d)
            if turnover_days > 160:
                zone = "paid"
            elif turnover_days > 120:
                zone = "warning"
            else:
                zone = "free"

            # Only charge storage for items in paid zone
            if zone == "free":
                est_monthly_cost = 0
                est_daily_cost = 0

            total_est_cost += est_monthly_cost

            # Days to sell out
            days_to_sell = round(info["total_stock"] / daily_sales) if daily_sales > 0 else None

            # Forecast 30d: stock decreases with sales
            forecast_30d = None
            if est_daily_cost > 0 and info["total_stock"] > 0:
                cost_per_unit = est_daily_cost / info["total_stock"]
                forecast = 0.0
                for day in range(30):
                    remaining = max(0, info["total_stock"] - daily_sales * day)
                    if remaining <= 0:
                        break
                    forecast += cost_per_unit * remaining
                forecast_30d = round(forecast, 2)

            # Ad info
            ad_info = ad_map.get(sku_id, {"has_ads": False, "spend_30d": 0, "orders_30d": 0})

            # Check for actual placement cost from Ozon report
            offer_id = info["offer_id"]
            actual_cost = actual_costs.get(offer_id)
            if actual_cost is not None:
                storage_source = "actual"
                display_cost = actual_cost
                total_actual_cost += actual_cost
            else:
                storage_source = "estimated"
                display_cost = est_monthly_cost

            # Build per-warehouse breakdown with cost_month + forecast
            wh_cost_map = actual_wh_costs.get(offer_id, {})
            wh_list = []
            for wh_info in info["warehouses"]:
                wh_name = wh_info["warehouse_name"]
                wh_stock = wh_info["stock"]
                # Try actual warehouse cost first; fallback to proportional estimate
                wh_actual_cost = wh_cost_map.get(wh_name)
                if wh_actual_cost is not None:
                    wh_cost_month = round(wh_actual_cost, 2)
                elif est_daily_cost > 0 and info["total_stock"] > 0:
                    wh_cost_month = round(est_daily_cost / info["total_stock"] * wh_stock * 30, 2)
                else:
                    wh_cost_month = 0.0

                # Per-warehouse forecast: proportional to stock share
                wh_forecast = None
                if forecast_30d is not None and info["total_stock"] > 0 and wh_stock > 0:
                    stock_share = wh_stock / info["total_stock"]
                    wh_forecast = round(forecast_30d * stock_share, 2)

                wh_list.append({
                    "warehouse": wh_name,
                    "stock": wh_stock,
                    "reserved": wh_info.get("reserved", 0),
                    "cost_month": wh_cost_month,
                    "forecast_30d": wh_forecast,
                })

            storage_skus.append({
                # WB-compatible fields for StorageSkusTable
                "nm_id": sku_id,
                "vendor_code": info["offer_id"],
                "name": info["name"],
                "vol_liters": vol_liters,
                "total_stock": info["total_stock"],
                "est_cost_month": round(display_cost, 2),
                "storage_source": storage_source,
                "daily_sales": round(daily_sales, 2),
                "daily_cost": round(est_daily_cost, 2) if est_daily_cost > 0 else None,
                "days_to_sell": days_to_sell,
                "forecast_30d": forecast_30d,
                "has_active_ads": ad_info["has_ads"],
                # Ozon-specific
                "offer_id": info["offer_id"],
                "turnover_days": round(turnover_days, 1) if turnover_days < 99999 else None,
                "zone": zone,
                "turnover_category": tr_data["turnover_category"] if tr_data else "",
                "warehouses": sorted(wh_list, key=lambda w: w["cost_month"], reverse=True),
            })

        # Sort: paid first, then warning, then by est_cost desc
        zone_order = {"paid": 0, "warning": 1, "free": 2}
        storage_skus.sort(key=lambda s: (zone_order.get(s["zone"], 2), -s["est_cost_month"]))

        # ── 7. KPI ──
        total_stock = sum(s["total_stock"] for s in storage_skus)
        total_daily = sum(s["daily_sales"] for s in storage_skus)
        avg_turnover = total_stock / total_daily if total_daily > 0 else None
        paid_count = sum(1 for s in storage_skus if s["zone"] == "paid")
        warning_count = sum(1 for s in storage_skus if s["zone"] == "warning")
        total_forecast = sum(s["forecast_30d"] or 0 for s in storage_skus)

        return {
            "kpi": {
                "total_skus": len(storage_skus),
                "total_stock": total_stock,
                "total_storage": round(
                    total_actual_cost if total_actual_cost > 0
                    else fallback_total_storage if fallback_total_storage > 0
                    else total_est_cost, 2
                ),
                "avg_turnover_days": round(avg_turnover, 1) if avg_turnover else None,
                "paid_zone_skus": paid_count,
                "warning_zone_skus": warning_count,
                "period_days": period,
                "forecast_30d": round(total_forecast, 2) if total_forecast > 0 else None,
                "has_actual_data": has_actual_data,
                "actual_period": actual_period,
            },
            "storage_skus": storage_skus,
        }

    finally:
        ch.close()


@router.post("/ozon/sync-placement-cost")
async def trigger_ozon_placement_sync(
    shop_id: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Trigger Ozon placement cost report sync.

    Creates a Celery task that:
    1. Requests placement cost Excel report from Ozon API
    2. Polls until ready, downloads, parses
    3. Inserts per-SKU costs into ClickHouse

    Limits: 5 reports per day per seller, max 31-day period.
    """
    from celery_app.tasks.tasks import sync_ozon_placement_cost
    from app.core.encryption import decrypt_api_key

    shop = await db.get(Shop, shop_id)
    if not shop or shop.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Shop not found")
    if shop.marketplace != "ozon":
        raise HTTPException(status_code=400, detail="Only Ozon shops supported")

    api_key = decrypt_api_key(shop.api_key_encrypted)
    client_id = shop.client_id

    if not api_key or not client_id:
        raise HTTPException(status_code=400, detail="Shop API key or Client ID not configured")

    task = sync_ozon_placement_cost.apply_async(
        kwargs={
            "shop_id": shop_id,
            "api_key": api_key,
            "client_id": client_id,
        },
        queue="heavy",
    )

    return {
        "task_id": task.id,
        "status": "queued",
        "message": "Placement cost sync started. Report generation may take 30-60 seconds.",
    }


@router.post("/ozon/backfill-placement-cost")
async def trigger_ozon_placement_backfill(
    shop_id: int = Query(...),
    months: int = Query(3, ge=1, le=6),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Trigger Ozon placement cost backfill for last N months.

    Creates a Celery task that:
    1. Splits the period into 30-day chunks (Ozon API limit: 31 days/report)
    2. For each chunk: creates report → polls → downloads → parses → inserts
    3. 3 months = 3 chunks = 3 reports (within 5 reports/day limit)

    Can take up to 10-15 minutes depending on Ozon API response time.
    """
    from celery_app.tasks.tasks import backfill_ozon_placement_cost
    from app.core.encryption import decrypt_api_key

    shop = await db.get(Shop, shop_id)
    if not shop or shop.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Shop not found")
    if shop.marketplace != "ozon":
        raise HTTPException(status_code=400, detail="Only Ozon shops supported")

    api_key = decrypt_api_key(shop.api_key_encrypted)
    client_id = shop.client_id

    if not api_key or not client_id:
        raise HTTPException(status_code=400, detail="Shop API key or Client ID not configured")

    task = backfill_ozon_placement_cost.apply_async(
        kwargs={
            "shop_id": shop_id,
            "api_key": api_key,
            "client_id": client_id,
            "months": months,
        },
        queue="heavy",
    )

    return {
        "task_id": task.id,
        "status": "queued",
        "months": months,
        "message": f"Backfill placement cost started for {months} months. May take 10-15 minutes.",
    }


# ══════════════════════════════════════════════════════════════
# Ozon Warehouse Overview (lightweight dashboard)
# ══════════════════════════════════════════════════════════════

@router.get("/ozon/overview")
async def ozon_warehouse_overview(
    shop_id: int = Query(...),
    period: int = Query(30, description="Period in days"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Overview dashboard for Ozon warehouses — similar to WB analytics overview.

    Returns:
    - KPI summary with trends (prev period comparison)
    - Costs breakdown by operation type
    - Per-warehouse breakdown with SKU details
    - Out-of-stock alerts
    """
    from app.core.clickhouse import get_clickhouse_client

    shop = await db.get(Shop, shop_id)
    if not shop or shop.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Shop not found")
    if shop.marketplace != "ozon":
        raise HTTPException(status_code=400, detail="Only Ozon shops supported")

    ch = get_clickhouse_client()
    today = date.today()
    d_start = today - timedelta(days=period)

    try:
        # ── 1. Current stocks per warehouse ──────────────────────
        stock_rows = ch.query("""
            SELECT warehouse_name, sku, offer_id, product_name,
                   free_to_sell, reserved
            FROM fact_ozon_warehouse_stocks FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND dt = (SELECT max(dt) FROM fact_ozon_warehouse_stocks WHERE shop_id = {shop_id:UInt32})
              AND warehouse_type = 'fbo'
              AND free_to_sell > 0
        """, parameters={"shop_id": shop_id}).result_rows

        # Aggregate: wh → {skus: [{sku, offer_id, name, stock, reserved}], total}
        wh_stocks: dict[str, dict] = {}
        all_skus_set: set[int] = set()
        for row in stock_rows:
            wh, sku_id, offer_id, name, fts, reserved = row[0], int(row[1]), row[2], row[3], int(row[4]), int(row[5])
            if wh not in wh_stocks:
                wh_stocks[wh] = {"skus": {}, "total": 0}
            wh_stocks[wh]["skus"][sku_id] = {
                "sku": sku_id, "offer_id": offer_id, "name": name,
                "stock": fts, "reserved": reserved,
            }
            wh_stocks[wh]["total"] += fts
            all_skus_set.add(sku_id)

        # ── 2. Orders per warehouse × cluster_to (current period) ──
        order_rows = ch.query("""
            SELECT warehouse_name, sku, cluster_to,
                   count() AS orders, sum(quantity) AS qty
            FROM fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= {d_start:Date}
              AND status NOT IN ('cancelled')
            GROUP BY warehouse_name, sku, cluster_to
        """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows

        # Aggregate orders
        wh_orders: dict[str, dict] = {}  # wh → {total, skus: {sku → {orders, cluster_detail}}}
        for row in order_rows:
            wh, sku_id, cluster_to, orders, qty = row[0], int(row[1]), row[2], int(row[3]), int(row[4])
            if wh not in wh_orders:
                wh_orders[wh] = {"total": 0, "skus": {}, "cluster_detail": {}}
            wh_orders[wh]["total"] += orders
            # Per-SKU
            sku_data = wh_orders[wh]["skus"].setdefault(sku_id, {"orders": 0, "cluster_detail": {}})
            sku_data["orders"] += orders
            if cluster_to:
                sku_data["cluster_detail"][cluster_to] = sku_data["cluster_detail"].get(cluster_to, 0) + orders
                wh_orders[wh]["cluster_detail"][cluster_to] = wh_orders[wh]["cluster_detail"].get(cluster_to, 0) + orders

        total_orders = sum(d["total"] for d in wh_orders.values())

        # ── 3. Previous period orders (for trend) ──────────────────
        prev_d_start = d_start - timedelta(days=period)
        prev_d_end = d_start
        prev_total_orders = 0
        try:
            prev_rows = ch.query("""
                SELECT count() FROM fact_ozon_orders FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND order_date >= {prev_d_start:Date}
                  AND order_date < {prev_d_end:Date}
                  AND status NOT IN ('cancelled')
            """, parameters={"shop_id": shop_id, "prev_d_start": prev_d_start, "prev_d_end": prev_d_end}).result_rows
            prev_total_orders = int(prev_rows[0][0]) if prev_rows else 0
        except Exception:
            pass

        # ── 4. Costs from fact_ozon_transactions ───────────────────
        # Two groups: amount-based costs and services_total-based costs
        # Logistics (OperationAgentDeliveredToCustomer) is in services_total (negative)
        # while warehouse costs (crossdocking, storage) are in amount (negative)
        amount_cost_types = {
            "MarketplaceServiceItemCrossdocking": {"label": "Кроссдокинг", "icon": "arrow-right-left"},
            "OperationMarketplaceServiceStorage": {"label": "Хранение", "icon": "boxes"},
            "OperationMarketplaceSupplyAdditional": {"label": "Приёмка / FBO обработка", "icon": "package"},
            "OperationMarketplaceSupplyExpirationDateProcessing": {"label": "ФБО обработка", "icon": "factory"},
            "OperationMarketplaceServiceSupplyInboundCargoShortage": {"label": "Недостача", "icon": "alert"},
            "OperationMarketplaceServiceSupplyInboundCargoSurplus": {"label": "Излишки", "icon": "package"},
            "DefectFineShipmentDelayRated": {"label": "Штрафы отгрузка", "icon": "alert-triangle"},
            "DefectFineShipmentDelay": {"label": "Штрафы отгрузка", "icon": "alert-triangle"},
            "DefectFineCancellation": {"label": "Штрафы отмены", "icon": "alert-triangle"},
        }
        service_cost_types = {
            "OperationAgentDeliveredToCustomer": {"label": "Логистика", "icon": "truck"},
            "OperationItemReturn": {"label": "Возвраты", "icon": "undo"},
            "OperationReturnGoodsFBSofRMS": {"label": "Возвраты FBS", "icon": "undo"},
        }

        all_cost_types = list(amount_cost_types.keys()) + list(service_cost_types.keys())
        all_op_list_sql = ", ".join(f"'{op}'" for op in all_cost_types)

        costs_rows = ch.query(f"""
            SELECT operation_type,
                   count() AS cnt,
                   sum(abs(amount)) AS total_amount,
                   sum(abs(services_total)) AS total_services
            FROM fact_ozon_transactions FINAL
            WHERE shop_id = {{shop_id:UInt32}}
              AND operation_date >= {{d_start:Date}}
              AND operation_type IN ({all_op_list_sql})
            GROUP BY operation_type
        """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows

        costs_by_type: dict[str, dict] = {}
        for row in costs_rows:
            op, cnt, total_amount, total_services = row[0], int(row[1]), float(row[2]), float(row[3])
            # For logistics/acquiring/returns: cost is in services_total
            # For warehouse costs: cost is in amount
            if op in service_cost_types:
                cost_val = total_services
            else:
                cost_val = total_amount
            costs_by_type[op] = {"count": cnt, "amount": round(cost_val, 2)}

        # Previous period costs
        prev_costs_by_type: dict[str, dict] = {}
        try:
            prev_costs_rows = ch.query(f"""
                SELECT operation_type,
                       count() AS cnt,
                       sum(abs(amount)) AS total_amount,
                       sum(abs(services_total)) AS total_services
                FROM fact_ozon_transactions FINAL
                WHERE shop_id = {{shop_id:UInt32}}
                  AND operation_date >= {{prev_d_start:Date}}
                  AND operation_date < {{prev_d_end:Date}}
                  AND operation_type IN ({all_op_list_sql})
                GROUP BY operation_type
            """, parameters={"shop_id": shop_id, "prev_d_start": prev_d_start, "prev_d_end": prev_d_end}).result_rows
            for row in prev_costs_rows:
                op = row[0]
                if op in service_cost_types:
                    cost_val = float(row[3])
                else:
                    cost_val = float(row[2])
                prev_costs_by_type[op] = {"count": int(row[1]), "amount": round(cost_val, 2)}
        except Exception:
            pass

        # ── 5. Actual storage from fact_ozon_placement_cost ────────
        actual_storage_total = 0.0
        actual_storage_by_wh: dict[str, float] = {}
        has_actual_storage = False
        try:
            ps_rows = ch.query("""
                SELECT warehouse_name, sum(total_cost) AS cost
                FROM fact_ozon_placement_cost FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND date >= {d_start:Date}
                GROUP BY warehouse_name
                HAVING cost > 0
            """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows
            for r in ps_rows:
                wh_name, cost = r[0], float(r[1])
                actual_storage_by_wh[wh_name] = cost
                actual_storage_total += cost
            if actual_storage_by_wh:
                has_actual_storage = True
        except Exception as e:
            logger.warning("Ozon overview: actual storage query failed: %s", e)

        # Previous period storage
        prev_actual_storage = 0.0
        try:
            if has_actual_storage:
                prev_ps_rows = ch.query("""
                    SELECT sum(total_cost)
                    FROM fact_ozon_placement_cost FINAL
                    WHERE shop_id = {shop_id:UInt32}
                      AND date >= {prev_d_start:Date}
                      AND date < {prev_d_end:Date}
                """, parameters={"shop_id": shop_id, "prev_d_start": prev_d_start, "prev_d_end": prev_d_end}).result_rows
                prev_actual_storage = float(prev_ps_rows[0][0]) if prev_ps_rows and prev_ps_rows[0][0] else 0.0
        except Exception:
            pass

        # ── 6. Build warehouse list ───────────────────────────────
        all_wh_names = set(wh_stocks.keys()) | set(wh_orders.keys())
        month_mult = 30 / period if period > 0 else 1.0

        warehouses_result = []
        total_cross_orders = 0

        for wh_name in sorted(all_wh_names):
            stk = wh_stocks.get(wh_name, {"skus": {}, "total": 0})
            ords = wh_orders.get(wh_name, {"total": 0, "skus": {}, "cluster_detail": {}})

            stock_total = stk["total"]
            orders_total = ords["total"]
            daily_sales = orders_total / period if period > 0 else 0
            turnover = stock_total / daily_sales if daily_sales > 0 else None

            # Cluster from WAREHOUSE_TO_CLUSTER
            wh_cluster = WAREHOUSE_TO_CLUSTER.get(wh_name, "")

            # Cross analysis: orders shipped to different clusters
            cluster_detail = ords.get("cluster_detail", {})
            local_orders = cluster_detail.get(wh_cluster, 0) if wh_cluster else 0
            cross_orders = orders_total - local_orders
            cross_pct = round(cross_orders / orders_total * 100, 1) if orders_total > 0 else 0
            total_cross_orders += cross_orders

            # Status
            if stock_total == 0 and orders_total > 0:
                wh_status = "empty"
            elif turnover is not None and turnover < 14:
                wh_status = "critical"
            elif turnover is not None and turnover < 30:
                wh_status = "attention"
            elif turnover is not None and turnover > 120:
                wh_status = "overstocked"
            else:
                wh_status = "ok"

            # Storage cost
            cd_cost = costs_by_type.get("MarketplaceServiceItemCrossdocking", {}).get("amount", 0)
            storage_cost_actual = actual_storage_by_wh.get(wh_name, 0)
            storage_cost_month = round(storage_cost_actual * month_mult, 2) if storage_cost_actual > 0 else 0

            # Per-SKU details
            skus_detail = []
            sku_orders_data = ords.get("skus", {})
            for sku_id, sku_info in stk.get("skus", {}).items():
                sku_ords_data = sku_orders_data.get(sku_id, {"orders": 0, "cluster_detail": {}})
                sku_total_orders = sku_ords_data["orders"]
                sku_daily = sku_total_orders / period if period > 0 else 0
                sku_days = sku_info["stock"] / sku_daily if sku_daily > 0 else None

                # Per-SKU cross analysis
                sku_cluster_detail = sku_ords_data.get("cluster_detail", {})
                sku_local = sku_cluster_detail.get(wh_cluster, 0) if wh_cluster else 0
                sku_cross = sku_total_orders - sku_local
                sku_cross_pct = round(sku_cross / sku_total_orders * 100, 1) if sku_total_orders > 0 else 0

                skus_detail.append({
                    "sku": sku_id,
                    "offer_id": sku_info["offer_id"],
                    "name": sku_info["name"],
                    "stock": sku_info["stock"],
                    "orders": sku_total_orders,
                    "daily_sales": round(sku_daily, 2),
                    "days_supply": round(sku_days, 1) if sku_days is not None else None,
                    "cross_pct": sku_cross_pct,
                    "cross_orders": sku_cross,
                })
            skus_detail.sort(key=lambda x: x["orders"], reverse=True)

            warehouses_result.append({
                "warehouse_name": wh_name,
                "cluster": wh_cluster,
                "status": wh_status,
                "stock": stock_total,
                "sku_count": len(stk.get("skus", {})),
                "orders": orders_total,
                "daily_sales": round(daily_sales, 2),
                "turnover_days": round(turnover, 1) if turnover is not None else None,
                "pct_of_total_sales": round(orders_total / total_orders * 100, 1) if total_orders > 0 else 0,
                "cross_pct": cross_pct,
                "cross_orders": cross_orders,
                "local_orders": local_orders,
                "storage_cost_actual": round(storage_cost_actual, 2),
                "storage_cost_month": storage_cost_month,
                "skus": skus_detail[:50],
            })

        warehouses_result.sort(key=lambda x: x["orders"], reverse=True)

        # ── 7. Costs summary (grouped, WB-compatible format) ─────
        # Group fines into one line, returns into one line
        fine_ops = ["DefectFineShipmentDelayRated", "DefectFineShipmentDelay", "DefectFineCancellation"]
        return_ops = ["OperationItemReturn", "OperationReturnGoodsFBSofRMS"]

        # Aggregate fines
        fine_total = 0
        fine_count = 0
        fine_details = []
        for op in fine_ops:
            d = costs_by_type.get(op)
            if d and d["amount"] > 0:
                fine_total += d["amount"]
                fine_count += d["count"]
                # Detail label
                detail_label = amount_cost_types.get(op, {}).get("label", op)
                fine_details.append({"reason": detail_label, "amount": round(d["amount"], 2), "count": d["count"]})

        # Aggregate returns
        returns_total = 0
        returns_count = 0
        for op in return_ops:
            d = costs_by_type.get(op)
            if d and d["amount"] > 0:
                returns_total += d["amount"]
                returns_count += d["count"]

        # Build ordered costs list (WB-compatible: label, icon, count, amount)
        # Order: Логистика → Кроссдокинг → Хранение → Приёмка → Возвраты → Штрафы → остальные
        ordered_costs = [
            ("OperationAgentDeliveredToCustomer", "Логистика", "truck"),
            ("MarketplaceServiceItemCrossdocking", "Кроссдокинг", "arrow-right-left"),
            ("_storage", "Хранение", "factory"),
            ("OperationMarketplaceSupplyAdditional", "Приёмка", "package"),
            ("_returns", "Возвраты", "ban"),
            ("_fines", "Штрафы", "alert"),
            ("OperationMarketplaceServiceSupplyInboundCargoShortage", "Недостача", "alert"),
            ("OperationMarketplaceServiceSupplyInboundCargoSurplus", "Излишки", "package"),
            ("OperationMarketplaceSupplyExpirationDateProcessing", "ФБО обработка", "factory"),
        ]

        costs_summary = []
        for op_key, label, icon in ordered_costs:
            if op_key == "_storage":
                # Storage: prefer actual, fallback to transaction
                if has_actual_storage and actual_storage_total > 0:
                    costs_summary.append({
                        "operation_type": "actual_storage",
                        "label": "Хранение (факт)",
                        "icon": icon, "count": 0,
                        "amount": round(actual_storage_total, 2),
                    })
                else:
                    d = costs_by_type.get("OperationMarketplaceServiceStorage")
                    if d and d["amount"] > 0:
                        costs_summary.append({
                            "operation_type": "OperationMarketplaceServiceStorage",
                            "label": "Хранение", "icon": icon,
                            "count": d["count"], "amount": d["amount"],
                        })
            elif op_key == "_returns":
                if returns_total > 0:
                    costs_summary.append({
                        "operation_type": "returns_grouped",
                        "label": label, "icon": icon,
                        "count": returns_count, "amount": round(returns_total, 2),
                    })
            elif op_key == "_fines":
                if fine_total > 0:
                    costs_summary.append({
                        "operation_type": "fines_grouped",
                        "label": label, "icon": icon,
                        "count": fine_count, "amount": round(fine_total, 2),
                    })
            else:
                d = costs_by_type.get(op_key)
                if d and d["amount"] > 0:
                    costs_summary.append({
                        "operation_type": op_key,
                        "label": label, "icon": icon,
                        "count": d["count"], "amount": d["amount"],
                    })

        # ── 8. Out-of-stock aggregation ───────────────────────────
        # Aggregate stock + daily sales globally per SKU
        global_sku_agg: dict[int, dict] = {}
        for wh_name, stk_data in wh_stocks.items():
            sku_orders_data = wh_orders.get(wh_name, {"skus": {}}).get("skus", {})
            for sku_id, sku_info in stk_data["skus"].items():
                if sku_id not in global_sku_agg:
                    global_sku_agg[sku_id] = {
                        "sku": sku_id, "offer_id": sku_info["offer_id"],
                        "name": sku_info["name"], "stock": 0, "orders": 0,
                    }
                global_sku_agg[sku_id]["stock"] += sku_info["stock"]
                global_sku_agg[sku_id]["orders"] += sku_orders_data.get(sku_id, {"orders": 0})["orders"]

        # Also add SKUs that have orders but NO stock (filtered by free_to_sell > 0)
        missing_sku_ids: set[int] = set()
        missing_sku_orders: dict[int, int] = {}  # sku_id → total_orders
        for wh_name, wh_ord_data in wh_orders.items():
            for sku_id, sku_ord in wh_ord_data.get("skus", {}).items():
                if sku_id not in global_sku_agg:
                    missing_sku_ids.add(sku_id)
                    missing_sku_orders[sku_id] = missing_sku_orders.get(sku_id, 0) + sku_ord["orders"]

        if missing_sku_ids:
            # Resolve offer_id / name for these zero-stock SKUs
            sku_info_rows = ch.query("""
                SELECT sku, offer_id, product_name
                FROM fact_ozon_warehouse_stocks FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND sku IN {sku_ids:Array(UInt64)}
            """, parameters={"shop_id": shop_id, "sku_ids": list(missing_sku_ids)}).result_rows
            sku_info_map = {int(r[0]): {"offer_id": r[1], "name": r[2]} for r in sku_info_rows}

            for sku_id in missing_sku_ids:
                info = sku_info_map.get(sku_id, {"offer_id": str(sku_id), "name": ""})
                global_sku_agg[sku_id] = {
                    "sku": sku_id, "offer_id": info["offer_id"],
                    "name": info["name"], "stock": 0,
                    "orders": missing_sku_orders.get(sku_id, 0),
                }

        out_of_stock_skus = []
        for agg in global_sku_agg.values():
            daily = agg["orders"] / period if period > 0 else 0
            if daily > 0:
                if agg["stock"] == 0:
                    # Already out-of-stock — most critical
                    out_of_stock_skus.append({
                        "offer_id": agg["offer_id"],
                        "name": agg["name"],
                        "stock": 0,
                        "daily": round(daily, 1),
                        "days_left": 0,
                    })
                elif (agg["stock"] / daily) < 14:
                    days_left = round(agg["stock"] / daily)
                    out_of_stock_skus.append({
                        "offer_id": agg["offer_id"],
                        "name": agg["name"],
                        "stock": agg["stock"],
                        "daily": round(daily, 1),
                        "days_left": days_left,
                    })
        out_of_stock_skus.sort(key=lambda x: (x["days_left"], -x["daily"]))

        # ── 9. KPI ────────────────────────────────────────────────
        total_stock = sum(w["stock"] for w in warehouses_result)
        total_daily = sum(w["daily_sales"] for w in warehouses_result)
        avg_turnover = total_stock / total_daily if total_daily > 0 else None
        cross_pct_global = round(total_cross_orders / total_orders * 100, 1) if total_orders > 0 else 0

        total_logistics = costs_by_type.get("OperationAgentDeliveredToCustomer", {}).get("amount", 0)
        total_crossdocking = costs_by_type.get("MarketplaceServiceItemCrossdocking", {}).get("amount", 0)
        total_storage = actual_storage_total if has_actual_storage else costs_by_type.get("OperationMarketplaceServiceStorage", {}).get("amount", 0)
        total_fbo = costs_by_type.get("OperationMarketplaceSupplyAdditional", {}).get("amount", 0)
        total_fines = round(fine_total, 2)  # grouped from fine_ops above

        prev_logistics = prev_costs_by_type.get("OperationAgentDeliveredToCustomer", {}).get("amount", 0)
        prev_crossdocking = prev_costs_by_type.get("MarketplaceServiceItemCrossdocking", {}).get("amount", 0)
        prev_storage_tx = prev_costs_by_type.get("OperationMarketplaceServiceStorage", {}).get("amount", 0)
        prev_storage = prev_actual_storage if has_actual_storage else prev_storage_tx

        # Total expenses = sum of all non-revenue items (excluding acquiring — not logistics-related)
        total_expenses = total_logistics + total_crossdocking + total_storage + total_fbo + returns_total + fine_total
        prev_expenses = prev_logistics + prev_crossdocking + prev_storage

        # Cross-cluster: list ALL problem warehouses (cross_pct > 30%)
        cross_problem_warehouses = []
        for w in warehouses_result:
            if w["orders"] > 0 and w["cross_pct"] > 30:
                cross_problem_warehouses.append({
                    "warehouse_name": w["warehouse_name"],
                    "cluster": w["cluster"],
                    "cross_pct": w["cross_pct"],
                    "cross_orders": w["cross_orders"],
                    "total_orders": w["orders"],
                })
        cross_problem_warehouses.sort(key=lambda x: x["cross_pct"], reverse=True)

        kpi = {
            "total_warehouses": len([w for w in warehouses_result if w["stock"] > 0 or w["orders"] > 0]),
            "total_stock": total_stock,
            "total_sku": len(all_skus_set),
            "avg_turnover_days": round(avg_turnover, 1) if avg_turnover is not None else None,
            "total_expenses": round(total_expenses, 2),
            "total_logistics": round(total_logistics, 2),
            "total_crossdocking": round(total_crossdocking, 2),
            "total_storage": round(total_storage, 2),
            "total_fbo": round(total_fbo, 2),
            "total_returns": round(returns_total, 2),
            "total_fines": total_fines,
            "fine_details": fine_details,
            "has_actual_storage": has_actual_storage,
            "cross_pct": cross_pct_global,
            "total_orders": total_orders,
            "period_days": period,
            "out_of_stock_skus": out_of_stock_skus,
            "cross_problem_warehouses": cross_problem_warehouses,
            "prev": {
                "total_expenses": round(prev_expenses, 2),
                "total_logistics": round(prev_logistics, 2),
                "total_crossdocking": round(prev_crossdocking, 2),
                "total_storage": round(prev_storage, 2),
                "total_orders": prev_total_orders,
            },
        }

        return {
            "kpi": kpi,
            "warehouses": warehouses_result,
            "costs": costs_summary,
        }

    finally:
        ch.close()


@router.get("/ozon/analytics")
async def ozon_warehouse_analytics(
    shop_id: int = Query(...),
    period: int = Query(30, description="Period in days for sales calculation"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Warehouse analytics for Ozon FBO.

    Returns:
    - KPI summary (total stock, warehouses, avg turnover, costs)
    - Per-warehouse breakdown (stock, orders, turnover, delivery speed, costs)
    - Per-warehouse SKU details
    """
    import os
    import clickhouse_connect

    # Verify shop ownership
    shop = await db.get(Shop, shop_id)
    if not shop or shop.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Shop not found")
    if shop.marketplace != "ozon":
        raise HTTPException(status_code=400, detail="Only Ozon shops supported")

    ch_host = os.getenv("CLICKHOUSE_HOST", "clickhouse")
    ch_port = int(os.getenv("CLICKHOUSE_PORT", 8123))
    ch_user = os.getenv("CLICKHOUSE_USER", "default")
    ch_pass = os.getenv("CLICKHOUSE_PASSWORD", "")

    ch = clickhouse_connect.get_client(
        host=ch_host, port=ch_port,
        username=ch_user, password=ch_pass,
        database="mms_analytics",
    )

    try:
        # ── 1. Current stocks per warehouse ──────────────────────
        stocks_data = ch.query("""
            SELECT warehouse_name, warehouse_type,
                   groupArray(sku) as skus,
                   groupArray(offer_id) as offer_ids,
                   groupArray(product_name) as names,
                   groupArray(free_to_sell) as frees,
                   groupArray(reserved) as reserveds,
                   count() as sku_count,
                   sum(free_to_sell) as total_free,
                   sum(reserved) as total_reserved
            FROM fact_ozon_warehouse_stocks FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND dt = (SELECT max(dt) FROM fact_ozon_warehouse_stocks WHERE shop_id = {shop_id:UInt32})
              AND warehouse_type = 'fbo'
            GROUP BY warehouse_name, warehouse_type
            ORDER BY total_free DESC
        """, parameters={"shop_id": shop_id})

        # ── 2. Orders per warehouse (last N days) ────────────────
        orders_data = ch.query("""
            SELECT warehouse_name,
                   count() as order_count,
                   sum(quantity) as total_qty,
                   sum(price * quantity) as revenue,
                   uniq(sku) as active_skus
            FROM fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= today() - {period:UInt32}
              AND status NOT IN ('cancelled')
            GROUP BY warehouse_name
        """, parameters={"shop_id": shop_id, "period": period})

        # ── 3. Geography: cluster_to distribution per warehouse ──
        geo_data = ch.query("""
            SELECT warehouse_name, cluster_to,
                   count() as orders, sum(quantity) as qty
            FROM fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= today() - {period:UInt32}
              AND status NOT IN ('cancelled')
              AND cluster_to != ''
            GROUP BY warehouse_name, cluster_to
            ORDER BY warehouse_name, qty DESC
        """, parameters={"shop_id": shop_id, "period": period})

        # ── 3b. Sales geography per SKU × cluster_to ─────────────
        sku_sales_geo_data = ch.query("""
            SELECT sku, cluster_to,
                   count() as orders, sum(quantity) as qty,
                   sum(price * quantity) as revenue
            FROM fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= today() - {period:UInt32}
              AND status NOT IN ('cancelled')
              AND cluster_to != ''
            GROUP BY sku, cluster_to
            ORDER BY sku, qty DESC
        """, parameters={"shop_id": shop_id, "period": period})

        # ── 3c. Per-SKU per-warehouse geography (for cross analysis) ──
        sku_wh_geo_data = ch.query("""
            SELECT warehouse_name, sku, cluster_to,
                   count() as orders, sum(quantity) as qty
            FROM fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= today() - {period:UInt32}
              AND status NOT IN ('cancelled')
              AND cluster_to != ''
            GROUP BY warehouse_name, sku, cluster_to
        """, parameters={"shop_id": shop_id, "period": period})

        # Index: wh → sku → {total_orders, cluster_detail: {cluster → count}}
        wh_sku_geo: dict[str, dict[int, dict]] = {}
        for row in sku_wh_geo_data.result_rows:
            wh, sku_id, cluster, orders, qty = row[0], int(row[1]), row[2], int(row[3]), int(row[4])
            entry = wh_sku_geo.setdefault(wh, {}).setdefault(sku_id, {"orders": 0, "cluster_detail": {}})
            entry["orders"] += orders
            entry["cluster_detail"][cluster] = entry["cluster_detail"].get(cluster, 0) + orders

        # ── 4. Logistics costs from transactions ─────────────────
        costs_data = ch.query("""
            SELECT operation_type, operation_type_name,
                   count() as cnt, sum(amount) as total, sum(services_total) as svcs
            FROM fact_ozon_transactions FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND operation_date >= today() - {period:UInt32}
              AND operation_type IN (
                  'MarketplaceServiceItemCrossdocking',
                  'OperationMarketplaceServiceStorage',
                  'OperationMarketplaceSupplyAdditional',
                  'OperationMarketplaceSupplyExpirationDateProcessing',
                  'OperationMarketplaceServiceSupplyInboundCargoShortage'
              )
            GROUP BY operation_type, operation_type_name
        """, parameters={"shop_id": shop_id, "period": period})

        # ── 4b. Costs per warehouse (via JOIN with orders) ────────
        costs_per_wh_data = ch.query("""
            SELECT o.warehouse_name,
                   t.operation_type,
                   count() as cnt,
                   sum(t.amount) as total
            FROM fact_ozon_transactions t FINAL
            INNER JOIN (
                SELECT DISTINCT posting_number, warehouse_name
                FROM fact_ozon_orders FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND warehouse_name != ''
            ) o USING (posting_number)
            WHERE t.shop_id = {shop_id:UInt32}
              AND t.operation_date >= today() - {period:UInt32}
              AND t.operation_type IN (
                  'MarketplaceServiceItemCrossdocking',
                  'OperationMarketplaceServiceStorage',
                  'OperationMarketplaceSupplyAdditional'
              )
              AND t.posting_number != ''
            GROUP BY o.warehouse_name, t.operation_type
        """, parameters={"shop_id": shop_id, "period": period})

        # ── 5. Turnover data (if available) ──────────────────────
        turnover_data = {}
        try:
            tr = ch.query("""
                SELECT sku, days_of_supply, avg_daily_sales,
                       stock_fbo, turnover_category
                FROM fact_ozon_turnover FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND dt = (SELECT max(dt) FROM fact_ozon_turnover WHERE shop_id = {shop_id:UInt32})
            """, parameters={"shop_id": shop_id})
            for row in tr.result_rows:
                turnover_data[row[0]] = {
                    "days_of_supply": float(row[1]),
                    "avg_daily_sales": float(row[2]),
                    "stock_fbo": int(row[3]),
                    "turnover_category": row[4],
                }
        except Exception:
            pass  # Table may not exist yet

        # ── 6. Product dimensions for storage cost estimation ────
        dims_data = {}
        try:
            dims = await db.execute(text("""
                SELECT product_id, length, width, height, weight
                FROM dim_ozon_products
                WHERE shop_id = :shop_id AND length > 0 AND width > 0 AND height > 0
            """), {"shop_id": shop_id})
            for row in dims.fetchall():
                # Volume in liters (dimensions in mm → cm → liters)
                vol_liters = (row[1] / 10) * (row[2] / 10) * (row[3] / 10) / 1000
                dims_data[row[0]] = {
                    "volume_liters": round(vol_liters, 2),
                    "weight_kg": round(row[4] / 1000, 2) if row[4] else 0,
                }
        except Exception:
            pass

        # ── 7. Storage risk SKUs (turnover > 120 days per SKU) ─────
        # Calculate which SKUs are in paid storage zone or approaching it
        storage_risk_query = ch.query("""
            SELECT ws.sku, ws.offer_id, ws.product_name, ws.warehouse_name,
                   ws.free_to_sell, ws.reserved,
                   COALESCE(ord.sold, 0) as sold_period,
                   COALESCE(ord.revenue, 0) as revenue_period
            FROM fact_ozon_warehouse_stocks ws FINAL
            LEFT JOIN (
                SELECT sku, sum(quantity) as sold, sum(price * quantity) as revenue
                FROM fact_ozon_orders FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND order_date >= today() - {period:UInt32}
                  AND status != 'cancelled'
                GROUP BY sku
            ) ord ON ws.sku = ord.sku
            WHERE ws.shop_id = {shop_id:UInt32}
              AND ws.dt = (SELECT max(dt) FROM fact_ozon_warehouse_stocks WHERE shop_id = {shop_id:UInt32})
              AND ws.warehouse_type = 'fbo'
              AND ws.free_to_sell > 0
            ORDER BY ws.sku, ws.free_to_sell DESC
        """, parameters={"shop_id": shop_id, "period": period})

        # Aggregate per SKU
        sku_agg: dict[int, dict] = {}
        for row in storage_risk_query.result_rows:
            sku_id = int(row[0])
            if sku_id not in sku_agg:
                sku_agg[sku_id] = {
                    "sku": sku_id,
                    "offer_id": row[1],
                    "name": row[2],
                    "total_stock": 0,
                    "total_reserved": 0,
                    "sold_period": int(row[6]),
                    "revenue_period": float(row[7]),
                    "warehouses": [],
                }
            sku_agg[sku_id]["total_stock"] += int(row[4])
            sku_agg[sku_id]["total_reserved"] += int(row[5])
            sku_agg[sku_id]["warehouses"].append({
                "warehouse_name": row[3],
                "stock": int(row[4]),
                "reserved": int(row[5]),
            })

        # Fetch ad data for storage risk SKUs
        ad_data_map: dict[int, dict] = {}
        storage_sku_ids = list(sku_agg.keys())
        if storage_sku_ids:
            try:
                ad_query = ch.query("""
                    SELECT sku,
                           sum(money_spent) AS spend_30d,
                           sum(orders) AS orders_30d,
                           sumIf(money_spent, dt >= today() - 7) AS spend_7d
                    FROM mms_analytics.fact_ozon_ad_daily FINAL
                    WHERE shop_id = {shop_id:UInt32}
                      AND dt >= today() - 30
                      AND sku IN {sku_ids:Array(UInt64)}
                    GROUP BY sku
                """, parameters={"shop_id": shop_id, "sku_ids": storage_sku_ids})
                for r in ad_query.result_rows:
                    ad_data_map[int(r[0])] = {
                        "spend_30d": float(r[1]),
                        "orders_30d": int(r[2]),
                        "spend_7d": float(r[3]),
                        "has_active_ads": float(r[3]) > 0,  # spend in last 7d = active
                    }
            except Exception:
                pass  # ad data is optional

        # Calculate turnover and storage cost for each SKU
        # Ozon tariff: ~0.07 ₽/л/день for items > 160 days turnover (approximate)
        STORAGE_TARIFF_PER_LITER_PER_DAY = 0.07
        storage_risk_skus = []

        for sku_id, info in sku_agg.items():
            daily_sales = info["sold_period"] / period if period > 0 else 0
            turnover_days = (
                info["total_stock"] / daily_sales if daily_sales > 0 else 99999
            )

            # Only include SKUs with turnover > 120 days (approaching or in paid zone)
            if turnover_days <= 120:
                continue

            # Estimate volume-based storage cost
            dims = dims_data.get(sku_id, {})  # product_id may differ from sku
            vol_liters = dims.get("volume_liters", 0.5)  # default 0.5L

            # Days over threshold (160 days = paid storage starts)
            days_over = max(0, turnover_days - 160) if turnover_days < 99999 else 0
            # Estimated daily cost for ALL items of this SKU
            est_daily_cost = info["total_stock"] * vol_liters * STORAGE_TARIFF_PER_LITER_PER_DAY
            est_monthly_cost = est_daily_cost * 30

            zone = "free"
            if turnover_days > 160:
                zone = "paid"
            elif turnover_days > 120:
                zone = "warning"

            # Per-SKU smart recommendation (ad-aware)
            sku_ad = ad_data_map.get(sku_id, {})
            has_ads = sku_ad.get("has_active_ads", False)
            ad_spend_30d = sku_ad.get("spend_30d", 0)
            ad_orders_30d = sku_ad.get("orders_30d", 0)

            rec_action = ""
            rec_reason = ""
            rec_severity = "medium"
            if zone == "paid":
                if daily_sales == 0:
                    if has_ads and ad_spend_30d > 500:
                        rec_action = "Вывезти — реклама не помогает"
                        rec_reason = f"Потрачено {int(ad_spend_30d)} руб за 30д, 0 продаж"
                        rec_severity = "critical"
                    elif has_ads:
                        rec_action = "Скидка 40-50% или вывоз"
                        rec_reason = f"Реклама идёт ({int(ad_spend_30d)} руб), но 0 продаж — нужна скидка"
                        rec_severity = "high"
                    else:
                        rec_action = "Запустить рекламу + скидка 20-30%"
                        rec_reason = "0 продаж и нет рекламы — сначала попробовать продать"
                        rec_severity = "high"
                elif daily_sales < 0.3:
                    sell_days = int(info["total_stock"] / daily_sales) if daily_sales > 0 else 99999
                    if not has_ads:
                        rec_action = "Запустить рекламу + скидка"
                        rec_reason = f"Нет рекламы, распродажа займёт {sell_days} дн"
                        rec_severity = "high"
                    else:
                        rec_action = "Скидка 30-50% или вывоз"
                        rec_reason = f"Реклама есть, но продажи {daily_sales:.1f}/д"
                        rec_severity = "high"
                elif turnover_days > 300:
                    if not has_ads:
                        rec_action = "Запустить рекламу + промо"
                        rec_reason = f"Нет рекламы, оборач. {int(turnover_days)} дн"
                        rec_severity = "high"
                    else:
                        rec_action = "Скидка 20-30% + промо"
                        rec_reason = f"Реклама крутится, оборач. {int(turnover_days)} дн"
                        rec_severity = "high"
                else:
                    if not has_ads:
                        rec_action = "Запустить рекламу"
                        rec_reason = "Нет рекламы — можно ускорить продажи"
                        rec_severity = "medium"
                    else:
                        rec_action = "Ускорить продажи"
                        rec_reason = "Реклама идёт, цель — снизить до 160 дн"
                        rec_severity = "medium"
            elif zone == "warning":
                days_to_paid = int(160 - turnover_days) if turnover_days < 160 else 0
                if daily_sales == 0:
                    if not has_ads:
                        rec_action = "Запустить рекламу + скидку"
                        rec_reason = "0 продаж и нет рекламы — скоро платное хранение"
                        rec_severity = "high"
                    else:
                        rec_action = "Снизить цену + промо"
                        rec_reason = "Реклама не работает, 0 продаж"
                        rec_severity = "high"
                elif daily_sales < 0.5:
                    if not has_ads:
                        rec_action = "Запустить рекламу"
                        rec_reason = f"Нет рекламы, ~{days_to_paid} дн до платного"
                        rec_severity = "medium"
                    else:
                        rec_action = "Снизить цену 15-20%"
                        rec_reason = f"Реклама есть, ~{days_to_paid} дн до платного"
                        rec_severity = "medium"
                else:
                    rec_action = "Не поставлять новые"
                    rec_reason = f"Продажи {daily_sales:.1f}/д, запас велик"
                    rec_severity = "low"

            storage_risk_skus.append({
                "sku": sku_id,
                "offer_id": info["offer_id"],
                "name": info["name"],
                "total_stock": info["total_stock"],
                "sold_period": info["sold_period"],
                "daily_sales": round(daily_sales, 2),
                "turnover_days": round(turnover_days, 1) if turnover_days < 99999 else None,
                "days_over_threshold": round(days_over, 0) if days_over > 0 else 0,
                "zone": zone,
                "volume_liters": vol_liters,
                "est_daily_cost": round(est_daily_cost, 2),
                "est_monthly_cost": round(est_monthly_cost, 2),
                "revenue_period": round(info["revenue_period"], 2),
                "ad_info": {
                    "has_ads": has_ads,
                    "spend_30d": round(ad_spend_30d, 2),
                    "orders_30d": ad_orders_30d,
                },
                "recommendation": {
                    "action": rec_action,
                    "reason": rec_reason,
                    "severity": rec_severity,
                },
                "warehouses": sorted(
                    info["warehouses"],
                    key=lambda w: w["stock"],
                    reverse=True,
                ),
            })

        # Sort: paid first, then by turnover descending
        storage_risk_skus.sort(
            key=lambda s: (0 if s["zone"] == "paid" else 1, -(s["turnover_days"] or 99999)),
        )

        # ── 8. Crossdocking analysis (SKUs selling on warehouses with no stock) ───
        # If a SKU sells on a warehouse where stock=0, it's delivered via crossdocking
        # Get FBO warehouse names whitelist — only consider orders shipping TO known FBO warehouses
        # This excludes FBS warehouses (e.g. "ООО ТЕЙЛОРД") which never have FBO stock
        fbo_warehouses: set[str] = set()
        try:
            fbo_wh_query = ch.query("""
                SELECT DISTINCT warehouse_name
                FROM fact_ozon_warehouse_stocks FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND warehouse_type = 'fbo'
                  AND warehouse_name != ''
            """, parameters={"shop_id": shop_id})
            fbo_warehouses = {row[0] for row in fbo_wh_query.result_rows}
        except Exception:
            pass

        # FBO whitelist clause — only include orders for known FBO warehouses
        fbo_filter_sql = ""
        if fbo_warehouses:
            escaped = ", ".join(f"'{w}'" for w in fbo_warehouses)
            fbo_filter_sql = f"AND o.warehouse_name IN ({escaped})"

        crossdocking_sql = """
            SELECT o.warehouse_name, o.sku, o.offer_id, o.product_name,
                   sum(o.quantity) as sold,
                   round(sum(o.price * o.quantity), 2) as revenue,
                   COALESCE(ws.free_to_sell, 0) as current_stock
            FROM fact_ozon_orders o FINAL
            LEFT JOIN (
                SELECT sku, warehouse_name, free_to_sell
                FROM fact_ozon_warehouse_stocks FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND dt = (SELECT max(dt) FROM fact_ozon_warehouse_stocks
                            WHERE shop_id = {shop_id:UInt32})
                  AND warehouse_type = 'fbo'
            ) ws ON o.sku = ws.sku AND o.warehouse_name = ws.warehouse_name
            WHERE o.shop_id = {shop_id:UInt32}
              AND o.order_date >= today() - {period:UInt32}
              AND o.status != 'cancelled'
              """ + fbo_filter_sql + """
            GROUP BY o.warehouse_name, o.sku, o.offer_id, o.product_name, ws.free_to_sell
            HAVING current_stock = 0 AND sold >= 2
            ORDER BY sold DESC
        """
        crossdocking_analysis_query = ch.query(crossdocking_sql, parameters={"shop_id": shop_id, "period": period})

        # Load real dimensions (depth, height, width in mm) from PostgreSQL for volume calculation
        sku_volumes: dict[int, float] = {}  # sku → volume in liters
        try:
            # Rollback any broken prior transaction to avoid InFailedSQLTransactionError
            try:
                await db.rollback()
            except Exception:
                pass
            vol_rows = await db.execute(
                text("""SELECT sku, depth, height, width
                        FROM dim_ozon_products
                        WHERE shop_id = :sid AND sku IS NOT NULL
                          AND depth > 0 AND height > 0 AND width > 0"""),
                {"sid": shop_id},
            )
            for vr in vol_rows.fetchall():
                if vr[0]:
                    # mm × mm × mm → mm³ → liters (÷ 1_000_000)
                    volume_liters = (float(vr[1]) * float(vr[2]) * float(vr[3])) / 1_000_000
                    sku_volumes[int(vr[0])] = round(volume_liters, 2)
        except Exception:
            try:
                await db.rollback()
            except Exception:
                pass

        # Fallback: if no dimensions in PG, try loading from Ozon API on-the-fly
        if not sku_volumes:
            try:
                from app.services.ozon_products_service import OzonProductsService
                shop_row = await db.execute(
                    text("SELECT api_key, client_id FROM shops WHERE id = :sid"),
                    {"sid": shop_id},
                )
                shop_creds = shop_row.fetchone()
                if shop_creds and shop_creds[0] and shop_creds[1]:
                    # Get all product_ids that need dimensions
                    pid_rows = await db.execute(
                        text("SELECT product_id, sku FROM dim_ozon_products WHERE shop_id = :sid AND sku IS NOT NULL"),
                        {"sid": shop_id},
                    )
                    all_pids = [(r[0], r[1]) for r in pid_rows.fetchall()]
                    product_ids = [p[0] for p in all_pids]
                    pid_to_sku = {p[0]: p[1] for p in all_pids}

                    if product_ids:
                        svc = OzonProductsService(db=db, shop_id=shop_id, api_key=shop_creds[0], client_id=shop_creds[1])
                        products_info = await svc.fetch_product_info(product_ids[:100])  # limit batch

                        for item in products_info:
                            pid = item.get("id")
                            d = float(item.get("depth", 0) or 0)
                            h = float(item.get("height", 0) or 0)
                            w = float(item.get("width", 0) or 0)
                            if pid and d > 0 and h > 0 and w > 0:
                                sku = pid_to_sku.get(pid)
                                if sku:
                                    volume_liters = (d * h * w) / 1_000_000
                                    sku_volumes[int(sku)] = round(volume_liters, 2)
                                # Save back to PG for future use
                                await db.execute(
                                    text("UPDATE dim_ozon_products SET depth = :d, height = :h, width = :w WHERE shop_id = :sid AND product_id = :pid"),
                                    {"d": d, "h": h, "w": w, "sid": shop_id, "pid": pid},
                                )
                        await db.commit()
            except Exception as e:
                import logging
                logging.getLogger(__name__).warning(f"Failed to fetch dimensions from Ozon API: {e}")

        # Build stock index: sku → {warehouse → stock} for detecting excess (separate query)
        stock_by_sku: dict[int, dict[str, int]] = {}
        sku_stock_raw = ch.query("""
            SELECT sku, warehouse_name, free_to_sell
            FROM fact_ozon_warehouse_stocks FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND dt = (SELECT max(dt) FROM fact_ozon_warehouse_stocks WHERE shop_id = {shop_id:UInt32})
              AND warehouse_type = 'fbo'
              AND free_to_sell > 0
        """, parameters={"shop_id": shop_id})
        for row in sku_stock_raw.result_rows:
            sk, wh, stock = int(row[0]), row[1], int(row[2])
            if sk not in stock_by_sku:
                stock_by_sku[sk] = {}
            stock_by_sku[sk][wh] = stock

        # Build sales index: sku → {warehouse → sold} for detecting low-selling warehouses
        sales_by_sku_wh: dict[int, dict[str, int]] = {}
        sku_sales_raw = ch.query("""
            SELECT sku, warehouse_name, sum(quantity) as sold
            FROM fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= today() - {period:UInt32}
              AND status != 'cancelled'
            GROUP BY sku, warehouse_name
        """, parameters={"shop_id": shop_id, "period": period})
        for row in sku_sales_raw.result_rows:
            sk, wh, sold = int(row[0]), row[1], int(row[2])
            if sk not in sales_by_sku_wh:
                sales_by_sku_wh[sk] = {}
            sales_by_sku_wh[sk][wh] = sold

        # ── CD order geography: cities of buyers & where it shipped from ──────
        cd_geo_data: dict[int, dict[str, dict]] = {}  # sku → wh → {cities, shipped_from}
        try:
            cd_geo_query = ch.query("""
                SELECT sku, warehouse_name, city, cluster_from,
                       count() as orders, sum(quantity) as qty
                FROM fact_ozon_orders FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND order_date >= today() - {period:UInt32}
                  AND status != 'cancelled'
                  AND city != ''
                GROUP BY sku, warehouse_name, city, cluster_from
                ORDER BY qty DESC
            """, parameters={"shop_id": shop_id, "period": period})
            for row in cd_geo_query.result_rows:
                sk = int(row[0])
                wh = row[1]
                city = row[2]
                cf = row[3]
                orders = int(row[4])
                qty = int(row[5])
                if sk not in cd_geo_data:
                    cd_geo_data[sk] = {}
                if wh not in cd_geo_data[sk]:
                    cd_geo_data[sk][wh] = {"cities": {}, "shipped_from": {}}
                cd_geo_data[sk][wh]["cities"][city] = cd_geo_data[sk][wh]["cities"].get(city, 0) + qty
                if cf:
                    cd_geo_data[sk][wh]["shipped_from"][cf] = cd_geo_data[sk][wh]["shipped_from"].get(cf, 0) + qty
        except Exception:
            pass

        # Ozon FBO transfer tariff function
        def calc_transfer_cost_per_unit(volume_liters: float) -> float:
            """Ozon FBO transfer tariff (2024-2026)"""
            if volume_liters <= 5:
                return 50.4
            elif volume_liters <= 175:
                return 50.4 + 5.6 * (volume_liters - 5)
            else:
                return 1008.0

        # Aggregate: per SKU → list of warehouses where it's sold via crossdocking
        cd_sku_agg: dict[int, dict] = {}
        for row in crossdocking_analysis_query.result_rows:
            sku_id = int(row[1])
            if sku_id not in cd_sku_agg:
                cd_sku_agg[sku_id] = {
                    "sku": sku_id,
                    "offer_id": row[2],
                    "name": row[3],
                    "total_sold_via_cd": 0,
                    "total_revenue": 0,
                    "demand_warehouses": [],
                }
            cd_sku_agg[sku_id]["total_sold_via_cd"] += int(row[4])
            cd_sku_agg[sku_id]["total_revenue"] += float(row[5])
            cd_sku_agg[sku_id]["demand_warehouses"].append({
                "warehouse_name": row[0],
                "sold": int(row[4]),
                "revenue": round(float(row[5]), 2),
                "current_stock": int(row[6]),
            })

        crossdocking_skus = []

        for sku_id, info in cd_sku_agg.items():
            daily_sales_cd = info["total_sold_via_cd"] / period if period > 0 else 0
            recommended_supply = int(daily_sales_cd * 60) if daily_sales_cd > 0 else info["total_sold_via_cd"]

            # Volume for tariff calculation
            vol = sku_volumes.get(sku_id, 1.0)  # default 1 liter if unknown
            transfer_cost_per_unit = calc_transfer_cost_per_unit(vol)

            # Find donor warehouses: has stock for this SKU but low/no sales
            demand_wh_names = {w["warehouse_name"] for w in info["demand_warehouses"]}
            sku_stocks = stock_by_sku.get(sku_id, {})
            sku_sales = sales_by_sku_wh.get(sku_id, {})

            source_warehouses = []
            total_available_for_transfer = 0
            for wh_name, stock_qty in sku_stocks.items():
                if wh_name in demand_wh_names or stock_qty <= 5:
                    continue
                wh_sales = sku_sales.get(wh_name, 0)
                wh_daily_sales = wh_sales / period if period > 0 else 0
                wh_turnover = stock_qty / wh_daily_sales if wh_daily_sales > 0 else 9999
                # Excess: stock beyond 90 days coverage
                safe_stock = int(wh_daily_sales * 90)
                excess = max(0, stock_qty - safe_stock)
                if excess > 10:  # worth transferring
                    source_warehouses.append({
                        "warehouse_name": wh_name,
                        "stock": stock_qty,
                        "sales": wh_sales,
                        "daily_sales": round(wh_daily_sales, 2),
                        "turnover_days": round(wh_turnover),
                        "excess": excess,
                    })
                    total_available_for_transfer += excess

            source_warehouses.sort(key=lambda s: -s["excess"])

            # Decide action
            if total_available_for_transfer >= recommended_supply * 0.3:
                # Enough excess to transfer (at least 30% of needed)
                transfer_qty = min(total_available_for_transfer, recommended_supply)
                supply_qty = max(0, recommended_supply - transfer_qty)
                action = "transfer"
                total_transfer_cost = round(transfer_qty * transfer_cost_per_unit, 2)
            else:
                # No significant excess — recommend direct supply
                transfer_qty = 0
                supply_qty = recommended_supply
                action = "supply"
                total_transfer_cost = 0

            crossdocking_skus.append({
                "sku": sku_id,
                "offer_id": info["offer_id"],
                "name": info["name"],
                "total_sold_via_cd": info["total_sold_via_cd"],
                "total_revenue": round(info["total_revenue"], 2),
                "daily_sales_cd": round(daily_sales_cd, 2),
                "est_cd_cost_monthly": round(info["total_sold_via_cd"] * transfer_cost_per_unit / period * 30, 2),
                "recommended_supply": recommended_supply,
                "warehouse_count": len(info["demand_warehouses"]),
                "action": action,
                "transfer_qty": transfer_qty,
                "supply_qty": supply_qty,
                "volume_liters": round(vol, 2),
                "transfer_cost_per_unit": round(transfer_cost_per_unit, 2),
                "total_transfer_cost": total_transfer_cost,
                "source_warehouses": source_warehouses[:3],  # top 3 donors
                "demand_warehouses": sorted(
                    info["demand_warehouses"],
                    key=lambda w: w["sold"],
                    reverse=True,
                ),
            })

        crossdocking_skus.sort(key=lambda s: -s["total_sold_via_cd"])

        # ── 8b. Distribution Plan: aggregate by DESTINATION warehouse ─────
        # Group crossdocking_skus by demand warehouses → unified action plan
        dist_plan_wh: dict[str, dict] = {}  # warehouse_name → plan

        for sku_info in crossdocking_skus:
            sku_stocks = stock_by_sku.get(sku_info["sku"], {})
            sku_sales = sales_by_sku_wh.get(sku_info["sku"], {})
            vol = sku_info["volume_liters"]
            transfer_cost_pu = sku_info["transfer_cost_per_unit"]

            for dw in sku_info["demand_warehouses"]:
                wh_name = dw["warehouse_name"]
                if wh_name not in dist_plan_wh:
                    dist_plan_wh[wh_name] = {
                        "warehouse_name": wh_name,
                        "items": [],
                        "total_cd_cost_monthly": 0,
                        "total_transfer_cost": 0,
                        "transfer_count": 0,
                        "supply_count": 0,
                        "total_qty": 0,
                    }

                # Calculate per-warehouse recommended qty (proportional to sold)
                total_sold = sku_info["total_sold_via_cd"]
                wh_sold = dw["sold"]
                if total_sold > 0 and sku_info["recommended_supply"] > 0:
                    wh_recommended = max(1, int(sku_info["recommended_supply"] * wh_sold / total_sold))
                else:
                    wh_recommended = wh_sold

                # Per-warehouse CD cost estimate
                wh_cd_cost = round(wh_sold * transfer_cost_pu / max(period, 1) * 30, 2)

                # Determine action for THIS warehouse: can we transfer from a donor?
                item_action = "supply"
                source_wh_name = None
                source_excess = 0

                for src in sku_info["source_warehouses"]:
                    if src["excess"] >= wh_recommended * 0.3:
                        item_action = "transfer"
                        source_wh_name = src["warehouse_name"]
                        source_excess = src["excess"]
                        break

                item = {
                    "sku": sku_info["sku"],
                    "offer_id": sku_info["offer_id"],
                    "name": sku_info["name"],
                    "action": item_action,
                    "qty": wh_recommended,
                    "sold_via_cd": wh_sold,
                    "daily_sales_cd": round(wh_sold / max(period, 1), 2),
                    "revenue": dw["revenue"],
                    "volume_liters": vol,
                    "transfer_cost_per_unit": transfer_cost_pu,
                    "est_cd_cost_monthly": wh_cd_cost,
                }

                # ── Enrich with geography context ──
                geo_info = cd_geo_data.get(sku_info["sku"], {}).get(wh_name, {})
                cities = geo_info.get("cities", {})
                shipped_from = geo_info.get("shipped_from", {})

                # Top demand cities for this SKU × warehouse
                top_cities = sorted(cities.items(), key=lambda x: -x[1])[:5]
                item["demand_cities"] = [{"city": c, "qty": q} for c, q in top_cities]

                # Where shipments actually came from (cluster_from)
                top_shipped = sorted(shipped_from.items(), key=lambda x: -x[1])[:3]
                item["shipped_from"] = [{"cluster": c, "qty": q} for c, q in top_shipped]

                # Build textual reason
                city_names = ", ".join([c for c, _ in top_cities[:3]]) if top_cities else wh_name
                shipped_clusters = ", ".join([c for c, _ in top_shipped[:2]]) if top_shipped else "другие склады"

                if item_action == "transfer":
                    src_short = (source_wh_name or "").replace("_РФЦ", "").replace("_МРФЦ", "")
                    item["reason"] = f"Спрос из {city_names}. Стока 0, товар едет из {shipped_clusters}. На {src_short} избыток {source_excess} шт."
                    payback = f"~{max(1, round(wh_recommended * transfer_cost_pu / max(wh_cd_cost, 1)))} мес" if wh_cd_cost > 0 else "быстро"
                    item["benefit"] = f"Экономия ~{round(wh_cd_cost)} ₽/мес на CD + ускорение доставки. Окупаемость перемещения: {payback}."
                else:
                    item["reason"] = f"Спрос из {city_names}. Стока 0, товар едет из {shipped_clusters}. Нет избытков ни на одном складе."
                    item["benefit"] = f"Прямая поставка сократит расход ~{round(wh_cd_cost)} ₽/мес и ускорит доставку."

                if item_action == "transfer":
                    item["source_warehouse"] = source_wh_name
                    item["source_excess"] = source_excess
                    item["transfer_cost"] = round(wh_recommended * transfer_cost_pu, 2)
                    dist_plan_wh[wh_name]["transfer_count"] += 1
                    dist_plan_wh[wh_name]["total_transfer_cost"] += item["transfer_cost"]
                else:
                    dist_plan_wh[wh_name]["supply_count"] += 1

                dist_plan_wh[wh_name]["items"].append(item)
                dist_plan_wh[wh_name]["total_cd_cost_monthly"] += wh_cd_cost
                dist_plan_wh[wh_name]["total_qty"] += wh_recommended

        # Sort: warehouses with highest CD cost first; items within by sold desc
        distribution_plan = sorted(dist_plan_wh.values(), key=lambda w: -w["total_cd_cost_monthly"])
        for wp in distribution_plan:
            wp["total_cd_cost_monthly"] = round(wp["total_cd_cost_monthly"], 2)
            wp["total_transfer_cost"] = round(wp["total_transfer_cost"], 2)
            wp["items"].sort(key=lambda i: -i["sold_via_cd"])
            # Aggregate top demand cities across all items in this warehouse
            wh_cities: dict[str, int] = {}
            wh_total_orders_cd = 0
            for it in wp["items"]:
                wh_total_orders_cd += it["sold_via_cd"]
                for dc in it.get("demand_cities", []):
                    wh_cities[dc["city"]] = wh_cities.get(dc["city"], 0) + dc["qty"]
            top_wh_cities = sorted(wh_cities.items(), key=lambda x: -x[1])[:5]
            wp["top_demand_cities"] = [{"city": c, "qty": q} for c, q in top_wh_cities]
            wp["total_orders_cd"] = wh_total_orders_cd


        # Index orders by warehouse
        orders_by_wh = {}
        for row in orders_data.result_rows:
            orders_by_wh[row[0]] = {
                "orders": int(row[1]),
                "qty": int(row[2]),
                "revenue": float(row[3]),
                "active_skus": int(row[4]),
            }

        # Index per-warehouse costs
        costs_by_wh: dict[str, dict] = {}
        for row in costs_per_wh_data.result_rows:
            wh = row[0]
            op = row[1]
            if wh not in costs_by_wh:
                costs_by_wh[wh] = {"crossdocking": 0, "storage": 0, "fbo_processing": 0,
                                    "crossdocking_cnt": 0, "storage_cnt": 0, "fbo_cnt": 0}
            amount = float(row[3])
            cnt = int(row[2])
            if op == "MarketplaceServiceItemCrossdocking":
                costs_by_wh[wh]["crossdocking"] += amount
                costs_by_wh[wh]["crossdocking_cnt"] += cnt
            elif op == "OperationMarketplaceServiceStorage":
                costs_by_wh[wh]["storage"] += amount
                costs_by_wh[wh]["storage_cnt"] += cnt
            elif op == "OperationMarketplaceSupplyAdditional":
                costs_by_wh[wh]["fbo_processing"] += amount
                costs_by_wh[wh]["fbo_cnt"] += cnt

        # Index geography by warehouse
        geo_by_wh: dict[str, list] = {}
        for row in geo_data.result_rows:
            wh = row[0]
            if wh not in geo_by_wh:
                geo_by_wh[wh] = []
            geo_by_wh[wh].append({
                "cluster": row[1],
                "orders": int(row[2]),
                "qty": int(row[3]),
            })

        # Costs summary
        costs_summary = {}
        total_crossdocking = 0
        total_storage = 0
        total_fbo_processing = 0
        for row in costs_data.result_rows:
            op_type = row[0]
            amount = float(row[3])
            costs_summary[op_type] = {
                "name": row[1],
                "count": int(row[2]),
                "amount": round(amount, 2),
            }
            if op_type == "MarketplaceServiceItemCrossdocking":
                total_crossdocking = amount
            elif op_type == "OperationMarketplaceServiceStorage":
                total_storage = amount
            elif op_type == "OperationMarketplaceSupplyAdditional":
                total_fbo_processing = amount

        # Build warehouse list
        warehouses = []
        total_stock = 0
        total_skus_set = set()

        for row in stocks_data.result_rows:
            wh_name = row[0]
            wh_type = row[1]
            skus = row[2]
            offer_ids = row[3]
            names = row[4]
            frees = row[5]
            reserveds = row[6]
            sku_count = int(row[7])
            stock_free = int(row[8])
            stock_reserved = int(row[9])

            total_stock += stock_free
            total_skus_set.update(skus)

            # Get orders for this warehouse
            wh_orders = orders_by_wh.get(wh_name, {})
            order_qty = wh_orders.get("qty", 0)
            order_revenue = wh_orders.get("revenue", 0)
            order_count = wh_orders.get("orders", 0)

            # Calculate daily sales and turnover
            daily_sales = order_qty / period if period > 0 else 0
            turnover_days = stock_free / daily_sales if daily_sales > 0 else 9999
            days_to_zero = int(turnover_days) if turnover_days < 9999 else None

            # Determine cluster
            cluster = _get_cluster_for_warehouse(wh_name)

            # Calculate average delivery speed
            delivery_speed_avg = 28  # default local
            geo = geo_by_wh.get(wh_name, [])
            if geo:
                cluster_routes = DELIVERY_HOURS.get(cluster, {})
                total_weighted = 0
                total_geo_qty = 0
                for g in geo:
                    dest_cluster = g["cluster"]
                    hours = cluster_routes.get(dest_cluster, 60)
                    total_weighted += hours * g["qty"]
                    total_geo_qty += g["qty"]
                if total_geo_qty > 0:
                    delivery_speed_avg = round(total_weighted / total_geo_qty, 1)

            # Storage cost estimation
            storage_risk = "ok"
            estimated_storage_cost = 0
            if turnover_days > 180:
                storage_risk = "critical"
                estimated_storage_cost = stock_free * 1.50  # rough per-day
            elif turnover_days > 160:
                storage_risk = "warning"
                estimated_storage_cost = stock_free * 0.75

            # Status
            if stock_free == 0 and order_qty > 0:
                wh_status = "empty"
            elif turnover_days < 14:
                wh_status = "critical"
            elif turnover_days < 30:
                wh_status = "attention"
            elif turnover_days > 180:
                wh_status = "overstocked"
            elif turnover_days > 160:
                wh_status = "storage_fee"
            else:
                wh_status = "ok"

            # Total sales for % calculation
            total_orders_qty = sum(o.get("qty", 0) for o in orders_by_wh.values())
            pct_of_sales = round(order_qty / total_orders_qty * 100, 1) if total_orders_qty > 0 else 0

            # SKU details (enriched with cross analysis)
            sku_details = []
            wh_sku_geo_data = wh_sku_geo.get(wh_name, {})
            for i in range(len(skus)):
                sku_id = int(skus[i])
                sku_free = int(frees[i])
                sku_turnover = turnover_data.get(sku_id, {})
                sku_daily = sku_turnover.get("avg_daily_sales", 0)
                sku_days_supply = sku_turnover.get("days_of_supply", 0)

                # Fallback: calculate from orders if no turnover data
                if sku_daily == 0 and daily_sales > 0 and sku_count > 0:
                    sku_daily = daily_sales / sku_count  # rough approximation

                # Per-SKU cross geography from this warehouse
                sku_geo_entry = wh_sku_geo_data.get(sku_id, {"orders": 0, "cluster_detail": {}})
                sku_total_orders = sku_geo_entry["orders"]
                sku_cluster_detail = sku_geo_entry["cluster_detail"]
                sku_local_orders = sku_cluster_detail.get(cluster, 0)
                sku_cross_orders = sku_total_orders - sku_local_orders
                sku_cross_pct = round(sku_cross_orders / sku_total_orders * 100, 1) if sku_total_orders > 0 else 0

                sku_geo_list = []
                for cl_name, cl_count in sorted(sku_cluster_detail.items(), key=lambda x: x[1], reverse=True):
                    cl_share = round(cl_count / sku_total_orders * 100, 1) if sku_total_orders > 0 else 0
                    sku_geo_list.append({
                        "cluster": cl_name,
                        "orders": cl_count,
                        "share": cl_share,
                        "is_local": cl_name == cluster,
                    })

                sku_details.append({
                    "sku": sku_id,
                    "offer_id": offer_ids[i],
                    "name": names[i],
                    "stock": sku_free,
                    "reserved": int(reserveds[i]),
                    "daily_sales": round(sku_daily, 2),
                    "days_supply": round(sku_days_supply, 1) if sku_days_supply > 0 else (
                        round(sku_free / sku_daily, 0) if sku_daily > 0 else None
                    ),
                    "turnover_category": sku_turnover.get("turnover_category", ""),
                    "orders": sku_total_orders,
                    "cross_orders": sku_cross_orders,
                    "cross_pct": sku_cross_pct,
                    "geography": sku_geo_list,
                })

            # Sort SKU details: highest orders first (for cross analysis)
            sku_details.sort(key=lambda x: x.get("orders", 0), reverse=True)

            # Geography with shares + cross analysis
            geo_total_orders = sum(g["orders"] for g in geo) if geo else 0
            geo_total = sum(g["qty"] for g in geo) if geo else 0
            wh_local_orders = 0
            clusters_served = []
            for g in sorted(geo, key=lambda x: x["qty"], reverse=True)[:10]:
                is_local_cluster = g["cluster"] == cluster
                if is_local_cluster:
                    wh_local_orders += g["orders"]
                clusters_served.append({
                    "cluster": g["cluster"],
                    "orders": g["orders"],
                    "qty": g["qty"],
                    "share": round(g["qty"] / geo_total * 100, 1) if geo_total > 0 else 0,
                    "is_local": is_local_cluster,
                })
            # Also count local from non-top-10 for accuracy
            for g in geo:
                if g["cluster"] == cluster and not any(c["cluster"] == g["cluster"] for c in clusters_served):
                    wh_local_orders += g["orders"]
            wh_cross_orders = geo_total_orders - wh_local_orders
            wh_cross_pct = round(wh_cross_orders / geo_total_orders * 100, 1) if geo_total_orders > 0 else 0

            # Per-warehouse costs
            wh_costs = costs_by_wh.get(wh_name, {})

            warehouses.append({
                "warehouse_name": wh_name,
                "cluster": cluster,
                "warehouse_type": wh_type,
                "stock_free": stock_free,
                "stock_reserved": stock_reserved,
                "sku_count": sku_count,
                "orders_period": order_count,
                "qty_period": order_qty,
                "revenue_period": round(order_revenue, 2),
                "daily_sales": round(daily_sales, 2),
                "turnover_days": round(turnover_days, 1) if turnover_days < 9999 else None,
                "days_to_zero": days_to_zero,
                "pct_of_total_sales": pct_of_sales,
                "delivery_speed_avg_h": delivery_speed_avg,
                "status": wh_status,
                "storage_risk": storage_risk,
                "estimated_storage_cost_day": round(estimated_storage_cost, 2),
                "cross_pct": wh_cross_pct,
                "cross_orders": wh_cross_orders,
                "local_orders": wh_local_orders,
                "costs": {
                    "crossdocking": round(wh_costs.get("crossdocking", 0), 2),
                    "crossdocking_cnt": wh_costs.get("crossdocking_cnt", 0),
                    "storage": round(wh_costs.get("storage", 0), 2),
                    "storage_cnt": wh_costs.get("storage_cnt", 0),
                    "fbo_processing": round(wh_costs.get("fbo_processing", 0), 2),
                    "fbo_cnt": wh_costs.get("fbo_cnt", 0),
                    "total": round(
                        wh_costs.get("crossdocking", 0)
                        + wh_costs.get("storage", 0)
                        + wh_costs.get("fbo_processing", 0), 2
                    ),
                },
                "clusters_served": clusters_served,
                "skus": sku_details,
            })

        # Sort: by daily sales descending
        warehouses.sort(key=lambda w: w["daily_sales"], reverse=True)

        # Averages
        wh_with_sales = [w for w in warehouses if w["daily_sales"] > 0]
        avg_turnover = (
            round(sum(w["turnover_days"] for w in wh_with_sales if w["turnover_days"])
                  / len(wh_with_sales), 1)
            if wh_with_sales else None
        )
        avg_delivery = (
            round(sum(w["delivery_speed_avg_h"] for w in wh_with_sales)
                  / len(wh_with_sales), 1)
            if wh_with_sales else None
        )

        critical_count = sum(1 for w in warehouses if w["status"] in ("critical", "empty"))
        overstocked_count = sum(1 for w in warehouses if w["status"] in ("overstocked", "storage_fee"))

        # ── Cross-map: warehouse × cluster_to matrix ─────────────
        # Include ALL warehouses with orders, not just those with stock
        _all_clusters_seen: set[str] = set()
        total_cross_orders_all = 0
        total_orders_all = 0
        for w in warehouses:
            total_cross_orders_all += w.get("cross_orders", 0)
            total_orders_all += w.get("cross_orders", 0) + w.get("local_orders", 0)
            for cs in w.get("clusters_served", []):
                _all_clusters_seen.add(cs["cluster"])

        # Also collect clusters from geo_data (warehouses with orders but without stock)
        _wh_in_warehouses = {w["warehouse_name"] for w in warehouses}
        _extra_wh_geo: dict[str, dict[str, int]] = {}  # wh_name → {cluster → orders}
        _extra_wh_orders: dict[str, int] = {}
        for row in geo_data.result_rows:
            wh_name_g, cluster_g, orders_g = row[0], row[1], int(row[2])
            _all_clusters_seen.add(cluster_g)
            if wh_name_g not in _wh_in_warehouses:
                _extra_wh_geo.setdefault(wh_name_g, {})[cluster_g] = \
                    _extra_wh_geo.get(wh_name_g, {}).get(cluster_g, 0) + orders_g
                _extra_wh_orders[wh_name_g] = _extra_wh_orders.get(wh_name_g, 0) + orders_g

        cluster_list = sorted(_all_clusters_seen)
        overall_cross_pct = round(total_cross_orders_all / total_orders_all * 100, 1) if total_orders_all > 0 else 0

        cross_map = []

        # 1) Warehouses from stocks (with stock data)
        for wh_data in warehouses:
            wh_geo_orders = wh_data.get("cross_orders", 0) + wh_data.get("local_orders", 0)
            if wh_geo_orders == 0:
                continue
            clusters_detail = {cs["cluster"]: cs["orders"] for cs in wh_data.get("clusters_served", [])}
            row_data = {
                "warehouse": wh_data["warehouse_name"],
                "home_cluster": wh_data["cluster"],
                "total_orders": wh_geo_orders,
                "clusters": {},
            }
            for cl_name in cluster_list:
                cnt = clusters_detail.get(cl_name, 0)
                is_local = cl_name == wh_data["cluster"]
                row_data["clusters"][cl_name] = {
                    "count": cnt,
                    "is_local": is_local,
                }
            cross_map.append(row_data)

        # 2) Extra warehouses from geo_data (orders exist, no FBO stock)
        for wh_name_extra, cluster_orders in _extra_wh_geo.items():
            wh_cluster_extra = _get_cluster_for_warehouse(wh_name_extra)
            total_extra_orders = _extra_wh_orders.get(wh_name_extra, 0)
            row_data = {
                "warehouse": wh_name_extra,
                "home_cluster": wh_cluster_extra,
                "total_orders": total_extra_orders,
                "clusters": {},
            }
            for cl_name in cluster_list:
                cnt = cluster_orders.get(cl_name, 0)
                is_local = cl_name == wh_cluster_extra
                row_data["clusters"][cl_name] = {
                    "count": cnt,
                    "is_local": is_local,
                }
            cross_map.append(row_data)
            # Add to totals
            local_extra = cluster_orders.get(wh_cluster_extra, 0)
            cross_extra = total_extra_orders - local_extra
            total_cross_orders_all += cross_extra
            total_orders_all += total_extra_orders

        # Sort cross_map by total orders desc
        cross_map.sort(key=lambda x: x["total_orders"], reverse=True)

        # ── Generate recommendations ─────────────────────────────
        recommendations = []

        # 1. Overstocked → recommend moving to low-stock warehouses (MULTI-target)
        overstocked_whs = [w for w in warehouses
                           if w["status"] in ("overstocked", "storage_fee")
                           and w["stock_free"] > 0]
        low_stock_whs = [w for w in warehouses
                         if w["status"] in ("critical", "attention")
                         and w["daily_sales"] > 0]

        for ow in overstocked_whs:
            ow_td = ow["turnover_days"] or 9999
            excess_days = ow_td - 90  # target: 90 days
            if excess_days > 0 and ow["daily_sales"] > 0:
                excess_qty = int(excess_days * ow["daily_sales"])
            else:
                excess_qty = ow["stock_free"]

            if excess_qty <= 0:
                continue

            is_paid = ow_td > 160

            # Find ALL target warehouses with common SKUs, sorted by urgency
            ow_sku_map = {s["sku"]: s for s in ow["skus"] if s["stock"] > 10}
            ow_skus = set(ow_sku_map.keys())

            targets = []
            for lw in low_stock_whs:
                if lw["warehouse_name"] == ow["warehouse_name"]:
                    continue
                lw_skus = {s["sku"] for s in lw["skus"]}
                common = ow_skus & lw_skus
                if not common:
                    continue
                lw_td = lw["turnover_days"] or 0
                lw_daily = lw["daily_sales"] or 0.01
                # How many units can this warehouse absorb in 90 days
                current_stock = lw["stock_free"]
                target_stock = int(lw_daily * 90)  # 90 days of stock
                can_absorb = max(0, target_stock - current_stock)
                if can_absorb <= 0:
                    continue
                targets.append({
                    "warehouse": lw["warehouse_name"],
                    "cluster": lw["cluster"],
                    "status": lw["status"],
                    "current_td": int(lw_td),
                    "daily_sales": lw_daily,
                    "current_stock": current_stock,
                    "can_absorb": can_absorb,
                    "common_skus": len(common),
                    "common_names": [ow_sku_map[s].get("name", ow_sku_map[s].get("offer_id", ""))
                                     for s in list(common)[:5]],
                })

            if not targets:
                continue

            # Sort by urgency (lowest turnover first — most need stock)
            targets.sort(key=lambda t: t["current_td"])

            # Distribute excess_qty across targets
            remaining = excess_qty
            distribution = []
            for t in targets:
                if remaining <= 0:
                    break
                move_qty = min(remaining, t["can_absorb"])
                if move_qty < 5:
                    continue  # not worth it
                new_stock = t["current_stock"] + move_qty
                new_td = int(new_stock / t["daily_sales"]) if t["daily_sales"] > 0 else 9999
                distribution.append({
                    **t,
                    "move_qty": move_qty,
                    "projected_td": new_td,
                })
                remaining -= move_qty

            if not distribution:
                continue

            total_moved = sum(d["move_qty"] for d in distribution)
            new_ow_stock = max(0, ow["stock_free"] - total_moved)
            new_ow_td = int(new_ow_stock / ow["daily_sales"]) if ow["daily_sales"] > 0 else 9999

            # Estimate savings
            est_storage_saved = 0
            if is_paid:
                est_storage_saved = int(total_moved * 0.5 * 0.07 * 30)

            # Build action items with detail per target
            action_items = []
            for d in distribution[:3]:  # max 3 targets
                action_items.append(
                    f"{ow['warehouse_name']} → {d['warehouse']} ({d['cluster']}): "
                    f"~{d['move_qty']} ед. "
                    f"[{d['warehouse']}: {d['current_td']}д → {d['projected_td']}д запаса]"
                )

            # Build impact text
            impact_parts = []
            if is_paid:
                impact_parts.append(
                    f"Вы платите за хранение товара, который не продаётся на {ow['warehouse_name']}."
                )
            impact_parts.append(
                f"Итого перемещение ~{total_moved} ед. по {len(distribution)} складам. "
                f"{ow['warehouse_name']}: {int(ow_td)}д → {new_ow_td}д запаса."
            )
            for d in distribution[:2]:
                impact_parts.append(
                    f"{d['warehouse']}: получит {d['move_qty']} ед., "
                    f"запас вырастет с {d['current_td']}д до {d['projected_td']}д "
                    f"(продаёт {d['daily_sales']:.1f} ед/день)."
                )
            if est_storage_saved > 0:
                impact_parts.append(f"Потенциальная экономия на хранении: ~{est_storage_saved:,} ₽/мес.")

            recommendations.append({
                "type": "move_stock",
                "severity": "high" if is_paid else "medium",
                "from_warehouse": ow["warehouse_name"],
                "from_cluster": ow["cluster"],
                "title": f"Перемещение со склада {ow['warehouse_name']}",
                "reason": (
                    f"На складе {ow['warehouse_name']} ({ow['cluster']}) оборачиваемость "
                    f"{int(ow_td)} дн — {'в зоне ПЛАТНОГО хранения (>160 дн)' if is_paid else 'сильно выше нормы (90 дн)'}. "
                    f"Можно распределить ~{total_moved} ед. по {len(distribution)} складам, "
                    f"у которых дефицит и общие SKU."
                ),
                "impact": " ".join(impact_parts),
                "action_items": action_items,
                "distribution": distribution[:5],
                "affected_skus": sum(d["common_skus"] for d in distribution),
                "excess_qty": total_moved,
                "est_savings": est_storage_saved,
            })

        # 2. Warehouses with high crossdocking → recommend direct supply
        for w in warehouses:
            cd_cost = abs(w["costs"].get("crossdocking", 0))
            if cd_cost > 5000 and w["daily_sales"] > 0.5:
                monthly_cd = round(cd_cost / period * 30)
                recommendations.append({
                    "type": "optimize_crossdocking",
                    "severity": "medium",
                    "warehouse": w["warehouse_name"],
                    "cluster": w["cluster"],
                    "crossdocking_cost": round(cd_cost, 2),
                    "daily_sales": w["daily_sales"],
                    "title": f"Высокий расход на кроссдокинг: {w['warehouse_name']}",
                    "reason": (
                        f"Склад {w['warehouse_name']} ({w['cluster']}) — расход на кроссдокинг "
                        f"{round(cd_cost):,} ₽ за {period} дн (~{monthly_cd:,} ₽/мес). "
                        f"Кроссдокинг означает, что ваш товар сначала едет на центральный склад, "
                        f"а потом перераспределяется — каждая перевалка стоит денег."
                    ),
                    "impact": (
                        f"Прямая поставка на {w['warehouse_name']} может сэкономить "
                        f"до {int(cd_cost * 0.6):,} ₽ за период (до 60% от расхода на кроссдокинг). "
                        f"При продажах {w['daily_sales']:.1f} ед/день склад достаточно активный "
                        f"для оправдания прямых поставок."
                    ),
                    "action_items": [
                        f"Рассмотреть создание прямого маршрута поставки на {w['warehouse_name']}",
                        f"Сравнить стоимость логистики прямой поставки vs кроссдокинг (~{monthly_cd:,} ₽/мес)",
                    ],
                    "est_savings": int(cd_cost * 0.6),
                })

        # 3. Warehouses approaching paid storage threshold
        for w in warehouses:
            td = w["turnover_days"]
            if td and 120 < td <= 160:
                days_left = int(160 - td)
                # Estimate what storage would cost once it crosses threshold
                est_monthly_if_paid = int(w["stock_free"] * 0.5 * 0.07 * 30)
                recommendations.append({
                    "type": "storage_warning",
                    "severity": "medium",
                    "warehouse": w["warehouse_name"],
                    "cluster": w["cluster"],
                    "turnover_days": round(td),
                    "days_to_paid": days_left,
                    "stock": w["stock_free"],
                    "title": f"Приближение к платному хранению: {w['warehouse_name']}",
                    "reason": (
                        f"Склад {w['warehouse_name']} ({w['cluster']}) — оборачиваемость {int(td)} дн. "
                        f"До порога платного хранения Ozon (160 дн) осталось ~{days_left} дней. "
                        f"На складе {w['stock_free']} ед. товара."
                    ),
                    "impact": (
                        f"Если ничего не делать, через ~{days_left} дн Ozon начнёт взимать плату "
                        f"за хранение. Прогноз расхода при текущих стоках: ~{est_monthly_if_paid:,} ₽/мес."
                    ),
                    "action_items": [
                        "Снизить цену для ускорения продаж на этом складе",
                        "Запустить акцию/промо для товаров с низкой оборачиваемостью",
                        f"Уменьшить поставки на {w['warehouse_name']} до стабилизации",
                    ],
                    "est_savings": est_monthly_if_paid,
                })

        # 4. (Removed: per-warehouse paid_storage recs — now handled per-SKU in storage_risk_skus)

        # Sort recs: high severity first, then by est_savings descending
        severity_order = {"high": 0, "medium": 1, "low": 2}
        recommendations.sort(key=lambda r: (severity_order.get(r["severity"], 9), -(r.get("est_savings", 0))))

        # ── Generate text summary of problems ─────────────────────
        summary_parts = []

        # Storage problems
        paid_skus = [s for s in storage_risk_skus if s.get("zone") == "paid"]
        warning_skus = [s for s in storage_risk_skus if s.get("zone") == "warning"]
        paid_monthly = sum(s.get("est_monthly_cost", 0) for s in paid_skus)
        if paid_skus:
            summary_parts.append(
                f"⚠️ {len(paid_skus)} SKU в зоне платного хранения — "
                f"прогноз расхода ~{paid_monthly:,.0f} ₽/мес."
            )
        if warning_skus:
            summary_parts.append(
                f"🟡 {len(warning_skus)} SKU приближаются к платному хранению — "
                f"нужно ускорить продажи или вывезти."
            )

        # Overstocked warehouses
        overstocked_whs = [w for w in warehouses if w.get("status") == "overstocked"]
        if overstocked_whs:
            names = ", ".join(w["warehouse_name"] for w in overstocked_whs[:3])
            summary_parts.append(
                f"📦 {len(overstocked_whs)} складов перезатарены: {names}."
            )

        # Crossdocking costs
        if crossdocking_skus:
            cd_total_monthly = sum(s.get("est_cd_cost_monthly", 0) for s in crossdocking_skus)
            transfer_count = sum(1 for s in crossdocking_skus if s.get("action") == "transfer")
            supply_count = sum(1 for s in crossdocking_skus if s.get("action") == "supply")
            parts = []
            if transfer_count:
                parts.append(f"{transfer_count} переместить")
            if supply_count:
                parts.append(f"{supply_count} поставить напрямую")
            summary_parts.append(
                f"🔄 {len(crossdocking_skus)} SKU доставляются через кроссдокинг — "
                f"расход ~{cd_total_monthly:,.0f} ₽/мес. "
                f"Рекомендация: {', '.join(parts)}."
            )

        # Savings potential
        total_savings = sum(r.get("est_savings", 0) for r in recommendations)
        if total_savings > 0:
            summary_parts.append(
                f"💰 Потенциал экономии: ~{total_savings:,.0f} ₽/мес при выполнении рекомендаций."
            )

        # Delivery speed
        if avg_delivery and avg_delivery > 48:
            summary_parts.append(
                f"🚚 Средняя скорость доставки {avg_delivery:.0f}ч — выше нормы (48ч). "
                f"Рассмотрите размещение стоков ближе к регионам спроса."
            )

        summary = " ".join(summary_parts) if summary_parts else "✅ Серьёзных проблем не обнаружено."

        # ── Build SKU geography (stocking + sales view) ────────────
        # First build sales-by-cluster map from sku_sales_geo_data
        sku_sales_map: dict[int, list] = {}
        for row in sku_sales_geo_data.result_rows:
            sku_id = int(row[0])
            if sku_id not in sku_sales_map:
                sku_sales_map[sku_id] = []
            sku_sales_map[sku_id].append({
                "cluster": row[1],
                "orders": int(row[2]),
                "qty": int(row[3]),
                "revenue": round(float(row[4]), 2),
            })

        sku_geo_map: dict[int, dict] = {}
        for w in warehouses:
            for sku_detail in w.get("skus", []):
                sku_id = sku_detail["sku"]
                if sku_id not in sku_geo_map:
                    sku_geo_map[sku_id] = {
                        "sku": sku_id,
                        "offer_id": sku_detail.get("offer_id", ""),
                        "name": sku_detail.get("name", ""),
                        "total_stock": 0,
                        "total_daily_sales": 0.0,
                        "warehouses": [],
                        "sales_clusters": [],
                    }
                sku_geo_map[sku_id]["total_stock"] += sku_detail.get("stock", 0)
                sku_geo_map[sku_id]["total_daily_sales"] += sku_detail.get("daily_sales", 0)
                sku_geo_map[sku_id]["warehouses"].append({
                    "warehouse_name": w["warehouse_name"],
                    "cluster": w["cluster"],
                    "stock": sku_detail.get("stock", 0),
                    "reserved": sku_detail.get("reserved", 0),
                    "daily_sales": round(sku_detail.get("daily_sales", 0), 2),
                    "days_supply": sku_detail.get("days_supply"),
                    "warehouse_status": w["status"],
                })

        # Attach sales_clusters to each SKU
        for sku_id, geo_entry in sku_geo_map.items():
            geo_entry["sales_clusters"] = sku_sales_map.get(sku_id, [])

        sku_geography = sorted(
            sku_geo_map.values(),
            key=lambda s: s["total_stock"],
            reverse=True,
        )

        return {
            "kpi": {
                "total_warehouses": len(warehouses),
                "total_stock": total_stock,
                "total_skus": len(total_skus_set),
                "avg_turnover_days": avg_turnover,
                "avg_delivery_h": avg_delivery,
                "total_crossdocking": round(total_crossdocking, 2),
                "total_storage_fee": round(total_storage, 2),
                "total_fbo_processing": round(total_fbo_processing, 2),
                "critical_warehouses": critical_count,
                "overstocked_warehouses": overstocked_count,
                "period_days": period,
                "cross_pct": overall_cross_pct,
            },
            "summary": summary,
            "costs": costs_summary,
            "warehouses": warehouses,
            "cross_map": cross_map,
            "cluster_list": cluster_list,
            "recommendations": recommendations,
            "storage_risk_skus": storage_risk_skus,
            "crossdocking_skus": crossdocking_skus,
            "distribution_plan": distribution_plan,
            "sku_geography": sku_geography,
        }

    finally:
        ch.close()


# ─── Excel export: distribution plan ──────────────────────────────────

@router.get("/ozon/analytics/distribution-plan/excel")
async def distribution_plan_excel(
    shop_id: int = Query(...),
    period: int = Query(30),
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Download distribution plan as formatted Excel workbook."""
    from fastapi.responses import StreamingResponse
    import io

    # Reuse the analytics handler to get distribution_plan data
    result = await ozon_warehouse_analytics(shop_id=shop_id, period=period, db=db, current_user=current_user)
    plan = result.get("distribution_plan", [])

    if not plan:
        raise HTTPException(404, "Нет данных для плана распределения")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Styles ──
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill_blue = PatternFill("solid", fgColor="2F5496")
    hdr_fill_green = PatternFill("solid", fgColor="2E7D32")
    hdr_fill_purple = PatternFill("solid", fgColor="6A1B9A")
    wh_fill = PatternFill("solid", fgColor="D9E2F3")
    wh_font = Font(bold=True, size=12)
    bold_font = Font(bold=True, size=11)
    bold14 = Font(bold=True, size=14)
    reason_font = Font(size=11, color="333333")
    benefit_font = Font(bold=True, size=11, color="1B5E20")
    transfer_font = Font(bold=True, size=11, color="6A1B9A")
    supply_font = Font(bold=True, size=11, color="2E7D32")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(bottom=thin, top=thin, left=thin, right=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    num_fmt = "#,##0"
    rub_fmt = '#,##0 "₽"'

    def style_header(ws, headers, fill):
        for ci, (name, w) in enumerate(headers, 1):
            c = ws.cell(1, ci, name)
            c.font = hdr_font
            c.fill = fill
            c.alignment = center
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ════════════════════════════════════════════════════════════════
    # Sheet 1: СВОДНЫЙ ПЛАН
    # ════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Сводный план"

    headers1 = [
        ("Склад назначения", 24), ("Действие", 14), ("Артикул", 22),
        ("Название", 40), ("Кол-во, шт", 12), ("Прод/день CD", 10),
        ("CD расход/мес", 14), ("Источник", 22),
        ("Стоим. перемещ.", 14), ("Города спроса", 30),
        ("Откуда едет сейчас", 22), ("Причина", 50), ("Выгода", 40),
    ]
    style_header(ws1, headers1, hdr_fill_blue)

    row = 2
    for wh in plan:
        wh_name = wh["warehouse_name"]
        for item in wh["items"]:
            action = "⇄ Переместить" if item["action"] == "transfer" else "↓ Поставить"
            cities = ", ".join([f'{d["city"]} ({d["qty"]})' for d in item.get("demand_cities", [])])
            shipped = ", ".join([f'{d["cluster"]} ({d["qty"]})' for d in item.get("shipped_from", [])])

            ws1.cell(row, 1, wh_name).font = bold_font
            c_act = ws1.cell(row, 2, action)
            c_act.font = transfer_font if item["action"] == "transfer" else supply_font
            ws1.cell(row, 3, item["offer_id"]).font = bold_font
            ws1.cell(row, 4, item.get("name", "")).alignment = wrap
            ws1.cell(row, 5, item["qty"]).number_format = num_fmt
            ws1.cell(row, 6, round(item["daily_sales_cd"], 1)).number_format = "0.0"
            ws1.cell(row, 7, round(item["est_cd_cost_monthly"])).number_format = rub_fmt
            ws1.cell(row, 8, (item.get("source_warehouse") or "Новая поставка").replace("_РФЦ", "").replace("_МРФЦ", ""))
            ws1.cell(row, 9, round(item.get("transfer_cost", 0))).number_format = rub_fmt
            ws1.cell(row, 10, cities).alignment = wrap
            ws1.cell(row, 11, shipped).alignment = wrap
            ws1.cell(row, 12, item.get("reason", "")).alignment = wrap
            ws1.cell(row, 12).font = reason_font
            ws1.cell(row, 13, item.get("benefit", "")).alignment = wrap
            ws1.cell(row, 13).font = benefit_font

            for ci in range(1, 14):
                ws1.cell(row, ci).border = border
            row += 1

    # ════════════════════════════════════════════════════════════════
    # Sheet 2: ПОСТАВКИ ПО СКЛАДАМ
    # ════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Поставки по складам")

    headers2 = [
        ("Склад назначения", 28), ("Артикул", 22), ("Название", 40),
        ("Кол-во, шт", 12), ("Прод/день CD", 10), ("CD расход/мес", 14),
        ("Города спроса", 30), ("Причина", 50),
    ]
    style_header(ws2, headers2, hdr_fill_green)

    row = 2
    for wh in plan:
        supply_items = [i for i in wh["items"] if i["action"] == "supply"]
        if not supply_items:
            continue

        wh_short = wh["warehouse_name"].replace("_РФЦ", "").replace("_МРФЦ", "")
        merge_start = row

        for item in supply_items:
            cities = ", ".join([f'{d["city"]} ({d["qty"]})' for d in item.get("demand_cities", [])])

            ws2.cell(row, 1, wh_short)
            ws2.cell(row, 2, item["offer_id"]).font = bold_font
            ws2.cell(row, 3, item.get("name", "")).alignment = wrap
            ws2.cell(row, 4, item["qty"]).number_format = num_fmt
            ws2.cell(row, 5, round(item["daily_sales_cd"], 1)).number_format = "0.0"
            ws2.cell(row, 6, round(item["est_cd_cost_monthly"])).number_format = rub_fmt
            ws2.cell(row, 7, cities).alignment = wrap
            ws2.cell(row, 8, item.get("reason", "")).alignment = wrap
            ws2.cell(row, 8).font = reason_font

            for ci in range(1, 9):
                ws2.cell(row, ci).border = border
            row += 1

        # Merge warehouse name cells
        if row - merge_start > 1:
            ws2.merge_cells(start_row=merge_start, start_column=1, end_row=row - 1, end_column=1)
        ws2.cell(merge_start, 1).font = wh_font
        ws2.cell(merge_start, 1).fill = wh_fill
        ws2.cell(merge_start, 1).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

        # Subtotal row
        total_qty = sum(i["qty"] for i in supply_items)
        total_cd = sum(i["est_cd_cost_monthly"] for i in supply_items)
        ws2.cell(row, 1, "").fill = PatternFill("solid", fgColor="E8F5E9")
        ws2.cell(row, 2, f"ИТОГО: {wh_short}").font = bold_font
        ws2.cell(row, 2).fill = PatternFill("solid", fgColor="E8F5E9")
        ws2.cell(row, 3, f"{len(supply_items)} SKU").fill = PatternFill("solid", fgColor="E8F5E9")
        ws2.cell(row, 4, total_qty).number_format = num_fmt
        ws2.cell(row, 4).font = bold14
        ws2.cell(row, 4).fill = PatternFill("solid", fgColor="E8F5E9")
        ws2.cell(row, 6, round(total_cd)).number_format = rub_fmt
        ws2.cell(row, 6).font = bold_font
        ws2.cell(row, 6).fill = PatternFill("solid", fgColor="E8F5E9")
        for ci in range(1, 9):
            ws2.cell(row, ci).border = border
            if ci not in (1, 2, 3, 4, 6):
                ws2.cell(row, ci).fill = PatternFill("solid", fgColor="E8F5E9")
        row += 1  # blank row between warehouses
        row += 1

    # ════════════════════════════════════════════════════════════════
    # Sheet 3: ПЕРЕМЕЩЕНИЯ
    # ════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Перемещения")

    headers3 = [
        ("Откуда", 24), ("Куда", 24), ("Артикул", 22), ("Название", 40),
        ("Кол-во, шт", 12), ("Прод/день CD", 10), ("Избыток на источн.", 14),
        ("Стоим. перемещ.", 14), ("CD расход/мес", 14),
        ("Окупаемость", 14), ("Причина", 50),
    ]
    style_header(ws3, headers3, hdr_fill_purple)

    # Group transfers by source warehouse
    transfers_by_source: dict[str, list] = {}
    for wh in plan:
        for item in wh["items"]:
            if item["action"] == "transfer" and item.get("source_warehouse"):
                src = item["source_warehouse"]
                if src not in transfers_by_source:
                    transfers_by_source[src] = []
                transfers_by_source[src].append({**item, "_dest_wh": wh["warehouse_name"]})

    row = 2
    for src_wh, items in sorted(transfers_by_source.items(), key=lambda x: -sum(i.get("transfer_cost", 0) for i in x[1])):
        src_short = src_wh.replace("_РФЦ", "").replace("_МРФЦ", "")
        merge_start = row

        for item in items:
            dest_short = item["_dest_wh"].replace("_РФЦ", "").replace("_МРФЦ", "")
            ws3.cell(row, 1, src_short)
            ws3.cell(row, 2, dest_short).font = bold_font
            ws3.cell(row, 3, item["offer_id"]).font = bold_font
            ws3.cell(row, 4, item.get("name", "")).alignment = wrap
            ws3.cell(row, 5, item["qty"]).number_format = num_fmt
            ws3.cell(row, 6, round(item["daily_sales_cd"], 1)).number_format = "0.0"
            ws3.cell(row, 7, item.get("source_excess", 0)).number_format = num_fmt
            ws3.cell(row, 8, round(item.get("transfer_cost", 0))).number_format = rub_fmt
            ws3.cell(row, 9, round(item["est_cd_cost_monthly"])).number_format = rub_fmt
            # payback
            cd_cost = item["est_cd_cost_monthly"]
            tr_cost = item.get("transfer_cost", 0)
            payback = f"~{max(1, round(tr_cost / max(cd_cost, 1)))} мес" if cd_cost > 0 else "—"
            ws3.cell(row, 10, payback).alignment = center
            ws3.cell(row, 11, item.get("reason", "")).alignment = wrap
            ws3.cell(row, 11).font = reason_font

            for ci in range(1, 12):
                ws3.cell(row, ci).border = border
            row += 1

        # Merge source warehouse cells
        if row - merge_start > 1:
            ws3.merge_cells(start_row=merge_start, start_column=1, end_row=row - 1, end_column=1)
        ws3.cell(merge_start, 1).font = wh_font
        ws3.cell(merge_start, 1).fill = PatternFill("solid", fgColor="E8D5F5")
        ws3.cell(merge_start, 1).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

        # Subtotal row
        total_qty = sum(i["qty"] for i in items)
        total_tr = sum(i.get("transfer_cost", 0) for i in items)
        total_cd = sum(i["est_cd_cost_monthly"] for i in items)
        ws3.cell(row, 1, "").fill = PatternFill("solid", fgColor="F3E5F5")
        ws3.cell(row, 2, f"ИТОГО из {src_short}").font = bold_font
        ws3.cell(row, 2).fill = PatternFill("solid", fgColor="F3E5F5")
        ws3.cell(row, 3, f"{len(items)} SKU").fill = PatternFill("solid", fgColor="F3E5F5")
        ws3.cell(row, 5, total_qty).number_format = num_fmt
        ws3.cell(row, 5).font = bold14
        ws3.cell(row, 5).fill = PatternFill("solid", fgColor="F3E5F5")
        ws3.cell(row, 8, round(total_tr)).number_format = rub_fmt
        ws3.cell(row, 8).font = bold_font
        ws3.cell(row, 8).fill = PatternFill("solid", fgColor="F3E5F5")
        ws3.cell(row, 9, round(total_cd)).number_format = rub_fmt
        ws3.cell(row, 9).font = bold_font
        ws3.cell(row, 9).fill = PatternFill("solid", fgColor="F3E5F5")
        for ci in range(1, 12):
            ws3.cell(row, ci).border = border
            if ci not in (1, 2, 3, 5, 8, 9):
                ws3.cell(row, ci).fill = PatternFill("solid", fgColor="F3E5F5")
        row += 1
        row += 1

    # ── Save to buffer ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"distribution_plan_shop{shop_id}_{period}d.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ══════════════════════════════════════════════════════════════
# WB Warehouse Analytics
# ══════════════════════════════════════════════════════════════

@router.get("/wb/analytics")
async def wb_warehouse_analytics(
    shop_id: int = Query(...),
    period: int = Query(30, description="Period in days for sales/cost calculation"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Warehouse analytics for WB shops.

    Returns:
    - KPI summary (warehouses, stock, turnover, costs, cross%)
    - Per-warehouse breakdown (stock, sales, turnover, cross%, logistics, tariffs)
    - Cross-map: warehouse × federal_okrug matrix
    - Costs breakdown by operation type
    - Top storage SKUs by estimated cost
    - Optimization recommendations
    """
    from app.core.clickhouse import get_clickhouse_client

    # Verify shop ownership
    shop = await db.get(Shop, shop_id)
    if not shop or shop.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Shop not found")
    if shop.marketplace != "wildberries":
        raise HTTPException(status_code=400, detail="Only WB shops supported")

    ch = get_clickhouse_client()
    today = date.today()
    d_start = today - timedelta(days=period)

    # ── 1. Current stocks per warehouse ──────────────────────
    stock_rows = ch.query(f"""
        SELECT
            warehouse_name,
            nm_id,
            argMax(quantity, fetched_at)   AS qty
        FROM mms_analytics.fact_inventory_snapshot
        WHERE shop_id = {{shop_id:UInt32}}
          AND warehouse_name NOT LIKE 'FBS:%'
        GROUP BY warehouse_name, nm_id
        HAVING qty > 0
    """, parameters={"shop_id": shop_id}).result_rows

    # Aggregate to warehouse level
    wh_stocks: dict[str, dict] = {}  # warehouse → {nm_ids, qtys, sku_count, total}
    for row in stock_rows:
        wh = row[0]
        nm_id = row[1]
        qty = row[2]
        if wh not in wh_stocks:
            wh_stocks[wh] = {"nm_ids": [], "qtys": {}, "sku_count": 0, "total_qty": 0}
        wh_stocks[wh]["nm_ids"].append(nm_id)
        wh_stocks[wh]["qtys"][nm_id] = qty
        wh_stocks[wh]["sku_count"] += 1
        wh_stocks[wh]["total_qty"] += qty

    # ── 2. Orders per warehouse × okrug (last N days) ────────
    order_rows = ch.query(f"""
        SELECT
            warehouse_name,
            oblast_okrug_name,
            count()                                AS order_count,
            sum(toFloat64(finished_price))         AS revenue,
            groupArray(nm_id)                      AS nm_ids_arr
        FROM mms_analytics.fact_orders_raw
        WHERE shop_id = {{shop_id:UInt32}}
          AND date >= {{d_start:Date}}
          AND is_cancel = 0
        GROUP BY warehouse_name, oblast_okrug_name
    """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows

    # Aggregate orders
    wh_orders: dict[str, dict] = {}  # warehouse → {total, revenue, okrug_detail: {okrug → {cnt, rev}}}
    for row in order_rows:
        wh, okrug, cnt, rev = row[0], row[1], row[2], row[3]
        if wh not in wh_orders:
            wh_orders[wh] = {"total": 0, "revenue": 0.0, "okrug_detail": {}, "nm_ids": set()}
        wh_orders[wh]["total"] += cnt
        wh_orders[wh]["revenue"] += rev
        wh_orders[wh]["okrug_detail"][okrug] = {"count": cnt, "revenue": rev}
        for nm in row[4]:
            wh_orders[wh]["nm_ids"].add(nm)

    # ── 3. Per-SKU orders per warehouse + okrug ────────────────
    sku_order_rows = ch.query(f"""
        SELECT
            warehouse_name,
            nm_id,
            oblast_okrug_name,
            count()                        AS orders,
            sum(toFloat64(finished_price))  AS revenue
        FROM mms_analytics.fact_orders_raw
        WHERE shop_id = {{shop_id:UInt32}}
          AND date >= {{d_start:Date}}
          AND is_cancel = 0
        GROUP BY warehouse_name, nm_id, oblast_okrug_name
    """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows

    # Aggregate: wh → nm_id → {orders, revenue, okrug_detail: {okrug → {count, revenue}}}
    wh_sku_orders: dict[str, dict[int, dict]] = {}
    for row in sku_order_rows:
        wh, nm, okrug, cnt, rev = row[0], row[1], row[2], row[3], row[4]
        sku_data = wh_sku_orders.setdefault(wh, {}).setdefault(nm, {"orders": 0, "revenue": 0.0, "okrug_detail": {}})
        sku_data["orders"] += cnt
        sku_data["revenue"] += rev
        sku_data["okrug_detail"][okrug] = {"count": cnt, "revenue": rev}


    # ── 4. Costs from fact_finances ──────────────────────────
    cost_rows = ch.query(f"""
        SELECT
            CASE
                WHEN operation_type = 'Удержание' AND JSONExtractString(raw_payload, 'bonus_type_name') LIKE '%WB Продвижение%'
                    THEN '__SKIP__'
                WHEN operation_type = 'Удержание' AND JSONExtractString(raw_payload, 'bonus_type_name') LIKE 'Списание за отзыв%'
                    THEN 'Списание за отзыв'
                WHEN operation_type = 'Удержание'
                    THEN 'Удержания'
                ELSE operation_type
            END                            AS op_type_key,
            warehouse_name,
            count()                    AS cnt,
            sum(logistics_total)       AS logistics,
            sum(storage_fee)           AS storage,
            sum(acceptance_fee)        AS acceptance,
            sum(penalty_total)         AS penalty,
            sum(payout_amount)         AS payout
        FROM mms_analytics.fact_finances
        WHERE shop_id = {{shop_id:UInt32}}
          AND event_date >= {{d_start:Date}}
        GROUP BY op_type_key, warehouse_name
    """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows

    # Aggregate costs by type and by warehouse
    cost_by_type: dict[str, dict] = {}
    cost_by_wh: dict[str, dict] = {}
    for row in cost_rows:
        op_type, wh_name = row[0], row[1]
        cnt, logistics, storage, acceptance, penalty, payout = row[2], row[3], row[4], row[5], row[6], row[7]

        # By type
        if op_type not in cost_by_type:
            cost_by_type[op_type] = {"count": 0, "logistics": 0, "storage": 0, "acceptance": 0, "penalty": 0, "payout": 0}
        cost_by_type[op_type]["count"] += cnt
        cost_by_type[op_type]["logistics"] += logistics
        cost_by_type[op_type]["storage"] += storage
        cost_by_type[op_type]["acceptance"] += acceptance
        cost_by_type[op_type]["penalty"] += penalty
        cost_by_type[op_type]["payout"] += payout

        # By warehouse (only for logistics — storage is not attributed)
        if wh_name and logistics:
            if wh_name not in cost_by_wh:
                cost_by_wh[wh_name] = {"logistics": 0, "logistics_cnt": 0}
            cost_by_wh[wh_name]["logistics"] += logistics
            cost_by_wh[wh_name]["logistics_cnt"] += cnt

    # ── 4a. PREVIOUS period costs (for trend comparison) ─────
    prev_d_start = d_start - timedelta(days=period)
    prev_d_end = d_start
    prev_cost_by_type: dict[str, dict] = {}
    prev_total_orders = 0
    try:
        prev_cost_rows = ch.query(f"""
            SELECT
                CASE
                    WHEN operation_type = 'Удержание' AND JSONExtractString(raw_payload, 'bonus_type_name') LIKE '%WB Продвижение%'
                        THEN '__SKIP__'
                    WHEN operation_type = 'Удержание' AND JSONExtractString(raw_payload, 'bonus_type_name') LIKE 'Списание за отзыв%'
                        THEN 'Списание за отзыв'
                    WHEN operation_type = 'Удержание'
                        THEN 'Удержания'
                    ELSE operation_type
                END                            AS op_type_key,
                count()                    AS cnt,
                sum(logistics_total)       AS logistics,
                sum(storage_fee)           AS storage,
                sum(acceptance_fee)        AS acceptance,
                sum(penalty_total)         AS penalty
            FROM mms_analytics.fact_finances
            WHERE shop_id = {{shop_id:UInt32}}
              AND event_date >= {{prev_d_start:Date}}
              AND event_date < {{prev_d_end:Date}}
            GROUP BY op_type_key
        """, parameters={"shop_id": shop_id, "prev_d_start": prev_d_start, "prev_d_end": prev_d_end}).result_rows
        for row in prev_cost_rows:
            op_type = row[0]
            if op_type == '__SKIP__':
                continue
            prev_cost_by_type[op_type] = {
                "count": row[1], "logistics": row[2], "storage": row[3],
                "acceptance": row[4], "penalty": row[5],
            }

        # Previous period orders count
        prev_orders_row = ch.query(f"""
            SELECT count() FROM mms_analytics.fact_orders_raw
            WHERE shop_id = {{shop_id:UInt32}}
              AND date >= {{prev_d_start:Date}} AND date < {{prev_d_end:Date}}
              AND is_cancel = 0
        """, parameters={"shop_id": shop_id, "prev_d_start": prev_d_start, "prev_d_end": prev_d_end}).result_rows
        prev_total_orders = int(prev_orders_row[0][0]) if prev_orders_row else 0
    except Exception as e:
        logger.warning("Prev-period costs query failed: %s", e)

    # ── 4b. Penalty details (breakdown by reason) ────────────
    penalty_details: list[dict] = []
    try:
        pen_rows = ch.query(f"""
            SELECT
                JSONExtractString(raw_payload, 'bonus_type_name') AS reason,
                count()                                           AS cnt,
                sum(abs(penalty_total))                           AS amount
            FROM mms_analytics.fact_finances
            WHERE shop_id = {{shop_id:UInt32}}
              AND event_date >= {{d_start:Date}}
              AND operation_type = 'Штраф'
              AND penalty_total != 0
            GROUP BY reason
            ORDER BY amount DESC
        """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows
        for r in pen_rows:
            reason = r[0] or "Без указания причины"
            penalty_details.append({"reason": reason, "count": int(r[1]), "amount": round(float(r[2]), 2)})
    except Exception as e:
        logger.warning("Penalty details query failed: %s", e)

    # ── 5. Tariffs ───────────────────────────────────────────
    tariffs: dict[str, dict] = {}
    try:
        tariff_rows = ch.query(f"""
            SELECT
                warehouse_name,
                argMax(storage_coef, updated_at)            AS storage_coef,
                argMax(storage_base_liter, updated_at)      AS storage_base_liter,
                argMax(coefficient, updated_at)             AS acceptance_coef,
                argMax(allow_unload, updated_at)            AS allow_unload
            FROM mms_analytics.fact_wb_acceptance_tariffs
            WHERE box_type_id = 2
            GROUP BY warehouse_name
        """).result_rows
        for r_t in tariff_rows:
            tariffs[r_t[0]] = {
                "storage_coef": _parse_ru_float(r_t[1]),
                "storage_base_liter": _parse_ru_float(r_t[2]),
                "acceptance_coef": float(r_t[3]),
                "allow_unload": int(r_t[4]),
            }
    except Exception:
        tariffs = {}

    # ── 6. Product info (dim_products) ───────────────────────
    all_nm_ids: set[int] = set()
    for wh_data in wh_stocks.values():
        all_nm_ids.update(wh_data["nm_ids"])
    # Also include nm_ids from orders (for OOS items that have stock=0 but orders>0)
    for wh_sku_data in wh_sku_orders.values():
        all_nm_ids.update(wh_sku_data.keys())

    products_map: dict[int, dict] = {}
    if all_nm_ids:
        nm_list = ", ".join(str(x) for x in all_nm_ids)
        pg_rows = (await db.execute(
            text(f"""
                SELECT nm_id, vendor_code, name, category,
                       COALESCE(length, 0), COALESCE(width, 0), COALESCE(height, 0)
                FROM dim_products
                WHERE shop_id = :sid AND nm_id IN ({nm_list})
            """),
            {"sid": shop_id},
        )).fetchall()

        for r in pg_rows:
            l, w, h = float(r[4]), float(r[5]), float(r[6])
            vol_liters = (l * w * h) / 1000.0
            if vol_liters > 10000:
                vol_liters = (l * w * h) / 1_000_000.0
            vol_liters = max(vol_liters, 0.1)
            products_map[r[0]] = {
                "nm_id": r[0],
                "vendor_code": r[1] or "",
                "name": r[2] or "",
                "category": r[3] or "",
                "vol_liters": vol_liters,
            }

    # ── 7. Build warehouse list ──────────────────────────────
    all_wh_names = set(wh_stocks.keys()) | set(wh_orders.keys())
    total_orders = sum(d["total"] for d in wh_orders.values())

    warehouses_result = []
    total_cross_orders = 0
    _all_okrugs_seen: set[str] = set()

    for wh_name in sorted(all_wh_names):
        stk = wh_stocks.get(wh_name, {"nm_ids": [], "qtys": {}, "sku_count": 0, "total_qty": 0})
        ords = wh_orders.get(wh_name, {"total": 0, "revenue": 0.0, "okrug_detail": {}})
        costs_wh = cost_by_wh.get(wh_name, {"logistics": 0, "logistics_cnt": 0})

        stock_total = stk["total_qty"]
        orders_total = ords["total"]
        daily_sales = orders_total / period if period > 0 else 0
        turnover = stock_total / daily_sales if daily_sales > 0 else None
        pct_sales = round(orders_total / total_orders * 100, 1) if total_orders > 0 else 0

        # Determine okrug from WAREHOUSE_TO_OKRUG
        wh_okrug = WAREHOUSE_TO_OKRUG.get(wh_name, "")

        # Cross analysis
        okrug_detail = ords.get("okrug_detail", {})
        local_orders = okrug_detail.get(wh_okrug, {}).get("count", 0) if wh_okrug else 0
        cross_orders = orders_total - local_orders
        cross_pct = round(cross_orders / orders_total * 100, 1) if orders_total > 0 else 0
        total_cross_orders += cross_orders

        for okrug_name in okrug_detail:
            _all_okrugs_seen.add(okrug_name)

        # Warehouse type
        if ": Питание" in wh_name or ":Питание" in wh_name:
            wh_type = "food"
        elif "СГТ" in wh_name:
            wh_type = "sgt"
        else:
            wh_type = "normal"

        # Tariff info
        t = tariffs.get(wh_name, {})
        storage_coef = t.get("storage_coef", 0)
        accept_coef = t.get("acceptance_coef", 0)
        accept_label = "Без коэфф." if accept_coef == 0 else ("Закрыт" if accept_coef < 0 else f"x{accept_coef:.0f}")

        # Status
        if stock_total == 0 and orders_total > 0:
            wh_status = "empty"
        elif turnover is not None and turnover < 14:
            wh_status = "critical"
        elif turnover is not None and turnover < 30:
            wh_status = "attention"
        elif turnover is not None and turnover > 120:
            wh_status = "overstocked"
        else:
            wh_status = "ok"

        # Per-SKU details for this warehouse
        skus_detail = []
        wh_sku_data = wh_sku_orders.get(wh_name, {})
        # Combine nm_ids from stock AND orders (FBS has no stock but has orders)
        all_sku_nm_ids = set(stk["nm_ids"]) | set(wh_sku_data.keys())
        for nm_id in all_sku_nm_ids:
            qty = stk["qtys"].get(nm_id, 0)
            sku_ords = wh_sku_data.get(nm_id, {"orders": 0, "revenue": 0.0, "okrug_detail": {}})
            sku_total_orders = sku_ords["orders"]
            sku_daily = sku_total_orders / period if period > 0 else 0
            sku_days = qty / sku_daily if sku_daily > 0 else None
            prod = products_map.get(nm_id, {})

            # Per-SKU geography from this warehouse
            sku_geo = []
            sku_cross_orders = 0
            sku_okrug_detail = sku_ords.get("okrug_detail", {})
            for ok_name, ok_data in sorted(sku_okrug_detail.items(), key=lambda x: x[1]["count"], reverse=True):
                ok_is_local = ok_name == wh_okrug
                ok_share = round(ok_data["count"] / sku_total_orders * 100, 1) if sku_total_orders > 0 else 0
                if not ok_is_local:
                    sku_cross_orders += ok_data["count"]
                sku_geo.append({
                    "okrug": ok_name,
                    "orders": ok_data["count"],
                    "share": ok_share,
                    "is_local": ok_is_local,
                })
            sku_cross_pct = round(sku_cross_orders / sku_total_orders * 100, 1) if sku_total_orders > 0 else 0

            skus_detail.append({
                "nm_id": nm_id,
                "vendor_code": prod.get("vendor_code", ""),
                "name": prod.get("name", ""),
                "stock": qty,
                "daily_sales": round(sku_daily, 2),
                "days_supply": round(sku_days, 1) if sku_days is not None else None,
                "orders": sku_total_orders,
                "cross_orders": sku_cross_orders,
                "cross_pct": sku_cross_pct,
                "geography": sku_geo,
            })
        skus_detail.sort(key=lambda x: x["orders"], reverse=True)

        # Geography detail
        geography = []
        for okrug_name, okrug_data in sorted(okrug_detail.items(), key=lambda x: x[1]["count"], reverse=True):
            is_local = okrug_name == wh_okrug
            share = round(okrug_data["count"] / orders_total * 100, 1) if orders_total > 0 else 0
            geography.append({
                "okrug": okrug_name,
                "orders": okrug_data["count"],
                "share": share,
                "is_local": is_local,
            })

        warehouses_result.append({
            "warehouse_name": wh_name,
            "okrug": wh_okrug,
            "warehouse_type": wh_type,
            "status": wh_status,
            "stock": stock_total,
            "sku_count": stk["sku_count"],
            "orders": orders_total,
            "revenue": round(ords["revenue"], 2),
            "daily_sales": round(daily_sales, 2),
            "turnover_days": round(turnover, 1) if turnover is not None else None,
            "pct_of_total_sales": pct_sales,
            "cross_pct": cross_pct,
            "cross_orders": cross_orders,
            "local_orders": local_orders,
            "logistics_cost": round(costs_wh["logistics"], 2),
            "logistics_count": costs_wh["logistics_cnt"],
            "storage_coef": storage_coef,
            "acceptance_coef": accept_coef,
            "acceptance": accept_label,
            "skus": skus_detail[:50],  # top 50
            "geography": geography,
        })

    warehouses_result.sort(key=lambda x: x["orders"], reverse=True)

    # ── 8. Cross-map: warehouse × okrug matrix ───────────────
    # Include ALL warehouses with orders, not just those with stock
    _wh_in_result = {w["warehouse_name"] for w in warehouses_result}

    # Collect extra okrugs from wh_orders (warehouses with orders but no stock)
    for wh_name_o, wh_data_o in wh_orders.items():
        for okrug_name_o in wh_data_o.get("okrug_detail", {}):
            _all_okrugs_seen.add(okrug_name_o)

    okrug_list = sorted(_all_okrugs_seen)
    cross_map = []

    # 1) Warehouses from stocks (with stock data)
    for wh_data in warehouses_result:
        if wh_data["orders"] == 0:
            continue
        okrug_detail = wh_orders.get(wh_data["warehouse_name"], {}).get("okrug_detail", {})
        row_data = {
            "warehouse": wh_data["warehouse_name"],
            "home_okrug": wh_data["okrug"],
            "total_orders": wh_data["orders"],
            "okrugs": {},
        }
        for okrug_name in okrug_list:
            cnt = okrug_detail.get(okrug_name, {}).get("count", 0)
            is_local = okrug_name == wh_data["okrug"]
            row_data["okrugs"][okrug_name] = {
                "count": cnt,
                "is_local": is_local,
            }
        cross_map.append(row_data)

    # 2) Extra warehouses from wh_orders (orders exist, no stock)
    for wh_name_extra, wh_data_extra in wh_orders.items():
        if wh_name_extra in _wh_in_result:
            continue
        if wh_data_extra["total"] == 0:
            continue
        wh_okrug_extra = WAREHOUSE_TO_OKRUG.get(wh_name_extra, "")
        okrug_detail_extra = wh_data_extra.get("okrug_detail", {})
        row_data = {
            "warehouse": wh_name_extra,
            "home_okrug": wh_okrug_extra,
            "total_orders": wh_data_extra["total"],
            "okrugs": {},
        }
        for okrug_name in okrug_list:
            cnt = okrug_detail_extra.get(okrug_name, {}).get("count", 0)
            is_local = okrug_name == wh_okrug_extra
            row_data["okrugs"][okrug_name] = {
                "count": cnt,
                "is_local": is_local,
            }
        cross_map.append(row_data)

    # Sort by total orders desc
    cross_map.sort(key=lambda x: x["total_orders"], reverse=True)

    # ── 9. Costs summary ─────────────────────────────────────
    costs_summary = []
    _type_map = {
        "Логистика": {"icon": "truck", "label": "Логистика"},
        "Хранение": {"icon": "package", "label": "Хранение"},
        "Штраф": {"icon": "alert", "label": "Штрафы"},
        "Удержания": {"icon": "ban", "label": "Удержания"},
        "Списание за отзыв": {"icon": "message-circle", "label": "Списание за отзыв"},
        "Обработка товара": {"icon": "factory", "label": "Приёмка"},
        "Возмещение издержек по перевозке/по складским операциям с товаром": {"icon": "circle", "label": "Возмещение"},
        "Возмещение за выдачу и возврат товаров на ПВЗ": {"icon": "circle", "label": "Возмещение ПВЗ"},
        "Компенсация скидки по программе лояльности": {"icon": "circle", "label": "Компенсация СПП"},
        "Добровольная компенсация при возврате": {"icon": "circle", "label": "Компенсация возвратов"},
    }
    for op_type, data in sorted(cost_by_type.items(), key=lambda x: abs(x[1].get("logistics", 0) + x[1].get("storage", 0) + x[1].get("penalty", 0) + x[1].get("acceptance", 0)), reverse=True):
        if op_type == '__SKIP__':
            continue
        amount = data["logistics"] + data["storage"] + data["acceptance"] + data["penalty"]
        if abs(amount) < 1:
            continue
        meta = _type_map.get(op_type, {"icon": "circle", "label": op_type})
        costs_summary.append({
            "operation_type": op_type,
            "label": meta["label"],
            "icon": meta["icon"],
            "count": data["count"],
            "amount": round(amount, 2),
        })

    # ── 10. Actual paid storage from fact_wb_paid_storage ─────
    #  Fetch REAL per-SKU per-warehouse storage costs from WB API data
    #  Fallback to tariff-based estimates if no paid storage data available

    # 10a. Query actual paid storage
    actual_storage_by_nm: dict[int, dict[str, float]] = {}  # nm_id → {warehouse → cost_period}
    actual_storage_by_wh: dict[str, float] = {}  # warehouse → total cost for period
    total_storage_actual_all = 0.0
    try:
        ps_rows = ch.query("""
            SELECT nm_id, warehouse,
                   SUM(warehouse_price) AS cost_period
            FROM mms_analytics.fact_wb_paid_storage FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND dt >= {d_start:Date}
              AND dt < today()
            GROUP BY nm_id, warehouse
            HAVING cost_period != 0
        """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows
        for r in ps_rows:
            nm_id_ps, wh_ps, cost_period = int(r[0]), r[1], float(r[2])
            actual_storage_by_nm.setdefault(nm_id_ps, {})[wh_ps] = cost_period
            actual_storage_by_wh[wh_ps] = actual_storage_by_wh.get(wh_ps, 0) + cost_period
            total_storage_actual_all += cost_period
        logger.info("WB analytics: loaded actual storage for %d nm_ids, %d warehouses from fact_wb_paid_storage",
                    len(actual_storage_by_nm), len(actual_storage_by_wh))
    except Exception as e:
        logger.warning("Actual paid storage query failed (will use tariff estimates): %s", e)

    has_actual_storage = len(actual_storage_by_nm) > 0
    month_mult = 30 / period if period > 0 else 1.0  # to extrapolate period → 30 days

    # 10a2. Per-SKU daily cost from last 7 days (for forecast)
    daily_cost_per_nm: dict[int, float] = {}  # nm_id → avg daily cost (₽/day)
    try:
        dc_rows = ch.query("""
            SELECT nm_id,
                   SUM(warehouse_price) / 7 AS daily_cost
            FROM mms_analytics.fact_wb_paid_storage FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND dt >= today() - 7
            GROUP BY nm_id
            HAVING daily_cost > 0
        """, parameters={"shop_id": shop_id}).result_rows
        for r in dc_rows:
            daily_cost_per_nm[int(r[0])] = float(r[1])
        logger.info("WB analytics: loaded daily storage cost for %d nm_ids (last 7d)",
                    len(daily_cost_per_nm))
    except Exception as e:
        logger.warning("Daily storage cost query failed: %s", e)

    # 10a3. Per-SKU total orders (for daily_sales in forecast)
    nm_orders_total: dict[int, int] = {}  # nm_id → total orders in period
    for wh_name, wh_sku_data in wh_sku_orders.items():
        for nm_id, sku_data in wh_sku_data.items():
            nm_orders_total[nm_id] = nm_orders_total.get(nm_id, 0) + sku_data["orders"]

    # 10a4. Per-SKU total stock
    nm_stock_total: dict[int, int] = {}  # nm_id → total stock
    for wh_name, stk_data in wh_stocks.items():
        for nm_id, qty in stk_data["qtys"].items():
            nm_stock_total[nm_id] = nm_stock_total.get(nm_id, 0) + qty

    # 10b. Add storage_cost_actual to each warehouse result
    for wh_data in warehouses_result:
        wh_name = wh_data["warehouse_name"]
        actual_period = actual_storage_by_wh.get(wh_name, 0)
        wh_data["storage_cost_actual"] = round(actual_period, 2)
        wh_data["storage_cost_month"] = round(actual_period * month_mult, 2)

    # 10c. Build storage SKU list with actual data (fallback to tariff estimate)
    sku_storage: dict[int, dict] = {}
    for wh_name, stk_data in wh_stocks.items():
        t = tariffs.get(wh_name, {})
        stor_base = t.get("storage_base_liter", 0)
        stor_coef = t.get("storage_coef", 100)
        coef_mult = stor_coef / 100.0 if stor_coef > 0 else 1.0
        for nm_id in stk_data["nm_ids"]:
            qty = stk_data["qtys"].get(nm_id, 0)
            if qty == 0:
                continue
            prod = products_map.get(nm_id, {})
            vol = prod.get("vol_liters", 1.0)

            # Actual cost from paid storage (for this nm_id + warehouse)
            actual_cost_period = actual_storage_by_nm.get(nm_id, {}).get(wh_name)
            if actual_cost_period is not None:
                cost_month = actual_cost_period * month_mult
                source = "actual"
            elif has_actual_storage:
                # Shop has actual paid storage data but this SKU/warehouse combo is missing
                # → likely zero cost (not charged). Don't use tariff fallback.
                cost_month = 0
                source = "actual"
            else:
                # Tariff-based fallback (no paid storage data at all)
                cost_day = stor_base * vol * qty * coef_mult
                cost_month = cost_day * 30
                source = "estimated"

            # Calculate cost for the actual selected period (no extrapolation)
            if actual_cost_period is not None:
                cost_period_actual = actual_cost_period
            elif has_actual_storage:
                cost_period_actual = 0
            else:
                cost_day = stor_base * vol * qty * coef_mult
                cost_period_actual = cost_day * period

            if nm_id not in sku_storage:
                sku_storage[nm_id] = {
                    "nm_id": nm_id,
                    "vendor_code": prod.get("vendor_code", ""),
                    "name": prod.get("name", ""),
                    "vol_liters": vol,
                    "total_stock": 0,
                    "est_cost_month": 0,
                    "est_cost_period": 0,
                    "storage_source": source,  # "actual" if ANY warehouse has actual data
                    "warehouses": [],
                }
            sku_storage[nm_id]["total_stock"] += qty
            sku_storage[nm_id]["est_cost_month"] += cost_month
            sku_storage[nm_id]["est_cost_period"] += cost_period_actual
            if source == "actual":
                sku_storage[nm_id]["storage_source"] = "actual"  # upgrade to actual
            sku_storage[nm_id]["warehouses"].append({
                "warehouse": wh_name,
                "stock": qty,
                "stor_base": stor_base,
                "cost_month": round(cost_month, 2),
                "source": source,
            })

    # 10d. Calculate forecast_30d for each SKU
    # Formula: for each day i in 0..29, remaining_stock = max(0, stock - daily_sales * i)
    # forecast_30d = Σ (cost_per_unit_per_day × remaining_stock)
    total_forecast_30d = 0.0
    for nm_id, sku_data in sku_storage.items():
        stock = sku_data["total_stock"]
        daily_cost = daily_cost_per_nm.get(nm_id, 0)
        orders_period = nm_orders_total.get(nm_id, 0)
        daily_sales = orders_period / period if period > 0 else 0

        if daily_cost > 0 and stock > 0:
            # cost_per_unit_per_day = daily_cost / current_stock
            cost_per_unit = daily_cost / stock
            forecast = 0.0
            for day in range(30):
                remaining = max(0, stock - daily_sales * day)
                if remaining <= 0:
                    break
                forecast += cost_per_unit * remaining
            sku_data["forecast_30d"] = round(forecast, 2)
            sku_data["daily_sales"] = round(daily_sales, 2)
            sku_data["daily_cost"] = round(daily_cost, 2)
            # days_to_sell = how many days until stock runs out
            sku_data["days_to_sell"] = round(stock / daily_sales) if daily_sales > 0 else None
            total_forecast_30d += forecast
        else:
            # No paid storage data → use est_cost_month as fallback
            sku_data["forecast_30d"] = None
            sku_data["daily_sales"] = round((orders_period / period) if period > 0 else 0, 2)
            sku_data["daily_cost"] = None
            sku_data["days_to_sell"] = round(stock / (orders_period / period)) if orders_period > 0 and period > 0 else None

    storage_skus = sorted(sku_storage.values(), key=lambda x: x["est_cost_month"], reverse=True)
    for s in storage_skus:
        s["est_cost_month"] = round(s["est_cost_month"], 2)
        s["est_cost_period"] = round(s.get("est_cost_period", 0), 2)

    # 10e. Active ads in last 3 days (for ad status icon in table)
    ad_active_nm_ids: set[int] = set()
    try:
        ad_3d_rows = ch.query("""
            SELECT DISTINCT nm_id
            FROM mms_analytics.fact_advert_stats_v3 FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND date >= today() - 3
              AND spend > 0
        """, parameters={"shop_id": shop_id}).result_rows
        ad_active_nm_ids = {int(r[0]) for r in ad_3d_rows}
        logger.info("WB analytics storage: %d nm_ids with active ads in last 3d", len(ad_active_nm_ids))
    except Exception as e:
        logger.warning("Ad active query for storage failed: %s", e)

    for s in storage_skus:
        s["has_active_ads"] = s["nm_id"] in ad_active_nm_ids

    # ── 11. Out-of-stock aggregation (across ALL warehouses, no truncation) ──
    # Aggregate stock + daily_sales per nm_id from full wh_stocks + wh_sku_orders
    global_sku_agg: dict[int, dict] = {}  # nm_id → {stock, orders, daily}
    for wh_name, stk_data in wh_stocks.items():
        wh_sku_data = wh_sku_orders.get(wh_name, {})
        for nm_id in stk_data["nm_ids"]:
            qty = stk_data["qtys"].get(nm_id, 0)
            sku_ords = wh_sku_data.get(nm_id, {"orders": 0})
            if nm_id not in global_sku_agg:
                prod = products_map.get(nm_id, {})
                global_sku_agg[nm_id] = {
                    "nm_id": nm_id,
                    "vendor_code": prod.get("vendor_code", ""),
                    "name": prod.get("name", ""),
                    "stock": 0,
                    "orders": 0,
                }
            global_sku_agg[nm_id]["stock"] += qty
            global_sku_agg[nm_id]["orders"] += sku_ords["orders"]

    # Also add items that have orders but NO stock (stock query filters qty > 0)
    for wh_name, sku_data_map in wh_sku_orders.items():
        for nm_id, sku_ords in sku_data_map.items():
            if nm_id not in global_sku_agg:
                prod = products_map.get(nm_id, {})
                global_sku_agg[nm_id] = {
                    "nm_id": nm_id,
                    "vendor_code": prod.get("vendor_code", ""),
                    "name": prod.get("name", ""),
                    "stock": 0,
                    "orders": sku_ords["orders"],
                }

    out_of_stock_skus = []
    for agg in global_sku_agg.values():
        daily = agg["orders"] / period if period > 0 else 0
        if daily > 0:
            if agg["stock"] == 0:
                # Already out-of-stock — most critical
                out_of_stock_skus.append({
                    "vendor_code": agg["vendor_code"],
                    "name": agg["name"],
                    "stock": 0,
                    "daily": round(daily, 1),
                    "days_left": 0,
                })
            elif (agg["stock"] / daily) < 14:
                days_left = round(agg["stock"] / daily)
                out_of_stock_skus.append({
                    "vendor_code": agg["vendor_code"],
                    "name": agg["name"],
                    "stock": agg["stock"],
                    "daily": round(daily, 1),
                    "days_left": days_left,
                })
    out_of_stock_skus.sort(key=lambda x: (x["days_left"], -x["daily"]))

    # ── 11b. Products summary (ALL products aggregated across warehouses) ──
    products_summary = []
    for agg in global_sku_agg.values():
        daily = round(agg["orders"] / period, 2) if period > 0 else 0
        days_supply = round(agg["stock"] / daily) if daily > 0 else None
        products_summary.append({
            "nm_id": agg["nm_id"],
            "vendor_code": agg["vendor_code"],
            "name": agg["name"],
            "stock": agg["stock"],
            "orders": agg["orders"],
            "daily": daily,
            "days_supply": days_supply,
        })
    # Sort: OOS first (stock=0 with sales), then by daily descending
    products_summary.sort(key=lambda x: (
        0 if x["stock"] == 0 and (x["daily"] or 0) > 0 else 1,
        -(x["daily"] or 0),
    ))

    # ── 12. KPI ──────────────────────────────────────────────
    total_stock = sum(w["stock"] for w in warehouses_result)
    total_daily = sum(w["daily_sales"] for w in warehouses_result)
    avg_turnover = total_stock / total_daily if total_daily > 0 else None
    total_logistics = sum(c["amount"] for c in costs_summary if c["operation_type"] == "Логистика")
    total_storage = sum(c["amount"] for c in costs_summary if c["operation_type"] == "Хранение")
    total_penalties = sum(c["amount"] for c in costs_summary if c["operation_type"] == "Штраф")
    cross_pct_global = round(total_cross_orders / total_orders * 100, 1) if total_orders > 0 else 0

    # Total actual storage: from fact_wb_paid_storage, real sum for the selected period
    total_storage_actual = round(total_storage_actual_all, 2) if has_actual_storage else None

    kpi = {
        "total_warehouses": len([w for w in warehouses_result if w["stock"] > 0 or w["orders"] > 0]),
        "total_stock": total_stock,
        "total_sku": len(all_nm_ids),
        "avg_turnover_days": round(avg_turnover, 1) if avg_turnover is not None else None,
        "total_logistics": round(total_logistics, 2),
        "total_storage": round(total_storage, 2),
        "total_storage_actual": total_storage_actual,
        "total_penalties": round(total_penalties, 2),
        "cross_pct": cross_pct_global,
        "total_orders": total_orders,
        "period_days": period,
        "has_actual_storage": has_actual_storage,
        "forecast_30d": round(total_forecast_30d, 2) if total_forecast_30d > 0 else None,
        "out_of_stock_skus": out_of_stock_skus,
        # Previous period comparison
        "prev": {
            "total_logistics": round(prev_cost_by_type.get("Логистика", {}).get("logistics", 0), 2),
            "total_storage": round(prev_cost_by_type.get("Хранение", {}).get("storage", 0), 2),
            "total_penalties": round(prev_cost_by_type.get("Штраф", {}).get("penalty", 0), 2),
            "total_orders": prev_total_orders,
        },
        "penalty_details": penalty_details,
    }

    # ── 12. Recommendations ──────────────────────────────────
    recommendations = []

    # High cross% warehouses
    for wh in warehouses_result:
        if wh["cross_pct"] > 60 and wh["orders"] > 10:
            # Find top cross-okrugs
            top_okrugs = sorted(
                [g for g in wh["geography"] if not g["is_local"]],
                key=lambda x: x["orders"], reverse=True,
            )[:3]
            target = ", ".join(g["okrug"].replace(" федеральный округ", "") for g in top_okrugs)
            recommendations.append({
                "type": "reduce_cross",
                "severity": "high" if wh["cross_pct"] > 70 else "medium",
                "title": f'{wh["warehouse_name"]}: {wh["cross_pct"]}% кросс-отправок',
                "reason": f'Склад {wh["warehouse_name"]} отправляет {wh["cross_pct"]}% заказов в другие округа ({wh["cross_orders"]} из {wh["orders"]}). Это увеличивает КТР и стоимость логистики.',
                "impact": f'Поставка товаров на склады в {target} снизит КТР и удешевит логистику.',
                "action_items": [
                    f'Рассмотрите размещение товара на складах в: {target}',
                    'Это снизит индекс локализации → удешевит логистику до 50%',
                ],
                "warehouse": wh["warehouse_name"],
            })

    # Closed warehouses with demand
    for wh in warehouses_result:
        if wh["acceptance_coef"] < 0 and wh["orders"] > 5:
            recommendations.append({
                "type": "closed_warehouse",
                "severity": "medium",
                "title": f'{wh["warehouse_name"]}: закрыт для приёмки, но есть спрос',
                "reason": f'Склад закрыт для приёмки коробов, но за период было {wh["orders"]} заказов.',
                "action_items": ['Рассмотрите приёмку через монопаллеты (box_type=5)', 'Мониторьте открытие приёмки'],
                "warehouse": wh["warehouse_name"],
            })

    # Overstocked warehouses
    for wh in warehouses_result:
        if wh["status"] == "overstocked" and wh["stock"] > 50:
            recommendations.append({
                "type": "overstock",
                "severity": "medium",
                "title": f'{wh["warehouse_name"]}: перезатарка ({wh["turnover_days"]:.0f} дн)' if wh["turnover_days"] else f'{wh["warehouse_name"]}: нет продаж',
                "reason": f'Оборачиваемость {wh["turnover_days"]:.0f} дней при stock={wh["stock"]}. Высокие расходы на хранение.' if wh["turnover_days"] else f'{wh["warehouse_name"]}: stock={wh["stock"]}, продаж нет.',
                "action_items": ['Снизить поставки на этот склад', 'Рассмотрите вывоз товара'],
                "warehouse": wh["warehouse_name"],
            })

    return {
        "kpi": kpi,
        "warehouses": warehouses_result,
        "products_summary": products_summary,
        "cross_map": cross_map,
        "okrug_list": okrug_list,
        "costs": costs_summary,
        "storage_skus": storage_skus,
        "recommendations": recommendations,
        "period_days": period,
    }


# ═══════════════════════════════════════════════════════════════
# WB Storage — Excel export (Хранение + ИИ-рекомендации)
# ═══════════════════════════════════════════════════════════════

@router.get("/wb/storage/export")
async def export_wb_storage_excel(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download WB storage analysis as formatted Excel workbook with AI recommendations."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    # Verify shop
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "wildberries":
        raise HTTPException(status_code=404, detail="Shop not found")

    # Get analytics data (reuse existing function)
    analytics = await wb_warehouse_analytics(
        shop_id=shop_id, period=period, db=db, current_user=current_user
    )
    storage_skus = analytics["storage_skus"]
    kpi = analytics["kpi"]

    # Try to get AI data from cache
    ai_data = None
    cache_key = f"wb_storage_ai_{shop_id}_{period}"
    if cache_key in _ai_cache:
        cached_ts, cached_data = _ai_cache[cache_key]
        ai_data = cached_data
    logger.info("Storage export: AI cache_key=%s, found=%s", cache_key, ai_data is not None)

    # ── Styles ──────────────────────────────────────────────
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    totals_fill = PatternFill("solid", fgColor="D9E2F3")
    totals_font = Font(bold=True, size=11)
    red_font = Font(bold=True, color="CC0000")
    green_font = Font(bold=True, color="006600")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(bottom=thin, left=thin, right=thin)
    num_fmt = "#,##0"
    money_fmt = '#,##0" ₽"'

    wb = openpyxl.Workbook()

    # ═══ Sheet 1: Хранение по SKU ═══
    ws = wb.active
    ws.title = "Хранение по SKU"

    headers = [
        ("Артикул", 24), ("nm_id", 14), ("Название", 40), ("Объём, л", 10),
        ("Остаток", 10), ("Прод/д", 8), ("Дней до распродажи", 16),
        (f"Хранение/{period}д", 15), ("Прогноз 30д", 14),
        ("Реклама", 10),
    ]
    for ci, (name, w) in enumerate(headers, 1):
        c = ws.cell(1, ci, name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    total_cost = 0
    total_forecast = 0
    total_stock = 0

    for ri, sku in enumerate(storage_skus, 2):
        est_cost = sku.get("est_cost_period", sku.get("est_cost_month", 0))
        forecast = sku.get("forecast_30d", 0) or 0
        total_cost += est_cost
        total_forecast += forecast
        total_stock += sku.get("total_stock", 0)

        has_ads = sku.get("has_active_ads", False)

        vals = [
            sku.get("vendor_code", ""),
            sku.get("nm_id", ""),
            sku.get("name", ""),
            round(sku.get("vol_liters", 0), 1),
            sku.get("total_stock", 0),
            round(sku.get("daily_sales", 0), 1),
            sku.get("days_to_sell"),
            est_cost,
            forecast if forecast else None,
            "Да" if has_ads else "Нет",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v if v is not None else "—")
            c.border = border
            c.alignment = Alignment(horizontal="center" if ci not in (1, 3) else "left")
            if ci == 5:
                c.number_format = num_fmt
            if ci in (8, 9):
                c.number_format = money_fmt
            if ci == 8 and isinstance(v, (int, float)) and v > 500:
                c.font = red_font
            if ci == 10:
                c.font = green_font if has_ads else Font(color="999999")

    # Totals row
    tr = len(storage_skus) + 2
    ws.cell(tr, 1, "ИТОГО").font = totals_font
    ws.cell(tr, 1).fill = totals_fill
    for ci in range(1, len(headers) + 1):
        ws.cell(tr, ci).fill = totals_fill
        ws.cell(tr, ci).font = totals_font
        ws.cell(tr, ci).border = Border(top=Side(style="medium", color="2F5496"), bottom=thin, left=thin, right=thin)
    ws.cell(tr, 5, total_stock).number_format = num_fmt
    ws.cell(tr, 8, round(total_cost)).number_format = money_fmt
    ws.cell(tr, 9, round(total_forecast)).number_format = money_fmt

    # KPI summary at top-right
    kpi_col = len(headers) + 2
    ws.cell(1, kpi_col, "Период").font = Font(bold=True, size=10)
    ws.cell(1, kpi_col + 1, f"{kpi['period_days']}д").font = Font(size=10)
    ws.cell(2, kpi_col, "Склады").font = Font(bold=True, size=10)
    ws.cell(2, kpi_col + 1, kpi['total_warehouses'])
    ws.cell(3, kpi_col, "Всего SKU").font = Font(bold=True, size=10)
    ws.cell(3, kpi_col + 1, kpi['total_sku'])
    ws.column_dimensions[get_column_letter(kpi_col)].width = 14
    ws.column_dimensions[get_column_letter(kpi_col + 1)].width = 12

    # ═══ Sheet 2: Склады (breakdown) ═══
    ws2 = wb.create_sheet("Детализация по складам")
    wh_headers = [
        ("Артикул", 24), ("nm_id", 14), ("Название", 36), ("Склад", 28),
        ("Остаток", 10), ("Баз.тариф", 10), ("Стоим./мес", 14),
    ]
    for ci, (name, w) in enumerate(wh_headers, 1):
        c = ws2.cell(1, ci, name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(wh_headers))}1"

    r2 = 2
    for sku in storage_skus:
        for wh in sku.get("warehouses", []):
            ws2.cell(r2, 1, sku.get("vendor_code", ""))
            ws2.cell(r2, 2, sku.get("nm_id", ""))
            ws2.cell(r2, 3, sku.get("name", ""))
            ws2.cell(r2, 4, wh.get("warehouse_name", wh.get("warehouse", "")))
            ws2.cell(r2, 5, wh.get("stock", 0)).number_format = num_fmt
            ws2.cell(r2, 6, round(wh.get("stor_base", 0), 2))
            ws2.cell(r2, 7, round(wh.get("cost_month", 0))).number_format = money_fmt
            for ci in range(1, len(wh_headers) + 1):
                ws2.cell(r2, ci).border = border
            r2 += 1

    # ═══ Sheet 3: ИИ-рекомендации ═══
    if ai_data:
        ws3 = wb.create_sheet("ИИ-рекомендации")
        ai_hdr_fill = PatternFill("solid", fgColor="7030A0")
        ai_hdr_font = Font(bold=True, size=11, color="FFFFFF")

        # Header info
        ws3.merge_cells("A1:F1")
        c_title = ws3.cell(1, 1, f"ИИ-Анализ хранения — {shop.name}")
        c_title.font = Font(bold=True, size=14, color="7030A0")
        c_title.alignment = Alignment(horizontal="left")

        severity_labels = {"critical": "🔴 Критично", "warning": "🟡 Внимание", "ok": "🟢 Всё ОК"}
        ws3.merge_cells("A2:F2")
        ws3.cell(2, 1, f"Статус: {severity_labels.get(ai_data.get('severity', 'ok'), ai_data.get('severity', ''))}").font = Font(size=11, bold=True)

        ws3.merge_cells("A3:F3")
        diag_text = f"Диагноз: {ai_data.get('diagnosis', '')}"
        ws3.cell(3, 1, diag_text).font = Font(size=11)
        # Auto-height: ~100 chars per line at merged width, 15px per line
        diag_lines = max(2, len(diag_text) // 90 + 1)
        ws3.row_dimensions[3].height = diag_lines * 18
        ws3.cell(3, 1).alignment = Alignment(wrap_text=True, vertical="top")

        analyzed_ts = ai_data.get("analyzed_at", 0)
        if analyzed_ts:
            ws3.cell(4, 1, f"Дата анализа: {datetime.fromtimestamp(analyzed_ts).strftime('%d.%m.%Y %H:%M')}").font = Font(size=10, color="666666")

        # Key metrics
        km = ai_data.get("key_metrics", {})
        if km:
            ws3.cell(4, 4, "Избыток хранения:").font = Font(bold=True, size=10)
            ws3.cell(4, 5, round(km.get("storage_excess", 0))).number_format = money_fmt
            ws3.cell(5, 4, "Потери кросс-логистики:").font = Font(bold=True, size=10)
            ws3.cell(5, 5, round(km.get("cross_logistics_loss", 0))).number_format = money_fmt
            ws3.cell(6, 4, "Убыточных SKU:").font = Font(bold=True, size=10)
            ws3.cell(6, 5, km.get("unprofitable_skus_count", 0))
            ws3.column_dimensions["D"].width = 22
            ws3.column_dimensions["E"].width = 14

        savings = ai_data.get("total_potential_savings", 0)
        if savings:
            ws3.merge_cells("A5:C5")
            ws3.cell(5, 1, f"Потенциальная экономия: {round(savings):,} ₽/мес").font = Font(bold=True, size=11, color="006600")

        # SKU actions — card layout
        sku_actions = ai_data.get("sku_actions", [])
        if sku_actions:
            row = 8

            # Set column widths for card layout
            ws3.column_dimensions["A"].width = 4    # № / marker
            ws3.column_dimensions["B"].width = 30   # Label / артикул
            ws3.column_dimensions["C"].width = 70   # Details / text
            ws3.column_dimensions["D"].width = 16   # Savings / cost
            ws3.column_dimensions["E"].width = 12   # Risk
            ws3.column_dimensions["F"].width = 10   # Recommended mark

            ws3.merge_cells(f"A{row}:F{row}")
            ws3.cell(row, 1, f"Рекомендации по товарам ({len(sku_actions)})").font = Font(bold=True, size=13, color="7030A0")
            row += 2

            risk_labels = {"low": "Низкий", "medium": "Средний", "high": "Высокий"}
            risk_colors = {"low": "006600", "medium": "CC6600", "high": "CC0000"}
            card_border = Border(
                left=Side(style="medium", color="7030A0"),
                right=Side(style="medium", color="7030A0"),
            )
            card_top = Border(
                top=Side(style="medium", color="7030A0"),
                left=Side(style="medium", color="7030A0"),
                right=Side(style="medium", color="7030A0"),
            )
            card_bottom = Border(
                bottom=Side(style="medium", color="7030A0"),
                left=Side(style="medium", color="7030A0"),
                right=Side(style="medium", color="7030A0"),
            )
            card_top_right = Border(
                top=Side(style="medium", color="7030A0"),
                right=Side(style="medium", color="7030A0"),
            )

            for ai_idx, action in enumerate(sku_actions):
                vendor = action.get("vendor_code", "")
                cost_val = action.get("storage_cost_month", action.get("current_storage_cost", 0))
                turnover = action.get("current_turnover_days", 0)
                stock = action.get("stock", 0)
                problem_text = action.get("problem", action.get("diagnosis", ""))
                recommended_idx = action.get("recommended_option", 0)
                options = action.get("options", [])

                # ── Card header: артикул + metrics ──
                ws3.merge_cells(f"A{row}:B{row}")
                c_vc = ws3.cell(row, 1, f"📦  {vendor}")
                c_vc.font = Font(bold=True, size=14, color="7030A0")
                c_vc.alignment = Alignment(vertical="center")
                c_vc.border = card_top

                ws3.cell(row, 2).border = card_top  # merged but set border
                ws3.cell(row, 3, f"Хранение: {round(cost_val):,} ₽/мес").font = Font(bold=True, size=11)
                ws3.cell(row, 3).border = card_top
                ws3.cell(row, 4, f"Оборач.: {turnover} дн").font = Font(size=11, color="666666")
                ws3.cell(row, 4).border = card_top
                ws3.cell(row, 5, f"Остаток: {stock} шт").font = Font(size=11, color="666666")
                ws3.cell(row, 5).border = card_top
                ws3.cell(row, 6).border = card_top_right
                ws3.row_dimensions[row].height = 28
                # Fill header
                hdr_card_fill = PatternFill("solid", fgColor="F3E8FF")  # light purple
                for ci in range(1, 7):
                    ws3.cell(row, ci).fill = hdr_card_fill
                row += 1

                # ── Problem description ──
                ws3.merge_cells(f"B{row}:F{row}")
                ws3.cell(row, 1, "").border = card_border
                c_prob = ws3.cell(row, 2, f"⚠️  {problem_text}")
                c_prob.font = Font(size=10, color="333333")
                c_prob.alignment = Alignment(wrap_text=True, vertical="top")
                c_prob.border = card_border
                # set right border on merged end
                for ci in [3, 4, 5]:
                    ws3.cell(row, ci).border = card_border
                ws3.cell(row, 6).border = Border(right=Side(style="medium", color="7030A0"))
                prob_lines = max(2, len(problem_text) // 80 + 1)
                ws3.row_dimensions[row].height = prob_lines * 16
                row += 1

                # ── Options header ──
                ws3.cell(row, 1, "").border = card_border
                ws3.cell(row, 2, "Вариант действий").font = Font(bold=True, size=10, color="666666")
                ws3.cell(row, 2).border = card_border
                ws3.cell(row, 3, "Описание").font = Font(bold=True, size=10, color="666666")
                ws3.cell(row, 3).border = card_border
                ws3.cell(row, 4, "Экономия").font = Font(bold=True, size=10, color="666666")
                ws3.cell(row, 4).border = card_border
                ws3.cell(row, 5, "Риск").font = Font(bold=True, size=10, color="666666")
                ws3.cell(row, 5).border = card_border
                ws3.cell(row, 6, "").border = Border(right=Side(style="medium", color="7030A0"))
                opt_hdr_fill = PatternFill("solid", fgColor="F5F5F5")
                for ci in range(1, 7):
                    ws3.cell(row, ci).fill = opt_hdr_fill
                ws3.row_dimensions[row].height = 20
                row += 1

                # ── Each option ──
                for oi, opt in enumerate(options):
                    is_rec = oi == recommended_idx
                    is_last_opt = oi == len(options) - 1
                    opt_label = opt.get("label", "")
                    detail_text = opt.get("detail", "")
                    savings = opt.get("expected_savings", 0)
                    risk = opt.get("risk", "")

                    # Marker column
                    marker = "✅" if is_rec else f"  {oi + 1}."
                    ws3.cell(row, 1, marker).alignment = Alignment(horizontal="center", vertical="top")

                    # Option label
                    c_lbl = ws3.cell(row, 2, opt_label)
                    c_lbl.alignment = Alignment(wrap_text=True, vertical="top")

                    # Detail text
                    c_det = ws3.cell(row, 3, detail_text)
                    c_det.alignment = Alignment(wrap_text=True, vertical="top")

                    # Savings
                    c_sav = ws3.cell(row, 4)
                    if savings > 0:
                        c_sav.value = round(savings)
                        c_sav.number_format = money_fmt
                    else:
                        c_sav.value = "—"
                    c_sav.alignment = Alignment(vertical="top")

                    # Risk
                    risk_text = risk_labels.get(risk, risk)
                    c_risk = ws3.cell(row, 5, risk_text)
                    c_risk.font = Font(color=risk_colors.get(risk, "333333"))
                    c_risk.alignment = Alignment(vertical="top")

                    # Recommended badge
                    if is_rec:
                        ws3.cell(row, 6, "Лучший").font = Font(bold=True, size=9, color="006600")
                        ws3.cell(row, 6).alignment = Alignment(horizontal="center", vertical="top")
                        # Highlight the recommended row
                        rec_fill = PatternFill("solid", fgColor="E8F5E9")
                        rec_font = Font(bold=True, size=11, color="1B5E20")
                        c_lbl.font = rec_font
                        for ci in range(1, 7):
                            ws3.cell(row, ci).fill = rec_fill
                    else:
                        ws3.cell(row, 6, "")
                        c_lbl.font = Font(size=10)

                    # Borders
                    opt_border = card_bottom if is_last_opt else card_border
                    for ci in range(1, 6):
                        ws3.cell(row, ci).border = opt_border
                    ws3.cell(row, 6).border = Border(
                        right=Side(style="medium", color="7030A0"),
                        bottom=Side(style="medium", color="7030A0") if is_last_opt else None,
                    )

                    # Row height from detail text
                    det_lines = max(2, len(detail_text) // 60 + 1)
                    ws3.row_dimensions[row].height = max(30, det_lines * 16)
                    row += 1

                # Spacer between cards
                row += 1

        # Transfers
        transfers = ai_data.get("transfers", [])
        if transfers:
            row += 1
            ws3.merge_cells(f"A{row}:F{row}")
            ws3.cell(row, 1, f"Перемещения между складами ({len(transfers)})").font = Font(bold=True, size=12, color="7030A0")
            row += 1
            tr_headers = [("Артикул", 24), ("Со склада", 28), ("Остаток", 10), ("Оставить", 10), ("Куда → кол-во", 50), ("Эффект", 40)]
            for ci, (name, w) in enumerate(tr_headers, 1):
                c = ws3.cell(row, ci, name)
                c.font = ai_hdr_font
                c.fill = ai_hdr_fill
                c.alignment = Alignment(horizontal="center", wrap_text=True)
            row += 1
            for tr_item in transfers:
                ws3.cell(row, 1, tr_item.get("vendor_code", ""))
                ws3.cell(row, 2, tr_item.get("from_warehouse", ""))
                ws3.cell(row, 3, tr_item.get("from_stock", 0)).number_format = num_fmt
                ws3.cell(row, 4, tr_item.get("keep_at_source", 0)).number_format = num_fmt
                dests = tr_item.get("destinations", [])
                dest_text = "; ".join(f'{d["warehouse"]}: {d["qty"]} шт' for d in dests)
                ws3.cell(row, 5, dest_text).alignment = Alignment(wrap_text=True)
                ws3.cell(row, 6, tr_item.get("expected_effect", "")).alignment = Alignment(wrap_text=True)
                ws3.row_dimensions[row].height = 40
                for ci in range(1, 7):
                    ws3.cell(row, ci).border = border
                row += 1

        # General tips
        general_tips = ai_data.get("general_tips", [])
        supply_tip = ai_data.get("supply_tip", "")
        if general_tips or supply_tip:
            row += 1
            ws3.merge_cells(f"A{row}:F{row}")
            ws3.cell(row, 1, "Общие рекомендации").font = Font(bold=True, size=12, color="7030A0")
            row += 1
            if supply_tip:
                ws3.merge_cells(f"A{row}:F{row}")
                ws3.cell(row, 1, f"📦 Поставки: {supply_tip}").font = Font(bold=True, size=11)
                ws3.cell(row, 1).alignment = Alignment(wrap_text=True)
                ws3.row_dimensions[row].height = 40
                row += 1
            for tip in general_tips:
                ws3.merge_cells(f"A{row}:F{row}")
                c_tip = ws3.cell(row, 1, f"• {tip}")
                c_tip.alignment = Alignment(wrap_text=True)
                ws3.row_dimensions[row].height = 40
                row += 1

    # ── Save & return ────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"storage_{shop_id}_{period}d.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ═══════════════════════════════════════════════════════════════
# Ozon Storage — Excel export (Хранение по SKU + ИИ-рекомендации)
# ═══════════════════════════════════════════════════════════════

@router.get("/ozon/storage/export")
async def export_ozon_storage_excel(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download Ozon storage analysis as formatted Excel workbook."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    # Verify shop
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "ozon":
        raise HTTPException(status_code=404, detail="Shop not found")

    # Get analytics data (reuse existing function)
    analytics = await ozon_storage_analytics(
        shop_id=shop_id, period=period, db=db, current_user=current_user
    )
    storage_skus = analytics["storage_skus"]
    kpi = analytics["kpi"]

    # Try to get AI data from cache
    ai_data = None
    cache_key = f"ozon_storage_ai_{shop_id}_{period}"
    if cache_key in _ai_cache:
        cached_ts, cached_data = _ai_cache[cache_key]
        ai_data = cached_data
    logger.info("Ozon Storage export: AI cache_key=%s, found=%s", cache_key, ai_data is not None)

    # ── Styles ──────────────────────────────────────────────
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1A56DB")
    totals_fill = PatternFill("solid", fgColor="DBEAFE")
    totals_font = Font(bold=True, size=11)
    red_font = Font(bold=True, color="CC0000")
    green_font = Font(bold=True, color="006600")
    amber_font = Font(bold=True, color="CC6600")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(bottom=thin, left=thin, right=thin)
    num_fmt = "#,##0"
    money_fmt = '#,##0" ₽"'

    wb = openpyxl.Workbook()

    # ═══ Sheet 1: Хранение по SKU ═══
    ws = wb.active
    ws.title = "Хранение по SKU"

    headers = [
        ("Артикул", 24), ("SKU", 14), ("Название", 40), ("Объём, л", 10),
        ("Остаток", 10), ("Прод/д", 8), ("Дней до распродажи", 16),
        (f"Хранение/{period}д", 15), ("Прогноз 30д", 14),
        ("Оборач., дн", 12), ("Зона", 10), ("Реклама", 10),
    ]
    for ci, (name, w) in enumerate(headers, 1):
        c = ws.cell(1, ci, name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    total_cost = 0
    total_forecast = 0
    total_stock = 0

    zone_labels = {"paid": "Платная", "warning": "Внимание", "free": "Бесплатная"}

    for ri, sku in enumerate(storage_skus, 2):
        est_cost = sku.get("est_cost_month", 0)
        # Scale to selected period (est_cost_month is 30d)
        est_cost_period = est_cost * period / 30 if period != 30 else est_cost
        forecast = sku.get("forecast_30d", 0) or 0
        total_cost += est_cost_period
        total_forecast += forecast
        total_stock += sku.get("total_stock", 0)

        has_ads = sku.get("has_active_ads", False)
        zone = sku.get("zone", "free")
        turnover = sku.get("turnover_days")

        vals = [
            sku.get("vendor_code", sku.get("offer_id", "")),
            sku.get("nm_id", ""),
            sku.get("name", ""),
            round(sku.get("vol_liters", 0), 1),
            sku.get("total_stock", 0),
            round(sku.get("daily_sales", 0), 1),
            sku.get("days_to_sell"),
            round(est_cost_period, 2),
            forecast if forecast else None,
            round(turnover, 1) if turnover is not None else None,
            zone_labels.get(zone, zone),
            "Да" if has_ads else "Нет",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v if v is not None else "—")
            c.border = border
            c.alignment = Alignment(horizontal="center" if ci not in (1, 3) else "left")
            if ci == 5:
                c.number_format = num_fmt
            if ci in (8, 9):
                c.number_format = money_fmt
            if ci == 8 and isinstance(v, (int, float)) and v > 500:
                c.font = red_font
            if ci == 11:
                if zone == "paid":
                    c.font = red_font
                elif zone == "warning":
                    c.font = amber_font
                else:
                    c.font = green_font
            if ci == 12:
                c.font = green_font if has_ads else Font(color="999999")

    # Totals row
    tr = len(storage_skus) + 2
    ws.cell(tr, 1, "ИТОГО").font = totals_font
    ws.cell(tr, 1).fill = totals_fill
    for ci in range(1, len(headers) + 1):
        ws.cell(tr, ci).fill = totals_fill
        ws.cell(tr, ci).font = totals_font
        ws.cell(tr, ci).border = Border(top=Side(style="medium", color="1A56DB"), bottom=thin, left=thin, right=thin)
    ws.cell(tr, 5, total_stock).number_format = num_fmt
    ws.cell(tr, 8, round(total_cost)).number_format = money_fmt
    ws.cell(tr, 9, round(total_forecast)).number_format = money_fmt

    # KPI summary at top-right
    kpi_col = len(headers) + 2
    ws.cell(1, kpi_col, "Период").font = Font(bold=True, size=10)
    ws.cell(1, kpi_col + 1, f"{kpi['period_days']}д").font = Font(size=10)
    ws.cell(2, kpi_col, "Всего SKU").font = Font(bold=True, size=10)
    ws.cell(2, kpi_col + 1, kpi['total_skus'])
    ws.cell(3, kpi_col, "Ср. оборач.").font = Font(bold=True, size=10)
    ws.cell(3, kpi_col + 1, f"{round(kpi['avg_turnover_days'])} дн" if kpi.get('avg_turnover_days') else "—")
    ws.cell(4, kpi_col, "Платная зона").font = Font(bold=True, size=10)
    ws.cell(4, kpi_col + 1, kpi.get('paid_zone_skus', 0))
    ws.cell(5, kpi_col, "Зона внимания").font = Font(bold=True, size=10)
    ws.cell(5, kpi_col + 1, kpi.get('warning_zone_skus', 0))
    ws.column_dimensions[get_column_letter(kpi_col)].width = 14
    ws.column_dimensions[get_column_letter(kpi_col + 1)].width = 12

    # ═══ Sheet 2: Детализация по складам ═══
    ws2 = wb.create_sheet("Детализация по складам")
    wh_headers = [
        ("Артикул", 24), ("SKU", 14), ("Название", 36), ("Склад", 28),
        ("Остаток", 10), ("Резерв", 10), ("Стоим./мес", 14),
    ]
    for ci, (name, w) in enumerate(wh_headers, 1):
        c = ws2.cell(1, ci, name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(wh_headers))}1"

    r2 = 2
    for sku in storage_skus:
        for wh in sku.get("warehouses", []):
            ws2.cell(r2, 1, sku.get("vendor_code", sku.get("offer_id", "")))
            ws2.cell(r2, 2, sku.get("nm_id", ""))
            ws2.cell(r2, 3, sku.get("name", ""))
            ws2.cell(r2, 4, wh.get("warehouse_name", ""))
            ws2.cell(r2, 5, wh.get("stock", 0)).number_format = num_fmt
            ws2.cell(r2, 6, wh.get("reserved", 0)).number_format = num_fmt
            ws2.cell(r2, 7, round(wh.get("cost_month", 0))).number_format = money_fmt
            for ci in range(1, len(wh_headers) + 1):
                ws2.cell(r2, ci).border = border
            r2 += 1

    # ═══ Sheet 3: ИИ-рекомендации (if cached) ═══
    if ai_data:
        ws3 = wb.create_sheet("ИИ-рекомендации")
        ai_hdr_fill = PatternFill("solid", fgColor="1A56DB")
        ai_hdr_font = Font(bold=True, size=11, color="FFFFFF")

        ws3.merge_cells("A1:F1")
        c_title = ws3.cell(1, 1, f"ИИ-Анализ хранения Ozon — {shop.name}")
        c_title.font = Font(bold=True, size=14, color="1A56DB")

        severity_labels = {"critical": "🔴 Критично", "warning": "🟡 Внимание", "ok": "🟢 Всё ОК"}
        ws3.merge_cells("A2:F2")
        ws3.cell(2, 1, f"Статус: {severity_labels.get(ai_data.get('severity', 'ok'), '')}").font = Font(size=11, bold=True)

        ws3.merge_cells("A3:F3")
        diag_text = f"Диагноз: {ai_data.get('diagnosis', '')}"
        ws3.cell(3, 1, diag_text).font = Font(size=11)
        ws3.cell(3, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws3.row_dimensions[3].height = max(36, len(diag_text) // 90 * 18)

        # SKU actions
        sku_actions = ai_data.get("sku_actions", [])
        if sku_actions:
            row = 5
            ws3.merge_cells(f"A{row}:F{row}")
            ws3.cell(row, 1, f"Рекомендации по товарам ({len(sku_actions)})").font = Font(bold=True, size=13, color="1A56DB")
            row += 1

            act_headers = [("Артикул", 24), ("Проблема", 50), ("Хранение/мес", 14), ("Оборач., дн", 12), ("Остаток", 10), ("Рекомендация", 50)]
            for ci, (name, w) in enumerate(act_headers, 1):
                c = ws3.cell(row, ci, name)
                c.font = ai_hdr_font
                c.fill = ai_hdr_fill
                c.alignment = Alignment(horizontal="center", wrap_text=True)
                ws3.column_dimensions[get_column_letter(ci)].width = w
            row += 1

            for action in sku_actions:
                ws3.cell(row, 1, action.get("vendor_code", ""))
                ws3.cell(row, 2, action.get("problem", action.get("diagnosis", ""))).alignment = Alignment(wrap_text=True)
                ws3.cell(row, 3, round(action.get("storage_cost_month", action.get("current_storage_cost", 0)))).number_format = money_fmt
                ws3.cell(row, 4, action.get("current_turnover_days", 0))
                ws3.cell(row, 5, action.get("stock", 0)).number_format = num_fmt
                recommended_idx = action.get("recommended_option", 0)
                options = action.get("options", [])
                rec_text = options[recommended_idx].get("label", "") + ": " + options[recommended_idx].get("detail", "") if recommended_idx < len(options) else ""
                ws3.cell(row, 6, rec_text).alignment = Alignment(wrap_text=True)
                ws3.row_dimensions[row].height = 40
                for ci in range(1, len(act_headers) + 1):
                    ws3.cell(row, ci).border = border
                row += 1

        # General tips
        general_tips = ai_data.get("general_tips", [])
        if general_tips:
            row = row + 1 if sku_actions else 5
            ws3.merge_cells(f"A{row}:F{row}")
            ws3.cell(row, 1, "Общие рекомендации").font = Font(bold=True, size=12, color="1A56DB")
            row += 1
            for tip in general_tips:
                ws3.merge_cells(f"A{row}:F{row}")
                ws3.cell(row, 1, f"• {tip}").alignment = Alignment(wrap_text=True)
                ws3.row_dimensions[row].height = 40
                row += 1

    # ── Save & return ────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"ozon_storage_{shop_id}_{period}d.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ═══════════════════════════════════════════════════════════════
# AI-powered Warehouse Diagnostics
# ═══════════════════════════════════════════════════════════════

import json
import os
import time
import httpx

# Simple in-memory cache: key → (timestamp, data)
_ai_cache: dict[str, tuple[float, dict]] = {}
_AI_CACHE_TTL = 24 * 3600  # 24 hours — analysis once per day

# ═══════════════════════════════════════════════════════════════════
# PROMPT 1: Проблемные SKU — дерево решений с чистой прибылью
# ═══════════════════════════════════════════════════════════════════
_AI_PROMPT_SKU_PROBLEMS = """Ты — эксперт по складской логистике Wildberries.

ЗАДАЧА: для каждого проблемного SKU дай КОНКРЕТНЫЕ варианты действий с финансовыми расчётами.
Ты анализируешь ТОЛЬКО проблемы платного хранения и оборачиваемости. НЕ давай общих советов по рекламе или ценообразованию вне контекста складского хранения.

## ДАННЫЕ, КОТОРЫЕ ТЕБЕ ДАНЫ
Для каждого SKU ты получаешь:
- **net_profit** — ЧИСТАЯ прибыль за период (после ВСЕХ вычетов: комиссия WB, логистика, хранение, удержания, реклама, себестоимость)
- **profit_margin_pct** — маржа чистой прибыли в % от выручки
- **revenue, payout, logistics, storage_fact, deductions, cogs** — компоненты P&L
- **est_storage_month** — оценка хранения в ₽/мес при текущем остатке
- **ad_spend, ad_orders, drr** — рекламные данные (для контекста оборачиваемости)

## ДЕРЕВО РЕШЕНИЙ
Для каждого проблемного SKU (оборачиваемость > 90д ИЛИ хранение > чистой прибыли):

1. Товар БЕЗ рекламы + не продаётся → «Запустить рекламу для ускорения оборачиваемости»
   - Рассчитай: при CPO ~X₽ и чистой марже Y₽, хранение окупится через Z дней
   - Показатель: хранение стоит A₽/мес, реклама поможет продать быстрее

2. Товар С рекламой + не продаётся → «Снизить цену для ускорения распродажи»
   - Рассчитай: при скидке N%, новая чистая маржа M₽, распродажа за K дней
   - Альтернатива: хранение за 60д стоит B₽, что больше/меньше потери от скидки

3. Товар с хранением > чистой прибыли → «Убыточное хранение, действовать срочно»
   - Покажи: хранение X₽/мес, прибыль Y₽/мес → убыток Z₽/мес
   - Варианты: перераспределить / вывезти / распродать со скидкой

4. Вариант «Ничего не делать» — ОБЯЗАТЕЛЕН для каждого SKU:
   - Рассчитай: через 30д хранение +A₽, через 60д +B₽
   - Покажи: при текущих продажах запас закончится через C дней / не закончится никогда

## ФАКТЫ О WB
- Хранение платное С ПЕРВОГО ДНЯ (бесплатный период отменён 05.01.2026)
- Оборачиваемость > 90д — хранение начинает съедать прибыль
- Оборачиваемость > 180д — критическая зона
- Вывоз: бесплатно до 5 коробов, далее платно

## ФОРМАТ ОТВЕТА — СТРОГО JSON:
{
  "severity": "critical" | "warning" | "ok",
  "diagnosis": "Ключевая проблема и размер убытка (1-2 предложения, крупно)",
  "total_potential_savings": 0,
  "key_metrics": {
    "cross_logistics_loss": 0,
    "storage_excess": 0,
    "unprofitable_skus_count": 0
  },
  "analysis_sections": [
    {
      "section": "cross_logistics",
      "severity": "critical" | "warning" | "ok",
      "summary": "Кросс-отправки 58% — основная проблема логистики. Потери ~12 000₽/мес",
      "action_text": "Перейти в Кросс-логистику"
    },
    {
      "section": "storage",
      "severity": "critical" | "warning" | "ok",
      "summary": "2 склада затоварены, 5 SKU с хранением > прибыли. Убыток ~8 000₽/мес",
      "action_text": "Перейти в Хранение"
    },
    {
      "section": "supply",
      "severity": "critical" | "warning" | "ok",
      "summary": "3 SKU заканчиваются < 14 дней на всех складах. Риск out-of-stock",
      "action_text": "Перейти в Поставки"
    },
    {
      "section": "geography",
      "severity": "critical" | "warning" | "ok",
      "summary": "Уральский и Сибирский ФО без покрытия — кросс-доставка 100%",
      "action_text": "Перейти в Географию"
    }
  ],
  "sku_actions": [
    {
      "vendor_code": "АРТИКУЛ",
      "name": "Название",
      "problem": "Оборачиваемость 500д, хранение 987₽/мес, чистая прибыль −200₽/мес → убыток",
      "storage_cost_month": 987,
      "net_profit_month": -200,
      "current_turnover_days": 500,
      "stock": 224,
      "options": [
        {
          "action": "discount",
          "label": "Скидка 25%",
          "detail": "Снизить цену с 1500₽ до 1125₽.",
          "expected_savings": 500,
          "risk": "medium"
        },
        {
          "action": "do_nothing",
          "label": "Оставить как есть",
          "detail": "Хранение 987₽/мес. Через 30д расходы +987₽.",
          "expected_savings": 0,
          "risk": "high"
        }
      ],
      "recommended_option": 0
    }
  ]
}

## ПРАВИЛА
- sku_actions: 5-10 самых проблемных SKU. Критерии (взять ВСЕ подходящие):
  1) Высокие расходы хранения (хранение > 1000₽/мес — ВСЕ такие SKU ОБЯЗАТЕЛЬНО!)
  2) Высокая оборачиваемость (> 90д)
  3) Хранение > чистой прибыли (убыточное хранение)
  4) Нет продаж (0 заказов за период)
  ВАЖНО: товары отсортированы по расходам хранения от наибольших к наименьшим. Первые 5-7 товаров ВСЕГДА должны попасть в sku_actions!
- Для каждого SKU: 2-3 варианта, ОБЯЗАТЕЛЬНО включая "do_nothing"
- vendor_code: ТОЛЬКО vendor_code, НИКОГДА nm_id
- options.action: "discount" | "launch_ads" | "withdraw" | "do_nothing" | "reduce_supply"
- "launch_ads" — ТОЛЬКО если товар БЕЗ рекламы И хранение дорогое. НЕ «оптимизируйте рекламу»
- expected_savings: экономия в ₽/мес (для do_nothing = 0)
- risk: "low" | "medium" | "high"
- recommended_option: индекс лучшего варианта (0-based)
- СЧИТАЙ ОТ ЧИСТОЙ ПРИБЫЛИ (net_profit), НЕ от маржи price−cost
- severity: "critical" если есть убыточные SKU, "warning" если оборач > 90д
- analysis_sections: ОБЯЗАТЕЛЬНО 4 секции (cross_logistics, storage, supply, geography). Каждая с severity + summary (1-2 предложения, конкретно с цифрами)
- Пиши НА РУССКОМ. Каждый текст — 2-3 предложения, конкретно и понятно
"""

# ═══════════════════════════════════════════════════════════════════
# PROMPT 2: Перераспределение по складам — конкретные трансферы
# ═══════════════════════════════════════════════════════════════════
_AI_PROMPT_REDISTRIBUTION = """Ты — эксперт по складской логистике Wildberries.

ЗАДАЧА: проанализируй остатки по складам и дай КОНКРЕТНЫЕ рекомендации по перераспределению.
Не давай общих советов. Указывай: какой SKU, откуда, куда, сколько штук, и почему.

## ФАКТЫ О WB
- Перераспределение (кросс-борд): комиссия +0.5% от ВСЕХ продаж на 60 дней. Лимиты ежедневные
- Кросс-доставка стоит в 1.5-2× дороже доставки «со своего склада»
- Склады с 100% кросс-отправок = весь товар доставляется издалека
- Перераспределение имеет смысл, если экономия на кросс-логистике > стоимости перемещения
- Вывоз: бесплатно до 5 коробов, далее платно

## ЛОГИКА АНАЛИЗА
1. Найди склады с большим остатком + высоким % кросс-отправок → они «кормят» другие регионы
2. Найди SKU, которые лежат только на 1-2 складах, а продаются в разных округах
3. Предложи перемещение: часть стока с «перегруженного» склада → на 2-3 склада ближе к спросу
4. Рассчитай ожидаемый эффект: снижение кросс-% и экономию в ₽

## ФОРМАТ ОТВЕТА — СТРОГО JSON:
{
  "transfers": [
    {
      "vendor_code": "АМ-СОБ-МЕЛ-ТЕЛ-1",
      "name": "Название товара",
      "from_warehouse": "Котовск",
      "from_stock": 2685,
      "keep_at_source": 400,
      "destinations": [
        {
          "warehouse": "Казань",
          "qty": 500,
          "reason": "Приволжский ФО: заказы есть, сток 0. Кросс из Котовска = переплата ~30₽/заказ"
        },
        {
          "warehouse": "Электросталь",
          "qty": 800,
          "reason": "Центральный ФО: основной спрос. Доставка со «своего» склада в 2× дешевле"
        }
      ],
      "expected_effect": "Кросс-логистика снизится с 100% до ~35%. Экономия ~12 000₽/мес"
    }
  ],
  "general_tips": [
    "Описание общей ситуации со складами и приоритетных действий. 2-3 предложения."
  ],
  "supply_tip": "Конкретные рекомендации для следующей поставки. Куда поставлять, каких SKU не хватает."
}

## ПРАВИЛА
- transfers: 2-5 самых импактных перемещений (наибольшая экономия)
- vendor_code: используй ТОЛЬКО vendor_code
- keep_at_source: сколько оставить на исходном складе (90д запас по текущим продажам)
- destinations: 1-3 склада-получателя с конкретными qty и причиной
- expected_effect: экономия в ₽/мес с объяснением
- general_tips: 1-3 совета по общей стратегии складов
- supply_tip: рекомендация по поставке (всегда упомяни раздел «Склады → Поставки»)
- Пиши НА РУССКОМ. Конкретно, с числами, понятным языком
"""


@router.get("/wb/ai-analysis")
async def get_wb_ai_analysis(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    force: bool = Query(False, description="Skip cache"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """AI-powered warehouse analysis using Gemini 2.5 Flash.
    Two parallel requests: SKU problems + redistribution."""
    import asyncio

    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "wildberries":
        raise HTTPException(status_code=404, detail="Shop not found")

    cache_key = f"wb_ai_v2_{shop_id}_{period}"

    if not force and cache_key in _ai_cache:
        ts, cached = _ai_cache[cache_key]
        if time.time() - ts < _AI_CACHE_TTL:
            return {**cached, "cached": True}

    api_key = os.getenv("KIE_AI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="AI API key not configured")

    try:
        from app.core.clickhouse import get_clickhouse_client

        ch = get_clickhouse_client()
        today = date.today()
        d_start = today - timedelta(days=period)

        # ── 1. Warehouse summary (3 simple queries merged in Python) ──
        # 1a. Stock per warehouse
        stock_rows = ch.query("""
            SELECT warehouse_name, sum(qty) AS total_stock
            FROM (
                SELECT warehouse_name, nm_id, argMax(quantity, fetched_at) AS qty
                FROM mms_analytics.fact_inventory_snapshot
                WHERE shop_id = {shop_id:UInt32}
                  AND warehouse_name NOT LIKE 'FBS:%'
                GROUP BY warehouse_name, nm_id
                HAVING qty > 0
            )
            GROUP BY warehouse_name
            ORDER BY total_stock DESC
        """, parameters={"shop_id": shop_id}).result_rows

        stock_by_wh: dict[str, int] = {}
        for r in stock_rows:
            stock_by_wh[r[0]] = int(r[1])

        # 1b. Orders & cross-regional per warehouse
        order_rows = ch.query("""
            SELECT
                warehouse_name,
                count() AS orders,
                countIf(
                    oblast_okrug_name != multiIf(
                        warehouse_name LIKE '%Коледино%' OR warehouse_name LIKE '%Подольск%' OR warehouse_name LIKE '%Электросталь%', 'Центральный',
                        warehouse_name LIKE '%Казань%' OR warehouse_name LIKE '%Набережные%', 'Приволжский',
                        warehouse_name LIKE '%Краснодар%' OR warehouse_name LIKE '%Ростов%', 'Южный',
                        warehouse_name LIKE '%Екатеринбург%', 'Уральский',
                        warehouse_name LIKE '%Новосибирск%', 'Сибирский',
                        warehouse_name LIKE '%Хабаровск%' OR warehouse_name LIKE '%Владивосток%', 'Дальневосточный',
                        warehouse_name LIKE '%Санкт-Петербург%' OR warehouse_name LIKE '%Тверь%', 'Северо-Западный',
                        warehouse_name LIKE '%Воронеж%' OR warehouse_name LIKE '%Котовск%', 'Центральный',
                        'Другой'
                    )
                ) AS cross_orders
            FROM mms_analytics.fact_orders_raw
            WHERE shop_id = {shop_id:UInt32} AND date >= {d_start:Date} AND is_cancel = 0
            GROUP BY warehouse_name
        """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows

        orders_by_wh: dict[str, dict] = {}
        for r in order_rows:
            orders_by_wh[r[0]] = {"orders": int(r[1]), "cross_orders": int(r[2])}

        # 1c. Total stock per nm_id (for turnover calculation)
        stock_per_nm = ch.query("""
            SELECT nm_id, sum(qty) AS total_qty
            FROM (
                SELECT nm_id, argMax(quantity, fetched_at) AS qty
                FROM mms_analytics.fact_inventory_snapshot
                WHERE shop_id = {shop_id:UInt32}
                  AND warehouse_name NOT LIKE 'FBS:%'
                GROUP BY warehouse_name, nm_id
                HAVING qty > 0
            )
            GROUP BY nm_id
        """, parameters={"shop_id": shop_id}).result_rows

        stock_nm_map: dict[int, int] = {}
        for r in stock_per_nm:
            stock_nm_map[int(r[0])] = int(r[1])

        # 1d. Orders per nm_id per warehouse (for turnover calc)
        nm_orders_rows = ch.query("""
            SELECT warehouse_name, nm_id, count() AS order_cnt
            FROM mms_analytics.fact_orders_raw
            WHERE shop_id = {shop_id:UInt32} AND date >= {d_start:Date} AND is_cancel = 0
            GROUP BY warehouse_name, nm_id
        """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows

        # Calculate avg turnover per warehouse
        wh_turnovers: dict[str, list[int]] = {}
        for r in nm_orders_rows:
            wh = r[0]
            nm = int(r[1])
            cnt = int(r[2])
            total_stock_nm = stock_nm_map.get(nm, 0)
            if cnt > 0 and total_stock_nm > 0:
                daily_sales = cnt / period
                turnover = round(total_stock_nm / daily_sales) if daily_sales > 0 else 9999
            else:
                turnover = 9999
            wh_turnovers.setdefault(wh, []).append(min(turnover, 9999))

        # Merge into wh_summary
        all_wh_names = set(stock_by_wh.keys()) | set(orders_by_wh.keys())
        wh_summary = []
        for wh_name in all_wh_names:
            stock = stock_by_wh.get(wh_name, 0)
            od = orders_by_wh.get(wh_name, {"orders": 0, "cross_orders": 0})
            orders = od["orders"]
            cross = od["cross_orders"]
            turnovers = wh_turnovers.get(wh_name, [])
            avg_turnover = round(sum(turnovers) / len(turnovers)) if turnovers else 999
            wh_summary.append({
                "warehouse": wh_name,
                "stock": stock,
                "orders": orders,
                "cross_orders": cross,
                "cross_pct": round(cross / orders * 100) if orders > 0 else 0,
                "avg_turnover": avg_turnover,
            })
        wh_summary.sort(key=lambda x: x["orders"], reverse=True)


        # ── 2. Overall costs ──
        cost_rows_ch = ch.query("""
            SELECT
                sum(abs(wb_delivery_rub)) AS logistics,
                sum(abs(storage_fee)) AS storage,
                sumIf(abs(penalty_total), operation_type != 'Удержание') AS penalties
            FROM mms_analytics.fact_finances FINAL
            WHERE shop_id = {shop_id:UInt32} AND marketplace = 1
              AND event_date >= {d_start:Date} AND event_date <= {d_end:Date}
        """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": today}).result_rows
        costs = {"logistics": 0, "storage": 0, "penalties": 0}
        if cost_rows_ch:
            costs["logistics"] = round(float(cost_rows_ch[0][0] or 0))
            costs["storage"] = round(float(cost_rows_ch[0][1] or 0))
            costs["penalties"] = round(float(cost_rows_ch[0][2] or 0))

        # ── 3. Per-SKU: orders, stock, advertising ──
        sku_rows = ch.query("""
            SELECT
                s.nm_id,
                s.total_stock,
                coalesce(o.orders, 0) AS orders,
                coalesce(o.revenue, 0) AS revenue,
                coalesce(a.spend, 0) AS ad_spend,
                coalesce(a.views, 0) AS ad_views,
                coalesce(a.clicks, 0) AS ad_clicks,
                coalesce(a.ad_orders, 0) AS ad_orders
            FROM (
                SELECT nm_id, sum(qty) AS total_stock
                FROM (
                    SELECT nm_id, argMax(quantity, fetched_at) AS qty
                    FROM mms_analytics.fact_inventory_snapshot
                    WHERE shop_id = {shop_id:UInt32}
                      AND warehouse_name NOT LIKE 'FBS:%'
                    GROUP BY warehouse_name, nm_id
                    HAVING qty > 0
                )
                GROUP BY nm_id
            ) AS s
            LEFT JOIN (
                SELECT nm_id, count() AS orders, sum(toFloat64(price_with_disc)) AS revenue
                FROM mms_analytics.fact_orders_raw
                WHERE shop_id = {shop_id:UInt32} AND date >= {d_start:Date} AND is_cancel = 0
                GROUP BY nm_id
            ) AS o ON o.nm_id = s.nm_id
            LEFT JOIN (
                SELECT nm_id, sum(spend) AS spend, sum(views) AS views,
                       sum(clicks) AS clicks, sum(orders) AS ad_orders
                FROM mms_analytics.fact_advert_stats_v3 FINAL
                WHERE shop_id = {shop_id:UInt32} AND date >= {d_start:Date}
                GROUP BY nm_id
            ) AS a ON a.nm_id = s.nm_id
        """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows

        # ── 4. Product info from PostgreSQL ──
        nm_ids = [int(r[0]) for r in sku_rows]
        products_map: dict[int, dict] = {}
        if nm_ids:
            nm_list = ", ".join(str(x) for x in nm_ids)
            pg_rows = (await db.execute(
                text(f"""
                    SELECT nm_id, vendor_code, name, current_price, current_discount,
                           COALESCE(length, 0) AS length, COALESCE(width, 0) AS width, COALESCE(height, 0) AS height
                    FROM dim_products
                    WHERE shop_id = :sid AND nm_id IN ({nm_list})
                """),
                {"sid": shop_id},
            )).fetchall()
            for r in pg_rows:
                l, w, h = float(r[5] or 0), float(r[6] or 0), float(r[7] or 0)
                vol = (l * w * h) / 1000.0 if (l > 0 and w > 0 and h > 0) else 1.0
                if vol > 10000:
                    vol = (l * w * h) / 1_000_000.0
                vol = max(vol, 0.1)
                products_map[r[0]] = {
                    "vendor_code": r[1] or "",
                    "name": (r[2] or "")[:60],
                    "vol_liters": vol,
                    "price": float(r[3]) if r[3] else 0,
                    "discount": int(r[4]) if r[4] else 0,
                }

        # ── 5. Cost prices ──
        cost_prices: dict[str, float] = {}
        try:
            cost_rows_pg = (await db.execute(
                text("SELECT offer_id, COALESCE(cost_price, 0) + COALESCE(packaging_cost, 0) AS total_cost FROM product_costs WHERE shop_id = :sid AND (cost_price > 0 OR packaging_cost > 0)"),
                {"sid": shop_id},
            )).fetchall()
            for r in cost_rows_pg:
                cost_prices[r[0]] = float(r[1])
        except Exception:
            pass

        # ── 6. ★ NEW: Per-SKU P&L from fact_finances ──
        sku_pnl: dict[str, dict] = {}  # vendor_code → {revenue, payout, logistics, storage, deductions, ...}
        try:
            pnl_rows = ch.query("""
                SELECT
                    vendor_code,
                    sumIf(JSONExtractFloat(raw_payload, 'retail_price_withdisc_rub'),
                        operation_type = 'Продажа') AS revenue,
                    sumIf(payout_amount, operation_type = 'Продажа')
                        - sumIf(payout_amount, operation_type = 'Возврат') AS payout,
                    sum(abs(wb_delivery_rub)) AS logistics,
                    sum(abs(storage_fee)) AS storage_fact,
                    sum(abs(acceptance_fee)) AS acceptance,
                    sumIf(abs(JSONExtractFloat(raw_payload, 'deduction')), 1) AS deductions,
                    sumIf(quantity, operation_type = 'Продажа' AND quantity > 0) AS sales_qty,
                    sumIf(quantity, operation_type = 'Возврат') AS ret_qty
                FROM mms_analytics.fact_finances FINAL
                WHERE shop_id = {shop_id:UInt32} AND marketplace = 1
                  AND event_date >= {d_start:Date} AND event_date <= {d_end:Date}
                  AND vendor_code != ''
                GROUP BY vendor_code
            """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": today}).result_rows
            for r in pnl_rows:
                vc = str(r[0] or "").strip()
                if not vc:
                    continue
                revenue = float(r[1] or 0)
                payout = float(r[2] or 0)
                logistics = float(r[3] or 0)
                storage_fact = float(r[4] or 0)
                acceptance = float(r[5] or 0)
                deductions = float(r[6] or 0)
                sales_qty = int(r[7] or 0)
                ret_qty = int(r[8] or 0)
                cogs_unit = cost_prices.get(vc, 0)
                cogs_total = cogs_unit * max(sales_qty - abs(ret_qty), 0)
                net_profit = payout - logistics - storage_fact - deductions - acceptance - cogs_total
                sku_pnl[vc] = {
                    "revenue": round(revenue),
                    "payout": round(payout),
                    "logistics": round(logistics),
                    "storage_fact": round(storage_fact),
                    "deductions": round(deductions),
                    "acceptance": round(acceptance),
                    "cogs": round(cogs_total),
                    "net_profit": round(net_profit),
                    "profit_margin_pct": round(net_profit / revenue * 100, 1) if revenue > 0 else 0,
                    "sales_qty": sales_qty,
                }
        except Exception as e:
            logger.warning("P&L per SKU query failed: %s", e)

        # ── 7. Per-warehouse stock distribution ──
        ch2 = get_clickhouse_client()
        wh_stock_map: dict[int, list] = {}
        if nm_ids:
            nm_list_ch = ",".join(str(x) for x in nm_ids[:60])
            wh_stock_rows = ch2.query(f"""
                SELECT nm_id, warehouse_name, argMax(quantity, fetched_at) AS qty
                FROM mms_analytics.fact_inventory_snapshot
                WHERE shop_id = {{shop_id:UInt32}} AND nm_id IN ({nm_list_ch})
                  AND warehouse_name NOT LIKE 'FBS:%'
                GROUP BY nm_id, warehouse_name
                HAVING qty > 0
                ORDER BY nm_id, qty DESC
            """, parameters={"shop_id": shop_id}).result_rows
            for wr in wh_stock_rows:
                wh_stock_map.setdefault(int(wr[0]), []).append({
                    "warehouse": wr[1],
                    "qty": int(wr[2]),
                })
        # Fetch warehouse tariffs for accurate storage cost calculation
        tariff_rows = ch2.query("""
            SELECT warehouse_name,
                   argMax(toFloat32OrZero(storage_base_liter), dt) AS stor_base,
                   argMax(toFloat32OrZero(storage_coef), dt) AS stor_coef
            FROM mms_analytics.fact_wb_acceptance_tariffs
            GROUP BY warehouse_name
        """).result_rows
        tariffs_map: dict[str, dict] = {}
        for tr in tariff_rows:
            tariffs_map[tr[0]] = {
                "stor_base": float(tr[1] or 0),
                "stor_coef": float(tr[2] or 100),
            }
        ch2.close()

        # ── 7b. Fetch ACTUAL paid storage from WB API data ──
        ch3 = get_clickhouse_client()
        actual_storage_map: dict[str, float] = {}  # vendor_code → actual storage ₽/30d
        try:
            actual_rows = ch3.query("""
                SELECT vendor_code,
                       round(SUM(warehouse_price) * (30 / {period:UInt32}), 2) AS storage_30d
                FROM mms_analytics.fact_wb_paid_storage FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND dt >= {d_start:Date}
                  AND dt <= {d_end:Date}
                GROUP BY vendor_code
                HAVING storage_30d != 0
            """, parameters={
                "shop_id": shop_id,
                "period": max(period, 1),
                "d_start": d_start.isoformat(),
                "d_end": today.isoformat(),
            }).result_rows
            for ar in actual_rows:
                vc = str(ar[0])
                if vc:
                    actual_storage_map[vc] = float(ar[1])
            logger.info("AI analysis: loaded actual storage for %d SKUs from fact_wb_paid_storage",
                        len(actual_storage_map))
        except Exception as e:
            logger.warning("Actual paid storage query failed (will use estimates): %s", e)
        ch3.close()

        ch.close()

        # ── 8. Build enriched SKU context with REAL storage costs ──
        # Fallback: average storage per unit if no tariffs
        total_stock_all = sum(s["stock"] for s in wh_summary) or 1
        avg_storage_per_unit = round(float(costs["storage"]) / total_stock_all, 2)
        has_actual_storage = len(actual_storage_map) > 0

        skus_context = []
        for r in sku_rows:
            nm_id = int(r[0])
            prod = products_map.get(nm_id, {})
            vc = prod.get("vendor_code", "")
            stock = int(r[1])
            orders = int(r[2])
            revenue = float(r[3])
            ad_spend = float(r[4])
            ad_orders = int(r[7])
            daily = orders / period if period > 0 else 0
            turnover = round(stock / daily) if daily > 0 else 9999
            drr = round(ad_spend / revenue * 100, 1) if revenue > 0 else 0
            cost_price = cost_prices.get(vc, 0)
            price = prod.get("price", 0)
            vol = prod.get("vol_liters", 1.0)

            # Calculate est_storage_month: prefer ACTUAL data, fallback to tariff estimate
            storage_source = "estimated"
            wh_list = wh_stock_map.get(nm_id, [])
            if vc in actual_storage_map:
                est_storage_month = round(actual_storage_map[vc])
                storage_source = "actual"
            else:
                # Tariff-based estimate (old method)
                wh_list = wh_stock_map.get(nm_id, [])
                est_storage_month = 0.0
                if wh_list and tariffs_map:
                    for wh_item in wh_list:
                        wh_name = wh_item["warehouse"]
                        wh_qty = wh_item["qty"]
                        t = tariffs_map.get(wh_name, {})
                        stor_base = t.get("stor_base", 0)
                        stor_coef = t.get("stor_coef", 100)
                        coef_mult = stor_coef / 100.0 if stor_coef > 0 else 1.0
                        if stor_base > 0:
                            est_storage_month += stor_base * vol * wh_qty * coef_mult * 30
                if est_storage_month < 1:
                    # Fallback: use P&L data or average estimate
                    pnl_st = sku_pnl.get(vc, {}).get("storage_fact", 0)
                    if pnl_st and pnl_st > 0:
                        est_storage_month = round(pnl_st / period * 30)
                    else:
                        est_storage_month = round(stock * avg_storage_per_unit)
                est_storage_month = round(est_storage_month)

            # P&L data from fact_finances
            pnl = sku_pnl.get(vc, {})

            skus_context.append({
                "vendor_code": vc,
                "name": prod.get("name", ""),
                "stock": stock,
                "orders": orders,
                "daily": round(daily, 2),
                "turnover_days": turnover,
                "revenue": round(revenue),
                "ad_spend": round(ad_spend),
                "ad_orders": ad_orders,
                "drr": drr,
                "in_ads": ad_spend > 0,
                "price": price,
                "cost_price": cost_price,
                "vol_liters": round(vol, 2),
                "est_storage_month": est_storage_month,
                "storage_source": storage_source,
                # P&L enrichment
                "net_profit": pnl.get("net_profit", None),
                "profit_margin_pct": pnl.get("profit_margin_pct", None),
                "pnl_revenue": pnl.get("revenue", None),
                "pnl_payout": pnl.get("payout", None),
                "pnl_logistics": pnl.get("logistics", None),
                "pnl_storage_fact": pnl.get("storage_fact", None),
                "pnl_deductions": pnl.get("deductions", None),
                "pnl_cogs": pnl.get("cogs", None),
                "warehouses": wh_list[:5],
            })

        # Sort by REAL storage cost DESC so AI always sees most expensive items first
        skus_context.sort(key=lambda x: x["est_storage_month"], reverse=True)

        # Mark top-5 by storage cost as MUST include in analysis
        for i, s in enumerate(skus_context[:5]):
            s["must_include"] = True
            s["storage_rank"] = i + 1

        # Limit to top-40 for AI prompt context
        skus_for_ai = skus_context[:40]

        # ── 8b. Compute key_metrics in Python (not relying on AI) ──
        total_cross_orders = sum(w["cross_orders"] for w in wh_summary)
        total_orders_all = sum(w["orders"] for w in wh_summary)
        # Cross-logistics loss estimate: avg logistics cost × cross shipment premium (~40% markup)
        avg_logistics_per_order = costs["logistics"] / total_orders_all if total_orders_all > 0 else 0
        cross_logistics_loss = round(total_cross_orders * avg_logistics_per_order * 0.4)  # ~40% premium for cross

        # Storage excess: sum of storage costs for SKUs with turnover > 90 days
        storage_excess = 0
        for s in skus_context:
            if s["turnover_days"] > 90 and s["est_storage_month"] > 0:
                # Excess = storage that could be saved if turnover were 60 days
                ideal_stock = round(s["daily"] * 60) if s["daily"] > 0 else 0
                excess_stock = max(s["stock"] - ideal_stock, 0)
                if s["stock"] > 0:
                    storage_excess += round(s["est_storage_month"] * excess_stock / s["stock"])

        total_potential_savings = cross_logistics_loss + storage_excess

        # ── 9. Build two prompts ──
        total_orders = sum(w["orders"] for w in wh_summary)
        total_cross = sum(w["cross_orders"] for w in wh_summary)
        total_stock = sum(w["stock"] for w in wh_summary)
        total_cross_pct = round(total_cross / total_orders * 100, 1) if total_orders > 0 else 0
        skus_in_ads = sum(1 for s in skus_context if s["in_ads"])
        skus_no_ads = sum(1 for s in skus_context if not s["in_ads"] and s["stock"] > 5)

        common_header = f"""Магазин: {shop.name} (Wildberries)
Период анализа: {period} дней (с {d_start} по {today})
Средняя стоимость хранения: ~{avg_storage_per_unit}₽/шт/мес

## ОБЩИЕ МЕТРИКИ:
- Всего заказов: {total_orders}
- Всего остаток: {total_stock} шт
- % кросс-отправок: {total_cross_pct}%
- Расходы логистика: {costs['logistics']}₽
- Расходы хранение: {costs['storage']}₽
- Штрафы: {costs['penalties']}₽
- SKU в рекламе: {skus_in_ads}, без рекламы: {skus_no_ads}
"""

        # --- Prompt 1: SKU problems ---
        prompt_skus = common_header + "\n## ТОВАРЫ (с полным P&L):\n"
        for i, s in enumerate(skus_for_ai):
            if s.get('must_include'):
                prompt_skus += f"\n### ⚠️ #{i+1}. {s['vendor_code']} — {s['name']} [ТОП-{s['storage_rank']} ПО ХРАНЕНИЮ — ОБЯЗАТЕЛЬНО ВКЛЮЧИТЬ В sku_actions!]\n"
            else:
                prompt_skus += f"\n### {i+1}. {s['vendor_code']} — {s['name']}\n"
            prompt_skus += f"Остаток: {s['stock']} шт | Заказов за {period}д: {s['orders']} ({s['daily']}/день)\n"
            prompt_skus += f"Оборачиваемость: {s['turnover_days']}д | Хранение {'факт' if s.get('storage_source') == 'actual' else 'оценка'}: ~{s['est_storage_month']}₽/мес | Объём: {s.get('vol_liters', '?')}л\n"
            prompt_skus += f"Цена: {s['price']}₽ | Себестоимость: {s['cost_price']}₽\n"
            if s['net_profit'] is not None:
                prompt_skus += f"★ ЧИСТАЯ ПРИБЫЛЬ за {period}д: {s['net_profit']}₽ (маржа {s['profit_margin_pct']}%)\n"
                prompt_skus += f"  P&L: выручка {s['pnl_revenue']}₽ → payout {s['pnl_payout']}₽ → логистика −{s['pnl_logistics']}₽ → хранение −{s['pnl_storage_fact']}₽ → удержания −{s['pnl_deductions']}₽ → COGS −{s['pnl_cogs']}₽\n"
            prompt_skus += f"Реклама: {s['ad_spend']}₽ ({s['ad_orders']} заказов) | DRR: {s['drr']}% | В рекламе: {'да' if s['in_ads'] else 'НЕТ'}\n"

        prompt_skus += "\nПроанализируй проблемные SKU и выдай JSON с sku_actions."

        # --- Prompt 2: Redistribution ---
        prompt_redistr = common_header + "\n## СКЛАДЫ:\n"
        for w in wh_summary[:15]:
            prompt_redistr += f"- {w['warehouse']}: заказов={w['orders']}, кросс={w['cross_pct']}%, остаток={w['stock']}, оборач={w['avg_turnover']}дн\n"

        prompt_redistr += "\n## РАСПРЕДЕЛЕНИЕ ТОВАРОВ ПО СКЛАДАМ:\n"
        for s in skus_for_ai[:20]:
            if s['warehouses']:
                wh_str = ", ".join(f"{wh['warehouse']}={wh['qty']}шт" for wh in s['warehouses'])
                prompt_redistr += f"- {s['vendor_code']}: {s['stock']} шт, {s['daily']}/день, оборач {s['turnover_days']}д | {wh_str}\n"

        prompt_redistr += "\nПроанализируй распределение и выдай JSON с transfers, general_tips и supply_tip."

        # ── 10. Call Gemini (2 parallel requests) ──
        KIE_AI_URL = "https://api.kie.ai/gemini-2.5-flash/v1/chat/completions"

        async def call_gemini(system_prompt: str, user_prompt: str) -> dict:
            async with httpx.AsyncClient(timeout=120.0) as client:
                resp = await client.post(
                    KIE_AI_URL,
                    headers={
                        "Content-Type": "application/json",
                        "Authorization": f"Bearer {api_key}",
                    },
                    json={
                        "messages": [
                            {"role": "system", "content": [{"type": "text", "text": system_prompt}]},
                            {"role": "user", "content": [{"type": "text", "text": user_prompt}]},
                        ],
                        "stream": False,
                        "include_thoughts": False,
                    },
                )

            if resp.status_code != 200:
                logger.error("Gemini API error %s: %s", resp.status_code, resp.text[:500])
                return {}

            resp_json = resp.json()
            content = resp_json.get("choices", [{}])[0].get("message", {}).get("content", "")

            # Strip markdown code fences
            content = content.strip()
            if content.startswith("```"):
                content = content.split("\n", 1)[1] if "\n" in content else content[3:]
            if content.endswith("```"):
                content = content[:-3]
            content = content.strip()
            if content.startswith("json"):
                content = content[4:].strip()

            try:
                return json.loads(content)
            except json.JSONDecodeError:
                logger.warning("Failed to parse AI JSON: %s", content[:500])
                return {}

        # Parallel execution
        result_skus, result_redistr = await asyncio.gather(
            call_gemini(_AI_PROMPT_SKU_PROBLEMS, prompt_skus),
            call_gemini(_AI_PROMPT_REDISTRIBUTION, prompt_redistr),
        )

        # ── 11. Merge results ──
        # Count unprofitable SKUs (storage > net_profit or no sales)
        unprofitable_count = sum(
            1 for s in skus_context
            if s["turnover_days"] > 90 and s["est_storage_month"] > 100
        )

        # ── 11a. Force-inject top-5 storage SKUs into sku_actions ──
        ai_sku_actions = result_skus.get("sku_actions", [])
        ai_vendor_codes = {sa.get("vendor_code", "") for sa in ai_sku_actions}

        # Get top-5 must_include SKUs that AI missed
        for s in skus_for_ai:
            if not s.get("must_include"):
                continue
            if s["vendor_code"] in ai_vendor_codes:
                continue  # AI already included this one

            # Build a proper sku_action entry from Python data
            storage_cost = s["est_storage_month"]
            net_profit = s.get("net_profit") or 0
            turnover = s["turnover_days"]
            stock = s["stock"]
            daily = s["daily"]
            vol = s.get("vol_liters", 0)

            # Determine severity
            is_loss = storage_cost > 0 and net_profit < storage_cost
            is_slow = turnover > 90

            diagnosis_parts = []
            diagnosis_parts.append(f"Оборачиваемость {turnover}д")
            diagnosis_parts.append(f"хранение {storage_cost}₽/мес")
            if net_profit is not None and net_profit != 0:
                diagnosis_parts.append(f"чистая прибыль {net_profit}₽/мес")
                if is_loss:
                    diagnosis_parts.append(f"→ убыток {abs(net_profit - storage_cost):.0f}₽/мес")
            diagnosis_text = ", ".join(diagnosis_parts) + "."

            options = []
            # Option 1: discount
            if stock > 0 and daily > 0:
                ideal_stock = round(daily * 60)
                excess = max(stock - ideal_stock, 0)
                savings_discount = round(storage_cost * 0.4) if excess > 0 else round(storage_cost * 0.2)
                options.append({
                    "action": "discount",
                    "label": "Снизить цену на 15-25%",
                    "detail": f"Крупная фасовка ({vol}л), остаток {stock} шт. Скидка ускорит продажи и снизит остаток. "
                              f"При росте продаж в 1.5× оборачиваемость упадёт до ~{max(turnover // 2, 30)}д. "
                              f"Экономия на хранении ~{savings_discount}₽/мес.",
                    "expected_savings": savings_discount,
                    "risk": "medium"
                })
            elif stock > 0:
                options.append({
                    "action": "discount",
                    "label": "Распродажа со скидкой 30-40%",
                    "detail": f"Нет продаж за период. Объём {vol}л, остаток {stock} шт. "
                              f"Глубокая скидка поможет распродать сток и сэкономить {storage_cost}₽/мес на хранении.",
                    "expected_savings": storage_cost,
                    "risk": "medium"
                })

            # Option 2: reduce supply
            if stock > 20:
                options.append({
                    "action": "reduce_supply",
                    "label": "Сократить поставки",
                    "detail": f"Объём {vol}л — крупная фасовка, занимает много места. "
                              f"Не поставлять новые партии до снижения остатка. "
                              f"Текущий запас {stock} шт ≈ {turnover}д оборачиваемости.",
                    "expected_savings": round(storage_cost * 0.3),
                    "risk": "low"
                })

            # Option 3: do nothing (always)
            months_stock = round(stock / daily, 1) if daily > 0 else 999
            options.append({
                "action": "do_nothing",
                "label": "Оставить как есть",
                "detail": f"Хранение {storage_cost}₽/мес. "
                          f"Запас {stock} шт при {daily} продажах/день ≈ {months_stock} мес. "
                          f"Через 3 мес расходы на хранение: +{storage_cost * 3}₽.",
                "expected_savings": 0,
                "risk": "high" if is_loss else "medium"
            })

            forced_action = {
                "vendor_code": s["vendor_code"],
                "name": s["name"],
                "diagnosis": diagnosis_text,
                "current_storage_cost": storage_cost,
                "current_turnover_days": turnover,
                "stock": stock,
                "options": options,
                "recommended_option": 0,
            }
            # Insert at the beginning (high priority)
            ai_sku_actions.insert(0, forced_action)
            logger.info("Force-injected SKU %s (storage_rank=%s, est_storage=%s₽/мес) into sku_actions",
                        s["vendor_code"], s.get("storage_rank"), storage_cost)

        # ── 11b. Build analysis_sections with Python fallback ──
        ai_sections = result_skus.get("analysis_sections", [])
        section_keys = {s.get("section") for s in ai_sections}

        # Fallback: ensure all 4 sections exist
        if "cross_logistics" not in section_keys:
            sev = "critical" if total_cross_pct > 50 else "warning" if total_cross_pct > 25 else "ok"
            ai_sections.append({
                "section": "cross_logistics",
                "severity": sev,
                "summary": f"Кросс-отправки {total_cross_pct}%. Потери ~{cross_logistics_loss}₽/мес" if total_cross_pct > 25 else f"Кросс-отправки {total_cross_pct}% — под контролем",
                "action_text": "Кросс-логистика →"
            })
        if "storage" not in section_keys:
            sev = "critical" if storage_excess > 5000 else "warning" if storage_excess > 1000 else "ok"
            ai_sections.append({
                "section": "storage",
                "severity": sev,
                "summary": f"Избыточное хранение: ~{storage_excess}₽/мес. {unprofitable_count} SKU с оборачиваемостью >90д" if storage_excess > 0 else "Хранение в норме",
                "action_text": "Хранение →"
            })
        if "supply" not in section_keys:
            # Count SKUs near out-of-stock globally
            oos_count = 0
            for s in skus_context:
                if s["daily"] > 0 and (s["stock"] == 0 or (s["stock"] / s["daily"]) < 14):
                    oos_count += 1
            sev = "critical" if oos_count > 2 else "warning" if oos_count > 0 else "ok"
            ai_sections.append({
                "section": "supply",
                "severity": sev,
                "summary": f"{oos_count} SKU с запасом <14 дней на всех складах. Риск out-of-stock" if oos_count > 0 else "Запасы в норме",
                "action_text": "Поставки →"
            })
        if "geography" not in section_keys:
            ai_sections.append({
                "section": "geography",
                "severity": "warning" if total_cross_pct > 40 else "ok",
                "summary": f"Кросс-доставка {total_cross_pct}% указывает на неоптимальное размещение" if total_cross_pct > 40 else "Географическое покрытие в норме",
                "action_text": "География →"
            })

        ai_result = {
            "severity": result_skus.get("severity", "warning"),
            "diagnosis": result_skus.get("diagnosis", "Анализ недоступен. Попробуйте позже."),
            "total_potential_savings": total_potential_savings,
            "key_metrics": {
                "cross_logistics_loss": cross_logistics_loss,
                "storage_excess": storage_excess,
                "unprofitable_skus_count": unprofitable_count,
            },
            "analysis_sections": ai_sections,
            "sku_actions": ai_sku_actions,
            "transfers": result_redistr.get("transfers", []),
            "general_tips": result_redistr.get("general_tips", []),
            "supply_tip": result_redistr.get("supply_tip", "Используйте раздел «Склады → Поставки» для расчёта оптимальных объёмов"),
        }

        # Add metadata
        ai_result["shop_name"] = shop.name
        ai_result["period_days"] = period
        ai_result["analyzed_at"] = int(time.time())
        ai_result["context"] = {
            "total_orders": total_orders,
            "total_stock": total_stock,
            "cross_pct": total_cross_pct,
            "costs_logistics": costs["logistics"],
            "costs_storage": costs["storage"],
            "costs_penalties": costs["penalties"],
            "skus_in_ads": skus_in_ads,
            "skus_no_ads": skus_no_ads,
            "warehouses_count": len(wh_summary),
            "has_actual_storage": has_actual_storage,
            "actual_storage_skus": len(actual_storage_map),
        }

        # ── 12. Cache ──
        _ai_cache[cache_key] = (time.time(), ai_result)

        return {**ai_result, "cached": False}

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("AI warehouse analysis failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Analysis error: {str(e)}")


# ═══════════════════════════════════════════════════════════════
# WB Sales Geography
# ═══════════════════════════════════════════════════════════════

@router.get("/wb/geography")
async def get_wb_geography(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    nm_ids: str = Query(None, description="Comma-separated nm_ids to filter by"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """WB sales geography — okrugs with stability, avg check, drill-down."""
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "wildberries":
        raise HTTPException(status_code=404, detail="Shop not found")

    from app.core.clickhouse import get_clickhouse_client

    ch = get_clickhouse_client()
    today = date.today()
    d_start = today - timedelta(days=period)

    # Parse multi-SKU filter
    nm_id_list: list[int] = []
    nm_filter = ""
    params: dict = {"shop_id": shop_id, "d_start": d_start, "period_days": period}
    if nm_ids:
        try:
            nm_id_list = [int(x.strip()) for x in nm_ids.split(",") if x.strip()]
        except ValueError:
            pass
        if nm_id_list:
            id_str = ", ".join(str(x) for x in nm_id_list)
            nm_filter = f"AND nm_id IN ({id_str})"

    # ── 1. Orders by okrug + region with stability ──
    rows = ch.query(f"""
        SELECT
            oblast_okrug_name AS okrug,
            region_name AS region,
            count() AS orders,
            sum(toFloat64(price_with_disc)) AS revenue,
            count(DISTINCT toDate(date)) AS active_days
        FROM mms_analytics.fact_orders_raw FINAL
        WHERE shop_id = {{shop_id:UInt32}}
          AND date >= {{d_start:Date}}
          AND is_cancel = 0
          {nm_filter}
        GROUP BY okrug, region
        ORDER BY orders DESC
    """, parameters=params).result_rows

    total_orders = sum(int(r[2]) for r in rows)
    total_revenue = sum(float(r[3]) for r in rows)

    # ── 1b. Stability per okrug (unique days with orders) ──
    okrug_stability_rows = ch.query(f"""
        SELECT
            oblast_okrug_name AS okrug,
            count(DISTINCT toDate(date)) AS active_days
        FROM mms_analytics.fact_orders_raw FINAL
        WHERE shop_id = {{shop_id:UInt32}}
          AND date >= {{d_start:Date}}
          AND is_cancel = 0
          {nm_filter}
        GROUP BY okrug
    """, parameters=params).result_rows
    okrug_active_days: dict[str, int] = {}
    for r in okrug_stability_rows:
        okrug_active_days[str(r[0]) or "Не определено"] = int(r[1])

    # Group by okrug
    okrug_data: dict[str, dict] = {}
    for r in rows:
        okrug = str(r[0]) or "Не определено"
        region = str(r[1]) or "Не определено"
        orders = int(r[2])
        revenue = float(r[3])
        active_days = int(r[4])

        if okrug not in okrug_data:
            okrug_data[okrug] = {
                "okrug": okrug,
                "orders": 0,
                "revenue": 0,
                "share_pct": 0,
                "avg_check": 0,
                "stability_pct": 0,
                "regions": [],
            }
        okrug_data[okrug]["orders"] += orders
        okrug_data[okrug]["revenue"] += revenue
        region_avg_check = round(revenue / orders, 2) if orders > 0 else 0
        region_stability = round(active_days / period * 100, 1)
        okrug_data[okrug]["regions"].append({
            "region": region,
            "orders": orders,
            "revenue": round(revenue, 2),
            "avg_check": region_avg_check,
            "stability_pct": region_stability,
            "share_pct": round(orders / total_orders * 100, 1) if total_orders > 0 else 0,
        })

    regions_result = []
    for okrug_name, od in sorted(okrug_data.items(), key=lambda x: x[1]["orders"], reverse=True):
        od["share_pct"] = round(od["orders"] / total_orders * 100, 1) if total_orders > 0 else 0
        od["revenue"] = round(od["revenue"], 2)
        od["avg_check"] = round(od["revenue"] / od["orders"], 2) if od["orders"] > 0 else 0
        active_d = okrug_active_days.get(okrug_name, 0)
        od["stability_pct"] = round(active_d / period * 100, 1)
        od["regions"] = sorted(od["regions"], key=lambda x: x["orders"], reverse=True)
        regions_result.append(od)

    # ── 2. Top products ──
    top_products = []
    prod_map: dict[int, dict] = {}
    if not nm_id_list:
        prod_rows = ch.query(f"""
            SELECT
                nm_id,
                count() AS orders,
                sum(toFloat64(price_with_disc)) AS revenue,
                count(DISTINCT oblast_okrug_name) AS okrug_count,
                count(DISTINCT region_name) AS region_count,
                uniqExact(toDate(date)) AS active_days
            FROM mms_analytics.fact_orders_raw FINAL
            WHERE shop_id = {{shop_id:UInt32}}
              AND date >= {{d_start:Date}}
              AND is_cancel = 0
            GROUP BY nm_id
            ORDER BY orders DESC
            LIMIT 50
        """, parameters=params).result_rows

        prod_nm_ids = [int(r[0]) for r in prod_rows]
        if prod_nm_ids:
            nm_list = ", ".join(str(x) for x in prod_nm_ids)
            pg_rows = (await db.execute(
                text(f"""
                    SELECT nm_id, vendor_code, name
                    FROM dim_products
                    WHERE shop_id = :sid AND nm_id IN ({nm_list})
                """),
                {"sid": shop_id},
            )).fetchall()
            for r in pg_rows:
                prod_map[r[0]] = {"vendor_code": r[1] or "", "name": (r[2] or "")[:80]}

        for r in prod_rows:
            nm = int(r[0])
            prod = prod_map.get(nm, {})
            nm_orders = int(r[1])
            nm_rev = float(r[2])
            active_days = int(r[5])
            top_products.append({
                "nm_id": nm,
                "vendor_code": prod.get("vendor_code", ""),
                "name": prod.get("name", ""),
                "orders": nm_orders,
                "revenue": round(nm_rev, 2),
                "avg_check": round(nm_rev / nm_orders, 2) if nm_orders > 0 else 0,
                "okrug_count": int(r[3]),
                "region_count": int(r[4]),
                "stability_pct": round(active_days / period * 100, 1),
                "share_pct": round(nm_orders / total_orders * 100, 1) if total_orders > 0 else 0,
            })

    # ── 3. Per-okrug top products ──
    okrug_top_products: dict[str, list] = {}
    if not nm_id_list:
        otop_rows = ch.query(f"""
            SELECT
                oblast_okrug_name AS okrug,
                nm_id,
                count() AS orders,
                sum(toFloat64(price_with_disc)) AS revenue,
                uniqExact(toDate(date)) AS active_days
            FROM mms_analytics.fact_orders_raw FINAL
            WHERE shop_id = {{shop_id:UInt32}}
              AND date >= {{d_start:Date}}
              AND is_cancel = 0
            GROUP BY okrug, nm_id
            ORDER BY okrug, orders DESC
        """, parameters=params).result_rows

        current_okrug = ""
        count = 0
        for r in otop_rows:
            okrug = str(r[0]) or "Не определено"
            if okrug != current_okrug:
                current_okrug = okrug
                count = 0
            count += 1
            if count <= 5:
                nm = int(r[1])
                prod = prod_map.get(nm, {})
                active_days = int(r[4])
                okrug_top_products.setdefault(okrug, []).append({
                    "nm_id": nm,
                    "vendor_code": prod.get("vendor_code", ""),
                    "name": prod.get("name", ""),
                    "orders": int(r[2]),
                    "revenue": round(float(r[3]), 2),
                    "avg_check": round(float(r[3]) / int(r[2]), 2) if int(r[2]) > 0 else 0,
                    "stability_pct": round(active_days / period * 100, 1),
                })

    # ── 4. SKU filter info ──
    sku_filter_info = []
    if nm_id_list:
        nm_list_str = ", ".join(str(x) for x in nm_id_list)
        pg_rows = (await db.execute(
            text(f"""
                SELECT nm_id, vendor_code, name
                FROM dim_products
                WHERE shop_id = :sid AND nm_id IN ({nm_list_str})
            """),
            {"sid": shop_id},
        )).fetchall()
        for r in pg_rows:
            sku_filter_info.append({
                "nm_id": r[0],
                "vendor_code": r[1] or "",
                "name": (r[2] or "")[:80],
            })

    # ── 5. Total unique regions count ──
    total_regions = len(set(
        str(r[1]) for r in rows if str(r[1])
    ))
    total_okrugs = len(okrug_data)

    avg_check = round(total_revenue / total_orders, 2) if total_orders > 0 else 0

    ch.close()

    return {
        "total_orders": total_orders,
        "total_revenue": round(total_revenue, 2),
        "avg_check": avg_check,
        "total_okrugs": total_okrugs,
        "total_regions": total_regions,
        "period_days": period,
        "regions": regions_result,
        "top_products": top_products,
        "okrug_top_products": okrug_top_products,
        "sku_filter": sku_filter_info,
    }


@router.get("/wb/geography/region-products")
async def get_wb_geography_region_products(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    region: str = Query(..., description="Region name (e.g. Московская область)"),
    nm_ids: str = Query(None, description="Comma-separated nm_ids to filter by"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Top products for a specific region (drill-down)."""
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "wildberries":
        raise HTTPException(status_code=404, detail="Shop not found")

    from app.core.clickhouse import get_clickhouse_client

    ch = get_clickhouse_client()
    today = date.today()
    d_start = today - timedelta(days=period)

    nm_filter = ""
    if nm_ids:
        try:
            nm_id_list = [int(x.strip()) for x in nm_ids.split(",") if x.strip()]
            if nm_id_list:
                id_str = ", ".join(str(x) for x in nm_id_list)
                nm_filter = f"AND nm_id IN ({id_str})"
        except ValueError:
            pass

    total_weeks = max(1, period // 7)

    prod_rows = ch.query(f"""
        SELECT
            nm_id,
            count() AS orders,
            sum(toFloat64(price_with_disc)) AS revenue,
            uniqExact(toMonday(date)) AS active_weeks
        FROM mms_analytics.fact_orders_raw FINAL
        WHERE shop_id = {{shop_id:UInt32}}
          AND date >= {{d_start:Date}}
          AND is_cancel = 0
          AND region_name = {{region:String}}
          {nm_filter}
        GROUP BY nm_id
        ORDER BY orders DESC
        LIMIT 20
    """, parameters={"shop_id": shop_id, "d_start": d_start, "region": region}).result_rows

    ch.close()

    nm_ids = [int(r[0]) for r in prod_rows]
    prod_map: dict[int, dict] = {}
    if nm_ids:
        nm_list = ", ".join(str(x) for x in nm_ids)
        pg_rows = (await db.execute(
            text(f"""
                SELECT nm_id, vendor_code, name
                FROM dim_products
                WHERE shop_id = :sid AND nm_id IN ({nm_list})
            """),
            {"sid": shop_id},
        )).fetchall()
        for r in pg_rows:
            prod_map[r[0]] = {"vendor_code": r[1] or "", "name": (r[2] or "")[:80]}

    result = []
    for r in prod_rows:
        nm = int(r[0])
        prod = prod_map.get(nm, {})
        nm_orders = int(r[1])
        nm_rev = float(r[2])
        active_weeks = int(r[3])
        stability_pct = round(active_weeks / total_weeks * 100, 1)
        result.append({
            "nm_id": nm,
            "vendor_code": prod.get("vendor_code", ""),
            "name": prod.get("name", ""),
            "orders": nm_orders,
            "revenue": round(nm_rev, 2),
            "avg_check": round(nm_rev / nm_orders, 2) if nm_orders > 0 else 0,
            "stability_pct": stability_pct,
        })

    return {"region": region, "products": result}



@router.get("/wb/geography/products-search")
async def get_wb_geography_products_search(
    shop_id: int = Query(...),
    q: str = Query("", description="Search query (name, vendor_code, or nm_id)"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Search products for autocomplete in geography filter."""
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "wildberries":
        raise HTTPException(status_code=404, detail="Shop not found")

    search = q.strip()
    if not search:
        # Return all products (limited)
        pg_rows = (await db.execute(
            text("""
                SELECT nm_id, vendor_code, name
                FROM dim_products
                WHERE shop_id = :sid
                ORDER BY name
                LIMIT 100
            """),
            {"sid": shop_id},
        )).fetchall()
    else:
        # Try numeric search (nm_id)
        try:
            nm_int = int(search)
            pg_rows = (await db.execute(
                text("""
                    SELECT nm_id, vendor_code, name
                    FROM dim_products
                    WHERE shop_id = :sid AND nm_id = :nm
                    LIMIT 20
                """),
                {"sid": shop_id, "nm": nm_int},
            )).fetchall()
            if not pg_rows:
                # Partial nm_id match
                pg_rows = (await db.execute(
                    text("""
                        SELECT nm_id, vendor_code, name
                        FROM dim_products
                        WHERE shop_id = :sid AND CAST(nm_id AS TEXT) LIKE :pattern
                        ORDER BY name
                        LIMIT 20
                    """),
                    {"sid": shop_id, "pattern": f"%{search}%"},
                )).fetchall()
        except ValueError:
            pg_rows = (await db.execute(
                text("""
                    SELECT nm_id, vendor_code, name
                    FROM dim_products
                    WHERE shop_id = :sid
                      AND (LOWER(name) LIKE :q OR LOWER(vendor_code) LIKE :q)
                    ORDER BY name
                    LIMIT 30
                """),
                {"sid": shop_id, "q": f"%{search.lower()}%"},
            )).fetchall()

    result = []
    for r in pg_rows:
        result.append({
            "nm_id": r[0],
            "vendor_code": r[1] or "",
            "name": (r[2] or "")[:80],
        })

    return {"products": result}


# ═══════════════════════════════════════════════════════════════════
# AI Geography Analysis — structured Gemini output
# ═══════════════════════════════════════════════════════════════════

_AI_PROMPT_GEOGRAPHY = """Ты — эксперт по ГЕОГРАФИИ ПРОДАЖ на Wildberries.

ТВОЯ ЗАДАЧА: проанализировать ГЕОГРАФИЧЕСКОЕ распределение спроса и дать конкретные выводы.

## СТРОГИЕ ГРАНИЦЫ АНАЛИЗА

АНАЛИЗИРУЙ ТОЛЬКО:
- Географическое распределение заказов по округам и регионам
- Тренды: какие регионы растут, какие стагнируют (по стабильности спроса)
- Концентрация: насколько равномерно распределены продажи
- Недопокрытые регионы: где есть спрос но мало/нет стока
- Географический охват товаров: в скольких регионах продаётся каждый товар
- Потенциал роста в новых регионах

НЕ АНАЛИЗИРУЙ (для этого есть ОТДЕЛЬНЫЕ разделы в системе):
- ❌ Рекламу, DRR, рекламные кампании → раздел "Реклама"
- ❌ Детали кросс-логистики, оптимизацию складов → раздел "Кросс-логистика"
- ❌ Финансовые показатели (маржа, прибыль) → раздел "Финансы"

ВАЖНО ПРО WILDBERRIES:
- На WB реклама НЕ ТАРГЕТИРУЕТСЯ по регионам — нельзя запустить рекламу "только на Сибирь"
- Единственный способ влиять на географию — РАЗМЕЩЕНИЕ СТОКА на складах ближе к спросу
- Не давай советов "запустить рекламу в регионе" — это технически невозможно на WB

## ЛОГИКА АНАЛИЗА

### 1. Концентрация (ОБЯЗАТЕЛЬНО)
- Посчитай: какой % выручки дают топ-3 округа
- Если > 70% — ВЫСОКАЯ концентрация (risk_level: "high")
- Если 40-70% — СРЕДНЯЯ (risk_level: "medium")
- Если < 40% — НИЗКАЯ, хорошая диверсификация (risk_level: "low")
- НАЗОВИ конкретные округа и их долю

### 2. Инсайты по товарам (2-5 штук) — ТОЛЬКО географические!
Для каждого товара определи тип:
- **stable_leader** — высокая стабильность (>30%) + много регионов → лидер, масштабировать географию
- **unstable_demand** — стабильность <15% → нестабильный спрос, возможны проблемы с наличием
- **regional_champion** — продаётся хорошо в 1-2 округах, но отсутствует в остальных → потенциал расширения географии
- **cross_delivery_problem** — товар продаётся в округе, но склад далеко → нужен сток ближе
- **dead_stock_risk** — есть на складе, но нет заказов из ближайших регионов → нет спроса в этой географии

Для КАЖДОГО инсайта:
- Укажи vendor_code и name
- Укажи конкретные числа: заказы, стабильность, в скольких регионах
- Дай конкретное действие (action): redistribute / increase_supply / monitor
- Рассчитай expected_effect — экономию или потенциальную выручку в РУБЛЯХ

### 3. Логистическое соответствие (1-3 самых важных)
Оценка: насколько расположение стока соответствует географии спроса.
Для каждого округа с высокой кросс-доставкой:
- Укажи: откуда реально доставляется (serving_warehouse)
- Укажи ближайший склад и сток на нём
- Дай рекомендацию по размещению: куда и сколько штук отправить

## ФОРМАТ ОТВЕТА — СТРОГО JSON:
{
  "severity": "critical" | "warning" | "ok",
  "diagnosis": "Краткий главный географический вывод из 1-2 предложений с конкретными числами",
  "key_metrics": {
    "concentration_pct": 80,
    "top_regions_count": 3,
    "total_regions": 64,
    "regions_with_stable_demand": 12,
    "underserved_okrugs": 2
  },
  "concentration": {
    "summary": "80% выручки из Центрального, Южного и Приволжского ФО",
    "top_regions": [
      {"region": "Центральный федеральный округ", "orders": 3276, "share_pct": 33.7, "stability_pct": 107}
    ],
    "risk_level": "high" | "medium" | "low",
    "recommendation": "Конкретная рекомендация по диверсификации географии через размещение стока"
  },
  "product_insights": [
    {
      "vendor_code": "АРТИКУЛ",
      "name": "Название",
      "insight_type": "stable_leader",
      "regions_count": 12,
      "stability_pct": 47,
      "orders": 21,
      "detail": "Географический анализ с числами: где продаётся, где не продаётся, что делать.",
      "action": "redistribute",
      "expected_effect": "Потенциальная выручка +50 000₽/мес при расширении на Сибирь"
    }
  ],
  "logistics_match": [
    {
      "okrug": "Приволжский ФО",
      "orders": 50,
      "nearest_warehouse": "Казань",
      "warehouse_stock": 0,
      "serving_warehouse": "Коледино",
      "cross_pct": 100,
      "detail": "Описание проблемы: спрос есть, стока нет, доставляется издалека",
      "recommendation": "Разместить N штук на складе X для покрытия спроса"
    }
  ],
  "general_tips": [
    "Конкретная географическая рекомендация, 1-2 предложения."
  ]
}

## ПРАВИЛА
- severity: "critical" если концентрация > 70% или есть округа с 100% кросс-доставкой, "warning" если концентрация 50-70%, "ok" если всё сбалансировано
- product_insights: 2-5 ГЕОГРАФИЧЕСКИХ инсайтов. НАЗЫВАЙ товары по vendor_code + name
- insight_type: ТОЛЬКО "stable_leader", "unstable_demand", "regional_champion", "cross_delivery_problem", "dead_stock_risk"
- action: ТОЛЬКО "redistribute", "increase_supply", "monitor" (НЕ "launch_ads" — на WB нельзя таргетировать рекламу по гео!)
- logistics_match: 1-3 проблемных округа. Если проблем нет — пустой массив
- general_tips: 2-3 совета ПО ГЕОГРАФИИ. Только про размещение стока и диверсификацию, НЕ про рекламу
- top_regions в concentration: топ-5 округов по выручке
- Все числа — из реальных данных, НЕ выдумывай
- Пиши НА РУССКОМ
"""


# Map WB warehouse names to approximate okrug
_WH_TO_OKRUG = {
    "Коледино": "Центральный", "Подольск": "Центральный", "Электросталь": "Центральный",
    "Котовск": "Центральный", "Тверь": "Центральный", "Белые Столбы": "Центральный",
    "Казань": "Приволжский", "Набережные Челны": "Приволжский",
    "Краснодар": "Южный", "Ростов": "Южный",
    "Екатеринбург": "Уральский",
    "Новосибирск": "Сибирский",
    "Хабаровск": "Дальневосточный", "Владивосток": "Дальневосточный",
    "Санкт-Петербург": "Северо-Западный", "СПб": "Северо-Западный",
    "Воронеж": "Центральный",
}


def _wh_to_okrug(wh_name: str) -> str:
    """Map warehouse name to approximate okrug."""
    for key, okrug in _WH_TO_OKRUG.items():
        if key.lower() in wh_name.lower():
            return okrug
    return "Другой"


@router.post("/wb/geography/ai-analysis")
async def get_wb_geography_ai_analysis(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    force: bool = Query(False, description="Skip cache"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """AI-powered geography sales analysis using Gemini 2.5 Flash."""
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "wildberries":
        raise HTTPException(status_code=404, detail="Shop not found")

    cache_key = f"geo_ai_{shop_id}_{period}"
    if not force and cache_key in _ai_cache:
        ts, cached = _ai_cache[cache_key]
        if time.time() - ts < _AI_CACHE_TTL:
            return {**cached, "cached": True}

    api_key = os.getenv("KIE_AI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="AI API key not configured")

    try:
        from app.core.clickhouse import get_clickhouse_client

        ch = get_clickhouse_client()
        today = date.today()
        d_start = today - timedelta(days=period)
        params = {"shop_id": shop_id, "d_start": d_start, "period_days": period}

        # ── 1. Orders by okrug + region with stability ──
        geo_rows = ch.query("""
            SELECT
                oblast_okrug_name AS okrug,
                region_name AS region,
                count() AS orders,
                sum(toFloat64(price_with_disc)) AS revenue,
                count(DISTINCT toDate(date)) AS active_days
            FROM mms_analytics.fact_orders_raw FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND date >= {d_start:Date}
              AND is_cancel = 0
            GROUP BY okrug, region
            ORDER BY orders DESC
        """, parameters=params).result_rows

        total_orders = sum(int(r[2]) for r in geo_rows)
        total_revenue = sum(float(r[3]) for r in geo_rows)

        # Build okrug summary
        okrug_summary: dict[str, dict] = {}
        for r in geo_rows:
            okrug = str(r[0]) or "Не определено"
            region = str(r[1]) or "Не определено"
            orders = int(r[2])
            revenue = float(r[3])
            active_days = int(r[4])

            if okrug not in okrug_summary:
                okrug_summary[okrug] = {"orders": 0, "revenue": 0, "regions": []}
            okrug_summary[okrug]["orders"] += orders
            okrug_summary[okrug]["revenue"] += revenue
            okrug_summary[okrug]["regions"].append({
                "region": region,
                "orders": orders,
                "revenue": round(revenue),
                "stability_pct": round(active_days / period * 100, 1),
            })

        # Okrug-level stability
        okrug_stab_rows = ch.query("""
            SELECT
                oblast_okrug_name AS okrug,
                count(DISTINCT toDate(date)) AS active_days
            FROM mms_analytics.fact_orders_raw FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND date >= {d_start:Date}
              AND is_cancel = 0
            GROUP BY okrug
        """, parameters=params).result_rows
        for r in okrug_stab_rows:
            okrug = str(r[0]) or "Не определено"
            if okrug in okrug_summary:
                okrug_summary[okrug]["stability_pct"] = round(int(r[1]) / period * 100, 1)

        # ── 2. Top products with geography spread ──
        prod_rows = ch.query("""
            SELECT
                nm_id,
                count() AS orders,
                sum(toFloat64(price_with_disc)) AS revenue,
                count(DISTINCT oblast_okrug_name) AS okrug_count,
                count(DISTINCT region_name) AS region_count,
                uniqExact(toDate(date)) AS active_days
            FROM mms_analytics.fact_orders_raw FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND date >= {d_start:Date}
              AND is_cancel = 0
            GROUP BY nm_id
            ORDER BY orders DESC
            LIMIT 30
        """, parameters=params).result_rows

        prod_nm_ids = [int(r[0]) for r in prod_rows]
        prod_map: dict[int, dict] = {}
        if prod_nm_ids:
            nm_list = ", ".join(str(x) for x in prod_nm_ids)
            pg_rows = (await db.execute(
                text(f"""
                    SELECT nm_id, vendor_code, name
                    FROM dim_products
                    WHERE shop_id = :sid AND nm_id IN ({nm_list})
                """),
                {"sid": shop_id},
            )).fetchall()
            for r in pg_rows:
                prod_map[r[0]] = {"vendor_code": r[1] or "", "name": (r[2] or "")[:60]}

        products_context = []
        for r in prod_rows:
            nm = int(r[0])
            prod = prod_map.get(nm, {})
            nm_orders = int(r[1])
            active_days = int(r[5])
            products_context.append({
                "nm_id": nm,
                "vendor_code": prod.get("vendor_code", ""),
                "name": prod.get("name", f"nm_id {nm}"),
                "orders": nm_orders,
                "revenue": round(float(r[2])),
                "okrug_count": int(r[3]),
                "region_count": int(r[4]),
                "stability_pct": round(active_days / period * 100, 1),
                "share_pct": round(nm_orders / total_orders * 100, 1) if total_orders > 0 else 0,
            })

        # ── 3. Per-okrug top products ──
        otop_rows = ch.query("""
            SELECT
                oblast_okrug_name AS okrug,
                nm_id,
                count() AS orders,
                uniqExact(toDate(date)) AS active_days
            FROM mms_analytics.fact_orders_raw FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND date >= {d_start:Date}
              AND is_cancel = 0
            GROUP BY okrug, nm_id
            ORDER BY okrug, orders DESC
        """, parameters=params).result_rows

        okrug_products: dict[str, list] = {}
        current_ok = ""
        cnt = 0
        for r in otop_rows:
            ok = str(r[0]) or "?"
            if ok != current_ok:
                current_ok = ok
                cnt = 0
            cnt += 1
            if cnt <= 5:
                nm = int(r[1])
                prod = prod_map.get(nm, {})
                okrug_products.setdefault(ok, []).append({
                    "vendor_code": prod.get("vendor_code", str(nm)),
                    "name": prod.get("name", ""),
                    "orders": int(r[2]),
                    "stability_pct": round(int(r[3]) / period * 100, 1),
                })

        # ── 4. Warehouse stock + cross-delivery data ──
        stock_rows = ch.query("""
            SELECT warehouse_name, sum(qty) AS total_stock
            FROM (
                SELECT warehouse_name, nm_id, argMax(quantity, fetched_at) AS qty
                FROM mms_analytics.fact_inventory_snapshot
                WHERE shop_id = {shop_id:UInt32}
                  AND warehouse_name NOT LIKE 'FBS:%'
                GROUP BY warehouse_name, nm_id
                HAVING qty > 0
            )
            GROUP BY warehouse_name
            ORDER BY total_stock DESC
        """, parameters={"shop_id": shop_id}).result_rows

        warehouses_context = []
        for r in stock_rows:
            wh_name = str(r[0])
            warehouses_context.append({
                "warehouse": wh_name,
                "stock": int(r[1]),
                "okrug": _wh_to_okrug(wh_name),
            })

        # Cross-delivery stats per okrug
        cross_rows = ch.query("""
            SELECT
                oblast_okrug_name AS order_okrug,
                warehouse_name,
                count() AS orders
            FROM mms_analytics.fact_orders_raw FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND date >= {d_start:Date}
              AND is_cancel = 0
            GROUP BY order_okrug, warehouse_name
            ORDER BY orders DESC
        """, parameters=params).result_rows

        # Calculate cross-delivery % per okrug
        okrug_cross: dict[str, dict] = {}  # okrug → {total, cross, main_wh}
        for r in cross_rows:
            order_okrug = str(r[0]) or "?"
            wh_name = str(r[1])
            orders = int(r[2])
            wh_okrug = _wh_to_okrug(wh_name)

            if order_okrug not in okrug_cross:
                okrug_cross[order_okrug] = {"total": 0, "cross": 0, "main_wh": ""}
            okrug_cross[order_okrug]["total"] += orders
            if wh_okrug != order_okrug and order_okrug != "Другой":
                okrug_cross[order_okrug]["cross"] += orders
            if not okrug_cross[order_okrug]["main_wh"] or orders > 0:
                okrug_cross[order_okrug]["main_wh"] = wh_name

        # ── 5. Advertising data removed ──
        # Реклама НЕ передаётся в географический анализ,
        # т.к. на WB нельзя таргетировать рекламу по регионам.
        # Анализ рекламы — отдельный раздел системы.

        ch.close()

        # ── 6. Build prompt ──
        prompt = f"""Магазин: {shop.name} (Wildberries)
Период: {period} дней (с {d_start} по {today})
Всего заказов: {total_orders}
Всего выручка: {round(total_revenue)}₽
Регионов с заказами: {len(set(str(r[1]) for r in geo_rows))}
Округов: {len(okrug_summary)}

## ОКРУГА И РЕГИОНЫ:
"""
        for ok_name, ok_data in sorted(okrug_summary.items(), key=lambda x: x[1]["orders"], reverse=True):
            share = round(ok_data["orders"] / total_orders * 100, 1) if total_orders > 0 else 0
            stab = ok_data.get("stability_pct", 0)
            prompt += f"\n### {ok_name}: {ok_data['orders']} заказов ({share}%), выручка {round(ok_data['revenue'])}₽, стабильность {stab}%\n"
            # Cross info
            cross = okrug_cross.get(ok_name, {})
            if cross:
                cross_pct = round(cross.get("cross", 0) / cross.get("total", 1) * 100)
                prompt += f"  Кросс-доставка: {cross_pct}%, основной склад: {cross.get('main_wh', '?')}\n"
            # Regions
            for reg in sorted(ok_data["regions"], key=lambda x: x["orders"], reverse=True)[:10]:
                prompt += f"  - {reg['region']}: {reg['orders']} зак, {reg['revenue']}₽, стаб {reg['stability_pct']}%\n"
            # Top products in okrug
            ok_prods = okrug_products.get(ok_name, [])
            if ok_prods:
                prompt += "  Топ товары в округе:\n"
                for p in ok_prods[:3]:
                    prompt += f"    · {p['vendor_code']} ({p['name'][:40]}): {p['orders']} зак, стаб {p['stability_pct']}%\n"

        prompt += "\n## СКЛАДЫ (остатки):\n"
        for wh in warehouses_context[:10]:
            prompt += f"- {wh['warehouse']}: {wh['stock']} шт (округ: {wh['okrug']})\n"

        prompt += "\n## ТОП ТОВАРЫ (общие):\n"
        for p in products_context[:15]:
            prompt += f"- {p['vendor_code']} ({p['name'][:40]}): {p['orders']} заказов, {p['revenue']}₽, стаб {p['stability_pct']}%, в {p['okrug_count']} округах, {p['region_count']} регионах\n"

        prompt += "\nПроанализируй ГЕОГРАФИЧЕСКОЕ распределение спроса и выдай JSON. Фокусируйся ТОЛЬКО на географии, НЕ анализируй рекламу."

        # ── 7. Call Gemini ──
        KIE_AI_URL = "https://api.kie.ai/gemini-2.5-flash/v1/chat/completions"

        async with httpx.AsyncClient(timeout=120.0) as client:
            resp = await client.post(
                KIE_AI_URL,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {api_key}",
                },
                json={
                    "messages": [
                        {"role": "system", "content": [{"type": "text", "text": _AI_PROMPT_GEOGRAPHY}]},
                        {"role": "user", "content": [{"type": "text", "text": prompt}]},
                    ],
                    "stream": False,
                    "include_thoughts": False,
                },
            )

        if resp.status_code != 200:
            logger.error("Gemini API error %s: %s", resp.status_code, resp.text[:500])
            raise HTTPException(status_code=502, detail="AI API error")

        resp_json = resp.json()
        content = resp_json.get("choices", [{}])[0].get("message", {}).get("content", "")

        # Strip markdown code fences
        content = content.strip()
        if content.startswith("```"):
            content = content.split("\n", 1)[1] if "\n" in content else content[3:]
        if content.endswith("```"):
            content = content[:-3]
        content = content.strip()
        if content.startswith("json"):
            content = content[4:].strip()

        try:
            ai_result = json.loads(content)
        except json.JSONDecodeError:
            logger.warning("Failed to parse AI geography JSON: %s", content[:500])
            raise HTTPException(status_code=502, detail="AI returned invalid JSON")

        # Enrich with context
        ai_result["period_days"] = period
        ai_result["analyzed_at"] = int(time.time())
        ai_result["context"] = {
            "total_orders": total_orders,
            "total_revenue": round(total_revenue),
            "total_okrugs": len(okrug_summary),
            "total_regions": len(set(str(r[1]) for r in geo_rows)),
            "warehouses_count": len(warehouses_context),
        }

        # Cache
        _ai_cache[cache_key] = (time.time(), ai_result)

        return ai_result

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Geography AI analysis failed")
        raise HTTPException(status_code=500, detail=f"Ошибка анализа: {str(e)}")


# ═══════════════════════════════════════════════════════════════
# Ozon Sales Geography
# ═══════════════════════════════════════════════════════════════

@router.get("/ozon/geography")
async def get_ozon_geography(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    skus: str = Query(None, description="Comma-separated SKUs to filter by"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Ozon sales geography — clusters with stability, avg check, drill-down by city."""
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "ozon":
        raise HTTPException(status_code=404, detail="Shop not found")

    from app.core.clickhouse import get_clickhouse_client

    ch = get_clickhouse_client()
    today = date.today()
    d_start = today - timedelta(days=period)

    # Parse multi-SKU filter
    sku_list: list[int] = []
    sku_filter = ""
    params: dict = {"shop_id": shop_id, "d_start": d_start, "period_days": period}
    if skus:
        try:
            sku_list = [int(x.strip()) for x in skus.split(",") if x.strip()]
        except ValueError:
            pass
        if sku_list:
            id_str = ", ".join(str(x) for x in sku_list)
            sku_filter = f"AND sku IN ({id_str})"

    # ── 1. Orders by cluster_to + city with stability ──
    rows = ch.query(f"""
        SELECT
            cluster_to AS cluster,
            city,
            count() AS orders,
            sum(toFloat64(price) * quantity) AS revenue,
            count(DISTINCT toDate(order_date)) AS active_days
        FROM mms_analytics.fact_ozon_orders FINAL
        WHERE shop_id = {{shop_id:UInt32}}
          AND order_date >= {{d_start:Date}}
          AND status NOT IN ('cancelled', 'canceled')
          AND cluster_to != ''
          {sku_filter}
        GROUP BY cluster, city
        ORDER BY orders DESC
    """, parameters=params).result_rows

    total_orders = sum(int(r[2]) for r in rows)
    total_revenue = sum(float(r[3]) for r in rows)

    # ── 1b. Stability per cluster (unique days with orders) ──
    cluster_stability_rows = ch.query(f"""
        SELECT
            cluster_to AS cluster,
            count(DISTINCT toDate(order_date)) AS active_days
        FROM mms_analytics.fact_ozon_orders FINAL
        WHERE shop_id = {{shop_id:UInt32}}
          AND order_date >= {{d_start:Date}}
          AND status NOT IN ('cancelled', 'canceled')
          AND cluster_to != ''
          {sku_filter}
        GROUP BY cluster
    """, parameters=params).result_rows
    cluster_active_days: dict[str, int] = {}
    for r in cluster_stability_rows:
        cluster_active_days[str(r[0]) or "Не определено"] = int(r[1])

    # Group by cluster
    cluster_data: dict[str, dict] = {}
    for r in rows:
        cluster = str(r[0]) or "Не определено"
        city = str(r[1]) or "Не определено"
        orders = int(r[2])
        revenue = float(r[3])
        active_days = int(r[4])

        if cluster not in cluster_data:
            cluster_data[cluster] = {
                "cluster": cluster,
                "orders": 0,
                "revenue": 0,
                "share_pct": 0,
                "avg_check": 0,
                "stability_pct": 0,
                "cities": [],
            }
        cluster_data[cluster]["orders"] += orders
        cluster_data[cluster]["revenue"] += revenue
        city_avg_check = round(revenue / orders, 2) if orders > 0 else 0
        city_stability = round(active_days / period * 100, 1)
        cluster_data[cluster]["cities"].append({
            "city": city,
            "orders": orders,
            "revenue": round(revenue, 2),
            "avg_check": city_avg_check,
            "stability_pct": city_stability,
            "share_pct": round(orders / total_orders * 100, 1) if total_orders > 0 else 0,
        })

    clusters_result = []
    for cluster_name, cd in sorted(cluster_data.items(), key=lambda x: x[1]["orders"], reverse=True):
        cd["share_pct"] = round(cd["orders"] / total_orders * 100, 1) if total_orders > 0 else 0
        cd["revenue"] = round(cd["revenue"], 2)
        cd["avg_check"] = round(cd["revenue"] / cd["orders"], 2) if cd["orders"] > 0 else 0
        active_d = cluster_active_days.get(cluster_name, 0)
        cd["stability_pct"] = round(active_d / period * 100, 1)
        cd["cities"] = sorted(cd["cities"], key=lambda x: x["orders"], reverse=True)
        clusters_result.append(cd)

    # ── 2. Top products ──
    top_products = []
    prod_map: dict[int, dict] = {}
    if not sku_list:
        prod_rows = ch.query(f"""
            SELECT
                sku,
                count() AS orders,
                sum(toFloat64(price) * quantity) AS revenue,
                count(DISTINCT cluster_to) AS cluster_count,
                count(DISTINCT city) AS city_count,
                uniqExact(toDate(order_date)) AS active_days
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {{shop_id:UInt32}}
              AND order_date >= {{d_start:Date}}
              AND status NOT IN ('cancelled', 'canceled')
              AND cluster_to != ''
            GROUP BY sku
            ORDER BY orders DESC
            LIMIT 50
        """, parameters=params).result_rows

        prod_skus = [int(r[0]) for r in prod_rows]
        if prod_skus:
            sku_list_str = ", ".join(str(x) for x in prod_skus)
            pg_rows = (await db.execute(
                text(f"""
                    SELECT sku, offer_id, name
                    FROM dim_ozon_products
                    WHERE shop_id = :sid AND sku IN ({sku_list_str})
                """),
                {"sid": shop_id},
            )).fetchall()
            for r in pg_rows:
                prod_map[r[0]] = {"offer_id": r[1] or "", "name": (r[2] or "")[:80]}

        for r in prod_rows:
            sku_id = int(r[0])
            prod = prod_map.get(sku_id, {})
            p_orders = int(r[1])
            p_rev = float(r[2])
            active_days = int(r[5])
            top_products.append({
                "sku": sku_id,
                "offer_id": prod.get("offer_id", ""),
                "name": prod.get("name", ""),
                "orders": p_orders,
                "revenue": round(p_rev, 2),
                "avg_check": round(p_rev / p_orders, 2) if p_orders > 0 else 0,
                "cluster_count": int(r[3]),
                "city_count": int(r[4]),
                "stability_pct": round(active_days / period * 100, 1),
                "share_pct": round(p_orders / total_orders * 100, 1) if total_orders > 0 else 0,
            })

    # ── 3. Per-cluster top products ──
    cluster_top_products: dict[str, list] = {}
    if not sku_list:
        ctop_rows = ch.query(f"""
            SELECT
                cluster_to AS cluster,
                sku,
                count() AS orders,
                sum(toFloat64(price) * quantity) AS revenue,
                uniqExact(toDate(order_date)) AS active_days
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {{shop_id:UInt32}}
              AND order_date >= {{d_start:Date}}
              AND status NOT IN ('cancelled', 'canceled')
              AND cluster_to != ''
            GROUP BY cluster, sku
            ORDER BY cluster, orders DESC
        """, parameters=params).result_rows

        current_cluster = ""
        count = 0
        for r in ctop_rows:
            cluster = str(r[0]) or "Не определено"
            if cluster != current_cluster:
                current_cluster = cluster
                count = 0
            count += 1
            if count <= 5:
                sku_id = int(r[1])
                prod = prod_map.get(sku_id, {})
                active_days = int(r[4])
                cluster_top_products.setdefault(cluster, []).append({
                    "sku": sku_id,
                    "offer_id": prod.get("offer_id", ""),
                    "name": prod.get("name", ""),
                    "orders": int(r[2]),
                    "revenue": round(float(r[3]), 2),
                    "avg_check": round(float(r[3]) / int(r[2]), 2) if int(r[2]) > 0 else 0,
                    "stability_pct": round(active_days / period * 100, 1),
                })

    # ── 4. SKU filter info ──
    sku_filter_info = []
    if sku_list:
        sku_list_str = ", ".join(str(x) for x in sku_list)
        pg_rows = (await db.execute(
            text(f"""
                SELECT sku, offer_id, name
                FROM dim_ozon_products
                WHERE shop_id = :sid AND sku IN ({sku_list_str})
            """),
            {"sid": shop_id},
        )).fetchall()
        for r in pg_rows:
            sku_filter_info.append({
                "sku": r[0],
                "offer_id": r[1] or "",
                "name": (r[2] or "")[:80],
            })

    # ── 5. Total unique cities count ──
    total_cities = len(set(
        str(r[1]) for r in rows if str(r[1])
    ))
    total_clusters = len(cluster_data)

    avg_check = round(total_revenue / total_orders, 2) if total_orders > 0 else 0

    ch.close()

    return {
        "total_orders": total_orders,
        "total_revenue": round(total_revenue, 2),
        "avg_check": avg_check,
        "total_clusters": total_clusters,
        "total_cities": total_cities,
        "period_days": period,
        "clusters": clusters_result,
        "top_products": top_products,
        "cluster_top_products": cluster_top_products,
        "sku_filter": sku_filter_info,
    }


@router.get("/ozon/geography/city-products")
async def get_ozon_geography_city_products(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    city: str = Query(..., description="City name"),
    skus: str = Query(None, description="Comma-separated SKUs to filter by"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Top products for a specific city (drill-down)."""
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "ozon":
        raise HTTPException(status_code=404, detail="Shop not found")

    from app.core.clickhouse import get_clickhouse_client

    ch = get_clickhouse_client()
    today = date.today()
    d_start = today - timedelta(days=period)

    sku_filter = ""
    if skus:
        try:
            sku_id_list = [int(x.strip()) for x in skus.split(",") if x.strip()]
            if sku_id_list:
                id_str = ", ".join(str(x) for x in sku_id_list)
                sku_filter = f"AND sku IN ({id_str})"
        except ValueError:
            pass

    total_weeks = max(1, period // 7)

    prod_rows = ch.query(f"""
        SELECT
            sku,
            count() AS orders,
            sum(toFloat64(price) * quantity) AS revenue,
            uniqExact(toMonday(order_date)) AS active_weeks
        FROM mms_analytics.fact_ozon_orders FINAL
        WHERE shop_id = {{shop_id:UInt32}}
          AND order_date >= {{d_start:Date}}
          AND status NOT IN ('cancelled', 'canceled')
          AND city = {{city:String}}
          {sku_filter}
        GROUP BY sku
        ORDER BY orders DESC
        LIMIT 20
    """, parameters={"shop_id": shop_id, "d_start": d_start, "city": city}).result_rows

    ch.close()

    sku_ids = [int(r[0]) for r in prod_rows]
    prod_map: dict[int, dict] = {}
    if sku_ids:
        sku_list_str = ", ".join(str(x) for x in sku_ids)
        pg_rows = (await db.execute(
            text(f"""
                SELECT sku, offer_id, name
                FROM dim_ozon_products
                WHERE shop_id = :sid AND sku IN ({sku_list_str})
            """),
            {"sid": shop_id},
        )).fetchall()
        for r in pg_rows:
            prod_map[r[0]] = {"offer_id": r[1] or "", "name": (r[2] or "")[:80]}

    result = []
    for r in prod_rows:
        sku_id = int(r[0])
        prod = prod_map.get(sku_id, {})
        p_orders = int(r[1])
        p_rev = float(r[2])
        active_weeks = int(r[3])
        stability_pct = round(active_weeks / total_weeks * 100, 1)
        result.append({
            "sku": sku_id,
            "offer_id": prod.get("offer_id", ""),
            "name": prod.get("name", ""),
            "orders": p_orders,
            "revenue": round(p_rev, 2),
            "avg_check": round(p_rev / p_orders, 2) if p_orders > 0 else 0,
            "stability_pct": stability_pct,
        })

    return {"city": city, "products": result}


@router.get("/ozon/geography/products-search")
async def get_ozon_geography_products_search(
    shop_id: int = Query(...),
    q: str = Query("", description="Search query (name, offer_id, or sku)"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Search Ozon products for autocomplete in geography filter."""
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "ozon":
        raise HTTPException(status_code=404, detail="Shop not found")

    search = q.strip()
    if not search:
        pg_rows = (await db.execute(
            text("""
                SELECT sku, offer_id, name
                FROM dim_ozon_products
                WHERE shop_id = :sid
                ORDER BY name
                LIMIT 100
            """),
            {"sid": shop_id},
        )).fetchall()
    else:
        # Try numeric search (sku)
        try:
            sku_int = int(search)
            pg_rows = (await db.execute(
                text("""
                    SELECT sku, offer_id, name
                    FROM dim_ozon_products
                    WHERE shop_id = :sid AND sku = :sku_val
                    LIMIT 20
                """),
                {"sid": shop_id, "sku_val": sku_int},
            )).fetchall()
            if not pg_rows:
                pg_rows = (await db.execute(
                    text("""
                        SELECT sku, offer_id, name
                        FROM dim_ozon_products
                        WHERE shop_id = :sid AND CAST(sku AS TEXT) LIKE :pattern
                        ORDER BY name
                        LIMIT 20
                    """),
                    {"sid": shop_id, "pattern": f"%{search}%"},
                )).fetchall()
        except ValueError:
            pg_rows = (await db.execute(
                text("""
                    SELECT sku, offer_id, name
                    FROM dim_ozon_products
                    WHERE shop_id = :sid
                      AND (LOWER(name) LIKE :q OR LOWER(offer_id) LIKE :q)
                    ORDER BY name
                    LIMIT 30
                """),
                {"sid": shop_id, "q": f"%{search.lower()}%"},
            )).fetchall()

    result = []
    for r in pg_rows:
        result.append({
            "sku": r[0],
            "offer_id": r[1] or "",
            "name": (r[2] or "")[:80],
        })

    return {"products": result}


# ═══════════════════════════════════════════════════════════════
# Ozon Geography AI Analysis
# ═══════════════════════════════════════════════════════════════

_AI_PROMPT_OZON_GEOGRAPHY = """Ты — эксперт по ГЕОГРАФИИ ПРОДАЖ на Ozon.

ТВОЯ ЗАДАЧА: проанализировать ГЕОГРАФИЧЕСКОЕ распределение спроса по кластерам доставки Ozon и выдать конкретные выводы.

## СТРОГИЕ ГРАНИЦЫ АНАЛИЗА

АНАЛИЗИРУЙ ТОЛЬКО:
- Географическое распределение заказов по кластерам доставки и городам
- Тренды: какие кластеры растут, какие стагнируют (по стабильности спроса)
- Концентрация: насколько равномерно распределены продажи
- Недопокрытые кластеры: где есть спрос но мало/нет стока на ближайших РФЦ
- Географический охват товаров: в скольких кластерах продаётся каждый товар
- Потенциал роста в новых кластерах

НЕ АНАЛИЗИРУЙ (для этого есть ОТДЕЛЬНЫЕ разделы в системе):
- ❌ Детали рекламы, DRR, рекламные кампании → раздел "Реклама"
- ❌ Финансовые показатели (маржа, прибыль) → раздел "Финансы"

ВАЖНО ПРО OZON:
- Ozon использует ~15-25 КЛАСТЕРОВ ДОСТАВКИ (Москва МО и Дальние регионы, Санкт-Петербург и СЗО, Казань, Краснодар и т.д.)
- Есть ~34 РФЦ (распределительных фулфилмент-центров), они маппятся на кластеры
- Товар отгружается с ближайшего РФЦ → если на нём нет стока, идёт кросс-доставка с другого РФЦ
- На Ozon можно частично таргетировать рекламу по регионам, но основной рычаг — РАЗМЕЩЕНИЕ СТОКА на нужных РФЦ
- offer_id — это артикул продавца (аналог vendor_code на WB)

## ЛОГИКА АНАЛИЗА

### 1. Концентрация (ОБЯЗАТЕЛЬНО)
- Посчитай: какой % выручки дают топ-3 кластера
- Если > 70% — ВЫСОКАЯ концентрация (risk_level: "high")
- Если 40-70% — СРЕДНЯЯ (risk_level: "medium")
- Если < 40% — НИЗКАЯ, хорошая диверсификация (risk_level: "low")
- НАЗОВИ конкретные кластеры и их долю

### 2. Инсайты по товарам (2-5 штук) — ТОЛЬКО географические!
Для каждого товара определи тип:
- **stable_leader** — высокая стабильность + много кластеров → лидер, масштабировать географию
- **unstable_demand** — стабильность <15% → нестабильный спрос, возможны проблемы с наличием
- **regional_champion** — продаётся хорошо в 1-2 кластерах, но отсутствует в остальных → потенциал расширения
- **cross_delivery_problem** — товар продаётся в кластере, но на его РФЦ нет стока → нужен сток ближе
- **dead_stock_risk** — есть на складе, но нет заказов из ближайших кластеров → нет спроса

Для КАЖДОГО инсайта:
- Укажи offer_id и name
- Укажи конкретные числа: заказы, стабильность, в скольких кластерах
- Дай конкретное действие (action): redistribute / increase_supply / monitor
- Рассчитай expected_effect — потенциальную выручку или экономию в РУБЛЯХ

### 3. Логистическое соответствие (1-3 самых важных)
Оценка: насколько расположение стока на РФЦ соответствует географии спроса.
Для каждого кластера где сток не на ближайшем РФЦ:
- Укажи: откуда реально отгружается
- Укажи ближайший РФЦ и сток на нём
- Дай рекомендацию: куда и сколько штук разместить

## ФОРМАТ ОТВЕТА — СТРОГО JSON:
{
  "severity": "critical" | "warning" | "ok",
  "diagnosis": "Краткий главный географический вывод из 1-2 предложений с конкретными числами",
  "key_metrics": {
    "concentration_pct": 80,
    "top_clusters_count": 3,
    "total_clusters": 15,
    "clusters_with_stable_demand": 8,
    "underserved_clusters": 2
  },
  "concentration": {
    "summary": "80% выручки из кластеров Москва, СПб и Краснодар",
    "top_regions": [
      {"region": "Москва, МО и Дальние регионы", "orders": 200, "share_pct": 33.7, "stability_pct": 95}
    ],
    "risk_level": "high" | "medium" | "low",
    "recommendation": "Конкретная рекомендация по диверсификации географии через размещение стока"
  },
  "product_insights": [
    {
      "vendor_code": "АРТИКУЛ (offer_id)",
      "name": "Название",
      "insight_type": "stable_leader",
      "regions_count": 12,
      "stability_pct": 47,
      "orders": 21,
      "detail": "Географический анализ с числами",
      "action": "redistribute",
      "expected_effect": "Потенциальная выручка +50 000₽/мес при расширении на кластер Екатеринбург"
    }
  ],
  "logistics_match": [
    {
      "okrug": "Казань (кластер)",
      "orders": 50,
      "nearest_warehouse": "Казань РФЦ",
      "warehouse_stock": 0,
      "serving_warehouse": "Тверь РФЦ",
      "cross_pct": 100,
      "detail": "Описание проблемы",
      "recommendation": "Разместить N штук на РФЦ X"
    }
  ],
  "general_tips": [
    "Конкретная географическая рекомендация, 1-2 предложения."
  ]
}

## ПРАВИЛА
- severity: "critical" если концентрация > 70% или есть кластеры с 100% кросс-доставкой, "warning" если концентрация 50-70%, "ok" если всё сбалансировано
- product_insights: 2-5 ГЕОГРАФИЧЕСКИХ инсайтов. НАЗЫВАЙ товары по offer_id + name
- insight_type: ТОЛЬКО "stable_leader", "unstable_demand", "regional_champion", "cross_delivery_problem", "dead_stock_risk"
- action: ТОЛЬКО "redistribute", "increase_supply", "monitor"
- logistics_match: 1-3 проблемных кластера. Если проблем нет — пустой массив
- general_tips: 2-3 совета ПО ГЕОГРАФИИ. Только про размещение стока на РФЦ и диверсификацию
- top_regions в concentration: топ-5 кластеров по выручке
- Все числа — из реальных данных, НЕ выдумывай
- Пиши НА РУССКОМ
"""


@router.post("/ozon/geography/ai-analysis")
async def get_ozon_geography_ai_analysis(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    force: bool = Query(False, description="Skip cache"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """AI-powered Ozon geography sales analysis using Gemini 2.5 Flash."""
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "ozon":
        raise HTTPException(status_code=404, detail="Shop not found")

    cache_key = f"ozon_geo_ai_{shop_id}_{period}"
    if not force and cache_key in _ai_cache:
        ts, cached = _ai_cache[cache_key]
        if time.time() - ts < _AI_CACHE_TTL:
            return {**cached, "cached": True}

    api_key = os.getenv("KIE_AI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="AI API key not configured")

    try:
        from app.core.clickhouse import get_clickhouse_client

        ch = get_clickhouse_client()
        today = date.today()
        d_start = today - timedelta(days=period)
        params = {"shop_id": shop_id, "d_start": d_start, "period_days": period}

        # ── 1. Orders by cluster_to + city with stability ──
        geo_rows = ch.query("""
            SELECT
                cluster_to AS cluster,
                city,
                count() AS orders,
                sum(toFloat64(price) * quantity) AS revenue,
                count(DISTINCT toDate(order_date)) AS active_days
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= {d_start:Date}
              AND status NOT IN ('cancelled', 'canceled')
              AND cluster_to != ''
            GROUP BY cluster, city
            ORDER BY orders DESC
        """, parameters=params).result_rows

        total_orders = sum(int(r[2]) for r in geo_rows)
        total_revenue = sum(float(r[3]) for r in geo_rows)

        # Build cluster summary
        cluster_summary: dict[str, dict] = {}
        for r in geo_rows:
            cluster = str(r[0]) or "Не определено"
            city = str(r[1]) or "Не определено"
            orders = int(r[2])
            revenue = float(r[3])
            active_days = int(r[4])

            if cluster not in cluster_summary:
                cluster_summary[cluster] = {"orders": 0, "revenue": 0, "cities": []}
            cluster_summary[cluster]["orders"] += orders
            cluster_summary[cluster]["revenue"] += revenue
            cluster_summary[cluster]["cities"].append({
                "city": city,
                "orders": orders,
                "revenue": round(revenue),
                "stability_pct": round(active_days / period * 100, 1),
            })

        # Cluster-level stability
        cluster_stab_rows = ch.query("""
            SELECT
                cluster_to AS cluster,
                count(DISTINCT toDate(order_date)) AS active_days
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= {d_start:Date}
              AND status NOT IN ('cancelled', 'canceled')
              AND cluster_to != ''
            GROUP BY cluster
        """, parameters=params).result_rows
        for r in cluster_stab_rows:
            cluster = str(r[0]) or "Не определено"
            if cluster in cluster_summary:
                cluster_summary[cluster]["stability_pct"] = round(int(r[1]) / period * 100, 1)

        # ── 2. Top products with geography spread ──
        prod_rows = ch.query("""
            SELECT
                sku,
                count() AS orders,
                sum(toFloat64(price) * quantity) AS revenue,
                count(DISTINCT cluster_to) AS cluster_count,
                count(DISTINCT city) AS city_count,
                uniqExact(toDate(order_date)) AS active_days
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= {d_start:Date}
              AND status NOT IN ('cancelled', 'canceled')
            GROUP BY sku
            ORDER BY orders DESC
            LIMIT 30
        """, parameters=params).result_rows

        prod_skus = [int(r[0]) for r in prod_rows]
        prod_map: dict[int, dict] = {}
        if prod_skus:
            sku_list = ", ".join(str(x) for x in prod_skus)
            pg_rows = (await db.execute(
                text(f"""
                    SELECT sku, offer_id, name
                    FROM dim_ozon_products
                    WHERE shop_id = :sid AND sku IN ({sku_list})
                """),
                {"sid": shop_id},
            )).fetchall()
            for r in pg_rows:
                prod_map[r[0]] = {"offer_id": r[1] or "", "name": (r[2] or "")[:60]}

        products_context = []
        for r in prod_rows:
            sku = int(r[0])
            prod = prod_map.get(sku, {})
            sku_orders = int(r[1])
            active_days = int(r[5])
            products_context.append({
                "sku": sku,
                "offer_id": prod.get("offer_id", ""),
                "name": prod.get("name", f"SKU {sku}"),
                "orders": sku_orders,
                "revenue": round(float(r[2])),
                "cluster_count": int(r[3]),
                "city_count": int(r[4]),
                "stability_pct": round(active_days / period * 100, 1),
                "share_pct": round(sku_orders / total_orders * 100, 1) if total_orders > 0 else 0,
            })

        # ── 3. Per-cluster top products ──
        ctop_rows = ch.query("""
            SELECT
                cluster_to AS cluster,
                sku,
                count() AS orders,
                uniqExact(toDate(order_date)) AS active_days
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= {d_start:Date}
              AND status NOT IN ('cancelled', 'canceled')
              AND cluster_to != ''
            GROUP BY cluster, sku
            ORDER BY cluster, orders DESC
        """, parameters=params).result_rows

        cluster_products: dict[str, list] = {}
        current_cl = ""
        cnt = 0
        for r in ctop_rows:
            cl = str(r[0]) or "?"
            if cl != current_cl:
                current_cl = cl
                cnt = 0
            cnt += 1
            if cnt <= 5:
                sku = int(r[1])
                prod = prod_map.get(sku, {})
                cluster_products.setdefault(cl, []).append({
                    "offer_id": prod.get("offer_id", str(sku)),
                    "name": prod.get("name", ""),
                    "orders": int(r[2]),
                    "stability_pct": round(int(r[3]) / period * 100, 1),
                })

        # ── 4. Warehouse stock per cluster (via WAREHOUSE_TO_CLUSTER) ──
        stock_rows = ch.query("""
            SELECT warehouse_name, sum(qty) AS total_stock
            FROM (
                SELECT warehouse_name, offer_id, argMax(free_to_sell, updated_at) AS qty
                FROM mms_analytics.fact_ozon_warehouse_stocks
                WHERE shop_id = {shop_id:UInt32}
                GROUP BY warehouse_name, offer_id
                HAVING qty > 0
            )
            GROUP BY warehouse_name
            ORDER BY total_stock DESC
        """, parameters={"shop_id": shop_id}).result_rows

        warehouses_context = []
        for r in stock_rows:
            wh_name = str(r[0])
            cluster = WAREHOUSE_TO_CLUSTER.get(wh_name, "Другой")
            warehouses_context.append({
                "warehouse": wh_name,
                "stock": int(r[1]),
                "cluster": cluster,
            })

        # ── 5. Cross-delivery stats: orders cluster vs warehouse cluster ──
        cross_rows = ch.query("""
            SELECT
                cluster_to AS order_cluster,
                warehouse_name,
                count() AS orders
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= {d_start:Date}
              AND status NOT IN ('cancelled', 'canceled')
              AND cluster_to != ''
              AND warehouse_name != ''
            GROUP BY order_cluster, warehouse_name
            ORDER BY orders DESC
        """, parameters=params).result_rows

        cluster_cross: dict[str, dict] = {}
        for r in cross_rows:
            order_cluster = str(r[0]) or "?"
            wh_name = str(r[1])
            orders = int(r[2])
            wh_cluster = WAREHOUSE_TO_CLUSTER.get(wh_name, "Другой")

            if order_cluster not in cluster_cross:
                cluster_cross[order_cluster] = {"total": 0, "cross": 0, "main_wh": ""}
            cluster_cross[order_cluster]["total"] += orders
            if wh_cluster != order_cluster and order_cluster != "Другой":
                cluster_cross[order_cluster]["cross"] += orders
            if not cluster_cross[order_cluster]["main_wh"] or orders > 0:
                cluster_cross[order_cluster]["main_wh"] = wh_name

        ch.close()

        # ── 6. Build prompt ──
        prompt = f"""Магазин: {shop.name} (Ozon)
Период: {period} дней (с {d_start} по {today})
Всего заказов: {total_orders}
Всего выручка: {round(total_revenue)}₽
Городов с заказами: {len(set(str(r[1]) for r in geo_rows))}
Кластеров: {len(cluster_summary)}

## КЛАСТЕРЫ ДОСТАВКИ И ГОРОДА:
"""
        for cl_name, cl_data in sorted(cluster_summary.items(), key=lambda x: x[1]["orders"], reverse=True):
            share = round(cl_data["orders"] / total_orders * 100, 1) if total_orders > 0 else 0
            stab = cl_data.get("stability_pct", 0)
            prompt += f"\n### {cl_name}: {cl_data['orders']} заказов ({share}%), выручка {round(cl_data['revenue'])}₽, стабильность {stab}%\n"
            # Cross info
            cross = cluster_cross.get(cl_name, {})
            if cross:
                cross_pct = round(cross.get("cross", 0) / cross.get("total", 1) * 100)
                prompt += f"  Кросс-доставка: {cross_pct}%, основной РФЦ: {cross.get('main_wh', '?')}\n"
            # Cities
            for city in sorted(cl_data["cities"], key=lambda x: x["orders"], reverse=True)[:10]:
                prompt += f"  - {city['city']}: {city['orders']} зак, {city['revenue']}₽, стаб {city['stability_pct']}%\n"
            # Top products
            cl_prods = cluster_products.get(cl_name, [])
            if cl_prods:
                prompt += "  Топ товары в кластере:\n"
                for p in cl_prods[:3]:
                    prompt += f"    · {p['offer_id']} ({p['name'][:40]}): {p['orders']} зак, стаб {p['stability_pct']}%\n"

        prompt += "\n## СКЛАДЫ (РФЦ, остатки):\n"
        for wh in warehouses_context[:15]:
            prompt += f"- {wh['warehouse']}: {wh['stock']} шт (кластер: {wh['cluster']})\n"

        prompt += "\n## ТОП ТОВАРЫ (общие):\n"
        for p in products_context[:15]:
            prompt += f"- {p['offer_id']} ({p['name'][:40]}): {p['orders']} заказов, {p['revenue']}₽, стаб {p['stability_pct']}%, в {p['cluster_count']} кластерах, {p['city_count']} городах\n"

        prompt += "\nПроанализируй ГЕОГРАФИЧЕСКОЕ распределение спроса по кластерам доставки и выдай JSON."

        # ── 7. Call Gemini ──
        KIE_AI_URL = "https://api.kie.ai/gemini-2.5-flash/v1/chat/completions"

        async with httpx.AsyncClient(timeout=120.0) as client:
            resp = await client.post(
                KIE_AI_URL,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {api_key}",
                },
                json={
                    "messages": [
                        {"role": "system", "content": [{"type": "text", "text": _AI_PROMPT_OZON_GEOGRAPHY}]},
                        {"role": "user", "content": [{"type": "text", "text": prompt}]},
                    ],
                    "stream": False,
                    "include_thoughts": False,
                },
            )

        if resp.status_code != 200:
            logger.error("Gemini API error %s: %s", resp.status_code, resp.text[:500])
            raise HTTPException(status_code=502, detail="AI API error")

        resp_json = resp.json()
        content = resp_json.get("choices", [{}])[0].get("message", {}).get("content", "")

        # Strip markdown code fences
        content = content.strip()
        if content.startswith("```"):
            content = content.split("\n", 1)[1] if "\n" in content else content[3:]
        if content.endswith("```"):
            content = content[:-3]
        content = content.strip()
        if content.startswith("json"):
            content = content[4:].strip()

        try:
            ai_result = json.loads(content)
        except json.JSONDecodeError:
            logger.warning("Failed to parse Ozon geo AI JSON: %s", content[:500])
            raise HTTPException(status_code=502, detail="AI returned invalid JSON")

        # Enrich with context
        ai_result["period_days"] = period
        ai_result["analyzed_at"] = int(time.time())
        ai_result["context"] = {
            "total_orders": total_orders,
            "total_revenue": round(total_revenue),
            "total_okrugs": len(cluster_summary),
            "total_regions": len(set(str(r[1]) for r in geo_rows)),
            "warehouses_count": len(warehouses_context),
        }

        # Cache
        _ai_cache[cache_key] = (time.time(), ai_result)

        return ai_result

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Ozon Geography AI analysis failed")
        raise HTTPException(status_code=500, detail=f"Ошибка анализа: {str(e)}")


# ═══════════════════════════════════════════════════════════════
# Ozon Cross-Logistics AI Analysis
# ═══════════════════════════════════════════════════════════════

_AI_PROMPT_OZON_CROSS = """Ты — эксперт по КРОСС-ЛОГИСТИКЕ на Ozon. Твоя задача — дать ОБЗОР кросс-проблем и ОЦЕНКУ складов.

## ВВОДНЫЕ
- **Кросс-доставка** — заказ из кластера X обслуживается складом из кластера Y (не ближайшим). Увеличивает стоимость и время.
- **offer_id** — артикул продавца.
- **daily_sales** — среднесуточные продажи (глобально).
- Ты НЕ СЧИТАЕШЬ конкретные количества для перемещений и поставок — это делает алгоритм в разделе «Поставки».

## ТЫ ОБЯЗАН
1. Оценить общую ситуацию с кросс-доставкой (severity, diagnosis)
2. Выделить проблемные SKU (где кросс выше 25%)
3. Оценить каждый склад — сколько кросса генерирует и почему
4. Дать текстовые рекомендации — что сделать, БЕЗ конкретных штук

## ФОРМАТ ОТВЕТА — строго JSON:
{
  "severity": "critical" | "warning" | "ok",
  "diagnosis": "1-3 предложения. Конкретные цифры кросса, кол-во проблемных SKU.",
  "key_metrics": {
    "cross_pct": 45,
    "cross_orders": 120,
    "total_orders": 267,
    "warehouses_with_cross": 5,
    "skus_with_high_cross": 8
  },
  "problem_skus": [
    {
      "offer_id": "АРТ-789",
      "name": "Название товара",
      "total_orders": 50,
      "cross_orders": 30,
      "cross_pct": 60,
      "stock_distribution": [
        {"warehouse": "ДОМОДЕДОВО_РФЦ", "stock": 150},
        {"warehouse": "САМАРА_РФЦ", "stock": 0}
      ],
      "top_cross_routes": [
        {"from_warehouse": "ДОМОДЕДОВО_РФЦ", "to_cluster": "Самара", "orders": 15}
      ],
      "recommendation": "Товар сконцентрирован на Домодедово, а 60% заказов уходит в Самару и Екатеринбург. Нужен довоз на склады этих кластеров."
    }
  ],
  "warehouse_assessments": [
    {
      "warehouse": "ГРИВНО_РФЦ",
      "cluster": "Москва, МО и Дальние регионы",
      "status": "critical" | "warning" | "ok",
      "total_orders": 200,
      "cross_orders": 80,
      "cross_pct": 40,
      "main_cross_destinations": ["Ярославль", "Беларусь", "Краснодар"],
      "assessment": "Склад генерирует 40% кросс. Основная нагрузка: Ярославль (25 заказов), Беларусь (18 заказов). Довоз товаров на склады в Ярославле и Краснодаре снизит кросс."
    }
  ],
  "priority_actions": [
    {
      "action": "Сформировать поставку на склады кластеров Ярославль, Краснодар и Самара для закрытия кросс-спроса",
      "impact": "Снизит кросс на ~120 заказов/месяц (~38% от текущего кросса)",
      "link_to_supply": true
    },
    {
      "action": "Перераспределить сток АМ-СОБ-МЕЛ-ЯГ-1 с Гривно на ближайшие к спросу склады",
      "impact": "Снизит кросс этого SKU с 67% до ~20%",
      "link_to_supply": false
    }
  ],
  "general_tips": [
    "Конкретный совет с цифрами."
  ]
}

## АБСОЛЮТНЫЕ ЗАПРЕТЫ
- ❌ НЕ СЧИТАЙ конкретные штуки для перемещений и поставок — для этого есть алгоритм в разделе «Поставки»
- ❌ НЕ выдумывай данные — используй ТОЛЬКО то, что есть в промпте
- ❌ НИКОГДА: «например», «допустим», «альтернативно», «рассмотреть»
- ✅ severity: "critical" если cross_pct > 40%, "warning" если 20-40%, "ok" если < 20%
- ✅ warehouse_assessments.status: "critical" если cross_pct склада > 50%, "warning" если 25-50%, "ok" если < 25%
- ✅ Пиши НА РУССКОМ
- ✅ В priority_actions ставь link_to_supply=true для действий, которые решаются через раздел «Поставки»
"""


@router.post("/ozon/cross/ai-analysis")
async def get_ozon_cross_ai_analysis(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    force: bool = Query(False, description="Skip cache"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """AI-powered Ozon cross-logistics analysis using Gemini 2.5 Flash."""
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "ozon":
        raise HTTPException(status_code=404, detail="Shop not found")

    cache_key = f"ozon_cross_ai_{shop_id}_{period}"
    if not force and cache_key in _ai_cache:
        ts, cached = _ai_cache[cache_key]
        if time.time() - ts < _AI_CACHE_TTL:
            return {**cached, "cached": True}

    api_key = os.getenv("KIE_AI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="AI API key not configured")

    try:
        from app.core.clickhouse import get_clickhouse_client

        ch = get_clickhouse_client()
        today = date.today()
        d_start = today - timedelta(days=period)
        params = {"shop_id": shop_id, "d_start": d_start, "period_days": period}

        # ── 1. Cross-delivery matrix: sku × warehouse × cluster_to ──
        cross_matrix_rows = ch.query("""
            SELECT
                sku, warehouse_name, cluster_to,
                count() AS orders, sum(quantity) AS qty
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= {d_start:Date}
              AND status NOT IN ('cancelled', 'canceled')
              AND cluster_to != '' AND warehouse_name != ''
            GROUP BY sku, warehouse_name, cluster_to
            ORDER BY orders DESC
        """, parameters=params).result_rows

        # Build per-SKU cross data
        sku_cross: dict[int, dict] = {}
        for r in cross_matrix_rows:
            sku = int(r[0])
            wh = str(r[1])
            cluster = str(r[2])
            orders = int(r[3])
            wh_cluster = _get_cluster_for_warehouse(wh)
            is_cross = wh_cluster != cluster

            if sku not in sku_cross:
                sku_cross[sku] = {"total": 0, "cross": 0, "routes": []}
            sku_cross[sku]["total"] += orders
            if is_cross:
                sku_cross[sku]["cross"] += orders
                sku_cross[sku]["routes"].append({
                    "from_wh": wh, "to_cluster": cluster, "orders": orders
                })

        # ── 2. Per-warehouse cross summary ──
        wh_cross_rows = ch.query("""
            SELECT
                warehouse_name,
                cluster_to,
                count() AS orders
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= {d_start:Date}
              AND status NOT IN ('cancelled', 'canceled')
              AND cluster_to != '' AND warehouse_name != ''
            GROUP BY warehouse_name, cluster_to
            ORDER BY orders DESC
        """, parameters=params).result_rows

        wh_summary: dict[str, dict] = {}
        total_orders = 0
        total_cross = 0
        for r in wh_cross_rows:
            wh = str(r[0])
            cluster = str(r[1])
            orders = int(r[2])
            wh_cluster = _get_cluster_for_warehouse(wh)
            is_cross = wh_cluster != cluster

            if wh not in wh_summary:
                wh_summary[wh] = {"cluster": wh_cluster, "total": 0, "cross": 0}
            wh_summary[wh]["total"] += orders
            total_orders += orders
            if is_cross:
                wh_summary[wh]["cross"] += orders
                total_cross += orders

        # ── 3. Stocks per SKU per warehouse ──
        stock_rows = ch.query("""
            SELECT warehouse_name, offer_id, sku,
                   argMax(free_to_sell, updated_at) AS stock
            FROM mms_analytics.fact_ozon_warehouse_stocks
            WHERE shop_id = {shop_id:UInt32}
            GROUP BY warehouse_name, offer_id, sku
            HAVING stock > 0
            ORDER BY stock DESC
        """, parameters={"shop_id": shop_id}).result_rows

        # Build stock map: sku -> {warehouse: stock}
        sku_stock: dict[int, dict[str, int]] = {}
        for r in stock_rows:
            wh = str(r[0])
            sku = int(r[2])
            stock = int(r[3])
            if sku not in sku_stock:
                sku_stock[sku] = {}
            sku_stock[sku][wh] = stock

        # ── 4. Product metadata from PostgreSQL (ALL SKUs with stock + cross) ──
        all_sku_set = set(sku_cross.keys()) | set(sku_stock.keys())
        all_skus = list(all_sku_set)
        prod_map: dict[int, dict] = {}
        # Query in batches of 500 to avoid SQL limit
        for i in range(0, len(all_skus), 500):
            batch = all_skus[i:i+500]
            sku_list = ", ".join(str(x) for x in batch)
            pg_rows = (await db.execute(
                text(f"""
                    SELECT sku, offer_id, name
                    FROM dim_ozon_products
                    WHERE shop_id = :sid AND sku IN ({sku_list})
                """),
                {"sid": shop_id},
            )).fetchall()
            for r in pg_rows:
                prod_map[r[0]] = {"offer_id": r[1] or "", "name": (r[2] or "")[:60]}

        # ── 5. Daily sales per SKU ──
        daily_sales_rows = ch.query("""
            SELECT sku,
                   count() AS total_orders,
                   sum(quantity) AS total_qty
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= {d_start:Date}
              AND status NOT IN ('cancelled', 'canceled')
            GROUP BY sku
        """, parameters=params).result_rows

        sku_daily_sales: dict[int, float] = {}
        for r in daily_sales_rows:
            sku = int(r[0])
            total_qty = int(r[2])
            sku_daily_sales[sku] = round(total_qty / period, 2) if period > 0 else 0

        ch.close()

        # ── 6. Build prompt ──
        overall_cross_pct = round(total_cross / total_orders * 100, 1) if total_orders > 0 else 0

        prompt = f"""Магазин: {shop.name} (Ozon)
Период: {period} дней (с {d_start} по {today})
Всего заказов: {total_orders}
Кросс-заказов: {total_cross} ({overall_cross_pct}%)
Складов: {len(wh_summary)}

## СКЛАДЫ И КРОСС-СТАТИСТИКА:
"""
        for wh_name, wh_data in sorted(wh_summary.items(), key=lambda x: x[1]["total"], reverse=True):
            wh_cross_pct = round(wh_data["cross"] / wh_data["total"] * 100) if wh_data["total"] > 0 else 0
            prompt += f"- {wh_name} (кластер: {wh_data['cluster']}): {wh_data['total']} заказов, {wh_data['cross']} кросс ({wh_cross_pct}%)\n"

        prompt += "\n## ПРОБЛЕМНЫЕ SKU (с кросс-заказами):\n"
        problem_skus = [
            (sku, data) for sku, data in sku_cross.items()
            if data["total"] >= 3 and data["cross"] > 0
        ]
        problem_skus.sort(key=lambda x: x[1]["cross"], reverse=True)

        for sku, data in problem_skus[:20]:
            prod = prod_map.get(sku, {})
            cross_pct = round(data["cross"] / data["total"] * 100) if data["total"] > 0 else 0
            stocks = sku_stock.get(sku, {})
            stock_str = ", ".join(f"{wh}: {s}" for wh, s in sorted(stocks.items(), key=lambda x: x[1], reverse=True)[:8])
            if not stock_str:
                stock_str = "нет стока"
            ds = sku_daily_sales.get(sku, 0)
            total_stock = sum(stocks.values())

            prompt += f"\n### {prod.get('offer_id', str(sku))} ({prod.get('name', f'SKU {sku}')[:45]})\n"
            prompt += f"  Заказов: {data['total']}, кросс: {data['cross']} ({cross_pct}%), daily_sales: {ds}\n"
            prompt += f"  Сток: {total_stock} шт → {stock_str}\n"

            # Top cross routes (simplified — no deficit calc)
            routes = sorted(data["routes"], key=lambda x: x["orders"], reverse=True)[:7]
            if routes:
                prompt += "  Кросс-маршруты:\n"
                for route in routes:
                    prompt += f"    {route['from_wh']} → {route['to_cluster']}: {route['orders']} кросс-заказов\n"

        skus_high_cross = sum(1 for _, d in sku_cross.items() if d["total"] >= 3 and d["cross"] / d["total"] > 0.3)
        whs_with_cross = sum(1 for _, d in wh_summary.items() if d["cross"] > 0)

        prompt += f"""
## ЗАДАНИЕ:
Период анализа: {period} дней.
1. Оцени общую ситуацию (severity, diagnosis) — конкретные цифры.
2. Выбери 5-10 проблемных SKU — где кросс > 25%. Для каждого: маршруты, стоки, текстовая рекомендация (БЕЗ конкретных штук).
3. Оцени каждый склад (warehouse_assessments) — сколько кросса, куда идёт, почему проблема.
4. Дай 3-5 приоритетных действий (priority_actions) — текстом, БЕЗ конкретных количеств. Если действие решается через довоз/поставку — ставь link_to_supply=true.
5. Выдай ТОЛЬКО JSON."""

        # ── 6. Call Gemini ──
        KIE_AI_URL = "https://api.kie.ai/gemini-2.5-flash/v1/chat/completions"

        async with httpx.AsyncClient(timeout=120.0) as client:
            resp = await client.post(
                KIE_AI_URL,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {api_key}",
                },
                json={
                    "messages": [
                        {"role": "system", "content": [{"type": "text", "text": _AI_PROMPT_OZON_CROSS}]},
                        {"role": "user", "content": [{"type": "text", "text": prompt}]},
                    ],
                    "stream": False,
                    "include_thoughts": False,
                },
            )

        if resp.status_code != 200:
            logger.error("Gemini API error %s: %s", resp.status_code, resp.text[:500])
            raise HTTPException(status_code=502, detail="AI API error")

        resp_json = resp.json()
        content = resp_json.get("choices", [{}])[0].get("message", {}).get("content", "")

        # Strip markdown code fences
        content = content.strip()
        if content.startswith("```"):
            content = content.split("\n", 1)[1] if "\n" in content else content[3:]
        if content.endswith("```"):
            content = content[:-3]
        content = content.strip()
        if content.startswith("json"):
            content = content[4:].strip()

        try:
            ai_result = json.loads(content)
        except json.JSONDecodeError:
            logger.warning("Failed to parse Ozon cross AI JSON: %s", content[:500])
            raise HTTPException(status_code=502, detail="AI returned invalid JSON")

        # Enrich with context
        ai_result["period_days"] = period
        ai_result["analyzed_at"] = int(time.time())
        ai_result["context"] = {
            "total_orders": total_orders,
            "total_cross": total_cross,
            "cross_pct": overall_cross_pct,
            "warehouses_count": len(wh_summary),
            "skus_analyzed": len(sku_cross),
        }

        # Cache
        _ai_cache[cache_key] = (time.time(), ai_result)

        return ai_result

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Ozon Cross AI analysis failed")
        raise HTTPException(status_code=500, detail=f"Ошибка анализа: {str(e)}")


# ═══════════════════════════════════════════════════════════════
# WB Cross-Logistics AI Analysis
# ═══════════════════════════════════════════════════════════════

_AI_PROMPT_WB_CROSS = """Ты — эксперт по КРОСС-ЛОГИСТИКЕ на Wildberries. Твоя задача — дать ОБЗОР кросс-проблем и ОЦЕНКУ складов.

## ВВОДНЫЕ
- **Кросс-доставка** — заказ из федерального округа X обслуживается складом из округа Y (не ближайшим). Увеличивает стоимость и время.
- **FBS (Склад продавца)** — продавец сам отправляет товар покупателю. ВСЕ FBS-заказы — это фактически кросс-доставка, потому что товар идёт с одного склада продавца. FBS стоит ДОРОЖЕ чем FBO. В рекомендациях указывай: «переместить объём с FBS на FBO-склады WB в регионе X для снижения кросс-логистики».
- **FBO (Склад WB)** — товар хранится на складе WB и доставляется покупателю из ближайшего склада. Это ДЕШЕВЛЕ и БЫСТРЕЕ.
- **nm_id** — артикул ВБ.
- **Питание** — товары из категорий «Корма», «Товары для животных», «Продукты питания» ДОЛЖНЫ поставляться только на склады с пометкой ": Питание" (например «Электросталь: Питание», «Казань: Питание»). Обычный склад «Электросталь» НЕ подходит для кормов. Это ВАЖНО для рекомендаций по довозу!
- **Округа** — Центральный, Северо-Западный, Приволжский, Южный, Уральский, Сибирский, Дальневосточный.

## ТЫ ОБЯЗАН
1. Оценить общую ситуацию с кросс-доставкой (severity, diagnosis)
2. Выделить проблемные SKU (где кросс выше 25%)
3. Оценить каждый склад — сколько кросса генерирует и почему
4. Дать текстовые рекомендации — что сделать, БЕЗ конкретных штук
5. ЕСЛИ товар из категории «Корма»/«Товары для животных» — указывать что нужна поставка на склад ": Питание" в рекомендациях

## ФОРМАТ ОТВЕТА — строго JSON:
{
  "severity": "critical" | "warning" | "ok",
  "diagnosis": "1-3 предложения. Конкретные цифры кросса, кол-во проблемных SKU.",
  "key_metrics": {
    "cross_pct": 45,
    "cross_orders": 120,
    "total_orders": 267,
    "warehouses_with_cross": 5,
    "skus_with_high_cross": 8
  },
  "problem_skus": [
    {
      "offer_id": "АРТ-789",
      "name": "Название товара",
      "total_orders": 50,
      "cross_orders": 30,
      "cross_pct": 60,
      "is_food": true,
      "stock_distribution": [
        {"warehouse": "Электросталь: Питание", "stock": 150}
      ],
      "top_cross_routes": [
        {"from_warehouse": "Электросталь: Питание", "to_okrug": "Приволжский", "orders": 15}
      ],
      "recommendation": "Корм сконцентрирован на Электросталь: Питание, а 60% заказов из Приволжского округа. Нужен довоз на склады Питание в Казани или Новосемейкино."
    }
  ],
  "warehouse_assessments": [
    {
      "warehouse": "Электросталь: Питание",
      "okrug": "Центральный федеральный округ",
      "status": "critical" | "warning" | "ok",
      "total_orders": 200,
      "cross_orders": 80,
      "cross_pct": 40,
      "main_cross_destinations": ["Приволжский", "Уральский"],
      "assessment": "Склад генерирует 40% кросс — основной спрос из Приволжского и Уральского округов."
    }
  ],
  "priority_actions": [
    {
      "action": "Сформировать поставку на склады Питание в Поволжье (Казань/Новосемейкино) для закрытия кросс-спроса",
      "impact": "Снизит кросс на ~80 заказов/месяц",
      "link_to_supply": true
    }
  ],
  "general_tips": [
    "Конкретный совет с цифрами."
  ]
}

## АБСОЛЮТНЫЕ ЗАПРЕТЫ
- ❌ НЕ СЧИТАЙ конкретные штуки для перемещений и поставок
- ❌ НЕ выдумывай данные
- ❌ НИКОГДА: «например», «допустим», «альтернативно»
- ✅ severity: "critical" если cross_pct > 40%, "warning" если 20-40%, "ok" если < 20%
- ✅ warehouse_assessments.status: "critical" если cross_pct склада > 50%, "warning" если 25-50%, "ok" если < 25%
- ✅ Пиши НА РУССКОМ
- ✅ В priority_actions ставь link_to_supply=true для действий, которые решаются через довоз
- ✅ Для кормов/товаров для животных ВСЕГДА указывай что нужна поставка на склад ": Питание"
"""


@router.post("/wb/cross/ai-analysis")
async def get_wb_cross_ai_analysis(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    force: bool = Query(False, description="Skip cache"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """AI-powered WB cross-logistics analysis using Gemini 2.5 Flash."""
    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "wildberries":
        raise HTTPException(status_code=404, detail="Shop not found")

    cache_key = f"wb_cross_ai_{shop_id}_{period}"
    if not force and cache_key in _ai_cache:
        ts, cached = _ai_cache[cache_key]
        if time.time() - ts < _AI_CACHE_TTL:
            return {**cached, "cached": True}

    api_key = os.getenv("KIE_AI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="AI API key not configured")

    try:
        from app.core.clickhouse import get_clickhouse_client

        ch = get_clickhouse_client()
        today = date.today()
        d_start = today - timedelta(days=period)
        params = {"shop_id": shop_id, "d_start": d_start}

        # ── 1. Cross-delivery matrix: nm_id × warehouse × okrug ──
        cross_matrix_rows = ch.query("""
            SELECT
                nm_id, warehouse_name, oblast_okrug_name,
                count() AS orders
            FROM mms_analytics.fact_orders_raw
            WHERE shop_id = {shop_id:UInt32}
              AND date >= {d_start:Date}
              AND is_cancel = 0
              AND warehouse_name != '' AND oblast_okrug_name != ''
            GROUP BY nm_id, warehouse_name, oblast_okrug_name
            ORDER BY orders DESC
        """, parameters=params).result_rows

        # Build per-SKU cross data
        nm_cross: dict[int, dict] = {}
        for r in cross_matrix_rows:
            nm_id = int(r[0])
            wh = str(r[1])
            okrug = str(r[2])
            orders = int(r[3])
            wh_okrug = WAREHOUSE_TO_OKRUG.get(wh, "")
            is_cross = wh_okrug != okrug and wh_okrug != ""

            if nm_id not in nm_cross:
                nm_cross[nm_id] = {"total": 0, "cross": 0, "routes": []}
            nm_cross[nm_id]["total"] += orders
            if is_cross:
                nm_cross[nm_id]["cross"] += orders
                nm_cross[nm_id]["routes"].append({
                    "from_wh": wh, "to_okrug": okrug, "orders": orders
                })

        # ── 2. Per-warehouse cross summary ──
        wh_cross_rows = ch.query("""
            SELECT
                warehouse_name,
                oblast_okrug_name,
                count() AS orders,
                any(warehouse_type) AS wh_type
            FROM mms_analytics.fact_orders_raw
            WHERE shop_id = {shop_id:UInt32}
              AND date >= {d_start:Date}
              AND is_cancel = 0
              AND warehouse_name != '' AND oblast_okrug_name != ''
            GROUP BY warehouse_name, oblast_okrug_name
            ORDER BY orders DESC
        """, parameters=params).result_rows

        wh_summary: dict[str, dict] = {}
        total_orders = 0
        total_cross = 0
        for r in wh_cross_rows:
            wh = str(r[0])
            okrug = str(r[1])
            orders = int(r[2])
            wh_type = str(r[3]) if len(r) > 3 else ""
            wh_okrug = WAREHOUSE_TO_OKRUG.get(wh, "")
            is_cross = wh_okrug != okrug and wh_okrug != ""

            if wh not in wh_summary:
                wh_summary[wh] = {"okrug": wh_okrug, "total": 0, "cross": 0, "is_fbs": False}
            wh_summary[wh]["total"] += orders
            if wh_type == "Склад продавца":
                wh_summary[wh]["is_fbs"] = True
            total_orders += orders
            if is_cross:
                wh_summary[wh]["cross"] += orders
                total_cross += orders

        # ── 3. Stocks per nm_id per warehouse ──
        stock_rows = ch.query("""
            SELECT warehouse_name, nm_id,
                   argMax(quantity, fetched_at) AS stock
            FROM mms_analytics.fact_inventory_snapshot
            WHERE shop_id = {shop_id:UInt32}
              AND warehouse_name NOT LIKE 'FBS:%'
            GROUP BY warehouse_name, nm_id
            HAVING stock > 0
            ORDER BY stock DESC
        """, parameters=params).result_rows

        nm_stock: dict[int, dict[str, int]] = {}
        for r in stock_rows:
            wh = str(r[0])
            nm_id = int(r[1])
            stock = int(r[2])
            if nm_id not in nm_stock:
                nm_stock[nm_id] = {}
            nm_stock[nm_id][wh] = stock

        # ── 4. Product metadata from PostgreSQL ──
        all_nm_set = set(nm_cross.keys()) | set(nm_stock.keys())
        all_nms = list(all_nm_set)
        prod_map: dict[int, dict] = {}
        for i in range(0, len(all_nms), 500):
            batch = all_nms[i:i+500]
            nm_list = ", ".join(str(x) for x in batch)
            pg_rows = (await db.execute(
                text(f"""
                    SELECT nm_id, vendor_code, name, category
                    FROM dim_products
                    WHERE shop_id = :sid AND nm_id IN ({nm_list})
                """),
                {"sid": shop_id},
            )).fetchall()
            for r in pg_rows:
                prod_map[r[0]] = {
                    "offer_id": r[1] or "",
                    "name": (r[2] or "")[:60],
                    "subject": r[3] or "",
                }

        # ── 5. Daily sales per nm_id ──
        daily_sales_rows = ch.query("""
            SELECT nm_id,
                   count() AS total_orders
            FROM mms_analytics.fact_orders_raw
            WHERE shop_id = {shop_id:UInt32}
              AND date >= {d_start:Date}
              AND is_cancel = 0
            GROUP BY nm_id
        """, parameters=params).result_rows

        nm_daily_sales: dict[int, float] = {}
        for r in daily_sales_rows:
            nm_id = int(r[0])
            nm_daily_sales[nm_id] = round(int(r[1]) / period, 2) if period > 0 else 0

        ch.close()

        # ── 6. Build prompt ──
        overall_cross_pct = round(total_cross / total_orders * 100, 1) if total_orders > 0 else 0

        prompt = f"""Магазин: {shop.name} (Wildberries)
Период: {period} дней (с {d_start} по {today})
Всего заказов: {total_orders}
Кросс-заказов: {total_cross} ({overall_cross_pct}%)
Складов: {len(wh_summary)}

## СКЛАДЫ И КРОСС-СТАТИСТИКА:
"""
        for wh_name, wh_data in sorted(wh_summary.items(), key=lambda x: x[1]["total"], reverse=True):
            wh_cross_pct = round(wh_data["cross"] / wh_data["total"] * 100) if wh_data["total"] > 0 else 0
            is_food_wh = ": Питание" in wh_name
            is_fbs = wh_data.get("is_fbs", False)
            tags = []
            if is_fbs:
                tags.append("FBS")
            if is_food_wh:
                tags.append("ПИТАНИЕ")
            tag_str = f" [{', '.join(tags)}]" if tags else ""
            prompt += f"- {wh_name}{tag_str} (округ: {wh_data['okrug']}): {wh_data['total']} заказов, {wh_data['cross']} кросс ({wh_cross_pct}%)\n"

        prompt += "\n## ПРОБЛЕМНЫЕ SKU (с кросс-заказами):\n"
        problem_skus = [
            (nm, data) for nm, data in nm_cross.items()
            if data["total"] >= 3 and data["cross"] > 0
        ]
        problem_skus.sort(key=lambda x: x[1]["cross"], reverse=True)

        for nm_id, data in problem_skus[:20]:
            prod = prod_map.get(nm_id, {})
            cross_pct = round(data["cross"] / data["total"] * 100) if data["total"] > 0 else 0
            stocks = nm_stock.get(nm_id, {})
            stock_str = ", ".join(f"{wh}: {s}" for wh, s in sorted(stocks.items(), key=lambda x: x[1], reverse=True)[:8])
            if not stock_str:
                stock_str = "нет стока"
            ds = nm_daily_sales.get(nm_id, 0)
            total_stock = sum(stocks.values())

            # Determine if food product
            subject = prod.get("subject", "")
            is_food = any(cat.lower() in subject.lower() for cat in _FOOD_CATEGORIES) if subject else False
            food_label = " [КОРМ/ПИТАНИЕ]" if is_food else ""

            prompt += f"\n### {prod.get('offer_id', str(nm_id))} ({prod.get('name', f'NM {nm_id}')[:45]}){food_label}\n"
            prompt += f"  Заказов: {data['total']}, кросс: {data['cross']} ({cross_pct}%), daily_sales: {ds}\n"
            prompt += f"  Сток: {total_stock} шт → {stock_str}\n"

            # Top cross routes
            routes = sorted(data["routes"], key=lambda x: x["orders"], reverse=True)[:7]
            if routes:
                prompt += "  Кросс-маршруты:\n"
                for route in routes:
                    prompt += f"    {route['from_wh']} → {route['to_okrug']}: {route['orders']} кросс-заказов\n"

        skus_high_cross = sum(1 for _, d in nm_cross.items() if d["total"] >= 3 and d["cross"] / d["total"] > 0.3)
        whs_with_cross = sum(1 for _, d in wh_summary.items() if d["cross"] > 0)

        prompt += f"""
## ЗАДАНИЕ:
Период анализа: {period} дней.
1. Оцени общую ситуацию (severity, diagnosis) — конкретные цифры.
2. Выбери 5-10 проблемных SKU — где кросс > 25%. Для каждого: маршруты, стоки, текстовая рекомендация (БЕЗ конкретных штук). Если товар из категории корм/питание — ОБЯЗАТЕЛЬНО укажи что нужен склад ": Питание".
3. Оцени каждый склад (warehouse_assessments) — сколько кросса, куда идёт, почему проблема.
4. Дай 3-5 приоритетных действий (priority_actions). Если действие решается через довоз — link_to_supply=true.
5. Выдай ТОЛЬКО JSON."""

        # ── 7. Call Gemini ──
        KIE_AI_URL = "https://api.kie.ai/gemini-2.5-flash/v1/chat/completions"

        async with httpx.AsyncClient(timeout=120.0) as client:
            resp = await client.post(
                KIE_AI_URL,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {api_key}",
                },
                json={
                    "messages": [
                        {"role": "system", "content": [{"type": "text", "text": _AI_PROMPT_WB_CROSS}]},
                        {"role": "user", "content": [{"type": "text", "text": prompt}]},
                    ],
                    "stream": False,
                    "include_thoughts": False,
                },
            )

        if resp.status_code != 200:
            logger.error("Gemini API error %s: %s", resp.status_code, resp.text[:500])
            raise HTTPException(status_code=502, detail="AI API error")

        resp_json = resp.json()
        content = resp_json.get("choices", [{}])[0].get("message", {}).get("content", "")

        # Strip markdown code fences
        content = content.strip()
        if content.startswith("```"):
            content = content.split("\n", 1)[1] if "\n" in content else content[3:]
        if content.endswith("```"):
            content = content[:-3]
        content = content.strip()
        if content.startswith("json"):
            content = content[4:].strip()

        try:
            ai_result = json.loads(content)
        except json.JSONDecodeError:
            logger.warning("Failed to parse WB cross AI JSON: %s", content[:500])
            raise HTTPException(status_code=502, detail="AI returned invalid JSON")

        # Enrich with context
        ai_result["period_days"] = period
        ai_result["analyzed_at"] = int(time.time())
        ai_result["context"] = {
            "total_orders": total_orders,
            "total_cross": total_cross,
            "cross_pct": overall_cross_pct,
            "warehouses_count": len(wh_summary),
            "skus_analyzed": len(nm_cross),
        }

        # Cache
        _ai_cache[cache_key] = (time.time(), ai_result)

        return ai_result

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("WB Cross AI analysis failed")
        raise HTTPException(status_code=500, detail=f"Ошибка анализа: {str(e)}")

# ═══════════════════════════════════════════════════════════════
# AI Storage Analysis
# ═══════════════════════════════════════════════════════════════

_AI_PROMPT_STORAGE = """Ты — эксперт по ОПТИМИЗАЦИИ ПЛАТНОГО ХРАНЕНИЯ на Wildberries.

ТВОЯ ЗАДАЧА: проанализировать расходы на хранение каждого товара и дать КОНКРЕТНЫЕ рассчитанные рекомендации по оптимизации.

## КОНТЕКСТ: ЭКОНОМИКА ХРАНЕНИЯ WB

### Тарифы хранения (2025):
- Базовая ставка: ~0.08₽ за 1 литр в день (тип «Короб»)
- Коэффициент зависит от склада (загруженность): от 0.5 до 5.0
- Первые 60 дней с момента поставки — БЕСПЛАТНО (с сентября 2025)
- Формула: стоимость = base_rate × volume_liters × qty × coef × days

### Стоимость ВЫВОЗА товара со склада:
- Возврат продавцу: ~50₽ за единицу товара
- Доставка в ПВЗ: базовая стоимость логистики (38₽ за первый литр + 9.5₽ за доп. литр)
- Доставка в пределах МКАД: +1500₽ за объём до 10 м³
- Срок обработки заявки: до 45 дней
- ВАЖНО: вывоз имеет смысл ТОЛЬКО если расходы на хранение за 3-6 месяцев ПРЕВЫШАЮТ стоимость вывоза

### Обратная логистика (возврат от покупателя):
- 33₽ за единицу для всех категорий

## СТРОГИЕ ГРАНИЦЫ АНАЛИЗА

АНАЛИЗИРУЙ ТОЛЬКО:
- Расходы на хранение по каждому SKU: факт vs прогноз
- Оборачиваемость и тренд продаж
- Маржинальность: покрывает ли прибыль расходы на хранение
- Влияние рекламы на продажи (запущена реклама → товар может подождать)
- Экономика вывоза: когда вывозить дешевле, чем хранить
- Оптимальный запас: сколько реально нужно держать на складе

НЕ АНАЛИЗИРУЙ:
- ❌ Географию продаж → отдельный раздел
- ❌ Кросс-логистику → отдельный раздел
- ❌ Рекламные кампании детально → раздел «Реклама»

## ЛОГИКА ПРИНЯТИЯ РЕШЕНИЙ ПО КАЖДОМУ SKU

### Дерево решений:
1. **Товар убыточен** (хранение > прибыли за период):
   - Если DRR разумный (< 30%) и реклама запущена → ПОДОЖДАТЬ, реклама может раскачать
   - Если нет продаж и нет рекламы → ЛИКВИДАЦИЯ или ГЛУБОКАЯ СКИДКА
   - Если есть продажи, но медленные → СКИДКА 15-25%

2. **Товар прибыльный, но затоварен** (оборачиваемость > 90д):
   - Рассчитай оптимальный запас = daily_sales × 60 дней
   - Избыток = stock - оптимальный запас
   - Если избыток × стоимость хранения/шт × 6 мес > избыток × 50₽ вывоза → ВЫВЕЗТИ излишки
   - Иначе → СОКРАТИТЬ ПОСТАВКИ и подождать

3. **Товар прибыльный и быстро продаётся** (оборачиваемость < 60д):
   - Всё ОК, не трогать
   - Если рекламы нет и маржа позволяет → рассмотреть запуск рекламы для ускорения

### РАСЧЁТ СТОИМОСТИ ВЫВОЗА:
Для каждого SKU при рекомендации «вывезти»:
- withdrawal_cost = qty × 50₽ + логистика (38₽ × объём в литрах × qty для первого литра, 9.5₽ для доп.)
- Упрощённо: withdrawal_cost ≈ qty × (50 + объём_литры × 15)₽
- Сравни с: ежемесячное хранение × 6 месяцев
- Вывоз выгоден если: withdrawal_cost < storage_monthly × 3

## ФОРМАТ ОТВЕТА — СТРОГО JSON:
{
  "severity": "critical" | "warning" | "ok",
  "diagnosis": "Главный вывод 1-2 предложения с цифрами: сколько тратится на хранение, сколько можно сэкономить",
  "key_metrics": {
    "total_storage_monthly": 25000,
    "potential_savings": 8000,
    "storage_roi_pct": -15,
    "overstock_skus": 12,
    "loss_making_skus": 5,
    "avg_turnover_days": 180
  },
  "sku_actions": [
    {
      "vendor_code": "АРТИКУЛ",
      "name": "Название товара",
      "diagnosis": "Конкретный диагноз с числами: оборачиваемость, хранение, прибыль, почему проблема",
      "current_storage_cost": 1500,
      "current_turnover_days": 210,
      "stock": 28,
      "options": [
        {
          "action": "discount",
          "label": "Снизить цену на 20%",
          "detail": "Расчёт: при росте продаж в 1.5× оборачиваемость снизится до ~140д. Экономия на хранении: ~600₽/мес. Потеря в марже: ~200₽/мес. Чистый выигрыш: ~400₽/мес.",
          "expected_savings": 400,
          "withdrawal_cost": 0,
          "risk": "medium"
        },
        {
          "action": "withdraw",
          "label": "Вывезти 15 шт со склада",
          "detail": "Стоимость вывоза: 15 × 65₽ = 975₽. Экономия на хранении: ~800₽/мес. Окупится через 1.2 мес. Оставить 13 шт — это на ~130 дней при текущих продажах.",
          "expected_savings": 800,
          "withdrawal_cost": 975,
          "risk": "medium"
        },
        {
          "action": "do_nothing",
          "label": "Оставить как есть",
          "detail": "Через 6 месяцев дополнительные расходы на хранение: 9000₽. Но товар продаётся и прибыль покрывает хранение.",
          "expected_savings": 0,
          "withdrawal_cost": 0,
          "risk": "low"
        }
      ],
      "recommended_option": 0
    }
  ],
  "general_tips": [
    "Конкретный совет с расчётом, 1-3 предложения."
  ]
}

## ПРАВИЛА
- severity: "critical" если > 30% SKU убыточны из-за хранения, "warning" если 10-30%, "ok" если < 10%
- sku_actions: ТЫ ОБЯЗАН ВКЛЮЧИТЬ РОВНО КАЖДЫЙ ТОВАР ИЗ СПИСКА НИЖЕ. Не пропускай ни одного! Для товаров с пометкой ⚠️ — они ОБЯЗАТЕЛЬНЫ.
- options: 2-4 варианта для КАЖДОГО SKU. ВСЕГДА включай "do_nothing" с расчётом потерь
- action: ТОЛЬКО "discount", "withdraw", "launch_ads", "reduce_supply", "liquidate", "do_nothing"
- При "withdraw": ОБЯЗАТЕЛЬНО указать withdrawal_cost и расчёт окупаемости
- При "liquidate": указать максимальную скидку для быстрой распродажи
- expected_savings: ЕЖЕМЕСЯЧНАЯ экономия в рублях
- withdrawal_cost: разовая стоимость вывоза в рублях (только для "withdraw")
- risk: "low" (безопасно), "medium" (есть нюансы), "high" (рискованно)
- general_tips: 3-5 советов по ОПТИМИЗАЦИИ ХРАНЕНИЯ. Конкретные, с расчётами
- Все числа — из реальных данных, НЕ выдумывай
- Пиши НА РУССКОМ
- НАЗЫВАЙ товары по vendor_code + name
- КРИТИЧЕСКИ ВАЖНО: отвечай ТОЛЬКО чистым JSON. Без вводного текста, без markdown, без комментариев до или после JSON. Первый символ ответа ДОЛЖЕН быть {
"""


@router.post("/wb/storage/ai-analysis")
async def get_wb_storage_ai_analysis(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    force: bool = Query(False, description="Skip cache"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """AI-powered storage cost optimization analysis using Gemini 2.5 Flash."""

    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "wildberries":
        raise HTTPException(status_code=404, detail="Shop not found")

    cache_key = f"wb_storage_ai_{shop_id}_{period}"

    if not force and cache_key in _ai_cache:
        ts, cached = _ai_cache[cache_key]
        if time.time() - ts < _AI_CACHE_TTL:
            return {**cached, "cached": True}

    api_key = os.getenv("KIE_AI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="AI API key not configured")

    try:
        from app.core.clickhouse import get_clickhouse_client

        ch = get_clickhouse_client()
        today = date.today()
        d_start = today - timedelta(days=period)

        # ── 1. Per-SKU storage from fact_wb_paid_storage ──
        actual_storage_rows = ch.query("""
            SELECT
                vendor_code,
                nm_id,
                round(SUM(warehouse_price), 2) AS storage_total,
                round(SUM(warehouse_price) * (30 / {period:UInt32}), 2) AS storage_30d
            FROM mms_analytics.fact_wb_paid_storage FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND dt >= {d_start:Date} AND dt <= {d_end:Date}
            GROUP BY vendor_code, nm_id
            HAVING storage_total != 0
            ORDER BY storage_total DESC
        """, parameters={
            "shop_id": shop_id, "period": max(period, 1),
            "d_start": d_start.isoformat(), "d_end": today.isoformat(),
        }).result_rows

        storage_by_vc: dict[str, dict] = {}
        for r in actual_storage_rows:
            vc = str(r[0]).strip()
            if vc:
                storage_by_vc[vc] = {
                    "nm_id": int(r[1]),
                    "storage_period": round(float(r[2])),
                    "storage_30d": round(float(r[3])),
                }

        # ── 2. Per-SKU orders, revenue, ad data ──
        sku_rows = ch.query("""
            SELECT
                s.nm_id,
                s.total_stock,
                coalesce(o.orders, 0) AS orders,
                coalesce(o.revenue, 0) AS revenue,
                coalesce(a.spend, 0) AS ad_spend,
                coalesce(a.ad_orders, 0) AS ad_orders
            FROM (
                SELECT nm_id, sum(qty) AS total_stock
                FROM (
                    SELECT nm_id, argMax(quantity, fetched_at) AS qty
                    FROM mms_analytics.fact_inventory_snapshot
                    WHERE shop_id = {shop_id:UInt32}
                      AND warehouse_name NOT LIKE 'FBS:%'
                    GROUP BY warehouse_name, nm_id
                    HAVING qty > 0
                )
                GROUP BY nm_id
            ) AS s
            LEFT JOIN (
                SELECT nm_id, count() AS orders, sum(toFloat64(price_with_disc)) AS revenue
                FROM mms_analytics.fact_orders_raw
                WHERE shop_id = {shop_id:UInt32} AND date >= {d_start:Date} AND is_cancel = 0
                GROUP BY nm_id
            ) AS o ON o.nm_id = s.nm_id
            LEFT JOIN (
                SELECT nm_id, sum(spend) AS spend, sum(orders) AS ad_orders
                FROM mms_analytics.fact_advert_stats_v3 FINAL
                WHERE shop_id = {shop_id:UInt32} AND date >= {d_start:Date}
                GROUP BY nm_id
            ) AS a ON a.nm_id = s.nm_id
        """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows

        sku_data: dict[int, dict] = {}
        for r in sku_rows:
            nm = int(r[0])
            sku_data[nm] = {
                "stock": int(r[1]),
                "orders": int(r[2]),
                "revenue": float(r[3]),
                "ad_spend": float(r[4]),
                "ad_orders": int(r[5]),
            }

        # ── 3. Product info from PostgreSQL ──
        nm_ids = list(sku_data.keys())
        products_map: dict[int, dict] = {}
        if nm_ids:
            nm_list = ", ".join(str(x) for x in nm_ids)
            pg_rows = (await db.execute(
                text(f"""
                    SELECT nm_id, vendor_code, name, current_price, current_discount,
                           COALESCE(length, 0), COALESCE(width, 0), COALESCE(height, 0)
                    FROM dim_products
                    WHERE shop_id = :sid AND nm_id IN ({nm_list})
                """),
                {"sid": shop_id},
            )).fetchall()
            for r in pg_rows:
                l, w, h = float(r[5] or 0), float(r[6] or 0), float(r[7] or 0)
                vol = (l * w * h) / 1000.0 if (l > 0 and w > 0 and h > 0) else 1.0
                if vol > 10000:
                    vol = (l * w * h) / 1_000_000.0
                vol = max(vol, 0.1)
                products_map[r[0]] = {
                    "vendor_code": r[1] or "",
                    "name": (r[2] or "")[:60],
                    "price": float(r[3]) if r[3] else 0,
                    "discount": int(r[4]) if r[4] else 0,
                    "vol_liters": round(vol, 2),
                }

        # ── 4. Cost prices ──
        cost_prices: dict[str, float] = {}
        try:
            cost_rows_pg = (await db.execute(
                text("SELECT offer_id, COALESCE(cost_price, 0) + COALESCE(packaging_cost, 0) FROM product_costs WHERE shop_id = :sid AND (cost_price > 0 OR packaging_cost > 0)"),
                {"sid": shop_id},
            )).fetchall()
            for r in cost_rows_pg:
                cost_prices[r[0]] = float(r[1])
        except Exception:
            pass

        # ── 5. Per-SKU P&L from fact_finances ──
        sku_pnl: dict[str, dict] = {}
        try:
            pnl_rows = ch.query("""
                SELECT
                    vendor_code,
                    sumIf(JSONExtractFloat(raw_payload, 'retail_price_withdisc_rub'),
                        operation_type = 'Продажа') AS revenue,
                    sumIf(payout_amount, operation_type = 'Продажа')
                        - sumIf(payout_amount, operation_type = 'Возврат') AS payout,
                    sum(abs(wb_delivery_rub)) AS logistics,
                    sum(abs(storage_fee)) AS storage_fact,
                    sumIf(abs(JSONExtractFloat(raw_payload, 'deduction')), 1) AS deductions,
                    sumIf(quantity, operation_type = 'Продажа' AND quantity > 0) AS sales_qty,
                    sumIf(quantity, operation_type = 'Возврат') AS ret_qty
                FROM mms_analytics.fact_finances FINAL
                WHERE shop_id = {shop_id:UInt32} AND marketplace = 1
                  AND event_date >= {d_start:Date} AND event_date <= {d_end:Date}
                  AND vendor_code != ''
                GROUP BY vendor_code
            """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": today}).result_rows
            for r in pnl_rows:
                vc = str(r[0] or "").strip()
                if not vc:
                    continue
                revenue = float(r[1] or 0)
                payout = float(r[2] or 0)
                logistics = float(r[3] or 0)
                storage_fact = float(r[4] or 0)
                deductions = float(r[5] or 0)
                sales_qty = int(r[6] or 0)
                ret_qty = int(r[7] or 0)
                cogs_unit = cost_prices.get(vc, 0)
                cogs_total = cogs_unit * max(sales_qty - abs(ret_qty), 0)
                net_profit = payout - logistics - storage_fact - deductions - cogs_total
                sku_pnl[vc] = {
                    "revenue": round(revenue),
                    "payout": round(payout),
                    "logistics": round(logistics),
                    "storage_fact": round(storage_fact),
                    "deductions": round(deductions),
                    "cogs": round(cogs_total),
                    "net_profit": round(net_profit),
                    "margin_pct": round(net_profit / revenue * 100, 1) if revenue > 0 else 0,
                    "sales_qty": sales_qty,
                }
        except Exception as e:
            logger.warning("Storage AI: P&L query failed: %s", e)

        # ── 6. Per-warehouse stock distribution ──
        ch2 = get_clickhouse_client()
        wh_stock_map: dict[int, list] = {}
        if nm_ids:
            nm_list_ch = ",".join(str(x) for x in nm_ids[:60])
            wh_stock_rows = ch2.query(f"""
                SELECT nm_id, warehouse_name, argMax(quantity, fetched_at) AS qty
                FROM mms_analytics.fact_inventory_snapshot
                WHERE shop_id = {{shop_id:UInt32}} AND nm_id IN ({nm_list_ch})
                  AND warehouse_name NOT LIKE 'FBS:%'
                GROUP BY nm_id, warehouse_name
                HAVING qty > 0
                ORDER BY nm_id, qty DESC
            """, parameters={"shop_id": shop_id}).result_rows
            for wr in wh_stock_rows:
                wh_stock_map.setdefault(int(wr[0]), []).append({
                    "warehouse": wr[1],
                    "qty": int(wr[2]),
                })
        ch2.close()
        ch.close()

        # ── 7. Build SKU context for AI ──
        skus_context = []
        total_storage_30d = 0
        total_net_profit = 0

        for nm_id, sd in sku_data.items():
            prod = products_map.get(nm_id, {})
            vc = prod.get("vendor_code", "")
            if not vc:
                continue

            stock = sd["stock"]
            orders = sd["orders"]
            revenue = sd["revenue"]
            ad_spend = sd["ad_spend"]
            ad_orders = sd["ad_orders"]
            daily = orders / period if period > 0 else 0
            turnover = round(stock / daily) if daily > 0 else 9999
            drr = round(ad_spend / revenue * 100, 1) if revenue > 0 else 0
            vol = prod.get("vol_liters", 1.0)
            price = prod.get("price", 0)
            cost_price = cost_prices.get(vc, 0)

            # Storage cost
            st = storage_by_vc.get(vc, {})
            storage_30d = st.get("storage_30d", 0)
            storage_period = st.get("storage_period", 0)
            # Fallback: estimate from P&L
            if not storage_30d:
                pnl_storage = sku_pnl.get(vc, {}).get("storage_fact", 0)
                if pnl_storage > 0:
                    storage_30d = round(pnl_storage / period * 30)

            total_storage_30d += storage_30d

            # Forecast 30d: accounts for stock decreasing with sales
            forecast_30d = storage_30d  # fallback = flat
            if storage_30d > 0 and stock > 0:
                daily_cost = storage_30d / 30.0
                cost_per_unit = daily_cost / stock
                forecast = 0.0
                for day in range(30):
                    remaining = max(0, stock - daily * day)
                    if remaining <= 0:
                        break
                    forecast += cost_per_unit * remaining
                forecast_30d = round(forecast)

            # P&L
            pnl = sku_pnl.get(vc, {})
            net_profit = pnl.get("net_profit", 0)
            total_net_profit += net_profit

            # Withdrawal cost estimate
            withdrawal_cost_per_unit = round(50 + vol * 15)  # ~50₽ + logistics

            wh_list = wh_stock_map.get(nm_id, [])

            skus_context.append({
                "vendor_code": vc,
                "name": prod.get("name", ""),
                "nm_id": nm_id,
                "stock": stock,
                "orders": orders,
                "daily": round(daily, 2),
                "turnover_days": turnover,
                "revenue": round(revenue),
                "ad_spend": round(ad_spend),
                "ad_orders": ad_orders,
                "drr": drr,
                "in_ads": ad_spend > 0,
                "price": price,
                "cost_price": cost_price,
                "vol_liters": vol,
                "storage_30d": storage_30d,
                "storage_period": storage_period,
                "forecast_30d": forecast_30d,
                "net_profit": net_profit,
                "margin_pct": pnl.get("margin_pct", 0),
                "pnl_payout": pnl.get("payout", 0),
                "pnl_logistics": pnl.get("logistics", 0),
                "pnl_storage_fact": pnl.get("storage_fact", 0),
                "pnl_deductions": pnl.get("deductions", 0),
                "pnl_cogs": pnl.get("cogs", 0),
                "withdrawal_cost_per_unit": withdrawal_cost_per_unit,
                "warehouses": wh_list[:5],
            })

        # Sort by storage cost DESC
        skus_context.sort(key=lambda x: x["forecast_30d"], reverse=True)

        # Deterministic pre-filter: only SKUs that actually need optimization
        # Criteria: storage > 30₽/month AND (overstock OR unprofitable OR no sales)
        skus_needing_action = []
        for s in skus_context:
            forecast = s["forecast_30d"]
            if forecast < 30:  # negligible storage cost
                continue
            is_overstock = s["turnover_days"] > 90
            is_unprofitable = s["net_profit"] < forecast  # storage eats profit
            no_sales = s["daily"] == 0 and s["stock"] > 0
            if is_overstock or is_unprofitable or no_sales:
                skus_needing_action.append(s)

        # If too few, add top storage SKUs regardless
        existing_vcs = {s["vendor_code"] for s in skus_needing_action}
        for s in skus_context:
            if len(skus_needing_action) >= 5:
                break
            if s["vendor_code"] not in existing_vcs and s["forecast_30d"] > 0:
                skus_needing_action.append(s)
                existing_vcs.add(s["vendor_code"])

        # Cap at 15 for reasonable AI output size
        skus_needing_action = skus_needing_action[:15]

        # Sort again by storage cost DESC
        skus_needing_action.sort(key=lambda x: x["forecast_30d"], reverse=True)

        # Mark all as must-include (deterministic)
        for i, s in enumerate(skus_needing_action):
            s["must_include"] = True
            s["storage_rank"] = i + 1

        skus_for_ai = skus_needing_action

        # ── 8. Build prompt ──
        total_stock = sum(s["stock"] for s in skus_context)
        total_forecast_30d = sum(s["forecast_30d"] for s in skus_context)
        overstock_count = sum(1 for s in skus_context if s["turnover_days"] > 90 and s["forecast_30d"] > 50)
        loss_making = sum(1 for s in skus_context if s["forecast_30d"] > 0 and s["net_profit"] < s["forecast_30d"])
        avg_turnover = round(sum(s["turnover_days"] for s in skus_context) / len(skus_context)) if skus_context else 0

        prompt = f"""Магазин: {shop.name} (Wildberries)
Период анализа: {period} дней (с {d_start} по {today})

## ОБЩИЕ МЕТРИКИ ХРАНЕНИЯ:
- Общее хранение (факт) за 30д: ~{round(total_storage_30d)}₽
- ПРОГНОЗ хранения на 30д (с учётом продаж): ~{round(total_forecast_30d)}₽
- Общая чистая прибыль за {period}д: {round(total_net_profit)}₽
- Всего SKU на складе: {len(skus_context)}
- Общий остаток: {total_stock} шт
- Затоваренных SKU (оборач > 90д): {overstock_count}
- Убыточных по хранению: {loss_making}
- Средняя оборачиваемость: {avg_turnover} дней

## ТОВАРЫ (отсортированы по стоимости хранения, от дорогих к дешёвым):
"""

        for i, s in enumerate(skus_for_ai):
            if s.get('must_include'):
                prompt += f"\n### ⚠️ #{i+1}. {s['vendor_code']} — {s['name']} [ТОП-{s['storage_rank']} ПО ХРАНЕНИЮ — ОБЯЗАТЕЛЬНО ВКЛЮЧИТЬ В sku_actions!]\n"
            else:
                prompt += f"\n### {i+1}. {s['vendor_code']} — {s['name']}\n"

            prompt += f"Остаток: {s['stock']} шт | Заказов за {period}д: {s['orders']} ({s['daily']}/день)\n"
            prompt += f"Оборачиваемость: {s['turnover_days']}д | Объём: {s['vol_liters']}л\n"
            prompt += f"Хранение (факт): {s['storage_30d']}₽/30д | ★ ПРОГНОЗ 30д (с учётом продаж): {s['forecast_30d']}₽\n"
            prompt += f"Факт за период: {s['storage_period']}₽\n"
            prompt += f"Цена: {s['price']}₽ | Себестоимость: {s['cost_price']}₽\n"

            if s['net_profit']:
                prompt += f"★ Чистая прибыль за {period}д: {s['net_profit']}₽ (маржа {s['margin_pct']}%)\n"
                prompt += f"  P&L: payout {s['pnl_payout']}₽ → логистика −{s['pnl_logistics']}₽ → хранение −{s['pnl_storage_fact']}₽ → удержания −{s['pnl_deductions']}₽ → COGS −{s['pnl_cogs']}₽\n"

            prompt += f"Реклама: {'ЗАПУЩЕНА, расход ' + str(s['ad_spend']) + '₽, DRR ' + str(s['drr']) + '%' if s['in_ads'] else 'НЕТ'}\n"
            prompt += f"Стоимость вывоза 1 шт: ~{s['withdrawal_cost_per_unit']}₽\n"

            if s['warehouses']:
                wh_str = ", ".join(f"{wh['warehouse']}={wh['qty']}шт" for wh in s['warehouses'])
                prompt += f"Склады: {wh_str}\n"

        prompt += f"\nПроанализируй КАЖДЫЙ из {len(skus_for_ai)} товаров выше и выдай JSON с sku_actions (ровно {len(skus_for_ai)} записей — по ОДНОЙ на КАЖДЫЙ товар), key_metrics и general_tips."

        # ── 9. Call Gemini ──
        KIE_AI_URL = "https://api.kie.ai/gemini-2.5-flash/v1/chat/completions"

        async with httpx.AsyncClient(timeout=120.0) as client:
            resp = await client.post(
                KIE_AI_URL,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {api_key}",
                },
                json={
                    "messages": [
                        {"role": "system", "content": [{"type": "text", "text": _AI_PROMPT_STORAGE}]},
                        {"role": "user", "content": [{"type": "text", "text": prompt}]},
                    ],
                    "stream": False,
                    "include_thoughts": False,
                    "temperature": 0,
                    "max_output_tokens": 16384,
                },
            )

        if resp.status_code != 200:
            logger.error("Gemini API error %s: %s", resp.status_code, resp.text[:500])
            raise HTTPException(status_code=502, detail="AI service unavailable")

        resp_json = resp.json()
        content = resp_json.get("choices", [{}])[0].get("message", {}).get("content", "")

        # Robust JSON extraction: find outermost { ... } block
        import re as _re
        content = content.strip()
        # Remove markdown code fences
        content = _re.sub(r'^```(?:json)?\s*', '', content)
        content = _re.sub(r'\s*```$', '', content)
        content = content.strip()
        # Remove leading non-JSON text (e.g. "Я проанализировала...\n---\n")
        first_brace = content.find('{')
        if first_brace > 0:
            content = content[first_brace:]
        # Remove trailing non-JSON text after the closing brace
        # Find matching closing brace by counting
        depth = 0
        last_brace = -1
        for ci_idx, ch in enumerate(content):
            if ch == '{':
                depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    last_brace = ci_idx
                    break
        if last_brace > 0:
            content = content[:last_brace + 1]

        try:
            ai_result = json.loads(content)
        except json.JSONDecodeError:
            # Last resort: try to find JSON block with regex
            json_match = _re.search(r'\{[\s\S]*"sku_actions"[\s\S]*\}', content)
            if json_match:
                try:
                    ai_result = json.loads(json_match.group())
                except json.JSONDecodeError:
                    logger.warning("Failed to parse AI storage JSON (retry): %s", content[:500])
                    raise HTTPException(status_code=502, detail="AI returned invalid JSON")
            else:
                logger.warning("Failed to parse AI storage JSON: %s", content[:500])
                raise HTTPException(status_code=502, detail="AI returned invalid JSON")

        # ── 10. Force-inject top-5 storage SKUs that AI missed ──
        ai_sku_actions = ai_result.get("sku_actions", [])
        ai_vendor_codes = {sa.get("vendor_code", "") for sa in ai_sku_actions}

        for s in skus_for_ai:
            if not s.get("must_include"):
                continue
            if s["vendor_code"] in ai_vendor_codes:
                continue

            storage_cost = s["storage_30d"]
            net_profit = s.get("net_profit", 0)
            stock = s["stock"]
            daily = s["daily"]
            vol = s["vol_liters"]
            wc = s["withdrawal_cost_per_unit"]
            turnover = s["turnover_days"]

            options = []
            # Option: discount
            if stock > 0:
                options.append({
                    "action": "discount",
                    "label": "Снизить цену на 20-30%",
                    "detail": f"Остаток {stock} шт, хранение {storage_cost}₽/мес. Скидка ускорит продажи. При росте в 1.5× оборач. снизится до ~{max(turnover // 2, 30)}д.",
                    "expected_savings": round(storage_cost * 0.3),
                    "withdrawal_cost": 0,
                    "risk": "medium",
                })

            # Option: withdraw
            if stock > 10 and daily > 0:
                ideal = round(daily * 60)
                excess = max(stock - ideal, 0)
                if excess > 0:
                    total_wc = excess * wc
                    monthly_savings = round(storage_cost * excess / stock)
                    options.append({
                        "action": "withdraw",
                        "label": f"Вывезти {excess} шт излишков",
                        "detail": f"Стоимость вывоза: {excess}×{wc}₽ = {total_wc}₽. Экономия: ~{monthly_savings}₽/мес. Окупится за {round(total_wc / monthly_savings, 1) if monthly_savings > 0 else '∞'} мес.",
                        "expected_savings": monthly_savings,
                        "withdrawal_cost": total_wc,
                        "risk": "medium",
                    })
            elif stock > 0 and daily == 0:
                total_wc = stock * wc
                options.append({
                    "action": "withdraw",
                    "label": f"Вывезти все {stock} шт",
                    "detail": f"Нет продаж. Стоимость вывоза: {stock}×{wc}₽ = {total_wc}₽. Экономия: {storage_cost}₽/мес. Окупится за {round(total_wc / storage_cost, 1) if storage_cost > 0 else '∞'} мес.",
                    "expected_savings": storage_cost,
                    "withdrawal_cost": total_wc,
                    "risk": "high",
                })

            # Option: do nothing
            options.append({
                "action": "do_nothing",
                "label": "Оставить как есть",
                "detail": f"Хранение {storage_cost}₽/мес. Через 6 мес: +{storage_cost * 6}₽ расходов.",
                "expected_savings": 0,
                "withdrawal_cost": 0,
                "risk": "high" if net_profit < storage_cost else "medium",
            })

            ai_sku_actions.insert(0, {
                "vendor_code": s["vendor_code"],
                "name": s["name"],
                "diagnosis": f"Оборачиваемость {turnover}д, хранение {storage_cost}₽/мес, прибыль {net_profit}₽.",
                "current_storage_cost": storage_cost,
                "current_turnover_days": turnover,
                "stock": stock,
                "options": options,
                "recommended_option": 0,
            })

        ai_result["sku_actions"] = ai_sku_actions

        # Enrich with context
        ai_result["period_days"] = period
        ai_result["analyzed_at"] = int(time.time())
        ai_result["context"] = {
            "total_storage_30d": round(total_storage_30d),
            "total_net_profit": round(total_net_profit),
            "total_skus": len(skus_context),
            "total_stock": total_stock,
            "overstock_skus": overstock_count,
            "loss_making_skus": loss_making,
            "avg_turnover_days": avg_turnover,
        }

        # Cache
        _ai_cache[cache_key] = (time.time(), ai_result)

        return {**ai_result, "cached": False}

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Storage AI analysis failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Ошибка анализа: {str(e)}")


# ═══════════════════════════════════════════════════════════
#  OZON  Storage AI Analysis
# ═══════════════════════════════════════════════════════════

@router.post("/ozon/storage/ai-analysis")
async def get_ozon_storage_ai_analysis(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    force: bool = Query(False, description="Skip cache"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """AI-powered storage cost optimization analysis for Ozon shops."""

    shop_result = await db.execute(
        select(Shop).where(Shop.id == shop_id, Shop.user_id == current_user.id)
    )
    shop = shop_result.scalar_one_or_none()
    if not shop or shop.marketplace != "ozon":
        raise HTTPException(status_code=404, detail="Shop not found")
    shop_name = shop.name or f"Shop {shop_id}"

    cache_key = f"ozon_storage_ai_{shop_id}_{period}"

    if not force and cache_key in _ai_cache:
        ts, cached = _ai_cache[cache_key]
        if time.time() - ts < _AI_CACHE_TTL:
            return {**cached, "cached": True}

    api_key = os.getenv("KIE_AI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="AI API key not configured")

    try:
        from app.core.clickhouse import get_clickhouse_client

        ch = get_clickhouse_client()
        today = date.today()
        d_start = today - timedelta(days=period)

        # ── 1. Per-SKU storage from fact_ozon_placement_cost ──
        actual_storage_rows = ch.query("""
            SELECT
                offer_id,
                round(SUM(placement_cost), 2) AS storage_total,
                round(SUM(placement_cost) * (30 / {period:UInt32}), 2) AS storage_30d
            FROM mms_analytics.fact_ozon_placement_cost FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND period_to = (
                  SELECT max(period_to) FROM mms_analytics.fact_ozon_placement_cost FINAL
                  WHERE shop_id = {shop_id:UInt32}
              )
            GROUP BY offer_id
            HAVING storage_total != 0
            ORDER BY storage_total DESC
        """, parameters={
            "shop_id": shop_id, "period": max(period, 1),
        }).result_rows

        storage_by_offer: dict[str, dict] = {}
        for r in actual_storage_rows:
            oid = str(r[0]).strip()
            if oid:
                storage_by_offer[oid] = {
                    "storage_period": round(float(r[1])),
                    "storage_30d": round(float(r[2])),
                }

        # ── 2. Per-SKU orders, revenue ──
        sku_rows = ch.query("""
            SELECT
                s.sku,
                s.total_stock,
                coalesce(o.orders, 0) AS orders,
                coalesce(o.revenue, 0) AS revenue,
                coalesce(a.spend, 0) AS ad_spend,
                coalesce(a.ad_orders, 0) AS ad_orders
            FROM (
                SELECT sku, sum(free_to_sell) AS total_stock
                FROM mms_analytics.fact_ozon_warehouse_stocks FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND dt = (SELECT max(dt) FROM mms_analytics.fact_ozon_warehouse_stocks WHERE shop_id = {shop_id:UInt32})
                  AND warehouse_type = 'fbo'
                  AND free_to_sell > 0
                GROUP BY sku
            ) AS s
            LEFT JOIN (
                SELECT sku, sum(quantity) AS orders,
                       sum(price * quantity) AS revenue
                FROM mms_analytics.fact_ozon_orders
                WHERE shop_id = {shop_id:UInt32} AND order_date >= {d_start:Date}
                  AND status NOT IN ('cancelled')
                GROUP BY sku
            ) AS o ON o.sku = s.sku
            LEFT JOIN (
                SELECT sku, sum(money_spent) AS spend, sum(orders) AS ad_orders
                FROM mms_analytics.fact_ozon_ad_daily FINAL
                WHERE shop_id = {shop_id:UInt32} AND dt >= {d_start:Date}
                GROUP BY sku
            ) AS a ON a.sku = s.sku
        """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows

        sku_data: dict[int, dict] = {}
        for r in sku_rows:
            nm = int(r[0])
            sku_data[nm] = {
                "stock": int(r[1]),
                "orders": int(r[2]),
                "revenue": float(r[3]),
                "ad_spend": float(r[4]),
                "ad_orders": int(r[5]),
            }

        # ── 3. Product info from dim_ozon_products ──
        sku_ids = list(sku_data.keys())
        products_map: dict[int, dict] = {}
        if sku_ids:
            sku_list = ", ".join(str(x) for x in sku_ids)
            try:
                await db.rollback()
            except Exception:
                pass
            pg_rows = (await db.execute(
                text(f"""
                    SELECT sku, offer_id, name, marketing_price, old_price,
                           COALESCE(depth, 0), COALESCE(height, 0), COALESCE(width, 0),
                           COALESCE(volume_weight, 0)
                    FROM dim_ozon_products
                    WHERE shop_id = :sid AND sku IN ({sku_list})
                """),
                {"sid": shop_id},
            )).fetchall()
            for r in pg_rows:
                d, h, w = float(r[5] or 0), float(r[6] or 0), float(r[7] or 0)
                vw = float(r[8] or 0)
                if d > 0 and h > 0 and w > 0:
                    vol = (d * h * w) / 1_000_000
                elif vw > 0:
                    vol = vw * 2.87
                else:
                    vol = 0.5
                vol = max(round(vol, 2), 0.1)
                price = float(r[3]) if r[3] else 0
                products_map[int(r[0])] = {
                    "offer_id": r[1] or "",
                    "name": (r[2] or "")[:60],
                    "price": price,
                    "vol_liters": vol,
                }

        # ── 4. Cost prices ──
        cost_prices: dict[str, float] = {}
        try:
            cost_rows_pg = (await db.execute(
                text("SELECT offer_id, COALESCE(cost_price, 0) + COALESCE(packaging_cost, 0) FROM product_costs WHERE shop_id = :sid AND (cost_price > 0 OR packaging_cost > 0)"),
                {"sid": shop_id},
            )).fetchall()
            for r in cost_rows_pg:
                cost_prices[r[0]] = float(r[1])
        except Exception:
            pass

        # ── 5. Per-SKU P&L from fact_ozon_transactions ──
        sku_pnl: dict[str, dict] = {}
        try:
            pnl_rows = ch.query("""
                SELECT
                    offer_id,
                    sumIf(accruals_for_sale, type_name = 'Доставка и обработка покупателю' OR type_name = 'Доставка возврата покупателю') AS payout,
                    sumIf(abs(sale_commission), 1) AS commission,
                    sumIf(abs(amount), type_name = 'Услуги доставки относительно других продавцов') AS logistics,
                    sumIf(abs(amount), type_name LIKE '%размещени%' OR type_name LIKE '%хранени%') AS storage_fact,
                    sumIf(amount, type_name = 'Доставка и обработка покупателю') AS sale_amount
                FROM mms_analytics.fact_ozon_transactions FINAL
                WHERE shop_id = {shop_id:UInt32}
                  AND operation_date >= {d_start:Date} AND operation_date <= {d_end:Date}
                  AND offer_id != ''
                GROUP BY offer_id
            """, parameters={"shop_id": shop_id, "d_start": d_start, "d_end": today}).result_rows
            for r in pnl_rows:
                oid = str(r[0] or "").strip()
                if not oid:
                    continue
                payout = float(r[1] or 0)
                commission = float(r[2] or 0)
                logistics = float(r[3] or 0)
                storage_fact = float(r[4] or 0)
                sale_amount = float(r[5] or 0)
                revenue = sale_amount if sale_amount > 0 else payout + commission
                cogs_unit = cost_prices.get(oid, 0)
                net_profit = payout - logistics - storage_fact - cogs_unit
                sku_pnl[oid] = {
                    "revenue": round(revenue),
                    "payout": round(payout),
                    "logistics": round(logistics),
                    "storage_fact": round(storage_fact),
                    "commission": round(commission),
                    "net_profit": round(net_profit),
                    "margin_pct": round(net_profit / revenue * 100, 1) if revenue > 0 else 0,
                }
        except Exception as e:
            logger.warning("Ozon Storage AI: P&L query failed: %s", e)

        # ── 6. Per-warehouse stock ──
        ch2 = get_clickhouse_client()
        wh_stock_map: dict[int, list] = {}
        if sku_ids:
            sku_list_ch = ",".join(str(x) for x in sku_ids[:60])
            wh_rows = ch2.query(f"""
                SELECT sku, warehouse_name, free_to_sell
                FROM mms_analytics.fact_ozon_warehouse_stocks FINAL
                WHERE shop_id = {{shop_id:UInt32}}
                  AND dt = (SELECT max(dt) FROM mms_analytics.fact_ozon_warehouse_stocks WHERE shop_id = {{shop_id:UInt32}})
                  AND warehouse_type = 'fbo'
                  AND sku IN ({sku_list_ch})
                  AND free_to_sell > 0
                ORDER BY sku, free_to_sell DESC
            """, parameters={"shop_id": shop_id}).result_rows
            for wr in wh_rows:
                wh_stock_map.setdefault(int(wr[0]), []).append({
                    "warehouse": wr[1],
                    "qty": int(wr[2]),
                })
        ch2.close()
        ch.close()

        # ── 7. Build SKU context for AI ──
        skus_context = []
        total_storage_30d = 0
        total_net_profit = 0

        for sku_id, sd in sku_data.items():
            prod = products_map.get(sku_id, {})
            offer_id = prod.get("offer_id", "")
            if not offer_id:
                continue

            stock = sd["stock"]
            orders = sd["orders"]
            revenue = sd["revenue"]
            ad_spend = sd["ad_spend"]
            ad_orders = sd["ad_orders"]
            daily = orders / period if period > 0 else 0
            turnover = round(stock / daily) if daily > 0 else 9999
            drr = round(ad_spend / revenue * 100, 1) if revenue > 0 else 0
            vol = prod.get("vol_liters", 1.0)
            price = prod.get("price", 0)
            cost_price = cost_prices.get(offer_id, 0)

            # Storage cost
            st = storage_by_offer.get(offer_id, {})
            storage_30d = st.get("storage_30d", 0)
            storage_period = st.get("storage_period", 0)
            if not storage_30d:
                # Estimate: ~0.14 ₽/liter/day
                storage_30d = round(vol * 0.14 * stock * 30 / max(stock, 1))

            total_storage_30d += storage_30d

            # Forecast 30d: accounts for stock decreasing with sales
            forecast_30d = storage_30d  # fallback = flat
            if storage_30d > 0 and stock > 0:
                daily_cost = storage_30d / 30.0
                cost_per_unit = daily_cost / stock
                forecast = 0.0
                for day in range(30):
                    remaining = max(0, stock - daily * day)
                    if remaining <= 0:
                        break
                    forecast += cost_per_unit * remaining
                forecast_30d = round(forecast)

            pnl = sku_pnl.get(offer_id, {})
            net_profit = pnl.get("net_profit", 0)
            total_net_profit += net_profit

            withdrawal_cost_per_unit = round(50 + vol * 15)

            wh_list = wh_stock_map.get(sku_id, [])

            skus_context.append({
                "vendor_code": offer_id,
                "name": prod.get("name", ""),
                "sku": sku_id,
                "stock": stock,
                "orders": orders,
                "daily": round(daily, 2),
                "turnover_days": turnover,
                "revenue": round(revenue),
                "ad_spend": round(ad_spend),
                "ad_orders": ad_orders,
                "drr": drr,
                "in_ads": ad_spend > 0,
                "price": price,
                "cost_price": cost_price,
                "vol_liters": vol,
                "storage_30d": storage_30d,
                "storage_period": storage_period,
                "forecast_30d": forecast_30d,
                "net_profit": net_profit,
                "margin_pct": pnl.get("margin_pct", 0),
                "pnl_payout": pnl.get("payout", 0),
                "pnl_logistics": pnl.get("logistics", 0),
                "pnl_storage_fact": pnl.get("storage_fact", 0),
                "pnl_commission": pnl.get("commission", 0),
                "withdrawal_cost_per_unit": withdrawal_cost_per_unit,
                "warehouses": wh_list[:5],
            })

        skus_context.sort(key=lambda x: x["forecast_30d"], reverse=True)

        # Pre-filter: SKUs needing optimization
        skus_needing_action = []
        for s in skus_context:
            forecast = s["forecast_30d"]
            if forecast < 30:
                continue
            is_overstock = s["turnover_days"] > 90
            is_unprofitable = s["net_profit"] < forecast
            no_sales = s["daily"] == 0 and s["stock"] > 0
            if is_overstock or is_unprofitable or no_sales:
                skus_needing_action.append(s)

        existing_vcs = {s["vendor_code"] for s in skus_needing_action}
        for s in skus_context:
            if len(skus_needing_action) >= 5:
                break
            if s["vendor_code"] not in existing_vcs and s["forecast_30d"] > 0:
                skus_needing_action.append(s)
                existing_vcs.add(s["vendor_code"])

        skus_needing_action = skus_needing_action[:15]
        skus_needing_action.sort(key=lambda x: x["forecast_30d"], reverse=True)

        for i, s in enumerate(skus_needing_action):
            s["must_include"] = True
            s["storage_rank"] = i + 1

        skus_for_ai = skus_needing_action

        # ── 8. Build prompt ──
        total_stock = sum(s["stock"] for s in skus_context)
        total_forecast_30d = sum(s["forecast_30d"] for s in skus_context)
        overstock_count = sum(1 for s in skus_context if s["turnover_days"] > 90 and s["forecast_30d"] > 50)
        loss_making = sum(1 for s in skus_context if s["forecast_30d"] > 0 and s["net_profit"] < s["forecast_30d"])
        avg_turnover = round(sum(s["turnover_days"] for s in skus_context) / len(skus_context)) if skus_context else 0

        prompt = f"""Магазин: {shop_name} (Ozon)
Период анализа: {period} дней (с {d_start} по {today})

## ОБЩИЕ МЕТРИКИ ХРАНЕНИЯ:
- Общее хранение (факт) за 30д: ~{round(total_storage_30d)}₽
- ПРОГНОЗ хранения на 30д (с учётом продаж): ~{round(total_forecast_30d)}₽
- Общая чистая прибыль за {period}д: {round(total_net_profit)}₽
- Всего SKU на складе: {len(skus_context)}
- Общий остаток: {total_stock} шт
- Затоваренных SKU (оборач > 90д): {overstock_count}
- Убыточных по хранению: {loss_making}
- Средняя оборачиваемость: {avg_turnover} дней

## ТОВАРЫ (отсортированы по стоимости хранения, от дорогих к дешёвым):
"""

        for i, s in enumerate(skus_for_ai):
            if s.get('must_include'):
                prompt += f"\n### ⚠️ #{i+1}. {s['vendor_code']} — {s['name']} [ТОП-{s['storage_rank']} ПО ХРАНЕНИЮ — ОБЯЗАТЕЛЬНО ВКЛЮЧИТЬ В sku_actions!]\n"
            else:
                prompt += f"\n### {i+1}. {s['vendor_code']} — {s['name']}\n"

            prompt += f"Остаток: {s['stock']} шт | Заказов за {period}д: {s['orders']} ({s['daily']}/день)\n"
            prompt += f"Оборачиваемость: {s['turnover_days']}д | Объём: {s['vol_liters']}л\n"
            prompt += f"Хранение (факт): {s['storage_30d']}₽/30д | ★ ПРОГНОЗ 30д (с учётом продаж): {s['forecast_30d']}₽\n"
            prompt += f"Факт за период: {s['storage_period']}₽\n"
            prompt += f"Цена: {s['price']}₽ | Себестоимость: {s['cost_price']}₽\n"

            if s['net_profit']:
                prompt += f"★ Чистая прибыль за {period}д: {s['net_profit']}₽ (маржа {s['margin_pct']}%)\n"
                prompt += f"  P&L: payout {s['pnl_payout']}₽ → логистика −{s['pnl_logistics']}₽ → хранение −{s['pnl_storage_fact']}₽ → комиссия −{s['pnl_commission']}₽\n"

            prompt += f"Реклама: {'ЗАПУЩЕНА, расход ' + str(s['ad_spend']) + '₽, DRR ' + str(s['drr']) + '%' if s['in_ads'] else 'НЕТ'}\n"
            prompt += f"Стоимость вывоза 1 шт: ~{s['withdrawal_cost_per_unit']}₽\n"

            if s['warehouses']:
                wh_str = ", ".join(f"{wh['warehouse']}={wh['qty']}шт" for wh in s['warehouses'])
                prompt += f"Склады: {wh_str}\n"

        prompt += f"\nПроанализируй КАЖДЫЙ из {len(skus_for_ai)} товаров выше и выдай JSON с sku_actions (ровно {len(skus_for_ai)} записей — по ОДНОЙ на КАЖДЫЙ товар), key_metrics и general_tips."

        # ── 9. Call Gemini ──
        KIE_AI_URL = "https://api.kie.ai/gemini-2.5-flash/v1/chat/completions"

        async with httpx.AsyncClient(timeout=120.0) as client:
            resp = await client.post(
                KIE_AI_URL,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {api_key}",
                },
                json={
                    "messages": [
                        {"role": "system", "content": [{"type": "text", "text": _AI_PROMPT_STORAGE}]},
                        {"role": "user", "content": [{"type": "text", "text": prompt}]},
                    ],
                    "stream": False,
                    "include_thoughts": False,
                    "temperature": 0,
                    "max_output_tokens": 16384,
                },
            )

        if resp.status_code != 200:
            logger.error("Gemini API error %s: %s", resp.status_code, resp.text[:500])
            raise HTTPException(status_code=502, detail="AI service unavailable")

        resp_json = resp.json()
        content = resp_json.get("choices", [{}])[0].get("message", {}).get("content", "")

        import re as _re
        content = content.strip()
        content = _re.sub(r'^```(?:json)?\s*', '', content)
        content = _re.sub(r'\s*```$', '', content)
        content = content.strip()
        first_brace = content.find('{')
        if first_brace > 0:
            content = content[first_brace:]
        depth = 0
        last_brace = -1
        for ci_idx, ci_ch in enumerate(content):
            if ci_ch == '{':
                depth += 1
            elif ci_ch == '}':
                depth -= 1
                if depth == 0:
                    last_brace = ci_idx
                    break
        if last_brace > 0:
            content = content[:last_brace + 1]

        try:
            ai_result = json.loads(content)
        except json.JSONDecodeError:
            json_match = _re.search(r'\{[\s\S]*"sku_actions"[\s\S]*\}', content)
            if json_match:
                try:
                    ai_result = json.loads(json_match.group())
                except json.JSONDecodeError:
                    raise HTTPException(status_code=502, detail="AI returned invalid JSON")
            else:
                raise HTTPException(status_code=502, detail="AI returned invalid JSON")

        # ── 10. Force-inject missing SKUs ──
        ai_sku_actions = ai_result.get("sku_actions", [])
        ai_vendor_codes = {sa.get("vendor_code", "") for sa in ai_sku_actions}

        for s in skus_for_ai:
            if not s.get("must_include") or s["vendor_code"] in ai_vendor_codes:
                continue

            storage_cost = s["storage_30d"]
            net_profit = s.get("net_profit", 0)
            stock = s["stock"]
            daily = s["daily"]
            vol = s["vol_liters"]
            wc = s["withdrawal_cost_per_unit"]
            turnover = s["turnover_days"]

            options = []
            if stock > 0:
                options.append({
                    "action": "discount",
                    "label": "Снизить цену на 20-30%",
                    "detail": f"Остаток {stock} шт, хранение {storage_cost}₽/мес. Скидка ускорит продажи.",
                    "expected_savings": round(storage_cost * 0.3),
                    "withdrawal_cost": 0,
                    "risk": "medium",
                })
            if stock > 10 and daily > 0:
                ideal = round(daily * 60)
                excess = max(stock - ideal, 0)
                if excess > 0:
                    total_wc = excess * wc
                    monthly_savings = round(storage_cost * excess / stock)
                    options.append({
                        "action": "withdraw",
                        "label": f"Вывезти {excess} шт излишков",
                        "detail": f"Стоимость вывоза: {total_wc}₽. Экономия: ~{monthly_savings}₽/мес.",
                        "expected_savings": monthly_savings,
                        "withdrawal_cost": total_wc,
                        "risk": "medium",
                    })
            elif stock > 0 and daily == 0:
                total_wc = stock * wc
                options.append({
                    "action": "withdraw",
                    "label": f"Вывезти все {stock} шт",
                    "detail": f"Нет продаж. Стоимость вывоза: {total_wc}₽. Экономия: {storage_cost}₽/мес.",
                    "expected_savings": storage_cost,
                    "withdrawal_cost": total_wc,
                    "risk": "high",
                })
            options.append({
                "action": "do_nothing",
                "label": "Оставить как есть",
                "detail": f"Хранение {storage_cost}₽/мес.",
                "expected_savings": 0,
                "withdrawal_cost": 0,
                "risk": "high" if net_profit < storage_cost else "medium",
            })

            ai_sku_actions.insert(0, {
                "vendor_code": s["vendor_code"],
                "name": s["name"],
                "diagnosis": f"Оборачиваемость {turnover}д, хранение {storage_cost}₽/мес, прибыль {net_profit}₽.",
                "current_storage_cost": storage_cost,
                "current_turnover_days": turnover,
                "stock": stock,
                "options": options,
                "recommended_option": 0,
            })

        ai_result["sku_actions"] = ai_sku_actions
        ai_result["period_days"] = period
        ai_result["analyzed_at"] = int(time.time())
        ai_result["context"] = {
            "total_storage_30d": round(total_storage_30d),
            "total_net_profit": round(total_net_profit),
            "total_skus": len(skus_context),
            "total_stock": total_stock,
            "overstock_skus": overstock_count,
            "loss_making_skus": loss_making,
            "avg_turnover_days": avg_turnover,
        }

        _ai_cache[cache_key] = (time.time(), ai_result)

        return {**ai_result, "cached": False}

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Ozon Storage AI analysis failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Ошибка анализа: {str(e)}")


# ══════════════════════════════════════════════════════════════
# Stock Report — Excel (WB + Ozon)
# ══════════════════════════════════════════════════════════════

def _build_stock_report_excel(
    warehouses: list[dict],
    wh_name_key: str,
    sku_id_key: str,
    sku_label_key: str,
    sku_name_key: str,
    period: int,
    shop_id: int,
    marketplace: str,
):
    """Build a stock report Excel workbook with two sheets: По складам & По товарам.

    Common logic for both WB and Ozon — differences parametrized via keys.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Styles ──
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_fill_green = PatternFill("solid", fgColor="2E7D32")
    bold_font = Font(bold=True, size=11)
    bold_sm = Font(bold=True, size=10)
    oos_fill = PatternFill("solid", fgColor="FDECEA")       # light red
    oos_font = Font(bold=True, size=11, color="C0392B")
    low_fill = PatternFill("solid", fgColor="FFF8E1")       # light amber
    low_font = Font(bold=True, size=11, color="E65100")
    wh_fill = PatternFill("solid", fgColor="D9E2F3")
    wh_font = Font(bold=True, size=12)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(bottom=thin, top=thin, left=thin, right=thin)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")
    num_fmt = "#,##0"

    def style_header(ws, headers, fill):
        for ci, (name, w) in enumerate(headers, 1):
            c = ws.cell(1, ci, name)
            c.font = hdr_font
            c.fill = fill
            c.alignment = center
            c.border = border
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ════════════════════════════════════════════════════════════
    # Sheet 1: ПО СКЛАДАМ
    # ════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "По складам"

    headers1 = [
        ("Склад", 24), ("Артикул", 20), ("Название", 40),
        ("Остаток", 12), ("Заказов", 12), ("В день", 10),
        ("Запас дн", 12), ("Статус", 14),
    ]
    style_header(ws1, headers1, hdr_fill)

    row = 2
    for wh in sorted(warehouses, key=lambda w: w.get("orders", 0), reverse=True):
        wh_name = wh[wh_name_key]
        skus = wh.get("skus", [])
        if not skus:
            continue

        # Warehouse header row
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws1.cell(row, 1, f"📦 {wh_name}  —  Остаток: {wh.get('stock', 0):,}  |  Заказов: {wh.get('orders', 0):,}  |  SKU: {wh.get('sku_count', len(skus))}")
        c.font = wh_font
        c.fill = wh_fill
        c.alignment = Alignment(vertical="center")
        for ci in range(1, 9):
            ws1.cell(row, ci).border = border
        row += 1

        for sku in sorted(skus, key=lambda s: s.get("orders", 0), reverse=True):
            stock = sku.get("stock", 0)
            orders = sku.get("orders", 0)
            daily = sku.get("daily_sales", 0)
            days = sku.get("days_supply")
            is_oos = stock == 0 and orders > 0
            is_low = not is_oos and days is not None and days < 14 and stock > 0

            # Status label
            if is_oos:
                status = "🔴 OOS"
            elif is_low:
                status = "🟡 Дефицит"
            elif days is not None and days > 120:
                status = "🟣 Излишек"
            else:
                status = "✅ Норма"

            ws1.cell(row, 1, wh_name)
            ws1.cell(row, 2, sku.get(sku_label_key, ""))
            ws1.cell(row, 3, sku.get(sku_name_key, "")).alignment = wrap
            ws1.cell(row, 4, stock).number_format = num_fmt
            ws1.cell(row, 5, orders).number_format = num_fmt
            ws1.cell(row, 6, round(daily, 2)).number_format = "0.00"
            ws1.cell(row, 7, round(days, 1) if days is not None else "—")
            ws1.cell(row, 8, status)

            # Apply highlighting
            for ci in range(1, 9):
                ws1.cell(row, ci).border = border
                if is_oos:
                    ws1.cell(row, ci).fill = oos_fill
                    if ci in (4, 7, 8):
                        ws1.cell(row, ci).font = oos_font
                elif is_low:
                    ws1.cell(row, ci).fill = low_fill
                    if ci in (7, 8):
                        ws1.cell(row, ci).font = low_font
            row += 1

        row += 1  # blank row between warehouses

    # ════════════════════════════════════════════════════════════
    # Sheet 2: ПО ТОВАРАМ
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("По товарам")

    # Pivot: product → {wh_name: stock}
    product_map: dict[str, dict] = {}  # keyed by sku_id
    wh_names_set: set[str] = set()

    for wh in warehouses:
        wh_name = wh[wh_name_key]
        for sku in wh.get("skus", []):
            sku_id = str(sku.get(sku_id_key, ""))
            if sku_id not in product_map:
                product_map[sku_id] = {
                    "label": sku.get(sku_label_key, ""),
                    "name": sku.get(sku_name_key, ""),
                    "total_stock": 0,
                    "total_orders": 0,
                    "total_daily": 0.0,
                    "per_wh": {},
                }
            product_map[sku_id]["total_stock"] += sku.get("stock", 0)
            product_map[sku_id]["total_orders"] += sku.get("orders", 0)
            product_map[sku_id]["total_daily"] += sku.get("daily_sales", 0)
            product_map[sku_id]["per_wh"][wh_name] = sku.get("stock", 0)
            wh_names_set.add(wh_name)

    wh_names_list = sorted(wh_names_set)

    headers2 = [
        ("Артикул", 20), ("Название", 40),
        ("Остаток", 12), ("Заказов", 12), ("В день", 10),
        ("Запас дн", 12), ("Скл.", 8), ("Статус", 14),
    ]
    # Add per-warehouse columns
    for wn in wh_names_list:
        short_name = wn[:16] + "…" if len(wn) > 16 else wn
        headers2.append((short_name, 14))

    style_header(ws2, headers2, hdr_fill_green)

    # Sorted products: by orders desc
    products_sorted = sorted(product_map.items(), key=lambda x: x[1]["total_orders"], reverse=True)

    row = 2
    for sku_id, p in products_sorted:
        total_stock = p["total_stock"]
        total_orders = p["total_orders"]
        daily = p["total_daily"]
        days = total_stock / daily if daily > 0 else None
        wh_count = sum(1 for s in p["per_wh"].values() if s > 0)
        is_oos = total_stock == 0 and total_orders > 0
        is_low = not is_oos and days is not None and days < 14 and total_stock > 0

        if is_oos:
            status = "🔴 OOS"
        elif is_low:
            status = "🟡 Дефицит"
        elif days is not None and days > 120:
            status = "🟣 Излишек"
        else:
            status = "✅ Норма"

        ws2.cell(row, 1, p["label"]).font = bold_sm
        ws2.cell(row, 2, p["name"]).alignment = wrap
        ws2.cell(row, 3, total_stock).number_format = num_fmt
        ws2.cell(row, 4, total_orders).number_format = num_fmt
        ws2.cell(row, 5, round(daily, 2)).number_format = "0.00"
        ws2.cell(row, 6, round(days, 1) if days is not None else "—")
        ws2.cell(row, 7, wh_count)
        ws2.cell(row, 8, status)

        # Per-warehouse stock
        for wi, wn in enumerate(wh_names_list):
            stock_at_wh = p["per_wh"].get(wn, 0)
            c = ws2.cell(row, 9 + wi, stock_at_wh if stock_at_wh > 0 else "")
            c.number_format = num_fmt
            c.alignment = center
            # Highlight zero stock where product has orders
            if stock_at_wh == 0 and total_orders > 0:
                c.fill = PatternFill("solid", fgColor="FCE4EC")
                c.font = Font(color="B71C1C", size=10)

        # Apply row highlighting
        total_cols = 8 + len(wh_names_list)
        for ci in range(1, total_cols + 1):
            ws2.cell(row, ci).border = border
            if ci <= 8:  # main columns only
                if is_oos:
                    ws2.cell(row, ci).fill = oos_fill
                    if ci in (3, 6, 8):
                        ws2.cell(row, ci).font = oos_font
                elif is_low:
                    ws2.cell(row, ci).fill = low_fill
                    if ci in (6, 8):
                        ws2.cell(row, ci).font = low_font
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/wb/analytics/stock-report/excel")
async def wb_stock_report_excel(
    shop_id: int = Query(...),
    period: int = Query(30),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download WB stock report as formatted Excel (two sheets: По складам & По товарам).
    Includes OOS products (zero stock but with orders)."""
    from fastapi.responses import StreamingResponse
    from app.core.clickhouse import get_clickhouse_client

    result = await wb_warehouse_analytics(shop_id=shop_id, period=period, db=db, current_user=current_user)
    warehouses = result.get("warehouses", [])

    if not warehouses:
        raise HTTPException(404, "Нет данных по складам")

    # ── Inject OOS products (stock=0 but orders>0) ─────────
    ch = get_clickhouse_client()
    today = date.today()
    d_start = today - timedelta(days=period)

    # Collect all nm_ids that already have stock in some warehouse
    stocked_nm_ids: set[int] = set()
    for wh in warehouses:
        for sku in wh.get("skus", []):
            stocked_nm_ids.add(sku.get("nm_id", 0))

    # Get all ordered nm_ids per warehouse in the period
    oos_rows = ch.query("""
        SELECT warehouse_name, nm_id, count() AS orders
        FROM mms_analytics.fact_orders_raw
        WHERE shop_id = {shop_id:UInt32}
          AND date >= {d_start:Date}
          AND is_cancel = 0
        GROUP BY warehouse_name, nm_id
    """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows

    # Find OOS: nm_ids with orders but NOT in stocked_nm_ids
    oos_by_wh: dict[str, dict[int, int]] = {}  # wh → {nm_id → orders}
    all_oos_nm_ids: set[int] = set()
    for row in oos_rows:
        wh_name, nm_id, orders = row[0], int(row[1]), int(row[2])
        if nm_id not in stocked_nm_ids:
            oos_by_wh.setdefault(wh_name, {})[nm_id] = orders
            all_oos_nm_ids.add(nm_id)

    if all_oos_nm_ids:
        # Resolve product names for OOS nm_ids (dim_products is PostgreSQL!)
        from sqlalchemy import text as sa_text
        prod_result = await db.execute(
            sa_text("SELECT nm_id, vendor_code, name FROM dim_products WHERE shop_id = :shop_id AND nm_id = ANY(:nm_ids)"),
            {"shop_id": shop_id, "nm_ids": list(all_oos_nm_ids)},
        )
        prod_map = {int(r[0]): {"vendor_code": r[1] or str(r[0]), "name": r[2] or ""} for r in prod_result.fetchall()}

        # Add OOS SKUs to each warehouse
        wh_map = {wh["warehouse_name"]: wh for wh in warehouses}
        for wh_name, nm_orders in oos_by_wh.items():
            if wh_name not in wh_map:
                continue
            wh = wh_map[wh_name]
            for nm_id, orders in nm_orders.items():
                prod = prod_map.get(nm_id, {"vendor_code": str(nm_id), "name": ""})
                daily = orders / period if period > 0 else 0
                wh["skus"].append({
                    "nm_id": nm_id,
                    "vendor_code": prod["vendor_code"],
                    "name": prod["name"],
                    "stock": 0,
                    "daily_sales": round(daily, 2),
                    "days_supply": None,
                    "orders": orders,
                    "cross_orders": 0,
                    "cross_pct": 0,
                    "geography": [],
                })

    buf = _build_stock_report_excel(
        warehouses=warehouses,
        wh_name_key="warehouse_name",
        sku_id_key="nm_id",
        sku_label_key="vendor_code",
        sku_name_key="name",
        period=period,
        shop_id=shop_id,
        marketplace="wb",
    )

    filename = f"stock_report_wb_shop{shop_id}_{period}d.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/ozon/overview/stock-report/excel")
async def ozon_stock_report_excel(
    shop_id: int = Query(...),
    period: int = Query(30),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download Ozon stock report as formatted Excel (two sheets: По складам & По товарам).
    Includes OOS products (zero stock but with orders)."""
    from fastapi.responses import StreamingResponse
    from app.core.clickhouse import get_clickhouse_client

    result = await ozon_warehouse_overview(shop_id=shop_id, period=period, db=db, current_user=current_user)
    warehouses = result.get("warehouses", [])

    if not warehouses:
        raise HTTPException(404, "Нет данных по складам")

    # ── Inject OOS products (stock=0 but orders>0) ─────────
    ch = get_clickhouse_client()
    today = date.today()
    d_start = today - timedelta(days=period)

    # Collect all SKUs that already have stock
    stocked_skus: set[int] = set()
    for wh in warehouses:
        for sku in wh.get("skus", []):
            stocked_skus.add(int(sku.get("sku", 0)))

    # Get all ordered SKUs per warehouse in the period
    oos_rows = ch.query("""
        SELECT warehouse_name, sku, count() AS orders
        FROM fact_ozon_orders FINAL
        WHERE shop_id = {shop_id:UInt32}
          AND order_date >= {d_start:Date}
          AND status NOT IN ('cancelled')
        GROUP BY warehouse_name, sku
    """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows

    oos_by_wh: dict[str, dict[int, int]] = {}
    all_oos_skus: set[int] = set()
    for row in oos_rows:
        wh_name, sku_id, orders = row[0], int(row[1]), int(row[2])
        if sku_id not in stocked_skus:
            oos_by_wh.setdefault(wh_name, {})[sku_id] = orders
            all_oos_skus.add(sku_id)

    if all_oos_skus:
        # Resolve product names for OOS SKUs
        prod_rows = ch.query("""
            SELECT sku, offer_id, product_name
            FROM fact_ozon_warehouse_stocks FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND sku IN {sku_ids:Array(UInt64)}
        """, parameters={"shop_id": shop_id, "sku_ids": list(all_oos_skus)}).result_rows
        prod_map = {int(r[0]): {"offer_id": r[1], "name": r[2]} for r in prod_rows}

        # Fallback: try dim_ozon_products (PostgreSQL!)
        missing = all_oos_skus - set(prod_map.keys())
        if missing:
            from sqlalchemy import text as sa_text
            prod_result2 = await db.execute(
                sa_text("SELECT product_id, offer_id, name FROM dim_ozon_products WHERE shop_id = :shop_id AND product_id = ANY(:ids)"),
                {"shop_id": shop_id, "ids": list(missing)},
            )
            for r in prod_result2.fetchall():
                prod_map[int(r[0])] = {"offer_id": r[1] or str(r[0]), "name": r[2] or ""}

        # Add OOS SKUs to each warehouse
        wh_map = {wh["warehouse_name"]: wh for wh in warehouses}
        for wh_name, sku_orders in oos_by_wh.items():
            if wh_name not in wh_map:
                continue
            wh = wh_map[wh_name]
            for sku_id, orders in sku_orders.items():
                prod = prod_map.get(sku_id, {"offer_id": str(sku_id), "name": ""})
                daily = orders / period if period > 0 else 0
                wh["skus"].append({
                    "sku": sku_id,
                    "offer_id": prod["offer_id"],
                    "name": prod["name"],
                    "stock": 0,
                    "daily_sales": round(daily, 2),
                    "days_supply": None,
                    "orders": orders,
                    "cross_orders": 0,
                    "cross_pct": 0,
                })

    buf = _build_stock_report_excel(
        warehouses=warehouses,
        wh_name_key="warehouse_name",
        sku_id_key="sku",
        sku_label_key="offer_id",
        sku_name_key="name",
        period=period,
        shop_id=shop_id,
        marketplace="ozon",
    )

    filename = f"stock_report_ozon_shop{shop_id}_{period}d.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═══════════════════════════════════════════════════════════════
# Cross-logistics Excel export (WB + Ozon)
# ═══════════════════════════════════════════════════════════════

def _build_cross_excel(
    analytics_data: dict,
    shop_name: str,
    period: int,
    marketplace: str,
    ai_data: dict | None = None,
):
    """Build a 5-6 sheet Excel workbook for cross-logistics analysis."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_fill_green = PatternFill("solid", fgColor="548235")
    hdr_fill_red = PatternFill("solid", fgColor="C00000")
    totals_fill = PatternFill("solid", fgColor="D9E2F3")
    totals_font = Font(bold=True, size=11)
    red_font = Font(bold=True, color="CC0000")
    green_font = Font(bold=True, color="006600")
    amber_font = Font(bold=True, color="CC6600")
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11, color="444444")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(bottom=thin, left=thin, right=thin)
    num_fmt = "#,##0"
    money_fmt = '#,##0" ₽"'
    pct_fmt = '0.0"%"'
    alt_fill = PatternFill("solid", fgColor="F5F7FA")
    local_fill = PatternFill("solid", fgColor="E2EFDA")
    cross_fill = PatternFill("solid", fgColor="FCE4EC")

    is_wb = marketplace == "wildberries"
    kpi = analytics_data.get("kpi", {})
    raw_warehouses = analytics_data.get("warehouses", [])
    cross_map_raw = analytics_data.get("cross_map", [])
    region_list = analytics_data.get("okrug_list" if is_wb else "cluster_list", [])

    # ── Normalize warehouses (Ozon uses different field names) ──
    warehouses = []
    for w in raw_warehouses:
        if is_wb:
            wh = {
                "warehouse_name": w.get("warehouse_name", ""),
                "region": w.get("okrug", ""),
                "orders": w.get("orders", 0),
                "cross_orders": w.get("cross_orders", 0),
                "local_orders": w.get("local_orders", 0),
                "cross_pct": w.get("cross_pct", 0),
                "logistics_cost": w.get("logistics_cost", 0),
                "stock": w.get("stock", 0),
                "sku_count": w.get("sku_count", 0),
                "turnover_days": w.get("turnover_days"),
                "status": w.get("status", "ok"),
                "revenue": w.get("revenue", 0),
                "geography": w.get("geography", []),
                "skus": w.get("skus", []),
            }
        else:
            costs = w.get("costs", {})
            logistics = abs(costs.get("crossdocking", 0)) + abs(costs.get("fbo_processing", 0))
            wh = {
                "warehouse_name": w.get("warehouse_name", ""),
                "region": w.get("cluster", ""),
                "orders": w.get("orders_period", w.get("orders", 0)),
                "cross_orders": w.get("cross_orders", 0),
                "local_orders": w.get("local_orders", 0),
                "cross_pct": w.get("cross_pct", 0),
                "logistics_cost": logistics,
                "stock": w.get("stock_free", w.get("stock", 0)),
                "sku_count": w.get("sku_count", 0),
                "turnover_days": w.get("turnover_days"),
                "status": w.get("status", "ok"),
                "revenue": w.get("revenue_period", w.get("revenue", 0)),
                "geography": [
                    {"okrug": cs.get("cluster", cs.get("okrug", "")),
                     "orders": cs.get("orders", 0),
                     "share": cs.get("share", 0),
                     "is_local": cs.get("is_local", False)}
                    for cs in w.get("clusters_served", w.get("geography", []))
                ],
                "skus": [],
            }
            # Normalize SKUs
            for s in w.get("skus", []):
                geo = s.get("geography", [])
                norm_geo = [
                    {"okrug": g.get("cluster", g.get("okrug", "")),
                     "orders": g.get("orders", 0),
                     "share": g.get("share", 0),
                     "is_local": g.get("is_local", False)}
                    for g in geo
                ]
                wh["skus"].append({
                    "nm_id": s.get("nm_id", s.get("sku", 0)),
                    "vendor_code": s.get("vendor_code", s.get("offer_id", "")),
                    "name": s.get("name", ""),
                    "stock": s.get("stock", 0),
                    "daily_sales": s.get("daily_sales", 0),
                    "days_supply": s.get("days_supply"),
                    "orders": s.get("orders", 0),
                    "cross_orders": s.get("cross_orders", 0),
                    "cross_pct": s.get("cross_pct", 0),
                    "geography": norm_geo,
                })
        warehouses.append(wh)

    # Normalize cross_map
    cross_map = []
    for row in cross_map_raw:
        regions_data = row.get("okrugs", row.get("clusters", {}))
        cross_map.append({
            "warehouse": row.get("warehouse", ""),
            "home_region": row.get("home_okrug", row.get("home_cluster", "")),
            "total_orders": row.get("total_orders", 0),
            "regions": regions_data,
        })

    # If region_list is empty, build from cross_map
    if not region_list:
        rl_set = set()
        for row in cross_map:
            for k in row["regions"]:
                rl_set.add(k)
        region_list = sorted(rl_set)

    workbook = openpyxl.Workbook()

    # ── Aggregate metrics ──
    cross_cost = 0
    total_cross_orders = 0
    total_orders = 0
    total_logistics = 0.0
    for w in warehouses:
        w_orders = w["orders"]
        w_cross = w["cross_orders"]
        total_cross_orders += w_cross
        total_orders += w_orders
        w_log = float(w.get("logistics_cost", 0) or 0)
        total_logistics += w_log
        if w_orders > 0 and w_log > 0:
            cross_cost += float(w_log) * (w_cross / w_orders)

    all_skus = []
    for w in warehouses:
        for s in w["skus"]:
            all_skus.append({**s, "_wh": w["warehouse_name"], "_wh_region": w["region"]})

    problem_skus = [s for s in all_skus if s.get("orders", 0) >= 5 and s.get("cross_pct", 0) > 40]
    critical_whs = [w for w in warehouses if w["cross_pct"] > 50 and w["orders"] >= 5]

    # ═══════════════════════════════════════════
    # Sheet 1: Сводка
    # ═══════════════════════════════════════════
    ws1 = workbook.active
    ws1.title = "Сводка"
    ws1.column_dimensions["A"].width = 40
    ws1.column_dimensions["B"].width = 25

    ws1.cell(1, 1, f"Кросс-логистика — {shop_name}").font = title_font
    mp_label = "Wildberries" if is_wb else "Ozon"
    ws1.cell(2, 1, f"Маркетплейс: {mp_label} • Период: {period} дней • {datetime.now().strftime('%d.%m.%Y')}").font = subtitle_font

    kpi_rows = [
        ("Средний кросс %", f"{kpi.get('cross_pct', 0)}%"),
        ("Кросс-заказов (межрегионых)", total_cross_orders),
        ("Всего заказов за период", total_orders),
        ("Локальных заказов", total_orders - total_cross_orders),
        ("", ""),
        ("Общая логистика ₽", round(total_logistics)),
        ("≈ Оценка кросс-логистики ₽", round(cross_cost)),
        ("≈ Потери на кросс (доля)", f"{round(cross_cost / total_logistics * 100, 1)}%" if total_logistics > 0 else "—"),
        ("", ""),
        ("Складов активных", len(warehouses)),
        ("Критических складов (кросс >50%)", len(critical_whs)),
        ("Проблемных SKU (кросс >40%)", len(problem_skus)),
        ("Общий остаток (шт)", kpi.get("total_stock", sum(w["stock"] for w in warehouses))),
    ]

    for ri, (label, value) in enumerate(kpi_rows, 4):
        if not label:
            continue
        ws1.cell(ri, 1, label).font = Font(bold=True, size=11)
        c = ws1.cell(ri, 2, value)
        c.font = Font(bold=True, size=11)
        if isinstance(value, (int, float)):
            c.number_format = num_fmt

    # ═══════════════════════════════════════════
    # Sheet 2: По складам
    # ═══════════════════════════════════════════
    ws2 = workbook.create_sheet("По складам")

    region_col = "Округ" if is_wb else "Кластер"
    wh_headers = [
        ("Склад", 24), (region_col, 20), ("Заказов", 10), ("Локальных", 10),
        ("Кросс", 10), ("Кросс %", 10), ("Логистика ₽", 14), ("≈ Кросс ₽", 14),
        ("Выручка ₽", 14), ("Остаток", 10), ("SKU", 8), ("Оборот, дн", 10),
    ]

    for ci, (name, w) in enumerate(wh_headers, 1):
        c = ws2.cell(1, ci, name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(wh_headers))}1"

    sorted_whs = sorted(warehouses, key=lambda w: w["cross_pct"], reverse=True)
    t_o = t_l = t_c = t_log = t_cc = t_rev = t_st = 0

    for ri, w in enumerate(sorted_whs, 2):
        o = w["orders"]
        c_ord = w["cross_orders"]
        l_ord = w["local_orders"] or (o - c_ord)
        cp = w["cross_pct"]
        log = float(w.get("logistics_cost", 0) or 0)
        cc = round(log * (c_ord / o)) if o > 0 and log > 0 else 0
        rev = float(w.get("revenue", 0) or 0)

        ws2.cell(ri, 1, w["warehouse_name"])
        ws2.cell(ri, 2, w["region"])
        ws2.cell(ri, 3, o).number_format = num_fmt
        ws2.cell(ri, 4, l_ord).number_format = num_fmt
        ws2.cell(ri, 5, c_ord).number_format = num_fmt
        pc = ws2.cell(ri, 6, cp)
        pc.number_format = pct_fmt
        pc.font = red_font if cp > 50 else (amber_font if cp > 25 else green_font)
        ws2.cell(ri, 7, round(log)).number_format = money_fmt
        ws2.cell(ri, 8, cc).number_format = money_fmt
        ws2.cell(ri, 9, round(rev)).number_format = money_fmt
        ws2.cell(ri, 10, w["stock"]).number_format = num_fmt
        ws2.cell(ri, 11, w["sku_count"])
        ws2.cell(ri, 12, round(w["turnover_days"]) if w["turnover_days"] else "—")

        if ri % 2 == 0:
            for ci in range(1, len(wh_headers) + 1):
                ws2.cell(ri, ci).fill = alt_fill

        t_o += o; t_l += l_ord; t_c += c_ord; t_log += log; t_cc += cc; t_rev += rev; t_st += w["stock"]

    tr = len(sorted_whs) + 2
    for ci in range(1, len(wh_headers) + 1):
        ws2.cell(tr, ci).fill = totals_fill
        ws2.cell(tr, ci).font = totals_font
    ws2.cell(tr, 1, "ИТОГО")
    ws2.cell(tr, 3, t_o).number_format = num_fmt
    ws2.cell(tr, 4, t_l).number_format = num_fmt
    ws2.cell(tr, 5, t_c).number_format = num_fmt
    ws2.cell(tr, 6, round(t_c / t_o * 100, 1) if t_o > 0 else 0).number_format = pct_fmt
    ws2.cell(tr, 7, round(t_log)).number_format = money_fmt
    ws2.cell(tr, 8, round(t_cc)).number_format = money_fmt
    ws2.cell(tr, 9, round(t_rev)).number_format = money_fmt
    ws2.cell(tr, 10, t_st).number_format = num_fmt

    # ═══════════════════════════════════════════
    # Sheet 3: Кросс-карта (Склад × Регион)
    # ═══════════════════════════════════════════
    if cross_map and region_list:
        ws4 = workbook.create_sheet("Кросс-карта")

        short = lambda s: s.replace(" федеральный округ", "").replace("Центральный", "ЦФО").replace("Северо-Западный", "СЗФО").replace("Южный", "ЮФО").replace("Приволжский", "ПФО").replace("Уральский", "УФО").replace("Сибирский", "СФО").replace("Дальневосточный", "ДФО").replace("Северо-Кавказский", "СКФО")

        # Row 1: Headers
        ws4.cell(1, 1, "Склад").font = Font(bold=True, size=11, color="FFFFFF")
        ws4.cell(1, 1).fill = hdr_fill
        ws4.cell(1, 1).alignment = Alignment(horizontal="center")
        ws4.column_dimensions["A"].width = 24

        ws4.cell(1, 2, "Дом. регион").font = Font(bold=True, size=10, color="FFFFFF")
        ws4.cell(1, 2).fill = PatternFill("solid", fgColor="548235")
        ws4.cell(1, 2).alignment = Alignment(horizontal="center")
        ws4.column_dimensions["B"].width = 14

        for ci, reg in enumerate(region_list, 3):
            c = ws4.cell(1, ci, short(reg))
            c.font = Font(bold=True, size=10, color="FFFFFF")
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            ws4.column_dimensions[get_column_letter(ci)].width = 11

        total_col = len(region_list) + 3
        local_col = total_col + 1
        cross_col_idx = total_col + 2
        pct_col = total_col + 3

        for ci, (lbl, clr) in [(total_col, ("Всего", "2F5496")), (local_col, ("Локал.", "548235")), (cross_col_idx, ("Кросс", "C00000")), (pct_col, ("Кросс %", "C00000"))]:
            c = ws4.cell(1, ci, lbl)
            c.font = Font(bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=clr)
            c.alignment = Alignment(horizontal="center")
            ws4.column_dimensions[get_column_letter(ci)].width = 10

        ws4.freeze_panes = "C2"

        # Column totals accumulator
        col_totals = {reg: {"local": 0, "cross": 0} for reg in region_list}
        grand_total = grand_local = grand_cross = 0

        for ri, row in enumerate(cross_map, 2):
            ws4.cell(ri, 1, row["warehouse"]).font = Font(bold=True, size=10)
            home = row.get("home_region", "")
            ws4.cell(ri, 2, short(home)).font = Font(bold=True, size=9, color="548235")
            ws4.cell(ri, 2).alignment = Alignment(horizontal="center")

            row_total = row["total_orders"]
            row_local = 0
            row_cross = 0

            for ci, reg in enumerate(region_list, 3):
                cell_data = row["regions"].get(reg, {})
                count = cell_data.get("count", 0) if isinstance(cell_data, dict) else 0
                is_local = cell_data.get("is_local", False) if isinstance(cell_data, dict) else False

                if count > 0:
                    c = ws4.cell(ri, ci, count)
                    c.number_format = num_fmt
                    c.alignment = Alignment(horizontal="center")
                    c.font = Font(bold=True, color="006600" if is_local else "CC0000")
                    c.fill = local_fill if is_local else cross_fill
                    if is_local:
                        row_local += count
                        col_totals[reg]["local"] += count
                    else:
                        row_cross += count
                        col_totals[reg]["cross"] += count
                else:
                    c = ws4.cell(ri, ci, "")
                    c.alignment = Alignment(horizontal="center")

            ws4.cell(ri, total_col, row_total).number_format = num_fmt
            ws4.cell(ri, total_col).font = totals_font
            ws4.cell(ri, total_col).alignment = Alignment(horizontal="center")

            ws4.cell(ri, local_col, row_local).number_format = num_fmt
            ws4.cell(ri, local_col).font = green_font
            ws4.cell(ri, local_col).alignment = Alignment(horizontal="center")

            ws4.cell(ri, cross_col_idx, row_cross).number_format = num_fmt
            ws4.cell(ri, cross_col_idx).font = red_font
            ws4.cell(ri, cross_col_idx).alignment = Alignment(horizontal="center")

            cross_pct_val = round(row_cross / row_total * 100, 1) if row_total > 0 else 0
            pc = ws4.cell(ri, pct_col, cross_pct_val)
            pc.number_format = pct_fmt
            pc.font = red_font if cross_pct_val > 50 else (amber_font if cross_pct_val > 25 else green_font)
            pc.alignment = Alignment(horizontal="center")

            grand_total += row_total
            grand_local += row_local
            grand_cross += row_cross

        # Totals row
        tr = len(cross_map) + 2
        ws4.cell(tr, 1, "ИТОГО").font = Font(bold=True, size=11)
        ws4.cell(tr, 1).fill = totals_fill
        ws4.cell(tr, 2).fill = totals_fill
        for ci, reg in enumerate(region_list, 3):
            ct = col_totals[reg]["local"] + col_totals[reg]["cross"]
            if ct > 0:
                c = ws4.cell(tr, ci, ct)
                c.number_format = num_fmt
                c.font = totals_font
            c = ws4.cell(tr, ci)
            c.fill = totals_fill
            c.alignment = Alignment(horizontal="center")

        ws4.cell(tr, total_col, grand_total).number_format = num_fmt
        ws4.cell(tr, total_col).font = totals_font
        ws4.cell(tr, total_col).fill = totals_fill
        ws4.cell(tr, local_col, grand_local).number_format = num_fmt
        ws4.cell(tr, local_col).font = green_font
        ws4.cell(tr, local_col).fill = totals_fill
        ws4.cell(tr, cross_col_idx, grand_cross).number_format = num_fmt
        ws4.cell(tr, cross_col_idx).font = red_font
        ws4.cell(tr, cross_col_idx).fill = totals_fill
        gp = round(grand_cross / grand_total * 100, 1) if grand_total > 0 else 0
        ws4.cell(tr, pct_col, gp).number_format = pct_fmt
        ws4.cell(tr, pct_col).font = totals_font
        ws4.cell(tr, pct_col).fill = totals_fill

    # ═══════════════════════════════════════════
    # Sheet 4: По товарам (SKU) + география
    # ═══════════════════════════════════════════
    ws3 = workbook.create_sheet("По товарам (SKU)")

    id_col = "Артикул" if is_wb else "Offer ID"
    id2_col = "nm_id" if is_wb else "SKU"
    sku_headers = [
        (id_col, 22), (id2_col, 14), ("Название", 38), ("Склад", 18),
        ("Остаток", 8), ("Заказов", 8), ("Кросс", 8), ("Кросс %", 8),
        ("≈ Потери ₽", 12), ("Откуда", 15), ("→ Куда (кросс-регионы)", 35),
        ("Рекомендация", 30),
    ]

    for ci, (name, w) in enumerate(sku_headers, 1):
        c = ws3.cell(1, ci, name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border
        ws3.column_dimensions[get_column_letter(ci)].width = w

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(sku_headers))}1"

    sorted_skus = sorted(all_skus, key=lambda s: s.get("cross_orders", 0), reverse=True)

    for ri, s in enumerate(sorted_skus, 2):
        s_orders = s.get("orders", 0)
        s_cross = s.get("cross_orders", 0)
        s_cross_pct = s.get("cross_pct", 0)
        vendor_code = s.get("vendor_code", "")
        nm_id = s.get("nm_id", 0)
        name = s.get("name", "")
        wh_name = s.get("_wh", "")
        wh_region = s.get("_wh_region", "")
        stock = s.get("stock", 0)

        wh_data = next((w for w in warehouses if w["warehouse_name"] == wh_name), None)
        sku_loss = 0
        if wh_data and wh_data["orders"] > 0 and wh_data["logistics_cost"] > 0:
            sku_loss = round(float(wh_data["logistics_cost"]) * (s_cross / wh_data["orders"]))

        geography = s.get("geography", [])
        cross_geos = [g for g in geography if not g.get("is_local", True)]
        cross_details = []
        for g in cross_geos[:5]:
            okr = g.get("okrug", "").replace(" федеральный округ", "")
            cross_details.append(f"{okr} ({g.get('orders', 0)} зак.)")

        # Recommendation
        rec = ""
        if s_cross_pct > 50 and s_orders >= 5:
            top_cross = [g.get("okrug", "").replace(" федеральный округ", "") for g in cross_geos[:2]]
            rec = f"Довезти на склад в {', '.join(top_cross)}" if top_cross else "Перераспределить"
        elif s_cross_pct > 25 and s_orders >= 5:
            rec = "Мониторить, возможен довоз"

        ws3.cell(ri, 1, vendor_code)
        ws3.cell(ri, 2, nm_id)
        ws3.cell(ri, 3, name)
        ws3.cell(ri, 4, wh_name)
        ws3.cell(ri, 5, stock).number_format = num_fmt
        ws3.cell(ri, 6, s_orders).number_format = num_fmt
        ws3.cell(ri, 7, s_cross).number_format = num_fmt
        pc = ws3.cell(ri, 8, s_cross_pct)
        pc.number_format = pct_fmt
        pc.font = red_font if s_cross_pct > 50 else (amber_font if s_cross_pct > 25 else green_font)
        ws3.cell(ri, 9, sku_loss).number_format = money_fmt
        ws3.cell(ri, 10, wh_region.replace(" федеральный округ", ""))
        ws3.cell(ri, 11, "; ".join(cross_details) if cross_details else "—")
        ws3.cell(ri, 12, rec)

        if ri % 2 == 0:
            for ci in range(1, len(sku_headers) + 1):
                ws3.cell(ri, ci).fill = alt_fill

    # ═══════════════════════════════════════════
    # Sheet 5: География складов (распределение заказов)
    # ═══════════════════════════════════════════
    ws5 = workbook.create_sheet("География складов")

    geo_headers = [
        ("Склад", 24), ("Домашний регион", 18), ("Регион доставки", 22),
        ("Тип", 8), ("Заказов", 10), ("Доля %", 8),
    ]

    for ci, (name, w) in enumerate(geo_headers, 1):
        c = ws5.cell(1, ci, name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border
        ws5.column_dimensions[get_column_letter(ci)].width = w

    ws5.freeze_panes = "A2"
    ws5.auto_filter.ref = f"A1:{get_column_letter(len(geo_headers))}1"

    ri = 2
    for w in sorted(warehouses, key=lambda w: w["cross_pct"], reverse=True):
        if not w["geography"]:
            continue
        sorted_geo = sorted(w["geography"], key=lambda g: g.get("orders", 0), reverse=True)
        for g in sorted_geo:
            okrug = g.get("okrug", "").replace(" федеральный округ", "")
            is_local = g.get("is_local", False)
            ws5.cell(ri, 1, w["warehouse_name"])
            ws5.cell(ri, 2, w["region"].replace(" федеральный округ", ""))
            ws5.cell(ri, 3, okrug)
            type_cell = ws5.cell(ri, 4, "СВОЙ" if is_local else "КРОСС")
            type_cell.font = green_font if is_local else red_font
            type_cell.fill = local_fill if is_local else cross_fill
            ws5.cell(ri, 5, g.get("orders", 0)).number_format = num_fmt
            ws5.cell(ri, 6, g.get("share", 0)).number_format = pct_fmt

            if ri % 2 == 0:
                for ci in [1, 2, 3, 5, 6]:
                    ws5.cell(ri, ci).fill = alt_fill
            ri += 1

    # ═══════════════════════════════════════════
    # Sheet 6: ИИ-анализ (if available)
    # ═══════════════════════════════════════════
    if ai_data and isinstance(ai_data, dict) and ai_data.get("severity"):
        ws_ai = workbook.create_sheet("ИИ-анализ")
        ws_ai.column_dimensions["A"].width = 20
        ws_ai.column_dimensions["B"].width = 70
        ws_ai.column_dimensions["C"].width = 15

        ws_ai.cell(1, 1, f"ИИ-анализ кросс-логистики — {shop_name}").font = title_font
        ws_ai.cell(2, 1, f"Дата анализа: {datetime.now().strftime('%d.%m.%Y %H:%M')}").font = subtitle_font

        severity = ai_data.get("severity", "")
        sev_colors = {"critical": "CC0000", "warning": "CC6600", "ok": "006600", "info": "2F5496"}
        ws_ai.cell(4, 1, "Степень проблемы").font = Font(bold=True, size=12)
        sev_cell = ws_ai.cell(4, 2, severity.upper())
        sev_cell.font = Font(bold=True, size=14, color=sev_colors.get(severity, "000000"))

        ws_ai.cell(5, 1, "Диагноз").font = Font(bold=True, size=11)
        diag = ai_data.get("diagnosis", "")
        ws_ai.cell(5, 2, diag).alignment = Alignment(wrap_text=True)

        # ── Priority Actions ──
        actions = ai_data.get("priority_actions", [])
        if actions:
            ri = 7
            ws_ai.cell(ri, 1, "🎯 ПРИОРИТЕТНЫЕ ДЕЙСТВИЯ").font = Font(bold=True, size=12, color="2F5496")
            ri += 1
            for ci, h in enumerate(["#", "Действие", "Приоритет"], 1):
                c = ws_ai.cell(ri, ci, h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = Alignment(horizontal="center")
            ri += 1
            for idx, act in enumerate(actions, 1):
                action_text = act if isinstance(act, str) else act.get("action", act.get("text", str(act)))
                priority = act.get("priority", "") if isinstance(act, dict) else ""
                ws_ai.cell(ri, 1, idx).alignment = Alignment(horizontal="center")
                ws_ai.cell(ri, 2, action_text).alignment = Alignment(wrap_text=True)
                ws_ai.cell(ri, 3, priority).alignment = Alignment(horizontal="center")
                if ri % 2 == 0:
                    for c in range(1, 4):
                        ws_ai.cell(ri, c).fill = alt_fill
                ri += 1
            ri += 1
        else:
            ri = 7

        # ── Warehouse Assessments ──
        wh_assess = ai_data.get("warehouse_assessments", [])
        if wh_assess:
            ws_ai.cell(ri, 1, "🏭 ОЦЕНКА СКЛАДОВ").font = Font(bold=True, size=12, color="2F5496")
            ri += 1
            for ci, h in enumerate(["Склад", "Оценка / Проблема", "Кросс %"], 1):
                c = ws_ai.cell(ri, ci, h)
                c.font = hdr_font
                c.fill = PatternFill("solid", fgColor="548235")
                c.alignment = Alignment(horizontal="center")
            ri += 1
            for wha in wh_assess:
                wh_name_ai = wha.get("warehouse", wha.get("name", ""))
                assessment = wha.get("assessment", wha.get("issue", wha.get("text", str(wha))))
                cross_pct_ai = wha.get("cross_pct", "")
                ws_ai.cell(ri, 1, wh_name_ai).font = Font(bold=True)
                ws_ai.cell(ri, 2, assessment).alignment = Alignment(wrap_text=True)
                if cross_pct_ai:
                    ws_ai.cell(ri, 3, f"{cross_pct_ai}%").font = red_font if (isinstance(cross_pct_ai, (int, float)) and cross_pct_ai > 50) else amber_font
                if ri % 2 == 0:
                    for c in range(1, 4):
                        ws_ai.cell(ri, c).fill = alt_fill
                ri += 1
            ri += 1

        # ── Problem SKUs from AI ──
        ai_skus = ai_data.get("problem_skus", [])
        if ai_skus:
            ws_ai.cell(ri, 1, "⚠️ ПРОБЛЕМНЫЕ ТОВАРЫ (ИИ)").font = Font(bold=True, size=12, color="CC0000")
            ri += 1
            ai_sku_headers = ["Товар", "Проблема / Рекомендация", "Кросс %"]
            for ci, h in enumerate(ai_sku_headers, 1):
                c = ws_ai.cell(ri, ci, h)
                c.font = hdr_font
                c.fill = PatternFill("solid", fgColor="C00000")
                c.alignment = Alignment(horizontal="center")
            ri += 1
            for sku_ai in ai_skus:
                sku_label = sku_ai.get("offer_id", sku_ai.get("sku", sku_ai.get("name", "")))
                sku_issue = sku_ai.get("recommendation", sku_ai.get("issue", sku_ai.get("text", str(sku_ai))))
                sku_routes = sku_ai.get("routes", sku_ai.get("cross_routes", []))
                sku_cross = sku_ai.get("cross_pct", "")
                # Build full text
                full_text = sku_issue
                if sku_routes:
                    route_strs = []
                    for rt in sku_routes[:5]:
                        if isinstance(rt, str):
                            route_strs.append(rt)
                        elif isinstance(rt, dict):
                            route_strs.append(f"{rt.get('from', rt.get('from_wh', ''))} → {rt.get('to', rt.get('to_cluster', ''))}: {rt.get('orders', '')} зак.")
                    full_text += "\nМаршруты: " + "; ".join(route_strs)
                ws_ai.cell(ri, 1, sku_label).font = Font(bold=True)
                ws_ai.cell(ri, 2, full_text).alignment = Alignment(wrap_text=True)
                if sku_cross:
                    ws_ai.cell(ri, 3, f"{sku_cross}%")
                if ri % 2 == 0:
                    for c in range(1, 4):
                        ws_ai.cell(ri, c).fill = alt_fill
                ri += 1

    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)
    return buf


@router.get("/wb/cross/excel")
async def wb_cross_excel(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download WB cross-logistics analysis as formatted Excel workbook."""
    analytics = await wb_warehouse_analytics(
        shop_id=shop_id, period=period, db=db, current_user=current_user
    )

    shop = await db.get(Shop, shop_id)
    shop_name = shop.name if shop else f"Shop {shop_id}"

    # Try to get AI analysis from cache (populated when user views AI on the page)
    ai_data = None
    import time as _time
    cache_key = f"wb_cross_ai_{shop_id}_{period}"
    if cache_key in _ai_cache:
        ts, cached = _ai_cache[cache_key]
        if _time.time() - ts < _AI_CACHE_TTL:
            ai_data = cached

    buf = _build_cross_excel(analytics, shop_name, period, "wildberries", ai_data=ai_data)

    filename = f"cross_logistics_wb_shop{shop_id}_{period}d.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/ozon/cross/excel")
async def ozon_cross_excel(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download Ozon cross-logistics analysis as formatted Excel workbook."""
    analytics = await ozon_warehouse_analytics(
        shop_id=shop_id, period=period, db=db, current_user=current_user
    )

    shop = await db.get(Shop, shop_id)
    shop_name = shop.name if shop else f"Shop {shop_id}"

    # Try to get AI analysis from cache (populated when user views AI on the page)
    ai_data = None
    import time as _time
    cache_key = f"ozon_cross_ai_{shop_id}_{period}"
    if cache_key in _ai_cache:
        ts, cached = _ai_cache[cache_key]
        if _time.time() - ts < _AI_CACHE_TTL:
            ai_data = cached

    buf = _build_cross_excel(analytics, shop_name, period, "ozon", ai_data=ai_data)

    filename = f"cross_logistics_ozon_shop{shop_id}_{period}d.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═══════════════════════════════════════════════════════════════
# Geography Sales Excel Builder
# ═══════════════════════════════════════════════════════════════


def _build_geo_excel(
    analytics: dict, shop_name: str, period: int, marketplace: str,
    ai_data: dict | None = None,
    sku_region_data: list | None = None,
) -> io.BytesIO:
    """Build a detailed Excel workbook for geography sales analysis.

    Sheets:
    1. Сводка — KPI, top okrugs
    2. Регионы — per-region breakdown with okrug grouping
    3. Топ товары — top products overall
    4. Товары по округам — per-okrug top SKU drill-down
    5. SKU × Регионы — per-SKU per-region matrix
    6. ИИ-анализ
    """
    import openpyxl
    from datetime import datetime
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="666666")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    section_font = Font(name="Calibri", size=13, bold=True, color="1F4E79")
    section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    green_font = Font(color="006600", bold=True)
    amber_font = Font(color="CC6600", bold=True)
    red_font = Font(color="CC0000", bold=True)
    kpi_label_font = Font(name="Calibri", size=12, bold=True, color="2F5496")
    kpi_val_font = Font(name="Calibri", size=12, bold=True)
    border = Border(
        bottom=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
    )
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    num_fmt = "#,##0"
    money_fmt = "#,##0 ₽"

    workbook = openpyxl.Workbook()

    total_orders = analytics.get("total_orders", 0)
    total_revenue = float(analytics.get("total_revenue", 0))
    avg_check = float(analytics.get("avg_check", 0))
    total_okrugs = analytics.get("total_okrugs", 0)
    total_regions = analytics.get("total_regions", 0)
    regions = analytics.get("regions", [])
    top_products = analytics.get("top_products", [])
    okrug_top_products = analytics.get("okrug_top_products", {})

    mp_label = "Wildberries" if marketplace == "wildberries" else "Ozon"
    okrug_label = "Фед. округ" if marketplace == "wildberries" else "Кластер"
    region_label = "Регион" if marketplace == "wildberries" else "Город"
    sku_label = "Артикул" if marketplace == "wildberries" else "Offer ID"

    # ═════════════════════════════════════════
    # Sheet 1: Сводка
    # ═════════════════════════════════════════
    ws1 = workbook.active
    ws1.title = "Сводка"

    ws1.cell(1, 1, f"📊 География продаж — {shop_name}").font = title_font
    ws1.cell(2, 1, f"Маркетплейс: {mp_label} • Период: {period} дней • {datetime.now().strftime('%d.%m.%Y')}").font = subtitle_font

    kpi_rows = [
        ("Всего заказов", f"{total_orders:,}"),
        ("Выручка", f"{round(total_revenue):,} ₽"),
        ("Средний чек", f"{round(avg_check):,} ₽"),
        (f"{okrug_label}ов", total_okrugs),
        (f"{region_label}ов", total_regions),
    ]
    for i, (label, val) in enumerate(kpi_rows, 4):
        ws1.cell(i, 1, label).font = kpi_label_font
        c = ws1.cell(i, 2, val)
        c.font = kpi_val_font
        c.alignment = Alignment(horizontal="right")

    # Топ округа/кластеры
    ri = len(kpi_rows) + 6
    ws1.cell(ri, 1, f"РАСПРЕДЕЛЕНИЕ ПО {okrug_label.upper()}АМ").font = section_font
    ri += 1

    okrug_agg: dict[str, dict] = {}
    for reg in regions:
        ok = reg.get("okrug", "?")
        if ok not in okrug_agg:
            okrug_agg[ok] = {"orders": 0, "revenue": 0.0, "regions": 0, "max_stab": 0}
        okrug_agg[ok]["orders"] += reg.get("orders", 0)
        okrug_agg[ok]["revenue"] += float(reg.get("revenue", 0))
        okrug_agg[ok]["regions"] += 1
        s = reg.get("stability_pct", 0)
        if s > okrug_agg[ok]["max_stab"]:
            okrug_agg[ok]["max_stab"] = s

    ok_headers = [okrug_label, "Заказов", "Выручка ₽", "Ср. чек ₽", f"{region_label}ов", "Доля %", "Доля выр. %"]
    for ci, h in enumerate(ok_headers, 1):
        c = ws1.cell(ri, ci, h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    ri += 1

    for ok_name, ok_data in sorted(okrug_agg.items(), key=lambda x: x[1]["orders"], reverse=True):
        ok_orders = ok_data["orders"]
        ok_revenue = ok_data["revenue"]
        ok_avg = round(ok_revenue / ok_orders) if ok_orders > 0 else 0
        ok_share = round(ok_orders / total_orders * 100, 1) if total_orders > 0 else 0
        ok_rev_share = round(ok_revenue / total_revenue * 100, 1) if total_revenue > 0 else 0
        ws1.cell(ri, 1, ok_name).font = Font(bold=True)
        ws1.cell(ri, 2, ok_orders).number_format = num_fmt
        ws1.cell(ri, 3, round(ok_revenue)).number_format = money_fmt
        ws1.cell(ri, 4, ok_avg).number_format = money_fmt
        ws1.cell(ri, 5, ok_data["regions"])
        sc = ws1.cell(ri, 6, ok_share)
        sc.number_format = "0.0"
        sc.font = green_font if ok_share >= 25 else (amber_font if ok_share >= 10 else red_font)
        ws1.cell(ri, 7, ok_rev_share).number_format = "0.0"
        for ci in range(1, len(ok_headers) + 1):
            ws1.cell(ri, ci).border = border
        ri += 1

    # ИТОГО строка
    ws1.cell(ri, 1, "ИТОГО").font = Font(bold=True, color="1F4E79")
    ws1.cell(ri, 2, total_orders).number_format = num_fmt
    ws1.cell(ri, 2).font = Font(bold=True)
    ws1.cell(ri, 3, round(total_revenue)).number_format = money_fmt
    ws1.cell(ri, 3).font = Font(bold=True)
    ws1.cell(ri, 4, round(avg_check)).number_format = money_fmt
    ws1.cell(ri, 5, total_regions)
    ws1.cell(ri, 6, 100.0).number_format = "0.0"
    ws1.cell(ri, 7, 100.0).number_format = "0.0"
    for ci in range(1, len(ok_headers) + 1):
        ws1.cell(ri, ci).border = Border(top=Side(style="medium", color="1F4E79"), bottom=Side(style="medium", color="1F4E79"))

    for ci, w in enumerate([30, 12, 16, 14, 12, 10, 12], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ═════════════════════════════════════════
    # Sheet 2: Регионы (полная детализация)
    # ═════════════════════════════════════════
    ws2 = workbook.create_sheet("Регионы")

    reg_headers = [okrug_label, region_label, "Заказов", "Выручка ₽", "Ср. чек ₽",
                   "Стаб. %", "Доля зак. %", "Доля выр. %"]
    for ci, h in enumerate(reg_headers, 1):
        c = ws2.cell(1, ci, h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(reg_headers))}1"

    sorted_regions = sorted(regions, key=lambda r: r.get("orders", 0), reverse=True)
    for ri, reg in enumerate(sorted_regions, 2):
        r_orders = reg.get("orders", 0)
        r_revenue = float(reg.get("revenue", 0))
        r_avg = round(r_revenue / r_orders) if r_orders > 0 else 0
        r_share = round(r_orders / total_orders * 100, 1) if total_orders > 0 else 0
        r_rev_share = round(r_revenue / total_revenue * 100, 1) if total_revenue > 0 else 0
        r_stab = reg.get("stability_pct", 0)
        ws2.cell(ri, 1, reg.get("okrug", ""))
        ws2.cell(ri, 2, reg.get("region", ""))
        ws2.cell(ri, 3, r_orders).number_format = num_fmt
        ws2.cell(ri, 4, round(r_revenue)).number_format = money_fmt
        ws2.cell(ri, 5, r_avg).number_format = money_fmt
        stab_c = ws2.cell(ri, 6, r_stab)
        stab_c.number_format = "0.0"
        stab_c.font = green_font if r_stab >= 50 else (amber_font if r_stab >= 20 else red_font)
        ws2.cell(ri, 7, r_share).number_format = "0.0"
        ws2.cell(ri, 8, r_rev_share).number_format = "0.0"
        for ci in range(1, len(reg_headers) + 1):
            ws2.cell(ri, ci).border = border
            if ri % 2 == 0:
                ws2.cell(ri, ci).fill = alt_fill

    for ci, w in enumerate([28, 28, 12, 16, 14, 10, 12, 12], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # (Sheet "Топ товары" удалён — это общий топ продаж, не относящийся к географии.
    #  Географическая детализация по товарам — в листах "Товары по округам" и "SKU × Регионы".)

    # ═════════════════════════════════════════
    # Sheet 4: Товары по округам
    # ═════════════════════════════════════════
    if okrug_top_products:
        ws4 = workbook.create_sheet(f"Товары по {okrug_label.lower()}ам")

        ok_prod_headers = [okrug_label, sku_label, "Название", "Заказов", "Выручка ₽", "Ср. чек ₽", "Стаб. %"]
        for ci, h in enumerate(ok_prod_headers, 1):
            c = ws4.cell(1, ci, h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        ws4.freeze_panes = "A2"
        ws4.auto_filter.ref = f"A1:{get_column_letter(len(ok_prod_headers))}1"

        ri = 2
        for ok_name in sorted(okrug_top_products.keys(), key=lambda k: okrug_agg.get(k, {}).get("orders", 0), reverse=True):
            prods = okrug_top_products[ok_name]
            if not isinstance(prods, list):
                continue
            # Section header
            for ci in range(1, len(ok_prod_headers) + 1):
                ws4.cell(ri, ci).fill = section_fill
            ok_info = okrug_agg.get(ok_name, {})
            ws4.cell(ri, 1, f"▶ {ok_name}").font = section_font
            ok_o = ok_info.get("orders", 0)
            ws4.cell(ri, 4, ok_o).number_format = num_fmt
            ws4.cell(ri, 4).font = Font(bold=True)
            ws4.cell(ri, 5, round(ok_info.get("revenue", 0))).number_format = money_fmt
            ws4.cell(ri, 5).font = Font(bold=True)
            ri += 1

            for prod in prods:
                if not isinstance(prod, dict):
                    continue
                p_orders = prod.get("orders", 0)
                p_revenue = float(prod.get("revenue", 0))
                p_avg = round(p_revenue / p_orders) if p_orders > 0 else 0
                ws4.cell(ri, 1, "")
                ws4.cell(ri, 2, prod.get("vendor_code", str(prod.get("nm_id", ""))))
                ws4.cell(ri, 3, prod.get("name", ""))
                ws4.cell(ri, 4, p_orders).number_format = num_fmt
                ws4.cell(ri, 5, round(p_revenue)).number_format = money_fmt
                ws4.cell(ri, 6, p_avg).number_format = money_fmt
                stab_c = ws4.cell(ri, 7, prod.get("stability_pct", 0))
                stab_c.number_format = "0.0"
                stab_c.font = green_font if prod.get("stability_pct", 0) >= 50 else (amber_font if prod.get("stability_pct", 0) >= 20 else red_font)
                for ci in range(1, len(ok_prod_headers) + 1):
                    ws4.cell(ri, ci).border = border
                    if ri % 2 == 0:
                        ws4.cell(ri, ci).fill = alt_fill
                ri += 1
            ri += 1  # gap between okrugs

        for ci, w in enumerate([28, 22, 40, 12, 16, 14, 10], 1):
            ws4.column_dimensions[get_column_letter(ci)].width = w

    # ═════════════════════════════════════════
    # Sheet 5: SKU × Регионы (матрица)
    # ═════════════════════════════════════════
    if sku_region_data:
        ws5 = workbook.create_sheet(f"SKU × {region_label}ы")

        sr_headers = [sku_label, "Название", okrug_label, region_label,
                      "Заказов", "Выручка ₽", "Ср. чек ₽", "Стаб. %",
                      "Доля SKU %"]
        for ci, h in enumerate(sr_headers, 1):
            c = ws5.cell(1, ci, h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        ws5.freeze_panes = "A2"
        ws5.auto_filter.ref = f"A1:{get_column_letter(len(sr_headers))}1"

        # Sort: by SKU orders desc, then region orders desc
        sorted_sr = sorted(sku_region_data, key=lambda x: (-x.get("sku_total_orders", 0), -x.get("orders", 0)))

        current_sku = None
        for ri_idx, sr in enumerate(sorted_sr):
            ri = ri_idx + 2
            sku_id = sr.get("sku_label", "")
            show_sku = sku_id != current_sku
            current_sku = sku_id

            if show_sku:
                ws5.cell(ri, 1, sku_id).font = Font(bold=True)
                ws5.cell(ri, 2, sr.get("name", "")).font = Font(bold=True)
            else:
                ws5.cell(ri, 1, "")
                ws5.cell(ri, 2, "")
            ws5.cell(ri, 3, sr.get("okrug", ""))
            ws5.cell(ri, 4, sr.get("region", ""))
            r_orders = sr.get("orders", 0)
            r_rev = float(sr.get("revenue", 0))
            r_avg = round(r_rev / r_orders) if r_orders > 0 else 0
            ws5.cell(ri, 5, r_orders).number_format = num_fmt
            ws5.cell(ri, 6, round(r_rev)).number_format = money_fmt
            ws5.cell(ri, 7, r_avg).number_format = money_fmt
            stab_c = ws5.cell(ri, 8, sr.get("stability_pct", 0))
            stab_c.number_format = "0.0"
            # Share within this SKU
            sku_total = sr.get("sku_total_orders", 1)
            sku_share = round(r_orders / sku_total * 100, 1) if sku_total > 0 else 0
            ws5.cell(ri, 9, sku_share).number_format = "0.0"

            for ci in range(1, len(sr_headers) + 1):
                ws5.cell(ri, ci).border = border
                if ri % 2 == 0:
                    ws5.cell(ri, ci).fill = alt_fill

        for ci, w in enumerate([22, 38, 26, 26, 12, 16, 14, 10, 12], 1):
            ws5.column_dimensions[get_column_letter(ci)].width = w

    # ═════════════════════════════════════════
    # Sheet 6: ИИ-анализ (если есть)
    # ═════════════════════════════════════════
    if ai_data and isinstance(ai_data, dict):
        import re as _re
        ws_ai = workbook.create_sheet("ИИ-анализ")
        ws_ai.column_dimensions["A"].width = 28
        ws_ai.column_dimensions["B"].width = 22
        ws_ai.column_dimensions["C"].width = 18
        ws_ai.column_dimensions["D"].width = 14
        ws_ai.column_dimensions["E"].width = 14
        ws_ai.column_dimensions["F"].width = 50
        wrap_al = Alignment(wrap_text=True, vertical="top")
        center_al = Alignment(horizontal="center", vertical="top")
        hdr_fill_dk = PatternFill("solid", fgColor="2C3E50")
        hdr_font_w = Font(bold=True, size=11, color="FFFFFF")
        section_font = Font(bold=True, size=13, color="1F4E79")

        # ── Title ──
        ws_ai.cell(1, 1, f"ИИ-анализ географии — {shop_name}").font = title_font
        ws_ai.merge_cells("A1:F1")
        ws_ai.cell(2, 1, f"Дата анализа: {datetime.now().strftime('%d.%m.%Y %H:%M')}").font = subtitle_font
        ws_ai.merge_cells("A2:F2")
        ri = 4

        # ── Severity ──
        severity = ai_data.get("severity", "")
        if severity:
            sev_map = {"critical": "🔴 Критично", "warning": "🟡 Внимание", "ok": "🟢 Всё ОК",
                        "medium": "🟡 Средний", "low": "🟢 Низкий", "high": "🔴 Высокий"}
            ws_ai.cell(ri, 1, "Статус").font = Font(bold=True, size=12)
            ws_ai.cell(ri, 2, sev_map.get(severity, severity)).font = Font(size=12)
            ri += 2

        # ── Key Metrics (KPI row) ──
        km = ai_data.get("key_metrics", ai_data.get("context", {}))
        if km and isinstance(km, dict):
            ws_ai.cell(ri, 1, "📊 КЛЮЧЕВЫЕ МЕТРИКИ").font = section_font
            ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
            ri += 1
            kpi_items = []
            for k, lbl in [("concentration_pct", "Концентрация"), ("top3_concentration_pct", "Концентрация"),
                           ("total_clusters", okrug_label + "ов"), ("total_okrugs", "Округов"),
                           ("clusters_with_stable_demand", "Стаб. спрос"), ("underserved_clusters", "Недообслуж."),
                           ("total_orders", "Заказов"), ("total_revenue", "Выручка")]:
                v = km.get(k)
                if v is not None and v != "" and (k, lbl) not in [(ik, il) for ik, il, _ in kpi_items]:
                    if k.endswith("_pct"):
                        kpi_items.append((k, lbl, f"{v}%"))
                    elif k == "total_revenue":
                        kpi_items.append((k, lbl, f"{round(v):,} ₽" if isinstance(v, (int, float)) else str(v)))
                    elif k == "total_orders" and isinstance(v, (int, float)):
                        kpi_items.append((k, lbl, f"{int(v):,}"))
                    else:
                        kpi_items.append((k, lbl, str(v)))
            # Deduplicate labels
            seen_labels = set()
            deduped = []
            for k, lbl, val in kpi_items:
                if lbl not in seen_labels:
                    seen_labels.add(lbl)
                    deduped.append((lbl, val))
            for ci, (lbl, _) in enumerate(deduped, 1):
                ws_ai.cell(ri, ci, lbl).font = Font(bold=True, size=10, color="666666")
                ws_ai.cell(ri, ci).alignment = center_al
            ri += 1
            for ci, (_, val) in enumerate(deduped, 1):
                ws_ai.cell(ri, ci, val).font = Font(bold=True, size=14, color="1F4E79")
                ws_ai.cell(ri, ci).alignment = center_al
            ri += 2

        # ── Diagnosis ──
        diagnosis = ai_data.get("diagnosis", "")
        if diagnosis:
            ws_ai.cell(ri, 1, "📋 ОБЩАЯ ОЦЕНКА").font = section_font
            ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
            ri += 1
            d_lines = diagnosis.replace("\\n", "\n").split("\n")
            for dline in d_lines:
                dline = dline.strip()
                if not dline:
                    continue
                if len(dline) > 140:
                    for s in _re.split(r'(?<=[.!?])\s+', dline):
                        if s.strip():
                            c = ws_ai.cell(ri, 1, s.strip())
                            c.alignment = wrap_al
                            c.font = Font(size=11)
                            ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
                            ws_ai.row_dimensions[ri].height = max(30, 15 * ((len(s) // 110) + 1))
                            ri += 1
                else:
                    c = ws_ai.cell(ri, 1, dline)
                    c.alignment = wrap_al
                    c.font = Font(size=11)
                    ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
                    ws_ai.row_dimensions[ri].height = 30
                    ri += 1
            ri += 1

        # ── Concentration ──
        conc_data = ai_data.get("concentration", {})
        if conc_data and isinstance(conc_data, dict):
            risk_map = {"high": "🔴 Высокий", "medium": "🟡 Средний", "low": "🟢 Низкий"}
            risk = conc_data.get("risk_level", "")
            ws_ai.cell(ri, 1, "📍 КОНЦЕНТРАЦИЯ ПРОДАЖ").font = section_font
            if risk:
                ws_ai.cell(ri, 3, risk_map.get(risk, risk)).font = Font(bold=True, size=12)
            ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=2)
            ri += 1
            conc_summary = conc_data.get("summary", "")
            if conc_summary:
                c = ws_ai.cell(ri, 1, conc_summary)
                c.alignment = wrap_al
                c.font = Font(size=11)
                ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
                ws_ai.row_dimensions[ri].height = max(30, 15 * ((len(conc_summary) // 110) + 1))
                ri += 1
            top_regs = conc_data.get("top_regions", [])
            if top_regs and isinstance(top_regs, list):
                ri += 1
                for ci, h in enumerate([okrug_label, "Заказов", "Доля %", "Стабильность %"], 1):
                    c = ws_ai.cell(ri, ci, h)
                    c.font = hdr_font_w
                    c.fill = hdr_fill_dk
                    c.alignment = center_al
                ri += 1
                for tr_item in top_regs:
                    if not isinstance(tr_item, dict):
                        continue
                    ws_ai.cell(ri, 1, tr_item.get("region", tr_item.get("okrug", ""))).font = Font(size=11)
                    ws_ai.cell(ri, 2, tr_item.get("orders", 0)).number_format = num_fmt
                    ws_ai.cell(ri, 3, tr_item.get("share_pct", 0)).number_format = "0.0"
                    stab_v = tr_item.get("stability_pct", 0)
                    sc = ws_ai.cell(ri, 4, stab_v)
                    sc.number_format = "0.0"
                    if isinstance(stab_v, (int, float)):
                        sc.font = green_font if stab_v >= 60 else amber_font if stab_v >= 30 else red_font
                    for ci in range(1, 5):
                        ws_ai.cell(ri, ci).border = border
                        ws_ai.cell(ri, ci).alignment = center_al
                    ri += 1
            conc_rec = conc_data.get("recommendation", "")
            if conc_rec:
                ri += 1
                c = ws_ai.cell(ri, 1, f"💡 {conc_rec}")
                c.alignment = wrap_al
                c.font = Font(size=11, italic=True, color="2E7D32")
                ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
                ws_ai.row_dimensions[ri].height = max(40, 15 * ((len(conc_rec) // 100) + 1))
                ri += 1
            ri += 1

        # ── Product Insights ──
        prod_ins = ai_data.get("product_insights", [])
        if prod_ins and isinstance(prod_ins, list):
            ws_ai.cell(ri, 1, f"🔎 ИНСАЙТЫ ПО ТОВАРАМ ({len(prod_ins)})").font = section_font
            ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
            ri += 1
            type_lbl = {"stable_leader": "🏆 Лидер", "unstable_demand": "⚠️ Нестабильный",
                        "regional_champion": "📍 Регион. чемпион", "cross_delivery_problem": "🚛 Кросс-проблема",
                        "dead_stock_risk": "📦 Риск залёживания"}
            action_lbl = {"redistribute": "Перераспределить", "increase_supply": "Увеличить поставки", "monitor": "Мониторить"}
            for pi in prod_ins:
                if not isinstance(pi, dict):
                    continue
                vc = pi.get("vendor_code", pi.get("offer_id", ""))
                pname = pi.get("name", "")
                itype = pi.get("insight_type", "")
                o_v = pi.get("orders", 0)
                rc = pi.get("regions_count", pi.get("clusters_count", 0))
                sp = pi.get("stability_pct", 0)
                # Header
                ws_ai.cell(ri, 1, f"{vc}  {pname}").font = Font(bold=True, size=12)
                ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=4)
                ws_ai.cell(ri, 5, type_lbl.get(itype, itype)).font = Font(bold=True, size=11)
                ws_ai.merge_cells(start_row=ri, start_column=5, end_row=ri, end_column=6)
                bg = "E8F5E9" if itype == "stable_leader" else "FFF3E0" if itype in ("unstable_demand", "cross_delivery_problem") else "F3F4F6"
                for ci in range(1, 7):
                    ws_ai.cell(ri, ci).fill = PatternFill("solid", fgColor=bg)
                ri += 1
                # Stats
                parts = []
                if o_v:
                    parts.append(f"{o_v} заказов")
                if rc:
                    parts.append(f"{rc} кластеров")
                if sp:
                    parts.append(f"стаб. {sp}%")
                if parts:
                    ws_ai.cell(ri, 1, " · ".join(parts)).font = Font(size=10, color="666666")
                    ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
                    ri += 1
                # Detail
                det = pi.get("detail", "")
                if det:
                    c = ws_ai.cell(ri, 1, det)
                    c.alignment = wrap_al
                    c.font = Font(size=11)
                    ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
                    ws_ai.row_dimensions[ri].height = max(40, 15 * ((len(det) // 100) + 1))
                    ri += 1
                # Action + effect
                act = pi.get("action", "")
                eff = pi.get("expected_effect", "")
                if act or eff:
                    txt_parts = []
                    if act:
                        txt_parts.append(f"Действие: {action_lbl.get(act, act)}")
                    if eff:
                        txt_parts.append(f"Эффект: {eff}")
                    c = ws_ai.cell(ri, 1, " | ".join(txt_parts))
                    c.font = Font(size=10, italic=True, color="1565C0")
                    c.alignment = wrap_al
                    ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
                    ws_ai.row_dimensions[ri].height = max(30, 15 * ((len(" | ".join(txt_parts)) // 110) + 1))
                    ri += 1
                ri += 1  # spacing

        # ── Logistics Match ──
        logistics = ai_data.get("logistics_match", [])
        if logistics and isinstance(logistics, list):
            ws_ai.cell(ri, 1, "🚛 ЛОГИСТИЧЕСКОЕ СООТВЕТСТВИЕ").font = section_font
            ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
            ri += 1
            for ci, h in enumerate([okrug_label, "Заказов", "Ближ. склад", "Сток", "Откуда отгр.", "Кросс %"], 1):
                c = ws_ai.cell(ri, ci, h)
                c.font = hdr_font_w
                c.fill = hdr_fill_dk
                c.alignment = center_al
            ri += 1
            for lg in logistics:
                if not isinstance(lg, dict):
                    continue
                ws_ai.cell(ri, 1, lg.get("okrug", "")).alignment = wrap_al
                ws_ai.cell(ri, 2, lg.get("orders", 0)).number_format = num_fmt
                ws_ai.cell(ri, 3, lg.get("nearest_warehouse", "")).alignment = wrap_al
                ws_ai.cell(ri, 4, lg.get("warehouse_stock", 0)).number_format = num_fmt
                ws_ai.cell(ri, 5, lg.get("serving_warehouse", "")).alignment = wrap_al
                cp = lg.get("cross_pct", 0)
                cc = ws_ai.cell(ri, 6, cp)
                cc.number_format = "0"
                if isinstance(cp, (int, float)):
                    cc.font = red_font if cp >= 50 else amber_font if cp >= 25 else green_font
                for ci in range(1, 7):
                    ws_ai.cell(ri, ci).border = border
                ri += 1
                # Detail + recommendation
                lg_d = lg.get("detail", "")
                lg_r = lg.get("recommendation", "")
                if lg_d or lg_r:
                    txt = f"{lg_d} → {lg_r}" if lg_d and lg_r else lg_d or lg_r
                    c = ws_ai.cell(ri, 1, txt)
                    c.alignment = wrap_al
                    c.font = Font(size=10, italic=True)
                    ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
                    ws_ai.row_dimensions[ri].height = max(30, 15 * ((len(txt) // 100) + 1))
                    ri += 1
            ri += 1

        # ── Recommendations ──
        recs = ai_data.get("recommendations", [])
        if recs:
            ws_ai.cell(ri, 1, "💡 РЕКОМЕНДАЦИИ").font = section_font
            ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
            ri += 1
            for ci, h in enumerate(["Действие", "Обоснование", "Приоритет"], 1):
                c = ws_ai.cell(ri, ci, h)
                c.font = hdr_font_w
                c.fill = hdr_fill_dk
                c.alignment = center_al
            ri += 1
            for idx, rec in enumerate(recs, 1):
                if isinstance(rec, dict):
                    act_t = f"{idx}. {rec.get('action', '')}"
                    rsn = rec.get("reason", "")
                    p = rec.get("priority", "")
                    p_l = {"high": "🔴 Высокий", "medium": "🟡 Средний", "low": "🟢 Низкий"}.get(p, p)
                    ws_ai.cell(ri, 1, act_t).font = Font(bold=True, size=11)
                    ws_ai.cell(ri, 1).alignment = wrap_al
                    ws_ai.cell(ri, 2, rsn).alignment = wrap_al
                    ws_ai.cell(ri, 3, p_l).alignment = center_al
                    m_l = max(len(act_t), len(rsn))
                    ws_ai.row_dimensions[ri].height = max(30, 15 * ((m_l // 80) + 1))
                elif isinstance(rec, str):
                    ws_ai.cell(ri, 1, f"{idx}. {rec}").alignment = wrap_al
                    ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=3)
                for ci in range(1, 4):
                    ws_ai.cell(ri, ci).border = border
                    if ri % 2 == 0:
                        ws_ai.cell(ri, ci).fill = alt_fill
                ri += 1
            ri += 1

        # ── General Tips ──
        tips = ai_data.get("general_tips", ai_data.get("insights", ai_data.get("key_insights", [])))
        if tips and isinstance(tips, list):
            ws_ai.cell(ri, 1, "🔑 ОБЩИЕ РЕКОМЕНДАЦИИ").font = section_font
            ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
            ri += 1
            for idx, tip in enumerate(tips, 1):
                txt = tip if isinstance(tip, str) else tip.get("insight", tip.get("text", "")) if isinstance(tip, dict) else str(tip)
                if txt:
                    c = ws_ai.cell(ri, 1, f"{idx}. {txt}")
                    c.alignment = wrap_al
                    c.font = Font(size=11)
                    ws_ai.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
                    ws_ai.row_dimensions[ri].height = max(30, 15 * ((len(txt) // 100) + 1))
                    for ci in range(1, 7):
                        ws_ai.cell(ri, ci).border = border
                    ri += 1

    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)
    return buf


@router.get("/wb/geography/excel")
async def wb_geography_excel(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download WB geography sales analysis as formatted Excel workbook."""
    analytics = await get_wb_geography(
        shop_id=shop_id, period=period, nm_ids=None, db=db, current_user=current_user
    )

    shop = await db.get(Shop, shop_id)
    shop_name = shop.name if shop else f"Shop {shop_id}"

    # Try to get AI analysis from cache
    ai_data = None
    cache_key = f"geo_ai_{shop_id}_{period}"
    if cache_key in _ai_cache:
        ts, cached = _ai_cache[cache_key]
        if time.time() - ts < _AI_CACHE_TTL:
            ai_data = cached

    # ── Collect per-SKU per-region matrix data ──
    sku_region_data = []
    try:
        from app.core.clickhouse import get_clickhouse_client
        ch = get_clickhouse_client()
        today = date.today()
        d_start = today - timedelta(days=period)
        total_weeks = max(1, period // 7)

        rows = ch.query("""
            SELECT
                nm_id,
                oblast_okrug_name AS okrug,
                region_name AS region,
                count() AS orders,
                sum(toFloat64(price_with_disc)) AS revenue,
                uniqExact(toMonday(date)) AS active_weeks
            FROM mms_analytics.fact_orders_raw FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND date >= {d_start:Date}
              AND is_cancel = 0
              AND region_name != ''
            GROUP BY nm_id, okrug, region
            ORDER BY nm_id, orders DESC
        """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows
        ch.close()

        # Get product names from PG
        nm_ids_in = list(set(int(r[0]) for r in rows))
        prod_map: dict[int, dict] = {}
        if nm_ids_in:
            nm_list = ", ".join(str(x) for x in nm_ids_in[:500])
            pg_rows = (await db.execute(
                text(f"""
                    SELECT nm_id, vendor_code, name
                    FROM dim_products
                    WHERE shop_id = :sid AND nm_id IN ({nm_list})
                """),
                {"sid": shop_id},
            )).fetchall()
            for r in pg_rows:
                prod_map[r[0]] = {"vendor_code": r[1] or "", "name": (r[2] or "")[:80]}

        # Compute per-SKU totals
        sku_totals: dict[int, int] = {}
        for r in rows:
            nm = int(r[0])
            sku_totals[nm] = sku_totals.get(nm, 0) + int(r[3])

        for r in rows:
            nm = int(r[0])
            prod = prod_map.get(nm, {})
            active_weeks = int(r[5])
            stab = round(active_weeks / total_weeks * 100, 1)
            sku_region_data.append({
                "sku_label": prod.get("vendor_code") or str(nm),
                "name": prod.get("name", ""),
                "okrug": str(r[1]),
                "region": str(r[2]),
                "orders": int(r[3]),
                "revenue": float(r[4]),
                "stability_pct": stab,
                "sku_total_orders": sku_totals.get(nm, 0),
            })
    except Exception as e:
        import logging
        logging.exception("WB geo Excel: sku_region_data collection failed: %s", e)

    buf = _build_geo_excel(analytics, shop_name, period, "wildberries",
                           ai_data=ai_data, sku_region_data=sku_region_data)

    filename = f"geography_wb_shop{shop_id}_{period}d.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/ozon/geography/excel")
async def ozon_geography_excel(
    shop_id: int = Query(...),
    period: int = Query(30, ge=7, le=90),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download Ozon geography sales analysis as formatted Excel workbook."""
    analytics = await get_ozon_geography(
        shop_id=shop_id, period=period, skus=None, db=db, current_user=current_user
    )

    # Normalize Ozon format (clusters/cities) → unified format (regions/okrugs)
    normalized_regions = []
    for cl in analytics.get("clusters", []):
        for city in cl.get("cities", []):
            normalized_regions.append({
                "okrug": cl.get("cluster", ""),
                "region": city.get("city", ""),
                "orders": city.get("orders", 0),
                "revenue": city.get("revenue", 0),
                "avg_check": city.get("avg_check", 0),
                "stability_pct": city.get("stability_pct", 0),
                "share_pct": city.get("share_pct", 0),
            })
    # Normalize top_products: sku→nm_id, offer_id→vendor_code, cluster_count→okrug_count, city_count→region_count
    normalized_products = []
    for p in analytics.get("top_products", []):
        normalized_products.append({
            "nm_id": p.get("sku", 0),
            "vendor_code": p.get("offer_id", ""),
            "name": p.get("name", ""),
            "orders": p.get("orders", 0),
            "revenue": p.get("revenue", 0),
            "avg_check": p.get("avg_check", 0),
            "okrug_count": p.get("cluster_count", 0),
            "region_count": p.get("city_count", 0),
            "stability_pct": p.get("stability_pct", 0),
            "share_pct": p.get("share_pct", 0),
        })
    # Normalize cluster_top_products → okrug_top_products with unified keys
    normalized_okrug_prods: dict[str, list] = {}
    for cluster_name, prods in analytics.get("cluster_top_products", {}).items():
        if not isinstance(prods, list):
            continue
        normalized_okrug_prods[cluster_name] = []
        for p in prods:
            if not isinstance(p, dict):
                continue
            normalized_okrug_prods[cluster_name].append({
                "nm_id": p.get("sku", 0),
                "vendor_code": p.get("offer_id", ""),
                "name": p.get("name", "") or p.get("offer_id", ""),
                "orders": p.get("orders", 0),
                "revenue": p.get("revenue", 0),
                "avg_check": p.get("avg_check", 0),
                "stability_pct": p.get("stability_pct", 0),
            })
    analytics_normalized = {
        "total_orders": analytics.get("total_orders", 0),
        "total_revenue": analytics.get("total_revenue", 0),
        "avg_check": analytics.get("avg_check", 0),
        "total_okrugs": analytics.get("total_clusters", 0),
        "total_regions": analytics.get("total_cities", 0),
        "regions": normalized_regions,
        "top_products": normalized_products,
        "okrug_top_products": normalized_okrug_prods,
    }

    shop = await db.get(Shop, shop_id)
    shop_name = shop.name if shop else f"Shop {shop_id}"

    # Try to get AI analysis from cache
    ai_data = None
    cache_key = f"ozon_geo_ai_{shop_id}_{period}"
    if cache_key in _ai_cache:
        ts, cached = _ai_cache[cache_key]
        if time.time() - ts < _AI_CACHE_TTL:
            ai_data = cached

    # ── Collect per-SKU per-city matrix data ──
    sku_region_data = []
    try:
        from app.core.clickhouse import get_clickhouse_client
        ch = get_clickhouse_client()
        today = date.today()
        d_start = today - timedelta(days=period)
        total_weeks = max(1, period // 7)

        rows = ch.query("""
            SELECT
                sku,
                cluster_to AS cluster,
                city,
                count() AS orders,
                sum(toFloat64(price) * quantity) AS revenue,
                uniqExact(toMonday(toDate(order_date))) AS active_weeks,
                any(offer_id) AS offer_id_val
            FROM mms_analytics.fact_ozon_orders FINAL
            WHERE shop_id = {shop_id:UInt32}
              AND order_date >= {d_start:Date}
              AND status NOT IN ('cancelled', 'canceled')
              AND city != ''
            GROUP BY sku, cluster, city
            ORDER BY sku, orders DESC
        """, parameters={"shop_id": shop_id, "d_start": d_start}).result_rows
        ch.close()

        # Build offer_id map from CH results (no PG needed — dim_ozon_products may be empty)
        sku_offer_map: dict[int, str] = {}
        for r in rows:
            s = int(r[0])
            if s not in sku_offer_map:
                sku_offer_map[s] = str(r[6]) if r[6] else ""

        # Per-SKU totals
        sku_totals: dict[int, int] = {}
        for r in rows:
            s = int(r[0])
            sku_totals[s] = sku_totals.get(s, 0) + int(r[3])

        for r in rows:
            s = int(r[0])
            oid = sku_offer_map.get(s, "")
            active_weeks = int(r[5])
            stab = round(active_weeks / total_weeks * 100, 1)
            sku_region_data.append({
                "sku_label": oid or str(s),
                "name": oid,  # offer_id serves as product identifier for Ozon
                "okrug": str(r[1]),
                "region": str(r[2]),
                "orders": int(r[3]),
                "revenue": float(r[4]),
                "stability_pct": stab,
                "sku_total_orders": sku_totals.get(s, 0),
            })
    except Exception as e:
        import logging
        logging.exception("Ozon geo Excel: sku_region_data collection failed: %s", e)

    buf = _build_geo_excel(analytics_normalized, shop_name, period, "ozon",
                           ai_data=ai_data, sku_region_data=sku_region_data)

    filename = f"geography_ozon_shop{shop_id}_{period}d.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

